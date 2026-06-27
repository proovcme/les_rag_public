from __future__ import annotations
"""
Е.Ж.И.К. + С.А.М.О.В.А.Р. // parquet_writer.py
=================================================
Нормализатор табличных документов для индексации в Qdrant.

Поддерживает:
- КС-2 (Акт выполненных работ)
- АОСР (Акт освидетельствования скрытых работ)
- Спецификации оборудования (ГОСТ 21.110)
- Ведомости чертежей
- Сметы (1С / Гранд-Смета / произвольный Excel)
- Произвольные таблицы (LLM-маппинг)

Стратегия:
1. Читаем XLSX/CSV через openpyxl/pandas
2. LLM (через /api/chat или локально) анализирует заголовки → определяет тип + маппинг колонок
3. Нормализуем в единую схему → сохраняем Parquet
4. Каждая строка → отдельный чанк в Qdrant с rich payload

Зависимости:
    pip install openpyxl pandas pyarrow --break-system-packages

Запуск вручную:
    python3 parquet_writer.py /path/to/smeta.xlsx --out /tmp/parquet_out
"""

import json
import logging
import os
import re
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Optional
from xml.etree import ElementTree

logger = logging.getLogger("les.parquet")

# ─────────────────────────────────────────
# СТАНДАРТНАЯ СХЕМА НОРМАЛИЗОВАННОЙ СТРОКИ
# ─────────────────────────────────────────

# Все табличные документы приводим к этой схеме.
# Поля которых нет в источнике — остаются пустыми.
STANDARD_SCHEMA = {
    # Идентификация
    "doc_type":      "",   # KS2 / AOSR / SPEC / VEDOMOST / SMETA / TABLE
    "doc_title":     "",   # Название документа/листа
    "source_file":   "",   # Имя исходного файла

    # Позиция
    "pos":           "",   # Номер позиции / строки
    "section":       "",   # Раздел / глава
    "subsection":    "",   # Подраздел

    # Наименование
    "name":          "",   # Наименование работы / оборудования
    "code":          "",   # Шифр, артикул, обозначение
    "mark":          "",   # Марка, тип

    # Количество
    "unit":          "",   # Единица измерения
    "qty":           None, # Количество (float)
    "qty_per_unit":  None, # Расход на единицу

    # Стоимость
    "price":         None, # Цена единицы
    "amount":        None, # Сумма
    "amount_mat":    None, # Материалы
    "amount_work":   None, # Работы

    # Для КС-2
    "work_volume":   None, # Объём работ по договору
    "work_done":     None, # Выполнено в отчётном периоде
    "work_since_start": None, # Выполнено с начала строительства

    # Для АОСР
    "work_name":     "",   # Наименование скрытых работ
    "norms_refs":    "",   # Ссылки на нормативы
    "materials_used":"",   # Применённые материалы
    "date_start":    "",   # Начало работ
    "date_end":      "",   # Конец работ

    # Для спецификаций
    "position":      "",   # Позиция по спецификации
    "designation":   "",   # Обозначение (ГОСТ/ТУ)
    "weight_unit":   None, # Масса единицы
    "weight_total":  None, # Масса общая

    # Примечания
    "note":          "",
    "raw_row":       "",   # Оригинальная строка как JSON (резерв)
    "source_page":   None,
    "table_index":   None,
    "needs_ocr":     False,
    "extractor":     "",
}

# ─────────────────────────────────────────
# ТИПЫ ДОКУМЕНТОВ
# ─────────────────────────────────────────

DOC_TYPES = {
    "KS2":      "Акт о приёмке выполненных работ (КС-2)",
    "AOSR":     "Акт освидетельствования скрытых работ (АОСР)",
    "SPEC":     "Спецификация оборудования (ГОСТ 21.110)",
    "VEDOMOST": "Ведомость чертежей / ссылочных документов",
    "SMETA":    "Смета / Локальный сметный расчёт",
    "TABLE":    "Произвольная таблица",
}

# Ключевые слова для автодетекта типа по заголовкам
DOC_TYPE_HINTS = {
    "KS2":      ["выполненных работ", "кс-2", "кс2", "акт приёмки", "период"],
    "AOSR":     ["скрытых работ", "аоср", "освидетельствования", "скрытые работы"],
    "SPEC":     ["спецификация", "обозначение", "масса", "поз.", "позиция", "гост 21.110"],
    "VEDOMOST": ["ведомость", "чертежей", "ссылочных документов", "лист", "формат"],
    "SMETA":    ["смета", "расценка", "норма", "гэсн", "фер", "тер", "гранд"],
}


# ─────────────────────────────────────────
# LLM-МАППИНГ КОЛОНОК
# ─────────────────────────────────────────

COLUMN_MAPPING_PROMPT = """Ты — эксперт по строительной документации России.
Тебе дан список заголовков колонок из Excel-файла.
Определи:
1. Тип документа (KS2 / AOSR / SPEC / VEDOMOST / SMETA / TABLE)
2. Маппинг каждого заголовка на стандартное поле схемы

Стандартные поля: pos, section, subsection, name, code, mark, unit, qty, qty_per_unit, 
price, amount, amount_mat, amount_work, work_volume, work_done, work_since_start,
work_name, norms_refs, materials_used, date_start, date_end, 
position, designation, weight_unit, weight_total, note

Если колонка не подходит ни к одному полю — используй "skip".

Заголовки: {headers}

Первые 3 строки данных для контекста:
{sample_rows}

Отвечай ТОЛЬКО JSON без пояснений:
{{
  "doc_type": "KS2",
  "mapping": {{
    "Наименование работ": "name",
    "Ед.изм.": "unit",
    "Кол-во": "qty",
    "Цена": "price",
    "Сумма": "amount",
    "Примечание": "note"
  }}
}}"""


def _detect_doc_type_simple(headers: list, sheet_name: str = "") -> str:
    """Быстрый детект типа по ключевым словам без LLM."""
    text = " ".join(headers + [sheet_name]).lower()
    if any(token in text for token in ("цена", "сумма", "стоимость", "расценка")):
        return "SMETA"
    if (
        any(token in text for token in ("позиция", "поз.", "поз "))
        and any(token in text for token in ("кол-во", "количество", "ед.изм", "единица"))
    ):
        return "SPEC"
    scores = {dt: 0 for dt in DOC_TYPES}
    for dt, hints in DOC_TYPE_HINTS.items():
        for hint in hints:
            if hint in text:
                scores[dt] += 1
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "TABLE"


def _map_columns_simple(headers: list) -> dict:
    """Детерминированный mapping типовых русских табличных заголовков."""
    mapping = {}
    patterns = [
        ("pos", ("№", "номер", "поз", "позиция", "п/п")),
        ("section", ("раздел", "глава", "этап")),
        ("subsection", ("подраздел", "подэтап")),
        # «Наименование …» — это всегда колонка наименования, даже если в тексте есть
        # «материалов»/«работ» (иначе amount_mat/amount_work перехватывают её как число
        # → имя теряется, строка отбрасывается). Высокий приоритет до amount-правил.
        ("name", ("наименование", "наимен.")),
        ("amount", ("сумма", "итого", "всего")),
        ("amount_mat", ("материал", "материалы")),
        ("amount_work", ("работа", "работы", "зп", "оплата труда")),
        ("work_done", ("выполнено", "отчет", "отчёт")),
        ("work_since_start", ("с начала", "накопительно")),
        ("price", ("цена", "стоимость ед", "единичная")),
        ("qty_per_unit", ("расход", "на единицу")),
        ("qty", ("кол-во", "количество", "объем", "объём", "qty")),
        ("unit", ("ед.изм", "единица", "изм.", "ед.")),
        ("name", ("наименование", "работ", "оборудован", "материал", "ресурс")),
        ("code", ("шифр", "код", "расценка", "артикул")),
        ("mark", ("марка", "тип", "модель")),
        ("norms_refs", ("норматив", "гост", "сп ", "снип", "ту")),
        ("designation", ("обозначение", "документ")),
        ("weight_unit", ("масса ед", "масса 1")),
        ("weight_total", ("масса общ", "общая масса")),
        ("note", ("примеч", "комментар", "основание")),
    ]
    for header in headers:
        h = str(header).strip()
        lower = h.lower()
        field = "skip"
        for candidate, hints in patterns:
            if any(hint in lower for hint in hints):
                field = candidate
                break
        mapping[h] = field
    return mapping


async def _llm_map_columns(
    headers: list,
    sample_rows: list,
    llm_url: str = "http://localhost:8050/api/chat"
) -> dict:
    """
    Запрашивает LLM для маппинга колонок.
    Возвращает {"doc_type": str, "mapping": {header: field}}
    """
    try:
        import httpx

        sample_str = "\n".join(
            json.dumps(row, ensure_ascii=False) for row in sample_rows[:3]
        )
        prompt = COLUMN_MAPPING_PROMPT.format(
            headers=json.dumps(headers, ensure_ascii=False),
            sample_rows=sample_str,
        )

        async with httpx.AsyncClient(timeout=60.0) as client:
            r = await client.post(llm_url, json={"question": prompt})
            r.raise_for_status()
            data = r.json()
            answer = data.get("answer", data.get("response", ""))

        # Извлекаем JSON из ответа
        m = re.search(r"\{.*\}", answer, re.DOTALL)
        if m:
            result = json.loads(m.group(0))
            return result
    except Exception as e:
        logger.warning(f"[PARQUET] LLM маппинг недоступен: {e}, используем простой детект")

    # Fallback: простой детект + пустой маппинг
    doc_type = _detect_doc_type_simple(headers)
    return {"doc_type": doc_type, "mapping": _map_columns_simple(headers)}


# ─────────────────────────────────────────
# ЧИТАЛКА XLSX
# ─────────────────────────────────────────

def _find_header_row(ws, max_scan: int = 20) -> int:
    """
    Ищем строку с заголовками — обычно это первая непустая строка
    с несколькими заполненными ячейками.
    Гранд-Смета и 1С часто имеют шапку на 3-5 строках выше данных.
    """
    best_row = 0
    best_score = 0
    max_row = ws.max_row or max_scan
    scan_rows = min(max_scan, max_row)
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True),
        start=1,
    ):
        non_empty = sum(1 for v in row if v is not None and str(v).strip())
        text_cells = sum(1 for v in row if v is not None and isinstance(v, str) and len(str(v).strip()) > 1)
        score = text_cells * 2 + non_empty
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def _xls_to_xlsx_tmp(xls_path: str) -> str:
    """openpyxl не читает старый .xls — конвертируем в .xlsx через pandas/xlrd
    (вычисленные значения, без формул) и дальше используем штатный openpyxl-конвейер."""
    import tempfile
    import re as _re
    import pandas as pd
    sheets = pd.read_excel(xls_path, sheet_name=None, header=None)  # engine=xlrd для .xls
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
        for idx, (name, df) in enumerate(sheets.items()):
            safe = _re.sub(r"[\[\]:*?/\\]", "_", str(name))[:31] or f"s{idx}"
            df.to_excel(writer, sheet_name=safe, header=False, index=False)
    return tmp.name


def read_xlsx_sheets(file_path: str) -> list:
    """
    Читает XLSX и возвращает list[dict] — один элемент на лист:
    {"sheet_name": str, "headers": list, "rows": list[dict], "raw_headers_row": int}
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl не установлен: pip install openpyxl --break-system-packages")

    if str(file_path).lower().endswith(".xls"):
        file_path = _xls_to_xlsx_tmp(file_path)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is not None and ws.max_row < 2:
            continue

        header_row_idx = _find_header_row(ws)
        if header_row_idx == 0:
            continue

        # Заголовки
        header_values = next(
            ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True),
            (),
        )
        headers = [
            str(v).strip() if v is not None and str(v).strip() else f"col_{idx}"
            for idx, v in enumerate(header_values, start=1)
        ]
        if not headers:
            continue

        # Строки данных
        rows = []
        for row_vals in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            # Пропускаем полностью пустые строки
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue
            row_dict = {}
            for i, h in enumerate(headers):
                v = row_vals[i] if i < len(row_vals) else None
                row_dict[h] = v
            rows.append(row_dict)

        if rows:
            sheets.append({
                "sheet_name": sheet_name,
                "headers": headers,
                "rows": rows,
                "header_row": header_row_idx,
            })

    wb.close()
    return sheets


def read_csv(file_path: str) -> list:
    """Читаем CSV, возвращаем в том же формате что и read_xlsx_sheets."""
    try:
        import csv
        rows = []
        with open(file_path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            for row in reader:
                rows.append(dict(row))
        if rows:
            return [{"sheet_name": "CSV", "headers": list(headers), "rows": rows, "header_row": 1}]
    except UnicodeDecodeError:
        # Пробуем cp1251
        try:
            import csv
            rows = []
            with open(file_path, encoding="cp1251", newline="") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                for row in reader:
                    rows.append(dict(row))
            if rows:
                return [{"sheet_name": "CSV", "headers": list(headers), "rows": rows, "header_row": 1}]
        except Exception as e:
            logger.error(f"[PARQUET] CSV read error: {e}")
    return []


def _unique_headers(headers: list) -> list:
    result = []
    seen = {}
    for i, header in enumerate(headers, 1):
        name = str(header or "").strip()
        name = _WS_RE.sub(" ", name) if name else f"col_{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        result.append(name)
    return result


_WS_RE = re.compile(r"\s+")


def _clean_pdf_table(raw_table: list) -> tuple[list, list[dict]]:
    """Очищает таблицу из PDF: склейка шапки, пустые колонки, raw rows."""
    rows = [
        ["" if cell is None else _WS_RE.sub(" ", str(cell).strip()) for cell in row]
        for row in raw_table
        if row and any(str(cell or "").strip() for cell in row)
    ]
    if len(rows) < 2:
        return [], []

    width = max(len(row) for row in rows)
    rows = [row + [""] * (width - len(row)) for row in rows]
    keep_cols = [
        idx for idx in range(width)
        if any(row[idx] for row in rows)
    ]
    if not keep_cols:
        return [], []
    rows = [[row[idx] for idx in keep_cols] for row in rows]

    header_depth = 1
    for idx, row in enumerate(rows[:3]):
        text_cells = sum(1 for cell in row if cell and not _safe_float(cell))
        numeric_cells = sum(1 for cell in row if _safe_float(cell) is not None)
        if idx > 0 and text_cells >= numeric_cells:
            header_depth = idx + 1
        elif idx > 0:
            break

    header_rows = rows[:header_depth]
    headers = []
    for col_idx in range(len(rows[0])):
        parts = [row[col_idx] for row in header_rows if row[col_idx]]
        headers.append(" ".join(parts) if parts else f"col_{col_idx + 1}")
    headers = _unique_headers(headers)

    data_rows = []
    for row in rows[header_depth:]:
        if not any(row):
            continue
        data_rows.append({header: row[idx] for idx, header in enumerate(headers)})
    return headers, data_rows


def _table_is_usable(headers: list, rows: list[dict]) -> bool:
    if not headers or not rows:
        return False
    useful_headers = sum(1 for h in headers if not str(h).startswith("col_"))
    non_empty_cells = sum(
        1
        for row in rows[:10]
        for value in row.values()
        if value is not None and str(value).strip()
    )
    return len(headers) >= 2 and non_empty_cells >= max(2, len(rows[:10])) and useful_headers >= 1


def _pdf_max_pages() -> int:
    try:
        return max(1, int(os.getenv("PDF_TABLE_MAX_PAGES", "30")))
    except ValueError:
        return 30


def _pdf_max_tables() -> int:
    try:
        return max(1, int(os.getenv("PDF_TABLE_MAX_TABLES", "50")))
    except ValueError:
        return 50


def _extract_tables_pymupdf(file_path: str, max_pages: int, max_tables: int) -> tuple[list, set[int]]:
    try:
        import fitz
    except ImportError:
        return [], set()

    sheets = []
    scanned_pages = set()
    doc = fitz.open(file_path)
    try:
        for page_no in range(min(max_pages, len(doc))):
            page_idx = page_no + 1
            page = doc[page_no]
            if not page.get_text("text").strip():
                scanned_pages.add(page_idx)
                continue
            finder = getattr(page, "find_tables", None)
            if not finder:
                continue
            try:
                tables = finder()
            except Exception as e:
                logger.debug("[PDF_TABLE] PyMuPDF page %s failed: %s", page_idx, e)
                continue
            for table_idx, table in enumerate(getattr(tables, "tables", []) or [], 1):
                try:
                    headers, rows = _clean_pdf_table(table.extract())
                except Exception as e:
                    logger.debug("[PDF_TABLE] PyMuPDF table cleanup failed: %s", e)
                    continue
                if not _table_is_usable(headers, rows):
                    continue
                sheets.append({
                    "sheet_name": f"page_{page_idx}_table_{table_idx}",
                    "headers": headers,
                    "rows": rows,
                    "header_row": 1,
                    "source_page": page_idx,
                    "table_index": table_idx,
                    "extractor": "pymupdf",
                })
                if len(sheets) >= max_tables:
                    return sheets, scanned_pages
    finally:
        doc.close()
    return sheets, scanned_pages


def _extract_tables_pdfplumber(file_path: str, max_pages: int, max_tables: int) -> list:
    try:
        import pdfplumber
    except ImportError:
        return []

    sheets = []
    with pdfplumber.open(file_path) as pdf:
        for page_idx, page in enumerate(pdf.pages[:max_pages], 1):
            try:
                raw_tables = page.extract_tables() or []
            except Exception as e:
                logger.debug("[PDF_TABLE] pdfplumber page %s failed: %s", page_idx, e)
                continue
            for table_idx, raw_table in enumerate(raw_tables, 1):
                headers, rows = _clean_pdf_table(raw_table)
                if not _table_is_usable(headers, rows):
                    continue
                sheets.append({
                    "sheet_name": f"page_{page_idx}_table_{table_idx}",
                    "headers": headers,
                    "rows": rows,
                    "header_row": 1,
                    "source_page": page_idx,
                    "table_index": table_idx,
                    "extractor": "pdfplumber",
                })
                if len(sheets) >= max_tables:
                    return sheets
    return sheets


def read_pdf_tables(file_path: str) -> dict:
    """Извлекает таблицы из PDF: PyMuPDF first, pdfplumber fallback."""
    max_pages = _pdf_max_pages()
    max_tables = _pdf_max_tables()
    t0 = time.time()
    sheets, scanned_pages = _extract_tables_pymupdf(file_path, max_pages, max_tables)
    extractor = "pymupdf"

    if not sheets:
        fallback = _extract_tables_pdfplumber(file_path, max_pages, max_tables)
        if fallback:
            sheets = fallback
            extractor = "pdfplumber"

    return {
        "sheets": sheets,
        "needs_ocr": bool(scanned_pages) and not sheets,
        "scanned_pages": sorted(scanned_pages),
        "extractor": extractor if sheets else "",
        "elapsed_sec": round(time.time() - t0, 3),
    }


def _docx_max_tables() -> int:
    try:
        return max(1, int(os.getenv("DOCX_TABLE_MAX_TABLES", "80")))
    except ValueError:
        return 80


def _docx_cell_text(cell: ElementTree.Element, ns: dict[str, str]) -> str:
    paragraphs: list[str] = []
    for paragraph in cell.findall(".//w:p", ns):
        parts = [text.text or "" for text in paragraph.findall(".//w:t", ns)]
        if parts:
            paragraphs.append("".join(parts))
    return _WS_RE.sub(" ", " ".join(part.strip() for part in paragraphs if part.strip())).strip()


def read_docx_tables(file_path: str) -> dict:
    """Извлекает обычные DOCX tables без внешних зависимостей."""
    t0 = time.time()
    max_tables = _docx_max_tables()
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    sheets: list[dict] = []
    try:
        with zipfile.ZipFile(file_path) as archive:
            xml = archive.read("word/document.xml")
    except Exception as error:
        logger.warning("[DOCX_TABLE] %s: cannot read document.xml: %s", Path(file_path).name, error)
        return {"sheets": [], "needs_ocr": False, "scanned_pages": [], "extractor": "", "elapsed_sec": round(time.time() - t0, 3)}

    root = ElementTree.fromstring(xml)
    for table_idx, table in enumerate(root.findall(".//w:tbl", ns), 1):
        raw_rows: list[list[str]] = []
        for tr in table.findall("./w:tr", ns):
            cells = [_docx_cell_text(tc, ns) for tc in tr.findall("./w:tc", ns)]
            if any(cells):
                raw_rows.append(cells)
        headers, rows = _clean_pdf_table(raw_rows)
        if not _table_is_usable(headers, rows):
            continue
        sheets.append({
            "sheet_name": f"docx_table_{table_idx}",
            "headers": headers,
            "rows": rows,
            "header_row": 1,
            "table_index": table_idx,
            "extractor": "docx_xml",
        })
        if len(sheets) >= max_tables:
            break

    return {
        "sheets": sheets,
        "needs_ocr": False,
        "scanned_pages": [],
        "extractor": "docx_xml" if sheets else "",
        "elapsed_sec": round(time.time() - t0, 3),
    }


# ─────────────────────────────────────────
# НОРМАЛИЗАТОР СТРОК
# ─────────────────────────────────────────

def _safe_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        s = str(v).replace(",", ".").replace(" ", "").replace("\xa0", "")
        return float(s)
    except Exception:
        return None


def normalize_row(raw_row: dict, mapping: dict, doc_type: str, doc_title: str, source_file: str) -> dict:
    """Нормализует одну строку Excel в стандартную схему."""
    norm = dict(STANDARD_SCHEMA)
    norm["doc_type"] = doc_type
    norm["doc_title"] = doc_title
    norm["source_file"] = source_file
    norm["raw_row"] = json.dumps(raw_row, ensure_ascii=False, default=str)
    for extra_key in ("source_page", "table_index", "needs_ocr", "extractor"):
        if extra_key in raw_row:
            norm[extra_key] = raw_row[extra_key]

    for orig_col, std_field in mapping.items():
        if std_field == "skip" or std_field not in norm:
            continue
        val = raw_row.get(orig_col)
        if val is None:
            continue

        # Числовые поля
        if std_field in ("qty", "qty_per_unit", "price", "amount", "amount_mat",
                         "amount_work", "work_volume", "work_done", "work_since_start",
                         "weight_unit", "weight_total"):
            norm[std_field] = _safe_float(val)
        else:
            norm[std_field] = str(val).strip() if val is not None else ""

    return norm


# ─────────────────────────────────────────
# КОНВЕРТЕР В ЧАНКИ ДЛЯ QDRANT
# ─────────────────────────────────────────

def row_to_chunk_text(row: dict) -> str:
    """
    Формирует текст чанка из нормализованной строки.
    Текст должен быть богатым — эмбеддинг по нему строится.
    """
    parts = []

    if row.get("doc_type") and row.get("doc_title"):
        parts.append(f"Документ: {DOC_TYPES.get(row['doc_type'], row['doc_type'])} — {row['doc_title']}")

    if row.get("section"):
        parts.append(f"Раздел: {row['section']}")

    if row.get("pos") or row.get("position"):
        parts.append(f"Позиция: {row.get('pos') or row.get('position')}")

    if row.get("name") or row.get("work_name"):
        parts.append(f"Наименование: {row.get('name') or row.get('work_name')}")

    if row.get("code") or row.get("designation"):
        parts.append(f"Обозначение/Артикул: {row.get('code') or row.get('designation')}")

    if row.get("mark"):
        parts.append(f"Марка/Тип: {row['mark']}")

    if row.get("unit") and row.get("qty") is not None:
        parts.append(f"Количество: {row['qty']} {row['unit']}")

    if row.get("price") is not None:
        parts.append(f"Цена: {row['price']:.2f}")

    if row.get("amount") is not None:
        parts.append(f"Сумма: {row['amount']:.2f}")

    if row.get("work_done") is not None:
        parts.append(f"Выполнено: {row['work_done']}")

    if row.get("norms_refs"):
        parts.append(f"Нормативы: {row['norms_refs']}")

    if row.get("materials_used"):
        parts.append(f"Материалы: {row['materials_used']}")

    if row.get("note"):
        parts.append(f"Примечание: {row['note']}")

    if len(parts) <= 1 and row.get("raw_row"):
        try:
            raw = json.loads(row["raw_row"])
        except Exception:
            raw = {}
        raw_parts = []
        for key, value in raw.items():
            if value is None or str(value).strip() == "":
                continue
            raw_parts.append(f"{key}: {str(value).strip()}")
            if len("; ".join(raw_parts)) > 1200:
                break
        if raw_parts:
            parts.append("Строка: " + "; ".join(raw_parts))

    return "\n".join(parts) if parts else json.dumps(row, ensure_ascii=False)


def rows_to_qdrant_chunks(rows: list, dataset_id: str = "") -> list:
    """
    Конвертирует нормализованные строки в чанки для Qdrant.
    Возвращает list[dict] с полями "text" и "metadata".
    """
    chunks = []
    for row in rows:
        text = row_to_chunk_text(row)
        if not text.strip():
            continue

        # Payload для фильтрации в Qdrant
        metadata = {
            "type": "table_row",
            "doc_type": row.get("doc_type", ""),
            "doc_title": row.get("doc_title", ""),
            "source_file": row.get("source_file", ""),
            "source_page": row.get("source_page"),
            "table_index": row.get("table_index"),
            "needs_ocr": row.get("needs_ocr", False),
            "extractor": row.get("extractor", ""),
            "section": row.get("section", ""),
            "name": row.get("name", "") or row.get("work_name", ""),
            "code": row.get("code", "") or row.get("designation", ""),
            "unit": row.get("unit", ""),
            "qty": row.get("qty"),
            "amount": row.get("amount"),
            "dataset_id": dataset_id,
        }

        chunks.append({
            "text": text,
            "metadata": metadata,
            "raw": row,  # полная нормализованная строка
        })
    return chunks


# ─────────────────────────────────────────
# СОХРАНЕНИЕ В PARQUET
# ─────────────────────────────────────────

def save_parquet(rows: list, output_path: str) -> int:
    """
    Сохраняет нормализованные строки в Parquet.
    Возвращает количество записей.
    """
    try:
        import pandas as pd
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ImportError:
        raise RuntimeError(
            "Установи: pip install pandas pyarrow --break-system-packages"
        )

    if not rows:
        logger.warning("[PARQUET] Нет строк для сохранения")
        return 0

    df = pd.DataFrame(rows)

    # Приводим числовые колонки
    numeric_cols = ["qty", "qty_per_unit", "price", "amount", "amount_mat",
                    "amount_work", "work_volume", "work_done", "work_since_start",
                    "weight_unit", "weight_total"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    table = pa.Table.from_pandas(df)
    pq.write_table(table, output_path, compression="snappy")
    logger.info(f"[PARQUET] Сохранено {len(rows)} строк → {output_path}")
    return len(rows)


def load_parquet(parquet_path: str) -> list:
    """Загружает Parquet → list[dict]."""
    try:
        import pyarrow.parquet as pq
        table = pq.read_table(parquet_path)
        return table.to_pydict()
    except Exception as e:
        logger.error(f"[PARQUET] Ошибка чтения {parquet_path}: {e}")
        return []


# ─────────────────────────────────────────
# ГЛАВНЫЙ ПАЙПЛАЙН
# ─────────────────────────────────────────

class TableNormalizer:
    """
    Полный пайплайн: XLSX/CSV → нормализация → Parquet + Qdrant чанки.

    Пример:
        norm = TableNormalizer(llm_url="http://localhost:8050/api/chat")
        result = await norm.process("smeta.xlsx", dataset_id="uuid-xxx")
        # result: {"chunks": [...], "parquet_path": "...", "rows": N, "doc_type": "SMETA"}
    """

    def __init__(
        self,
        llm_url: str = "http://localhost:8050/api/chat",
        parquet_dir: str = "/tmp/ejik_parquet",
        use_llm: bool = True,
    ):
        self.llm_url = llm_url
        self.parquet_dir = Path(parquet_dir)
        self.parquet_dir.mkdir(parents=True, exist_ok=True)
        self.use_llm = use_llm

    async def process(self, file_path: str, dataset_id: str = "", doc_type_override: str | None = None) -> dict:
        """
        Обрабатывает один файл XLSX/CSV.
        Возвращает {"chunks": list, "parquet_path": str, "rows": int, "doc_type": str, "sheets": int}
        """
        fpath = Path(file_path)
        ext = fpath.suffix.lower()

        pdf_meta = {}
        if ext in (".xlsx", ".xls"):
            sheets = read_xlsx_sheets(str(fpath))
        elif ext == ".csv":
            sheets = read_csv(str(fpath))
        elif ext == ".pdf":
            pdf_meta = read_pdf_tables(str(fpath))
            sheets = pdf_meta["sheets"]
        elif ext == ".docx":
            pdf_meta = read_docx_tables(str(fpath))
            sheets = pdf_meta["sheets"]
        else:
            raise ValueError(f"Неподдерживаемый формат: {ext}")

        if not sheets:
            if pdf_meta.get("needs_ocr"):
                logger.warning(f"[PARQUET] {fpath.name}: скан без текстового слоя, нужна OCR/VLM очередь")
            else:
                logger.warning(f"[PARQUET] Нет данных в {fpath.name}")
            return {
                "chunks": [],
                "parquet_path": "",
                "rows": 0,
                "doc_type": "TABLE",
                "sheets": 0,
                "needs_ocr": bool(pdf_meta.get("needs_ocr")),
                "scanned_pages": pdf_meta.get("scanned_pages", []),
            }

        all_normalized = []
        all_chunks = []
        doc_type_final = "TABLE"

        for sheet in sheets:
            headers = sheet["headers"]
            rows = sheet["rows"]
            sheet_name = sheet["sheet_name"]

            if not rows:
                continue

            # Маппинг колонок
            if self.use_llm:
                mapping_result = await _llm_map_columns(headers, rows[:3], self.llm_url)
            else:
                # Простой детект без LLM
                doc_type = _detect_doc_type_simple(headers, sheet_name)
                mapping_result = {
                    "doc_type": doc_type,
                    "mapping": _map_columns_simple(headers),
                }

            doc_type = doc_type_override or mapping_result.get("doc_type", "TABLE")
            mapping = mapping_result.get("mapping", {})
            doc_type_final = doc_type

            logger.info(f"[PARQUET] {fpath.name} / {sheet_name}: тип={doc_type}, строк={len(rows)}")

            # Нормализация строк
            for raw_row in rows:
                raw_row = dict(raw_row)
                for extra_key in ("source_page", "table_index", "extractor"):
                    if extra_key in sheet:
                        raw_row[extra_key] = sheet[extra_key]
                norm = normalize_row(
                    raw_row=raw_row,
                    mapping=mapping,
                    doc_type=doc_type,
                    doc_title=f"{fpath.stem} / {sheet_name}",
                    source_file=fpath.name,
                )
                all_normalized.append(norm)

        # Чанки для Qdrant
        all_chunks = rows_to_qdrant_chunks(all_normalized, dataset_id=dataset_id)

        # Сохраняем Parquet
        parquet_path = ""
        if all_normalized:
            parquet_name = f"{fpath.stem}.parquet"
            parquet_path = str(self.parquet_dir / parquet_name)
            save_parquet(all_normalized, parquet_path)

        return {
            "chunks": all_chunks,
            "parquet_path": parquet_path,
            "rows": len(all_normalized),
            "doc_type": doc_type_final,
            "sheets": len(sheets),
            "needs_ocr": bool(pdf_meta.get("needs_ocr")),
            "scanned_pages": pdf_meta.get("scanned_pages", []),
        }


# ─────────────────────────────────────────
# CLI: тест вручную
# ─────────────────────────────────────────
if __name__ == "__main__":
    import argparse
    import asyncio

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    parser = argparse.ArgumentParser(description="Е.Ж.И.К. Parquet Normalizer")
    parser.add_argument("file", help="XLSX или CSV файл")
    parser.add_argument("--out", default="/tmp/parquet_out", help="Папка для Parquet")
    parser.add_argument("--no-llm", action="store_true", help="Без LLM маппинга")
    parser.add_argument("--llm", default="http://localhost:8050/api/chat", help="URL LLM")
    args = parser.parse_args()

    norm = TableNormalizer(
        llm_url=args.llm,
        parquet_dir=args.out,
        use_llm=not args.no_llm,
    )

    result = asyncio.run(norm.process(args.file))
    print(f"\n[ИТОГ]")
    print(f"  Тип документа: {result['doc_type']} ({DOC_TYPES.get(result['doc_type'], '?')})")
    print(f"  Листов:        {result['sheets']}")
    print(f"  Строк:         {result['rows']}")
    print(f"  Чанков:        {len(result['chunks'])}")
    print(f"  Parquet:       {result['parquet_path']}")

    if result["chunks"]:
        print(f"\n[ПЕРВЫЙ ЧАНК]")
        print(result["chunks"][0]["text"])

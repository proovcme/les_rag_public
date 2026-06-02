"""Fast deterministic document routing for ingestion."""

from __future__ import annotations

import csv
import os
import re
import zipfile
from dataclasses import dataclass, field
from xml.etree import ElementTree
from pathlib import Path
from typing import Any


TABLE_SUFFIXES = {".xlsx", ".xls", ".csv"}
PDF_SUFFIXES = {".pdf"}
EMAIL_SUFFIXES = {".eml", ".emlx", ".msg"}
CAD_BIM_SUFFIXES = {".dwg", ".rvt", ".ifc", ".ifczip"}


@dataclass
class DocumentProbe:
    path: Path
    suffix: str
    size_bytes: int
    page_count: int = 0
    text_sample: str = ""
    has_text_layer: bool = True
    has_tables: bool = False
    table_count_hint: int = 0
    sheet_count: int = 0
    row_count_hint: int = 0
    column_count_hint: int = 0
    needs_ocr: bool = False
    signals: dict[str, Any] = field(default_factory=dict)


@dataclass
class DocumentRoute:
    domain: str
    dataset_name: str
    doc_type: str
    content_type: str
    complexity: str
    pipeline: str
    metadata: dict[str, Any]


def _sample_limit() -> int:
    try:
        return max(1, int(os.getenv("DOC_ROUTER_SAMPLE_PAGES", "3")))
    except ValueError:
        return 3


def probe_document(path: Path) -> DocumentProbe:
    suffix = path.suffix.lower()
    stat = path.stat()
    if suffix in PDF_SUFFIXES:
        return _probe_pdf(path, stat.st_size)
    if suffix in TABLE_SUFFIXES:
        return _probe_table(path, stat.st_size)
    if suffix == ".docx":
        return _probe_docx(path, stat.st_size)
    return _probe_text_like(path, stat.st_size)


def route_document(path: Path) -> DocumentRoute:
    return classify_document(probe_document(path))


def classify_document(probe: DocumentProbe) -> DocumentRoute:
    doc_type = _classify_doc_type(probe)
    domain = _classify_domain(probe, doc_type)
    content_type = _classify_content_type(probe)
    complexity = _classify_complexity(probe, content_type)
    pipeline = _select_pipeline(probe, content_type, complexity)
    dataset_name = f"{domain}_Index"
    return DocumentRoute(
        domain=domain,
        dataset_name=dataset_name,
        doc_type=doc_type,
        content_type=content_type,
        complexity=complexity,
        pipeline=pipeline,
        metadata={
            "domain": domain,
            "dataset_name": dataset_name,
            "doc_type": doc_type,
            "content_type": content_type,
            "complexity": complexity,
            "pipeline": pipeline,
            "has_tables": probe.has_tables,
            "needs_ocr": probe.needs_ocr,
            "page_count": probe.page_count,
            "sheet_count": probe.sheet_count,
            "row_count_hint": probe.row_count_hint,
            "column_count_hint": probe.column_count_hint,
        },
    )


def _probe_pdf(path: Path, size_bytes: int) -> DocumentProbe:
    probe = DocumentProbe(path=path, suffix=".pdf", size_bytes=size_bytes)
    try:
        import fitz

        doc = fitz.open(path)
        try:
            probe.page_count = len(doc)
            sample_text = []
            text_pages = 0
            table_count = 0
            for page_no in range(min(_sample_limit(), len(doc))):
                page = doc[page_no]
                text = page.get_text("text") or ""
                if text.strip():
                    text_pages += 1
                    sample_text.append(text[:2000])
                finder = getattr(page, "find_tables", None)
                if finder:
                    try:
                        table_count += len(getattr(finder(), "tables", []) or [])
                    except Exception:
                        pass
            probe.text_sample = "\n".join(sample_text)[:6000]
            probe.has_text_layer = text_pages > 0
            probe.needs_ocr = not probe.has_text_layer and probe.page_count > 0
            probe.has_tables = table_count > 0 or _text_has_table_signals(probe.text_sample)
            probe.table_count_hint = table_count
        finally:
            doc.close()
    except Exception as e:
        probe.signals["probe_error"] = str(e)
    return probe


def _probe_table(path: Path, size_bytes: int) -> DocumentProbe:
    probe = DocumentProbe(path=path, suffix=path.suffix.lower(), size_bytes=size_bytes, has_tables=True)
    try:
        if probe.suffix == ".csv":
            with open(path, encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                rows = []
                for idx, row in enumerate(reader):
                    rows.append(row)
                    if idx >= 20:
                        break
            probe.sheet_count = 1
            probe.row_count_hint = max(0, len(rows) - 1)
            probe.column_count_hint = max((len(row) for row in rows), default=0)
            probe.text_sample = "\n".join(",".join(row) for row in rows[:5])
        else:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                probe.sheet_count = len(wb.sheetnames)
                samples = []
                rows_total = 0
                max_cols = 0
                for sheet_name in wb.sheetnames[:3]:
                    ws = wb[sheet_name]
                    rows_total += ws.max_row or 0
                    max_cols = max(max_cols, ws.max_column or 0)
                    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 5), values_only=True):
                        samples.append(" | ".join("" if v is None else str(v) for v in row))
                probe.row_count_hint = rows_total
                probe.column_count_hint = max_cols
                probe.text_sample = "\n".join(samples)[:6000]
            finally:
                wb.close()
    except UnicodeDecodeError:
        try:
            with open(path, encoding="cp1251", newline="") as f:
                reader = csv.reader(f)
                rows = [row for _, row in zip(range(20), reader)]
            probe.sheet_count = 1
            probe.row_count_hint = max(0, len(rows) - 1)
            probe.column_count_hint = max((len(row) for row in rows), default=0)
            probe.text_sample = "\n".join(",".join(row) for row in rows[:5])
        except Exception as e:
            probe.signals["probe_error"] = str(e)
    except Exception as e:
        probe.signals["probe_error"] = str(e)
    return probe


def _probe_docx(path: Path, size_bytes: int) -> DocumentProbe:
    probe = DocumentProbe(path=path, suffix=".docx", size_bytes=size_bytes)
    try:
        with zipfile.ZipFile(path) as docx:
            xml_names = [
                name
                for name in docx.namelist()
                if name == "word/document.xml"
                or (name.startswith("word/header") and name.endswith(".xml"))
            ]
            samples = []
            for name in xml_names:
                raw_xml = docx.read(name)
                if name == "word/document.xml":
                    probe.table_count_hint = len(re.findall(rb"<w:tbl(?:\s|>)", raw_xml))
                root = ElementTree.fromstring(raw_xml)
                for node in root.iter():
                    if node.tag.endswith("}t") and node.text:
                        samples.append(node.text)
                        if sum(len(item) for item in samples) >= 6000:
                            break
                if sum(len(item) for item in samples) >= 6000:
                    break
            probe.text_sample = " ".join(samples)[:6000]
            probe.has_text_layer = bool(probe.text_sample.strip())
            probe.has_tables = probe.table_count_hint > 0 or _text_has_table_signals(probe.text_sample)
    except Exception as e:
        probe.signals["probe_error"] = str(e)
        return _probe_text_like(path, size_bytes)
    return probe


def _probe_text_like(path: Path, size_bytes: int) -> DocumentProbe:
    probe = DocumentProbe(path=path, suffix=path.suffix.lower(), size_bytes=size_bytes)
    try:
        probe.text_sample = path.read_text(encoding="utf-8", errors="ignore")[:6000]
        probe.has_tables = _text_has_table_signals(probe.text_sample)
    except Exception as e:
        probe.signals["probe_error"] = str(e)
    return probe


def _classify_doc_type(probe: DocumentProbe) -> str:
    text = f"{probe.path.name}\n{probe.text_sample}".lower()
    name = probe.path.name.lower()
    if _is_cad_bim_source(probe):
        return "CAD_BIM"
    if probe.suffix in EMAIL_SUFFIXES:
        return "EMAIL"
    if _looks_like_book(probe):
        return "BOOK"
    normative_name_prefixes = ("гост", "сп ", "снип", "санпин", "постановление", "приказ")
    if name.startswith(normative_name_prefixes):
        return "NORMATIVE"
    if probe.suffix not in TABLE_SUFFIXES and _has_strong_normative_signal(text):
        return "NORMATIVE"
    has_price_amount = any(token in text for token in ("цена", "сумма", "стоимость", "расценка"))
    has_position_qty = (
        any(token in text for token in ("позиция", "поз.", "поз,", "поз "))
        and any(token in text for token in ("кол-во", "количество", "ед.изм", "единица"))
    )
    if any(token in text for token in ("кс-2", "кс2", "акт о приемке", "акт о приёмке")):
        return "KS2"
    if any(token in text for token in ("аоср", "скрытых работ", "освидетельствования")):
        return "AOSR"
    if any(token in text for token in ("постановление", "федеральный закон", "приказ росстандарта", "свод правил")):
        return "NORMATIVE"
    if any(token in text for token in ("смета", "локальный сметный", "гэсн", "фер", "тер", "расценка")):
        return "SMETA"
    if has_position_qty and not has_price_amount:
        return "SPEC"
    if any(token in text for token in ("спецификация", "ведомость оборудования", "масса единицы")):
        return "SPEC"
    if has_price_amount and probe.has_tables:
        return "SMETA"
    if any(token in text for token in ("гост", "сп ", "снип", "санпин", "норматив", "постановление")):
        return "NORMATIVE"
    if probe.suffix in TABLE_SUFFIXES:
        return "TABLE"
    return "DOCUMENT"


def _has_strong_normative_signal(text: str) -> bool:
    if any(token in text for token in ("национальный стандарт", "межгосударственный стандарт", "свод правил")):
        return True
    return bool(re.search(r"\b(гост|гост\s*р|сп|снип|санпин)\s*(?:iec|iso|р)?\s*\d", text))


def _classify_domain(probe: DocumentProbe, doc_type: str) -> str:
    text = f"{' '.join(probe.path.parts)}\n{probe.text_sample}".casefold()
    name = probe.path.name.casefold()

    if doc_type == "CAD_BIM" or _is_cad_bim_source(probe):
        return "CAD_BIM"

    if doc_type == "EMAIL" or probe.suffix in EMAIL_SUFFIXES:
        return "MAIL"
    if doc_type == "BOOK" or _looks_like_book(probe):
        return "BOOKS"

    if any(token in name for token in ("гкрф", "градостроительный кодекс", "постановление 87", "пп 87", "pp87")):
        return "GKRF"
    if "постановление 87" in text and "градостро" in text:
        return "GKRF"
    if (
        ("постановление 87" in text or "пп 87" in text)
        and "состав" in text
        and "раздел" in text
        and "проектн" in text
    ):
        return "GKRF"
    if "87" in name and "постановлен" in name:
        return "GKRF"

    if doc_type in {"KS2", "AOSR", "SMETA", "SPEC", "TABLE"}:
        return f"TABLE_{doc_type}"

    if _is_industrial_chimney_norm(text, name):
        return "NTD_STRUCTURAL"

    if _has_any(name, _FIRE_TOKENS):
        return "NTD_FIRE"
    if _has_any(name, _ELECTRICAL_TOKENS):
        return "NTD_ELECTRICAL"
    if _is_spds_norm(name, text):
        return "NTD_SPDS"
    if _has_any(name, _GEOTECH_TOKENS):
        return "NTD_GEOTECH"
    if _has_any(name, _TRANSPORT_TOKENS):
        return "NTD_TRANSPORT"
    if _has_any(name, _HVAC_TOKENS):
        return "NTD_HVAC"
    if _has_any(name, _WATER_TOKENS):
        return "NTD_WATER"
    if _has_any(name, _PIPELINE_TOKENS):
        return "NTD_PIPELINES"
    if _has_any(name, _BIM_OPERATION_TOKENS):
        return "NTD_BIM_OPERATION"
    if _has_any(name, _CONSTRUCTION_TOKENS):
        return "NTD_CONSTRUCTION"
    if _has_any(name, _MATERIALS_TOKENS):
        return "NTD_MATERIALS"
    if _has_any(name, _ARCH_URBAN_TOKENS):
        return "NTD_ARCH_URBAN"
    if _has_any(name, _SAFETY_TOKENS):
        return "NTD_SAFETY"
    if _has_any(name, _STRUCTURAL_TOKENS):
        return "NTD_STRUCTURAL"

    if _has_any(text, _FIRE_TEXT_TOKENS):
        return "NTD_FIRE"
    if _has_any(text, _ELECTRICAL_TEXT_TOKENS):
        return "NTD_ELECTRICAL"
    if _has_any(text, _GEOTECH_TEXT_TOKENS):
        return "NTD_GEOTECH"
    if _has_any(text, _TRANSPORT_TEXT_TOKENS):
        return "NTD_TRANSPORT"
    if _has_any(text, _HVAC_TEXT_TOKENS):
        return "NTD_HVAC"
    if _has_any(text, _WATER_TEXT_TOKENS):
        return "NTD_WATER"
    if _has_any(text, _PIPELINE_TEXT_TOKENS):
        return "NTD_PIPELINES"
    if _has_any(text, _BIM_OPERATION_TEXT_TOKENS):
        return "NTD_BIM_OPERATION"
    if _has_any(text, _CONSTRUCTION_TEXT_TOKENS):
        return "NTD_CONSTRUCTION"
    if _has_any(text, _MATERIALS_TEXT_TOKENS):
        return "NTD_MATERIALS"
    if _has_any(text, _ARCH_URBAN_TEXT_TOKENS):
        return "NTD_ARCH_URBAN"
    if _has_any(text, _SAFETY_TEXT_TOKENS):
        return "NTD_SAFETY"
    if _has_any(text, _STRUCTURAL_TEXT_TOKENS):
        return "NTD_STRUCTURAL"

    # Backward-compatible broad buckets kept for older abbreviated filenames.
    if any(
        token in name
        for token in (
            "13130",
            "пожар",
            "пожаротуш",
            "огнев",
            "огнестойк",
            "огнезащит",
            "огнепреград",
            "эвакуац",
            "эвакуа",
            "противодым",
            "противопожар",
            "пожарной безопасности",
            "дымоудален",
        )
    ):
        return "NTD_FIRE"
    if any(
        token in name
        for token in (
            "пуэ",
            "iec",
            "мэк",
            "электр",
            "кабел",
            "заземл",
            "молниезащит",
            "освещен",
            "напряжен",
            "светиль",
            "выключател",
            "предохранител",
            "низковоль",
            "электроустанов",
        )
    ):
        return "NTD_ELECTRICAL"
    if any(
        token in name
        for token in (
            "конструкц",
            "нагрузк",
            "фундамент",
            "основан",
            "железобетон",
            "бетон",
            "грунт",
            "здани",
            "сооруж",
            "сейсми",
        )
    ):
        return "NTD_STRUCTURAL"
    if doc_type == "NORMATIVE" and (
        "электроустановки" in text
        or "пожарной безопасности" in text
        or ("пожарн" in text and "безопас" in text)
    ):
        if "пожарной безопасности" in text or ("пожарн" in text and "безопас" in text):
            return "NTD_FIRE"
        return "NTD_ELECTRICAL"
    if doc_type == "NORMATIVE":
        return "NTD_GENERAL"
    if _is_ntd_source(probe):
        return "NTD_GENERAL"
    return "DOCS_OTHER"


_FIRE_TOKENS = (
    "13130",
    "59637",
    "59638",
    "59639",
    "59640",
    "пожар",
    "пожаротуш",
    "огнев",
    "огнестойк",
    "огнезащит",
    "огнепреград",
    "эвакуац",
    "эвакуа",
    "противодым",
    "противопожар",
    "дымоудален",
    "взрывопожар",
    "огн.",
    "горюч",
    "воспламен",
    "пенного пожаротушения",
)
_FIRE_TEXT_TOKENS = (
    "пожарной безопасности",
    "требования пожарной безопасности",
    "огнестойкости",
    "пожарная опасность",
)

_ELECTRICAL_TOKENS = (
    "пуэ",
    "iec",
    "мэк",
    "электр",
    "кабел",
    "заземл",
    "молниезащит",
    "освещен",
    "напряжен",
    "светиль",
    "выключател",
    "предохранител",
    "низковоль",
    "электроустанов",
    "60364",
    "50571",
    "30331",
    "60079",
    "31610",
    "60968",
    "61008",
)
_ELECTRICAL_TEXT_TOKENS = ("электроустановки", "электрические сети", "электроснабжение")

_SPDS_TOKENS = (
    "гост 21.",
    "гост р 21.",
    "спдс",
    "система проектной документации",
    "проектной документации для строитель",
    "рабочая документация",
)

_GEOTECH_TOKENS = (
    "грунт",
    "геотехник",
    "основан",
    "фундамент",
    "сейсми",
    "землетряс",
    "оползн",
    "карст",
    "мерзлот",
    "подпорн",
    "геофизик",
)
_GEOTECH_TEXT_TOKENS = ("механика грунтов", "основания зданий", "основания и фундаменты")

_TRANSPORT_TOKENS = (
    "дорог",
    "мост",
    "тоннел",
    "метрополитен",
    "железн",
    "аэродром",
    "улиц",
    "транспорт",
    "габарит",
    "путепровод",
    "биопереход",
)
_TRANSPORT_TEXT_TOKENS = ("автомобильные дороги", "железные дороги", "мосты и трубы")

_HVAC_TOKENS = (
    "отоп",
    "вентиля",
    "кондицион",
    "теплов",
    "теплоснаб",
    "воздух",
    "дымоудален",
    "шум",
    "акуст",
    "микроклимат",
)
_HVAC_TEXT_TOKENS = ("отопление вентиляция", "тепловые сети", "защита от шума")

_WATER_TOKENS = (
    "водоснаб",
    "водоотвед",
    "канализац",
    "гидротех",
    "мелиоратив",
    "водопропуск",
    "водоочист",
    "очистн",
    "морские причаль",
    "гидроаэродром",
)
_WATER_TEXT_TOKENS = ("системы водоснабжения", "гидротехнические сооружения")

_PIPELINE_TOKENS = (
    "трубопровод",
    "промыслов",
    "магистральн",
    "газопровод",
    "нефтепровод",
    "морские трубопроводы",
)
_PIPELINE_TEXT_TOKENS = ("магистральные трубопроводы", "промысловые трубопроводы")

_BIM_OPERATION_TOKENS = (
    "информационное моделирован",
    "bim",
    "обследован",
    "мониторинг",
    "эксплуатац",
    "техническ",
    "технич",
    "надзор",
)
_BIM_OPERATION_TEXT_TOKENS = ("информационная модель", "техническое состояние")

_CONSTRUCTION_TOKENS = (
    "организация строительства",
    "производства работ",
    "приемк",
    "приёмк",
    "земляные работы",
    "изоляционные и отделочные",
    "механизация строительства",
    "свароч",
    "снип iii",
    "iii-",
)
_CONSTRUCTION_TEXT_TOKENS = ("правила производства и приемки", "организация строительного производства")

_MATERIALS_TOKENS = (
    "материал",
    "издел",
    "изоляц",
    "опалуб",
    "полы",
    "стены",
    "покрыт",
    "пластмасс",
    "ограждающ",
    "панел",
    "кровл",
    "тепловая изоля",
)
_MATERIALS_TEXT_TOKENS = ("материалы строительные", "строительные материалы")

_ARCH_URBAN_TOKENS = (
    "жил",
    "обществен",
    "градостро",
    "планировк",
    "территор",
    "доступность",
    "учрежден",
    "образователь",
    "детск",
    "больниц",
    "спорт",
    "парк",
    "общежит",
    "полици",
    "наемные дома",
    "малоэтаж",
    "высотн",
)
_ARCH_URBAN_TEXT_TOKENS = ("жилые здания", "общественные здания", "городская среда")

_SAFETY_TOKENS = (
    "12.",
    "ссбт",
    "безопасност",
    "охрана труда",
    "опасн",
    "защитные сооружения",
    "гражданск",
    "аварийн",
    "химическ",
)
_SAFETY_TEXT_TOKENS = ("система стандартов безопасности труда", "защитные сооружения гражданской обороны")

_STRUCTURAL_TOKENS = (
    "конструкц",
    "нагрузк",
    "железобетон",
    "бетон",
    "стальные конструкции",
    "каменные конструкции",
    "деревянн",
    "сооруж",
    "резервуар",
    "силос",
    "дымовые трубы",
)
_STRUCTURAL_TEXT_TOKENS = ("строительные конструкции", "несущие конструкции")


def _has_any(haystack: str, tokens: tuple[str, ...]) -> bool:
    return any(token in haystack for token in tokens)


def _is_spds_norm(name: str, text: str) -> bool:
    if _has_any(name, ("гост 21.", "гост р 21.", "спдс")):
        return True
    if _has_any(name, ("система проектной документации", "проектной документации для строитель")):
        return True
    return "гост 21" in text and _has_any(text, _SPDS_TOKENS)


def _is_ntd_source(probe: DocumentProbe) -> bool:
    return any(part.casefold() == "ntd" for part in probe.path.parts)


def _is_books_source(probe: DocumentProbe) -> bool:
    return any(part.casefold() == "books" for part in probe.path.parts)


def _looks_like_book(probe: DocumentProbe) -> bool:
    name = probe.path.name.casefold()
    return _is_books_source(probe) or (
        probe.suffix in PDF_SUFFIXES
        and probe.page_count >= 200
        and any(token in name for token in ("рук-во", "руководство", "пособие", "справочник", "учебник", "book"))
    )


def _is_industrial_chimney_norm(text: str, name: str) -> bool:
    haystack = f"{name}\n{text}"
    chimney_phrases = (
        "трубы промышленные дымовые",
        "промышленные дымовые трубы",
        "дымовые промышленные трубы",
        "дымовая промышленная труба",
        "дымовых промышленных труб",
        "smoke stack",
        "smokestack",
        "industrial chimney",
    )
    if any(phrase in haystack for phrase in chimney_phrases):
        return True
    return "дымовые" in haystack and "труб" in haystack and "противодым" not in haystack


def _classify_content_type(probe: DocumentProbe) -> str:
    if _is_cad_bim_source(probe):
        return "cad_bim"
    if probe.suffix in EMAIL_SUFFIXES:
        return "email"
    if probe.needs_ocr:
        return "scan"
    if probe.suffix in TABLE_SUFFIXES:
        return "table"
    if probe.suffix in PDF_SUFFIXES and _looks_like_book(probe):
        return "mixed"
    if probe.suffix in PDF_SUFFIXES and probe.has_tables:
        return "mixed"
    if probe.suffix == ".docx" and probe.has_tables:
        return "mixed"
    return "text"


def _classify_complexity(probe: DocumentProbe, content_type: str) -> str:
    if probe.needs_ocr:
        return "needs_ocr"
    if probe.size_bytes > 50 * 1024 * 1024 or probe.page_count > 200:
        return "heavy"
    if content_type in ("table", "mixed") or probe.row_count_hint > 2000:
        return "structured"
    return "simple"


def _select_pipeline(probe: DocumentProbe, content_type: str, complexity: str) -> str:
    if complexity == "needs_ocr":
        return "markdown_needs_ocr"
    if content_type == "cad_bim":
        return "json_graph_projection"
    if probe.suffix in TABLE_SUFFIXES:
        return "parquet"
    if probe.suffix in PDF_SUFFIXES and content_type == "mixed":
        return "markdown_pdf_tables"
    return "markdown"


def _is_cad_bim_source(probe: DocumentProbe) -> bool:
    parts = {part.casefold() for part in probe.path.parts}
    return probe.suffix in CAD_BIM_SUFFIXES or "cad_bim" in parts


def _text_has_table_signals(text: str) -> bool:
    lower = text.lower()
    keywords = ("наименование", "кол-во", "количество", "ед.изм", "сумма", "цена", "поз.")
    keyword_hits = sum(1 for keyword in keywords if keyword in lower)
    numeric_lines = sum(1 for line in text.splitlines() if len(re.findall(r"\d+[,.]?\d*", line)) >= 3)
    return keyword_hits >= 2 or numeric_lines >= 3

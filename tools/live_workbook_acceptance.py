"""Fail-closed live acceptance for the ordinary workbook-chat workflow.

This runner is deliberately opt-in. Only the real entrypoint constructs an
HTTP client and writes a ``live_runtime`` receipt; contract exercises are
separate, non-persistent evidence.
"""
from __future__ import annotations

import argparse
import hashlib
from io import BytesIO
import json
import math
import os
import re
import time
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Protocol

import httpx
from openpyxl import load_workbook


UPLOAD_PATH = "/api/chat/attachments"
CHAT_STREAM_PATH = "/api/chat/stream"
ARTIFACT_PATH = "/api/artifacts"
VERSION_PATH = "/api/version"
REPORT_SCHEMA = "les.live_workbook_acceptance.v1"
LIVE_RUNTIME = "live_runtime"
CONTRACT_TEST = "contract_test"
PRESET_IDENTITIES = {
    "qwen-9b": "qwen-9b-restrictive",
    "qwen-35b": "qwen-35b-extended",
}
_REDACTED_KEYS = frozenset({
    "api_key", "apikey", "authorization", "auth_header", "secret", "token",
    "prompt", "prompt_dump", "source_path", "attachment_path", "exception", "traceback",
})
_LOCAL_PATH_RE = re.compile(r"(?:[A-Za-z]:[\\/]|/(?:home|users|private|var|tmp)/)", re.IGNORECASE)
_SECRET_VALUE_RE = re.compile(
    r"(?:bearer\s+\S+|(?:api[_-]?key|authorization|auth|secret|token)\s*[=:])",
    re.IGNORECASE,
)
_UNREDACTED_TEXT_RE = re.compile(r"(?:traceback|most recent call last|prompt(?:\s+dump)?)", re.IGNORECASE)
_RELATIVE_SOURCE_PATH_RE = re.compile(
    r"(?:^|[\s'\"])(?:\.?[\\/])?(?:data|storage|rag_content|tests|attachments|workspace)[\\/]",
    re.IGNORECASE,
)
_FULL_COMMIT_RE = re.compile(r"^[0-9a-f]{40}$", re.IGNORECASE)


class LiveAcceptanceError(ValueError):
    """The live workflow did not provide the required acceptance evidence."""


class HttpResponse(Protocol):
    content: bytes

    def raise_for_status(self) -> None: ...

    def json(self) -> Any: ...

    def iter_lines(self) -> Any: ...


class HttpClient(Protocol):
    def request(self, method: str, url: str, **kwargs: Any) -> HttpResponse: ...


@dataclass(frozen=True)
class AcceptanceConfig:
    attachment: Path
    base_url: str
    profile_revision_id: str
    model_preset: str
    out: Path
    api_key: str | None
    max_elapsed_seconds: float = 1800.0


def _non_empty(value: Any, label: str) -> str:
    result = str(value or "").strip()
    if not result:
        raise LiveAcceptanceError(f"missing {label}")
    return result


def _hex_sha(value: Any, label: str) -> str:
    result = _non_empty(value, label)
    if len(result) != 64 or any(char not in "0123456789abcdef" for char in result.lower()):
        raise LiveAcceptanceError(f"invalid {label}")
    return result.lower()


def _as_list(value: Any, label: str) -> list[Any]:
    if not isinstance(value, list):
        raise LiveAcceptanceError(f"{label} must be an array")
    return value


def _fixture_path(path: Path) -> bool:
    return "tests/fixtures" in path.resolve().as_posix().casefold()


def _base_url(base_url: str) -> str:
    value = _non_empty(base_url, "base URL").rstrip("/")
    if not value.startswith(("http://", "https://")):
        raise LiveAcceptanceError("base URL must use http or https")
    return value


def _url(base_url: str, path: str) -> str:
    return f"{_base_url(base_url)}{path}"


def _headers(config: AcceptanceConfig, *, idempotency: bool = False) -> dict[str, str]:
    headers = {"Accept": "application/json"}
    if config.api_key:
        headers["Authorization"] = f"Bearer {config.api_key}"
    if idempotency:
        headers["Idempotency-Key"] = f"live-workbook-{uuid.uuid4().hex}"
    return headers


def _request(client: HttpClient, method: str, url: str, **kwargs: Any) -> HttpResponse:
    response = client.request(method, url, **kwargs)
    response.raise_for_status()
    return response


def _stream_final(response: HttpResponse) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    event = "message"
    data: list[str] = []
    progress: list[dict[str, Any]] = []
    for raw_line in response.iter_lines():
        line = raw_line.decode("utf-8") if isinstance(raw_line, bytes) else str(raw_line)
        if not line:
            if event == "error":
                raise LiveAcceptanceError("ordinary chat returned an SSE error")
            if event == "final":
                try:
                    payload = json.loads("\n".join(data))
                except json.JSONDecodeError as error:
                    raise LiveAcceptanceError("ordinary chat final event was not JSON") from error
                if not isinstance(payload, dict):
                    raise LiveAcceptanceError("ordinary chat final event was not an object")
                return payload, progress
            if event == "tool_progress":
                try:
                    payload = json.loads("\n".join(data))
                except json.JSONDecodeError as error:
                    raise LiveAcceptanceError("tool progress was not JSON") from error
                if not isinstance(payload, dict):
                    raise LiveAcceptanceError("tool progress was not an object")
                progress.append(payload)
            event, data = "message", []
            continue
        if line.startswith("event:"):
            event = line.split(":", 1)[1].strip()
        elif line.startswith("data:"):
            data.append(line.split(":", 1)[1].strip())
    raise LiveAcceptanceError("ordinary chat stream ended without final event")


def _artifact_from_final(final: Mapping[str, Any], attachment_id: str, source_sha256: str) -> dict[str, Any]:
    artifact = final.get("artifact")
    checkpoint = final.get("checkpoint")
    retry = final.get("attachment_retry")
    source = final.get("source")
    if not isinstance(artifact, dict) or not isinstance(checkpoint, dict):
        raise LiveAcceptanceError("ordinary chat final lacks workbook artifact or checkpoint")
    if not isinstance(retry, dict) or str(retry.get("attachment_id") or "") != attachment_id:
        raise LiveAcceptanceError("ordinary chat did not preserve the uploaded attachment")
    if not isinstance(source, dict) or str(source.get("attachment_id") or "") != attachment_id:
        raise LiveAcceptanceError("workbook source attachment lineage is missing")
    if _hex_sha(source.get("sha256"), "source SHA-256") != source_sha256:
        raise LiveAcceptanceError("workbook source hash differs from uploaded attachment")
    checkpoint_id = _non_empty(checkpoint.get("checkpoint_id"), "checkpoint ID")
    if str(checkpoint.get("status") or "") != "complete":
        raise LiveAcceptanceError("workbook checkpoint is not complete")
    if str(artifact.get("decision_checkpoint_id") or "") != checkpoint_id:
        raise LiveAcceptanceError("artifact checkpoint lineage differs from final checkpoint")
    return dict(artifact)


def _model_identity(final: Mapping[str, Any]) -> str:
    model = final.get("model_connection")
    if not isinstance(model, dict):
        raise LiveAcceptanceError("ordinary chat final lacks observed model identity")
    return _non_empty(model.get("model_id"), "observed model identity")


def _runtime_version(client: HttpClient, config: AcceptanceConfig) -> dict[str, Any]:
    response = _request(client, "GET", _url(config.base_url, VERSION_PATH), headers=_headers(config))
    payload = response.json()
    if not isinstance(payload, dict):
        raise LiveAcceptanceError("runtime version was not an object")
    source_commit_full = str(payload.get("git_commit_full") or "").strip().lower()
    deployed_commit = str(payload.get("deployed_commit") or "").strip().lower()
    if not _FULL_COMMIT_RE.fullmatch(source_commit_full):
        source_commit_full = deployed_commit
    if not _FULL_COMMIT_RE.fullmatch(source_commit_full):
        raise LiveAcceptanceError("runtime full source commit is invalid")
    build_number = payload.get("build_number")
    if isinstance(build_number, bool) or not isinstance(build_number, int) or build_number <= 0:
        raise LiveAcceptanceError("runtime build number is invalid")
    if payload.get("repo_dirty") is not False:
        raise LiveAcceptanceError("runtime repository is dirty")
    alignment = payload.get("runtime_alignment")
    if not isinstance(alignment, dict) or alignment.get("status") != "aligned":
        raise LiveAcceptanceError("runtime alignment is not verified")
    return {
        "source_commit_full": source_commit_full.lower(),
        "build_number": build_number,
        "runtime_alignment": "aligned",
    }


def _validate_downloaded_xlsx(content: bytes) -> dict[str, int]:
    """Reject a hash-valid blob that is not a minimally meaningful workbook."""
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001 - malformed untrusted artifact
        raise LiveAcceptanceError("downloaded artifact is not a readable XLSX") from error
    try:
        visible = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        if not visible:
            raise LiveAcceptanceError("downloaded XLSX has no visible worksheet")
        template_seen = False
        for sheet in visible:
            rows = iter(sheet.iter_rows(values_only=True))
            for row in rows:
                header_cells = sum(
                    cell is not None and (not isinstance(cell, str) or bool(cell.strip()))
                    for cell in row
                )
                if header_cells < 2:
                    continue
                template_seen = True
                data_rows = sum(
                    any(
                        cell is not None and (not isinstance(cell, str) or bool(cell.strip()))
                        for cell in data_row
                    )
                    for data_row in rows
                )
                if data_rows:
                    return {
                        "visible_sheet_count": len(visible),
                        "header_cell_count": header_cells,
                        "data_row_count": data_rows,
                    }
        if template_seen:
            raise LiveAcceptanceError("downloaded XLSX has no populated data row beneath its header")
        raise LiveAcceptanceError("downloaded XLSX has no nontrivial header")
    finally:
        workbook.close()


def _downloaded_revision(
    client: HttpClient,
    config: AcceptanceConfig,
    artifact: Mapping[str, Any],
    *,
    attachment_id: str,
    model_identity: str,
) -> dict[str, Any]:
    revision_id = _non_empty(artifact.get("revision_id"), "artifact revision ID")
    metadata_response = _request(
        client,
        "GET",
        _url(config.base_url, f"{ARTIFACT_PATH}/{revision_id}"),
        headers=_headers(config),
    )
    metadata = metadata_response.json()
    if not isinstance(metadata, dict):
        raise LiveAcceptanceError("artifact metadata was not an object")
    for field in ("artifact_id", "revision_id", "revision_no", "parent_revision_id", "sha256", "byte_size"):
        if metadata.get(field) != artifact.get(field):
            raise LiveAcceptanceError(f"artifact metadata differs from ordinary chat for {field}")
    source_scope = _as_list(metadata.get("source_scope"), "artifact source_scope")
    if f"attachment:{attachment_id}" not in source_scope:
        raise LiveAcceptanceError("artifact provenance does not include the uploaded attachment")
    if _non_empty(metadata.get("profile_revision_id"), "artifact profile revision") != config.profile_revision_id:
        raise LiveAcceptanceError("artifact profile revision differs from requested profile")
    if _non_empty(metadata.get("model_identity"), "artifact model identity") != model_identity:
        raise LiveAcceptanceError("artifact model identity differs from ordinary chat")
    expected_preset = PRESET_IDENTITIES[config.model_preset]
    if _non_empty(metadata.get("model_preset"), "artifact model preset") != expected_preset:
        raise LiveAcceptanceError("artifact model preset differs from requested preset")
    checkpoint_id = _non_empty(metadata.get("decision_checkpoint_id"), "artifact checkpoint ID")
    if checkpoint_id != _non_empty(artifact.get("decision_checkpoint_id"), "chat artifact checkpoint ID"):
        raise LiveAcceptanceError("artifact metadata checkpoint differs from ordinary chat")
    missing = _as_list(metadata.get("missing"), "artifact missing")
    blockers = _as_list(metadata.get("blockers"), "artifact blockers")
    download_response = _request(
        client,
        "GET",
        _url(config.base_url, f"{ARTIFACT_PATH}/{revision_id}/download"),
        headers=_headers(config),
    )
    download_sha256 = hashlib.sha256(download_response.content).hexdigest()
    sha256 = _hex_sha(metadata.get("sha256"), "artifact SHA-256")
    if download_sha256 != sha256:
        raise LiveAcceptanceError("downloaded artifact SHA-256 differs from metadata")
    xlsx_structure = _validate_downloaded_xlsx(download_response.content)
    return {
        "artifact_id": _non_empty(metadata.get("artifact_id"), "artifact ID"),
        "revision_id": revision_id,
        "revision_no": metadata.get("revision_no"),
        "parent_revision_id": metadata.get("parent_revision_id"),
        "sha256": sha256,
        "download_sha256": download_sha256,
        "byte_size": metadata.get("byte_size"),
        "source_scope": list(source_scope),
        "profile_revision_id": config.profile_revision_id,
        "model_identity": model_identity,
        "model_preset": expected_preset,
        "decision_checkpoint_id": checkpoint_id,
        # Runtime wording may include input filenames, local paths, prompts, or
        # provider diagnostics. The receipt proves that both arrays were present
        # without persisting their free text.
        "missing_count": len(missing),
        "blocker_count": len(blockers),
        **xlsx_structure,
    }


def _chat_payload(
    *, attachment_id: str, profile_revision_id: str, question: str, session_id: str,
) -> dict[str, Any]:
    return {
        "question": question,
        "attachment_id": attachment_id,
        "profile_revision_id": profile_revision_id,
        "session_id": session_id,
        "candidate_acceptance": True,
    }


def _run_chat(
    client: HttpClient,
    config: AcceptanceConfig,
    payload: Mapping[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    response = _request(
        client,
        "POST",
        _url(config.base_url, CHAT_STREAM_PATH),
        headers=_headers(config),
        json=dict(payload),
    )
    return _stream_final(response)


def _require_completed_progress(progress: list[dict[str, Any]], checkpoint_id: str) -> None:
    bound = [item for item in progress if str(item.get("checkpoint_id") or "") == checkpoint_id]
    if not bound:
        raise LiveAcceptanceError("ordinary chat did not emit checkpoint-bound tool progress")
    previous = -1
    total = 0
    for item in bound:
        completed = item.get("completed")
        current_total = item.get("total")
        if not isinstance(completed, int) or not isinstance(current_total, int) or current_total <= 0:
            raise LiveAcceptanceError("tool progress is incomplete")
        if completed < previous or completed > current_total:
            raise LiveAcceptanceError("tool progress is not monotonic")
        previous = completed
        total = current_total
    if previous < total:
        raise LiveAcceptanceError("tool progress did not complete")


def _write_report(path: Path, report: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        temporary.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)


def _assert_redacted(value: Any) -> None:
    if isinstance(value, Mapping):
        for key, nested in value.items():
            normalized = str(key).casefold().replace("-", "_")
            if normalized in _REDACTED_KEYS:
                raise LiveAcceptanceError("acceptance report must remain redacted")
            _assert_redacted(nested)
    elif isinstance(value, list):
        for nested in value:
            _assert_redacted(nested)
    elif isinstance(value, str):
        if (
            "\n" in value
            or "\r" in value
            or _LOCAL_PATH_RE.search(value)
            or _RELATIVE_SOURCE_PATH_RE.search(value)
            or _SECRET_VALUE_RE.search(value)
            or _UNREDACTED_TEXT_RE.search(value)
        ):
            raise LiveAcceptanceError("acceptance report must remain redacted")


def _only_keys(value: Mapping[str, Any], allowed: frozenset[str], label: str) -> None:
    unknown = set(value) - allowed
    if unknown:
        raise LiveAcceptanceError(f"acceptance report has unknown {label} fields")


def _positive_int(value: Any, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise LiveAcceptanceError(f"{label} is invalid")
    return value


def _nonnegative_int(value: Any, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise LiveAcceptanceError(f"{label} is invalid")
    return value


def _string_list(value: Any, label: str) -> list[str]:
    result = _as_list(value, label)
    if any(not isinstance(item, str) or not item for item in result):
        raise LiveAcceptanceError(f"{label} must contain non-empty strings")
    return result


def _max_elapsed_seconds(config: AcceptanceConfig) -> float:
    value = config.max_elapsed_seconds
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value) or value <= 0:
        raise LiveAcceptanceError("configured deadline is invalid")
    return float(value)


def validate_report(
    report: Mapping[str, Any], *, expected_evidence_kind: str = LIVE_RUNTIME
) -> None:
    if not isinstance(report, Mapping):
        raise ValueError("acceptance report must be an object")
    if report.get("evidence_kind") != expected_evidence_kind:
        raise ValueError(f"acceptance report evidence_kind must be {expected_evidence_kind}")
    if report.get("schema") != REPORT_SCHEMA:
        raise ValueError(f"acceptance report schema must be {REPORT_SCHEMA}")
    try:
        _assert_redacted(report)
    except LiveAcceptanceError as error:
        raise ValueError(str(error)) from error
    runtime = report.get("runtime")
    attachment = report.get("attachment")
    first = report.get("revision_1")
    second = report.get("revision_2")
    if not all(isinstance(item, Mapping) for item in (runtime, attachment, first, second)):
        raise ValueError("acceptance report lacks runtime, attachment, or revision metadata")
    try:
        _only_keys(
            report,
            frozenset({"schema", "evidence_kind", "runtime", "attachment", "revision_1", "revision_2", "elapsed_seconds"}),
            "root",
        )
        _only_keys(
            runtime,
            frozenset({"source_commit_full", "build_number", "runtime_alignment", "profile_revision_id", "model_preset", "observed_model_preset", "model_identity"}),
            "runtime",
        )
        _only_keys(attachment, frozenset({"attachment_id", "sha256"}), "attachment")
        revision_fields = frozenset({
            "artifact_id", "revision_id", "revision_no", "parent_revision_id", "sha256",
            "download_sha256", "byte_size", "source_scope", "profile_revision_id",
            "model_identity", "model_preset", "decision_checkpoint_id", "missing_count",
            "blocker_count", "visible_sheet_count", "header_cell_count", "data_row_count",
        })
        _only_keys(first, revision_fields, "revision")
        _only_keys(second, revision_fields, "revision")
        profile_revision_id = _non_empty(runtime.get("profile_revision_id"), "runtime profile revision")
        source_commit_full = _non_empty(runtime.get("source_commit_full"), "runtime full source commit")
        if not _FULL_COMMIT_RE.fullmatch(source_commit_full):
            raise LiveAcceptanceError("runtime full source commit is invalid")
        _positive_int(runtime.get("build_number"), "runtime build number")
        if runtime.get("runtime_alignment") != "aligned":
            raise LiveAcceptanceError("runtime alignment is invalid")
        model_identity = _non_empty(runtime.get("model_identity"), "runtime model identity")
        model_preset = _non_empty(runtime.get("model_preset"), "runtime model preset")
        observed_preset = _non_empty(runtime.get("observed_model_preset"), "observed model preset")
        if PRESET_IDENTITIES.get(model_preset) != observed_preset:
            raise LiveAcceptanceError("runtime preset identity is not canonical")
        attachment_id = _non_empty(attachment.get("attachment_id"), "attachment ID")
        _hex_sha(attachment.get("sha256"), "attachment SHA-256")
        for expected_no, revision in ((1, first), (2, second)):
            if not isinstance(revision.get("revision_no"), int) or isinstance(revision["revision_no"], bool) or revision["revision_no"] != expected_no:
                raise LiveAcceptanceError(f"revision {expected_no} number is invalid")
            _non_empty(revision.get("artifact_id"), f"revision {expected_no} artifact ID")
            _non_empty(revision.get("revision_id"), f"revision {expected_no} ID")
            sha256 = _hex_sha(revision.get("sha256"), f"revision {expected_no} SHA-256")
            if _hex_sha(revision.get("download_sha256"), f"revision {expected_no} download SHA-256") != sha256:
                raise LiveAcceptanceError(f"revision {expected_no} downloaded hash differs")
            _positive_int(revision.get("byte_size"), f"revision {expected_no} byte size")
            scope = _string_list(revision.get("source_scope"), f"revision {expected_no} source scope")
            if f"attachment:{attachment_id}" not in scope:
                raise LiveAcceptanceError(f"revision {expected_no} lacks attachment provenance")
            if revision.get("profile_revision_id") != profile_revision_id:
                raise LiveAcceptanceError(f"revision {expected_no} profile lineage differs")
            if revision.get("model_identity") != model_identity:
                raise LiveAcceptanceError(f"revision {expected_no} model identity differs")
            if revision.get("model_preset") != observed_preset:
                raise LiveAcceptanceError(f"revision {expected_no} model preset differs")
            _non_empty(revision.get("decision_checkpoint_id"), f"revision {expected_no} checkpoint ID")
            if "missing" in revision or "blockers" in revision:
                raise LiveAcceptanceError("acceptance report must remain redacted")
            for key in ("missing_count", "blocker_count"):
                _nonnegative_int(revision.get(key), f"revision {expected_no} {key}")
            _positive_int(revision.get("visible_sheet_count"), f"revision {expected_no} visible sheet count")
            if _positive_int(revision.get("header_cell_count"), f"revision {expected_no} header cell count") < 2:
                raise LiveAcceptanceError(f"revision {expected_no} header cell count is invalid")
            _positive_int(revision.get("data_row_count"), f"revision {expected_no} data row count")
        if first.get("sha256") == second.get("sha256"):
            raise LiveAcceptanceError("immutable correction must change the workbook hash")
        if second.get("parent_revision_id") != first.get("revision_id"):
            raise LiveAcceptanceError("revision 2 parent lineage is invalid")
        elapsed = report.get("elapsed_seconds")
        if (
            isinstance(elapsed, bool)
            or not isinstance(elapsed, (int, float))
            or not math.isfinite(elapsed)
            or elapsed < 0
        ):
            raise LiveAcceptanceError("elapsed time is invalid")
    except LiveAcceptanceError as error:
        raise ValueError(str(error)) from error


def _exercise_contract_with_client(
    config: AcceptanceConfig,
    *,
    client: HttpClient,
) -> dict[str, Any]:
    attachment = Path(config.attachment)
    if _fixture_path(attachment):
        raise LiveAcceptanceError("tests/fixtures input cannot be used as live acceptance evidence")
    if not attachment.is_file():
        raise LiveAcceptanceError("attachment must be an existing regular file")
    if config.model_preset not in PRESET_IDENTITIES:
        raise LiveAcceptanceError("model preset must be qwen-9b or qwen-35b")
    deadline = _max_elapsed_seconds(config)
    _base_url(config.base_url)
    profile_revision_id = _non_empty(config.profile_revision_id, "profile revision")
    source_sha256 = hashlib.sha256(attachment.read_bytes()).hexdigest()
    started = time.monotonic()
    active_client = client
    try:
        version = _runtime_version(active_client, config)
        with attachment.open("rb") as stream:
            upload_response = _request(
                active_client,
                "POST",
                _url(config.base_url, UPLOAD_PATH),
                headers=_headers(config, idempotency=True),
                files={"file": (attachment.name, stream, "application/octet-stream")},
                data={"candidate_acceptance": "true"},
            )
        upload = upload_response.json()
        if not isinstance(upload, dict):
            raise LiveAcceptanceError("attachment upload response was not an object")
        attachment_id = _non_empty(upload.get("attachment_id"), "uploaded attachment ID")
        session_id = f"live-workbook-{uuid.uuid4().hex}"
        first_final, first_progress = _run_chat(
            active_client,
            config,
            _chat_payload(
                attachment_id=attachment_id,
                profile_revision_id=profile_revision_id,
                session_id=session_id,
                question="Собери проверяемый черновик ВОР из прикреплённого документа через доступный workbook tool.",
            ),
        )
        first_artifact = _artifact_from_final(first_final, attachment_id, source_sha256)
        _require_completed_progress(first_progress, str(first_artifact.get("decision_checkpoint_id") or ""))
        model_identity = _model_identity(first_final)
        revision_1 = _downloaded_revision(
            active_client,
            config,
            first_artifact,
            attachment_id=attachment_id,
            model_identity=model_identity,
        )
        second_final, second_progress = _run_chat(
            active_client,
            config,
            _chat_payload(
                attachment_id=attachment_id,
                profile_revision_id=profile_revision_id,
                session_id=session_id,
                question=(
                    "Сделай корректировку того же workbook draft. Используй тот же прикреплённый "
                    f"документ и создай новую immutable revision с parent revision {revision_1['revision_id']}."
                ),
            ),
        )
        second_artifact = _artifact_from_final(second_final, attachment_id, source_sha256)
        _require_completed_progress(second_progress, str(second_artifact.get("decision_checkpoint_id") or ""))
        second_model_identity = _model_identity(second_final)
        if second_model_identity != model_identity:
            raise LiveAcceptanceError("correction used a different observed model identity")
        revision_2 = _downloaded_revision(
            active_client,
            config,
            second_artifact,
            attachment_id=attachment_id,
            model_identity=model_identity,
        )
        elapsed_seconds = round(time.monotonic() - started, 3)
        if elapsed_seconds > deadline:
            raise LiveAcceptanceError("live acceptance exceeded configured deadline")
        report = {
            "schema": REPORT_SCHEMA,
            "evidence_kind": CONTRACT_TEST,
            "runtime": {
                **version,
                "profile_revision_id": profile_revision_id,
                "model_preset": config.model_preset,
                "observed_model_preset": PRESET_IDENTITIES[config.model_preset],
                "model_identity": model_identity,
            },
            "attachment": {"attachment_id": attachment_id, "sha256": source_sha256},
            "revision_1": revision_1,
            "revision_2": revision_2,
            "elapsed_seconds": elapsed_seconds,
        }
        validate_report(report, expected_evidence_kind=CONTRACT_TEST)
        return report
    finally:
        pass


def exercise_contract(config: AcceptanceConfig, *, client: HttpClient) -> dict[str, Any]:
    """Exercise public boundaries with a fake transport without live evidence I/O."""
    return _exercise_contract_with_client(
        config,
        client=client,
    )


def run_acceptance(config: AcceptanceConfig) -> dict[str, Any]:
    """Run real acceptance only; fake clients cannot use this public entrypoint."""
    with httpx.Client(timeout=httpx.Timeout(_max_elapsed_seconds(config))) as client:
        contract_report = _exercise_contract_with_client(
            config,
            client=client,
        )
    report = dict(contract_report)
    report["evidence_kind"] = LIVE_RUNTIME
    validate_report(report, expected_evidence_kind=LIVE_RUNTIME)
    _write_report(config.out, report)
    return report


def parse_args(argv: list[str] | None = None) -> AcceptanceConfig:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--attachment", type=Path, required=True, help="real user-owned source document")
    parser.add_argument("--base-url", required=True, help="running LES API origin")
    parser.add_argument("--profile-revision", required=True, help="immutable ordinary-chat profile revision")
    parser.add_argument("--model-preset", required=True, choices=sorted(PRESET_IDENTITIES))
    parser.add_argument("--out", type=Path, required=True, help="redacted JSON receipt path")
    parser.add_argument("--max-elapsed-seconds", type=float, default=1800.0)
    parser.add_argument(
        "--api-key-env",
        default="LES_LIVE_WORKBOOK_ACCEPTANCE_API_KEY",
        help="environment variable containing an optional API key; its value is never printed",
    )
    args = parser.parse_args(argv)
    return AcceptanceConfig(
        attachment=args.attachment,
        base_url=args.base_url,
        profile_revision_id=args.profile_revision,
        model_preset=args.model_preset,
        out=args.out,
        api_key=os.getenv(args.api_key_env) or None,
        max_elapsed_seconds=args.max_elapsed_seconds,
    )


def main(argv: list[str] | None = None) -> int:
    config = parse_args(argv)
    report = run_acceptance(config)
    print(f"live workbook acceptance receipt written: {config.out}")
    print(f"evidence_kind={report['evidence_kind']} elapsed_seconds={report['elapsed_seconds']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

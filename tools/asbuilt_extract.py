#!/usr/bin/env python3
"""asbuilt_extract.py — CLI приёмки смонтированного объёма из исполнительных схем (сканов).

Прогоняет PDF/папку через `asbuilt_intake_service`: рендер → поворот → vision-OCR таблиц →
строки → (опц.) журнал объёмов → свод по системам. Единственный LLM-шаг — OCR ячеек; числа
считает код (ADR-11).

Примеры:
  uv run python tools/asbuilt_extract.py "/path/АУПС-СОУЭ" --preview
  uv run python tools/asbuilt_extract.py "/path/lист.pdf" --engine cloud --rotate 90
  uv run python tools/asbuilt_extract.py "/path/АУПС-СОУЭ" --write --status pending --xlsx /tmp/svod.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from proxy.services.asbuilt_intake_service import process_path  # noqa: E402


def _fmt_qty(v) -> str:
    if v is None:
        return ""
    return (f"{v:.3f}".rstrip("0").rstrip(".")).replace(".", ",")


def main() -> int:
    ap = argparse.ArgumentParser(description="Приёмка смонтированного объёма из исполнительных схем (сканов)")
    ap.add_argument("path", help="PDF-файл или папка со сканами")
    ap.add_argument("--engine", choices=["local", "cloud"], default="local",
                    help="local=gemma4:12b (дефолт) | cloud=gpt-4.1 через proxyapi")
    ap.add_argument("--model", default="", help="переопределить модель OCR (иначе из env по движку)")
    ap.add_argument("--rotate", default="auto", help="auto | 0 | 90 | 180 | 270")
    ap.add_argument("--write", action="store_true", help="записать строки в журнал объёмов (status)")
    ap.add_argument("--preview", action="store_true", help="только показать (по умолчанию, без записи)")
    ap.add_argument("--status", default="pending", choices=["pending", "confirmed", "rejected"])
    ap.add_argument("--project-id", type=int, default=0)
    ap.add_argument("--xlsx", default="", help="сохранить свод в xlsx по этому пути")
    args = ap.parse_args()

    out = process_path(
        args.path, rotate=args.rotate, engine=args.engine, model=args.model or None,
        write=args.write and not args.preview, status=args.status, project_id=args.project_id,
    )

    print(f"\nДвижок OCR: {out['engine']} / {out['model']}")
    for f in out["files"]:
        ctx = f["ctx"]
        tag = "/".join(s for s in (ctx.get("floor"), ctx.get("system"), ctx.get("line")) if s) or "—"
        err = f"  ⚠ {f['error']}" if f["error"] else ""
        print(f"\n■ {f['pdf']}  [{tag}]  поворот={f['rotation_used']}°  строк={f['raw_count']} "
              f"(объём={f['kept']}, пропущено={f['skipped']}){err}")

    print("\n── Строки смонтированного объёма ──")
    print(f"{'Система':<7}{'Этаж/линия':<12}{'Наименование':<46}{'Тип':<26}{'Ед':<5}{'Кол-во':>10}")
    for r in out["rows"]:
        fl = "/".join(s for s in (r["floor"], r["line"]) if s)
        print(f"{r['system']:<7}{fl:<12}{r['name'][:44]:<46}{r['type'][:24]:<26}{r['unit']:<5}{_fmt_qty(r['qty']):>10}")

    print("\n── Свод по системам (SUM, код) ──")
    for c in out["consolidation"]:
        print(f"{c['system']:<7}{c['name'][:50]:<52}{c['unit']:<5}{_fmt_qty(c['total']):>12}  ({c['rows']} строк)")

    if args.write and not args.preview:
        print(f"\n✔ записано в журнал: {out['written']} строк (status={out['status']})")

    if args.xlsx:
        _write_xlsx(Path(args.xlsx), out)
        print(f"✔ свод xlsx: {args.xlsx}")
    return 0


def _write_xlsx(path: Path, out: dict) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Свод"
    hdr = PatternFill("solid", fgColor="1F4E78")
    hf = Font(bold=True, color="FFFFFF")
    ws.append(["Система", "Наименование", "Ед.", "Кол-во (смонтировано)", "Строк"])
    for c in range(1, 6):
        ws.cell(row=1, column=c).fill = hdr
        ws.cell(row=1, column=c).font = hf
    for c in out["consolidation"]:
        ws.append([c["system"], c["name"], c["unit"], c["total"], c["rows"]])
    ws2 = wb.create_sheet("Строки")
    ws2.append(["PDF", "Этаж", "Система", "Линия", "Наименование", "Тип", "Ед.", "Кол-во"])
    for r in out["rows"]:
        ws2.append([r["pdf"], r["floor"], r["system"], r["line"], r["name"], r["type"], r["unit"], r["qty"]])
    for w in (wb["Свод"], ws2):
        for col in w.columns:
            width = max((len(str(c.value)) for c in col if c.value), default=10)
            w.column_dimensions[col[0].column_letter].width = min(width + 2, 50)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    raise SystemExit(main())

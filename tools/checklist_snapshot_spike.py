#!/usr/bin/env python3
"""Snapshot-спайк чек-листов Glorax (T0.1).

Read-only анализ XLSX-чек-листов входного контроля ПД/РД: классифицирует
строки каждого содержательного листа на "критерий", "заголовок блока" и
"пустая/прочее", сравнивает с данными Data Validation и вторичными
визуальными признаками (заливка/жирность/номер строки), и печатает
Markdown-отчёт по каждому листу + сводку по файлу + список спорных строк.

Скрипт НИЧЕГО не записывает во входной xlsx (load_workbook(..., data_only=False),
workbook.save() не вызывается).

Использование:
    uv run python tools/checklist_snapshot_spike.py \
        --xlsx "путь/к/файлу.xlsx" --stage PD --out docs/checklist_review/SNAPSHOT_PD.md
"""

from __future__ import annotations

import argparse
import datetime as dt
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet

MAX_COL = 6  # колонки A..F достаточно; листы объявляют до 16383 колонок мусора
BLOCK_HEADER_FILL = "FF3200F0"  # фиолетовая заливка ячейки B у заголовков блоков
SERVICE_SHEETS = {"Правила заполнения", "Сводная"}
SUMMARY_TAIL_MARKERS = (
    "Всего пунктов:",
    "Чек лист заполнен на:",
    "Из них, соответствуют критериям:",
    "Из них, не соответствуют критериям:",
)
# Строки-заголовки раздела верхнего уровня (не блок, не критерий):
# A содержит текст (не число/не "X.Y"), B пустая. Пример: "ИСХОДНАЯ ДОКУМЕНТАЦИЯ".
SECTION_HEADER_HINT = "ИСХОДНАЯ ДОКУМЕНТАЦИЯ"

# Известные (ожидаемые) итоговые числа критериев для ПД (по ТЗ) — для самопроверки.
EXPECTED_PD_TOTALS = {
    "Общее": 10,
    "СПОЗУ": 29,
    "АР": 87,
    "КР": 45,
    "ЭОМ": 39,
    "ЭН": 19,
    " ВК и НВК": 66,
    "ОВиК": 80,
    "СС": 65,
    "ПБ2 (АППЗ)": 35,
}
EXPECTED_PD_TOTAL_ALL = 533


@dataclass
class RowClass:
    row: int
    kind: str  # "critere" | "block_header" | "block_header_no_fill" | "section_header" | "empty_other"
    a_value: object
    b_text: str
    dv_covered: bool
    bold: bool
    purple_fill: bool
    has_number_dotted: bool  # A похоже на "N.M" (подпункт критерия)
    ambiguous: bool = False
    ambiguous_reason: str = ""


@dataclass
class SheetReport:
    name: str
    total_rows: int = 0
    criteria: int = 0
    block_headers: int = 0
    block_headers_no_fill: int = 0  # B непустая, DV нет, заливки нет — заголовок блока без заливки
    empty_other: int = 0
    section_headers: int = 0
    b_nonempty_no_tail: int = 0
    dv_value_sets: list[str] = field(default_factory=list)
    rows: list[RowClass] = field(default_factory=list)


def expand_sqref_to_cells(sqref) -> set[str]:
    """Разворачивает MultiCellRange/строку sqref в множество координат ячеек."""
    cells: set[str] = set()
    ranges = str(sqref).split()
    for rng in ranges:
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cells.add(f"{openpyxl.utils.get_column_letter(c)}{r}")
    return cells


def is_dotted_number(value) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    return "." in s and all(part.isdigit() for part in s.split(".") if part != "")


def is_plain_number(value) -> bool:
    if isinstance(value, (int, float)):
        return True
    if value is None:
        return False
    s = str(value).strip()
    return s.isdigit()


def truncate(text: str, limit: int = 120) -> str:
    text = " ".join(str(text).split())
    if len(text) <= limit:
        return text
    return text[: limit - 1] + "…"


def classify_sheet(ws: Worksheet) -> SheetReport:
    report = SheetReport(name=ws.title)

    # Собираем DV-множества: value_set_string -> set(cells)
    dv_cell_map: dict[str, set[str]] = {}
    all_dv_cells: set[str] = set()
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        value_set = str(dv.formula1 or "").strip('"')
        cells = expand_sqref_to_cells(dv.sqref)
        dv_cell_map.setdefault(value_set, set()).update(cells)
        all_dv_cells.update(cells)
    report.dv_value_sets = sorted(dv_cell_map.keys())

    max_row = ws.max_row
    for row_idx in range(1, max_row + 1):
        report.total_rows += 1
        if row_idx <= 2:
            continue  # служебные строки: заголовок листа / шапка таблицы

        a_cell = ws.cell(row=row_idx, column=1)
        b_cell = ws.cell(row=row_idx, column=2)
        c_cell = ws.cell(row=row_idx, column=3)

        a_val = a_cell.value
        b_val = b_cell.value
        b_text = "" if b_val is None else str(b_val)

        if b_val is None or b_text.strip() == "":
            # Возможный заголовок раздела верхнего уровня (A есть текст, B пусто)
            if isinstance(a_val, str) and a_val.strip():
                report.section_headers += 1
                report.rows.append(
                    RowClass(
                        row=row_idx,
                        kind="section_header",
                        a_value=a_val,
                        b_text="",
                        dv_covered=False,
                        bold=bool(a_cell.font and a_cell.font.bold),
                        purple_fill=False,
                        has_number_dotted=False,
                    )
                )
            else:
                report.empty_other += 1
            continue

        # Строки итоговой мини-таблицы в конце листа ("Всего пунктов:" и т.п.)
        if any(b_text.strip().startswith(marker) for marker in SUMMARY_TAIL_MARKERS):
            report.empty_other += 1
            continue

        # Доп. метрика: все непустые B, начиная со строки 3, кроме хвостовых итоговых строк
        # (уже отфильтрованы выше). Считает и критерии, и заголовки блоков вместе — это ближе
        # к грубой оценке "~400", упомянутой в ТЗ как альтернатива числу 533.
        report.b_nonempty_no_tail += 1

        b_font = b_cell.font
        bold = bool(b_font and b_font.bold)
        fill = b_cell.fill
        purple = False
        try:
            purple = bool(fill and fill.fgColor and fill.fgColor.rgb == BLOCK_HEADER_FILL)
        except Exception:
            purple = False

        c_coord = f"C{row_idx}"
        dv_covered = c_coord in all_dv_cells
        dotted = is_dotted_number(a_val)
        plain_num = is_plain_number(a_val)

        # Основная эвристика: заголовок блока = заливка ячейки B фиолетовая (FF3200F0).
        # Замечено на ПД (доп. признак bold=True) и на РД (bold=False, только заливка) —
        # поэтому решающий признак: purple fill; bold используется как вторичный/подтверждающий.
        is_block_header = purple

        ambiguous = False
        ambiguous_reason = ""

        # Решение оператора 2026-07-04: критерий = B непустая И answer-ячейка C
        # покрыта data validation. Любая непустая B без DV-покрытия = заголовок
        # (блока — если есть заливка FF3200F0, иначе "заголовок блока (без заливки)").
        # Прежнее правило "номер N/N.M без DV = критерий" (ветка (b)) — убрано по решению
        # оператора: строки СПОЗУ:6/КР:7 и аналогичные — это заголовки блоков, не критерии.
        if is_block_header:
            report.block_headers += 1
            kind = "block_header"
        elif dv_covered:
            kind = "critere"
            report.criteria += 1
        else:
            kind = "block_header_no_fill"
            report.block_headers_no_fill += 1

        report.rows.append(
            RowClass(
                row=row_idx,
                kind=kind,
                a_value=a_val,
                b_text=b_text,
                dv_covered=dv_covered,
                bold=bold,
                purple_fill=purple,
                has_number_dotted=dotted,
                ambiguous=ambiguous,
                ambiguous_reason=ambiguous_reason,
            )
        )

    return report


def build_report_md(
    xlsx_path: Path,
    stage: str,
    sheet_reports: list[SheetReport],
    excluded_sheets: list[str],
) -> str:
    lines: list[str] = []
    today = dt.date.today().isoformat()
    lines.append(f"# Snapshot-спайк чек-листа {stage} (T0.1)")
    lines.append("")
    lines.append(f"- Дата: {today}")
    lines.append(f"- Файл-источник: `{xlsx_path}`")
    lines.append(f"- Скрипт: `tools/checklist_snapshot_spike.py` (read-only, openpyxl, data_only=False)")
    lines.append(
        "- Исключённые служебные листы: " + ", ".join(f"`{s}`" for s in excluded_sheets)
    )
    lines.append("")
    lines.append(
        "**Решение оператора 2026-07-04: строки без DV = заголовки.** Все строки, ранее "
        "помеченные как «спорные» (СПОЗУ:6, КР:7 и любые аналогичные), — это заголовки блоков "
        "(например АР:17 — заголовок блока строк 18–26), не критерии. Итоговое правило "
        "классификации: **критерий = ячейка B непустая И answer-ячейка C покрыта data "
        "validation; любая непустая B без DV-покрытия = заголовок** (блока или раздела). "
        "Прежнее правило «номер N/N.M без DV = критерий» — убрано."
    )
    lines.append("")
    lines.append("## Использованная эвристика")
    lines.append("")
    lines.append(
        "1. Строки 1-2 каждого листа — служебные (название раздела, шапка колонок), не классифицируются."
    )
    lines.append(
        "2. **Заголовок блока (с заливкой)** — решающий признак: заливка ячейки B равна `FF3200F0` "
        "(фиолетовый). На листах ПД заголовки блоков дополнительно жирные (`bold=True`), "
        "на листах РД жирность у заголовков блоков **не выставлена** (`bold=False`) — поэтому "
        "жирность используется только как вторичное подтверждение, а не как решающий признак."
    )
    lines.append(
        "3. **Критерий** — B непустая, заливка НЕ purple, И answer-ячейка в колонке C покрыта "
        "каким-либо data validation (`worksheet.data_validations.dataValidation`, sqref "
        "развёрнут в множество ячеек)."
    )
    lines.append(
        "4. **Заголовок блока (без заливки)** — B непустая, заливка НЕ purple, DV на "
        "answer-ячейке C нет. По решению оператора 2026-07-04 это тоже заголовок блока, не "
        "критерий: строки вида «Наличие необходимых расчётов:» на некоторых листах не "
        "выделены заливкой, но структурно являются заголовком группы подпунктов. Считается "
        "отдельной категорией, чтобы был виден количественный эффект решения оператора."
    )
    lines.append(
        "5. **Заголовок раздела верхнего уровня** (напр. `ИСХОДНАЯ ДОКУМЕНТАЦИЯ`) — B пустая, "
        "A содержит текст. Отдельная категория, не считается ни критерием, ни заголовком блока."
    )
    lines.append(
        "6. Итоговые строки в конце листа (`Всего пунктов:`, `Чек лист заполнен на:`, "
        "`Из них, соответствуют критериям:`, `Из них, не соответствуют критериям:`) — "
        "'пусто/прочее', не критерии, хотя B у них непустая и жирная."
    )
    lines.append(
        "7. ВАЖНО про DV-диапазоны: на большинстве листов DV НЕ покрывает сплошняком всю "
        "колонку C — есть разрывы ровно на строках заголовков блоков (напр. СПОЗУ: "
        "`C4:C5 C7 C9` и `C11:C12 C14:C18 ...` пропускают строки 10, 13, 19, 25 — это и есть "
        "заголовки блоков). DV-покрытие C — единственный решающий признак критерия (см. п.3); "
        "заливка (п.2) — диагностический подтип заголовка, номер в A более не участвует "
        "в классификации (см. решение оператора выше)."
    )
    lines.append("")
    lines.append("## Сводная таблица по листам")
    lines.append("")
    lines.append(
        "| Лист | Всего строк | Критериев | Заголовков блоков (с заливкой) | "
        "Заголовков блоков (без заливки) | Заголовков раздела | Пусто/прочее | DV-значения |"
    )
    lines.append("|---|---|---|---|---|---|---|---|")
    total_criteria = 0
    total_blocks = 0
    total_blocks_no_fill = 0
    total_rows = 0
    total_empty = 0
    total_section = 0
    for sr in sheet_reports:
        dv_str = "; ".join(f"«{v}»" for v in sr.dv_value_sets) if sr.dv_value_sets else "—"
        lines.append(
            f"| {sr.name} | {sr.total_rows} | {sr.criteria} | {sr.block_headers} | "
            f"{sr.block_headers_no_fill} | {sr.section_headers} | {sr.empty_other} | {dv_str} |"
        )
        total_criteria += sr.criteria
        total_blocks += sr.block_headers
        total_blocks_no_fill += sr.block_headers_no_fill
        total_rows += sr.total_rows
        total_empty += sr.empty_other
        total_section += sr.section_headers
    lines.append(
        f"| **ИТОГО** | {total_rows} | {total_criteria} | {total_blocks} | "
        f"{total_blocks_no_fill} | {total_section} | {total_empty} | |"
    )
    lines.append("")

    total_blocks_all = total_blocks + total_blocks_no_fill
    lines.append("## Итоговые числа по файлу")
    lines.append("")
    lines.append(f"- Критериев всего: **{total_criteria}**")
    lines.append(
        f"- Заголовков блоков всего: **{total_blocks_all}** "
        f"(с заливкой: {total_blocks}, без заливки: {total_blocks_no_fill})"
    )
    lines.append(f"- Заголовков разделов верхнего уровня: **{total_section}**")
    lines.append(f"- Пустых/прочих строк: **{total_empty}**")
    lines.append("")

    if stage.upper() == "PD":
        lines.append("## Самопроверка против ожидаемых чисел (ТЗ)")
        lines.append("")
        lines.append(
            "ТЗ (implementation_plan.md, со ссылкой на docs/CHECKLIST_REVIEW_PD_TASK.md) даёт "
            "по листам числа Общее(10)/СПОЗУ(29)/АР(87)/КР(45)/ЭОМ(39)/ЭН(19)/ВК и НВК(66)/"
            "ОВиК(80)/СС(65)/ПБ2(35), всего ~533, и явно помечает их как **criteria-like строки, "
            "включающие заголовки блоков** — то есть это не то же самое, что колонка «Критериев» "
            "в таблице выше. Ниже сравнение с суммой «критерии + все заголовки блоков (с "
            "заливкой и без)» (моя классификация после решения оператора 2026-07-04) и с "
            "независимой метрикой «все непустые B от строки 3, без хвостовых итоговых строк "
            "листа» (`Всего пунктов:` и т.п.), которая ближе к альтернативной оценке «~400», "
            "тоже упомянутой в ТЗ."
        )
        lines.append("")
        lines.append(
            "| Лист | Критерии+Заголовки (моя классификация) | B-непустых без хвоста (доп. метрика) | Ожидалось по ТЗ (533-набор) |"
        )
        lines.append("|---|---|---|---|")
        found_by_name = {
            sr.name: (sr.criteria + sr.block_headers + sr.block_headers_no_fill)
            for sr in sheet_reports
        }
        b_nonempty_by_name = {sr.name: sr.b_nonempty_no_tail for sr in sheet_reports}
        combined_total = 0
        b_nonempty_total = 0
        for name, expected in EXPECTED_PD_TOTALS.items():
            combined = found_by_name.get(name)
            b_ne = b_nonempty_by_name.get(name)
            combined_total += combined or 0
            b_nonempty_total += b_ne or 0
            lines.append(f"| {name} | {combined if combined is not None else 'нет листа'} | {b_ne} | {expected} |")
        lines.append(f"| **ВСЕГО** | {combined_total} | {b_nonempty_total} | {EXPECTED_PD_TOTAL_ALL} |")
        lines.append("")
        lines.append(
            f"«Критерии+заголовки» ({combined_total}) совпадает с «B-непустых без хвоста» "
            f"({b_nonempty_total}) — обе метрики считают все непустые B-строки, но не "
            f"совпадают с ТЗ-набором ({EXPECTED_PD_TOTAL_ALL}). Расхождение системное и "
            "ожидаемо: сам ТЗ-план фиксирует, что «533» и альтернативная оценка «~400» уже "
            "расходились между собой ДО этого прогона (см. implementation_plan.md, строка про "
            "«382» от внешнего разбора Codex) и что точное число должен зафиксировать importer "
            "снапшот-тестом на Phase 1, а не этот разведочный скрипт. Решением оператора "
            "2026-07-04 закрыт вопрос трактовки строк без DV (в т.ч. секции «ИСХОДНАЯ "
            "ДОКУМЕНТАЦИЯ»): все они — заголовки, не критерии; актуальное число критериев "
            f"ПД по этому прогону — {total_criteria} (было 337 до решения оператора, если "
            "строки без DV засчитывались как критерии по номеру)."
        )
        lines.append("")

    lines.append("## Спорные строки")
    lines.append("")
    any_ambiguous = False
    for sr in sheet_reports:
        amb_rows = [r for r in sr.rows if r.ambiguous]
        if not amb_rows:
            continue
        any_ambiguous = True
        for r in amb_rows:
            text = truncate(r.b_text or str(r.a_value))
            lines.append(f"- `{sr.name}:{r.row}` — {text} — {r.ambiguous_reason}")
    if not any_ambiguous:
        lines.append(
            "Нет — правило закрыто решением оператора 2026-07-04 (строки без DV = заголовки; "
            "см. секцию «Полный перечень заголовков блоков без заливки» ниже)."
        )
    lines.append("")

    if stage.upper() == "PD":
        lines.append("## Полный перечень заголовков блоков (для сверки оператором)")
        lines.append("")
        for sr in sheet_reports:
            block_rows = [r for r in sr.rows if r.kind == "block_header"]
            if not block_rows:
                continue
            lines.append(f"### {sr.name}")
            for r in block_rows:
                text = truncate(r.b_text)
                lines.append(f"- `{sr.name}:{r.row}` — {text}")
            lines.append("")

    lines.append("## Полный перечень заголовков блоков без заливки (эффект решения оператора)")
    lines.append("")
    any_no_fill = False
    for sr in sheet_reports:
        no_fill_rows = [r for r in sr.rows if r.kind == "block_header_no_fill"]
        if not no_fill_rows:
            continue
        any_no_fill = True
        lines.append(f"### {sr.name}")
        for r in no_fill_rows:
            text = truncate(r.b_text)
            lines.append(f"- `{sr.name}:{r.row}` — {text}")
        lines.append("")
    if not any_no_fill:
        lines.append("Нет строк этой категории на данном чек-листе.")
        lines.append("")

    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, type=Path, help="Путь к исходному XLSX (read-only)")
    parser.add_argument("--stage", required=True, choices=["PD", "RD"], help="Стадия: PD или RD")
    parser.add_argument("--out", required=True, type=Path, help="Путь к выходному Markdown-отчёту")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=False, read_only=False)

    excluded = [s for s in wb.sheetnames if s in SERVICE_SHEETS]
    content_sheets = [s for s in wb.sheetnames if s not in SERVICE_SHEETS]

    sheet_reports: list[SheetReport] = []
    for sheet_name in content_sheets:
        ws = wb[sheet_name]
        sheet_reports.append(classify_sheet(ws))

    report_md = build_report_md(args.xlsx, args.stage, sheet_reports, excluded)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(report_md, encoding="utf-8")

    print(f"OK: отчёт записан в {args.out}")
    total_criteria = sum(sr.criteria for sr in sheet_reports)
    total_blocks = sum(sr.block_headers for sr in sheet_reports)
    total_blocks_no_fill = sum(sr.block_headers_no_fill for sr in sheet_reports)
    print(
        f"Критериев: {total_criteria}, заголовков блоков: {total_blocks + total_blocks_no_fill} "
        f"(с заливкой: {total_blocks}, без заливки: {total_blocks_no_fill})"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

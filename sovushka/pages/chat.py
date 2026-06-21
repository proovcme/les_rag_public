"""
С.О.В.У.Ш.К.А. v5.0 — премиальная рабочая вкладка AI ЧАТ
"""
from __future__ import annotations

import asyncio
import json
import re
import time
from typing import Optional

from nicegui import context, ui

from sovushka.components.charts import _html, esc
from sovushka.safe_markup import sanitize_svg
from sovushka.state import (
    add_log,
    api_delete,
    api_get,
    api_get_bytes,
    api_patch,
    api_post,
    api_post_file,
    api_post_stream,
    last_api_error_text,
    refresh_indexing_mode,
    refresh_samovar,
    state,
)


async def _copy_text(text: str) -> None:
    """Скопировать в буфер. navigator.clipboard работает только в secure-context
    (https/localhost); за туннелем по http — fallback на execCommand через textarea."""
    payload = json.dumps(text or "")
    js = (
        "(async (t) => {"
        "  try { if (navigator.clipboard && window.isSecureContext) {"
        "    await navigator.clipboard.writeText(t); return true; } } catch (e) {}"
        "  try { const ta = document.createElement('textarea'); ta.value = t;"
        "    ta.style.position='fixed'; ta.style.top='0'; ta.style.opacity='0';"
        "    document.body.appendChild(ta); ta.focus(); ta.select();"
        "    const ok = document.execCommand('copy'); document.body.removeChild(ta);"
        "    return ok; } catch (e) { return false; }"
        f"}})({payload})"
    )
    try:
        ok = await ui.run_javascript(js, timeout=5.0)
    except Exception:
        ok = False
    ui.notify("Скопировано" if ok else "Не удалось скопировать (выдели и Ctrl+C)",
              type="positive" if ok else "warning")


def _is_spec_request(q: str) -> bool:
    """«Собери/составь/сделай … спецификацию …» — намерение собрать спеку → GOST-форма."""
    import re as _re
    return bool(_re.search(r"\b(собери|составь|сделай|сформируй|подготовь)\b.{0,40}специфика",
                           (q or "").lower()))


def _rows_to_csv(rows: list[dict]) -> bytes:
    """list[dict] → CSV для рус-Excel: разделитель «;», UTF-8 BOM (кириллица не ломается)."""
    import csv
    import io
    if not rows:
        return "﻿".encode("utf-8")
    keys: list[str] = []
    for r in rows:
        if isinstance(r, dict):
            for k in r:
                if k not in keys:
                    keys.append(k)
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=keys, extrasaction="ignore", delimiter=";")
    w.writeheader()
    for r in rows:
        if isinstance(r, dict):
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in keys})
    return ("﻿" + buf.getvalue()).encode("utf-8")


OUTPUT_FORMATS = {
    "text": ("Текст", "Свободный ответ"),
    "spec": ("Спецификация", "JSON-таблица изделий"),
    "schema": ("Схема", "Иерархия или дерево"),
    "structure": ("Структура", "JSON-объект"),
    "table": ("Таблица", "JSON-массив строк"),
    "mermaid": ("Диаграмма", "Mermaid"),
    "svg": ("SVG", "Векторная схема"),
    "template": ("По образцу", "Структура файла"),
    "verify": ("Верификация", "Сверка скана с распознанным"),
}


def _verify_path(q: str) -> str | None:
    """Путь к скану: в кавычках или абсолютный от первого '/'.

    Устойчиво к пробелам в имени, усечённому расширению (.pd) и хвосту «стр N» —
    лучше отдать путь backend'у (он честно скажет «файл не найден»), чем молча
    увести запрос в обычный RAG.
    """
    q = q or ""
    m = re.search(r'["«]([^"»]+)["»]', q)
    if m and m.group(1).strip().startswith("/"):
        cand = m.group(1).strip()
    else:
        idx = q.find("/")
        if idx < 0:
            return None
        cand = q[idx:].strip()
    # отрезать хвостовое «стр N» / «страница N» / «page N»
    cand = re.sub(r"\s+(?:стр\.?|страниц\w*|page)\s*\d+\s*$", "", cand, flags=re.IGNORECASE).strip()
    return cand or None


def _verify_page(q: str) -> int:
    m = re.search(r"(?:стр\.?|страниц\w*|page)\s*(\d+)", q or "", re.IGNORECASE)
    return max(0, int(m.group(1)) - 1) if m else 0  # оператор думает 1-based


def _is_verify_request(q: str) -> bool:
    """«проверь/сверь объёмы … <путь к скану>» → артефакт верификации (tool, не LLM)."""
    ql = (q or "").casefold()
    has_kw = ("провер" in ql or "свер" in ql or "верифиц" in ql) and (
        "объём" in ql or "объем" in ql or "скан" in ql or "таблиц" in ql
    )
    return bool(has_kw and _verify_path(q))


def should_skip_chat_resource_gate(question: str, dataset_filter: str | None = None) -> bool:
    selected_filter = dataset_filter if dataset_filter and dataset_filter != "(все датасеты)" else None
    try:
        from proxy.services.kot_service import analyze_question
        from proxy.services.query_router import route_query

        intent = route_query(question, dataset_filter=selected_filter)
        kot = analyze_question(question)
        effective_filter = selected_filter or intent.dataset_filter or kot.dataset_filter
        return intent.channel in {"mail", "table"} or effective_filter in {"MAIL", "TABLE"}
    except Exception:
        q = question.casefold()
        table_hint = any(token in q for token in ("смет", "таблиц", "строк", "стоимост", "итого"))
        aggregate_hint = any(token in q for token in ("посчитай", "сумм", "сколько", "покажи"))
        mail_hint = any(token in q for token in ("почт", "письм", "email", "mail", "dropbox"))
        return mail_hint or (table_hint and aggregate_hint)


def build_chat(is_admin: bool, tabs=None, tab_mermaid=None):
    """Строит автономный экран чата: история слева, чат по центру, артефакты справа."""

    out_mode_val = {"v": "text"}
    selected_session_card = {"el": None}
    project_state = {"id": None}  # W17.1: активный объект (None = обычный RAG по всему)

    # Резиновый layout: тащим разделитель → меняем ширину панели артефактов (CSS-var),
    # ширина сохраняется в localStorage. Деградирует мягко (нет JS → разделитель статичен).
    ui.add_body_html("""
    <script>
    (function(){
      function init(){
        var shell=document.querySelector('.sov-chat-shell');
        var div=document.querySelector('.sov-resize-divider');
        if(!shell||!div){ setTimeout(init,400); return; }
        if(div.dataset.bound){ return; } div.dataset.bound='1';
        try{ var s=localStorage.getItem('sovArtW'); if(s) shell.style.setProperty('--sov-artifacts-w', s); }catch(e){}
        var drag=false;
        div.addEventListener('pointerdown', function(e){ drag=true; try{div.setPointerCapture(e.pointerId);}catch(_){} e.preventDefault(); });
        window.addEventListener('pointermove', function(e){
          if(!drag) return;
          var r=shell.getBoundingClientRect();
          var w=r.right - e.clientX - 14;
          w=Math.max(280, Math.min(820, w));
          shell.style.setProperty('--sov-artifacts-w', w+'px');
        });
        window.addEventListener('pointerup', function(){
          if(!drag) return; drag=false;
          var w=shell.style.getPropertyValue('--sov-artifacts-w');
          try{ if(w) localStorage.setItem('sovArtW', w.trim()); }catch(e){}
        });
      }
      if(document.readyState!=='loading'){ init(); } else { document.addEventListener('DOMContentLoaded', init); }
    })();
    </script>
    """)

    with ui.element("div").classes("sov-chat-shell"):
        history_drawer = ui.element("aside").classes("sov-history-drawer")
        history_drawer.set_visibility(False)

        with history_drawer:
            with ui.row().classes("w-full items-center justify-between"):
                _html('<div class="sov-panel-title">История</div>')
                ui.button(icon="o_close", on_click=lambda: history_drawer.set_visibility(False)).props(
                    'flat round dense aria-label="Закрыть историю"'
                ).classes("sov-icon-btn")
            sessions_col = ui.column().classes("w-full gap-2 sov-history-list")

        # Задачи/объёмы прямо в чат-шелле: ввод — командами чата («поставь задачу…»,
        # «запиши объём…»), просмотр — здесь, рядом, без ухода в админ-консоль.
        work_drawer = ui.element("aside").classes("sov-history-drawer")
        work_drawer.set_visibility(False)
        with work_drawer:
            with ui.row().classes("w-full items-center justify-between"):
                _html('<div class="sov-panel-title">Задачи и объёмы</div>')
                ui.button(icon="o_close", on_click=lambda: work_drawer.set_visibility(False)).props(
                    'flat round dense aria-label="Закрыть"'
                ).classes("sov-icon-btn")
            _html(
                '<div class="sov-muted" style="font-size:.64rem;line-height:1.4;">'
                'Ввод — командами в чате: «поставь задачу…», «запиши объём…». Здесь — просмотр.</div>'
            )
            work_body = ui.column().classes("w-full gap-1 sov-history-list")

        # W18.1: файл-вьювер — дерево RAG_Content + просмотр (текст/код/картинка/PDF).
        files_drawer = ui.element("aside").classes("sov-history-drawer")
        files_drawer.set_visibility(False)
        with files_drawer:
            with ui.row().classes("w-full items-center justify-between"):
                _html('<div class="sov-panel-title">Файлы</div>')
                ui.button(icon="o_close", on_click=lambda: files_drawer.set_visibility(False)).props(
                    'flat round dense aria-label="Закрыть"'
                ).classes("sov-icon-btn")
            files_tree_box = ui.column().classes("w-full gap-0").style("max-height:38%;overflow:auto;")
            ui.separator().style("border-color:var(--border);margin:6px 0;")
            files_view_box = ui.column().classes("w-full gap-1 sov-history-list")
            with files_view_box:
                _html('<div class="sov-muted" style="font-size:.62rem;">Выбери файл в дереве для просмотра.</div>')

        with ui.element("main").classes("sov-chat-main"):
            with ui.row().classes("sov-chat-topbar"):
                with ui.row().classes("items-center gap-2"):
                    ui.button(icon="o_history", on_click=lambda: _toggle_history()).props(
                        'flat round dense aria-label="История чата"'
                    ).classes("sov-icon-btn")
                    ui.button(icon="o_checklist", on_click=lambda: _toggle_work()).props(
                        'flat round dense aria-label="Задачи и объёмы"'
                    ).classes("sov-icon-btn")
                    ui.button(icon="o_folder_open", on_click=lambda: _toggle_files()).props(
                        'flat round dense aria-label="Файлы"'
                    ).classes("sov-icon-btn")
                    _html('<div class="sov-chat-title">С.О.В.У.Ш.К.А.</div>')
                    _html('<div class="sov-chat-subtitle">нормативный RAG-диспетчер</div>')
                with ui.row().classes("items-center gap-2"):
                    # W17.1: режим объекта — сужает ретрив к датасетам проекта; «— весь RAG —» = обычный поиск.
                    project_select = ui.select({0: "— весь RAG —"}, value=0).props(
                        'dense outlined options-dense aria-label="Объект (режим проекта)"'
                    ).style("min-width:150px;font-size:.66rem;background:var(--input-bg);")
                    project_select.on(
                        "update:model-value",
                        lambda e: project_state.__setitem__("id", project_select.value or None),
                    )

                    async def _load_projects():
                        data = await api_get("/api/projects") or {}
                        opts = {0: "— весь RAG —"}
                        for p in (data.get("projects") or []):
                            if p.get("id") and p.get("status") != "archived":
                                opts[int(p["id"])] = f"🏗 {p.get('name', 'объект')}"
                        project_select.set_options(opts, value=project_state["id"] or 0)

                    asyncio.create_task(_load_projects())
                    # W17.5: КАРТА ОБЪЕКТА — паспорт выбранного объекта.
                    ui.button(icon="o_dashboard", on_click=lambda: _open_dossier()).props(
                        'flat round dense aria-label="Карта объекта"'
                    ).classes("sov-icon-btn")
                    mode_chip = ui.label("RAG").classes("sov-chip")
                    validation_chip = ui.label("CRAG ON").classes("sov-chip")
                    ui.button(icon="o_delete_sweep", on_click=lambda: _clear_chat()).props(
                        'flat round dense aria-label="Очистить чат"'
                    ).classes("sov-icon-btn")

            chat_scroll = ui.scroll_area().classes("sov-chat-scroll")
            with chat_scroll:
                chat_column = ui.column().classes("sov-chat-thread")
                with chat_column:
                    _html('<div class="chat-msg-sys">Система активирована. Ожидание запросов.</div>')

            indexing_banner = ui.label("").classes("sov-indexing-banner")
            indexing_banner.set_visibility(False)

            # Скрепка чата: прикрепить документ — быстрая сверка / индексация (W11.8)
            attach_state = {"id": None, "name": "", "mode": ""}
            with ui.dialog() as attach_dialog, ui.card().classes("sov-control-card").style("min-width:360px"):
                _html('<div class="section-title">Прикрепить документ</div>')
                attach_mode = ui.toggle(
                    {"quick": "Быстрая сверка (таблицы)", "index": "Индексация (в базу)"},
                    value="quick",
                ).props("no-caps")
                ui.label(
                    "Быстрая сверка: парс таблиц без векторов — сразу можно «сверь с ВОР». "
                    "Индексация: документ войдёт в базу знаний (поиск в чате)."
                ).style("font-size:.66rem;color:var(--dim);")
                ui.upload(
                    auto_upload=True,
                    on_upload=lambda e: asyncio.create_task(_do_attach(e)),
                ).props("flat accept=.xlsx,.xls,.csv,.pdf,.docx").classes("w-full")
                attach_status = ui.label("").style("font-size:.7rem;color:var(--ok);")

            async def _do_attach(e):
                try:
                    content = e.content.read()
                except Exception:
                    ui.notify("Не удалось прочитать файл", type="negative")
                    return
                attach_status.text = f"Загрузка «{e.name}»…"
                add_log(f"[СКРЕПКА] {e.name} mode={attach_mode.value}")
                d = await api_post_file("/api/rag/attach", content, e.name, params={"mode": attach_mode.value})
                if not isinstance(d, dict):
                    ui.notify(last_api_error_text("Не удалось прикрепить файл"), type="negative")
                    attach_status.text = ""
                    return
                attach_state.update({"id": d.get("attachment_id"), "name": d.get("name", e.name), "mode": d.get("mode")})
                if d.get("mode") == "quick":
                    msg = f"📎 {d.get('name')} — {d.get('rows', 0)} строк, готово к сверке"
                else:
                    msg = f"📎 {d.get('name')} — индексируется в «{d.get('dataset_name', 'базу')}»"
                attach_chip.text = msg
                attach_chip.visible = True
                attach_status.text = "Готово"
                attach_dialog.close()
                ui.notify(msg, type="positive")
                if d.get("mode") == "quick":
                    ui.notify("Спроси «сверь это с ВОР» — файл уже участвует в сверке", type="info")

            with ui.element("div").classes("sov-composer") as composer_box:
                chat_input = ui.textarea(
                    placeholder="Спросить по нормативам, проекту или базе знаний..."
                ).classes("sov-composer-input").props("rows=2 autogrow borderless")
                attach_chip = ui.label("").style("font-size:.7rem;color:var(--accent);font-weight:700;")
                attach_chip.visible = False
                with ui.row().classes("sov-composer-actions"):
                    with ui.row().classes("sov-guard-controls"):
                        validation_sw = ui.switch("Т.О.С.К.А.", value=True).props("dense")
                        validation_state = ui.label("ON").classes("sov-chip")
                    # W11.17: палитра /-команд (как у взрослых)
                    with ui.button(icon="o_terminal").props("no-caps flat round").tooltip("Команды /"):
                        with ui.menu():
                            cmd_menu_box = ui.column().classes("gap-0")
                            with cmd_menu_box:
                                ui.menu_item("Загрузка команд…", on_click=lambda: None)
                    ui.button(
                        icon="o_attach_file",
                        on_click=lambda: attach_dialog.open(),
                    ).props("no-caps flat round").tooltip("Прикрепить документ")
                    advanced_btn = ui.button(
                        "Расширенный запрос",
                        icon="o_tune",
                        on_click=lambda: advanced_dialog.open(),
                    ).props("no-caps flat")
                    send_btn = ui.button(
                        "Отправить",
                        icon="o_send",
                        on_click=lambda: asyncio.create_task(send_chat()),
                    ).props("no-caps")

        # Резиновый layout: разделитель между чатом и артефактами (таскать по ширине).
        ui.element("div").classes("sov-resize-divider")

        with ui.element("aside").classes("sov-artifacts-panel"):
            with ui.row().classes("w-full items-center justify-between"):
                _html('<div class="sov-panel-title">Артефакты</div>')
                _html('<span class="sov-chip sov-chip-soft">live</span>')
            artifact_panel = ui.column().classes("sov-artifacts-body")
            with artifact_panel:
                _html(
                    '<div class="sov-artifact-empty">'
                    '<div class="sov-artifact-empty-title">Пока пусто</div>'
                    '<div class="sov-muted">Структурированные ответы, таблицы, SVG и диаграммы появятся здесь.</div>'
                    '</div>'
                )

    with ui.dialog() as advanced_dialog:
        with ui.card().classes("sov-advanced-dialog"):
            with ui.row().classes("w-full items-center justify-between"):
                with ui.column().classes("gap-0"):
                    _html('<div class="sov-panel-title">Расширенный запрос</div>')
                    _html('<div class="sov-muted">формат, датасет, стиль и образец выдачи</div>')
                ui.button(icon="o_close", on_click=advanced_dialog.close).props('flat round dense aria-label="Закрыть"').classes("sov-icon-btn")

            with ui.scroll_area().classes("sov-advanced-scroll"):
                with ui.column().classes("w-full gap-3"):
                    with ui.card().classes("sov-control-card"):
                        _html('<div class="section-title">Формат выдачи</div>')
                        format_hint_lbl = ui.label(OUTPUT_FORMATS["text"][1]).classes("sov-muted")
                        format_btns = {}
                        with ui.grid(columns=4).classes("w-full gap-2"):
                            for key, (label, hint) in OUTPUT_FORMATS.items():
                                btn = ui.button(label).props("no-caps flat").classes("sov-format-btn")
                                format_btns[key] = btn

                    with ui.card().classes("sov-control-card"):
                        _html('<div class="section-title">Параметры форматов</div>')
                        mermaid_opts_row = ui.column().classes("w-full gap-2")
                        with mermaid_opts_row:
                            mermaid_type = ui.select(
                                [
                                    "flowchart TD",
                                    "flowchart LR",
                                    "sequenceDiagram",
                                    "erDiagram",
                                    "gantt",
                                    "classDiagram",
                                    "mindmap",
                                ],
                                value="flowchart TD",
                                label="Тип диаграммы",
                            ).classes("w-full")
                        mermaid_opts_row.set_visibility(False)

                        svg_opts_row = ui.column().classes("w-full gap-2")
                        with svg_opts_row:
                            svg_type = ui.select(
                                [
                                    "Аксонометрическая схема",
                                    "План помещения",
                                    "Функциональная схема",
                                    "Принципиальная схема",
                                    "Организационная структура",
                                    "Диаграмма потоков",
                                ],
                                value="Функциональная схема",
                                label="Тип SVG",
                            ).classes("w-full")
                            svg_size = ui.select(
                                ["800×600", "1200×800", "600×400", "1600×900"],
                                value="800×600",
                                label="Размер",
                            ).classes("w-full")
                        svg_opts_row.set_visibility(False)

                        spec_opts_row = ui.column().classes("w-full gap-2")
                        with spec_opts_row:
                            spec_type = ui.select(
                                [
                                    "Спецификация оборудования (по ГОСТ 21.110)",
                                    "Ведомость чертежей (ГОСТ 21.101)",
                                    "Ведомость ссылочных документов",
                                    "Спецификация материалов",
                                    "Перечень элементов (ПЭ3)",
                                ],
                                value="Спецификация оборудования (по ГОСТ 21.110)",
                                label="Тип спецификации",
                            ).classes("w-full")
                            spec_group = ui.switch("Группировать по разделам")
                            spec_gost = ui.switch("Строгий формат ГОСТ", value=True)
                        spec_opts_row.set_visibility(False)

                        schema_opts_row = ui.column().classes("w-full gap-2")
                        with schema_opts_row:
                            schema_depth = ui.number("Глубина вложенности", value=3, min=1, max=6, step=1).classes("w-full")
                            schema_format = ui.select(
                                ["JSON дерево", "Маркированный список", "Нумерованный список", "YAML"],
                                value="JSON дерево",
                                label="Формат схемы",
                            ).classes("w-full")
                        schema_opts_row.set_visibility(False)

                        template_row = ui.column().classes("w-full gap-2")
                        with template_row:
                            ui.label("Файл-образец: JSON, CSV или XLSX").classes("sov-muted")
                            ui.upload(
                                auto_upload=True,
                                on_upload=lambda e: asyncio.create_task(load_output_template(e)),
                            ).props("flat accept=.json,.csv,.xlsx").classes("w-full")
                            template_lbl = ui.label("").style("color:var(--ok);font-size:.72rem;")
                            template_preview = _html("").classes("sov-template-preview")
                        template_row.set_visibility(False)

                    with ui.card().classes("sov-control-card"):
                        _html('<div class="section-title">Детали запроса</div>')
                        detail_dataset = ui.select([], label="Датасет").classes("w-full")
                        detail_depth = ui.select(
                            [
                                "Кратко (1-2 абзаца)",
                                "Стандарт (3-5 абзацев)",
                                "Подробно (развёрнутый ответ)",
                                "Максимум (полный анализ)",
                            ],
                            value="Стандарт (3-5 абзацев)",
                            label="Детальность",
                        ).classes("w-full")
                        detail_lang = ui.select(
                            [
                                "Русский (технический)",
                                "Русский (нормативный ГОСТ)",
                                "Краткие тезисы",
                                "Для презентации",
                            ],
                            value="Русский (технический)",
                            label="Стиль",
                        ).classes("w-full")
                        reranker_sw = ui.switch("Реранкер", value=False)
                        detail_extra = ui.textarea(label="Дополнительные требования").props("rows=3").classes("w-full")

                    with ui.card().classes("sov-control-card"):
                        with ui.row().classes("w-full items-center justify-between"):
                            _html('<div class="section-title">Промпт</div>')
                            ui.button(icon="o_refresh", on_click=lambda: _update_prompt_preview()).props("flat round dense").classes("sov-icon-btn")
                        prompt_preview = _html("").classes("sov-prompt-preview")

            with ui.row().classes("w-full justify-end gap-2"):
                ui.button("Закрыть", on_click=advanced_dialog.close).props("no-caps flat")
                apply_btn = ui.button(
                    "Применить и отправить",
                    icon="o_send",
                    on_click=lambda: asyncio.create_task(send_with_form()),
                ).props("no-caps")

    def select_format(key: str):
        out_mode_val["v"] = key
        label, hint = OUTPUT_FORMATS[key]
        format_hint_lbl.set_text(hint)
        for fmt_key, btn in format_btns.items():
            if fmt_key == key:
                btn.classes(add="sov-format-btn-active")
            else:
                btn.classes(remove="sov-format-btn-active")
        mermaid_opts_row.set_visibility(key == "mermaid")
        svg_opts_row.set_visibility(key == "svg")
        spec_opts_row.set_visibility(key == "spec")
        schema_opts_row.set_visibility(key == "schema")
        template_row.set_visibility(key == "template")
        _html_set_artifact_mode(label, hint)
        _update_prompt_preview()

    for _key in OUTPUT_FORMATS:
        format_btns[_key].on("click", lambda k=_key: select_format(k))

    async def _load_datasets_select():
        await refresh_samovar()
        names = [s.get("folder", "") for s in state["sources"]]
        detail_dataset.options = ["(все датасеты)"] + names
        detail_dataset.value = "(все датасеты)"
        # W5.7-v2: переход из графа знаний — /classic?dataset=<папка> предвыбирает фильтр.
        try:
            preset = (context.client.request.query_params.get("dataset") or "").strip()
            if preset and preset in names:
                detail_dataset.value = preset
                ui.notify(f"Фильтр из графа: {preset}", type="info")
        except Exception:
            pass
        detail_dataset.update()

    asyncio.create_task(_load_datasets_select())

    def _toggle_history():
        work_drawer.set_visibility(False)
        files_drawer.set_visibility(False)
        history_drawer.set_visibility(not history_drawer.visible)
        if history_drawer.visible:
            asyncio.create_task(_load_sessions())

    def _toggle_work():
        history_drawer.set_visibility(False)
        files_drawer.set_visibility(False)
        work_drawer.set_visibility(not work_drawer.visible)
        if work_drawer.visible:
            asyncio.create_task(_refresh_work())

    async def _refresh_work():
        """Просмотр задач/объёмов в чат-шелле. Данные те же, что у чат-команд и
        вкладок ЗАДАЧИ/ОБЪЁМЫ (/api/tasks, /api/field) — без LLM."""
        tdata = await api_get("/api/tasks?limit=100") or {}
        tasks = [t for t in (tdata.get("tasks") or []) if t.get("status") in ("open", "in_progress")]
        fdata = await api_get("/api/field?limit=50") or {}
        entries = [e for e in (fdata.get("entries") or []) if e.get("status") == "confirmed"][:20]
        work_body.clear()
        with work_body:
            ui.label("Открытые задачи").classes("section-title")
            if not tasks:
                ui.label("нет открытых задач").style("font-size:.68rem;color:var(--dim);")
            for t in tasks[:30]:
                with ui.row().classes("w-full items-center gap-2").style(
                    "border-bottom:1px dashed var(--border);padding:3px 0;"
                ):
                    ui.label(f"#{t.get('id','?')} {t.get('title', '—')}").classes("flex-1").style("font-size:.74rem;")
                    ui.label(str(t.get("status", ""))).style("font-size:.62rem;color:var(--dim);width:90px;flex-shrink:0;")
            ui.label("Объёмы (подтверждённые)").classes("section-title").style("margin-top:12px;")
            if not entries:
                ui.label("журнал пуст").style("font-size:.68rem;color:var(--dim);")
            for e in entries:
                with ui.row().classes("w-full items-center gap-2").style(
                    "border-bottom:1px dashed var(--border);padding:3px 0;"
                ):
                    ui.label(f"#{e.get('id','?')} {e.get('position', '—')}").classes("flex-1").style("font-size:.72rem;")
                    ui.label(f"{e.get('volume', '')} {e.get('unit', '')}").style(
                        "font-size:.72rem;font-weight:700;width:110px;text-align:right;flex-shrink:0;"
                    )

    # W18.1: файл-вьювер (дерево RAG_Content + просмотр текст/код/картинка/PDF).
    def _toggle_files():
        history_drawer.set_visibility(False)
        work_drawer.set_visibility(False)
        files_drawer.set_visibility(not files_drawer.visible)
        if files_drawer.visible:
            asyncio.create_task(_load_file_tree())

    # Состояние файл-дерева: карта «папка ли» + множество уже-дозагруженных папок
    # + ссылка на сам ui.tree (для ленивой дозагрузки поддеревьев по раскрытию).
    _file_state: dict = {"is_dir": {}, "loaded": set(), "tree": None}
    _LAZY = "\x00lazy"  # маркер заглушки-ребёнка нераскрытой папки

    def _mk_node(api_node: dict) -> dict:
        path = api_node.get("path")
        nid = path if path else "/"
        _file_state["is_dir"][nid] = bool(api_node.get("dir"))
        node = {"id": nid, "label": api_node.get("name", "?")}
        kids = api_node.get("children")
        if kids:
            node["children"] = [_mk_node(c) for c in kids]
            _file_state["loaded"].add(nid)
        elif api_node.get("dir"):
            # заглушка → у папки появляется стрелка раскрытия (ленивая дозагрузка)
            node["children"] = [{"id": nid + _LAZY, "label": "…"}]
        return node

    def _find_node(nodes: list, nid: str):
        for n in nodes:
            if n.get("id") == nid:
                return n
            found = _find_node(n.get("children") or [], nid)
            if found is not None:
                return found
        return None

    async def _load_file_tree():
        files_tree_box.clear()
        _file_state.update({"is_dir": {}, "loaded": set(), "tree": None})
        data = await api_get("/api/rag/tree?depth=2")
        if not isinstance(data, dict):
            with files_tree_box:
                _html('<div class="sov-muted" style="font-size:.62rem;">Не удалось загрузить дерево файлов.</div>')
            return
        root = _mk_node(data)
        with files_tree_box:
            tree = ui.tree([root], label_key="label", node_key="id",
                           on_select=lambda e: _on_file_select(e.value),
                           on_expand=lambda e: asyncio.create_task(_on_tree_expand(e.value)))
            _file_state["tree"] = tree
            tree.expand([root["id"]])

    async def _on_tree_expand(expanded: list) -> None:
        # Дозагружаем детей для раскрытых папок, которых ещё не грузили (depth=1).
        from urllib.parse import quote
        tree = _file_state["tree"]
        if tree is None:
            return
        changed = False
        for nid in (expanded or []):
            if (not nid or nid.endswith(_LAZY) or nid in _file_state["loaded"]
                    or _file_state["is_dir"].get(nid) is False):
                continue
            data = await api_get(f"/api/rag/tree?path={quote(nid)}&depth=1")
            if not isinstance(data, dict):
                continue
            node = _find_node(tree._props["nodes"], nid)
            if node is None:
                continue
            node["children"] = [_mk_node(c) for c in (data.get("children") or [])]
            _file_state["loaded"].add(nid)
            changed = True
        if changed:
            tree.update()

    def _on_file_select(path) -> None:
        if path and not path.endswith(_LAZY) and not _file_state["is_dir"].get(path, False):
            asyncio.create_task(_view_file(path))

    async def _view_file(path: str) -> None:
        from urllib.parse import quote
        files_view_box.clear()
        ext = ("." + path.rsplit(".", 1)[-1].lower()) if "." in path else ""
        url = f"/lite-api/rag/file/raw?path={quote(path)}"
        with files_view_box:
            ui.label(path).style("font-size:.62rem;color:var(--accent);font-weight:700;word-break:break-all;")
            if ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
                ui.image(url).style("width:100%;border:1px solid var(--border);border-radius:6px;")
            elif ext == ".pdf":
                _html(f'<iframe src="{url}" style="width:100%;height:520px;border:1px solid var(--border);border-radius:6px;"></iframe>')
            elif ext in (".txt", ".md", ".json", ".jsonl", ".csv", ".tsv", ".xml", ".yaml",
                         ".yml", ".log", ".html", ".svg", ".py", ".ini", ".cfg", ".sql"):
                d = await api_get(f"/api/rag/file/text?path={quote(path)}")
                if isinstance(d, dict):
                    ui.codemirror(d.get("content", ""), language=None).props("readonly").style(
                        "width:100%;height:480px;"
                    )
                else:
                    _html('<div class="sov-muted">Не удалось прочитать файл (или он бинарный).</div>')
            else:
                ui.link("Скачать файл", url).props("target=_blank").style("font-size:.72rem;color:var(--accent);")

    # W17.5: КАРТА ОБЪЕКТА — паспорт объекта (досье), собирается из /api/projects/{id}/dossier (0 LLM).
    async def _open_dossier():
        pid = project_state["id"]
        if not pid:
            ui.notify("Сначала выбери объект в селекторе сверху", type="info")
            return
        d = await api_get(f"/api/projects/{pid}/dossier")
        if not isinstance(d, dict):
            ui.notify(last_api_error_text("Не удалось загрузить карту объекта"), type="negative")
            return
        # все датасеты тянем ДО построения диалога: await внутри `with ui...` рвёт слот-контекст.
        all_ds = await api_get("/api/rag/datasets") or []
        with ui.dialog() as dlg, ui.card().classes("sov-advanced-dialog"):
            proj = d.get("project", {})
            with ui.row().classes("w-full items-center justify-between"):
                ui.label(f"КАРТА ОБЪЕКТА — {proj.get('name', '')}").classes("sov-panel-title")
                ui.button(icon="o_close", on_click=dlg.close).props(
                    'flat round dense aria-label="Закрыть"'
                ).classes("sov-icon-btn")
            ui.label(f"{proj.get('code') or ''} · {proj.get('address') or ''} · статус: {proj.get('status') or ''}").style(
                "font-size:.62rem;color:var(--dim);"
            )
            with ui.scroll_area().classes("sov-advanced-scroll"):
                with ui.column().classes("w-full gap-2"):
                    para = d.get("para", {})
                    with ui.row().classes("gap-2"):
                        for lbl, val in (("активные задачи", para.get("active", 0)),
                                         ("нормативы", para.get("resources", 0)),
                                         ("в архиве", para.get("archive", 0))):
                            with ui.column().classes("kpi-box").style("min-width:auto;padding:8px 14px;"):
                                ui.label(str(val)).classes("kpi-val").style("font-size:1.2rem;")
                                ui.label(lbl).classes("kpi-lbl")
                    ui.label("Датасеты объекта").classes("section-title").style("margin-top:8px;")
                    ui.label("Привязка сужает поиск к объекту. «Данные»: P0 — только локально; "
                             "P1/P2 — можно в облако (быстрее). P0-данные облако не увидит.").style(
                        "font-size:.6rem;color:var(--dim);line-height:1.35;")
                    ds = d.get("datasets_in_scope", [])
                    sens_by_id = {x.get("id"): (x.get("sensitivity") or "P0") for x in all_ds}
                    bound_ids = {x.get("id") for x in ds}

                    async def _reopen_dossier():
                        dlg.close()
                        await _open_dossier()

                    async def _set_ds_sens(dsid: str, level: str):
                        from urllib.parse import quote as _q
                        r = await api_patch(f"/api/rag/datasets/{_q(dsid, safe='')}/sensitivity?sensitivity={level}")
                        if r is not None:
                            ui.notify(f"Данные → {level}" + (" (можно в облако)" if level != "P0" else " (только локально)"),
                                      type="positive")
                        else:
                            ui.notify(last_api_error_text("Не удалось сменить уровень"), type="negative")

                    async def _unbind_ds(dsid: str):
                        r = await api_delete(f"/api/projects/{pid}/links", data={"kind": "dataset", "ref": dsid})
                        if r is not None:
                            ui.notify("Отвязано", type="positive")
                            await _reopen_dossier()
                        else:
                            ui.notify(last_api_error_text("Не удалось отвязать"), type="negative")

                    async def _bind_ds(dsid: str):
                        if not dsid:
                            return
                        r = await api_post(f"/api/projects/{pid}/links", {"kind": "dataset", "ref": dsid})
                        if r is not None:
                            ui.notify("Привязано", type="positive")
                            await _reopen_dossier()
                        else:
                            ui.notify(last_api_error_text("Не удалось привязать"), type="negative")

                    if not ds:
                        ui.label("датасеты ещё не привязаны — привяжи ниже ↓").style("font-size:.66rem;color:var(--dim);")
                    for x in ds:
                        _did = x.get("id")
                        with ui.row().classes("w-full items-center gap-2").style("flex-wrap:nowrap;"):
                            ui.label(f"• {x.get('name')} — {x.get('files', 0)}ф / {x.get('chunks', 0)}ч").style(
                                "font-size:.7rem;flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;")
                            ui.select(["P0", "P1", "P2"], value=sens_by_id.get(_did, "P0"),
                                      on_change=lambda e, i=_did: asyncio.create_task(_set_ds_sens(i, e.value))).props(
                                "dense outlined options-dense").style("font-size:.62rem;min-width:64px;")
                            ui.button(icon="o_link_off", on_click=lambda i=_did: asyncio.create_task(_unbind_ds(i))).props(
                                'flat round dense aria-label="Отвязать"').classes("sov-icon-btn")
                    # привязать новый датасет
                    _unbound = {x.get("id"): f"{x.get('name')} [{x.get('sensitivity') or 'P0'}]"
                                for x in all_ds if x.get("id") not in bound_ids}
                    if _unbound:
                        with ui.row().classes("w-full items-center gap-2").style("margin-top:4px;"):
                            _bind_sel = ui.select(_unbound, label="привязать датасет…").props(
                                "dense outlined options-dense").style("font-size:.62rem;flex:1;min-width:0;")
                            ui.button("Привязать", icon="o_add_link",
                                      on_click=lambda: asyncio.create_task(_bind_ds(_bind_sel.value))).props(
                                "no-caps flat dense").style("font-size:.62rem;color:var(--accent);")
                    ui.label("Открытые задачи").classes("section-title").style("margin-top:8px;")
                    tasks = d.get("open_tasks", [])
                    if not tasks:
                        ui.label("нет").style("font-size:.66rem;color:var(--dim);")
                    for t in tasks[:20]:
                        ui.label(f"#{t.get('id')} {t.get('title')} [{t.get('status')}]").style("font-size:.7rem;")
                    vol = d.get("volumes", {})
                    ui.label(f"Объёмы (подтв.): {vol.get('groups', 0)} позиций · всего {vol.get('total', 0)}").classes(
                        "section-title"
                    ).style("margin-top:8px;")
                    for r in (vol.get("by_position") or [])[:10]:
                        ui.label(f"• {r.get('position')}: {r.get('total')} {r.get('unit', '')}").style("font-size:.7rem;")
                    # W17.4 — решения проекта (RFI-стиль)
                    decisions = d.get("decisions") or []
                    if decisions:
                        ui.label(f"Решения по объекту ({d.get('decisions_count', len(decisions))})").classes(
                            "section-title"
                        ).style("margin-top:8px;")
                        _dmark = {"open": "◌", "decided": "●", "superseded": "⊘"}
                        for dec in decisions[:10]:
                            line = f"{_dmark.get(dec.get('status'), '·')} #{dec.get('id')} {dec.get('decision', '')[:140]}"
                            ui.label(line).style("font-size:.7rem;")
                            if dec.get("rationale"):
                                ui.label(f"   ↳ {dec['rationale'][:160]}").style("font-size:.64rem;color:var(--dim);")
                    ui.label(
                        f"Связи: {d.get('edges_count', 0)} · Заметки: {d.get('notes_count', 0)}"
                    ).classes("section-title").style("margin-top:8px;")
                    bim = d.get("bim")
                    if isinstance(bim, dict) and bim:
                        ui.label("BIM: " + ", ".join(f"{k}={v}" for k, v in list(bim.items())[:4])).style(
                            "font-size:.66rem;color:var(--dim);"
                        )
                    # W17.3 — классификационный хребет (Floor→System→Category)
                    clf = d.get("classification") or {}
                    clf_tot = clf.get("totals") or {}
                    if clf_tot.get("elements"):
                        ui.label(
                            f"Классификационный хребет: {clf_tot.get('elements', 0)} элементов · "
                            f"этажей {clf_tot.get('floors', 0)} · систем {clf_tot.get('systems', 0)} · "
                            f"категорий {clf_tot.get('categories', 0)}"
                        ).classes("section-title").style("margin-top:8px;")
                        for f in (clf.get("top_floors") or [])[:6]:
                            systems = ", ".join(f.get("top_systems") or []) or "—"
                            ui.label(f"• {f.get('floor')} ({f.get('elements', 0)}): {systems}").style("font-size:.7rem;")
                    # W17.3 — состояния документов CDE (ISO 19650)
                    cde = d.get("cde") or {}
                    if any(cde.values()):
                        ui.label(
                            "Документы (CDE): " + " · ".join(f"{k} {cde.get(k, 0)}" for k in ("WIP", "Shared", "Published", "Archived"))
                        ).classes("section-title").style("margin-top:8px;")
                    # W17.3 — захватки (LBS-хабы) из журнала объёмов
                    lbs = d.get("lbs") or []
                    if lbs:
                        ui.label("Захватки (факт. объёмы)").classes("section-title").style("margin-top:8px;")
                        for h in lbs[:10]:
                            ui.label(
                                f"• {h.get('zahvatka')}: {h.get('total', 0)} ({h.get('entries', 0)} записей)"
                            ).style("font-size:.7rem;")
        dlg.open()

    async def _load_sessions():
        sessions_col.clear()
        data = await api_get("/api/chat/sessions?limit=40")
        if not data:
            with sessions_col:
                _html('<div class="sov-muted" style="padding:14px;">Нет сохранённых сессий</div>')
            return
        with sessions_col:
            for session in data:
                _render_session_card(session)

    def _render_session_card(session: dict):
        sid = session["session_id"]
        first_q = session.get("first_question") or "Без названия"
        msg_count = session.get("msg_count", 0)
        started_at = (session.get("started_at") or "")[:16].replace("T", " ")
        with ui.element("button").classes("sov-session-card") as card:
            _html(f'<span class="sov-session-title">{esc(first_q[:90])}</span>')
            _html(f'<span class="sov-session-meta">{esc(started_at)} · {msg_count} сообщ.</span>')

        async def _open(session_id=sid, el=card):
            await _open_session(session_id, el)

        card.on("click", _open)

    async def _open_session(session_id: str, el=None):
        add_log(f"[ИСТОРИЯ] Загружаю сессию {session_id[:8]}...")
        msgs = await api_get(f"/api/chat/history?session_id={session_id}")
        if msgs is None:
            add_log("[ИСТОРИЯ] Ошибка загрузки сессии")
            return
        state["chat_history"] = msgs
        state["load_session_id"] = session_id
        state["session_id"] = session_id
        if selected_session_card["el"]:
            selected_session_card["el"].classes(remove="sov-session-card-active")
        if el:
            el.classes(add="sov-session-card-active")
            selected_session_card["el"] = el
        _render_chat_history("Сессия загружена из истории.")
        history_drawer.set_visibility(False)
        chat_scroll.scroll_to(percent=1)

    def _source_label(source) -> str:
        if isinstance(source, dict):
            return str(source.get("file") or source.get("name") or source)
        return str(source)

    def _render_source_tags(srcs: list, crag: str = "", meta: dict | None = None):
        if not srcs and not crag and not meta:
            return
        with ui.row().classes("msg-srcs"):
            for source in srcs:
                ui.label(_source_label(source)).classes("src-tag")
            if crag:
                if crag == "VERIFIED":
                    cls = "src-tag"
                    label = "Т.О.С.К.А.: VERIFIED"
                elif crag == "UNVALIDATED":
                    cls = "src-tag src-tag-warn"
                    label = "Т.О.С.К.А.: OFF"
                else:
                    cls = "src-tag src-tag-err"
                    label = f"Т.О.С.К.А.: {crag}"
                ui.label(label).classes(cls)
            if meta:
                query_route = meta.get("query_route") if isinstance(meta.get("query_route"), dict) else {}
                kot = query_route.get("kot") if isinstance(query_route.get("kot"), dict) else {}
                trace = meta.get("retrieval_trace") if isinstance(meta.get("retrieval_trace"), dict) else {}
                cache = meta.get("cache") or "miss"
                validation = meta.get("validation") if isinstance(meta.get("validation"), dict) else {}
                if kot:
                    kdf = kot.get("dataset_filter") or "AUTO"
                    conf = kot.get("confidence", 0)
                    ui.label(f"KOT {kdf} {conf}").classes("src-tag")
                if trace:
                    mode = str(trace.get("mode") or "vector").upper()
                    quality = trace.get("quality_status") or trace.get("quality", {}).get("status") or "?"
                    ui.label(f"{mode} {quality}").classes("src-tag")
                    context_window = trace.get("context_window") if isinstance(trace.get("context_window"), dict) else {}
                    if context_window:
                        expanded = context_window.get("expanded_count", 0)
                        total = context_window.get("input_count", 0)
                        cls = "src-tag" if expanded else "src-tag src-tag-warn"
                        ui.label(f"CTX {expanded}/{total}").classes(cls)
                ui.label(f"CACHE {str(cache).upper()}").classes("src-tag")
                if validation:
                    ui.label("VALIDATOR ON" if validation.get("enabled") else "VALIDATOR OFF").classes(
                        "src-tag" if validation.get("enabled") else "src-tag src-tag-warn"
                    )
                history_id = meta.get("history_id")
                if history_id:
                    async def _feedback(status: str):
                        result = await api_post(f"/api/chat/history/{history_id}/feedback", {"feedback": status})
                        if result:
                            ui.notify("Оценка сохранена", type="positive")
                        else:
                            ui.notify(last_api_error_text("Не удалось сохранить оценку"), type="warning")

                    ui.button(
                        icon="thumb_up",
                        on_click=lambda: asyncio.create_task(_feedback("correct")),
                    ).props("flat dense round").tooltip("Ответ корректен")
                    ui.button(
                        "Плохой ответ",
                        icon="thumb_down",
                        on_click=lambda: asyncio.create_task(_feedback("bad_answer")),
                    ).props("flat dense").tooltip("Плохой ответ: сохранить для разбора")
                    ui.button(
                        icon="travel_explore",
                        on_click=lambda: asyncio.create_task(_feedback("wrong_dataset")),
                    ).props("flat dense round").tooltip("Источник не из того датасета")

    def _render_suggestions(meta: dict | None):
        if not meta:
            return
        questions = meta.get("clarifying_questions") or []
        filters = meta.get("suggested_filters") or []
        class_suggestions = meta.get("class_suggestions") or []
        if not questions and not filters and not class_suggestions:
            return

        with ui.column().classes("w-full gap-2 mt-2 pt-2 border-t border-dashed border-gray-700"):
            ui.label("Подсказки для уточнения:").classes("text-xs font-semibold text-gray-400 uppercase tracking-wider")

            # ADR-12 мультикласс: вопрос задел несколько классов — предложим переспрос в их области.
            if class_suggestions:
                with ui.row().classes("gap-2 items-center flex-wrap"):
                    ui.label("Посмотреть как:").classes("text-xs text-gray-500")
                    for cs in class_suggestions:
                        def _make_click_class(query=cs.get("query", ""), filt=cs.get("dataset_filter")):
                            async def click_class():
                                if filt and detail_dataset.options:
                                    if filt in detail_dataset.options:
                                        detail_dataset.value = filt
                                    else:
                                        matched = [o for o in detail_dataset.options
                                                   if filt.lower() in str(o).lower()]
                                        if matched:
                                            detail_dataset.value = matched[0]
                                chat_input.value = query
                                _update_prompt_preview()
                                await send_chat()
                            return click_class

                        ui.button(cs.get("label", "?"), on_click=_make_click_class()).props(
                            "outline dense size=sm color=secondary"
                        ).classes("text-xs normal-case")

            if filters:
                with ui.row().classes("gap-2 items-center flex-wrap"):
                    ui.label("Выбрать датасет:").classes("text-xs text-gray-500")
                    for f in filters:
                        f_name = f
                        if f == "NTD_FIRE":
                            f_name = "🔥 Пожарная безопасность"
                        elif f == "NTD_ELECTRICAL":
                            f_name = "⚡ Электрика"
                        elif f == "NTD_STRUCTURAL":
                            f_name = "🏗️ Конструкции"
                        elif f == "TABLE_SMETA":
                            f_name = "📊 Сметы"
                        elif f == "GKRF":
                            f_name = "⚖️ Постановление 87 / ГК РФ"
                        elif f == "NTD":
                            f_name = "📚 Стандарты (СП/ГОСТ)"

                        def _make_click_filter(f_code=f, name=f_name):
                            async def click_filter():
                                if f_code in detail_dataset.options:
                                    detail_dataset.value = f_code
                                elif name in detail_dataset.options:
                                    detail_dataset.value = name
                                else:
                                    matched = [opt for opt in detail_dataset.options if f_code in opt or f_code.lower() in opt.lower()]
                                    if matched:
                                        detail_dataset.value = matched[0]
                                ui.notify(f"Выбран датасет: {name}", type="info")
                                _update_prompt_preview()
                            return click_filter

                        ui.button(f_name, on_click=_make_click_filter(f, f_name)).props("outline dense size=sm").classes("text-xs text-white border-blue-500")

            if questions:
                with ui.row().classes("gap-2 items-center flex-wrap"):
                    ui.label("Уточнить вопрос:").classes("text-xs text-gray-500")
                    for q in questions:
                        def _make_click_question(q_val=q):
                            async def click_question():
                                chat_input.value = q_val
                                _update_prompt_preview()
                                await send_chat()
                            return click_question

                        ui.button(q, on_click=_make_click_question(q)).props("outline dense size=sm color=primary").classes("text-xs text-left normal-case")

    def _render_excerpts(meta: dict | None):
        """Конкретные фрагменты норм/документов, на которые опёрся ответ — «вот это
        место». Раскрываемо; ссылка «открыть» ведёт в файл (W18.1) если есть путь."""
        if not meta:
            return
        excerpts = meta.get("source_excerpts") or []
        if not excerpts:
            return
        from urllib.parse import quote
        with ui.expansion(f"Цитаты из источников ({len(excerpts)})", icon="format_quote").props(
            "dense"
        ).classes("w-full mt-2").style("font-size:.66rem;"):
            for ex in excerpts:
                doc = ex.get("doc", "") or ""
                with ui.column().classes("w-full gap-1").style(
                    "border-left:2px solid var(--accent);padding:3px 0 8px 10px;margin-top:6px;"
                ):
                    with ui.row().classes("w-full items-center gap-2"):
                        ui.label(doc.rsplit("/", 1)[-1] or "источник").style(
                            "font-size:.64rem;color:var(--accent);font-weight:700;word-break:break-all;"
                        )
                        if ex.get("score") is not None:
                            ui.label(f"score {ex['score']}").style("font-size:.56rem;color:var(--dim);")
                        if "/" in doc:
                            ui.link("открыть", f"/lite-api/rag/file/raw?path={quote(doc)}").props(
                                "target=_blank"
                            ).style("font-size:.58rem;color:var(--ok);margin-left:auto;")
                    ui.label(ex.get("text", "")).style(
                        "font-size:.7rem;line-height:1.5;color:var(--text);white-space:pre-wrap;"
                    )

    def _artifact_present(ans: str, mode: str) -> bool:
        """Есть ли в ответе рендеримый артефакт (таблица/спека/диаграмма/svg)."""
        ans = ans or ""
        if mode in ("spec", "table", "structure", "template", "schema"):
            return bool(_parse_table_from_ai(ans) or _parse_json_from_ai(ans))
        if mode == "mermaid":
            return bool(_parse_mermaid_from_ai(ans))
        if mode == "svg":
            return bool(_parse_svg_from_ai(ans))
        # text/дефолт: авто-детект таблицы или диаграммы в обычном ответе
        return bool(_parse_table_from_ai(ans) or _parse_markdown_table(ans)
                    or _parse_mermaid_from_ai(ans) or _parse_svg_from_ai(ans))

    def _show_artifact(ans: str, mode: str) -> None:
        """Открыть артефакт сообщения в панели «Артефакты» (как карточка в Claude Desktop)."""
        _render_result(ans, mode if mode in OUTPUT_FORMATS else "text", artifact_panel)
        try:
            ui.run_javascript(
                "document.querySelector('.sov-artifacts-panel')?.scrollIntoView({behavior:'smooth',block:'start'})"
            )
        except Exception:
            pass

    def _artifact_button(ans: str, mode: str) -> None:
        """Кнопка-карточка артефакта в пузыре ответа (если артефакт есть)."""
        if not _artifact_present(ans, mode):
            return
        lbl = OUTPUT_FORMATS[mode][0] if (mode in OUTPUT_FORMATS and mode != "text") else "Таблица"
        ui.button(
            f"Артефакт: {lbl} — открыть",
            icon="o_table_view",
            on_click=lambda a=ans, m=mode: _show_artifact(a, m),
        ).props("no-caps flat dense").classes("sov-artifact-chip").style(
            "margin-top:6px;border:1px solid var(--border);border-radius:8px;"
            "background:var(--bg);color:var(--accent);font-size:.68rem;font-weight:700;padding:4px 10px;"
        )

    def _bubble_text(ans: str, mode: str) -> str:
        """Текст для пузыря: если есть артефакт — убираем дублирующую таблицу/код-блоки
        (они уже в артефакте), оставляем только коммент модели. Олег: «текст-то зачем?»."""
        ans = ans or ""
        if not _artifact_present(ans, mode):
            return ans
        t = re.sub(r"```.*?```", "", ans, flags=re.DOTALL)           # fenced json/svg/mermaid
        t = "\n".join(ln for ln in t.splitlines() if not ln.strip().startswith("|"))  # md-таблица
        t = re.sub(r"\n{3,}", "\n\n", t).strip()
        return t or "Готово — результат в артефакте (кнопка ниже)."

    def _render_chat_bubble(
        text: str,
        class_name: str,
        srcs: list | None = None,
        crag: str = "",
        meta: dict | None = None,
    ):
        _mode = (meta or {}).get("out_mode", "text")
        _disp = _bubble_text(str(text or ""), _mode) if (meta and "chat-msg-ai" in class_name) else str(text or "")
        with ui.element("div").classes(class_name) as bubble:
            ui.label(_disp).classes("sov-chat-message-text")
            _render_source_tags(srcs or [], crag, meta)
            if meta:
                _render_suggestions(meta)
                _render_excerpts(meta)
                if "chat-msg-ai" in class_name:
                    _artifact_button(str(text or ""), _mode)
        return bubble

    def _render_ai_placeholder(text: str):
        with ui.element("div").classes("chat-msg-ai typing") as bubble:
            label = ui.label(text).classes("sov-chat-message-text")
        return bubble, label

    def _finish_ai_placeholder(
        bubble,
        label,
        text: str,
        srcs: list | None = None,
        crag: str = "",
        error: bool = False,
        meta: dict | None = None,
    ):
        bubble.classes(remove="typing")
        if error:
            bubble.classes(add="chat-msg-error")
        _mode = (meta or {}).get("out_mode", "text")
        label.set_text(_bubble_text(str(text or ""), _mode) if (meta and not error) else str(text or ""))
        with bubble:
            _render_source_tags(srcs or [], crag, meta)
            if meta:
                _render_suggestions(meta)
                _render_excerpts(meta)
                if not error:
                    _artifact_button(str(text or ""), meta.get("out_mode", "text"))

    def _render_msg(msg):
        if msg.get("role") == "user":
            _render_chat_bubble(msg.get("text", ""), "chat-msg-user")
            return
        _render_chat_bubble(
            msg.get("text", ""),
            "chat-msg-ai",
            msg.get("srcs", []),
            msg.get("crag", ""),
            msg.get("meta"),
        )

    def _render_chat_history(system_msg: str = "История загружена."):
        chat_column.clear()
        with chat_column:
            _render_chat_bubble(system_msg, "chat-msg-sys")
            for msg in state.get("chat_history", []):
                _render_msg(msg)
            if state.get("chat_pending"):
                pending_q = state["chat_pending"].get("question", "")
                _render_chat_bubble(f"Запрос выполняется: {pending_q[:80]}", "chat-msg-ai typing")

    def _apply_loaded_session() -> bool:
        """Подхват сессии, выбранной в ИСТОРИИ. Вызывается и при построении,
        и хуком из вкладки истории (фикс «чат из истории не открывается»)."""
        if not state.get("load_session_id"):
            return False
        state["session_id"] = state["load_session_id"]
        state["load_session_id"] = None
        _render_chat_history("Сессия загружена из истории.")
        chat_scroll.scroll_to(percent=1)
        return True

    # Хук для вкладки ИСТОРИЯ: после выбора сессии чат перерисовывается сразу,
    # а не «при следующем рендере» (которого без хука не наступало).
    state["chat_reload_hook"] = _apply_loaded_session

    async def _load_history():
        if _apply_loaded_session():
            return
        if not state.get("chat_history"):
            hist = await api_get("/api/chat/history?limit=40")
            if hist:
                state["chat_history"] = hist
                _render_chat_history()
                chat_scroll.scroll_to(percent=1)

    asyncio.create_task(_load_history())

    async def load_output_template(e):
        content = e.content.read()
        fname = e.name
        try:
            if fname.endswith(".json"):
                data = json.loads(content.decode("utf-8"))
                state["output_template"] = data if isinstance(data, list) else [data]
            elif fname.endswith(".csv"):
                lines = content.decode("utf-8").strip().split("\n")
                keys = [k.strip() for k in lines[0].split(",")]
                rows = [dict(zip(keys, [v.strip() for v in row.split(",")])) for row in lines[1:] if row.strip()]
                state["output_template"] = rows
            elif fname.endswith(".xlsx"):
                import tempfile
                import openpyxl

                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
                    tf.write(content)
                    tf.flush()
                    wb = openpyxl.load_workbook(tf.name)
                    ws = wb.active
                    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
                    rows = []
                    for row in list(ws.iter_rows(min_row=2, values_only=True))[:20]:
                        rows.append(dict(zip(headers, [str(v or "") for v in row])))
                    state["output_template"] = rows
            else:
                ui.notify("Поддерживаются JSON, CSV, XLSX", type="warning")
                return

            tmpl = state["output_template"]
            template_lbl.set_text(f"{fname} · {len(tmpl)} строк")
            if tmpl:
                preview_str = json.dumps(tmpl[0], ensure_ascii=False, indent=2)
                template_preview.set_content(f'<pre>{esc(preview_str)}</pre>')
            add_log(f"[ШАБЛОН] Загружен {fname} · {len(tmpl)} строк")
            _update_prompt_preview()
        except Exception as ex:
            ui.notify(f"Ошибка парсинга: {ex}", type="negative")
            add_log(f"[ШАБЛОН] Ошибка: {ex}")

    def _build_extra_prompt(question: str) -> str:
        mode = out_mode_val["v"]
        depth_map = {
            "Кратко (1-2 абзаца)": "Ответь кратко — 1-2 абзаца.",
            "Стандарт (3-5 абзацев)": "Ответь развёрнуто — 3-5 абзацев.",
            "Подробно (развёрнутый ответ)": "Дай полный развёрнутый ответ со всеми деталями.",
            "Максимум (полный анализ)": "Проведи максимально подробный анализ. Не сокращай.",
        }
        style_map = {
            "Русский (технический)": "Пиши профессиональным техническим языком.",
            "Русский (нормативный ГОСТ)": "Пиши в нормативном стиле ГОСТ: чёткие формулировки, без лирики.",
            "Краткие тезисы": "Отвечай тезисами — каждый пункт одна мысль.",
            "Для презентации": "Формат для слайдов: заголовок + маркированный список.",
        }

        parts = []
        if depth_map.get(detail_depth.value):
            parts.append(depth_map[detail_depth.value])
        if style_map.get(detail_lang.value):
            parts.append(style_map[detail_lang.value])

        if mode == "spec":
            gost_str = " строго по форме ГОСТ 21.110-2013" if spec_gost.value else ""
            group_str = " Группируй по разделам." if spec_group.value else ""
            parts.append(
                f"\n\nВЫВЕДИ ОТВЕТ В ФОРМАТЕ СПЕЦИФИКАЦИИ{gost_str}.\n"
                f"Тип: {spec_type.value}.{group_str}\n"
                "Верни JSON-массив объектов. Обязательные поля: "
                "поз, обозначение, наименование, тип_марка, ед_изм, кол_во, масса_ед, примечание.\n"
                "Оберни в ```json ... ```"
            )
        elif mode == "schema":
            depth = int(schema_depth.value) if schema_depth.value else 3
            fmt = schema_format.value
            if fmt == "JSON дерево":
                parts.append(
                    f"\n\nВЫВЕДИ ОТВЕТ В ВИДЕ JSON-ДЕРЕВА, глубина {depth}. "
                    "Структура узла: {\"name\": str, \"children\": [...], \"desc\": str}. "
                    "Оберни в ```json ... ```"
                )
            elif fmt == "YAML":
                parts.append(f"\n\nВЫВЕДИ ОТВЕТ В ВИДЕ YAML-ДЕРЕВА, глубина {depth}. Оберни в ```yaml ... ```")
            else:
                parts.append(f"\n\nВЫВЕДИ ОТВЕТ В ВИДЕ {fmt.upper()}, глубина {depth} уровней.")
        elif mode == "structure":
            parts.append("\n\nВЫВЕДИ ОТВЕТ В ВИДЕ СТРУКТУРИРОВАННОГО JSON-ОБЪЕКТА. Оберни в ```json ... ```")
        elif mode == "table":
            parts.append("\n\nВЫВЕДИ ОТВЕТ В ВИДЕ ТАБЛИЦЫ: JSON-массив объектов. Оберни в ```json ... ```")
        elif mode == "mermaid":
            parts.append(
                f"\n\nВЫВЕДИ ОТВЕТ В ВИДЕ MERMAID-ДИАГРАММЫ типа {mermaid_type.value}. "
                "Оберни в ```mermaid ... ```. Пиши на русском, метки узлов короткие."
            )
        elif mode == "svg":
            w, h = svg_size.value.split("×") if "×" in svg_size.value else ("800", "600")
            parts.append(
                f"\n\nВЫВЕДИ ОТВЕТ В ВИДЕ SVG-СХЕМЫ ({svg_type.value}). "
                f"Размер viewBox: 0 0 {w} {h}. Оберни в ```svg ... ```"
            )
        elif mode == "template":
            tmpl = state.get("output_template")
            if tmpl:
                parts.append(
                    "\n\nОТВЕЧАЙ СТРОГО ПО СТРУКТУРЕ ОБРАЗЦА (JSON-массив).\n"
                    f"Образец:\n```json\n{json.dumps(tmpl[:3], ensure_ascii=False, indent=2)}\n```\n"
                    "Оберни в ```json ... ```"
                )
            else:
                parts.append("\n\nОТВЕЧАЙ В ВИДЕ JSON-МАССИВА ОБЪЕКТОВ. Оберни в ```json ... ```")

        if detail_extra.value and detail_extra.value.strip():
            parts.append(f"\n\nДОПОЛНИТЕЛЬНО: {detail_extra.value.strip()}")
        return " ".join(parts[:2]) + "".join(parts[2:])

    def _update_prompt_preview():
        q = chat_input.value.strip() or "[текст запроса]"
        extra = _build_extra_prompt(q)
        preview_text = (q + extra)[:1000] + ("..." if len(q + extra) > 1000 else "")
        prompt_preview.set_content(f'<pre>{esc(preview_text)}</pre>')

    chat_input.on("input", lambda: _update_prompt_preview())

    def _clear_chat():
        from sovushka.state import _new_session_id

        chat_column.clear()
        with chat_column:
            _html('<div class="chat-msg-sys">Чат очищен. Новая сессия готова.</div>')
        artifact_panel.clear()
        with artifact_panel:
            _render_empty_artifacts()
        state["chat_history"].clear()
        state["session_id"] = _new_session_id()
        state["load_session_id"] = None
        state["chat_pending"] = None
        add_log("[ЧАТ] История очищена, новая сессия")

    _sending = {"v": False}
    _resource_blocked = {"v": False, "reason": ""}

    def _indexing_summary(data: dict) -> str:
        rag = state.get("rag_health", {}) if isinstance(state.get("rag_health"), dict) else {}
        totals = rag.get("totals", {}) if isinstance(rag, dict) else {}
        indexed = totals.get("indexed_files", "—")
        pending = totals.get("pending_files", "—")
        errors = totals.get("error_files", "—")
        chunks = totals.get("chunks", "—")
        reason = data.get("chat_generation_reason") or "Индексация активна."
        if not data.get("active"):
            return f"Чат временно заблокирован защитой runtime. {reason}"
        return (
            f"Индексация активна: чат заблокирован. "
            f"indexed={indexed} · pending={pending} · errors={errors} · chunks={chunks}. "
            f"{reason}"
        )

    def _set_chat_blocked(blocked: bool, reason: str = ""):
        _resource_blocked["v"] = blocked
        _resource_blocked["reason"] = reason
        if blocked:
            advanced_btn.props("disabled")
            if not _sending["v"]:
                send_btn.props(remove="disabled")
                apply_btn.props(remove="disabled")
                chat_input.props(remove="disabled")
            composer_box.classes(add="sov-composer-blocked")
            indexing_banner.set_text(reason)
            indexing_banner.set_visibility(True)
            mode_chip.set_text("PAUSED")
            mode_chip.classes(add="sov-chip-soft")
            return
        if not _sending["v"]:
            send_btn.props(remove="disabled")
            apply_btn.props(remove="disabled")
            advanced_btn.props(remove="disabled")
            chat_input.props(remove="disabled")
        composer_box.classes(remove="sov-composer-blocked")
        indexing_banner.set_visibility(False)
        mode_chip.set_text("RAG")
        mode_chip.classes(remove="sov-chip-soft")

    def _sync_validation_ui():
        enabled = bool(validation_sw.value)
        validation_chip.set_text("CRAG ON" if enabled else "CRAG OFF")
        validation_state.set_text("ON" if enabled else "OFF")
        if enabled:
            validation_chip.classes(remove="sov-chip-warn")
            validation_state.classes(remove="sov-chip-warn")
        else:
            validation_chip.classes(add="sov-chip-warn")
            validation_state.classes(add="sov-chip-warn")

    validation_sw.on("update:model-value", lambda _e: _sync_validation_ui())

    async def _refresh_resource_gate() -> bool:
        data = await refresh_indexing_mode()
        allowed = bool(data.get("chat_generation_allowed", True)) if isinstance(data, dict) else True
        blocked = not allowed
        _set_chat_blocked(blocked, _indexing_summary(data) if blocked else "")
        return allowed

    async def _do_verify(question: str):
        """Верификация объёмов: распознать скан и открыть спец-артефакт сверки.

        Tool-действие (не LLM-ответ): сплит «скан ↔ распознанное» в панели артефактов.
        Сохранённый оператором результат = принятая выписка + ground truth для бенча.
        """
        path = _verify_path(question)
        page = _verify_page(question)
        with chat_column:
            _render_chat_bubble(question, "chat-msg-user")
            ph, ph_label = _render_ai_placeholder("Распознаю таблицу объёмов…")
        chat_scroll.scroll_to(percent=1)
        try:
            _render_artifact_loading("verify", question)
            res = await api_post("/api/verify/extract", {"path": path, "page": page, "engine": "local"})
            if not isinstance(res, dict):
                _finish_ai_placeholder(ph, ph_label, f"Не удалось распознать скан «{path}» (файл найден? см. лог).", [], "", meta={})
                return
            rows = res.get("rows") or []
            payload = {
                "token": res.get("token"),
                "rows": rows,
                "columns": res.get("columns") or [],
                "source": path,
                "page": page,
            }
            ans = json.dumps(payload, ensure_ascii=False)
            text = (
                f"Распознал таблицу объёмов: {len(rows)} строк из «{path}» (стр.{page + 1}). "
                "Сверь со сканом и подтверди в артефакте справа →"
            )
            _finish_ai_placeholder(ph, ph_label, text, [], "", meta={"out_mode": "text"})
            with ph:  # кнопка переоткрытия артефакта (как у Клода)
                ui.button(
                    "Артефакт: Верификация — открыть",
                    icon="o_table_view",
                    on_click=lambda a=ans: _show_artifact(a, "verify"),
                ).props("no-caps flat dense").classes("sov-artifact-chip").style(
                    "margin-top:6px;border:1px solid var(--border);border-radius:8px;"
                    "background:var(--bg);color:var(--accent);font-size:.68rem;font-weight:700;padding:4px 10px;"
                )
            _render_result(ans, "verify", artifact_panel)
            chat_scroll.scroll_to(percent=1)
            add_log(f"[ВЕРИФ] {path} стр.{page}: {len(rows)} строк → артефакт")
        except Exception as ex:  # любая ошибка — в чат текстом, не в тишину
            import traceback
            add_log(f"[ВЕРИФ] ОШИБКА: {ex}")
            add_log(traceback.format_exc()[:700])
            try:
                _finish_ai_placeholder(ph, ph_label, f"⚠️ Верификация упала: {type(ex).__name__}: {ex}", [], "", meta={})
            except Exception:
                with chat_column:
                    _render_chat_bubble(f"⚠️ Верификация упала: {ex}", "chat-msg-ai")

    async def _do_send(question: str):
        if _sending["v"]:
            return
        # Верификация объёмов — tool-действие: распознать скан + открыть спец-артефакт.
        if _is_verify_request(question):
            await _do_verify(question)
            return
        selected_dataset_filter = (
            detail_dataset.value
            if detail_dataset.value and detail_dataset.value != "(все датасеты)"
            else None
        )
        skip_resource_gate = should_skip_chat_resource_gate(question, selected_dataset_filter)
        if skip_resource_gate:
            _set_chat_blocked(False)
        else:
            await _refresh_resource_gate()
        _sending["v"] = True
        send_btn.props("disabled")
        apply_btn.props("disabled")
        advanced_btn.props("disabled")
        chat_input.props("disabled")
        # Авто-GOST: «собери/составь спецификацию …» → формат спеки (ГОСТ 21.110), чтобы
        # артефакт был чистой таблицей по форме, а не прозой. Не липко — селектор вернём.
        _orig_mode = out_mode_val["v"]
        if _orig_mode == "text" and _is_spec_request(question):
            out_mode_val["v"] = "spec"
        out_mode = out_mode_val["v"]

        state["chat_history"].append({"role": "user", "text": question})
        state["chat_pending"] = {"question": question, "started_at": time.time()}

        with chat_column:
            _render_chat_bubble(question, "chat-msg-user")
            ai_placeholder, ai_placeholder_label = _render_ai_placeholder("Генерирую... 0с")
        chat_scroll.scroll_to(percent=1)
        _render_artifact_loading(out_mode, question)
        add_log(f'[AI] Запрос: "{question[:60]}"')

        # Формат/стиль ответа — ОТДЕЛЬНЫМ полем (не клеим в текст вопроса): иначе бэкенд
        # видит мусор-шаблон как вопрос → авто-заметки/роутинг/ретрив плывут.
        extra_prompt = "" if skip_resource_gate else _build_extra_prompt(question)
        out_mode_val["v"] = _orig_mode  # вернуть пользователю его выбор формата (авто-GOST не липкий)
        payload = {
            "question": question,
            "output_directive": extra_prompt or None,
            "reranker_enabled": reranker_sw.value,
            "validation_enabled": bool(validation_sw.value),
            "session_id": state.get("session_id"),
        }
        if detail_dataset.value and detail_dataset.value != "(все датасеты)":
            payload["dataset_filter"] = detail_dataset.value
        if project_state["id"]:  # W17.1: режим объекта — сузить ретрив к датасетам проекта
            payload["project_id"] = project_state["id"]

        _t0 = time.monotonic()
        _stop_tick = {"v": False}

        async def _tick():
            while not _stop_tick["v"]:
                elapsed = int(time.monotonic() - _t0)
                ai_placeholder_label.set_text(f"Генерирую... {elapsed}с")
                await asyncio.sleep(1)

        _tick_task = asyncio.create_task(_tick())

        async def _gen_form_from_command(cmd: dict) -> None:
            """Создать документ по /-команде: генерация формы + скачивание файла."""
            fid = cmd.get("form_id")
            fmt = cmd.get("fmt", "xlsx")
            gen = await api_post(f"/api/forms/{fid}/generate", {"fmt": fmt})
            if not isinstance(gen, dict) or not gen.get("download"):
                ui.notify(last_api_error_text("Не удалось создать документ"), type="negative")
                return
            res = await api_get_bytes(gen["download"])
            if not res:
                ui.notify(last_api_error_text("Файл не готов"), type="negative")
                return
            data, fname = res
            ui.download(data, fname)
            ui.notify(f"Документ создан: {cmd.get('title', fid)} ({fmt})", type="positive")

        def _apply_chat_result(d: dict) -> None:
            """Применяет финальный payload (общий для стрима и нестриминга):
            форматированный ответ, источники, вердикт, артефакт."""
            ans = d.get("answer", d.get("response", "Нет ответа"))
            srcs = d.get("sources", [])
            crag = d.get("crag_status", "")
            meta = {
                "query_route": d.get("query_route") or {},
                "retrieval_trace": d.get("retrieval_trace") or {},
                "cache": d.get("cache", "miss"),
                "validation": d.get("validation") or {"enabled": bool(validation_sw.value)},
                "history_id": d.get("history_id"),
                "table_query": d.get("table_query"),
                "out_mode": out_mode,
                "clarifying_questions": d.get("clarifying_questions") or [],
                "suggested_filters": d.get("suggested_filters") or [],
                "class_suggestions": d.get("class_suggestions") or [],
                "source_excerpts": d.get("source_excerpts") or [],
            }
            state["chat_history"].append({"role": "ai", "text": ans, "srcs": srcs, "crag": crag, "meta": meta})
            _finish_ai_placeholder(ai_placeholder, ai_placeholder_label, ans, srcs, crag, meta=meta)
            _render_result(ans, out_mode, artifact_panel, table_query=d.get("table_query"))
            add_log(f"[AI] Формат:{out_mode} CRAG:{crag or 'N/A'} src:{len(srcs)}")
            # W11.17: команда «создать документ» → генерируем форму и скачиваем файл.
            cmd = d.get("command") or {}
            if cmd.get("action") == "generate_form" and cmd.get("form_id"):
                asyncio.create_task(_gen_form_from_command(cmd))
            # Команда задачника/журнала могла изменить данные — обновим открытую панель.
            if work_drawer.visible:
                asyncio.create_task(_refresh_work())

        # W5.1: SSE-стрим — токены в пузырь по мере генерации; финальное событие
        # несёт авторитетный payload (вердикт валидации в crag_status).
        stream_state = {"text": "", "got_token": False, "final": None, "error": None}

        def _on_sse(event: str, payload) -> None:
            if event == "token":
                if not stream_state["got_token"]:
                    stream_state["got_token"] = True
                    _stop_tick["v"] = True  # глушим тикер «Генерирую… Nс»
                stream_state["text"] += payload if isinstance(payload, str) else ""
                ai_placeholder_label.set_text(stream_state["text"])
            elif event == "reset":
                stream_state["text"] = ""
                ai_placeholder_label.set_text("")
            elif event == "final":
                stream_state["final"] = payload if isinstance(payload, dict) else {}
            elif event == "error":
                stream_state["error"] = payload if isinstance(payload, dict) else {"detail": str(payload)}

        completed = False
        try:
            await api_post_stream("/api/chat/stream", payload, _on_sse)
            _stop_tick["v"] = True
            d = stream_state["final"]
            if d:
                completed = True
                _apply_chat_result(d)
            elif stream_state["got_token"]:
                # Токены пришли, но финал потерян (обрыв середины стрима) —
                # не перегенерируем (дорого), показываем честную ошибку.
                completed = True
                err = stream_state["error"] or {}
                message = (
                    f"{err.get('status', '')}: {err.get('detail', '')}".strip(": ")
                    or last_api_error_text("Соединение прервано — ответ получен не полностью")
                )
                if err.get("status") == 409:
                    await _refresh_resource_gate()
                _finish_ai_placeholder(ai_placeholder, ai_placeholder_label, message, error=True)
                _render_artifact_error(message)
            else:
                # Ни одного токена (стрим-эндпоинт недоступен/ошибка до первого токена) —
                # безопасный откат на нестриминговый /api/chat.
                d = await api_post("/api/chat", payload)
                completed = True
                if d:
                    _apply_chat_result(d)
                else:
                    err = state.get("last_api_error") or {}
                    serr = stream_state["error"] or {}
                    message = last_api_error_text(
                        serr.get("detail") or "Ошибка запроса"
                    )
                    if err.get("status_code") == 409 or serr.get("status") == 409:
                        await _refresh_resource_gate()
                    _finish_ai_placeholder(ai_placeholder, ai_placeholder_label, message, error=True)
                    _render_artifact_error(message)
        except Exception as ex:
            completed = True
            _finish_ai_placeholder(ai_placeholder, ai_placeholder_label, f"Ошибка: {ex}", error=True)
            _render_artifact_error(str(ex))
        finally:
            if completed:
                state["chat_pending"] = None
            _stop_tick["v"] = True
            _tick_task.cancel()
            _sending["v"] = False
            await _refresh_resource_gate()
            chat_scroll.scroll_to(percent=1)

    async def send_chat():
        q = chat_input.value.strip()
        if not q:
            return
        chat_input.value = ""
        _update_prompt_preview()
        await _do_send(q)

    async def send_with_form():
        q = chat_input.value.strip()
        if not q:
            ui.notify("Введите текст запроса", type="warning")
            return
        advanced_dialog.close()
        chat_input.value = ""
        _update_prompt_preview()
        await _do_send(q)

    def _html_set_artifact_mode(label: str, hint: str):
        artifact_panel.clear()
        with artifact_panel:
            _html(
                '<div class="sov-artifact-empty">'
                f'<div class="sov-artifact-empty-title">{esc(label)}</div>'
                f'<div class="sov-muted">{esc(hint)}</div>'
                '</div>'
            )

    def _render_empty_artifacts():
        _html(
            '<div class="sov-artifact-empty">'
            '<div class="sov-artifact-empty-title">Пока пусто</div>'
            '<div class="sov-muted">Структурированные ответы, таблицы, SVG и диаграммы появятся здесь.</div>'
            '</div>'
        )

    def _render_artifact_loading(mode: str, question: str):
        artifact_panel.clear()
        label = OUTPUT_FORMATS.get(mode, ("Артефакт", ""))[0]
        with artifact_panel:
            _html(
                '<div class="sov-artifact-empty">'
                f'<div class="sov-artifact-empty-title">{esc(label)}</div>'
                f'<div class="sov-muted">Готовлю артефакт по запросу: {esc(question[:100])}</div>'
                '<div class="sov-artifact-loader"></div>'
                '</div>'
            )

    def _render_artifact_error(detail: str):
        artifact_panel.clear()
        with artifact_panel:
            _html(
                '<div class="sov-artifact-empty">'
                '<div class="sov-artifact-empty-title" style="color:var(--err);">Ошибка</div>'
                f'<div class="sov-muted">{esc(detail)}</div>'
                '</div>'
            )

    def _render_result(ans: str, mode: str, container, table_query: dict | None = None):
        container.clear()
        with container:
            with ui.card().classes("sov-artifact-card"):
                label = "Интерактивная таблица" if table_query else OUTPUT_FORMATS.get(mode, ("Ответ", ""))[0]
                # text-режим с таблицей внутри → заголовок «Таблица», а не «Текст»
                if not table_query and mode == "text" and (_parse_table_from_ai(ans) or _parse_markdown_table(ans)):
                    label = "Таблица"
                with ui.row().classes("w-full items-center justify-between"):
                    _html(f'<div class="sov-panel-title">{esc(label)}</div>')
                    ui.button("Копировать", icon="o_content_copy", on_click=lambda: asyncio.create_task(_copy_text(ans))).props(
                        "no-caps flat dense"
                    )

                if table_query:
                    _render_table_query(table_query)
                elif mode == "text":
                    # Если в ответе есть таблица — артефакт = ТОЛЬКО таблица + CSV (прозу
                    # видно в чате; Олег: «артефакт только таблицу, текст я и так вижу»).
                    auto = _parse_table_from_ai(ans) or _parse_markdown_table(ans)
                    if isinstance(auto, list) and auto and isinstance(auto[0], dict):
                        _render_table(auto)
                    else:
                        ui.markdown(ans).classes("sov-artifact-markdown")
                elif mode == "spec":
                    data = _parse_table_from_ai(ans)
                    if isinstance(data, list) and data and isinstance(data[0], dict):
                        _render_gost_spec(data)
                    else:
                        ui.markdown(ans).classes("sov-artifact-markdown")
                elif mode == "verify":
                    from sovushka.pages.verify import render_verify_artifact
                    render_verify_artifact(_parse_json_from_ai(ans))
                elif mode == "schema":
                    data = _parse_json_from_ai(ans)
                    if data:
                        with ui.column().classes("w-full gap-1"):
                            _render_tree(data)
                    else:
                        ui.markdown(ans).classes("sov-artifact-markdown")
                elif mode in ("structure", "table", "template"):
                    data = _parse_table_from_ai(ans) or _parse_json_from_ai(ans)
                    if isinstance(data, list) and data:
                        _render_table(data if isinstance(data[0], dict) else [{"значение": str(r)} for r in data])
                    elif isinstance(data, dict):
                        ui.markdown(f"```json\n{json.dumps(data, ensure_ascii=False, indent=2)}\n```").classes("sov-artifact-markdown")
                    else:
                        ui.markdown(ans).classes("sov-artifact-markdown")
                elif mode == "mermaid":
                    code = _parse_mermaid_from_ai(ans)
                    if code:
                        state["mermaid_last"] = code
                        ui.mermaid(code).classes("w-full")
                        with ui.row().classes("gap-2"):
                            ui.button("Код", icon="o_content_copy", on_click=lambda c=code: asyncio.create_task(_copy_text(c))).props("no-caps flat dense")
                            if tabs and tab_mermaid:
                                ui.button("В редактор", icon="o_open_in_new", on_click=lambda: tabs.set_value(tab_mermaid)).props("no-caps flat dense")
                    else:
                        ui.markdown(ans).classes("sov-artifact-markdown")
                elif mode == "svg":
                    svg_code = _parse_svg_from_ai(ans)
                    if svg_code:
                        _html(f'<div class="sov-svg-preview">{svg_code}</div>')
                        ui.button("Копировать SVG", icon="o_content_copy", on_click=lambda c=svg_code: asyncio.create_task(_copy_text(c))).props("no-caps flat dense")
                    else:
                        ui.markdown(ans).classes("sov-artifact-markdown")

    def _parse_table_from_ai(text: str):
        match = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(1))
            except Exception:
                pass
        return None

    def _parse_markdown_table(text: str):
        """Первую markdown-таблицу (| a | b | + строка-разделитель) → list[dict]. Иначе None."""
        lines = [ln.rstrip() for ln in (text or "").splitlines()]
        block: list[str] = []
        for ln in lines:
            s = ln.strip()
            if s.startswith("|") and s.count("|") >= 2:
                block.append(s)
            elif block:
                break  # таблица закончилась — берём первую целостную
        if len(block) < 2:
            return None

        def _cells(row: str) -> list[str]:
            return [c.strip() for c in row.strip().strip("|").split("|")]

        # вторая строка должна быть разделителем (---|:--: и т.п.)
        if not re.match(r"^\s*\|?[\s:|-]+\|?\s*$", block[1]):
            return None
        headers = _cells(block[0])
        rows: list[dict] = []
        for row in block[2:]:
            vals = _cells(row)
            if not any(vals):
                continue
            vals = (vals + [""] * len(headers))[: len(headers)]
            rows.append({h or f"col{i+1}": vals[i] for i, h in enumerate(headers)})
        return rows or None

    def _parse_json_from_ai(text: str):
        match = re.search(r"```(?:json)?\s*(\{.*?\}|\[.*?\])\s*```", text, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(1))
            except Exception:
                pass
        try:
            start = text.find("{") if "{" in text else text.find("[")
            if start >= 0:
                end = text.rfind("}") if "{" in text else text.rfind("]")
                return json.loads(text[start : end + 1])
        except Exception:
            pass
        return None

    def _parse_mermaid_from_ai(text: str) -> Optional[str]:
        match = re.search(r"```mermaid\s*(.*?)```", text, re.DOTALL)
        return match.group(1).strip() if match else None

    def _parse_svg_from_ai(text: str) -> Optional[str]:
        match = re.search(r"```svg\s*(.*?)```", text, re.DOTALL)
        if match:
            return sanitize_svg(match.group(1).strip())
        match = re.search(r"(<svg[\s\S]*?</svg>)", text, re.IGNORECASE)
        return sanitize_svg(match.group(1).strip()) if match else None

    def _render_table(data: list[dict]):
        keys = list(data[0].keys()) if data else []
        cols = [{"name": k, "label": k, "field": k, "align": "left", "sortable": True} for k in keys]
        ui.table(columns=cols, rows=data, pagination=8).classes("sov-artifact-table")
        with ui.row().classes("gap-2"):
            ui.button(
                "CSV",
                icon="o_download",
                on_click=lambda d=data: ui.download(_rows_to_csv(d), "specification.csv"),
            ).props("no-caps flat dense").tooltip("Скачать CSV (рус-Excel: ; + UTF-8 BOM)")
            ui.button(
                "JSON",
                icon="o_content_copy",
                on_click=lambda d=data: asyncio.create_task(_copy_text(json.dumps(d, ensure_ascii=False, indent=2))),
            ).props("no-caps flat dense")

    def _render_gost_spec(data: list[dict]):
        """Рендер спецификации по форме ГОСТ 21.110-2013: графы с правильными
        заголовками + рамка + заголовок формы + CSV/JSON с ГОСТ-шапкой."""
        gost = [
            ("поз", "Поз."),
            ("обозначение", "Обозначение"),
            ("наименование", "Наименование и техническая характеристика"),
            ("тип_марка", "Тип, марка, обозначение документа / опросного листа"),
            ("код", "Код продукции / завод-изготовитель"),
            ("ед_изм", "Ед. изм."),
            ("кол_во", "Кол."),
            ("масса_ед", "Масса ед., кг"),
            ("примечание", "Примечание"),
        ]
        rows = [r for r in data if isinstance(r, dict)]
        present = [(k, h) for k, h in gost if any(k in r for r in rows)]
        if not present:  # модель отдала иные ключи — показываем как есть
            present = [(k, k) for k in (rows[0].keys() if rows else [])]
        _html('<div style="font-size:.72rem;font-weight:900;color:var(--text);'
              'text-transform:uppercase;letter-spacing:.04em;margin-bottom:6px;">'
              'Спецификация оборудования, изделий и материалов · ГОСТ 21.110-2013</div>')
        cols = [{"name": k, "label": h, "field": k, "align": "left"} for k, h in present]
        ui.table(columns=cols, rows=rows, pagination=20, row_key="поз").props(
            "bordered dense flat"
        ).classes("sov-artifact-table")
        # CSV/JSON — с ГОСТ-заголовками граф (рус-Excel читает кириллицу).
        csv_rows = [{h: r.get(k, "") for k, h in present} for r in rows]
        with ui.row().classes("gap-2"):
            ui.button("CSV", icon="o_download",
                      on_click=lambda d=csv_rows: ui.download(_rows_to_csv(d), "specification_gost.csv")
                      ).props("no-caps flat dense").tooltip("Скачать CSV по ГОСТ (; + UTF-8 BOM)")
            ui.button("JSON", icon="o_content_copy",
                      on_click=lambda d=rows: asyncio.create_task(_copy_text(json.dumps(d, ensure_ascii=False, indent=2)))
                      ).props("no-caps flat dense")

    def _render_table_query(table_query: dict):
        try:
            rows = table_query.get("rows") or []
            if not rows:
                ui.markdown("_Нет данных для отображения в таблице_").classes("sov-artifact-markdown")
                return
            
            # Curate keys: hide internal ones and map to pretty labels
            sample = rows[0]
            visible_keys = [k for k in sample.keys() if not k.startswith("_") and k != "raw_row"]
            
            pretty_labels = {
                "pos": "№",
                "code": "Код",
                "name": "Наименование работ",
                "work_name": "Наименование работ",
                "unit": "Ед. изм.",
                "qty": "Кол-во",
                "price": "Цена",
                "amount": "Сумма",
                "amount_mat": "Материалы",
                "amount_work": "Работы",
                "work_done": "Выполнено",
                "weight_total": "Масса"
            }
            
            # Generate column definitions for AG Grid
            column_defs = []
            for k in visible_keys:
                label = pretty_labels.get(k.lower(), k)
                column_defs.append({
                    "headerName": label,
                    "field": k,
                    "filter": True,
                    "sortable": True,
                    "resizable": True
                })
                
            aggrid_options = {
                "columnDefs": column_defs,
                "rowData": rows,
                "pagination": True,
                "paginationPageSize": 10,
                "domLayout": "autoHeight"
            }
            
            # UI elements for table query
            operation = table_query.get("operation") or "list"
            total = table_query.get("total")
            count = table_query.get("count", 0)
            
            with ui.column().classes("w-full gap-2"):
                # Summary label
                summary_text = f"**Операция:** {operation.upper()}"
                if total is not None:
                    summary_text += f" | **Итого:** {total:,.2f}".replace(",", " ")
                summary_text += f" | **Строк:** {count}"
                ui.markdown(summary_text)
                
                # Render AG Grid
                ui.aggrid(aggrid_options).classes("w-full").style("margin-top: 5px;")
                
                # Export actions
                with ui.row().classes("gap-2"):
                    ui.button("Копировать JSON", icon="o_content_copy", on_click=lambda: asyncio.create_task(_copy_text(json.dumps(rows, ensure_ascii=False, indent=2)))).props("no-caps flat dense")
                    
        except Exception as e:
            logger.error(f"Error rendering AG Grid table query: {e}")
            try:
                markdown_lines = []
                sample = rows[0]
                cols = [k for k in sample.keys() if not k.startswith("_") and k != "raw_row"]
                markdown_lines.append("| " + " | ".join(cols) + " |")
                markdown_lines.append("| " + " | ".join(["---"] * len(cols)) + " |")
                for r in rows:
                    markdown_lines.append("| " + " | ".join(str(r.get(c) or "") for c in cols) + " |")
                ui.markdown("\n".join(markdown_lines)).classes("sov-artifact-markdown")
            except Exception:
                ui.markdown("_Ошибка отображения таблицы. Данные повреждены._").classes("sov-artifact-markdown")

    def _render_tree(data, level: int = 0):
        if isinstance(data, dict):
            name = data.get("name", data.get("title", data.get("id", "—")))
            desc = data.get("desc", data.get("description", ""))
            children = data.get("children", data.get("items", []))
            indent = level * 14
            _html(
                f'<div class="sov-tree-row" style="margin-left:{indent}px;">'
                f'<span class="sov-tree-mark">{"▸" if children else "•"}</span>'
                f'<span class="sov-tree-name">{esc(name)}</span>'
                f'<span class="sov-tree-desc">{esc(desc)}</span>'
                '</div>'
            )
            for child in children if isinstance(children, list) else []:
                _render_tree(child, level + 1)
        elif isinstance(data, list):
            for item in data:
                _render_tree(item, level)

    select_format("text")
    async def _load_commands() -> None:  # build_chat-уровень: зовётся ниже при сборке страницы
        d = await api_get("/api/commands")
        cmds = (d or {}).get("commands", []) if isinstance(d, dict) else []
        cmd_menu_box.clear()
        with cmd_menu_box:
            if not cmds:
                ui.menu_item("Команды недоступны", on_click=lambda: None)
            for c in cmds:
                label = f"{c['cmd']} — {c['desc']}"

                def _run(cmd=c["cmd"]):
                    chat_input.value = cmd
                    _update_prompt_preview()
                    asyncio.create_task(send_chat())

                ui.menu_item(label, on_click=_run)

    asyncio.create_task(_load_commands())  # W11.17: наполнить /-палитру
    asyncio.create_task(_refresh_resource_gate())
    resource_gate_timer = ui.timer(5.0, lambda: asyncio.create_task(_refresh_resource_gate()))
    context.client.on_disconnect(lambda *_: resource_gate_timer.cancel())
    chat_input.on(
        "keydown.enter.prevent",
        lambda e: asyncio.create_task(send_chat()) if not (e.args or {}).get("shiftKey") and not _resource_blocked["v"] else None,
    )

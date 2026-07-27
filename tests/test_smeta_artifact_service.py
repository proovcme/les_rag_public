from proxy.services.smeta_artifact_service import (
    _select_pricebook_for_question,
    build_checked_rim_form_from_visible_rows,
    build_norm_candidate_artifact_from_lookup,
    build_smeta_artifact,
    build_lsr_form,
    compact_smeta_answer,
    extract_smeta_tables,
    persist_smeta_artifact_exports,
)


def test_norm_candidate_artifact_formats_lookup_trace_for_excel_roundtrip(tmp_path):
    trace = {
        "enabled": True,
        "results": [
            {
                "call": {
                    "tool": "search_norm",
                    "args": {
                        "work_description": "прокладка кабеля UTP",
                        "work_family": "low_current",
                        "unit_hint": "м",
                    },
                },
                "result": {
                    "status": "ambiguous",
                    "work_family": "low_current",
                    "candidates": [
                        {
                            "norm_code": "ГЭСНм10-06-034-01",
                            "title": "Прокладка кабеля связи",
                            "collection": "ГЭСНм10",
                            "measure_unit": "100 м",
                            "unit_compatible": True,
                            "applicability_status": "accepted",
                            "score_total": 13.75,
                        },
                        {
                            "norm_code": "ГЭСНм10-06-037-01",
                            "title": "Прокладка кабеля в коробах",
                            "collection": "ГЭСНм10",
                            "measure_unit": "100 м",
                            "unit_compatible": True,
                            "applicability_status": "candidate",
                            "score_total": 11.5,
                        },
                    ],
                },
            },
            {
                "call": {
                    "tool": "search_norm",
                    "args": {
                        "work_description": "неизвестная работа",
                        "work_family": "low_current",
                        "unit_hint": "шт",
                    },
                },
                "result": {
                    "status": "not_found",
                    "hint": "переформулируй work_description",
                    "candidates": [],
                },
            },
        ],
    }

    artifact = build_norm_candidate_artifact_from_lookup(trace, question="сделай кандидатов")
    exported = persist_smeta_artifact_exports(artifact, output_dir=tmp_path, prefix="candidates")

    assert artifact is not None
    assert artifact["stage"] == "norm_candidates"
    assert artifact["tables"][0]["kind"] == "norm_candidates"
    assert "Код ГЭСН" in artifact["tables"][0]["headers"]
    assert artifact["tables"][0]["rows"] == 3
    assert "ГЭСНм10-06-034-01" in artifact["content"]
    assert "не ЛСР и не расчет денег" in artifact["content"]
    assert "деньги по ним" in artifact["content"]
    assert "ручной приемки" not in artifact["content"]
    assert exported is not None
    xlsx_path = tmp_path / exported["downloads"]["xlsx"].split("path=")[1]
    csv_path = tmp_path / exported["downloads"]["csv"].split("path=")[1]
    assert xlsx_path.is_file()
    assert "ГЭСНм10-06-037-01" in csv_path.read_text(encoding="utf-8-sig")


def test_smeta_artifact_extracts_work_cost_table_and_totals():
    answer = """
**ВОР и стоимость работ**
| Работа | Объём | Ставка сценария | Сумма | Комментарий |
|---|---:|---:|---:|---|
| Прокладка медного кабеля U/UTP | 38 600 м | 120 руб./м | 4 632 000 руб. | без поставки |
| Оконцевание портов RJ-45 | 518 порт | 1 250 руб./порт | 647 500 руб. | измерения отдельно |
| Итого |  |  | 5 279 500 руб. |  |

**Итог**
Предварительная оценка работ.
"""

    tables = extract_smeta_tables(answer)
    artifact = build_smeta_artifact(answer, question="Оцени СКС")

    assert len(tables) == 1
    assert tables[0].kind == "work_cost"
    assert tables[0].amount_total == 5_279_500
    assert artifact is not None
    assert artifact["mode"] == "markdown"
    assert "Сметный артефакт" in artifact["content"]
    assert "5 279 500 руб." in artifact["content"]
    assert artifact["tables"][0]["rows"] == 3


def test_smeta_artifact_uses_default_system_pricebook_without_region(monkeypatch):
    class FakePriceBook:
        pass

    monkeypatch.setattr(
        "proxy.services.fgis_price_service.available_pricebooks",
        lambda *args, **kwargs: ["/tmp/moskva_2kv2026.parquet", "/tmp/sankt-peterburg_2kv2026.parquet"],
    )
    monkeypatch.setattr("proxy.services.fgis_price_service.get_pricebook", lambda path: FakePriceBook())

    book, name = _select_pricebook_for_question("сделай ЛСР по выбранным нормам")

    assert isinstance(book, FakePriceBook)
    assert name == "sankt-peterburg_2kv2026"


def test_smeta_artifact_prefers_full_spb_pricebook_over_refresh_without_period(monkeypatch):
    class FakePriceBook:
        pass

    monkeypatch.setattr(
        "proxy.services.fgis_price_service.available_pricebooks",
        lambda *args, **kwargs: ["/tmp/spb_refresh.parquet", "/tmp/sankt-peterburg_2kv2026.parquet", "/tmp/spb_2kv2025.parquet"],
    )
    monkeypatch.setattr("proxy.services.fgis_price_service.get_pricebook", lambda path: FakePriceBook())

    book, name = _select_pricebook_for_question("объект СПб, БАП, теперь деньги по ним")

    assert isinstance(book, FakePriceBook)
    assert name == "sankt-peterburg_2kv2026"


def test_smeta_artifact_parses_english_money_separators():
    answer = """
**ЛСР**
| № п/п | Обоснование | Наименование работ и затрат | Ед. изм. | Кол-во на ед. | коэф. | Кол-во всего | Базис на ед., руб. | Индекс | Текущий на ед., руб. | коэф. | Текущий всего, руб. |
|---:|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|
| 1 | ГЭСНм10, кандидат | Монтаж шкафа | шт | 1 | 1 | 2 | 8,500.00 | 1.00 | 8,500.00 | 1.00 | 17,000.00 |
| 2 | ГЭСН15, кандидат | Отделка | м2 | 1 | 1 | 3 | 1,250.50 | 1.00 | 1,250.50 | 1.00 | 3,751.50 |
"""

    artifact = build_smeta_artifact(answer, question="сделай ЛСР")

    assert artifact is not None
    assert artifact["rim_lsr_form"]["amount_total"] == 20_751.5
    assert artifact["rim_lsr_form"]["rows"][0]["quantity"] == "2"
    assert artifact["rim_lsr_form"]["rows"][0]["unit_price"] == "8,500.00"


def test_smeta_artifact_skips_visible_total_row_inside_lsr_table():
    answer = """
**ЛСР**
| № п/п | Обоснование | Наименование работ и затрат | Ед. изм. | Кол-во на ед. | коэф. | Кол-во всего | Базис на ед., руб. | Индекс | Текущий на ед., руб. | коэф. | Текущий всего, руб. |
|---:|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|
| 1 | ГЭСНм08, кандидат | Прокладка кабеля | м | 1 | 1 | 100 | 120,00 | 1 | 120,00 | 1 | 12 000,00 |
| 2 | ГЭСН15, кандидат | Окраска | м2 | 1 | 1 | 10 | 300,00 | 1 | 300,00 | 1 | 3 000,00 |
|  |  | **ВСЕГО по смете** |  |  |  |  |  |  |  |  | **15 000,00** |
"""

    artifact = build_smeta_artifact(answer, question="сделай ЛСР")

    assert artifact is not None
    assert len(artifact["rim_lsr_form"]["rows"]) == 2
    assert artifact["rim_lsr_form"]["amount_total"] == 15_000
    assert artifact["rim_lsr_form"]["rows"][0]["quantity"] == "100"


def test_compact_smeta_answer_keeps_long_tables_visible_by_default(monkeypatch):
    monkeypatch.delenv("LES_SMETA_COMPACT_CHAT_TABLES", raising=False)
    rows = "\n".join(
        f"| Работа {idx} | {idx} шт | {idx * 1000} руб. |"
        for idx in range(1, 7)
    )
    answer = (
        "**Что понял**\n"
        "Даю сметную оценку работ.\n\n"
        "**Стоимость работ**\n"
        "| Работа | Объём | Сумма |\n"
        "|---|---:|---:|\n"
        f"{rows}\n\n"
        "**Итог**\n"
        "Сумма предварительная."
    )

    artifact = build_smeta_artifact(answer, question="смета")
    compact = compact_smeta_answer(answer, artifact)

    assert artifact is not None
    assert "Работа 6" in artifact["content"]
    assert "Работа 6" in compact
    assert "Таблица вынесена в артефакт" not in compact
    assert "Сумма предварительная" in compact


def test_compact_smeta_answer_moves_long_tables_to_artifact_marker_when_enabled(monkeypatch):
    monkeypatch.setenv("LES_SMETA_COMPACT_CHAT_TABLES", "1")
    rows = "\n".join(
        f"| Работа {idx} | {idx} шт | {idx * 1000} руб. |"
        for idx in range(1, 7)
    )
    answer = (
        "**Что понял**\n"
        "Даю сметную оценку работ.\n\n"
        "**Стоимость работ**\n"
        "| Работа | Объём | Сумма |\n"
        "|---|---:|---:|\n"
        f"{rows}\n\n"
        "**Итог**\n"
        "Сумма предварительная."
    )

    artifact = build_smeta_artifact(answer, question="смета")
    compact = compact_smeta_answer(answer, artifact)

    assert artifact is not None
    assert "Таблица вынесена в артефакт" in compact
    assert "6 строк" in compact
    assert "Работа 6" not in compact
    assert "Сумма предварительная" in compact


def test_compact_smeta_answer_drops_conflicting_manual_total_when_enabled(monkeypatch):
    monkeypatch.setenv("LES_SMETA_COMPACT_CHAT_TABLES", "1")
    rows = "\n".join(
        f"| {idx} | Работа {idx} | {idx} | шт | ГЭСН 08, кандидат | 1 000 руб./шт | {idx * 1000} руб. | scenario_assumption |"
        for idx in range(1, 7)
    )
    answer = (
        "**ЛСР**\n"
        "| № | Работа | Кол-во | Ед. | Норма/источник | Ставка/допущение | Сумма | Комментарий |\n"
        "|---:|---|---:|---:|---|---:|---:|---|\n"
        f"{rows}\n\n"
        "**Итог**\n"
        "Итого стоимость работ: 31 000 руб.\n"
        "Сумма предварительная."
    )

    artifact = build_smeta_artifact(answer, question="сделай ЛСР")
    compact = compact_smeta_answer(answer, artifact)

    assert artifact is not None
    assert artifact["rim_lsr_form"]["amount_total"] == 21_000
    assert "ЛСР-форма вынесена в артефакт/XLSX: 6 строк, сумма 21 000 руб." in compact
    assert "31 000 руб" not in compact
    assert "Сумма предварительная" in compact


def test_smeta_artifact_persists_xlsx_and_csv_downloads(tmp_path):
    answer = """
**ВОР**
| № | Раздел | Работа | Ед. | Кол-во | Основание | Статус |
|---:|---|---|---:|---:|---|---|
| 1 | СКС | Прокладка кабеля | м | 100 | спецификация | измеримо |

**Оценка стоимости работ**
| № | Работа | Кол-во | Ед. | Норма/источник | Ставка/допущение | Сумма | Комментарий |
|---:|---|---:|---:|---|---:|---:|---|
| 1 | Прокладка кабеля | 100 | м | ГЭСНм10, кандидат по кабелям связи | 120 руб./м | 12 000 руб. | сценарно |
"""

    artifact = build_smeta_artifact(answer, question="Дай ВОР")
    exported = persist_smeta_artifact_exports(artifact, output_dir=tmp_path, prefix="unit")

    assert exported is not None
    assert exported["downloads"]["xlsx"].endswith(".xlsx")
    assert exported["downloads"]["csv"].endswith(".csv")
    xlsx_path = tmp_path / exported["downloads"]["xlsx"].split("path=")[1]
    csv_path = tmp_path / exported["downloads"]["csv"].split("path=")[1]
    assert xlsx_path.is_file()
    assert csv_path.is_file()

    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    assert "Свод" in wb.sheetnames
    assert any("Оценка стоимости" in name for name in wb.sheetnames)
    assert "ГЭСНм10" in csv_path.read_text(encoding="utf-8-sig")


def test_smeta_artifact_adds_lsr_form_without_replacing_tables(tmp_path):
    answer = """
**Оценка стоимости работ**
| № | Работа | Кол-во | Ед. | Норма/источник | Ставка/допущение | Сумма | Комментарий |
|---:|---|---:|---:|---|---:|---:|---|
| 1 | Прокладка кабеля | 100 | м | ГЭСНм 10-06-037-01, аналог | 120 руб./м | 12 000 руб. | scenario_assumption |
| 2 | Оконцевание портов | 24 | порт | ГЭСНм 10, кандидат | 1 250 руб./порт | 30 000 руб. | scenario_assumption |
| Итого |  |  |  |  |  | 42 000 руб. |  |
"""

    tables = extract_smeta_tables(answer)
    lsr_form = build_lsr_form(tables)
    artifact = build_smeta_artifact(answer, question="Дай ВОР и оценку")
    exported = persist_smeta_artifact_exports(artifact, output_dir=tmp_path, prefix="lsr")

    assert lsr_form is not None
    assert lsr_form["schema"] == "lsr_rim_display_form_v1"
    assert lsr_form["rows"][0]["basis"] == "ГЭСНм 10-06-037-01"
    assert lsr_form["is_priced_final"] is False
    assert lsr_form["finality"] == "scenario_display"
    assert artifact is not None
    assert "## 1. Оценка стоимости работ" in artifact["content"]
    assert "## ЛСР РИМ (форма 421/пр)" in artifact["content"]
    assert "Приложение № 3" in artifact["content"]
    assert "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ (СМЕТА)" in artifact["content"]
    assert "Сметная стоимость: **42 000 руб.**" in artifact["content"]
    assert "| 1 | 2 | 3 | 4 | 5 | 6 | 7 | 8 | 9 | 10 | 11 | 12 |" in artifact["content"]
    assert "ВСЕГО по смете" in artifact["content"]
    assert "не финальная ЛСР" in artifact["content"]
    assert artifact["content"].index("## ЛСР РИМ (форма 421/пр)") < artifact["content"].index("## 1. Оценка стоимости работ")
    assert artifact["lsr_form"]["amount_total"] == 42_000
    assert artifact["rim_lsr_form"]["amount_total"] == 42_000

    import openpyxl

    xlsx_path = tmp_path / exported["downloads"]["xlsx"].split("path=")[1]
    csv_path = tmp_path / exported["downloads"]["csv"].split("path=")[1]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    assert "ЛСР РИМ" in wb.sheetnames
    assert wb.sheetnames[0] == "ЛСР РИМ"
    assert "Источники ЛСР" in wb.sheetnames
    ws = wb["ЛСР РИМ"]
    assert ws.cell(row=1, column=10).value == "Приложение № 3"
    assert "421/пр" in ws.cell(row=2, column=8).value
    assert ws.cell(row=7, column=1).value == "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ (СМЕТА) № ____"
    assert "не финальная ЛСР" in ws.cell(row=10, column=1).value
    assert ws.cell(row=15, column=3).value == "Наименование работ и затрат"
    assert ws.cell(row=15, column=12).value == "Сметная стоимость в текущем уровне цен всего, руб."
    assert ws.cell(row=16, column=12).value == 12
    assert ws.cell(row=17, column=3).value == "Прокладка кабеля"
    assert ws.cell(row=18, column=12).value == 30_000
    assert ws.cell(row=19, column=3).value == "ВСЕГО по смете"
    assert ws.cell(row=19, column=12).value == 42_000
    sources = wb["Источники ЛСР"]
    assert sources.cell(row=2, column=2).value == "scenario_assumption"
    csv_text = csv_path.read_text(encoding="utf-8-sig")
    assert "ЛСР РИМ (форма 421/пр)" in csv_text
    assert "ВСЕГО по смете" in csv_text
    assert "Источники ЛСР" in csv_text


def test_smeta_artifact_lsr_uses_one_primary_cost_table_not_duplicate_partial_lsr():
    cost_rows = "\n".join(
        f"| {idx} | Работа {idx} | 1 | шт. | ГЭСН 08, кандидат | {idx * 10} руб./шт. | {idx * 10} руб. | scenario_assumption |"
        for idx in range(1, 20)
    )
    partial_lsr_rows = "\n".join(
        f"| {idx} | ГЭСН 08 | Работа {idx} | шт. | 1 | {idx * 10} руб. | {idx * 10} руб. | предварительно |"
        for idx in range(1, 13)
    )
    answer = f"""
**Оценка стоимости работ**
| № | Работа | Кол-во | Ед. | Норма/источник | Ставка/допущение | Сумма | Комментарий |
|---:|---|---:|---:|---|---:|---:|---|
{cost_rows}

## ЛСР (предварительная форма)
| № | Обоснование | Наименование работ и затрат | Ед. | Кол-во | Цена ед. | Стоимость всего | Статус/источник |
|---:|---|---|---:|---:|---:|---:|---|
{partial_lsr_rows}
"""

    artifact = build_smeta_artifact(answer, question="сделай ЛСР")

    assert artifact is not None
    assert artifact["rim_lsr_form"]["source_tables"] == ["Оценка стоимости работ"]
    assert len(artifact["rim_lsr_form"]["rows"]) == 19
    assert artifact["rim_lsr_form"]["amount_total"] == 1_900
    assert "Сметная стоимость: **1 900 руб.**" in artifact["content"]
    assert "Сумма выбранной ЛСР-формы: **1 900 руб.**" in artifact["content"]
    assert "Сумма выбранной ЛСР-формы: **2 680 руб.**" not in artifact["content"]


def test_smeta_artifact_prefers_rim_trace_when_model_selected_norm_code():
    answer = """
**ЛСР**
| № п/п | Обоснование | Наименование работ и затрат | Ед. изм. | Кол-во на ед. | коэф. | Кол-во всего | Базис на ед., руб. | Индекс | Текущий на ед., руб. | коэф. | Текущий всего, руб. |
|---:|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|
| 1 | ГЭСН 12-01-034-02 | Устройство обрешетки | м2 | 1 | 1 | 61 | 1 000 | 1 | 1 000 | 1 | 61 000 |
"""

    artifact = build_smeta_artifact(answer, question="выдай цену в виде ЛСР по ценам СПб 2026")

    assert artifact is not None
    assert artifact["rim_lsr_form"]["schema"] == "lsr_rim_trace_form_v1"
    assert artifact["rim_lsr_form"]["amount_total"] == 11_813.04
    assert artifact["rim_lsr_form"]["is_priced_final"] is True
    assert artifact["rim_lsr_form"]["rows"][0]["basis"] == "ГЭСН12-01-034-02"
    assert artifact["model_lsr_form"]["amount_total"] == 61_000
    assert "11 813 руб." in artifact["content"]


def test_compact_smeta_answer_shows_checked_trace_lsr_without_model_placeholders(monkeypatch):
    monkeypatch.delenv("LES_SMETA_COMPACT_CHAT_TABLES", raising=False)
    answer = """
**ЛСР**
| № п/п | Обоснование | Наименование работ и затрат | Ед. изм. | Кол-во на ед. | коэф. | Кол-во всего | Базис на ед., руб. | Индекс | Текущий на ед., руб. | коэф. | Текущий всего, руб. |
|---:|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|
| 1 | ГЭСН 12-01-034-02 | Устройство обрешетки | м2 | 1 | 1 | 61 | 0.00 | 0.00 | 0.00 | 1 | 0.00 |
|  |  | **ВСЕГО по смете** |  |  |  |  |  |  |  |  | **0.00** |
"""

    artifact = build_smeta_artifact(answer, question="выдай цену в виде ЛСР по ценам СПб 2026")
    compact = compact_smeta_answer(answer, artifact)

    assert artifact is not None
    assert compact.startswith("ЛСР РИМ сформирована по расчетной трассе")
    assert "11 813 руб." in compact
    assert "Приложение № 3" in compact
    assert "Сметная стоимость в текущем уровне цен всего" in compact
    assert "**ВСЕГО по смете**" in compact
    assert "0.00" not in compact
    assert "placeholders" not in compact


def test_smeta_artifact_trace_does_not_invent_norms_for_unbound_rows():
    answer = """
**ЛСР**
| № п/п | Обоснование | Наименование работ и затрат | Ед. изм. | Кол-во на ед. | коэф. | Кол-во всего | Базис на ед., руб. | Индекс | Текущий на ед., руб. | коэф. | Текущий всего, руб. |
|---:|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|
| 1 | ГЭСН 12-01-034-02 | Устройство обрешетки | м2 | 1 | 1 | 61 | 1 000 | 1 | 1 000 | 1 | 61 000 |
| 2 | кандидат ГЭСНм 10 | Работа без выбранного шифра | шт | 1 | 1 | 2 | 5 000 | 1 | 5 000 | 1 | 10 000 |
"""

    artifact = build_smeta_artifact(answer, question="выдай цену в виде ЛСР по ценам СПб 2026")

    assert artifact is not None
    assert artifact["rim_lsr_form"]["schema"] == "lsr_rim_trace_form_v1"
    assert artifact["rim_lsr_form"]["amount_total"] == 11_813.04
    assert artifact["rim_lsr_form"]["finality"] == "priced_partial"
    assert len(artifact["rim_lsr_form"]["rows"]) == 2
    assert artifact["rim_lsr_form"]["rows"][1]["title"] == "Работа без выбранного шифра"
    assert artifact["rim_lsr_form"]["rows"][1]["basis"] == "нужен подбор нормы"
    assert artifact["rim_lsr_form"]["rows"][1]["amount"] == 0.0
    assert artifact["rim_lsr_form"]["rows"][1]["status"] == "norm_selection_required"
    flags = artifact["rim_lsr_form"]["trace"]["summary"]["flags"]
    assert "norm_selection_required: row-2" in flags

    compact = compact_smeta_answer(answer, artifact)
    assert compact.startswith("ЛСР РИМ сформирована по расчетной трассе")
    assert "1/2 строк ВОР рассчитано" in compact.split("## ЛСР РИМ", 1)[0]
    assert "Работа без выбранного шифра" in compact
    assert "Добор/проверка:" not in compact.split("## ЛСР РИМ", 1)[0]
    assert "Примечания: есть незакрытые строки/ресурсные цены/ставки" in compact


def test_smeta_artifact_keeps_calculated_rows_when_source_rows_are_partial():
    answer = """
**ЛСР**
| № п/п | Обоснование | Наименование работ и затрат | Ед. изм. | Кол-во на ед. | коэф. | Кол-во всего | Базис на ед., руб. | Индекс | Текущий на ед., руб. | коэф. | Текущий всего, руб. |
|---:|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|
| 1 | ГЭСН 12-01-034-02 | [SRC: Раздел 1; 1] Устройство обрешетки | м2 | 1 | 1 | 61 | 0.00 | 0.00 | 0.00 | 1 | 0.00 |
| 2 | кандидат ГЭСНм 10 | [SRC: Раздел 1; 2] Работа без выбранного шифра | шт | 1 | 1 | 2 | 0.00 | 0.00 | 0.00 | 1 | 0.00 |
"""
    question = (
        'Сделай ЛСР по строкам ВОР: [{"section":"Раздел 1","source_no":"1","name":"Обрешетка"},'
        '{"section":"Раздел 1","source_no":"2","name":"Работа без выбранного шифра"}]'
    )

    artifact = build_smeta_artifact(answer, question=question)

    assert artifact is not None
    assert artifact["rim_lsr_form"]["amount_total"] == 11_813.04
    assert artifact["rim_lsr_form"]["finality"] == "priced_partial"
    assert len(artifact["rim_lsr_form"]["rows"]) == 2
    assert artifact["rim_lsr_form"]["rows"][0]["basis"] == "ГЭСН12-01-034-02"
    assert artifact["rim_lsr_form"]["rows"][0]["amount"] == 11_813.04
    assert artifact["rim_lsr_form"]["rows"][1]["basis"] == "нужен подбор нормы"
    assert artifact["rim_lsr_form"]["rows"][1]["amount"] == 0.0
    summary = artifact["rim_lsr_form"]["trace"]["summary"]
    assert summary["bound_rows"] == 1
    assert summary["input_rows"] == 2


def test_checked_rim_form_keeps_source_row_order_across_sections():
    rows = [
        {
            "basis": "ГЭСН12-01-034-02",
            "title": "Устройство обрешетки",
            "unit": "м2",
            "quantity": 61,
            "section": "Раздел А",
        },
        {
            "basis": "ГЭСН15-04-006-02",
            "title": "Грунтовка потолков",
            "unit": "м2",
            "quantity": 3.2,
            "section": "Раздел Б",
        },
        {
            "basis": "ГЭСН15-02-038-02",
            "title": "Шпатлевка потолков",
            "unit": "м2",
            "quantity": 3.2,
            "section": "Раздел А",
        },
    ]

    rim_form = build_checked_rim_form_from_visible_rows(rows, question="ЛСР СПб 2026")

    assert rim_form is not None
    assert [row["basis"] for row in rim_form["rows"]] == [
        "ГЭСН12-01-034-02",
        "ГЭСН15-04-006-02",
        "ГЭСН15-02-038-02",
    ]
    assert [row["title"] for row in rim_form["rows"]] == [
        "Устройство обрешетки",
        "Грунтовка потолков",
        "Шпатлевка потолков",
    ]
    assert rim_form["trace"]["summary"]["input_rows"] == 3
    assert rim_form["trace"]["summary"]["bound_rows"] == 3

import openpyxl

from proxy.services import lsr_assembly_service as lsr_assembly
from proxy.services.rim_lsr_trace_service import build_lsr_trace
from proxy.services.rim_session_xlsx_service import render_session_lsr_xlsx


def test_session_xlsx_contains_rim_form_requirements_and_audit(tmp_path):
    trace = build_lsr_trace(
        [
            {
                "code": "ГЭСН12-01-034-02",
                "qty": 0.61,
                "section": "Кровля",
                "work_id": "vor-001",
            }
        ],
        pricebook=lsr_assembly._resolve_book(None),
        name="Тестовая ЛСР",
    )
    requirement = {
        "requirement_id": "req-1",
        "kind": "kac",
        "severity": "blocking",
        "finality_policy": "blocks_final",
        "work_id": "vor-001",
        "resource_code": "01.1",
        "description": "Нет текущей цены",
        "required_fields": ["supplier_offer_refs"],
        "status": "open",
        "source_refs": [],
    }
    audit = {
        "session": {
            "session_id": "session-1",
            "project_id": "project-1",
            "current_vor_revision_id": "vor-rev",
            "current_mapping_revision_id": "mapping-rev",
            "mapping_lock_revision_id": "mapping-lock",
            "current_scenario_revision_id": "scenario-rev",
            "current_pricing_revision_id": "pricing-rev",
            "final_lock_revision_id": "",
            "normative_base_version": "fsnb-2022",
            "pricebook_id": "pricebook-1",
            "region_code": "77",
            "price_period": "2026-Q2",
        },
        "revisions": [
            {
                "revision_id": "vor-rev",
                "parent_revision_id": "",
                "revision_kind": "vor_revision",
                "created_by": "model",
                "created_at": "2026-07-29T00:00:00+00:00",
                "payload_sha256": "abc",
            }
        ],
    }
    target = render_session_lsr_xlsx(
        trace,
        [requirement],
        audit,
        tmp_path / "draft.xlsx",
        is_final=False,
    )
    workbook = openpyxl.load_workbook(target, data_only=False)
    try:
        assert workbook.sheetnames == [
            "ЛСР РИМ",
            "Проверка",
            "Недостающие данные",
            "Аудит",
        ]
        assert workbook["Недостающие данные"]["A2"].value == "req-1"
        assert workbook["Аудит"]["B1"].value == "ЧЕРНОВИК"
        headers = [workbook["ЛСР РИМ"].cell(row=21, column=index).value for index in range(1, 13)]
        assert len(headers) == 12
    finally:
        workbook.close()


"""Родные xlsx-шаблоны форм: подстановка {{key}} + строки от якоря {{rows}}."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from proxy.services import forms_service


def _blob(path: str) -> str:
    ws = load_workbook(path).active
    return " | ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)


def test_vor_uses_native_template(tmp_path: Path):
    out = tmp_path / "vor.xlsx"
    res = forms_service.generate("vor", "xlsx", project_id=None,
                                 manual={"doc_code": "4-02-АС19"}, out_path=out)
    assert Path(res["path"]).is_file()
    blob = _blob(res["path"])
    assert "Ведомость объёмов работ" in blob       # шапка из родного бланка
    assert "4-02-АС19" in blob                       # {{doc_code}} подставлен
    assert "{{" not in blob                          # плейсхолдеры/якорь не утекли
    assert "Наименование работ" in blob              # колонка бланка


def test_ks_forms_registered():
    ids = {f["id"] for f in forms_service.list_forms()}
    assert {"vor", "ks2", "ks3"} <= ids


def test_template_anchor_writes_rows(tmp_path: Path):
    # строки данных пишутся от якоря {{rows}} (resolve_fields кладёт пустые при mode=blank)
    out = tmp_path / "ks2.xlsx"
    res = forms_service.generate("ks2", "xlsx", project_id=None,
                                 manual={"contractor": "ООО Подрядчик"}, out_path=out)
    blob = _blob(res["path"])
    assert "Акт о приёмке выполненных работ" in blob
    assert "ООО Подрядчик" in blob
    assert "{{" not in blob

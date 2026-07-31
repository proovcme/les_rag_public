"""Тесты `checklist_template_importer` (Phase 1, T1.1 — TDD, канон §10 CHECKLIST_REVIEW_PD_TASK.md).

Фикстура — синтетическая мини-книга XLSX, генерируемая openpyxl прямо в tmp_path (НЕ бинарник
в git): 4 листа — «Правила заполнения» (служебный), «АР» и «КР» (содержательные), «Сводная»
(служебный). Содержательные листы повторяют реальную структуру Glorax (docs/checklist_review/
SNAPSHOT_PD.md, docs/ALGO-glorax-checklist-review.md):

  row1  — название раздела (служебная строка)
  row2  — шапка колонок (№ / Критерий / Отметка / Примечание)
  затем:
    - строка-заголовок раздела: A непустая, B пустая
    - строка-заголовок блока С заливкой PatternFill(fgColor='FF3200F0') в B
    - строка-заголовок блока БЕЗ заливки и БЕЗ DV на C
    - 4-5 строк-критериев: B текст, C покрыта DataValidation list ('"Да,Нет,Не требуется"'
      либо '"Да,Нет"')
  в конце — итоговая строка «Всего пунктов:» с формулой COUNTA (игнорируется importer'ом).

Правило классификации (решение оператора 2026-07-04, ALGO-glorax-checklist-review.md):
  критерий        = B непустая И answer-ячейка C покрыта data validation
  заголовок блока = B непустая И DV на C нет (заливка FF3200F0 — вторичный диагностический
                     признак подтипа, не отдельная ветка правила)
  заголовок раздела = B пустая, A непустая
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from proxy.services.checklist_template_importer import import_checklist_template

BLOCK_FILL = PatternFill(fill_type="solid", fgColor="FF3200F0")


def _add_dv(ws, sqref: str, formula: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(sqref)


def _build_sheet(wb, name: str, *, dv_formula: str = '"Да,Нет,Не требуется"') -> None:
    """Строит содержательный лист реальной структуры (см. docstring модуля)."""
    ws = wb.create_sheet(name)
    ws["A1"] = f"Раздел «{name}»"
    ws["A2"] = "№"
    ws["B2"] = "Критерий"
    ws["C2"] = "Отметка"
    ws["D2"] = "Примечание"

    row = 3
    # заголовок раздела верхнего уровня: A непустая, B пустая
    ws.cell(row=row, column=1, value="ИСХОДНАЯ ДОКУМЕНТАЦИЯ")
    section_row = row
    row += 1

    # заголовок блока С заливкой FF3200F0
    block1_row = row
    ws.cell(row=row, column=2, value="Наличие необходимых расчётов:")
    ws.cell(row=row, column=2).fill = BLOCK_FILL
    row += 1

    # критерии блока 1 (4 штуки), DV на C
    block1_first_item_row = row
    criteria_block1 = [
        "Приложен отчет об инженерно-геодезических изысканиях",
        "Выполнен теплотехнический расчет",
        "Проверено соответствие требованиям ТЗ",
        "Соответствие армирования расчету",
    ]
    for text in criteria_block1:
        ws.cell(row=row, column=1, value=str(row - block1_first_item_row + 1))
        ws.cell(row=row, column=2, value=text)
        _add_dv(ws, f"C{row}", dv_formula)
        row += 1

    # заголовок блока БЕЗ заливки и БЕЗ DV
    block2_row = row
    ws.cell(row=row, column=2, value="Прочее")
    row += 1

    # критерии блока 2 (5 штук), DV на C
    block2_first_item_row = row
    criteria_block2 = [
        "Утеплитель соответствует вендор-листу",
        "Марка бетона соответствует расчету",
        "Приложен отчет об инженерно-геологических изысканиях",
        "Расположение корзин для кондиционеров согласовано",
        "Прочие требования выполнены",
    ]
    for text in criteria_block2:
        ws.cell(row=row, column=1, value=str(row - block2_first_item_row + 1))
        ws.cell(row=row, column=2, value=text)
        _add_dv(ws, f"C{row}", dv_formula)
        row += 1

    # итоговые строки — не критерии, игнорируются
    ws.cell(row=row, column=2, value="Всего пунктов:")
    ws.cell(row=row, column=3, value=f"=COUNTA(C{block1_first_item_row}:C{row - 1})")
    row += 1
    ws.cell(row=row, column=2, value="Из них, соответствуют критериям:")
    row += 1
    ws.cell(row=row, column=2, value="Чек лист заполнен на:")

    # атрибуты, полезные тестам через closure-словарь на самом листе (не часть XLSX-модели,
    # только для удобства ассертов — сохраняем как note в комментарии не будем; вместо этого
    # тесты вычисляют номера строк сами по формуле построения выше).


def _make_workbook(tmp_path, *, dv_ar='"Да,Нет,Не требуется"', dv_kr='"Да,Нет"'):
    wb = Workbook()
    # дефолтный лист openpyxl -> переиспользуем как служебный "Правила заполнения"
    rules_ws = wb.active
    rules_ws.title = "Правила заполнения"
    rules_ws["A1"] = "Инструкция по заполнению чек-листа"
    rules_ws["A2"] = "Отметьте Да/Нет/Не требуется в столбце C"

    _build_sheet(wb, "АР", dv_formula=dv_ar)
    _build_sheet(wb, "КР", dv_formula=dv_kr)

    summary_ws = wb.create_sheet("Сводная")
    summary_ws["A1"] = "Сводная таблица"
    summary_ws["A2"] = "Раздел"
    summary_ws["B2"] = "Критериев"
    summary_ws["A3"] = "АР"
    summary_ws["B3"] = "=COUNTA('АР'!C3:C100)"

    path = tmp_path / "checklist_synthetic.xlsx"
    wb.save(path)
    return path


# --- Ожидаемые координаты строк-критериев на построенном листе (см. _build_sheet) ---
# row3 = заголовок раздела "ИСХОДНАЯ ДОКУМЕНТАЦИЯ"
# row4 = заголовок блока (с заливкой) "Наличие необходимых расчётов:"
# row5-8 = 4 критерия блока 1
# row9 = заголовок блока (без заливки) "Прочее"
# row10-14 = 5 критериев блока 2
# row15 = "Всего пунктов:"
# row16 = "Из них, соответствуют критериям:"
# row17 = "Чек лист заполнен на:"

SECTION_ROW = 3
BLOCK1_HEADER_ROW = 4
BLOCK1_ITEM_ROWS = (5, 6, 7, 8)
BLOCK2_HEADER_ROW = 9
BLOCK2_ITEM_ROWS = (10, 11, 12, 13, 14)
ALL_ITEM_ROWS = BLOCK1_ITEM_ROWS + BLOCK2_ITEM_ROWS
TOTAL_ROW = 15


def test_checklist_template_importer_pd_extracts_sections(tmp_path):
    path = _make_workbook(tmp_path)
    template = import_checklist_template(path, stage="PD")

    assert template["stage"] == "PD"
    disciplines = set(template["disciplines"])
    assert disciplines == {"АР", "КР"}

    sheets_seen = {item["sheet_name"] for item in template["items"]}
    assert sheets_seen == {"АР", "КР"}


def test_checklist_template_importer_skips_rules_and_summary_sheets(tmp_path):
    path = _make_workbook(tmp_path)
    template = import_checklist_template(path, stage="PD")

    sheets_seen = {item["sheet_name"] for item in template["items"]}
    assert "Правила заполнения" not in sheets_seen
    assert "Сводная" not in sheets_seen
    assert "Правила заполнения" not in template["disciplines"]
    assert "Сводная" not in template["disciplines"]


def test_checklist_template_importer_reads_allowed_answers(tmp_path):
    path = _make_workbook(
        tmp_path, dv_ar='"Да,Нет,Не требуется"', dv_kr='"Да,Нет"'
    )
    template = import_checklist_template(path, stage="PD")

    ar_items = [i for i in template["items"] if i["sheet_name"] == "АР"]
    kr_items = [i for i in template["items"] if i["sheet_name"] == "КР"]
    assert ar_items, "ожидались items для листа АР"
    assert kr_items, "ожидались items для листа КР"

    assert all(i["allowed_answers"] == ["Да", "Нет", "Не требуется"] for i in ar_items)
    assert all(i["allowed_answers"] == ["Да", "Нет"] for i in kr_items)


def test_checklist_template_importer_stable_ids(tmp_path):
    path = _make_workbook(tmp_path)
    template1 = import_checklist_template(path, stage="PD")
    template2 = import_checklist_template(path, stage="PD")

    ids1 = [i["id"] for i in template1["items"]]
    ids2 = [i["id"] for i in template2["items"]]
    assert ids1 == ids2, "повторный импорт должен давать те же ids (детерминизм)"

    ar_items = sorted(
        (i for i in template1["items"] if i["sheet_name"] == "АР"),
        key=lambda i: i["row"],
    )
    # первый критерий листа АР — строка BLOCK1_ITEM_ROWS[0] = 5 -> PD-AR-005
    assert ar_items[0]["row"] == BLOCK1_ITEM_ROWS[0]
    assert ar_items[0]["id"] == f"PD-AR-{BLOCK1_ITEM_ROWS[0]:03d}"

    kr_items = sorted(
        (i for i in template1["items"] if i["sheet_name"] == "КР"),
        key=lambda i: i["row"],
    )
    assert kr_items[0]["id"] == f"PD-KR-{BLOCK1_ITEM_ROWS[0]:03d}"


def test_checklist_template_importer_headers_not_items(tmp_path):
    path = _make_workbook(tmp_path)
    template = import_checklist_template(path, stage="PD")

    ar_items = [i for i in template["items"] if i["sheet_name"] == "АР"]
    item_rows = {i["row"] for i in ar_items}

    # ни один из заголовков (раздела и обоих блоков) не должен попасть в items
    assert SECTION_ROW not in item_rows
    assert BLOCK1_HEADER_ROW not in item_rows
    assert BLOCK2_HEADER_ROW not in item_rows
    # итоговая строка тоже не item
    assert TOTAL_ROW not in item_rows
    # ровно 9 критериев (4 + 5) на листе АР
    assert item_rows == set(ALL_ITEM_ROWS)

    # заголовки должны попасть в section_path последующих критериев
    first_block1_item = next(i for i in ar_items if i["row"] == BLOCK1_ITEM_ROWS[0])
    assert any("ИСХОДНАЯ ДОКУМЕНТАЦИЯ" in p for p in first_block1_item["section_path"])
    assert any("Наличие необходимых расчётов" in p for p in first_block1_item["section_path"])

    first_block2_item = next(i for i in ar_items if i["row"] == BLOCK2_ITEM_ROWS[0])
    assert any("Прочее" in p for p in first_block2_item["section_path"])
    # заголовок блока 1 не должен "утечь" в section_path критериев блока 2
    assert not any(
        "Наличие необходимых расчётов" in p for p in first_block2_item["section_path"]
    )


def test_checklist_template_importer_classifies_basic_kinds(tmp_path):
    path = _make_workbook(tmp_path)
    template = import_checklist_template(path, stage="PD")
    ar_items = {i["row"]: i for i in template["items"] if i["sheet_name"] == "АР"}

    # "Приложен отчет об инженерно-геодезических изысканиях" -> presence
    presence_row = BLOCK1_ITEM_ROWS[0]
    assert ar_items[presence_row]["kind"] == "presence"

    # "Выполнен теплотехнический расчет" -> calculation
    calc_row = BLOCK1_ITEM_ROWS[1]
    assert ar_items[calc_row]["kind"] == "calculation"

    # "Проверено соответствие требованиям ТЗ" -> cross_section
    cross_row = BLOCK1_ITEM_ROWS[2]
    assert ar_items[cross_row]["kind"] == "cross_section"

    # "Соответствие армирования расчету" -> дефолт (не presence/calculation по ключевым словам
    # заголовка "расчет" — тут "соответствие" тоже триггерит cross_section) — проверим явным
    # дефолтным примером из блока 2 вместо неоднозначного.
    default_row = BLOCK2_ITEM_ROWS[4]  # "Прочие требования выполнены"
    assert ar_items[default_row]["kind"] == "manual_required"


def test_checklist_template_importer_classifies_all_seven_kinds(tmp_path):
    """T1.4: расширенная эвристика `_classify_kind` покрывает все 7 kind канона
    (docs/ALGO-glorax-checklist-review.md). Примеры близки к формулировкам реального
    ПД-чек-листа Glorax (config/checklists/glorax_pd_2026.json), но собраны в один
    синтетический лист, чтобы не зависеть от боевого JSON.
    """
    wb = Workbook()
    rules_ws = wb.active
    rules_ws.title = "Правила заполнения"

    ws = wb.create_sheet("АР")
    ws["A1"] = "Раздел «АР»"
    ws["A2"] = "№"
    ws["B2"] = "Критерий"
    ws["C2"] = "Отметка"

    examples = [
        # (criterion, expected_kind)
        ("Приложен отчет об инженерно-геологических изысканиях", "presence"),
        ("Наличие расчета инсоляции и КЕО", "presence"),
        ("Представлен ГПЗУ", "presence"),
        ("Выполнен теплотехнический расчет", "calculation"),
        ("Выполнен расчет скорости лифтов.", "calculation"),
        ("Расчетные параметры внутреннего воздуха соответствуют ТЗ.", "cross_section"),
        ("Пирог кровли здания соответствует ТЗ.", "cross_section"),
        ("Марка бетона соответствует требованиям по водонепроницаемости для ростверка (W12).",
         "parametric"),
        ("Толщина стяжки пола — не менее 80 мм.", "parametric"),
        # "сечение"/"диаметр"/"толщина" триггерят parametric только РЯДОМ с числом (реальные
        # данные glorax_pd_2026.json: "Проведён расчёт сечений проводников по допустимому
        # длительному току" без числа в тексте — это calculation, конкретное значение
        # получается уже в результате расчёта, не задано в самом критерии).
        ("Сечение кабеля не менее 2.5 мм2.", "parametric"),
        ("Ворота паркинга НЕ размещены над жилыми помещениями.", "spatial_visual"),
        ("Кухня НЕ размещена над жилыми комнатами.", "spatial_visual"),
        ("Места установки корзин для кондиционеров согласованы.", "spatial_visual"),
        ("Выполнен расчёт для зоны установки термовкладышей.", "spatial_visual"),
        ("Ведомость основного комплекта рабочих чертежей приложена.", "spds_formal"),
        ("Основная надпись в штампе заполнена в соответствии с ГОСТ 21.101.", "spds_formal"),
        ("Состав проекта соответствует общим данным тома.", "spds_formal"),
        # экспертный критерий: "соответствует концепции" — качественная архитектурная оценка,
        # намеренно НЕ попадает в cross_section (реальные данные glorax_pd_2026.json,
        # PD-SPOZU-014..018, все manual_required) — честная граница, не пробел эвристики.
        ("Решения по озеленению соответствуют концепции.", "manual_required"),
    ]

    for row_idx, (text, _expected) in enumerate(examples, start=3):
        ws.cell(row=row_idx, column=1, value=str(row_idx - 2))
        ws.cell(row=row_idx, column=2, value=text)
        _add_dv(ws, f"C{row_idx}", '"Да,Нет,Не требуется"')

    path = tmp_path / "checklist_all_kinds.xlsx"
    wb.save(path)

    template = import_checklist_template(path, stage="PD")
    items_by_row = {i["row"]: i for i in template["items"]}

    mismatches = []
    for row_idx, (text, expected) in enumerate(examples, start=3):
        actual = items_by_row[row_idx]["kind"]
        if actual != expected:
            mismatches.append(f"row={row_idx} text={text!r} expected={expected} actual={actual}")

    assert not mismatches, "\n".join(mismatches)


def test_checklist_template_importer_overrides_apply_after_heuristic(tmp_path):
    """T1.4: `overrides` (item_id -> {kind, comment}) перекрывает эвристику `_classify_kind`,
    применяется ПОСЛЕ автоматической классификации. Используется для ручной правки инженера
    без изменения кода эвристики (docs/ALGO-glorax-checklist-review.md).
    """
    path = _make_workbook(tmp_path)

    # без overrides — эвристика классифицирует блок2/row0 как manual_required (дефолт для
    # "Утеплитель соответствует вендор-листу" в текущей фикстуре: "соответств..." без ТЗ/ОПР/
    # .../вендор-паттерна matches vendor-триггер -> проверим фактическое поведение через baseline).
    baseline = import_checklist_template(path, stage="PD")
    ar_items_baseline = {i["row"]: i for i in baseline["items"] if i["sheet_name"] == "АР"}
    target_row = BLOCK1_ITEM_ROWS[0]  # "Приложен отчет об инженерно-геодезических изысканиях"
    target_id = ar_items_baseline[target_row]["id"]
    assert ar_items_baseline[target_row]["kind"] == "presence"

    overrides = {
        target_id: {"kind": "manual_required", "comment": "ручная правка инженера (тест)"},
    }

    template = import_checklist_template(path, stage="PD", overrides=overrides)
    ar_items = {i["row"]: i for i in template["items"] if i["sheet_name"] == "АР"}

    assert ar_items[target_row]["kind"] == "manual_required"
    assert ar_items[target_row]["comment"] == "ручная правка инженера (тест)"

    # правка точечная: остальные items с той же эвристикой не задеты
    other_row = BLOCK1_ITEM_ROWS[1]  # "Выполнен теплотехнический расчет"
    assert ar_items[other_row]["kind"] == "calculation"
    assert "comment" not in ar_items[other_row] or not ar_items[other_row]["comment"]


def test_checklist_template_importer_fill_wins_over_dv_conflict(tmp_path):
    """Реальный ПД-XLSX (T1.3, лист ОВиК, строки 15/18) содержит 2 строки с заливкой
    заголовка блока FF3200F0 И одновременно DV на answer-ячейке C (артефакт исходника —
    видимо протянутое форматирование DV при создании шаблона). Такая строка должна
    классифицироваться как заголовок блока, а не как критерий: заливка приоритетнее DV.
    """
    wb = Workbook()
    rules_ws = wb.active
    rules_ws.title = "Правила заполнения"

    ws = wb.create_sheet("ЭН")
    ws["A1"] = "Раздел «ЭН»"
    ws["A2"] = "№"
    ws["B2"] = "Критерий"
    ws["C2"] = "Отметка"

    # заголовок раздела
    ws.cell(row=3, column=1, value="ИСХОДНАЯ ДОКУМЕНТАЦИЯ")

    # заголовок блока С заливкой FF3200F0 И (артефактно) DV на C -> должен остаться заголовком
    conflict_row = 4
    ws.cell(row=conflict_row, column=1, value="1")
    ws.cell(row=conflict_row, column=2, value="Климатические условия")
    ws.cell(row=conflict_row, column=2).fill = BLOCK_FILL
    _add_dv(ws, f"C{conflict_row}", '"Да,Нет"')

    # обычный критерий этого блока (заливки нет, DV есть)
    item_row = 5
    ws.cell(row=item_row, column=1, value="1.1")
    ws.cell(row=item_row, column=2, value="Указаны расчётные параметры")
    _add_dv(ws, f"C{item_row}", '"Да,Нет"')

    path = tmp_path / "checklist_fill_conflict.xlsx"
    wb.save(path)

    template = import_checklist_template(path, stage="PD")
    en_items = [i for i in template["items"] if i["sheet_name"] == "ЭН"]
    rows_seen = {i["row"] for i in en_items}

    assert conflict_row not in rows_seen, (
        "строка с заливкой FF3200F0 И DV одновременно должна быть заголовком блока, "
        "не критерием (заливка приоритетнее DV)"
    )
    assert item_row in rows_seen

    item = next(i for i in en_items if i["row"] == item_row)
    assert any("Климатические условия" in p for p in item["section_path"])


def test_checklist_template_importer_item_fields(tmp_path):
    path = _make_workbook(tmp_path)
    template = import_checklist_template(path, stage="PD")

    required_fields = {
        "id",
        "stage",
        "discipline",
        "sheet_name",
        "row",
        "criterion",
        "allowed_answers",
        "kind",
        "answer_cell",
    }
    assert template["items"], "importer должен вернуть хотя бы один item"
    for item in template["items"]:
        missing = required_fields - set(item.keys())
        assert not missing, f"item {item.get('id')} не содержит полей: {missing}"
        assert item["stage"] == "PD"
        assert item["answer_cell"] == f"C{item['row']}"

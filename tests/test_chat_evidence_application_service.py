import asyncio
import hashlib
import inspect
import json
import sqlite3
import time
from dataclasses import fields
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

from proxy.routers import chat
from proxy.services import chat_evidence_application_service as service
from proxy.services.chat_attachment_service import preserve_read_attachment
from proxy.services.tool_harness_service import ToolHarness


def test_attachment_workbook_tools_survive_restrictive_model_shortlist():
    tools = [
        "dataset_map",
        "search_sources",
        "read_source",
        "read_excel_source",
        "search_project_tables",
        "read_project_table",
        "build_lsr_workbook",
        "build_vor_workbook",
    ]

    ordered = service.prioritize_workbook_tools(tools, workbook_phase=True)
    shortlist = ToolHarness().shortlist(
        "Собери ЛСР",
        allowed_tools=ordered,
        limit=5,
        dataset_ids=("selected",),
        workflow_phase="draft",
        model_preset="qwen-9b-restrictive",
        runtime_available=frozenset(tools),
        attachment_ids=("read_123456abcdef",),
    )

    assert ordered[:2] == ["build_lsr_workbook", "build_vor_workbook"]
    assert set(ordered) == set(tools)
    assert [item["name"] for item in shortlist["tools"][:2]] == [
        "build_lsr_workbook",
        "build_vor_workbook",
    ]


def test_tool_selector_request_declares_bound_attachment_without_copying_its_text():
    payload = service.tool_selector_request_payload(
        question="build workbook",
        mode="estimator",
        dataset_ids=[],
        target_file_ref={},
        round_no=1,
        attachment_id="read_123456abcdef",
    )

    assert payload["attachment"] == {
        "bound": True,
        "attachment_id": "read_123456abcdef",
    }
    assert "attachment_context" not in payload


def test_native_model_tool_schemas_preserve_registered_contract():
    schemas = service.native_model_tool_schemas(
        [
            {
                "name": "search_sources",
                "summary": "Search selected sources",
                "input_schema": {
                    "type": "object",
                    "properties": {"q": {"type": "string"}},
                    "required": ["q"],
                    "additionalProperties": False,
                },
            }
        ]
    )

    assert schemas == [
        {
            "type": "function",
            "function": {
                "name": "search_sources",
                "description": "Search selected sources",
                "parameters": {
                    "type": "object",
                    "properties": {"q": {"type": "string"}},
                    "required": ["q"],
                    "additionalProperties": False,
                },
            },
        }
    ]


def test_model_authored_search_batch_is_not_cut_to_five_calls():
    raw = json.dumps(
        {
            "calls": [
                {"tool": "search_sources", "args": {"q": f"work item {index}"}}
                for index in range(10)
            ]
        }
    )

    calls = chat._parse_model_tool_calls(
        raw,
        allowed_tools={"search_sources"},
        max_calls=48,
    )

    assert [call["args"]["q"] for call in calls] == [
        f"work item {index}" for index in range(10)
    ]


def test_model_rag_batch_preserves_every_model_query_without_a_row_limit():
    queries = [f"точный запрос {index}" for index in range(30)]

    assert service.parse_model_rag_queries(
        json.dumps({"queries": queries}, ensure_ascii=False)
    ) == queries


def test_model_rag_batch_reads_plain_model_authored_lines_without_a_row_limit():
    queries = [f"точный запрос {index}" for index in range(30)]

    assert service.parse_model_rag_queries("\n".join(queries)) == queries


def test_existing_estimator_profile_keeps_model_authored_query_workflow_without_new_flag():
    assert service.profile_uses_model_driven_retrieval(
        {
            "revision_id": "user:profile:estimator:legacy",
            "mode": "estimator",
            "rag_policy": {"grounded": True, "system_datasets": ["smeta"]},
        }
    ) is True
    assert service.profile_uses_model_driven_retrieval(
        {"mode": "search", "rag_policy": {"grounded": True}}
    ) is False


def test_model_rag_evidence_keeps_six_candidates_for_every_model_query():
    query_hits = []
    for query_index in range(1, 31):
        hits = [
            SimpleNamespace(
                content=(
                    f"Шифр: ГЭСН{query_index:02d}-00-000-{hit_index:02d} "
                    f"Наименование: кандидат {hit_index} " + "состав работ " * 80
                ),
                doc_id=f"doc-{query_index}-{hit_index}",
                doc_name="smeta_norm_cards.v1",
                score=1.0 / hit_index,
                meta={"dataset_id": "smeta"},
            )
            for hit_index in range(1, 9)
        ]
        query_hits.append((f"работа {query_index}", hits))

    groups, chunks = service.build_model_rag_evidence_groups(
        query_hits,
        max_chars=70_000,
    )

    rendered = "\n\n".join(groups)
    assert len(groups) == 30
    assert len(chunks) == 180
    assert "[Поисковый запрос Q30] работа 30" in rendered
    assert "[Q30.H6 | smeta_norm_cards.v1" in rendered
    assert "ГЭСН30-00-000-06" in rendered
    assert "ГЭСН30-00-000-07" not in rendered
    assert len(rendered) <= 70_000


def test_model_rag_result_preserves_all_rows_and_domain_fields_unchanged():
    rows = [
        {
            "source_row": index,
            "title": f"Работа {index}",
            "unit": "шт.",
            "quantity": index,
            "norm_code": f"ГЭСНм00-00-000-{index:02d}",
            "evidence_refs": [f"Источник {index}"],
        }
        for index in range(1, 31)
    ]

    answer = "\n".join(
        [
            "Готово",
            "",
            "| source_row | section | title | unit | quantity | norm_code | analogue | coverage | coefficient | evidence_refs |",
            "|---:|---|---|---|---:|---|---|---|---:|---|",
            *[
                f"| {index} | | Работа {index} | шт. | {index} | "
                f"ГЭСНм00-00-000-{index:02d} | | | | Источник {index} |"
                for index in range(1, 31)
            ],
        ]
    )

    assert service.parse_model_rag_result(answer) == (answer, rows)


def test_model_rag_result_reads_the_models_ordinary_sectioned_answer():
    answer = """### Раздел 1. Демонтажные работы ЭОМ
**Строка 1: Защитное укрытие пленкой, 116 м².**
* **Выбранный аналог:** `ГЭСН26-01-055-01` (Установка пароизоляционного слоя из пленки).
* **Обоснование:** Ближайший по составу работ аналог [Q1.H2].

### Раздел 2. Монтажные работы ЭОМ
**Строка 1: Монтаж блока аварийного питания, 16 шт.**
* **Выбранный аналог:** `ГЭСНм34-01-071-01` (Светильник с аварийным питанием).
* **Обоснование:** Аналог по сложности монтажа Q7.H1; Q7.H3.
"""

    assert service.parse_model_rag_result(answer) == (
        answer,
        [
            {
                "source_row": 1,
                "section": "Раздел 1. Демонтажные работы ЭОМ",
                "title": "Защитное укрытие пленкой",
                "unit": "м²",
                "quantity": 116,
                "norm_code": "ГЭСН26-01-055-01",
                "analogue": "Установка пароизоляционного слоя из пленки",
                "coverage": "Ближайший по составу работ аналог [Q1.H2].",
                "evidence_refs": ["Q1.H2"],
            },
            {
                "source_row": 2,
                "section": "Раздел 2. Монтажные работы ЭОМ",
                "title": "Монтаж блока аварийного питания",
                "unit": "шт.",
                "quantity": 16,
                "norm_code": "ГЭСНм34-01-071-01",
                "analogue": "Светильник с аварийным питанием",
                "coverage": "Аналог по сложности монтажа Q7.H1; Q7.H3.",
                "evidence_refs": ["Q7.H1", "Q7.H3"],
            },
        ],
    )


def test_model_rag_result_packages_plain_labelled_lines_without_a_markdown_table():
    answer = """Строка 1 — Прокладка контрольного кабеля; раздел: Монтаж; ед. изм.: м; количество: 120; norm_code: ГЭСНм08-02-146-01; аналог: Кабель контрольный; обоснование: прямое соответствие составу; коэффициент: 1; evidence: Q1.H2.
Строка 2 — Монтаж шкафа управления; раздел: Автоматика; ед. изм.: шт.; количество: 2; norm_code: ГЭСНм11-03-001-01; аналог: Шкаф управления; обоснование: совпадают измеритель и операции; evidence_refs: Q2.H1, Q2.H3."""

    parsed = service.parse_model_rag_result(answer)

    assert parsed == (
        answer,
        [
            {
                "source_row": 1,
                "section": "Монтаж",
                "title": "Прокладка контрольного кабеля",
                "unit": "м",
                "quantity": 120,
                "norm_code": "ГЭСНм08-02-146-01",
                "analogue": "Кабель контрольный",
                "coverage": "прямое соответствие составу",
                "coefficient": 1,
                "evidence_refs": ["Q1.H2"],
            },
            {
                "source_row": 2,
                "section": "Автоматика",
                "title": "Монтаж шкафа управления",
                "unit": "шт.",
                "quantity": 2,
                "norm_code": "ГЭСНм11-03-001-01",
                "analogue": "Шкаф управления",
                "coverage": "совпадают измеритель и операции",
                "evidence_refs": ["Q2.H1", "Q2.H3"],
            },
        ],
    )


def test_model_rag_queries_ignore_only_plain_text_presentation_wrappers():
    raw = (
        "Вот список поисковых запросов.\n"
        "```text\n"
        '"демонтаж кабеля в гофре"\n'
        '"монтаж аварийного светильника"\n'
        "```"
    )

    assert service.parse_model_rag_queries(raw) == [
        "демонтаж кабеля в гофре",
        "монтаж аварийного светильника",
    ]


def test_model_rag_result_reads_an_ordinary_russian_table_with_combined_norm_cell():
    answer = """| № п/п | Наименование работ (из ВОР) | Ед. изм. | Кол-во | Нормативная база (Шифр нормы) | Обоснование выбора и аналог | Примечания к составу работ |
|:---:|:---|:---:|:---:|:---|:---|:---|
| **Раздел 1. Демонтажные работы** | | | | | | |
| 1 | Защитное укрытие | м² | 116 | **ГЭСН26-01-055-01**<br>[Q3.H2] | Аналог: пароизоляционный слой | Материал отдельно |
| 2 | Разработка проема | шт. | 10 | *(Нет прямого аналога)*<br>Аналог: **ГЭСН15-02-035-03**<br>[Q6.H1] | Аналог по составу | Обратная операция |
"""

    parsed = service.parse_model_rag_result(answer)

    assert parsed is not None
    assert parsed[0] == answer
    assert parsed[1] == [
        {
            "source_row": 1,
            "section": "Раздел 1. Демонтажные работы",
            "title": "Защитное укрытие",
            "unit": "м²",
            "quantity": 116,
            "norm_code": "ГЭСН26-01-055-01",
            "analogue": "Аналог: пароизоляционный слой",
            "coverage": "Материал отдельно",
            "evidence_refs": ["Q3.H2"],
        },
        {
            "source_row": 2,
            "section": "Раздел 1. Демонтажные работы",
            "title": "Разработка проема",
            "unit": "шт.",
            "quantity": 10,
            "norm_code": "ГЭСН15-02-035-03",
            "analogue": "Аналог по составу",
            "coverage": "Обратная операция",
            "evidence_refs": ["Q6.H1"],
        },
    ]


def test_model_rag_result_collects_every_ordinary_table_across_sections():
    answer = """### Раздел 1. Демонтаж
| № пп | Наименование работ | Ед. изм. | Кол-во | Нормативная база (Norm Code) | Обоснование выбора и аналог | Примечания к выбору |
|:---|:---|:---:|:---:|:---|:---|:---|
| 1 | Работа 1 | м² | 116 | **ГЭСН26-02-010-02** | Аналог 1 | **Evidence:** Q1.H3 |

### Раздел 2. Монтаж
| № пп | Наименование работ | Ед. изм. | Кол-во | Нормативная база (Norm Code) | Обоснование выбора и аналог | Примечания к выбору |
|:---|:---|:---:|:---:|:---|:---|:---|
| 1 | Работа 2 | шт. | 10 | **ГЭСН17-01-010-01** | Аналог 2 | **Evidence:** Q3.H3 |
| 2 | Работа 3 | м | 160 | **ГЭСНм08-02-146-02** | Аналог 3 | **Evidence:** Q9.H4 |
"""

    parsed = service.parse_model_rag_result(answer)

    assert parsed is not None
    assert parsed[0] == answer
    assert [row["source_row"] for row in parsed[1]] == [1, 2, 3]
    assert [row["section"] for row in parsed[1]] == [
        "Раздел 1. Демонтаж",
        "Раздел 2. Монтаж",
        "Раздел 2. Монтаж",
    ]
    assert [row["norm_code"] for row in parsed[1]] == [
        "ГЭСН26-02-010-02",
        "ГЭСН17-01-010-01",
        "ГЭСНм08-02-146-02",
    ]
    assert [row["evidence_refs"] for row in parsed[1]] == [
        ["Q1.H3"],
        ["Q3.H3"],
        ["Q9.H4"],
    ]


def test_model_rag_result_accepts_the_models_short_ordinary_column_names():
    answer = """### Раздел 1. Работы
| № пп | Наименование | Ед. изм. | Кол-во | Примечание | norm_code (Шифр) | Аналог/Обоснование | Coverage (Соответствие) | Coefficient (Кэф.) | Evidence_refs |
|:---|:---|:---|:---|:---|:---|:---|:---|:---|:---|
| 1 | Работа | м² | 116 | Текст модели | ГЭСН26-01-055-01 | Аналог модели | Частичное | 1.0 | [Q1.H2] |
"""

    parsed = service.parse_model_rag_result(answer)

    assert parsed is not None
    assert parsed[1] == [
        {
            "source_row": 1,
            "section": "Раздел 1. Работы",
            "title": "Работа",
            "unit": "м²",
            "quantity": 116,
            "norm_code": "ГЭСН26-01-055-01",
            "analogue": "Аналог модели",
            "coverage": "Частичное",
            "coefficient": 1,
            "evidence_refs": ["Q1.H2"],
        }
    ]


async def _async_append(target, value):
    target.append(value)


async def _fake_workbook_execution(call, context, progress, _events):
    assert call["tool"] in {"build_lsr_workbook", "build_vor_workbook"}
    assert context["attachment_id"] == "read_123456abcdef"
    await progress({
        "call_id": call.get("call_id") or "workbook-call-1",
        "checkpoint_id": "cp-1",
        "phase": "rows",
        "completed": 3,
        "total": 10,
        "label": "Собираю строки ВОР",
    })
    return {
        "schema": "les.workbook_tool_result.v1",
        "tool": call["tool"],
        "status": "complete",
        "artifact": {
            "revision_id": "rev-2",
            "artifact_id": "artifact-1",
            "revision_no": 2,
            "filename": "vor-r2.xlsx",
            "download_url": "/api/artifacts/rev-2/download",
            "source_scope": ["attachment:read_123456abcdef"],
            "decision_checkpoint_id": "cp-1",
        },
        "source": {"attachment_id": "read_123456abcdef", "sha256": "a" * 64},
        "checkpoint": {"checkpoint_id": "cp-1", "status": "complete"},
        "missing": [],
        "blockers": [],
    }


async def _tracked_workbook_execution(call, context, progress, events, calls):
    calls.append(call)
    return await _fake_workbook_execution(call, context, progress, events)


async def _rejected_workbook_execution(call, calls):
    calls.append(call)
    return {
        "schema": "les.workbook_tool_result.v1",
        "tool": "build_vor_workbook",
        "status": "rejected",
        "code": "TOOL_SCOPE_VIOLATION",
        "message": "foreign attachment must not execute",
        "missing": [],
        "blockers": [],
    }


def test_workbook_result_harvests_revision_retry_and_checkpoint():
    harvested = service.harvest_workbook_tool_result(
        {
            "schema": "les.workbook_tool_result.v1",
            "tool": "build_vor_workbook",
            "status": "complete",
            "artifact": {
                "revision_id": "rev-2",
                "artifact_id": "artifact-1",
                "revision_no": 2,
                "filename": "vor-r2.xlsx",
                "download_url": "/api/artifacts/rev-2/download",
                "source_scope": ["attachment:read_123456abcdef"],
                "decision_checkpoint_id": "cp-1",
            },
            "source": {"attachment_id": "read_123456abcdef", "sha256": "a" * 64},
            "checkpoint": {"checkpoint_id": "cp-1", "status": "complete"},
        }
    )

    assert harvested["artifact"]["revision_id"] == "rev-2"
    assert harvested["artifact"]["filename"] == "vor-r2.xlsx"
    assert harvested["attachment_retry"]["attachment_id"] == "read_123456abcdef"
    assert harvested["source"] == {
        "attachment_id": "read_123456abcdef",
        "sha256": "a" * 64,
    }
    assert harvested["attachment_retry"]["preserved"] is True
    assert harvested["checkpoint"]["checkpoint_id"] == "cp-1"


def test_workbook_harvest_rejects_artifact_without_bound_attachment_checkpoint_lineage():
    harvested = service.harvest_workbook_tool_result(
        {
            "schema": "les.workbook_tool_result.v1",
            "tool": "build_vor_workbook",
            "status": "complete",
            "artifact": {
                "revision_id": "rev-2",
                "artifact_id": "artifact-1",
                "source_scope": ["attachment:read_foreign"],
                "decision_checkpoint_id": "cp-other",
            },
            "source": {"attachment_id": "read_123456abcdef"},
            "checkpoint": {"checkpoint_id": "cp-1", "status": "complete"},
        }
    )

    assert harvested == {}


def test_workbook_history_projection_redacts_failures_paths_and_unbounded_results():
    raw = {
        "schema": "les.workbook_tool_result.v1",
        "tool": "build_vor_workbook",
        "status": "failed",
        "code": "WORKBOOK_GENERATION_FAILED",
        "message": "adapter failed at C:/private/workbook.xlsx",
        "artifact": {
            "artifact_id": "art-1",
            "revision_id": "rev-1",
            "download_url": "/api/artifacts/rev-1/download",
            "source_scope": ["attachment:read_123456abcdef"],
            "decision_checkpoint_id": "cp-1",
            "filename": "C:/private/workbook.xlsx",
            "tool_calls": [{"private": "unbounded result"}],
        },
        "checkpoint": {"checkpoint_id": "cp-1", "status": "failed"},
        "source": {"attachment_id": "read_123456abcdef", "name": "C:/private/source.xlsx"},
        "blockers": ["RuntimeError: private filesystem detail"],
        "result": {"raw": "unbounded result"},
    }

    safe = service.safe_workbook_history_projection(raw)

    assert safe == {
        "schema": "les.workbook_tool_result.v1",
        "tool": "build_vor_workbook",
        "status": "failed",
        "code": "WORKBOOK_GENERATION_FAILED",
        "artifact": {
            "artifact_id": "art-1",
            "revision_id": "rev-1",
            "download_url": "/api/artifacts/rev-1/download",
            "source_scope": ["attachment:read_123456abcdef"],
            "decision_checkpoint_id": "cp-1",
        },
        "checkpoint": {"checkpoint_id": "cp-1", "status": "failed"},
        "source": {"attachment_id": "read_123456abcdef"},
    }
    serialized = json.dumps(safe, ensure_ascii=False)
    assert "private" not in serialized
    assert "unbounded result" not in serialized
    assert "RuntimeError" not in serialized


@pytest.mark.asyncio
async def test_chat_workbook_executor_builds_revision_and_streams_checkpoint(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    attachment_root = tmp_path / "attachments"
    monkeypatch.setenv("LES_CHAT_ATTACHMENT_ROOT", str(attachment_root))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(["Раздел", "Наименование", "Ед. изм.", "Количество"])
    workbook.active.append(["ОВ", "Воздуховод", "м", 12.5])
    workbook.save(source)
    workbook.close()
    attachment = preserve_read_attachment(
        source,
        attachment_id="read_123456abcdef",
        original_name="source.xlsx",
        root=attachment_root,
    )
    progress = []

    async def on_progress(event):
        progress.append(event)

    result = await chat._execute_chat_workbook_tool(
        {
            "call_id": "call-1",
            "tool": "build_vor_workbook",
            "args": {"attachment_id": attachment["attachment_id"]},
        },
        {
            "session_id": "session-1",
            "question": "Собери ВОР",
            "attachment_id": attachment["attachment_id"],
            "dataset_ids": [],
            "project_id": None,
            "profile_revision_id": "profile-1",
            "model_identity": "qwen-local",
            "model_preset": "qwen-9b",
        },
        on_progress,
    )

    assert result["status"] == "complete"
    assert result["artifact"]["revision_no"] == 1
    assert result["source"]["attachment_id"] == attachment["attachment_id"]
    assert f"attachment:{attachment['attachment_id']}" in result["artifact"]["source_scope"]
    assert result["artifact"]["decision_checkpoint_id"] == result["checkpoint"]["checkpoint_id"]
    assert progress[-1]["checkpoint_id"] == result["checkpoint"]["checkpoint_id"]
    assert progress[-1]["label"] == "Собираю строки ВОР"
    assert Path("storage/artifacts/meta.db").is_file()


@pytest.mark.asyncio
async def test_chat_lsr_executor_runs_thin_model_decision_adapter(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    attachment_root = tmp_path / "attachments"
    monkeypatch.setenv("LES_CHAT_ATTACHMENT_ROOT", str(attachment_root))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(["Наименование", "Ед. изм.", "Количество"])
    workbook.active.append(["Воздуховод", "м", 12.5])
    workbook.save(source)
    workbook.close()
    attachment = preserve_read_attachment(
        source,
        attachment_id="read_123456abcdef",
        original_name="source.xlsx",
        root=attachment_root,
    )
    decisions = [{
        "source_row": 1,
        "section": "ОВ",
        "title": "Монтаж воздуховода",
        "unit": "м",
        "quantity": 12.5,
        "norm_code": "ГЭСН20-01-001-01",
        "evidence_refs": ["dataset:fsnb:card:20-01-001-01"],
    }]

    def calculate(rows, **_kwargs):
        assert rows == decisions
        return {
            "schema": "rim_lsr_v1",
            "sections": [],
            "summary": {"input_rows": 1, "bound_rows": 1, "flags": []},
            "row_bindings": [{"row": 1, "status": "bound"}],
        }

    def render(_trace, output_path, **_kwargs):
        workbook = openpyxl.Workbook()
        workbook.active.append(["ЛСР"])
        workbook.save(output_path)
        workbook.close()
        return output_path

    monkeypatch.setattr(
        "proxy.services.lsr_workbook_adapter_service.build_lsr_trace_from_visible_rows",
        calculate,
    )
    monkeypatch.setattr(
        "proxy.services.lsr_workbook_adapter_service.render_lsr_xlsx",
        render,
    )

    result = await chat._execute_chat_workbook_tool(
        {
            "call_id": "call-lsr-1",
            "tool": "build_lsr_workbook",
            "args": {"attachment_id": attachment["attachment_id"], "decisions": decisions},
        },
        {
            "session_id": "session-1",
            "question": "Собери ЛСР",
            "attachment_id": attachment["attachment_id"],
            "dataset_ids": ["fsnb"],
            "project_id": None,
            "profile_revision_id": "profile-1",
            "model_identity": "qwen-local",
            "model_preset": "qwen-9b",
        },
        lambda _event: asyncio.sleep(0),
    )

    assert result["status"] == "complete"
    assert result["tool"] == "build_lsr_workbook"
    assert result["source"]["rows"] == 1
    assert "dataset:fsnb" in result["artifact"]["source_scope"]
    assert result["blockers"] == []


@pytest.mark.asyncio
async def test_chat_workbook_executor_rejects_model_dataset_outside_chat_scope(
    tmp_path, monkeypatch
):
    monkeypatch.chdir(tmp_path)
    attachment_root = tmp_path / "attachments"
    monkeypatch.setenv("LES_CHAT_ATTACHMENT_ROOT", str(attachment_root))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(["Наименование", "Ед. изм.", "Количество"])
    workbook.active.append(["Кабель", "м", 2])
    workbook.save(source)
    workbook.close()
    preserve_read_attachment(
        source,
        attachment_id="read_123456abcdef",
        original_name="source.xlsx",
        root=attachment_root,
    )

    async def on_progress(_event):
        raise AssertionError("rejected execution must not start")

    result = await chat._execute_chat_workbook_tool(
        {
            "tool": "build_vor_workbook",
            "args": {
                "attachment_id": "read_123456abcdef",
                "dataset_ids": ["foreign-dataset"],
            },
        },
        {
            "session_id": "session-1",
            "question": "Собери ВОР",
            "attachment_id": "read_123456abcdef",
            "dataset_ids": [],
            "project_id": None,
            "profile_revision_id": "profile-1",
            "model_identity": "qwen-local",
            "model_preset": "qwen-9b",
        },
        on_progress,
    )

    assert result["code"] == "TOOL_SCOPE_VIOLATION"
    assert "artifact" not in result
    assert not Path("storage/workbook_checkpoints.db").exists()
    assert not Path("storage/artifacts/meta.db").exists()


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("model_dataset_ids", "expected_status"),
    ((["selected-a"], "rejected"), (["selected-b", "selected-a"], "complete")),
)
async def test_chat_workbook_executor_binds_exact_request_dataset_scope(
    tmp_path, monkeypatch, model_dataset_ids, expected_status
):
    """A workbook artifact always records the full server-owned request scope."""
    monkeypatch.chdir(tmp_path)
    attachment_root = tmp_path / "attachments"
    monkeypatch.setenv("LES_CHAT_ATTACHMENT_ROOT", str(attachment_root))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(["Наименование", "Ед. изм.", "Количество"])
    workbook.active.append(["Кабель", "м", 2])
    workbook.save(source)
    workbook.close()
    attachment = preserve_read_attachment(
        source,
        attachment_id="read_123456abcdef",
        original_name="source.xlsx",
        root=attachment_root,
    )

    async def on_progress(_event):
        return None

    result = await chat._execute_chat_workbook_tool(
        {
            "tool": "build_vor_workbook",
            "args": {
                "attachment_id": attachment["attachment_id"],
                "dataset_ids": model_dataset_ids,
            },
        },
        {
            "session_id": "session-1",
            "question": "Собери ВОР",
            "attachment_id": attachment["attachment_id"],
            "dataset_ids": ["selected-a", "selected-b"],
            "project_id": None,
            "profile_revision_id": "profile-1",
            "model_identity": "qwen-local",
            "model_preset": "qwen-9b",
        },
        on_progress,
    )

    assert result["status"] == expected_status
    if expected_status == "rejected":
        assert result["code"] == "TOOL_SCOPE_VIOLATION"
        assert "artifact" not in result
        return
    assert result["artifact"]["source_scope"] == [
        f"attachment:{attachment['attachment_id']}",
        "dataset:selected-a",
        "dataset:selected-b",
    ]


@pytest.mark.asyncio
async def test_chat_workbook_executor_rejects_foreign_preserved_attachment_before_execution(
    tmp_path, monkeypatch
):
    """A model cannot switch a request to another server-owned attachment."""
    monkeypatch.chdir(tmp_path)
    attachment_root = tmp_path / "attachments"
    monkeypatch.setenv("LES_CHAT_ATTACHMENT_ROOT", str(attachment_root))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(["Наименование", "Ед. изм.", "Количество"])
    workbook.active.append(["Кабель", "м", 2])
    workbook.save(source)
    workbook.close()
    bound = preserve_read_attachment(
        source,
        attachment_id="read_111111111111",
        original_name="bound.xlsx",
        root=attachment_root,
    )
    foreign = preserve_read_attachment(
        source,
        attachment_id="read_222222222222",
        original_name="foreign.xlsx",
        root=attachment_root,
    )

    result = await chat._execute_chat_workbook_tool(
        {
            "tool": "build_vor_workbook",
            "args": {"attachment_id": foreign["attachment_id"]},
        },
        {
            "session_id": "session-1",
            "question": "Собери ВОР",
            "attachment_id": bound["attachment_id"],
            "dataset_ids": [],
            "project_id": None,
            "profile_revision_id": "profile-1",
            "model_identity": "qwen-local",
            "model_preset": "qwen-9b",
        },
        lambda _event: None,
    )

    assert result["status"] == "rejected"
    assert result["code"] == "TOOL_SCOPE_VIOLATION"
    assert not Path("storage/workbook_checkpoints.db").exists()
    assert not Path("storage/artifacts/meta.db").exists()
from proxy.services.canonical_route_service import CanonicalRouteMode
from proxy.services.context_governor_service import (
    ContextKind,
    ContextObject,
    ContextRequiredSectionOverflow,
)
from proxy.services.model_execution_preset_service import ModelExecutionPreset


def _governor_preset(
    *, input_tokens: int = 6000, preset_id: str = "fixture-restrictive"
) -> ModelExecutionPreset:
    return ModelExecutionPreset(
        preset_id=preset_id,
        model_family="fixture",
        input_token_limit=input_tokens,
        generation_reserve_tokens=20,
        safety_reserve_tokens=20,
        normal_tool_count=3,
        max_tools=5,
        max_batch_items=5,
        parallel_read_limit=1,
        reasoning_enabled=False,
        source_chain=("test",),
    )


def test_governed_inference_trace_exposes_exact_evidence_but_not_private_memory() -> None:
    messages, packet = service.govern_inference_messages(
        preset=_governor_preset(),
        profile_prefix="bound profile",
        request_payload={"question": "current request"},
        shortlist=[{"name": "read_source"}],
        checkpoint=[ContextObject("checkpoint:1", {"status": "continue"})],
        working_memory=[
            ContextObject("memory:1", {"context_role": "advisory_state", "secret": "hidden"})
        ],
        evidence=[{"source": "evidence payload"}],
        source_map=[{"label": "Источник 1"}],
        tool_exchange=[{"status": "ok"}],
        dialogue=["prior turn"],
    )

    assert [section.kind for section in packet.sections] == list(ContextKind)
    assert [message["role"] for message in messages] == ["system", "user"]
    trace = service.context_packet_trace(packet, purpose="answer")
    assert trace["purpose"] == "answer"
    assert [section["kind"] for section in trace["sections"]] == [kind.value for kind in ContextKind]
    assert "hidden" not in json.dumps(trace)
    evidence = next(section for section in trace["sections"] if section["kind"] == "evidence")
    source_map = next(section for section in trace["sections"] if section["kind"] == "source_map")
    assert evidence["objects"][0]["payload"] == {"source": "evidence payload"}
    assert evidence["objects"][0]["text"] == '{"source":"evidence payload"}'
    assert len(evidence["objects"][0]["sha256"]) == 64
    assert source_map["objects"][0]["payload"] == {"label": "Источник 1"}


def test_context_trace_records_exact_omitted_evidence_ids_and_cursor() -> None:
    _messages, packet = service.govern_inference_messages(
        preset=_governor_preset(input_tokens=90),
        profile_prefix="p",
        request_payload="q",
        evidence=["x" * 200],
        source_map=[{"label": "Источник 1"}],
    )

    trace = service.context_packet_trace(packet, purpose="tool_decision")
    omitted = next(item for item in trace["omissions"] if item["kind"] == "evidence")

    assert omitted["object_ids"] == ["evidence:0"]
    assert omitted["cursor"].startswith("ctx:evidence:")


def test_governed_inference_rejects_required_overflow_before_provider_call() -> None:
    provider_calls = []

    with pytest.raises(ContextRequiredSectionOverflow):
        messages, _packet = service.govern_inference_messages(
            preset=_governor_preset(input_tokens=50),
            profile_prefix="profile is too large for this deliberately tiny fixture",
            request_payload={"question": "also required"},
        )
        provider_calls.append(messages)

    assert provider_calls == []


def test_cloud_fallback_re_resolves_and_repacks_before_local_provider() -> None:
    source = inspect.getsource(service._execute_chat_evidence_application)

    resolve_at = source.index("fallback_preset = resolve_transport_execution_profile(")
    repack_at = source.index("fallback_messages, fallback_packet = govern_inference_messages(")
    send_at = source.index("fallback_body = {")
    assert resolve_at < repack_at < send_at
    assert '"messages": fallback_messages' in source
    assert "fallback_preset.generation_reserve_tokens" in source
    assert 'purpose="answer_fallback"' in source


def test_general_evidence_execution_is_outside_http_router():
    router_source = inspect.getsource(chat._run_chat)
    evidence_source = inspect.getsource(service._execute_chat_evidence_application)

    assert "retrieval_trace = retrieval.payload()" not in router_source
    assert "build_retrieval_evidence_packet(" not in router_source
    assert "async with gen_semaphore" not in router_source
    assert "retrieval_trace = retrieval.payload()" in evidence_source
    assert "build_retrieval_evidence_packet(" in evidence_source
    assert "async with gen_semaphore" in evidence_source


def test_evidence_boundary_is_typed_and_does_not_capture_namespaces():
    router_source = inspect.getsource(chat._run_chat)
    service_source = inspect.getsource(service)

    assert "globals()" not in router_source
    assert "locals()" not in router_source
    assert "global_scope" not in service_source
    assert "local_scope" not in service_source
    assert {field.name for field in fields(service.EvidenceRequestContext)} >= {
        "req", "dataset_ids", "query_route_payload", "target_file_ref", "topic_retrieval_plan"
    }
    assert {field.name for field in fields(service.EvidenceRuntimeDeps)} >= {
        "state", "rag_backend", "cache", "llm_runtime", "table_query_response"
    }
    assert {field.name for field in fields(service.ResponseBoundary)} == {
        "save_chat_history", "token_sink", "version_stamp"
    }


def test_router_builds_explicit_evidence_contracts():
    router_source = inspect.getsource(chat._run_chat)

    assert "EvidenceRequestContext(" in router_source
    assert "EvidenceRuntimeDeps(" in router_source
    assert "ResponseBoundary(" in router_source
    assert "run_chat_evidence_application(" in router_source


@pytest.mark.asyncio
async def test_typed_contract_maps_every_internal_binding(monkeypatch):
    dummy = lambda *_args, **_kwargs: None
    request = service.EvidenceRequestContext(
        req=SimpleNamespace(question="test"),
        dataset_ids=["ds"],
        effective_dataset_filter="DS",
        resolved_dataset_names=["Dataset"],
        dataset_name_by_id={"ds": "Dataset"},
        query_route_payload={"channel": "rag"},
        target_doc_filter=[],
        target_file_ref=None,
        topic_doc_filter=[],
        topic_retrieval_plan=None,
        inventory_requested=False,
        study_requested=False,
        memory_block="",
        session_block="",
        class_suggestions=[],
        use_semantic_cache=False,
        use_validation=False,
        validation_skip_reason="test",
        route=SimpleNamespace(intent="rag"),
        table_result=None,
        request_started_at=1.0,
    )
    runtime = service.EvidenceRuntimeDeps(
        state="state", rag_backend="rag", cache="cache", cache_embedding=None,
        cache_marker="miss", cache_scope="scope",
        assistant_text=dummy, augment_model_tool_args=dummy,
        chat_model_final_answer=dummy, cloud_body_for_model=dummy,
        compact_tool_result_for_prompt=dummy, dataset_ids_from_chunks=dummy,
        dataset_sensitivities=dummy, env_bool=dummy, env_float=dummy, env_int=dummy,
        expand_context_windows=dummy,
        format_tool_results_for_model=dummy, generation_token_budget=dummy,
        llm_runtime=dummy, local_context_budget=dummy, mlx_runtime=dummy,
        names_for_dataset_ids=dummy, notebook_study_validation_status=dummy,
        ollama_native_complete=dummy, parse_model_tool_calls=dummy,
        prepare_notebook_reader_memory=dummy, record_cloud_cost=dummy,
        retrieve_chat_chunks=dummy,
        source_excerpts=dummy, table_query_response=dummy,
        cloud_fallback_models=dummy, cloud_model_timeout=dummy,
    )
    boundary = service.ResponseBoundary(
        save_chat_history=dummy,
        token_sink=None,
        version_stamp=dummy,
    )
    captured = {}

    async def fake_execute(**kwargs):
        captured.update(kwargs)
        return {"answer": "ok"}

    monkeypatch.setattr(service, "_execute_chat_evidence_application", fake_execute)

    result = await service.run_chat_evidence_application(request, runtime, boundary)

    assert result == {"answer": "ok"}
    assert captured["req"] is request.req
    assert captured["_dataset_ids"] == ["ds"]
    assert captured["state"] == "state"
    assert captured["source_excerpts"] is dummy
    assert captured["save_chat_history"] is dummy


@pytest.mark.asyncio
async def test_shadow_candidate_executes_only_one_decision_and_redacts_result() -> None:
    calls = []

    class FakeHarness:
        async def call_async(self, tool, args, **policy):
            calls.append((tool, args, policy))
            return {
                "schema": "les_tool_result_v1",
                "status": "ok",
                "result": {"secret_text": "must not enter shadow trace"},
                "execution": {
                    "schema": "les_tool_execution_v1",
                    "status": "ok",
                    "code": "TOOL_OK",
                },
            }

    trace = await service.execute_canonical_shadow_decision(
        proposed_calls=[
            {"tool": "read_source", "args": {"doc_id": "d1"}},
            {"tool": "read_source", "args": {"doc_id": "d2"}},
        ],
        allowed_tools={"read_source"},
        dataset_ids=["selected"],
        tool_harness=FakeHarness(),
    )

    assert len(calls) == 1
    assert calls[0][2]["shadow"] is True
    assert trace["executed_calls"] == 1
    assert trace["pending_calls"] == 1
    assert trace["user_visible"] is False
    assert "secret_text" not in str(trace)


def test_grounded_application_never_uses_answer_cache_or_code_no_data_final() -> None:
    source = inspect.getsource(service._execute_chat_evidence_application)

    assert "if document_grounding_enabled:" in source
    assert "use_semantic_cache = False" in source
    assert '"empty_retrieval_model_first_v1"' in source
    assert "empty_scoped_retrieval_no_data_v1" not in source
    assert "if not chunks and effective_dataset_filter:" not in source
    assert "if not chunks and target_file_ref" not in source


def test_plain_ai_scope_removes_only_document_evidence_tools() -> None:
    tools = service.tools_for_document_scope(
        ["search_sources", "read_source", "calculate", "send_mail"],
        enabled=False,
    )

    assert tools == ["calculate", "send_mail"]
    assert service.tools_for_document_scope(
        ["search_sources", "read_source", "calculate"], enabled=True
    ) == ["search_sources", "read_source", "calculate"]


def test_estimator_role_uses_model_authored_initial_rag_query() -> None:
    assert service.profile_uses_model_driven_retrieval(
        {
            "mode": "estimator",
            "rag_policy": {
                "system_datasets": ["smeta"],
                "model_authored_initial_query": True,
            },
        }
    ) is True


def test_native_tool_selector_keeps_attachment_rows_inside_9b_context() -> None:
    attachment = "\n".join(
        f"{row}. Работа {row} | шт. | {row}" for row in range(1, 31)
    )
    evidence = service.selector_evidence_payload(
        attachment_context=attachment,
        rendered_context="[ФРАГМЕНТЫ ИЗ ИСТОЧНИКОВ]\nНет найденных фрагментов.",
    )
    messages, packet = service.govern_inference_messages(
        preset=_governor_preset(input_tokens=6000, preset_id="qwen-9b-restrictive"),
        profile_prefix=("Роль сметчика. " * 380),
        request_payload={"question": "Собери ЛСР", "attachment_id": "read_123456abcdef"},
        shortlist=service.selector_context_shortlist(
            [
                {"name": "search_sources", "input_schema": {"type": "object"}},
                {"name": "build_lsr_workbook", "input_schema": {"type": "object"}},
            ],
            native_tool_schemas=True,
        ),
        evidence=evidence,
    )

    visible = str(messages)
    assert "1. Работа 1 | шт. | 1" in visible
    assert "5. Работа 5 | шт. | 5" in visible
    assert not any(section.kind is ContextKind.TOOL_SHORTLIST for section in packet.sections)


def test_model_authored_first_search_is_not_labelled_as_empty_retrieval() -> None:
    empty_status = "[СТАТУС ИСТОЧНИКОВ: ФРАГМЕНТЫ НЕ НАЙДЕНЫ]"

    assert service.initial_selector_context(
        empty_status,
        model_authored_initial_query=True,
    ) == ""
    assert service.initial_selector_context(
        empty_status,
        model_authored_initial_query=False,
    ) == empty_status


def test_tool_selector_uses_thin_role_contract_and_bound_skill() -> None:
    prompt = service.profile_tool_selector_prompt(
        {
            "mode": "estimator",
            "prompt_text": "LONG ANSWER PROMPT MUST NOT ENTER TOOL DECISION",
            "skill_text": "Сначала изучи ВОР, затем ищи нормы.",
        }
    )

    assert "профиль estimator" in prompt
    assert "Сначала изучи ВОР, затем ищи нормы." in prompt
    assert "LONG ANSWER PROMPT" not in prompt
    assert "предметные решения" in prompt
    assert service.profile_uses_model_driven_retrieval(
        {
            "mode": "estimator",
            "rag_policy": {"system_datasets": ["smeta"]},
        }
    ) is True
    assert service.profile_uses_model_driven_retrieval(
        {
            "mode": "search",
            "rag_policy": {"model_authored_initial_query": True},
        }
    ) is True


def test_application_routes_model_search_sources_through_canonical_research_service() -> None:
    source = inspect.getsource(service._execute_chat_evidence_application)

    assert "ModelResearchToolService(" in source
    assert "model_research_tools.execute(call)" in source
    assert "retrieve=retrieve_chat_chunks" in source


def test_model_research_loop_is_evidence_first_and_model_stopped() -> None:
    source = inspect.getsource(service._execute_chat_evidence_application)

    initial_packet = source.index("initial_evidence_packet,")
    selector_loop = source.index("while time.monotonic() < research_deadline:")
    assert initial_packet < selector_loop
    assert "evidence=selector_evidence" in source
    assert "source_map=selector_source_map" in source
    assert "research_result.chunks" in source
    assert "seen_call_signatures" not in source
    assert 'stop_reason = "model_stop"' in source


def test_ordinary_chat_does_not_semantically_validate_or_rewrite_model_answer() -> None:
    source = inspect.getsource(service._execute_chat_evidence_application)

    assert "use_validation = False" in source
    assert "_chat_model_final_answer(answer" not in source
    assert "_notebook_study_validation_status(" not in source
    assert 'final_evidence_packet["evidence_status"] = "partial"' not in source
    assert 'crag_status = "UNVALIDATED"' not in source[source.index("citation_check ="):]


@pytest.mark.asyncio
async def test_shadow_failure_is_redacted_and_cannot_escape_to_legacy_path() -> None:
    class ThrowingHarness:
        async def call_async(self, *_args, **_kwargs):
            raise RuntimeError("secret candidate failure")

    trace = await service.safe_execute_canonical_shadow_decision(
        proposed_calls=[{"tool": "read_source", "args": {"doc_id": "d1"}}],
        allowed_tools={"read_source"},
        dataset_ids=["selected"],
        tool_harness=ThrowingHarness(),
    )

    assert trace["status"] == "error"
    assert trace["error_type"] == "RuntimeError"
    assert trace["attempted_calls"] == 1
    assert "secret candidate failure" not in str(trace)


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "scenario",
    [
        "normal",
        "normal_unscoped",
        "selector_overflow",
        "cloud_retry",
        "active",
        "active_workbook",
        "active_workbook_private_arg",
        "active_workbook_rejected",
        "active_model_rag_result",
        "active_model_rag_recovery",
        "candidate_workbook",
        "shadow_workbook",
        "legacy_workbook",
    ],
)
async def test_actual_chat_shadow_failure_preserves_legacy_answer_history_and_model_count(
    monkeypatch,
    tmp_path,
    scenario,
) -> None:
    selector_overflow = scenario == "selector_overflow"
    cloud_retry = scenario == "cloud_retry"
    workbook_profile = scenario in {
        "active_workbook",
        "active_workbook_private_arg",
        "active_workbook_rejected",
        "active_model_rag_result",
        "active_model_rag_recovery",
        "candidate_workbook",
        "shadow_workbook",
        "legacy_workbook",
    }
    active_workbook = scenario in {
        "active_workbook",
        "active_workbook_private_arg",
        "active_workbook_rejected",
        "active_model_rag_result",
        "active_model_rag_recovery",
        "candidate_workbook",
    }
    candidate_acceptance = scenario == "candidate_workbook"
    rejected_workbook = scenario == "active_workbook_rejected"
    model_rag_recovery = scenario == "active_model_rag_recovery"
    model_rag_result = scenario in {
        "active_model_rag_result",
        "active_model_rag_recovery",
    }
    active = scenario in {
        "active",
        "active_workbook",
        "active_workbook_private_arg",
        "active_workbook_rejected",
        "active_model_rag_result",
        "active_model_rag_recovery",
        "candidate_workbook",
    }
    inactive_workbook = workbook_profile and not active
    model_calls = []
    shortlist_policies = []
    shadow_calls = []
    executor_codes = []
    legacy_calls = []
    history_rows = []
    progress_events = []
    workbook_executor_calls = []
    rag_queries = []
    exact_decisions = [
        {
            "source_row": 1,
            "section": "ЭОМ",
            "title": "монтаж шкафа управления",
            "unit": "шт.",
            "quantity": 2.0,
            "norm_code": "ГЭСНм08-02-401-01",
            "analogue": "Монтаж оборудования по найденной карточке нормы",
            "coverage": "прямое соответствие",
            "evidence_refs": ["Q1.H1"],
        }
    ]
    protected_db = tmp_path / "protected.db"
    with sqlite3.connect(protected_db) as conn:
        conn.execute("CREATE TABLE protected_events (value TEXT NOT NULL)")

    def protected_hash():
        with sqlite3.connect(protected_db) as conn:
            rows = conn.execute(
                "SELECT value FROM protected_events ORDER BY rowid"
            ).fetchall()
        return hashlib.sha256(json.dumps(rows).encode("utf-8")).hexdigest()

    before_protected = protected_hash()

    class FakeResponse:
        status_code = 200

        def __init__(self, content):
            self._content = content

        def raise_for_status(self):
            return None

        def json(self):
            return {
                "choices": [{"message": {"content": self._content}}],
                "usage": {"completion_tokens": 7},
            }

    class FakeAsyncClient:
        def __init__(self, **_kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, *_args):
            return False

        async def post(self, _url, *, json, **_kwargs):
            system_text = str(json["messages"][0]["content"])
            if "исследовательским чтением LES" in system_text:
                model_calls.append("selector")
                if model_calls.count("selector") > 1:
                    return FakeResponse('{"calls":[]}')
                return FakeResponse(
                    '{"calls":['
                    '{"tool":"read_source","args":{"doc_id":"d1"}},'
                    '{"tool":"read_source","args":{"doc_id":"d2"}}]}'
                )
            if cloud_retry and "cloud.fixture" in _url:
                model_calls.append("cloud_final")
                raise service.httpx.ConnectError("fixture cloud failure")
            if cloud_retry:
                model_calls.append("local_final")
                if model_calls.count("local_final") == 1:
                    return FakeResponse("")
                return FakeResponse("legacy visible answer")
            model_calls.append("final")
            return FakeResponse("legacy visible answer")

    from proxy.services.tool_contract_service import (
        EffectClass,
        IdempotencyPolicy,
        ResultBudget,
        RetryPolicy,
        ToolContract,
    )
    from proxy.services.tool_registry_service import ToolRegistration, ToolRegistry
    from proxy.services.trusted_executor_service import (
        ExecutionRequest,
        TrustedExecutor,
    )

    def persistence_probe_handler(args):
        with sqlite3.connect(protected_db) as conn:
            conn.execute("INSERT INTO protected_events VALUES (?)", (args["doc_id"],))
        return {
            "schema": "les_tool_result_v1",
            "tool": "read_source",
            "operation": "read",
            "inputs": [dict(args)],
            "status": "ok",
            "result": {},
            "sources": [],
            "missing": [],
            "warnings": [],
            "trace": "persistence probe",
            "decision_required_from_model": True,
        }

    registry = ToolRegistry(
        [
            ToolRegistration(
                contract=ToolContract(
                    name="read_source",
                    version="1.0.0",
                    title="Read source",
                    category="source",
                    summary="Read one source",
                    input_schema={
                        "type": "object",
                        "required": ["doc_id"],
                        "properties": {"doc_id": {"type": "string"}},
                        "additionalProperties": False,
                    },
                    result_schema="les_tool_result_v1",
                    effect=EffectClass.READ,
                    scopes=("dataset",),
                    timeout_seconds=30,
                    retry=RetryPolicy.SAFE,
                    idempotency=IdempotencyPolicy.DERIVED,
                    result_budget=ResultBudget(max_chars=7000, max_items=20),
                    model_owned_fields=(),
                    provenance="source_refs_required",
                    tags=("shadow_validate_only",),
                ),
                handler=persistence_probe_handler,
            )
        ]
    )
    executor = TrustedExecutor(
        registry,
        scope_resolver=lambda _contract, _args: ("selected",),
    )

    class ExecutorBackedHarness:
        def shortlist(self, *_args, **kwargs):
            shortlist_policies.append(dict(kwargs))
            if model_rag_result:
                return {
                    "schema": "les_tool_shortlist_v1",
                    "tools": [
                        *(
                            [
                                {
                                    "name": "build_vor_workbook",
                                    "summary": "Extract a VOR workbook without estimating it",
                                    "input_schema": {
                                        "type": "object",
                                        "properties": {"attachment_id": {"type": "string"}},
                                        "required": ["attachment_id"],
                                        "additionalProperties": False,
                                    },
                                }
                            ]
                            if model_rag_recovery
                            else []
                        ),
                        {
                            "name": "search_sources",
                            "summary": "Search frozen datasets with native RRF",
                            "input_schema": {
                                "type": "object",
                                "properties": {"q": {"type": "string"}},
                                "required": ["q"],
                                "additionalProperties": False,
                            },
                        },
                        {
                            "name": "build_lsr_workbook",
                            "summary": "Render model decisions as XLSX",
                            "input_schema": {
                                "type": "object",
                                "properties": {
                                    "attachment_id": {"type": "string"},
                                    "decisions": {"type": "array"},
                                },
                                "required": ["attachment_id", "decisions"],
                                "additionalProperties": False,
                            },
                        },
                    ],
                }
            return {
                "schema": "les_tool_shortlist_v1",
                "tools": [{"name": "build_vor_workbook" if workbook_profile else "read_source"}],
            }

        async def call_async(self, tool, args, **policy):
            shadow_calls.append((tool, dict(args)))
            envelope = await executor.execute(
                ExecutionRequest(
                    call_id="shadow-call-1",
                    tool_name=tool,
                    arguments=args,
                    allowed_dataset_ids=tuple(policy["allowed_dataset_ids"]),
                    actor_id=str(policy["actor_id"]),
                    actor_role=str(policy["actor_role"]),
                    approval_receipt_id=None,
                    idempotency_key=None,
                    deadline_monotonic=float(
                        policy.get("deadline_monotonic", time.monotonic() + 120)
                    ),
                    shadow=bool(policy["shadow"]),
                )
            )
            executor_codes.append(envelope.code)
            if not workbook_profile:
                assert envelope.code == (
                    "TOOL_SCOPE_VIOLATION"
                    if scenario == "normal_unscoped"
                    else "TOOL_WOULD_EXECUTE"
                )
            raise RuntimeError("secret shadow failure")

        def call(self, tool, args):
            legacy_calls.append((tool, dict(args)))
            return {
                "schema": "les_tool_result_v1",
                "tool": tool,
                "operation": "read",
                "inputs": [dict(args)],
                "status": "missing",
                "result": {},
                "sources": [],
                "missing": ["test fixture"],
                "warnings": [],
                "trace": "legacy fixture",
                "decision_required_from_model": True,
            }

    class FakeRetrieval:
        trace = SimpleNamespace(status="ok", error_code="")
        quality = SimpleNamespace(status="weak", top_score=0.0)

        def __init__(self, chunks=()):
            self.chunks = list(chunks)

        def payload(self):
            return {"schema": "retrieval_trace_v1", "status": "ok"}

    class FakeWindows:
        def __init__(self, chunks=()):
            self.chunks = list(chunks)

        def payload(self):
            return {"schema": "context_windows_v1", "count": 0}

    runtime_config = SimpleNamespace(
        provider="openai" if cloud_retry else "mlx",
        model="cloud-fixture-model" if cloud_retry else "fixture-model",
        base_url="http://cloud.fixture" if cloud_retry else "http://fixture.invalid",
        chat_url=(
            "http://cloud.fixture/v1/chat/completions"
            if cloud_retry else "http://fixture.invalid/v1/chat/completions"
        ),
        api_key="",
        supports_validation=not cloud_retry,
    )
    mlx_runtime_config = SimpleNamespace(
        provider="mlx",
        model="local-fixture-model",
        base_url="http://local.fixture",
        chat_url="http://local.fixture/v1/chat/completions",
        api_key="",
        supports_validation=True,
    )
    state = SimpleNamespace(
        reranker_available=False,
        reranker_cls=None,
        llm_semaphore=asyncio.Semaphore(1),
        metrics_cache={},
        crag_stats={"no_data": 0, "hallucination": 0, "verified": 0},
        chat_metrics={
            "latency_search": [],
            "latency_gen": [],
            "tokens": [],
            "crag_fail": 0,
        },
    )

    from proxy.services import tool_harness_service

    monkeypatch.setattr(tool_harness_service, "harness", lambda: ExecutorBackedHarness())
    monkeypatch.setattr(service.httpx, "AsyncClient", FakeAsyncClient)
    monkeypatch.setattr(
        service,
        "rules_pre_verdict",
        lambda *_args, **_kwargs: pytest.fail("ordinary chat must not invoke semantic validator"),
    )
    monkeypatch.setattr(service, "maybe_answer_table_query", lambda *_a, **_k: None)
    monkeypatch.setattr(service, "dataset_memory_prompt_excerpt", lambda *_a, **_k: "")
    active_transport = None
    active_connection = None
    if active:
        from proxy.services.canonical_route_service import (
            CanonicalRouteDecision,
            CanonicalRouteMode,
        )
        from proxy.services.model_connection_contracts import ConnectionLocality, ConnectionRole
        from proxy.services.openai_compatible_transport_service import InferenceResponse

        active_connection = SimpleNamespace(
            connection_id="conn:active",
            revision_id="conn:active:r3",
            display_name="Active fixture",
            model_id="active-model",
            locality=ConnectionLocality.LOOPBACK,
            base_url="http://127.0.0.1:1919/v1",
            secret_ref=None,
            effective_preset=_governor_preset(
                input_tokens=6000,
                preset_id="active-preset",
            ),
        )

        class ActiveResolver:
            def resolve(self, role, **_kwargs):
                assert role is ConnectionRole.ANSWER
                return active_connection

            def resolve_fallback(self, *_args, **_kwargs):
                raise AssertionError("fallback must not be used")

        class ActiveTransport:
            def __init__(self):
                self.revisions = []
                self.requests = []
                if model_rag_result:
                    self.responses = [
                        InferenceResponse(
                            text=(
                                "монтаж шкафа управления\n"
                                "прокладка контрольного кабеля"
                            ),
                            tool_calls=(),
                            finish_reason="stop",
                            usage={},
                        ),
                        InferenceResponse(
                            text=(
                                "active visible answer\n\n"
                                "| source_row | section | title | unit | quantity | norm_code | analogue | coverage | coefficient | evidence_refs |\n"
                                "|---:|---|---|---|---:|---|---|---|---:|---|\n"
                                "| 1 | ЭОМ | монтаж шкафа управления | шт. | 2 | ГЭСНм08-02-401-01 | "
                                "Монтаж оборудования по найденной карточке нормы | прямое соответствие | | Q1.H1 |"
                            ),
                            tool_calls=(),
                            finish_reason="stop",
                            usage={"completion_tokens": 5},
                        ),
                    ]
                else:
                    self.responses = [
                        InferenceResponse(
                            text="",
                            tool_calls=(
                                (
                                    {"id": "workbook-call-1", "type": "function", "function": {"name": "build_vor_workbook", "arguments": '{"attachment_id":"read_123456abcdef"}'} },
                                    {"id": "workbook-call-2", "type": "function", "function": {"name": "build_vor_workbook", "arguments": '{"attachment_id":"read_123456abcdef"}'} },
                                )
                                if rejected_workbook
                                    else ({"id": "workbook-call-1", "type": "function", "function": {"name": "build_vor_workbook", "arguments": ('{"attachment_id":"read_123456abcdef","question":"C:/private/les/SECRET_TRACE_TOKEN"}' if scenario == "active_workbook_private_arg" else '{"attachment_id":"read_123456abcdef"}')} },)
                            ) if active_workbook else (
                                {"id": "c1", "type": "function", "function": {"name": "read_source", "arguments": '{"doc_id":"d1"}'}},
                                {"id": "c2", "type": "function", "function": {"name": "read_source", "arguments": '{"doc_id":"d2"}'}},
                            ),
                            finish_reason="tool_calls",
                            usage={},
                        ),
                        InferenceResponse(
                            text=(
                                '{"calls":[]}'
                                if rejected_workbook or not active_workbook
                                else "active visible answer"
                            ),
                            tool_calls=(),
                            finish_reason="stop",
                            usage={"completion_tokens": 5},
                        ),
                        InferenceResponse(
                            text="active visible answer",
                            tool_calls=(),
                            finish_reason="stop",
                            usage={"completion_tokens": 5},
                        ),
                    ]

            async def complete(self, connection, _request):
                self.revisions.append(connection.revision_id)
                self.requests.append(_request)
                return self.responses.pop(0)

        active_transport = ActiveTransport()
        monkeypatch.setattr(
            service,
            "resolve_canonical_route",
            lambda **_kwargs: CanonicalRouteDecision(
                requested=(
                    CanonicalRouteMode.SHADOW
                    if candidate_acceptance
                    else CanonicalRouteMode.ACTIVE
                ),
                effective=(
                    CanonicalRouteMode.SHADOW
                    if candidate_acceptance
                    else CanonicalRouteMode.ACTIVE
                ),
                source="test",
                reason=("receipt_missing" if candidate_acceptance else "exact_test_receipt"),
                restart_required=False,
            ),
        )
    elif scenario == "legacy_workbook":
        from proxy.services.canonical_route_service import (
            CanonicalRouteDecision,
            CanonicalRouteMode,
        )

        monkeypatch.setattr(
            service,
            "resolve_canonical_route",
            lambda **_kwargs: CanonicalRouteDecision(
                requested=CanonicalRouteMode.LEGACY,
                effective=CanonicalRouteMode.LEGACY,
                source="test",
                reason="explicit_legacy",
                restart_required=False,
            ),
        )
    if cloud_retry:
        monkeypatch.setattr(
            service,
            "decide_provider",
            lambda *_a, **_k: SimpleNamespace(
                downgraded=False,
                sensitivity="P0",
                reason="fixture cloud route",
            ),
        )
        monkeypatch.setattr(
            service,
            "resolve_transport_execution_profile",
            lambda *, provider, **_kwargs: _governor_preset(
                input_tokens=12_000 if provider == "openai" else 6_000,
                preset_id="cloud-large" if provider == "openai" else "local-restrictive",
            ),
        )
    if selector_overflow:
        original_govern = service.govern_inference_messages

        def overflow_selector(**kwargs):
            if kwargs.get("shortlist"):
                raise ContextRequiredSectionOverflow(
                    object_ids=("profile:bound", "request:current"),
                    budget=10,
                    required_tokens=11,
                )
            return original_govern(**kwargs)

        monkeypatch.setattr(service, "govern_inference_messages", overflow_selector)

    def save_history(**row):
        history_rows.append(row)
        return "history-1"

    request = service.EvidenceRequestContext(
        req=SimpleNamespace(
            question="Собери ЛСР" if model_rag_result else "Проверь документы",
            mode="estimator" if model_rag_result else "rag",
            response_length="short",
            output_directive="",
            session_id="session-1",
            dataset_filter=None,
                project_id=0,
                attachment_id="read_123456abcdef" if workbook_profile else None,
                attachment_context=(
                    "Строка 1: монтаж шкафа управления — 2 шт."
                    if model_rag_result
                    else ""
                ),
            candidate_acceptance=candidate_acceptance,
            reranker_enabled=None,
        ),
        dataset_ids=None if scenario == "normal_unscoped" else ["selected"],
        effective_dataset_filter="",
        resolved_dataset_names=[],
        dataset_name_by_id={},
        query_route_payload={"channel": "rag"},
        target_doc_filter=[],
        target_file_ref=None,
        topic_doc_filter=[],
        topic_retrieval_plan=None,
        inventory_requested=False,
        study_requested=False,
        memory_block="",
        session_block="",
        class_suggestions=[],
        use_semantic_cache=False,
        use_validation=False,
        validation_skip_reason="test",
        route=SimpleNamespace(intent="rag"),
        table_result=None,
        request_started_at=1.0,
        profile_snapshot={
            "revision_id": "fixture-profile:1",
            "mode": "estimator" if model_rag_result else "rag",
            "tools": (
                ["search_sources", "build_lsr_workbook"]
                if model_rag_result
                else ["build_vor_workbook" if workbook_profile else "read_source"]
            ),
            "rag_policy": (
                {
                    "iterative": True,
                    "system_datasets": ["smeta"],
                    "model_authored_initial_query": True,
                }
                if model_rag_result
                else {"iterative": False}
            ),
            "prompt_text": "Answer only from evidence.",
        },
    )

    def retrieve_fixture(**kwargs):
        if not model_rag_result:
            return _async_value(FakeRetrieval())
        pytest.fail("estimator model-RAG must not call the general les_rag retriever")

    def retrieve_smeta_fixture(query, *, limit):
        assert limit == 6
        rag_queries.append(query)
        return {
            "backend": "typed_sqlite_fts+smeta_norm_qdrant_hybrid",
            "cards": [
                {
                    "norm_code": "ГЭСНм08-02-401-01",
                    "title": f"Карточка нормы для {query}",
                    "measure_unit": "шт.",
                    "work_steps": ["Монтаж оборудования"],
                    "source_ref": f"fsnb#query={len(rag_queries)}",
                }
            ],
            "retrieval_trace": {
                "rag": {
                    "status": "ok",
                    "collection": "les_smeta_norm_cards",
                    "retrieval_channels": ["dense", "bm25_sparse"],
                    "fusion": "rrf",
                }
            },
        }

    monkeypatch.setattr(service, "retrieve_smeta_norm_cards", retrieve_smeta_fixture)

    runtime = service.EvidenceRuntimeDeps(
        state=state,
        rag_backend=SimpleNamespace(collection_name="fixture"),
        cache=SimpleNamespace(),
        cache_embedding=None,
        cache_marker="miss",
        cache_scope="",
        assistant_text=lambda message: str(message.get("content") or ""),
        augment_model_tool_args=lambda call, **_kwargs: call,
        chat_model_final_answer=lambda _answer, status: ("rewritten by code", status, {}),
        cloud_body_for_model=lambda body, *_args: body,
        compact_tool_result_for_prompt=lambda item, **_kwargs: item,
        dataset_ids_from_chunks=lambda _chunks: [],
        dataset_sensitivities=lambda _ids: [],
        env_bool=lambda _key, default=False: default,
        env_float=lambda _key, default=0.0: default,
        env_int=lambda _key, default=0: default,
        expand_context_windows=lambda chunks, *_args, **_kwargs: FakeWindows(chunks),
        format_tool_results_for_model=lambda rows: json.dumps(rows, ensure_ascii=False),
        generation_token_budget=lambda **_kwargs: 128,
        llm_runtime=lambda: runtime_config,
        local_context_budget=lambda **_kwargs: {
            "focus_max_chunks": 0,
            "context_max_chunks": 0,
            "context_chars_limit": 4000,
            "context_window_chars": 1000,
        },
        mlx_runtime=lambda: mlx_runtime_config if cloud_retry else runtime_config,
        names_for_dataset_ids=lambda *_args: [],
        notebook_study_validation_status=lambda status, **_kwargs: status,
        ollama_native_complete=lambda *_args, **_kwargs: None,
        parse_model_tool_calls=(
            lambda raw, *_args, **_kwargs: (
                [{"tool": "build_vor_workbook", "args": {"attachment_id": "read_123456abcdef"}}]
                    if workbook_profile and not active
                    else json.loads(raw).get("calls", [])
            )
        ),
        prepare_notebook_reader_memory=lambda *_args, **_kwargs: None,
        record_cloud_cost=lambda *_args, **_kwargs: None,
        retrieve_chat_chunks=retrieve_fixture,
        source_excerpts=lambda *_args, **_kwargs: [],
        table_query_response=lambda **_kwargs: None,
        cloud_fallback_models=lambda runtime: [runtime.model] if cloud_retry else [],
        cloud_model_timeout=lambda: 1.0,
        model_connection_resolver=(
            (lambda: (ActiveResolver(), object())) if active else None
        ),
        model_connection_transport=(
            (lambda _client, _secret_store: active_transport) if active else None
        ),
        workbook_tool_executor=(
            (lambda call, context, progress: _rejected_workbook_execution(call, workbook_executor_calls))
            if rejected_workbook
            else (
                (
                    lambda call, context, progress: (
                        _rejected_workbook_execution(call, workbook_executor_calls)
                        if model_rag_recovery and call.get("tool") == "build_vor_workbook"
                        else _tracked_workbook_execution(
                            call, context, progress, progress_events, workbook_executor_calls
                        )
                    )
                )
                if active_workbook else None
            )
        ),
    )
    boundary = service.ResponseBoundary(
        save_chat_history=save_history,
        token_sink=(lambda event: _async_append(progress_events, event)) if active_workbook else None,
        version_stamp=lambda: {},
    )

    if selector_overflow:
        with pytest.raises(service.HTTPException) as error:
            await service.run_chat_evidence_application(request, runtime, boundary)
        assert error.value.status_code == 422
        assert error.value.detail["code"] == "CONTEXT_REQUIRED_SECTION_OVERFLOW"
        assert model_calls == []
        assert history_rows == []
        assert protected_hash() == before_protected
        return

    result = await service.run_chat_evidence_application(request, runtime, boundary)

    after_protected = protected_hash()
    if model_rag_result:
        assert result["answer"].startswith("active visible answer\n\n| source_row |")
        assert rag_queries == [
            "монтаж шкафа управления",
            "прокладка контрольного кабеля",
        ]
        assert "Собери ЛСР" not in rag_queries
        assert workbook_executor_calls[-1]["args"]["decisions"] == exact_decisions
        assert active_transport.revisions == ["conn:active:r3"] * 2
        assert active_transport.requests[0].tools == ()
        assert active_transport.requests[1].tools == ()
        assert "Строка 1: монтаж шкафа управления — 2 шт." in str(
            active_transport.requests[0].messages
        )
        final_messages = str(active_transport.requests[1].messages)
        assert "Строка 1: монтаж шкафа управления — 2 шт." in final_messages
        assert "Карточка нормы для монтаж шкафа управления" in final_messages
        assert "Карточка нормы для прокладка контрольного кабеля" in final_messages
        assert "[Поисковый запрос Q1] монтаж шкафа управления" in final_messages
        assert "Q1.H1" in final_messages
        assert "[Поисковый запрос Q2] прокладка контрольного кабеля" in final_messages
        assert "Q2.H1" in final_messages
        first_messages = str(active_transport.requests[0].messages)
        assert "Строка 1: монтаж шкафа управления — 2 шт." in first_messages
        assert '"queries"' not in first_messages
        assert "верни только сами поисковые запросы" in first_messages
        assert "фиксированного лимита нет" in first_messages
        assert "по всем фактическим строкам" in first_messages
        assert "Код не подтверждает и не меняет твой выбор" in final_messages
        assert history_rows[0]["retrieval_trace"]["status"] == "model_driven"
        assert history_rows[0]["retrieval_trace"]["reason"] == (
            "awaiting_model_authored_query"
        )
        tool_loop = history_rows[0]["retrieval_trace"]["tool_loop"]
        assert tool_loop["schema"] == "les_model_rag_batch_v1"
        assert tool_loop["model_queries"] == rag_queries
        assert tool_loop["evidence_groups"] == ["Q1", "Q2"]
        assert "rounds" not in tool_loop
        assert "review" not in json.dumps(history_rows[0], ensure_ascii=False).casefold()
        assert "confirm" not in json.dumps(history_rows[0], ensure_ascii=False).casefold()
        return
    if inactive_workbook:
        assert result["answer"] == "legacy visible answer"
        assert shortlist_policies == []
        assert shadow_calls == []
        assert legacy_calls == []
        assert executor_codes == []
        assert after_protected == before_protected
        return
    if active:
        assert result["answer"] == "active visible answer"
        assert result["model_connection"] == {
            "connection_id": "conn:active",
            "revision_id": "conn:active:r3",
            "display_name": "Active fixture",
            "model_id": "active-model",
                "locality": "loopback",
                "fallback_used": False,
                "pending_tool_calls": 0,
        }
        assert result["source_scope"] == {
            "requested": ["selected"],
            "resolved": ["selected"],
            "used": [],
            "used_names": [],
        }
        assert active_transport.revisions == ["conn:active:r3", "conn:active:r3"] + (
            ["conn:active:r3"]
            if rejected_workbook or not active_workbook
            else []
        )
        if active_workbook:
            assert len(workbook_executor_calls) == (2 if rejected_workbook else 1)
            if rejected_workbook:
                assert "artifact" not in result
                assert history_rows[0]["retrieval_trace"]["tool_loop"]["stop_reason"] == "model_stop"
                return
            assert result["artifact"]["revision_id"] == "rev-2"
            assert result["attachment_retry"]["attachment_id"] == "read_123456abcdef"
            assert history_rows[0]["artifact"]["revision_id"] == "rev-2"
            if scenario == "active_workbook_private_arg":
                selected_calls = result["retrieval_trace"]["tool_loop"]["selected_calls"]
                assert selected_calls == history_rows[0]["retrieval_trace"]["tool_loop"]["selected_calls"]
                assert selected_calls[0]["tool"] == "build_vor_workbook"
                assert selected_calls[0]["call_id"] == "workbook-call-1"
                assert len(selected_calls[0]["arguments_sha256"]) == 64
                assert "args" not in selected_calls[0]
                assert "SECRET_TRACE_TOKEN" not in json.dumps(result, ensure_ascii=False)
                assert "C:/private/les" not in json.dumps(history_rows[0], ensure_ascii=False)
            assert any(event.get("event") == "tool_progress" for event in progress_events)
            assert legacy_calls == []
            if candidate_acceptance:
                trace = history_rows[0]["retrieval_trace"]
                assert trace["canonical_route"]["effective"] == "shadow"
                assert trace["candidate_acceptance"] == {
                    "enabled": True,
                    "execution_mode": "active",
                    "promotion_receipt": "not_used",
                    "state_root": "process_cwd_isolated",
                }
                assert trace["route_comparison"]["candidate_acceptance"] is True
                assert trace["route_comparison"]["legacy_output_authoritative"] is False
        else:
            assert [call[1]["doc_id"] for call in legacy_calls] == ["d1", "d2"]
        assert history_rows[0]["retrieval_trace"]["context_governor"]["preset_id"] == "active-preset"
        assert after_protected == before_protected
        return
    assert result["answer"] == "legacy visible answer"
    if cloud_retry:
        assert model_calls == [
            "selector",
            "selector",
            "cloud_final",
            "local_final",
            "local_final",
        ]
        calls = history_rows[0]["retrieval_trace"]["context_governor"]["calls"]
        assert [call["purpose"] for call in calls] == [
            "tool_decision",
            "tool_decision",
            "answer",
            "answer_fallback",
            "answer",
        ]
        assert [call["preset_id"] for call in calls] == [
            "cloud-large",
            "cloud-large",
            "cloud-large",
            "local-restrictive",
            "local-restrictive",
        ]
        assert history_rows[0]["retrieval_trace"]["canonical_shadow"]["persisted_effects"] == 0
        assert after_protected == before_protected
        return
    if scenario == "normal_unscoped":
        assert model_calls == ["final"]
        assert shortlist_policies == []
        assert shadow_calls == []
        assert executor_codes == []
        assert legacy_calls == []
        assert history_rows[0]["retrieval_trace"]["status"] == "skipped"
        assert history_rows[0]["retrieval_trace"]["reason"] == "scope_none"
        assert after_protected == before_protected
        return
    assert model_calls == ["selector", "selector", "final"]
    assert shortlist_policies == [
        {
            "mode": "rag",
            "allowed_tools": ["read_source"],
            "limit": 5,
            "dataset_ids": () if scenario == "normal_unscoped" else ("selected",),
            "workflow_phase": "research",
            "model_preset": "qwen-9b-restrictive",
            "runtime_available": frozenset({"read_source"}),
            "calls_remaining": 5,
            "result_chars_remaining": 35_000,
        }
    ]
    assert [call[1]["doc_id"] for call in shadow_calls] == ["d1"]
    assert executor_codes == [
        "TOOL_SCOPE_VIOLATION" if scenario == "normal_unscoped" else "TOOL_WOULD_EXECUTE"
    ]
    assert [call[1]["doc_id"] for call in legacy_calls] == ["d1", "d2"]
    assert len(history_rows) == 1
    assert history_rows[0]["answer"] == "legacy visible answer"
    assert history_rows[0]["retrieval_trace"]["canonical_shadow"]["status"] == "error"
    assert history_rows[0]["retrieval_trace"]["canonical_shadow"]["attempted_calls"] == 1
    assert history_rows[0]["retrieval_trace"]["canonical_shadow"]["pending_calls"] == 1
    assert history_rows[0]["retrieval_trace"]["canonical_shadow"]["persisted_effects"] == 0
    assert history_rows[0]["retrieval_trace"]["canonical_shadow"]["profile_revision"] == "fixture-profile:1"
    assert history_rows[0]["retrieval_trace"]["route_comparison"] == {
        "schema": "les.canonical-route-comparison.v1",
        "requested": "shadow",
        "effective": "shadow",
        "legacy_output_authoritative": True,
        "same_request": True,
        "profile_revision": "fixture-profile:1",
        "canonical_provider_calls_added": 0,
        "persisted_effects": 0,
    }
    governor_trace = history_rows[0]["retrieval_trace"]["context_governor"]
    assert governor_trace["preset_id"] == "qwen-9b-restrictive"
    assert [call["purpose"] for call in governor_trace["calls"]] == [
        "tool_decision",
        "tool_decision",
        "answer",
    ]
    assert "Проверь документы" not in json.dumps(governor_trace, ensure_ascii=False)
    assert "legacy fixture" not in json.dumps(governor_trace, ensure_ascii=False)
    assert "secret shadow failure" not in str(history_rows[0])
    assert after_protected == before_protected


def test_estimator_packages_lsr_or_vor_from_the_operator_request_not_both():
    tools = ["search_sources", "build_lsr_workbook", "build_vor_workbook"]
    assert service.estimator_workbook_packaging_tools("Собери ЛСР", tools) == [
        "build_lsr_workbook"
    ]
    assert service.estimator_workbook_packaging_tools("Собери ВОР", tools) == [
        "build_vor_workbook"
    ]
    assert service.estimator_workbook_packaging_tools("Собери ЛСР и ВОР", tools) == [
        "build_lsr_workbook",
        "build_vor_workbook",
    ]


def test_workbook_filename_projection_keeps_basename_and_drops_paths():
    assert service._safe_workbook_filename("LSR_source_2026-09-01_2147.xlsx") == (
        "LSR_source_2026-09-01_2147.xlsx"
    )
    assert service._safe_workbook_filename("C:/private/workbook.xlsx") == ""
    assert service._safe_workbook_filename("artifact.xlsx") == ""
    assert service._chat_workbook_filename(
        "",
        artifact_kind="vor_workbook",
    ).startswith("VOR_")
    assert "artifact.xlsx" not in service._chat_workbook_filename(
        "artifact.xlsx",
        artifact_kind="lsr_workbook",
    )


def test_model_rag_result_reads_normative_code_header_and_skips_missing_numbers():
    answer = """
| № | Наименование | Ед. изм. | Кол-во | Нормативный код (ГЭСН/ЕР) | Коэфф. | Примечание |
|---:|---|---|---:|---|---|---|
| 1 | Кабель | м | 12 | ГЭСНм08-02-409-09 | — |  |
| 2 | Плёнка | м2 |  | ГЭСН26-01-055-02 | 1/100 | нет прямого аналога |
"""
    parsed = service.parse_model_rag_result(answer)
    assert parsed is not None
    _, rows = parsed
    assert rows[0]["norm_code"] == "ГЭСНм08-02-409-09"
    assert "coefficient" not in rows[0]
    assert rows[1]["coefficient"] == 0.01
    assert "quantity" not in rows[1]


def test_compact_draft_rows_keep_model_owned_codes():
    compact = service.compact_estimator_draft_rows(
        [{"source_row": 3, "title": "Кабель", "quantity": 12, "unit": "м", "norm_code": "ГЭСНм08-02-409-09"}]
    )
    assert compact == [{
        "work_id": "3",
        "source_row": 3,
        "section": "",
        "title": "Кабель",
        "quantity": 12,
        "unit": "м",
        "norm_code": "ГЭСНм08-02-409-09",
        "decision": "ГЭСНм08-02-409-09",
    }]


async def _async_value(value):
    return value

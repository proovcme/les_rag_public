"""Тесты каркаса checklist_review_service (T2.1 presence + T2.2 calculation, TDD red->green).

АРХИТЕКТУРА (implementation_plan.md §2-§4, docs/ALGO-glorax-checklist-review.md):
run_checklist_review принимает template (dict, как из checklist_template_importer) и
инжектированные inventory_provider/search_provider/workbook_provider — НЕ монкипатчит живую
MetaDB/Qdrant. Паттерн инъекции зависимостей — по образцу run_review(doc_set, review_map, ...) в
doc_review_service.py: чистая функция без обращений к живым сервисам, окружение приходит
параметрами. Продакшн-провайдеры (default_inventory_provider/default_search_provider/
default_workbook_provider) — обёртки с lazy-импортами, в тестах НЕ вызываются.

T2.1 реализовал presence-механизм + честные заглушки остальных kind. T2.2 добавляет
calculation-механизм (openpyxl formula-count, 0 LLM) поверх нового workbook_provider.
T3.1 добавляет parametric-механизм (glorax_param_rules.yaml). T3.2 (A) подключает
параметрические правила поверх ЛЮБОГО kind (составные manual_required-пункты с числовым
порогом внутри) и (B) реализует two-sided gate для cross_section (project_doc/source_doc
evidence через source_dataset_ids). spds_formal остаётся заглушкой (review_needed/not_run).
Без LLM-вызовов вообще.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from proxy.services.checklist_review_service import (
    _anchor_terms,
    load_checklist_template,
    run_checklist_review,
)

# ── фикстуры: мини-template (не читаем реальный JSON) ──────────────────────────────────


def _mini_template() -> dict:
    return {
        "name": "mini_test_template",
        "stage": "PD",
        "title": "Мини-шаблон для теста",
        "source_file_name": "mini.xlsx",
        "version": "",
        "disciplines": ["Общее", "АР"],
        "items": [
            {
                "id": "PD-OB-003",
                "stage": "PD",
                "discipline": "Общее",
                "sheet_name": "Общее",
                "row": 3,
                "item_no": "1",
                "section_path": [],
                "criterion": "Приложен отчет об инженерно-геологических изысканиях",
                "answer_cell": "C3",
                "note_cell": "D3",
                "allowed_answers": ["Да", "Нет"],
                "kind": "presence",
            },
            {
                "id": "PD-OB-004",
                "stage": "PD",
                "discipline": "Общее",
                "sheet_name": "Общее",
                "row": 4,
                "item_no": "2",
                "section_path": [],
                "criterion": "Приложен отчет об инженерно-геодезических изысканиях",
                "answer_cell": "C4",
                "note_cell": "D4",
                "allowed_answers": ["Да", "Нет"],
                "kind": "presence",
            },
            {
                "id": "PD-AR-013",
                "stage": "PD",
                "discipline": "АР",
                "sheet_name": "АР",
                "row": 13,
                "item_no": "5",
                "section_path": ["Архитектурные решения"],
                "criterion": "Теплотехнический расчет",
                "answer_cell": "C13",
                "note_cell": "D13",
                "allowed_answers": ["Да", "Нет"],
                "kind": "calculation",
            },
            {
                "id": "PD-KR-020",
                "stage": "PD",
                "discipline": "КР",
                "sheet_name": "КР",
                "row": 20,
                "item_no": "6",
                "section_path": ["Конструктивные решения"],
                "criterion": "Марка бетона соответствует требованиям по водонепроницаемости для ростверка",
                "answer_cell": "C20",
                "note_cell": "D20",
                "allowed_answers": ["Да", "Нет"],
                "kind": "parametric",
            },
            {
                "id": "PD-SPOZU-010",
                "stage": "PD",
                "discipline": "СПОЗУ",
                "sheet_name": "СПОЗУ",
                "row": 10,
                "item_no": "7",
                "section_path": [],
                "criterion": "Соответствие пирогов благоустройства ТЗ и техническому стандарту Glorax",
                "answer_cell": "C10",
                "note_cell": "D10",
                "allowed_answers": ["Да", "Нет"],
                "kind": "cross_section",
            },
            {
                "id": "PD-AR-049",
                "stage": "PD",
                "discipline": "АР",
                "sheet_name": "АР",
                "row": 49,
                "item_no": "8",
                "section_path": ["Архитектурные решения"],
                "criterion": "Расположение корзин для кондиционеров",
                "answer_cell": "C49",
                "note_cell": "D49",
                "allowed_answers": ["Да", "Нет"],
                "kind": "spatial_visual",
            },
            {
                "id": "PD-AR-005",
                "stage": "PD",
                "discipline": "АР",
                "sheet_name": "АР",
                "row": 5,
                "item_no": "1.1.1",
                "section_path": ["Общие данные"],
                "criterion": "Общие данные комплекта приложены",
                "answer_cell": "C5",
                "note_cell": "D5",
                "allowed_answers": ["Да", "Нет"],
                "kind": "spds_formal",
            },
        ],
    }


class _FakeSearchProvider:
    """Инъектируемый поисковый провайдер: dataset_id x terms -> hits.

    Каждый экземпляр настраивается через ``hits_by_terms_key`` (тестовая заглушка), НЕ дергает
    lexical_index_service/Qdrant/MetaDB — соответствует контракту "инъекция зависимостей, не
    монкипатчинг живой инфраструктуры" из промпта.
    """

    def __init__(self, hits: dict[str, list[dict]] | None = None):
        self._hits = hits or {}
        self.calls: list[tuple[str, tuple[str, ...]]] = []

    def __call__(self, dataset_id: str, terms: list[str]) -> list[dict]:
        self.calls.append((dataset_id, tuple(terms)))
        return list(self._hits.get(dataset_id, []))


class _FakeInventoryProvider:
    def __init__(self, files: dict[str, list[dict]] | None = None):
        self._files = files or {}
        self.calls: list[str] = []

    def __call__(self, dataset_id: str) -> list[dict]:
        self.calls.append(dataset_id)
        return list(self._files.get(dataset_id, []))


def _empty_search(dataset_id: str, terms: list[str]) -> list[dict]:
    return []


def _empty_inventory(dataset_id: str) -> list[dict]:
    return []


# ── _anchor_terms: детерминированный вывод якорей из критерия ──────────────────────────


@pytest.mark.parametrize(
    "criterion,expected",
    [
        (
            "Приложен отчет об инженерно-геологических изысканиях",
            ["отчет", "инженерно-геологических", "изысканиях"],
        ),
        (
            "Приложен отчет об инженерно-геодезических изысканиях",
            ["отчет", "инженерно-геодезических", "изысканиях"],
        ),
        (
            "Теплотехнический расчет",
            ["теплотехнический", "расчет"],
        ),
    ],
)
def test_anchor_terms_extracts_content_words(criterion, expected):
    terms = _anchor_terms(criterion)
    # регистр и порядок значимых слов — детерминированы; стоп-слова/предлоги отфильтрованы
    assert terms == expected


def test_anchor_terms_drops_stopwords_and_punctuation():
    terms = _anchor_terms("Приложен отчет об инженерно-геологических изысканиях")
    assert "об" not in terms
    assert all(t == t.lower() for t in terms)
    assert all(t.strip(",.;:()") == t for t in terms)


# ── load_checklist_template ─────────────────────────────────────────────────────────────


def test_load_checklist_template_reads_real_glorax_pd():
    template = load_checklist_template("glorax_pd_2026")
    assert template["name"] == "glorax_pd_2026"
    assert template["stage"] == "PD"
    assert len(template["items"]) == 335


# ── run_checklist_review: контракт верхнего уровня ──────────────────────────────────────


def test_run_produces_checklist_review_v1_contract():
    template = _mini_template()
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
    )
    assert result["schema"] == "checklist_review_v1"
    assert result["dataset_id"] == "ds-pd"
    assert result["template"] == "mini_test_template"
    assert result["summary"]["total"] == len(template["items"])
    assert len(result["items"]) == len(template["items"])
    for item in result["items"]:
        assert "item_id" in item
        assert "status" in item
        assert "suggested_answer" in item
        assert "confidence" in item
        assert "document_evidence" in item
        assert "model_note" in item
        assert item["human_decision"] == "unset"


def test_run_defense_and_remarks_are_present_with_contract_fields():
    """T2.4: normalized_remarks/defense формируются содержательно (не пустые заглушки, как было
    в T2.1) — контракт implementation_plan.md §4 полностью заполнен. Мини-template содержит
    spatial_visual (manual_required) -> хотя бы одна remark ожидается даже без search/inventory
    хитов (правило _REMARK_STATUSES включает manual_required)."""
    template = _mini_template()
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
    )
    assert result["normalized_remarks"]
    assert all(r["schema"] == "normalized_remark_v1" for r in result["normalized_remarks"])
    assert "defense" in result
    assert result["defense"]["schema"] == "defense_contract_v1"
    assert "workflow_plan" in result


# ── pp87_composition: top-level поле контракта (T2.5) ──────────────────────────────────
#
# implementation_plan.md/промпт T2.5: 5 items дисциплины «Общее» в glorax_pd_2026.json — все
# про инженерные изыскания/СТУ, ни один не про состав разделов ПД (проверено grep по якорям
# "состав проектной документации"/"87"/"разделы ПД" — 0 совпадений). Решение — консервативное:
# composition-результат живёт в НОВОМ top-level поле контракта ``pp87_composition``, НЕ
# привязан к items. Поле вычисляется, только если вызывающий код передал ``pp87_config`` —
# без него run_checklist_review не тянет pp87_composition_service вовсе (чистая функция,
# обратная совместимость со всеми существующими вызовами/тестами).


def test_pp87_composition_absent_by_default_without_config():
    template = _mini_template()
    result = run_checklist_review(
        template, dataset_id="ds-pd",
        inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    assert result["pp87_composition"] is None


def test_pp87_composition_populated_when_config_passed():
    template = _mini_template()
    inventory = _FakeInventoryProvider({
        "ds-pd": [
            {"file_name": "Раздел ПД №1. Часть 1. ПЗ.pdf"},
            {"file_name": "Раздел ПД №3. Часть 1. АР1.pdf"},
        ],
    })
    pp87_config = {
        "meta": {"version": 1},
        "sections": [
            {
                "code": "1", "title": "ПЗ", "required": "always",
                "filename_patterns": [r"раздел\s*пд\s*№?\s*1\b"],
                "requirement_ref": "ПП РФ №87 от 16.02.2008, п.12, раздел 1",
            },
            {
                "code": "4", "title": "КР", "required": "always",
                "filename_patterns": [r"раздел\s*пд\s*№?\s*4\b"],
                "requirement_ref": "ПП РФ №87 от 16.02.2008, п.12, раздел 4",
            },
        ],
    }
    result = run_checklist_review(
        template, dataset_id="ds-pd",
        inventory_provider=inventory, search_provider=_empty_search,
        pp87_config=pp87_config,
    )
    composition = result["pp87_composition"]
    assert composition is not None
    assert composition["schema"] == "pp87_composition_v1"
    by_code = {s["code"]: s for s in composition["sections"]}
    assert by_code["1"]["status"] == "present"
    assert by_code["4"]["status"] == "missing"
    assert composition["summary"]["all_required_present"] is False
    # inventory_provider вызывается только для dataset_id прогона (не для source_dataset_ids) —
    # presence-items мини-шаблона тоже используют inventory_provider, поэтому проверяем состав
    # звонков, а не их точное количество.
    assert set(inventory.calls) == {"ds-pd"}


def test_pp87_composition_not_linked_to_items_stays_separate_field():
    """Явная проверка консервативного решения T2.5: даже когда pp87_config передан, items
    дисциплины «Общее» НЕ получают computed-evidence из composition (нет якорного маппинга) —
    их status/suggested_answer не меняются относительно прогона без pp87_config."""
    template = _mini_template()
    pp87_config = {
        "meta": {"version": 1},
        "sections": [
            {
                "code": "1", "title": "ПЗ", "required": "always",
                "filename_patterns": [r"раздел\s*пд\s*№?\s*1\b"],
                "requirement_ref": "ПП РФ №87 от 16.02.2008, п.12, раздел 1",
            },
        ],
    }
    baseline = run_checklist_review(
        template, dataset_id="ds-pd",
        inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    with_pp87 = run_checklist_review(
        template, dataset_id="ds-pd",
        inventory_provider=_empty_inventory, search_provider=_empty_search,
        pp87_config=pp87_config,
    )
    baseline_ob_items = [it for it in baseline["items"] if it["discipline"] == "Общее"]
    with_pp87_ob_items = [it for it in with_pp87["items"] if it["discipline"] == "Общее"]
    assert [it["status"] for it in baseline_ob_items] == [it["status"] for it in with_pp87_ob_items]
    assert [it["suggested_answer"] for it in baseline_ob_items] == [
        it["suggested_answer"] for it in with_pp87_ob_items
    ]


# ── presence: контентный хит → supported_by_evidence/yes ───────────────────────────────


def test_presence_content_hit_gives_supported_yes_with_evidence():
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:otchet_igi.pdf#page=1",
                "snippet": "Отчет об инженерно-геологических изысканиях выполнен в 2026 году.",
                "file_name": "otchet_igi.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-OB-003")
    assert item["status"] == "supported_by_evidence"
    assert item["suggested_answer"] == "yes"
    assert item["document_evidence"]
    assert item["document_evidence"][0]["source_ref"]
    assert item["document_evidence"][0]["snippet"]


def test_presence_yes_evidence_has_non_empty_file_name_bug2():
    """T2.6 (баг 2, SESSION_LOG Запись 18-19): в SMOKE_PD_O_FULL.json presence-yes items первый
    evidence имел ПУСТОЙ file_name при заполненном source_ref — _run_presence строила evidence-
    словарь без ключа file_name вообще, хотя default_search_provider отдаёт его в каждом hit
    (``{"source_ref", "snippet", "file_name"}``). Evidence обязан нести оба поля непустыми, когда
    search_provider их даёт — file_name нужен UI/оператору отдельно от source_ref (человеко-
    читаемое имя файла против внутреннего chunk-адреса)."""
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:otchet_igi.pdf#page=1",
                "snippet": "Отчет об инженерно-геологических изысканиях выполнен в 2026 году.",
                "file_name": "otchet_igi.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-OB-003")
    assert item["status"] == "supported_by_evidence"
    assert item["suggested_answer"] == "yes"
    ev = item["document_evidence"][0]
    assert ev["source_ref"] == "ds-pd:otchet_igi.pdf#page=1"
    assert ev["file_name"] == "otchet_igi.pdf", (
        "file_name не должен теряться при сборке presence-evidence из hit'а search_provider"
    )


# ── presence: filename-match без контента → review_needed/unknown (правило 1b) ─────────


def test_presence_filename_match_without_content_hit_is_review_needed():
    template = _mini_template()
    inventory = _FakeInventoryProvider({
        "ds-pd": [{"file_name": "Отчет_ИГИ_2026.pdf", "doc_type": "IGI_REPORT"}]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=inventory,
        search_provider=_empty_search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-OB-003")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] == "unknown"
    assert "имя совпало" in item["model_note"]
    assert "содержимое не подтверждено" in item["model_note"]


# ── presence: ничего не найдено → review_needed/unknown (не no) ────────────────────────


def test_presence_nothing_found_is_review_needed_not_no():
    template = _mini_template()
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-OB-004")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] == "unknown"
    assert item["suggested_answer"] != "no"


# ── spatial_visual → manual_required ────────────────────────────────────────────────────


def test_spatial_visual_is_manual_required():
    template = _mini_template()
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-049")
    assert item["status"] == "manual_required"
    assert item["suggested_answer"] == "manual_required"


# ── calculation/parametric/cross_section/spds_formal → review_needed, not_run (T2.1 заглушки) ──


@pytest.mark.parametrize(
    "item_id",
    ["PD-AR-013", "PD-SPOZU-010", "PD-AR-005"],
)
def test_unimplemented_kinds_are_review_needed_with_not_run(item_id):
    template = _mini_template()
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
    )
    item = next(it for it in result["items"] if it["item_id"] == item_id)
    assert item["status"] == "review_needed"
    assert item["computed_check"]["status"] == "not_run"


# ── фильтр discipline ────────────────────────────────────────────────────────────────────


def test_discipline_filter_only_touches_matching_items():
    template = _mini_template()
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        discipline="АР",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
    )
    ids = {it["item_id"] for it in result["items"]}
    assert ids == {"PD-AR-013", "PD-AR-049", "PD-AR-005"}
    assert result["summary"]["total"] == 3


def test_discipline_filter_absent_runs_all_items():
    template = _mini_template()
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
    )
    assert result["summary"]["total"] == len(template["items"])


# ── safety: yes/no всегда с source_ref ──────────────────────────────────────────────────


def test_safety_no_yes_no_suggested_answer_without_source_ref():
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:otchet_igi.pdf#page=1",
                "snippet": "Отчет об инженерно-геологических изысканиях выполнен в 2026 году.",
                "file_name": "otchet_igi.pdf",
            }
        ]
    })
    inventory = _FakeInventoryProvider({
        "ds-pd": [{"file_name": "Отчет_ИГДИ.pdf", "doc_type": "IGDI_REPORT"}]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=inventory,
        search_provider=search,
    )
    for item in result["items"]:
        if item["suggested_answer"] in ("yes", "no"):
            refs = [ev.get("source_ref") for ev in item["document_evidence"]]
            assert any(refs), (
                f"{item['item_id']}: suggested_answer={item['suggested_answer']} "
                "без непустого document_evidence[].source_ref"
            )


# ── summary: by_status/by_kind/suggested сходятся с items ──────────────────────────────


def test_summary_counts_match_items():
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:otchet_igi.pdf#page=1",
                "snippet": "Отчет об инженерно-геологических изысканиях выполнен в 2026 году.",
                "file_name": "otchet_igi.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=search,
    )
    items = result["items"]
    summary = result["summary"]

    by_status: dict[str, int] = {}
    by_kind: dict[str, int] = {}
    suggested: dict[str, int] = {}
    for it in items:
        by_status[it["status"]] = by_status.get(it["status"], 0) + 1
        by_kind[it["kind"]] = by_kind.get(it["kind"], 0) + 1
        suggested[it["suggested_answer"]] = suggested.get(it["suggested_answer"], 0) + 1

    assert summary["total"] == len(items)
    assert summary["by_status"] == by_status
    assert summary["by_kind"] == by_kind
    # suggested в контракте имеет фиксированные ключи yes/no/not_required/manual_required/unknown
    for key in ("yes", "no", "not_required", "manual_required", "unknown"):
        assert summary["suggested"].get(key, 0) == suggested.get(key, 0)


# ══════════════════════════════════════════════════════════════════════════════════════════
# T2.2 — calculation-механизм (openpyxl formula-count, 0 LLM)
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# Item PD-AR-013 (_mini_template) — kind=calculation, criterion="Теплотехнический расчет".
# workbook_provider(dataset_id) -> [{file_name, path}] — инъектируемый провайдер, как
# inventory_provider/search_provider. Ни один из этих тестов не трогает реальную MetaDB/диск
# датасета — только tmp_path синтетические книги.


class _FakeWorkbookProvider:
    def __init__(self, files: dict[str, list[dict]] | None = None):
        self._files = files or {}
        self.calls: list[str] = []

    def __call__(self, dataset_id: str) -> list[dict]:
        self.calls.append(dataset_id)
        return list(self._files.get(dataset_id, []))


def _empty_workbook(dataset_id: str) -> list[dict]:
    return []


def _make_xlsx_with_formulas(tmp_path, name: str = "Теплотехнический_расчет.xlsx") -> str:
    """Синтетическая книга с формульными ячейками на листе 'Расчет'. T2.6 (баг 1): заголовок в
    A1 называет тему расчёта дословно — реалистичное отражение боевых книг (тема расчёта видна
    в заголовке/шапке таблицы), нужное для прохождения новой контентной проверки книги
    (``_workbook_has_specific_content``, см. модульный докстринг ``_run_calculation``)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Расчет"
    ws["A1"] = "Теплотехнический расчёт наружной стены"
    ws["A2"] = "Параметр"
    ws["B2"] = "Значение"
    ws["A3"] = "Толщина утеплителя"
    ws["B3"] = 0.15
    ws["A4"] = "Сопротивление теплопередаче"
    ws["B4"] = "=B3/0.04"
    ws["A5"] = "Итого"
    ws["B5"] = "=SUM(B3:B4)"
    path = tmp_path / name
    wb.save(str(path))
    return str(path)


def _make_xlsx_without_formulas(tmp_path, name: str = "Теплотехнический_расчет_таблица.xlsx") -> str:
    """Синтетическая книга — только значения, ни одной формулы. Заголовок в A1 — см. докстринг
    ``_make_xlsx_with_formulas`` (T2.6, контентная проверка книги)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"
    ws["A1"] = "Теплотехнический расчёт наружной стены"
    ws["A2"] = "Параметр"
    ws["B2"] = "Значение"
    ws["A3"] = "Толщина утеплителя"
    ws["B3"] = 0.15
    ws["A4"] = "Сопротивление теплопередаче"
    ws["B4"] = 3.75
    path = tmp_path / name
    wb.save(str(path))
    return str(path)


# ── (c) xlsx-кандидат с формулами -> supported_by_evidence/yes + computed evidence ─────


def test_calculation_with_formulas_is_supported_by_evidence(tmp_path):
    path = _make_xlsx_with_formulas(tmp_path)
    template = _mini_template()
    workbook = _FakeWorkbookProvider({
        "ds-pd": [{"file_name": "Теплотехнический_расчет.xlsx", "path": path}],
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["status"] == "supported_by_evidence"
    assert item["suggested_answer"] == "yes"
    assert item["computed_check"]["name"] == "excel_formulas_present"
    assert item["computed_check"]["status"] == "ok"
    assert item["document_evidence"]
    ev = item["document_evidence"][0]
    assert ev["kind"] == "computed"
    assert ev["source_ref"]
    assert "Теплотехнический_расчет.xlsx" in ev["source_ref"]
    assert "sheet=" in ev["source_ref"]
    assert "formula cells" in ev["value"] or "formula" in ev["value"].lower()


# ── (d) xlsx-кандидат найден, формул нет -> computed_issue/no + evidence (safety-исключение) ──


def test_calculation_without_formulas_is_computed_issue_no_with_evidence(tmp_path):
    path = _make_xlsx_without_formulas(tmp_path)
    template = _mini_template()
    workbook = _FakeWorkbookProvider({
        "ds-pd": [{"file_name": "Теплотехнический_расчет_таблица.xlsx", "path": path}],
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["status"] == "computed_issue"
    assert item["suggested_answer"] == "no"
    assert item["computed_check"]["name"] == "excel_formulas_present"
    assert item["computed_check"]["status"] != "ok"
    # safety-исключение: no подкреплён computed evidence с непустым source_ref
    assert item["document_evidence"]
    ev = item["document_evidence"][0]
    assert ev["source_ref"]
    assert "Теплотехнический_расчет_таблица.xlsx" in ev["source_ref"]


# ── (e) xlsx не найден, но контентный хит расчёта в PDF -> review_needed + пометка ──────


def test_calculation_content_hit_in_non_excel_is_review_needed_with_note():
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:Теплотехническая_записка.pdf#page=3",
                "snippet": "Теплотехнический расчет наружной стены выполнен согласно СП 50.13330.",
                "file_name": "Теплотехническая_записка.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=search,
        workbook_provider=_empty_workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["status"] == "review_needed"
    note = item["model_note"] + " " + str(item.get("model_note", ""))
    assert "не в excel" in item["model_note"].lower() or "не в Excel".lower() in item["model_note"].lower()


# ── (f) совсем ничего не найдено -> review_needed ───────────────────────────────────────


def test_calculation_nothing_found_is_review_needed():
    template = _mini_template()
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=_empty_workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["status"] == "review_needed"


# ── (g) .xls legacy — честный legacy_unsupported, НЕ падает ─────────────────────────────


def test_calculation_legacy_xls_is_legacy_unsupported_review_needed(tmp_path):
    # .xls не может быть реально создан openpyxl (пишет только xlsx) — для теста достаточно
    # файла с расширением .xls на диске (importer обязан распознать РАСШИРЕНИЕ и не пытаться
    # парсить его openpyxl, что и так упало бы с BadZipFile/InvalidFileException).
    path = tmp_path / "Теплотехнический_расчет.xls"
    path.write_bytes(b"\xd0\xcf\x11\xe0not-a-real-ole-file")
    template = _mini_template()
    workbook = _FakeWorkbookProvider({
        "ds-pd": [{"file_name": "Теплотехнический_расчет.xls", "path": str(path)}],
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["status"] == "review_needed"
    assert "legacy_unsupported" in item["model_note"]


# ── anchor-based candidate selection: файл найден по якорным термам в имени файла ───────


def test_calculation_selects_candidate_by_anchor_terms_in_filename(tmp_path):
    """Кандидат отбирается по совпадению якорных термов критерия с именем файла — файл с
    нерелевантным именем среди прочих не должен ложно матчиться."""
    relevant_path = _make_xlsx_with_formulas(tmp_path, name="Теплотехнический_расчет_стен.xlsx")
    irrelevant_path = _make_xlsx_without_formulas(tmp_path, name="Смета_материалов.xlsx")
    template = _mini_template()
    workbook = _FakeWorkbookProvider({
        "ds-pd": [
            {"file_name": "Смета_материалов.xlsx", "path": irrelevant_path},
            {"file_name": "Теплотехнический_расчет_стен.xlsx", "path": relevant_path},
        ],
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["status"] == "supported_by_evidence"
    assert "Теплотехнический_расчет_стен.xlsx" in item["document_evidence"][0]["source_ref"]


# ══════════════════════════════════════════════════════════════════════════════════════════
# T2.6 (баг 1, SESSION_LOG Запись 18-19): жадный calculation-отбор кандидата
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# Боевой smoke (SMOKE_PD_O_FULL.json): 9 РАЗНЫХ критериев (звукоизоляция, электронагрузки, ТВО,
# влагоперенос, категории помещений, освещённость, теплотехническая однородность, воздухообмен)
# получили yes от ОДНОГО файла «РАЗДЕЛ 3/Экспликация помещений/Итоговая таблица_28.09.2022.xlsx».
# Root cause (подтверждено вручную): _workbook_candidates матчит якорные термы против ПОЛНОГО
# пути файла (workbook_provider отдаёт file_name = относительный путь, не basename) — термы вроде
# «помещений», «раздел», «таблица» (общие домен-слова критерия или структуры проекта: «РАЗДЕЛ 3»,
# «Экспликация помещений») случайно совпадают с именами ПАПОК в пути, не с темой самого расчёта.
# Фикс: (1) курируемый стоп-лист общих домен-термов (_CALC_DOMAIN_STOPWORDS) исключается из
# якорных термов ПЕРЕД матчингом имени файла — остаются только специфичные термы темы расчёта;
# (2) контентная проверка книги (_workbook_has_specific_content): имена листов/текст первых строк
# должны содержать хотя бы один специфичный терм — иначе кандидат отклоняется, даже если имя
# файла совпало.


def _make_xlsx_titled(tmp_path, name: str, sheet_title: str, header_a1: str,
                       with_formulas: bool = True) -> str:
    """Синтетическая книга с явным заголовком листа + текстом в A1 (контентная проверка книги
    читает именно эти два источника)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws["A1"] = header_a1
    ws["A2"] = "Параметр"
    ws["B2"] = "Значение"
    ws["A3"] = "Показатель"
    if with_formulas:
        ws["B3"] = "=1+1"
    else:
        ws["B3"] = 2
    path = tmp_path / name
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)


def test_calculation_bug1_generic_terms_do_not_match_unrelated_itogovaya_tablitsa(tmp_path):
    """RED-воспроизведение бага 1: книга «Итоговая таблица.xlsx» (лист «Лист1», лежит в пути
    «.../Экспликация помещений/...», реалистично отражающем боевой корпус) НЕ должна давать yes
    для критерия «Выполнен расчёт звукоизоляции...» — общие термы «расчет»/«выполнен»/«помещений»
    не являются специфичными для темы звукоизоляции, и лист/заголовок книги не содержат ничего
    про звукоизоляцию. ДО фикса (курируемый стоп-лист + контентная проверка) этот тест падал:
    _workbook_candidates матчил критерий на терм «помещений» (совпадает с именем папки в пути),
    _run_calculation слепо возвращал yes по формулам «Лист1» без проверки релевантности контента."""
    unrelated_path = _make_xlsx_titled(
        tmp_path,
        name="ПД_PDF/РАЗДЕЛ 3/Экспликация помещений/Итоговая таблица_28.09.2022.xlsx",
        sheet_title="Лист1",
        header_a1="Экспликация помещений",
    )
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-013":
            item["criterion"] = (
                "Выполнен расчёт звукоизоляции нормируемых помещений от источников шума: "
                "ИТП, насосная, трансформаторная, венткамеры и тд."
            )
    workbook = _FakeWorkbookProvider({
        "ds-pd": [
            {
                "file_name": "ПД_PDF/РАЗДЕЛ 3/Экспликация помещений/Итоговая таблица_28.09.2022.xlsx",
                "path": unrelated_path,
            },
        ],
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["status"] == "review_needed", (
        "баг 1: критерий про звукоизоляцию не должен матчить «Итоговую таблицу» экспликации "
        "помещений — нет специфичного терма ни в имени (после стоп-листа), ни в контенте книги"
    )
    assert item["suggested_answer"] == "unknown"


@pytest.mark.parametrize(
    "criterion",
    [
        "Расчет категории помещений",
        "Приведён расчёт одномерного влагопереноса для защиты от переувлажнения ограждающих "
        "конструкций (СП 50, раздел 8).",
        "Выполнен расчёт звукоизоляции нормируемых помещений от источников шума: ИТП, насосная, "
        "трансформаторная, венткамеры и тд.",
        "Выполнен расчет электрических нагрузок, представлена таблица электрических нагрузок в "
        "формате excel, приведена сводная таблица нагрузок по ЖК.",
        "Выполнен расчёт освещённости на нормируемых поверхностях помещений.",
        "Проведены расчёты коэффициента теплотехнической однородности r (по СП 230, раздел 5 и "
        "СП 50, приложение Г) для наружных стен.",
        "Приложена ТВО для коммерческих помещений по форме Заказчика (в формате Excel c "
        "сохранением расчетных зависимостей в ячейках).",
        "Выполнены расчёты воздухообмена на ассимиляцию вредностей для помещений автостоянок"
        "(при наличии).",
    ],
)
def test_calculation_bug1_all_nine_real_smoke_criteria_reject_itogovaya_tablitsa(tmp_path, criterion):
    """Все 9 реальных критериев боевого бага (дословно из SMOKE_PD_O_FULL.json) — против одной и
    той же нерелевантной «Итоговой таблицы» ни один не должен дать yes."""
    unrelated_path = _make_xlsx_titled(
        tmp_path,
        name="ПД_PDF/РАЗДЕЛ 3/Экспликация помещений/Итоговая таблица_28.09.2022.xlsx",
        sheet_title="Лист1",
        header_a1="Экспликация помещений",
    )
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-013":
            item["criterion"] = criterion
    workbook = _FakeWorkbookProvider({
        "ds-pd": [
            {
                "file_name": "ПД_PDF/РАЗДЕЛ 3/Экспликация помещений/Итоговая таблица_28.09.2022.xlsx",
                "path": unrelated_path,
            },
        ],
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["suggested_answer"] != "yes", (
        f"критерий {criterion!r} не должен матчить нерелевантную «Итоговую таблицу»"
    )


def test_calculation_bug1_positive_case_electrical_loads_book_with_matching_sheet(tmp_path):
    """Позитивный кейс из промпта: книга «Расчет электрических нагрузок.xlsx» с листом
    «Нагрузки» ДОЛЖНА давать yes для критерия про расчёт электрических нагрузок — специфичные
    термы («электрических», «нагрузок») совпадают и с именем файла, и с содержимым (лист
    «Нагрузки» + заголовок)."""
    relevant_path = _make_xlsx_titled(
        tmp_path,
        name="Расчет электрических нагрузок.xlsx",
        sheet_title="Нагрузки",
        header_a1="Расчёт электрических нагрузок по ЖК",
    )
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-013":
            item["criterion"] = (
                "Выполнен расчет электрических нагрузок, представлена таблица электрических "
                "нагрузок в формате excel, приведена сводная таблица нагрузок по ЖК."
            )
    workbook = _FakeWorkbookProvider({
        "ds-pd": [{"file_name": "Расчет электрических нагрузок.xlsx", "path": relevant_path}],
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["status"] == "supported_by_evidence"
    assert item["suggested_answer"] == "yes"
    assert "Расчет электрических нагрузок.xlsx" in item["document_evidence"][0]["source_ref"]


def test_calculation_bug1_candidate_by_name_rejected_without_matching_content(tmp_path):
    """Контентная проверка — самостоятельный гейт: даже если имя файла совпало по специфичному
    терму, отсутствие этого терма в листах/заголовке книги должно отклонить кандидата (не yes)."""
    misleading_path = _make_xlsx_titled(
        tmp_path,
        name="Расчет электрических нагрузок.xlsx",
        sheet_title="Общие данные",
        header_a1="Пояснительная записка проекта",
    )
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-013":
            item["criterion"] = (
                "Выполнен расчет электрических нагрузок, представлена таблица электрических "
                "нагрузок в формате excel."
            )
    workbook = _FakeWorkbookProvider({
        "ds-pd": [{"file_name": "Расчет электрических нагрузок.xlsx", "path": misleading_path}],
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["suggested_answer"] != "yes", (
        "имя файла совпало по слову «электрических»/«нагрузок», но ни лист, ни заголовок книги "
        "не подтверждают тему — контентная проверка должна отклонить кандидата"
    )


def test_calculation_bug1_no_candidates_after_stoplist_falls_back_to_content_hit_branch(tmp_path):
    """Если после ужесточения (стоп-лист + контентная проверка) кандидатов по имени нет вовсе, но
    lexical-поиск (search_provider) даёт контентный хит расчёта в PDF — работает существующая
    ветка (e) «расчёт найден не в Excel» (review_needed), а не молчаливый провал."""
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-013":
            item["criterion"] = "Выполнен расчёт звукоизоляции нормируемых помещений."
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:zapiska.pdf#page=5",
                "snippet": "Расчёт звукоизоляции нормируемых помещений выполнен согласно СП 51.",
                "file_name": "zapiska.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=search,
        workbook_provider=_empty_workbook,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert item["status"] == "review_needed"
    assert "не в excel" in item["model_note"].lower() or "не в Excel".lower() in item["model_note"].lower()


# ── safety: calculation (d) computed_issue/no проходит общий safety-инвариант ───────────


def test_safety_invariant_allows_no_with_computed_evidence(tmp_path):
    """Общий safety-тест (yes/no без source_ref запрещён) должен пропускать computed_issue/no —
    это допустимое исключение (evidence есть, просто computed, не document)."""
    path = _make_xlsx_without_formulas(tmp_path)
    template = _mini_template()
    workbook = _FakeWorkbookProvider({
        "ds-pd": [{"file_name": "Теплотехнический_расчет_таблица.xlsx", "path": path}],
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
        workbook_provider=workbook,
    )
    for item in result["items"]:
        if item["suggested_answer"] in ("yes", "no"):
            refs = [ev.get("source_ref") for ev in item["document_evidence"]]
            assert any(refs), (
                f"{item['item_id']}: suggested_answer={item['suggested_answer']} "
                "без непустого document_evidence[].source_ref"
            )
    # явно убеждаемся, что именно PD-AR-013 дал no+evidence (не пропущен молча)
    calc_item = next(it for it in result["items"] if it["item_id"] == "PD-AR-013")
    assert calc_item["suggested_answer"] == "no"
    assert calc_item["document_evidence"][0]["source_ref"]


# ══════════════════════════════════════════════════════════════════════════════════════════
# T2.2 — default-провайдеры (inventory/search/workbook) над временной SQLite MetaDB
# ══════════════════════════════════════════════════════════════════════════════════════════


def _make_meta_db(tmp_path, rows: list[dict]) -> str:
    import sqlite3

    db_path = str(tmp_path / "les_meta_test.db")
    conn = sqlite3.connect(db_path)
    conn.execute(
        """
        CREATE TABLE documents (
            id          TEXT PRIMARY KEY,
            dataset_id  TEXT,
            file_name   TEXT,
            status      TEXT,
            file_hash   TEXT,
            file_mtime  REAL,
            file_size   INTEGER,
            chunk_count INTEGER DEFAULT 0,
            source_path TEXT DEFAULT ''
        )
        """
    )
    for i, row in enumerate(rows):
        conn.execute(
            "INSERT INTO documents (id, dataset_id, file_name, status, source_path) "
            "VALUES (?, ?, ?, ?, ?)",
            (
                f"doc-{i}",
                row["dataset_id"],
                row["file_name"],
                row.get("status", "DONE"),
                row.get("source_path", ""),
            ),
        )
    conn.commit()
    conn.close()
    return db_path


def test_default_workbook_provider_filters_by_extension_from_metadb(tmp_path, monkeypatch):
    from proxy.services.checklist_review_service import default_workbook_provider

    db_path = _make_meta_db(tmp_path, [
        {"dataset_id": "ds-1", "file_name": "raschet.xlsx", "source_path": str(tmp_path / "raschet.xlsx")},
        {"dataset_id": "ds-1", "file_name": "staryj_raschet.xls", "source_path": str(tmp_path / "staryj_raschet.xls")},
        {"dataset_id": "ds-1", "file_name": "makro_raschet.xlsm", "source_path": str(tmp_path / "makro_raschet.xlsm")},
        {"dataset_id": "ds-1", "file_name": "poyasnitelnaya.pdf", "source_path": str(tmp_path / "poyasnitelnaya.pdf")},
        {"dataset_id": "ds-2", "file_name": "chужой.xlsx", "source_path": str(tmp_path / "chужой.xlsx")},
    ])
    monkeypatch.setenv("RAG_META_DB_PATH", db_path)

    files = default_workbook_provider("ds-1")
    names = {f["file_name"] for f in files}
    assert names == {"raschet.xlsx", "staryj_raschet.xls", "makro_raschet.xlsm"}
    assert all(f.get("path") for f in files)


def test_default_inventory_provider_reads_metadb(tmp_path, monkeypatch):
    from proxy.services.checklist_review_service import default_inventory_provider

    db_path = _make_meta_db(tmp_path, [
        {"dataset_id": "ds-1", "file_name": "otchet_igi.pdf"},
        {"dataset_id": "ds-1", "file_name": "raschet.xlsx"},
        {"dataset_id": "ds-2", "file_name": "chужой.pdf"},
    ])
    monkeypatch.setenv("RAG_META_DB_PATH", db_path)

    files = default_inventory_provider("ds-1")
    names = {f["file_name"] for f in files}
    assert names == {"otchet_igi.pdf", "raschet.xlsx"}


# ══════════════════════════════════════════════════════════════════════════════════════════
# T3.1 — parametric-механизм (glorax_param_rules.yaml + checklist_param_rules, 0 LLM)
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# Item PD-KR-020 (_mini_template) — kind=parametric, criterion "Марка бетона... (W12)" —
# правило concrete_water_resistance_rostverk_w12 в боевом glorax_param_rules.yaml.
# search_provider даёт хиты -> extract_value по правилу item'а -> compare кодом:
#   значение найдено и ok    -> supported_by_evidence/yes + evidence(snippet) + computed_check ok
#   значение найдено и issue -> computed_issue/no + evidence с source_ref (safety-исключение)
#   конфликт значений        -> computed_issue + ОБА evidence + model_note про конфликт
#   значение не найдено      -> review_needed (НЕ no)
#   item kind=parametric БЕЗ правила в реестре -> review_needed + model_note "нет правила"


def test_parametric_value_found_ok_gives_supported_by_evidence_yes():
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:spec_kr.pdf#page=4",
                "snippet": "Принят бетон класса В30 W12 для ростверка согласно спецификации КР.",
                "file_name": "spec_kr.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-KR-020")
    assert item["status"] == "supported_by_evidence"
    assert item["suggested_answer"] == "yes"
    assert item["computed_check"]["name"] == "concrete_water_resistance_rostverk_w12"
    assert item["computed_check"]["status"] == "ok"
    assert item["document_evidence"]
    ev = item["document_evidence"][0]
    assert ev["source_ref"]
    assert "W12" in ev["snippet"]


def test_parametric_value_found_issue_gives_computed_issue_no_with_source_ref():
    """Safety-исключение как в calculation: computed_issue/no ОБЯЗАН иметь непустой source_ref —
    W8 вместо требуемого W12 — реальное расхождение, не отсутствие evidence."""
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:spec_kr.pdf#page=4",
                "snippet": "Принят бетон класса В30 W8 для ростверка согласно спецификации КР.",
                "file_name": "spec_kr.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-KR-020")
    assert item["status"] == "computed_issue"
    assert item["suggested_answer"] == "no"
    assert item["computed_check"]["status"] == "issue"
    assert item["document_evidence"]
    ev = item["document_evidence"][0]
    assert ev["source_ref"]
    assert "W8" in ev["snippet"]


def test_parametric_conflicting_values_gives_computed_issue_with_both_evidence():
    """Два хита с разными значениями (W12 в одном источнике, W8 в другом) -> computed_issue,
    ОБА evidence присутствуют, model_note явно упоминает конфликт."""
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:spec_kr.pdf#page=4",
                "snippet": "Принят бетон класса В30 W12 для ростверка согласно спецификации КР.",
                "file_name": "spec_kr.pdf",
            },
            {
                "source_ref": "ds-pd:poyasnitelnaya.pdf#page=9",
                "snippet": "Марка бетона ростверка W8 указана в пояснительной записке.",
                "file_name": "poyasnitelnaya.pdf",
            },
        ]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-KR-020")
    assert item["status"] == "computed_issue"
    assert len(item["document_evidence"]) >= 2
    refs = {ev["source_ref"] for ev in item["document_evidence"]}
    assert "ds-pd:spec_kr.pdf#page=4" in refs
    assert "ds-pd:poyasnitelnaya.pdf#page=9" in refs
    assert "конфликт" in item["model_note"].lower()


def test_parametric_value_not_found_is_review_needed_not_no():
    template = _mini_template()
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-KR-020")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] != "no"


def test_parametric_item_without_registered_rule_is_review_needed_with_note():
    """PD-KR-020 переименован в id без носителя в реестре — item kind=parametric, но
    load_param_rules() не находит правило -> честный review_needed, НЕ падение/manual_required."""
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-KR-020":
            item["id"] = "PD-KR-999-NO-RULE"
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=_empty_search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-KR-999-NO-RULE")
    assert item["status"] == "review_needed"
    assert "нет параметрического правила" in item["model_note"].lower()


def test_parametric_screed_thickness_rule_positive_and_negative():
    """Независимая проверка второго обязательного правила (PD-AR-047, стяжка >=80мм) через
    полный run_checklist_review — не только unit-тест checklist_param_rules."""
    template = _mini_template()
    template["items"].append({
        "id": "PD-AR-047",
        "stage": "PD",
        "discipline": "АР",
        "sheet_name": "АР",
        "row": 47,
        "item_no": "45",
        "section_path": ["ИСХОДНАЯ ДОКУМЕНТАЦИЯ", "Подземный паркинг"],
        "criterion": "Толщина стяжки пола — не менее 80 мм.",
        "answer_cell": "C47",
        "note_cell": "D47",
        "allowed_answers": ["Да", "Нет", "Не требуется"],
        "kind": "parametric",
    })
    search_ok = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:ar_pояснительная.pdf#page=12",
                "snippet": "Толщина стяжки пола в паркинге выполнена 100 мм по проекту.",
                "file_name": "ar_poyasnitelnaya.pdf",
            }
        ]
    })
    result_ok = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=search_ok,
    )
    item_ok = next(it for it in result_ok["items"] if it["item_id"] == "PD-AR-047")
    assert item_ok["status"] == "supported_by_evidence"
    assert item_ok["suggested_answer"] == "yes"

    search_bad = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:ar_pояснительная.pdf#page=12",
                "snippet": "Толщина стяжки пола в паркинге выполнена 50 мм по проекту.",
                "file_name": "ar_poyasnitelnaya.pdf",
            }
        ]
    })
    result_bad = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=search_bad,
    )
    item_bad = next(it for it in result_bad["items"] if it["item_id"] == "PD-AR-047")
    assert item_bad["status"] == "computed_issue"
    assert item_bad["suggested_answer"] == "no"


def test_parametric_safety_invariant_no_status_has_source_ref():
    """Общий safety-инвариант (yes/no без source_ref запрещён) должен выполняться и для
    parametric-computed_issue/no."""
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:spec_kr.pdf#page=4",
                "snippet": "Принят бетон класса В30 W8 для ростверка согласно спецификации КР.",
                "file_name": "spec_kr.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template,
        dataset_id="ds-pd",
        inventory_provider=_empty_inventory,
        search_provider=search,
    )
    for item in result["items"]:
        if item["suggested_answer"] in ("yes", "no"):
            refs = [ev.get("source_ref") for ev in item["document_evidence"]]
            assert any(refs), (
                f"{item['item_id']}: suggested_answer={item['suggested_answer']} "
                "без непустого document_evidence[].source_ref"
            )


# ══════════════════════════════════════════════════════════════════════════════════════════
# T3.2 (A) — параметрические правила поверх ЛЮБОГО kind (gap T3.1, диспетчер _run_item)
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# PD-EOM-032 (kind=manual_required в боевом glorax_pd_2026.json, реестр glorax_param_rules.yaml
# содержит rule cable_fire_protection_parking_ei150 для этого item_id) — реальный носитель из
# T3.1 gap: 7 из 9 правил висят на manual_required items, диспетчер их раньше не запускал.
# Используем боевой template (load_checklist_template("glorax_pd_2026")), не мини-фикстуру —
# правило должно найтись именно по item_id из реального реестра.


def test_composite_item_parametric_rule_ok_gives_supported_with_manual_required_suggested():
    """Составной пункт (kind=manual_required) с найденным параметрическим правилом, значение ok:
    status=supported_by_evidence (параметрическая часть подтверждена кодом), НО
    suggested_answer остаётся manual_required (не yes) — остальная часть критерия (расчёт/
    методика/полнота) не проверена кодом, финал за инженером."""
    template = load_checklist_template("glorax_pd_2026")
    item_src = next(it for it in template["items"] if it["id"] == "PD-EOM-032")
    assert item_src["kind"] == "manual_required"  # предпосылка теста, не подгонка

    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:eom_spec.pdf#page=7",
                "snippet": "Транзитные кабельные линии через паркинг защищены до предела EI150.",
                "file_name": "eom_spec.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=search,
        discipline="ЭОМ",
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-EOM-032")
    assert item["status"] == "supported_by_evidence"
    assert item["suggested_answer"] == "manual_required"
    assert item["computed_check"]["name"] == "cable_fire_protection_parking_ei150"
    assert item["computed_check"]["status"] == "ok"
    assert item["document_evidence"]
    assert item["document_evidence"][0]["source_ref"]
    note = item["model_note"].lower()
    assert "параметрическ" in note
    assert "cable_fire_protection_parking_ei150".lower() in note
    assert "инженер" in note


def test_composite_item_parametric_rule_issue_gives_computed_issue_with_manual_required_suggested():
    """Составной пункт, значение issue (EI90 вместо EI150): status=computed_issue,
    suggested_answer ОСТАЁТСЯ manual_required (НЕ no, промпт явно требует не no для составных),
    evidence присутствует, model_note называет доказанное нарушение параметра."""
    template = load_checklist_template("glorax_pd_2026")
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:eom_spec.pdf#page=7",
                "snippet": "Транзитные кабельные линии через паркинг защищены до предела EI90.",
                "file_name": "eom_spec.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=search,
        discipline="ЭОМ",
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-EOM-032")
    assert item["status"] == "computed_issue"
    assert item["suggested_answer"] == "manual_required"
    assert item["suggested_answer"] != "no"
    assert item["computed_check"]["status"] == "issue"
    assert item["document_evidence"]
    ev = item["document_evidence"][0]
    assert ev["source_ref"]
    assert "EI90" in ev["snippet"]
    note = item["model_note"].lower()
    assert "наруш" in note or "issue" in note or "не соответств" in note


def test_composite_item_parametric_rule_not_found_falls_back_to_kind_manual_required():
    """Значение параметра не найдено ни в одном хите (правило есть, но evidence нет): составной
    пункт идёт по ПРЕЖНЕМУ пути kind (manual_required) — не выдумываем computed-статус без evidence."""
    template = load_checklist_template("glorax_pd_2026")
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
        discipline="ЭОМ",
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-EOM-032")
    assert item["status"] == "manual_required"
    assert item["suggested_answer"] == "manual_required"


def test_kind_parametric_still_behaves_as_before_after_dispatcher_change(tmp_path):
    """Регресс T3.1: kind=parametric items продолжают вести себя как раньше (ok->supported/yes,
    issue->computed_issue/no) — диспетчер-изменение не должно затронуть чистый parametric-путь."""
    template = _mini_template()
    search_ok = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:spec_kr.pdf#page=4",
                "snippet": "Принят бетон класса В30 W12 для ростверка согласно спецификации КР.",
                "file_name": "spec_kr.pdf",
            }
        ]
    })
    result_ok = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=search_ok,
    )
    item_ok = next(it for it in result_ok["items"] if it["item_id"] == "PD-KR-020")
    assert item_ok["status"] == "supported_by_evidence"
    assert item_ok["suggested_answer"] == "yes"

    search_bad = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:spec_kr.pdf#page=4",
                "snippet": "Принят бетон класса В30 W8 для ростверка согласно спецификации КР.",
                "file_name": "spec_kr.pdf",
            }
        ]
    })
    result_bad = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=search_bad,
    )
    item_bad = next(it for it in result_bad["items"] if it["item_id"] == "PD-KR-020")
    assert item_bad["status"] == "computed_issue"
    assert item_bad["suggested_answer"] == "no"  # НЕ manual_required — kind сам по себе parametric


# ══════════════════════════════════════════════════════════════════════════════════════════
# T3.2 (B) — two-sided gate для cross_section (project_doc + source_doc evidence)
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# PD-SPOZU-010 (_mini_template) — kind=cross_section, criterion "Соответствие пирогов
# благоустройства ТЗ и техническому стандарту Glorax". search_provider вызывается и для
# dataset_id (project-хит), и для каждого source_dataset_ids (source-хит) — фейковый provider
# различает датасеты по dataset_id-ключу в hits_by_terms_key (как _FakeSearchProvider выше,
# просто добавляем записи под разными ключами dataset_id: "ds-pd" и "ds-tz").


def _cross_section_search(project_hit: bool, source_hit: bool) -> _FakeSearchProvider:
    # Snippet'ы намеренно повторяют >=2 якорных терма критерия («пирогов»/«пироги»,
    # «благоустройства», «техническому», «стандарту») — _content_hit требует контентного
    # покрытия, одного случайного слова недостаточно (анти-галлюцинация, см. _run_presence).
    hits: dict[str, list[dict]] = {}
    if project_hit:
        hits["ds-pd"] = [
            {
                "source_ref": "ds-pd:spozu.pdf#page=2",
                "snippet": "Пироги благоустройства СПОЗУ соответствуют техническому стандарту Glorax.",
                "file_name": "spozu.pdf",
            }
        ]
    if source_hit:
        hits["ds-tz"] = [
            {
                "source_ref": "ds-tz:tz.pdf#page=5",
                "snippet": "ТЗ: пироги благоустройства должны соответствовать техническому стандарту Glorax.",
                "file_name": "tz.pdf",
            }
        ]
    return _FakeSearchProvider(hits)


def test_cross_section_both_sides_found_gives_review_needed_unknown_with_both_evidence():
    """Project-хит И source-хит есть -> review_needed/unknown (LLM-связывание — Phase 5, здесь НЕ
    зовём), оба evidence приложены с разными kind (project_doc/source_doc), model_note фиксирует,
    что обе стороны найдены и сверка за инженером/моделью."""
    template = _mini_template()
    search = _cross_section_search(project_hit=True, source_hit=True)
    result = run_checklist_review(
        template, dataset_id="ds-pd", source_dataset_ids=["ds-tz"],
        inventory_provider=_empty_inventory, search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-SPOZU-010")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] == "unknown"
    kinds = {ev["kind"] for ev in item["document_evidence"]}
    assert kinds == {"project_doc", "source_doc"}
    refs = {ev["source_ref"] for ev in item["document_evidence"]}
    assert "ds-pd:spozu.pdf#page=2" in refs
    assert "ds-tz:tz.pdf#page=5" in refs
    note = item["model_note"].lower()
    assert "обе стороны" in note or "обе стороны найдены" in note


def test_cross_section_only_project_side_found_review_needed_notes_missing_source():
    """Только project-хит есть, source-хита нет -> review_needed, model_note явно называет,
    какая сторона отсутствует (source/исходники)."""
    template = _mini_template()
    search = _cross_section_search(project_hit=True, source_hit=False)
    result = run_checklist_review(
        template, dataset_id="ds-pd", source_dataset_ids=["ds-tz"],
        inventory_provider=_empty_inventory, search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-SPOZU-010")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] == "unknown"
    note = item["model_note"].lower()
    assert "source" in note or "исходник" in note or "тз" in note or "опр" in note or "аго" in note or "сту" in note


def test_cross_section_only_source_side_found_review_needed_notes_missing_project():
    """Только source-хит есть, project-хита нет -> review_needed, model_note называет, что
    отсутствует именно проектная (раздел ПД) сторона."""
    template = _mini_template()
    search = _cross_section_search(project_hit=False, source_hit=True)
    result = run_checklist_review(
        template, dataset_id="ds-pd", source_dataset_ids=["ds-tz"],
        inventory_provider=_empty_inventory, search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-SPOZU-010")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] == "unknown"
    note = item["model_note"].lower()
    assert "проект" in note or "раздел пд" in note or "project" in note


def test_cross_section_no_source_dataset_ids_review_needed_with_blocker():
    """source_dataset_ids не задан (None/пусто) -> review_needed, model_note явно называет
    отсутствие выбранных исходников (ТЗ/ОПР/АГО/СТУ), в workflow_plan появляется blocker."""
    template = _mini_template()
    search = _cross_section_search(project_hit=True, source_hit=False)
    result = run_checklist_review(
        template, dataset_id="ds-pd",
        inventory_provider=_empty_inventory, search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-SPOZU-010")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] == "unknown"
    note = item["model_note"].lower()
    assert "исходник" in note
    assert "тз" in note or "опр" in note or "аго" in note or "сту" in note
    assert result["workflow_plan"]["blockers"], "ожидался blocker в workflow_plan"


def test_cross_section_no_source_dataset_ids_empty_list_also_blocks():
    """source_dataset_ids=[] (пустой список, не None) — тот же путь, что None."""
    template = _mini_template()
    result = run_checklist_review(
        template, dataset_id="ds-pd", source_dataset_ids=[],
        inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-SPOZU-010")
    assert item["status"] == "review_needed"
    assert result["workflow_plan"]["blockers"]


def test_cross_section_nothing_found_is_review_needed():
    """Ни project-, ни source-хита нет (но source_dataset_ids задан) -> review_needed (общее
    "ничего не найдено", не путать с веткой "исходники не выбраны")."""
    template = _mini_template()
    result = run_checklist_review(
        template, dataset_id="ds-pd", source_dataset_ids=["ds-tz"],
        inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-SPOZU-010")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] == "unknown"


def test_cross_section_search_provider_called_for_project_and_each_source_dataset():
    """search_provider вызывается и для dataset_id, и для КАЖДОГО source_dataset_id — проверяем
    фактические вызовы фейкового провайдера, не только результат."""
    template = _mini_template()
    search = _FakeSearchProvider({})
    run_checklist_review(
        template, dataset_id="ds-pd", source_dataset_ids=["ds-tz", "ds-ago"],
        inventory_provider=_empty_inventory, search_provider=search,
    )
    called_datasets = {ds for ds, _terms in search.calls}
    assert "ds-pd" in called_datasets
    assert "ds-tz" in called_datasets
    assert "ds-ago" in called_datasets


def test_cross_section_suggested_answer_yes_is_never_produced_safety():
    """Safety-тест T3.2: suggested_answer=yes для cross_section ЗАПРЕЩЁН всегда в этой фазе — нет
    механизма сверки СОДЕРЖИМОГО обеих сторон (LLM-связывание — Phase 5), даже когда обе стороны
    найдены и выглядят согласованными по snippet'ам. Честно review_needed/unknown, не yes."""
    template = _mini_template()
    search = _cross_section_search(project_hit=True, source_hit=True)
    result = run_checklist_review(
        template, dataset_id="ds-pd", source_dataset_ids=["ds-tz"],
        inventory_provider=_empty_inventory, search_provider=search,
    )
    for item in result["items"]:
        if item["kind"] == "cross_section":
            assert item["suggested_answer"] != "yes", (
                f"{item['item_id']}: cross_section suggested_answer=yes запрещён в T3.2"
            )


# ══════════════════════════════════════════════════════════════════════════════════════════
# T2.4 (A) — spds_formal через переиспользование doc_review (doc_review_provider)
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# PD-AR-005 (_mini_template) — kind=spds_formal, criterion "Общие данные комплекта приложены".
# doc_review_provider(dataset_id) -> dict|None — результат review_dataset в формате
# {"items": [{"rule_id", "status", "severity", "source_ref"(опц.)}]}, как реально отдаёт
# review_to_json()["items"] (doc_review_service.ReviewItem.to_dict(), см. review_defense_pack/
# review_to_normalized_remarks в doc_review_service.py — оттуда взят формат status/severity/
# document_evidence[].source_ref). Курируемый словарь якорных термов -> rule_ids задан в коде
# checklist_review_service._SPDS_ANCHOR_TO_RULE_IDS (реальные id из
# config/normcontrol/gost_r_21_101_2026.yaml + NK-коды normcontrol_service как алиасы того же
# семейства проверок).


def _doc_review_item(rule_id: str, status: str, severity: str = "warning", source_ref: str = "") -> dict:
    return {
        "rule_id": rule_id,
        "status": status,
        "severity": severity,
        "document_evidence": ([{"source_ref": source_ref, "snippet": ""}] if source_ref else []),
    }


def test_spds_formal_all_mapped_rules_supported_gives_supported_yes_with_evidence():
    """Все doc_review-правила, замапленные на якоря критерия («ведомость» -> D2-001..003/NK-04),
    имеют status=supported_by_evidence -> spds_formal item тоже supported_by_evidence/yes, с
    evidence (kind=computed, source_ref=rule_id) и заполненным doc_review_item_ids."""
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-005":
            item["criterion"] = "Ведомость общих данных комплекта приложена"

    def doc_review_provider(dataset_id):
        return {"items": [
            _doc_review_item("G21.101-2026-D2-001", "supported_by_evidence", source_ref="ds-pd:vedomost.pdf#page=1"),
            _doc_review_item("G21.101-2026-D2-002", "supported_by_evidence", source_ref="ds-pd:vedomost.pdf#page=1"),
            _doc_review_item("G21.101-2026-D2-003", "supported_by_evidence", source_ref="ds-pd:vedomost.pdf#page=1"),
        ]}

    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
        doc_review_provider=doc_review_provider,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-005")
    assert item["status"] == "supported_by_evidence"
    assert item["suggested_answer"] == "yes"
    assert item["document_evidence"]
    for ev in item["document_evidence"]:
        assert ev["source_ref"]
    assert set(item["doc_review_item_ids"]) == {
        "G21.101-2026-D2-001", "G21.101-2026-D2-002", "G21.101-2026-D2-003",
    }


def test_spds_formal_computed_issue_with_source_ref_gives_computed_issue_no():
    """Есть замапленный doc_review item со статусом computed_issue И непустым source_ref ->
    spds_formal item computed_issue/no (safety-исключение как в calculation/parametric)."""
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-005":
            item["criterion"] = "Ведомость общих данных комплекта приложена"

    def doc_review_provider(dataset_id):
        return {"items": [
            _doc_review_item("G21.101-2026-D2-002", "computed_issue", severity="error",
                              source_ref="ds-pd:vedomost.pdf#page=1"),
        ]}

    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
        doc_review_provider=doc_review_provider,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-005")
    assert item["status"] == "computed_issue"
    assert item["suggested_answer"] == "no"
    assert item["document_evidence"]
    assert item["document_evidence"][0]["source_ref"]
    assert "G21.101-2026-D2-002" in item["doc_review_item_ids"]


def test_spds_formal_computed_issue_without_source_ref_gives_manual_required():
    """computed_issue есть, но у doc_review item НЕТ source_ref -> safety не позволяет suggested_answer=no
    без source_ref, поэтому manual_required (НЕ no без evidence)."""
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-005":
            item["criterion"] = "Ведомость общих данных комплекта приложена"

    def doc_review_provider(dataset_id):
        return {"items": [
            _doc_review_item("G21.101-2026-D2-002", "computed_issue", severity="error", source_ref=""),
        ]}

    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
        doc_review_provider=doc_review_provider,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-005")
    assert item["status"] == "manual_required"
    assert item["suggested_answer"] == "manual_required"
    assert "G21.101-2026-D2-002" in item["doc_review_item_ids"]


def test_spds_formal_provider_none_gives_review_needed_with_model_note():
    """doc_review_provider отсутствует (None-параметр по умолчанию) -> review_needed + model_note
    объясняет, что doc_review недоступен."""
    template = _mini_template()
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-005")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] == "unknown"
    assert item["model_note"]


def test_spds_formal_provider_returns_none_gives_review_needed():
    """doc_review_provider вызывается, но возвращает None (doc_review недоступен для датасета,
    например нет документов) -> review_needed, НЕ падение."""
    template = _mini_template()

    def doc_review_provider(dataset_id):
        return None

    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
        doc_review_provider=doc_review_provider,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-005")
    assert item["status"] == "review_needed"
    assert item["suggested_answer"] == "unknown"


def test_spds_formal_no_anchor_mapping_gives_review_needed_with_model_note():
    """Критерий не матчит ни один якорный терм из курируемого словаря -> review_needed +
    model_note объясняет, что маппинг пуст (не выдумываем doc_review_item_ids)."""
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-005":
            item["criterion"] = "Нечто совершенно не связанное с якорями словаря спдс проверки"

    def doc_review_provider(dataset_id):
        return {"items": [_doc_review_item("G21.101-2026-D2-001", "supported_by_evidence", source_ref="x")]}

    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
        doc_review_provider=doc_review_provider,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-005")
    assert item["status"] == "review_needed"
    assert item["doc_review_item_ids"] == []


def test_spds_formal_anchor_cipher_maps_to_expected_rule_ids():
    """Якорь «шифр/обозначение» матчит D1-011/D3-001..003/NK-03 (курируемый словарь промпта)."""
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-005":
            item["criterion"] = "Шифр и обозначение документа соответствуют требованиям"

    def doc_review_provider(dataset_id):
        return {"items": [
            _doc_review_item("G21.101-2026-D1-011", "supported_by_evidence", source_ref="ds-pd:x.pdf#page=1"),
            _doc_review_item("G21.101-2026-D3-001", "supported_by_evidence", source_ref="ds-pd:x.pdf#page=1"),
            _doc_review_item("G21.101-2026-D3-002", "supported_by_evidence", source_ref="ds-pd:x.pdf#page=1"),
            _doc_review_item("G21.101-2026-D3-003", "supported_by_evidence", source_ref="ds-pd:x.pdf#page=1"),
        ]}

    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
        doc_review_provider=doc_review_provider,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-005")
    assert item["status"] == "supported_by_evidence"
    assert item["suggested_answer"] == "yes"
    assert set(item["doc_review_item_ids"]) >= {
        "G21.101-2026-D1-011", "G21.101-2026-D3-001", "G21.101-2026-D3-002", "G21.101-2026-D3-003",
    }


def test_spds_formal_partial_mapping_missing_rule_in_provider_is_review_needed():
    """Словарь матчит несколько rule_ids, но doc_review_provider вернул items только для части
    из них (остальные не найдены в результате doc_review) -> недостаточно данных для supported
    (не все замапленные правила проверены) -> review_needed, НЕ выдуманный yes."""
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-005":
            item["criterion"] = "Ведомость общих данных комплекта приложена"

    def doc_review_provider(dataset_id):
        return {"items": [
            _doc_review_item("G21.101-2026-D2-001", "supported_by_evidence", source_ref="ds-pd:v.pdf#page=1"),
            # D2-002/D2-003 отсутствуют в результате doc_review вовсе
        ]}

    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
        doc_review_provider=doc_review_provider,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-AR-005")
    assert item["status"] == "review_needed"


def test_spds_formal_safety_invariant_no_yes_no_without_source_ref():
    """Общий safety-инвариант держится и для spds_formal: yes/no без source_ref запрещён."""
    template = _mini_template()
    for item in template["items"]:
        if item["id"] == "PD-AR-005":
            item["criterion"] = "Ведомость общих данных комплекта приложена"

    def doc_review_provider(dataset_id):
        return {"items": [
            _doc_review_item("G21.101-2026-D2-001", "supported_by_evidence", source_ref="ds-pd:v.pdf#page=1"),
            _doc_review_item("G21.101-2026-D2-002", "supported_by_evidence", source_ref="ds-pd:v.pdf#page=1"),
            _doc_review_item("G21.101-2026-D2-003", "supported_by_evidence", source_ref="ds-pd:v.pdf#page=1"),
        ]}

    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
        doc_review_provider=doc_review_provider,
    )
    for item in result["items"]:
        if item["suggested_answer"] in ("yes", "no"):
            refs = [ev.get("source_ref") for ev in item["document_evidence"]]
            assert any(refs)


def test_default_doc_review_provider_is_lazy_and_not_called_in_tests():
    """default_doc_review_provider существует как lazy-обёртка над
    doc_review_service.review_dataset — импортируется без вызова, тест не гоняет реальный
    doc_review (нет живой MetaDB/Qdrant в unit-тестах)."""
    from proxy.services.checklist_review_service import default_doc_review_provider

    assert callable(default_doc_review_provider)


# ══════════════════════════════════════════════════════════════════════════════════════════
# T2.4 (B) — normalized_remarks + defense_contract_v1 (содержательное наполнение)
# ══════════════════════════════════════════════════════════════════════════════════════════


def test_normalized_remarks_empty_when_no_items_need_remark():
    """Пустой прогон (ничего не найдено ни по одному item) -> normalized_remarks не крэшит,
    но т.к. presence/spatial_visual/spds_formal-заглушки дают review_needed/manual_required,
    remarks формируются и для этих статусов (computed_issue/supported_by_evidence/manual_required)."""
    template = {"name": "empty_template", "stage": "PD", "items": []}
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    assert result["normalized_remarks"] == []
    assert result["defense"]["claims"] == []


def test_normalized_remark_for_computed_issue_item_has_source_refs_and_checklist_ref():
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:spec_kr.pdf#page=4",
                "snippet": "Принят бетон класса В30 W8 для ростверка согласно спецификации КР.",
                "file_name": "spec_kr.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=search,
    )
    calc_item = next(it for it in result["items"] if it["item_id"] == "PD-KR-020")
    assert calc_item["status"] == "computed_issue"
    remark = next(r for r in result["normalized_remarks"] if r["id"] == "REM-CHK-PD-KR-020")
    assert remark["schema"] == "normalized_remark_v1"
    assert remark["source"] == "checklist"
    assert remark["category"] == "checklist"
    assert remark["checklist_ref"] == {"template": template["name"], "item_id": "PD-KR-020"}
    assert remark["source_refs"]
    assert any(ref for ref in remark["source_refs"])
    assert remark["finality"] == "proposed"
    assert calc_item["normalized_remark_ids"] == ["REM-CHK-PD-KR-020"]


def test_normalized_remark_for_supported_item_has_proposed_finality():
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:otchet_igi.pdf#page=1",
                "snippet": "Отчет об инженерно-геологических изысканиях выполнен в 2026 году.",
                "file_name": "otchet_igi.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=search,
    )
    item = next(it for it in result["items"] if it["item_id"] == "PD-OB-003")
    assert item["status"] == "supported_by_evidence"
    remark = next(r for r in result["normalized_remarks"] if r["id"] == "REM-CHK-PD-OB-003")
    assert remark["finality"] == "proposed"
    assert remark["human_decision"] == "unset"


def test_normalized_remarks_not_built_for_review_needed_or_not_applicable_by_default():
    """review_needed (нет evidence вовсе) не порождает remark по умолчанию — отсутствие evidence
    не является замечанием (implementation_plan.md §4 правило 2)."""
    template = _mini_template()
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    review_needed_ids = {it["item_id"] for it in result["items"] if it["status"] == "review_needed"}
    assert review_needed_ids  # предпосылка: в мини-template есть review_needed items
    remark_ids = {r["checklist_ref"]["item_id"] for r in result["normalized_remarks"]}
    assert not (review_needed_ids & remark_ids)


def test_normalized_remark_manual_required_item_present():
    """manual_required (spatial_visual) тоже порождает remark — инженеру нужно явно увидеть
    пункт в списке замечаний для решения, даже без кодового evidence."""
    template = _mini_template()
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    remark_ids = {r["checklist_ref"]["item_id"] for r in result["normalized_remarks"]}
    assert "PD-AR-049" in remark_ids  # spatial_visual -> manual_required


def test_defense_schema_and_domain():
    template = _mini_template()
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    defense = result["defense"]
    assert defense["schema"] == "defense_contract_v1"
    assert defense["domain"] == "normcontrol.checklist_review"
    assert defense["summary"]["human_final_required"] is True
    assert "by_status" in defense["summary"]
    assert defense["required_actions"]


def test_defense_claims_count_matches_remarks_and_have_source_refs_for_yes_no():
    template = _mini_template()
    search = _FakeSearchProvider({
        "ds-pd": [
            {
                "source_ref": "ds-pd:otchet_igi.pdf#page=1",
                "snippet": "Отчет об инженерно-геологических изысканиях выполнен в 2026 году.",
                "file_name": "otchet_igi.pdf",
            }
        ]
    })
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=search,
    )
    defense = result["defense"]
    assert len(defense["claims"]) == len(result["normalized_remarks"])
    for claim in defense["claims"]:
        item = next(it for it in result["items"] if it["item_id"] == claim["id"])
        if item["suggested_answer"] in ("yes", "no"):
            assert claim["source_refs"], f"{claim['id']}: claim без source_refs при yes/no"


def test_defense_empty_run_has_no_claims_and_does_not_crash():
    template = {"name": "empty_template", "stage": "PD", "items": []}
    result = run_checklist_review(
        template, dataset_id="ds-pd", inventory_provider=_empty_inventory, search_provider=_empty_search,
    )
    assert result["defense"]["claims"] == []
    assert result["defense"]["summary"]["total"] == 0


# --- Audit response 2026-07-04: центральный runtime-guard yes/no => source_ref ---

def test_evidence_guard_demotes_yes_without_source_ref():
    """Guard: будущий механизм, вернувший yes без source_ref, понижается до review_needed."""
    from proxy.services.checklist_review_service import _enforce_evidence_guard

    bad = {"item_id": "X-1", "status": "supported_by_evidence", "suggested_answer": "yes",
           "document_evidence": [{"kind": "document", "source_ref": "", "snippet": "..."}],
           "model_note": ""}
    out = _enforce_evidence_guard(dict(bad))
    assert out["suggested_answer"] == "unknown"
    assert out["status"] == "review_needed"
    assert "guard" in out["model_note"].lower()


def test_evidence_guard_keeps_valid_yes_and_manual():
    from proxy.services.checklist_review_service import _enforce_evidence_guard

    ok = {"item_id": "X-2", "status": "supported_by_evidence", "suggested_answer": "yes",
          "document_evidence": [{"kind": "document", "source_ref": "ds:file.pdf#page=1", "snippet": "s"}],
          "model_note": "n"}
    assert _enforce_evidence_guard(dict(ok)) == ok

    manual = {"item_id": "X-3", "status": "manual_required", "suggested_answer": "manual_required",
              "document_evidence": [], "model_note": ""}
    assert _enforce_evidence_guard(dict(manual)) == manual


def test_evidence_guard_demotes_no_without_any_evidence():
    from proxy.services.checklist_review_service import _enforce_evidence_guard

    bad = {"item_id": "X-4", "status": "computed_issue", "suggested_answer": "no",
           "document_evidence": [], "model_note": ""}
    out = _enforce_evidence_guard(dict(bad))
    assert out["suggested_answer"] == "unknown"
    assert out["status"] == "review_needed"

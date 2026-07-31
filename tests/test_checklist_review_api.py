"""Тест checklist-review API (T4.1): роутер /api/checklist-review/* + persist решений/прогонов.

По образцу test_doc_review_api.py, но здесь — через FastAPI TestClient (изолированный app с одним
роутером, без живого прокси/require_user-инфраструктуры), т.к. промпт T4.1 требует именно
TestClient/httpx стиль. Провайдеры (inventory/search/workbook/doc_review) подменяются через
monkeypatch модуля роутера (proxy.routers.checklist_review) — сам роутер зовёт
default_*_provider из checklist_review_service, поэтому патчим ИХ имена в неймспейсе роутера
(тот же трюк, что monkeypatch.setattr(drv, "_OUT_DIR", tmp_path) в test_doc_review_api.py, только
для провайдеров, а не для пути).

0 LLM: весь run — presence/calculation/parametric/... детерминированные механизмы
checklist_review_service, здесь мы просто фейково гоняем шаблон с 2 items (presence).
"""

from __future__ import annotations

import json

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from proxy.routers import checklist_review as crv
from proxy.security import RequestUser, require_user


def _fake_user():
    return RequestUser(role="user", holder="test", source="test")


@pytest.fixture()
def app(tmp_path, monkeypatch):
    """Изолированное FastAPI-приложение с одним роутером + persist в tmp_path."""
    monkeypatch.setattr(crv, "_OUT_DIR", tmp_path / "checklist_review")

    fastapi_app = FastAPI()
    fastapi_app.include_router(crv.router)
    fastapi_app.dependency_overrides[require_user] = _fake_user
    return fastapi_app


@pytest.fixture()
def client(app):
    return TestClient(app)


def _mini_template() -> dict:
    return {
        "name": "mini_api_template",
        "stage": "PD",
        "title": "Мини-шаблон для API-теста",
        "source_file_name": "mini.xlsx",
        "version": "",
        "disciplines": ["Общее"],
        "items": [
            {
                "id": "PD-OB-001",
                "stage": "PD",
                "discipline": "Общее",
                "sheet_name": "Общее",
                "row": 1,
                "item_no": "1",
                "section_path": [],
                "criterion": "Приложен отчет об инженерно-геологических изысканиях",
                "answer_cell": "C1",
                "note_cell": "D1",
                "allowed_answers": ["Да", "Нет"],
                "kind": "presence",
            },
            {
                "id": "PD-OB-002",
                "stage": "PD",
                "discipline": "Общее",
                "sheet_name": "Общее",
                "row": 2,
                "item_no": "2",
                "section_path": [],
                "criterion": "Приложен отчет об инженерно-геодезических изысканиях",
                "answer_cell": "C2",
                "note_cell": "D2",
                "allowed_answers": ["Да", "Нет"],
                "kind": "presence",
            },
        ],
    }


@pytest.fixture()
def patched_providers(monkeypatch):
    """Фейковые провайдеры инъецируются в неймспейс роутера (0 живых сервисов)."""
    def fake_load_template(name, base=None):
        if name != "mini_api_template":
            raise FileNotFoundError(name)
        return _mini_template()

    def fake_inventory(dataset_id):
        return []

    def fake_search(dataset_id, terms):
        if dataset_id == "ds-pd" and "изысканиях" in terms:
            return [{
                "source_ref": "ds-pd:otchet_igi.pdf#page=1",
                "snippet": "Отчет об инженерно-геологических изысканиях выполнен в 2026 году.",
                "file_name": "otchet_igi.pdf",
            }]
        return []

    monkeypatch.setattr(crv, "load_checklist_template", fake_load_template)
    monkeypatch.setattr(crv, "default_inventory_provider", fake_inventory)
    monkeypatch.setattr(crv, "default_search_provider", fake_search)
    monkeypatch.setattr(crv, "default_workbook_provider", lambda dataset_id: [])
    monkeypatch.setattr(crv, "default_doc_review_provider", lambda dataset_id: None)
    return fake_load_template


# ── GET /templates ──────────────────────────────────────────────────────────────────────


def test_templates_lists_glorax_pd_2026(client):
    r = client.get("/api/checklist-review/templates")
    assert r.status_code == 200
    body = r.json()
    names = [t["name"] for t in body["templates"]]
    assert "glorax_pd_2026" in names
    tpl = next(t for t in body["templates"] if t["name"] == "glorax_pd_2026")
    assert tpl["stage"] == "PD"
    assert tpl["title"]
    assert tpl["items_count"] == 335
    assert isinstance(tpl["disciplines"], list) and len(tpl["disciplines"]) == 10


# ── POST /{dataset_id}/run ──────────────────────────────────────────────────────────────


def test_run_returns_durable_run_id_and_summary(client, patched_providers):
    r = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template", "source_dataset_ids": [], "discipline": None},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["run_id"]
    assert body["dataset_id"] == "ds-pd"
    assert body["summary"]["total"] == 2
    assert len(body["items"]) == 2


def test_run_persists_to_disk(client, patched_providers, tmp_path):
    r = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    )
    run_id = r.json()["run_id"]
    run_path = tmp_path / "checklist_review" / "ds-pd" / run_id / "run.json"
    assert run_path.exists()
    saved = json.loads(run_path.read_text(encoding="utf-8"))
    assert saved["run_id"] == run_id
    assert saved["dataset_id"] == "ds-pd"


def test_run_respects_limit_param(client, patched_providers):
    r = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template", "limit": 1},
    )
    assert r.status_code == 200
    body = r.json()
    assert len(body["items"]) == 1
    assert body["summary"]["total"] == 1


def test_run_respects_discipline_filter(client, patched_providers):
    r = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template", "discipline": "Общее"},
    )
    assert r.status_code == 200
    assert r.json()["summary"]["total"] == 2

    r2 = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template", "discipline": "НетТакойДисциплины"},
    )
    assert r2.status_code == 200
    assert r2.json()["summary"]["total"] == 0


def test_run_unknown_template_is_404(client, patched_providers):
    r = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "no-such-template"},
    )
    assert r.status_code == 404


# ── GET /{dataset_id}/runs/{run_id} ─────────────────────────────────────────────────────


def test_get_run_returns_full_result_with_items(client, patched_providers):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]

    r = client.get(f"/api/checklist-review/ds-pd/runs/{run_id}")
    assert r.status_code == 200
    body = r.json()
    assert body["run_id"] == run_id
    assert len(body["items"]) == 2
    item = next(it for it in body["items"] if it["item_id"] == "PD-OB-001")
    assert item["status"] == "supported_by_evidence"
    assert item["suggested_answer"] == "yes"


def test_get_run_unknown_run_id_is_404(client):
    r = client.get("/api/checklist-review/ds-pd/runs/no-such-run")
    assert r.status_code == 404


# ── GET /{dataset_id}/runs/{run_id}/download ────────────────────────────────────────────


def test_download_json_returns_file(client, patched_providers):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]

    r = client.get(f"/api/checklist-review/ds-pd/runs/{run_id}/download", params={"fmt": "json"})
    assert r.status_code == 200
    assert "json" in r.headers.get("content-type", "")
    body = r.json()
    assert body["run_id"] == run_id


def test_download_html_returns_text_html(client, patched_providers):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]

    r = client.get(f"/api/checklist-review/ds-pd/runs/{run_id}/download", params={"fmt": "html"})
    assert r.status_code == 200
    assert "text/html" in r.headers.get("content-type", "")
    assert "<table" in r.text.lower()
    assert "PD-OB-001" in r.text


def test_download_xlsx_returns_readable_workbook(client, patched_providers, tmp_path):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]

    r = client.get(f"/api/checklist-review/ds-pd/runs/{run_id}/download", params={"fmt": "xlsx"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    assert "attachment" in r.headers.get("content-disposition", "")

    from openpyxl import load_workbook

    xlsx_path = tmp_path / "downloaded.xlsx"
    xlsx_path.write_bytes(r.content)
    wb = load_workbook(xlsx_path)
    assert "Чек-лист" in wb.sheetnames
    assert "Сводка" in wb.sheetnames
    assert "normalized_remarks" in wb.sheetnames


def test_download_bad_fmt_is_400(client, patched_providers):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]

    r = client.get(f"/api/checklist-review/ds-pd/runs/{run_id}/download", params={"fmt": "pdf"})
    assert r.status_code == 400


# ── POST /{dataset_id}/runs/{run_id}/items/{item_id}/decision ──────────────────────────


def test_decision_persists_and_applies_on_get(client, patched_providers):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]

    r = client.post(
        f"/api/checklist-review/ds-pd/runs/{run_id}/items/PD-OB-002/decision",
        json={"human_answer": "yes", "human_note": "Подтверждено инженером"},
    )
    assert r.status_code == 200
    assert r.json()["ok"] is True

    got = client.get(f"/api/checklist-review/ds-pd/runs/{run_id}").json()
    item = next(it for it in got["items"] if it["item_id"] == "PD-OB-002")
    assert item["human_answer"] == "yes"
    assert item["human_note"] == "Подтверждено инженером"
    # human_answer сильнее suggested_answer — summary содержит блок human
    assert got["summary"]["human"]["yes"] >= 1


def test_decision_unknown_item_id_is_404(client, patched_providers):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]

    r = client.post(
        f"/api/checklist-review/ds-pd/runs/{run_id}/items/NO-SUCH-ITEM/decision",
        json={"human_answer": "yes"},
    )
    assert r.status_code == 404


def test_decision_bad_human_answer_enum_rejected(client, patched_providers):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]

    r = client.post(
        f"/api/checklist-review/ds-pd/runs/{run_id}/items/PD-OB-001/decision",
        json={"human_answer": "definitely_yes"},
    )
    assert r.status_code in (400, 422)


def test_decision_unset_clears_previous(client, patched_providers):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]

    client.post(
        f"/api/checklist-review/ds-pd/runs/{run_id}/items/PD-OB-001/decision",
        json={"human_answer": "no", "human_note": "test"},
    )
    r = client.post(
        f"/api/checklist-review/ds-pd/runs/{run_id}/items/PD-OB-001/decision",
        json={"human_answer": "unset"},
    )
    assert r.status_code == 200
    got = client.get(f"/api/checklist-review/ds-pd/runs/{run_id}").json()
    item = next(it for it in got["items"] if it["item_id"] == "PD-OB-001")
    assert item["human_answer"] == "unset"


# ── path-guard: dataset_id/run_id с `..`/слэшами ────────────────────────────────────────


def test_path_guard_dataset_id_traversal_rejected(client, patched_providers):
    r = client.post(
        "/api/checklist-review/..%2F..%2Fetc/run",
        json={"template": "mini_api_template"},
    )
    assert r.status_code in (400, 404)


def test_path_guard_run_id_traversal_on_get_rejected(client, patched_providers):
    run_id = client.post(
        "/api/checklist-review/ds-pd/run",
        json={"template": "mini_api_template"},
    ).json()["run_id"]
    assert run_id  # sanity, датасет реально создан

    r = client.get("/api/checklist-review/ds-pd/runs/..%2F..%2Fsecret")
    assert r.status_code in (400, 404)


def test_path_guard_run_id_traversal_on_decision_rejected(client, patched_providers):
    r = client.post(
        "/api/checklist-review/ds-pd/runs/..%2F..%2Fsecret/items/PD-OB-001/decision",
        json={"human_answer": "yes"},
    )
    assert r.status_code in (400, 404)


def test_concurrent_decisions_do_not_lose_records(client, patched_providers):
    """Review-fix 2026-07-07 (гонка decisions): параллельные decision'ы по разным items
    одного run не теряют записи (read-modify-write под локом, атомарный os.replace)."""
    import concurrent.futures

    run = client.post("/api/checklist-review/ds-pd/run",
                      json={"template": "mini_api_template"}).json()
    run_id = run["run_id"]
    targets = [it["item_id"] for it in run["items"]][:2]
    assert len(targets) == 2, "фикстурный template должен давать >=2 items"

    def _post(iid):
        return client.post(
            f"/api/checklist-review/ds-pd/runs/{run_id}/items/{iid}/decision",
            json={"human_answer": "yes", "human_note": f"note-{iid}"},
        )

    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as ex:
        codes = [r.status_code for r in ex.map(_post, targets)]
    assert all(c == 200 for c in codes)

    got = client.get(f"/api/checklist-review/ds-pd/runs/{run_id}").json()
    answered = {it["item_id"]: it.get("human_answer") for it in got["items"]}
    for iid in targets:
        assert answered.get(iid) == "yes", f"решение по {iid} потеряно"

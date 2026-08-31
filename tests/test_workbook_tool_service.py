from __future__ import annotations

import hashlib
import inspect
from pathlib import Path

import openpyxl
import pytest

from proxy.services.artifact_revision_service import ArtifactRevisionStore
from proxy.services.chat_attachment_service import preserve_read_attachment
from proxy.services.workflow_checkpoint_service import WorkflowCheckpointService
from proxy.services.workbook_tool_service import (
    WorkbookExecutionContext,
    build_lsr_workbook,
    build_vor_workbook,
)
from proxy.services.lsr_workbook_adapter_service import build_lsr_workbook_from_decisions


def _source_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Раздел", "Наименование", "Ед. изм.", "Количество", "Код"])
    sheet.append(["ОВ", "Воздуховод", "м", 12.5, "A-1"])
    sheet.append(["ОВ", "Решётка", None, None, "A-2"])
    workbook.save(path)
    workbook.close()
    return path


def _context(tmp_path: Path, *, key: str = "key-1", decision: str = "decision-1",
             lsr_adapter=None) -> WorkbookExecutionContext:
    return WorkbookExecutionContext(
        session_id="session-1",
        idempotency_key=key,
        model_decision_revision=decision,
        profile_revision_id="profile-1",
        model_identity="qwen-local",
        model_preset="qwen-9b",
        attachment_root=tmp_path / "attachments",
        work_dir=tmp_path / "work",
        checkpoints=WorkflowCheckpointService(tmp_path / "checkpoints.db"),
        artifacts=ArtifactRevisionStore(tmp_path / "artifacts.db", tmp_path / "artifacts"),
        lsr_adapter=lsr_adapter,
    )


def _attachment(tmp_path: Path, *, ident: str = "read_123456abcdef") -> dict:
    source = _source_workbook(tmp_path / f"{ident}.xlsx")
    return preserve_read_attachment(
        source,
        attachment_id=ident,
        original_name="source.xlsx",
        root=tmp_path / "attachments",
    )


@pytest.mark.asyncio
async def test_vor_handler_preserves_source_rows_and_creates_revision(tmp_path):
    meta = _attachment(tmp_path)
    ctx = _context(tmp_path)

    result = await build_vor_workbook({"attachment_id": meta["attachment_id"]}, ctx)

    assert result["schema"] == "les.workbook_tool_result.v1"
    assert result["status"] == "complete"
    assert result["artifact"]["revision_no"] == 1
    assert result["source"]["sha256"] == meta["sha256"]
    assert result["source"]["rows"] == 2
    assert result["missing"] == ["row:2:unit", "row:2:quantity"]
    stored = ctx.artifacts.resolve_path(result["artifact"]["revision_id"])
    workbook = openpyxl.load_workbook(stored, data_only=True, read_only=True)
    try:
        values = list(workbook.active.iter_rows(min_row=4, values_only=True))
    finally:
        workbook.close()
    assert values[0][1:6] == ("ОВ", "Воздуховод", "A-1", "м", 12.5)
    assert values[1][1:6] == ("ОВ", "Решётка", "A-2", None, None)
    assert "source.xlsx#1" in values[0][6]
    assert "file_path" not in str(result)


@pytest.mark.asyncio
async def test_lsr_handler_does_not_accept_model_supplied_prices(tmp_path):
    meta = _attachment(tmp_path)
    result = await build_lsr_workbook(
        {"attachment_id": meta["attachment_id"], "rows": [{"price": 1}], "prices": [1]},
        _context(tmp_path),
    )

    assert result["status"] == "rejected"
    assert result["code"] == "MODEL_DECISION_FIELD_NOT_ALLOWED"


@pytest.mark.asyncio
async def test_lsr_handler_requires_an_explicit_application_adapter(tmp_path):
    meta = _attachment(tmp_path)

    result = await build_lsr_workbook({"attachment_id": meta["attachment_id"]}, _context(tmp_path))

    assert result["status"] == "failed"
    assert result["code"] == "WORKBOOK_ADAPTER_UNAVAILABLE"


@pytest.mark.asyncio
async def test_thin_lsr_adapter_renders_only_model_selected_norms(tmp_path, monkeypatch):
    decisions = [{
        "source_row": 1,
        "section": "ОВ",
        "title": "Монтаж воздуховода",
        "unit": "м",
        "quantity": 12.5,
        "norm_code": "ГЭСН20-01-001-01",
    }]
    captured = {}

    def calculate(rows, **kwargs):
        captured["rows"] = rows
        captured["kwargs"] = kwargs
        return {
            "schema": "rim_lsr_v1",
            "sections": [],
            "summary": {"input_rows": 1, "bound_rows": 1, "flags": []},
            "row_bindings": [{"row": 1, "status": "bound"}],
        }

    def render(trace, output_path, **_kwargs):
        captured["trace"] = trace
        _source_workbook(output_path)
        return output_path

    monkeypatch.setattr(
        "proxy.services.lsr_workbook_adapter_service.build_lsr_trace_from_visible_rows",
        calculate,
    )
    monkeypatch.setattr(
        "proxy.services.lsr_workbook_adapter_service.render_lsr_xlsx",
        render,
    )

    result = await build_lsr_workbook_from_decisions(
        tmp_path / "source.xlsx",
        {"decisions": decisions, "question": "Собери ЛСР"},
        tmp_path / "result.xlsx",
        lambda *_args: None,
    )

    assert captured["rows"] == decisions
    assert captured["kwargs"]["name"] == "Собери ЛСР"
    assert result["source_rows"] == 1
    assert result["missing"] == []
    assert result["blockers"] == []
    assert result["file_path"].exists()


def test_workbook_boundary_does_not_import_old_document_workflow():
    import proxy.services.workbook_tool_service as service

    source = inspect.getsource(service)
    assert "smeta_chat_application_service" not in source
    assert "proxy.smeta_core" not in source


@pytest.mark.asyncio
async def test_retry_returns_completed_revision_without_running_adapter_twice(tmp_path):
    meta = _attachment(tmp_path)
    calls = 0

    async def adapter(_path, _args, output_path, _progress):
        nonlocal calls
        calls += 1
        _source_workbook(output_path)
        return {"file_path": output_path, "missing": [], "blockers": [], "source_rows": 2}

    ctx = _context(tmp_path, lsr_adapter=adapter)
    args = {"attachment_id": meta["attachment_id"]}
    first = await build_lsr_workbook(args, ctx)
    second = await build_lsr_workbook(args, ctx)

    assert calls == 1
    assert second["artifact"]["revision_id"] == first["artifact"]["revision_id"]
    assert second["source"]["rows"] == first["source"]["rows"] == 2
    assert second["checkpoint"]["resumed"] is True


@pytest.mark.asyncio
async def test_parent_revision_creates_correction_without_overwriting_first(tmp_path):
    meta = _attachment(tmp_path)
    ctx1 = _context(tmp_path, key="key-1")
    first = await build_vor_workbook({"attachment_id": meta["attachment_id"]}, ctx1)
    first_bytes = ctx1.artifacts.read_bytes(first["artifact"]["revision_id"])

    ctx2 = _context(tmp_path, key="key-2", decision="decision-2")
    second = await build_vor_workbook(
        {
            "attachment_id": meta["attachment_id"],
            "parent_revision_id": first["artifact"]["revision_id"],
            "question": "Исправленная редакция",
        },
        ctx2,
    )

    assert second["artifact"]["artifact_id"] == first["artifact"]["artifact_id"]
    assert second["artifact"]["revision_no"] == 2
    assert second["artifact"]["parent_revision_id"] == first["artifact"]["revision_id"]
    assert ctx2.artifacts.read_bytes(first["artifact"]["revision_id"]) == first_bytes


@pytest.mark.asyncio
async def test_attachment_hash_drift_is_rejected_before_generation(tmp_path):
    meta = _attachment(tmp_path)
    stored = tmp_path / "attachments" / "read_123456abcdef.xlsx"
    stored.write_bytes(b"changed")

    result = await build_vor_workbook({"attachment_id": meta["attachment_id"]}, _context(tmp_path))

    assert result["status"] == "rejected"
    assert result["code"] == "ATTACHMENT_INVALID"


@pytest.mark.asyncio
async def test_adapter_failure_publishes_no_artifact_and_keeps_retryable_checkpoint(tmp_path):
    meta = _attachment(tmp_path)

    async def broken(*_args):
        raise RuntimeError("adapter stopped")

    ctx = _context(tmp_path, lsr_adapter=broken)
    result = await build_lsr_workbook({"attachment_id": meta["attachment_id"]}, ctx)

    assert result["status"] == "failed"
    assert result["code"] == "WORKBOOK_GENERATION_FAILED"
    checkpoint = ctx.checkpoints.get(result["checkpoint"]["checkpoint_id"])
    assert checkpoint.status == "failed"
    assert checkpoint.artifact_revision_id is None
    assert not list((tmp_path / "artifacts").rglob("*.xlsx"))


@pytest.mark.asyncio
async def test_failed_checkpoint_resumes_and_can_publish_on_retry(tmp_path):
    meta = _attachment(tmp_path)
    calls = 0

    async def flaky(_path, _args, output_path, _progress):
        nonlocal calls
        calls += 1
        if calls == 1:
            raise RuntimeError("interrupted")
        _source_workbook(output_path)
        return {"file_path": output_path, "missing": [], "blockers": [], "source_rows": 2}

    ctx = _context(tmp_path, lsr_adapter=flaky)
    first = await build_lsr_workbook({"attachment_id": meta["attachment_id"]}, ctx)
    second = await build_lsr_workbook({"attachment_id": meta["attachment_id"]}, ctx)

    assert first["status"] == "failed"
    assert second["status"] == "complete"
    assert second["checkpoint"]["checkpoint_id"] == first["checkpoint"]["checkpoint_id"]
    assert calls == 2


@pytest.mark.asyncio
async def test_unsupported_attachment_is_rejected_without_checkpoint(tmp_path):
    source = tmp_path / "source.txt"
    source.write_text("not a workbook", encoding="utf-8")
    meta = preserve_read_attachment(
        source,
        attachment_id="read_123456abcdef",
        original_name="source.txt",
        root=tmp_path / "attachments",
    )
    ctx = _context(tmp_path)

    result = await build_vor_workbook({"attachment_id": meta["attachment_id"]}, ctx)

    assert result["status"] == "rejected"
    assert result["code"] == "UNSUPPORTED_ATTACHMENT_TYPE"
    assert not (tmp_path / "checkpoints.db").read_bytes() == b""


def test_source_fixture_hash_is_stable(tmp_path):
    meta = _attachment(tmp_path)
    payload = (tmp_path / "attachments" / "read_123456abcdef.xlsx").read_bytes()
    assert hashlib.sha256(payload).hexdigest() == meta["sha256"]

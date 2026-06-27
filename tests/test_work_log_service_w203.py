"""W20.3 — Общий журнал работ (ОЖР): раздел 3 из журнала объёмов бит-в-бит. 0 LLM."""
import importlib

import pytest


@pytest.fixture()
def svc(tmp_path, monkeypatch):
    (tmp_path / "data").mkdir()
    monkeypatch.setenv("RAG_META_DB_PATH", str(tmp_path / "data" / "les_meta.db"))
    monkeypatch.setenv("LES_WORKLOG_OUT_DIR", str(tmp_path / "data" / "worklog_out"))
    import backend.rag_config as rc
    importlib.reload(rc)
    import proxy.services.field_intake_service as fis
    importlib.reload(fis)
    import proxy.services.project_service as ps
    importlib.reload(ps)
    import proxy.services.work_log_service as wl
    importlib.reload(wl)
    return wl, fis, ps


def test_section3_from_journal_bit_for_bit(svc):
    wl, fis, ps = svc
    pid = ps.create_project("БЦ Тест")["id"]
    fis.create_entry("Бетонирование плиты", 25.0, "м3", entry_date="2026-06-02", zahvatka="З-1", status="confirmed", project_id=pid)
    fis.create_entry("Кладка стен", 120.0, "м2", entry_date="2026-06-01", status="confirmed", project_id=pid)
    fis.create_entry("Чужой объект", 9.0, "м3", entry_date="2026-06-01", status="confirmed", project_id=999)
    fis.create_entry("Не подтверждено", 5.0, "м3", entry_date="2026-06-03", status="pending", project_id=pid)

    s3 = wl.build_section3(pid)
    # только confirmed этого объекта, хронологически
    assert [r["work"] for r in s3] == ["Кладка стен", "Бетонирование плиты"]
    assert s3[0]["volume"] == 120.0 and s3[0]["unit"] == "м2"
    assert s3[1]["zahvatka"] == "З-1"


def test_build_work_log_structure(svc):
    wl, fis, ps = svc
    pid = ps.create_project("БЦ «Банкрот»")["id"]
    fis.create_entry("Монтаж воздуховодов", 80.0, "м", status="confirmed", project_id=pid)
    log = wl.build_work_log(pid)
    assert log["title"].startswith("ОБЩИЙ ЖУРНАЛ РАБОТ")
    assert log["header"]["object_name"] == "БЦ «Банкрот»"  # из карточки проекта
    assert log["section3_count"] == 1
    assert log["total_volume"] == 80.0


def test_meta_set_get(svc):
    wl, _, ps = svc
    pid = ps.create_project("О")["id"]
    wl.set_work_log_meta(pid, customer="ООО Заказчик", itr=["Иванов И.И. — прораб"], spec_journals=["Журнал бетонных работ"])
    meta = wl.get_work_log_meta(pid)
    assert meta["customer"] == "ООО Заказчик"
    assert meta["itr"] == ["Иванов И.И. — прораб"]
    assert meta["spec_journals"] == ["Журнал бетонных работ"]
    # шапка подхватывается в build_work_log
    log = wl.build_work_log(pid)
    assert log["itr"] == ["Иванов И.И. — прораб"]


def test_export_xlsx(svc):
    wl, fis, ps = svc
    pid = ps.create_project("О")["id"]
    fis.create_entry("Работа", 3.0, "шт", status="confirmed", project_id=pid)
    path = wl.export_xlsx(pid)
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    flat = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("ОБЩИЙ ЖУРНАЛ РАБОТ" in str(v) for v in flat)
    assert any("Раздел 3" in str(v) for v in flat)
    assert "Работа" in flat

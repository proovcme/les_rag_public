from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from proxy.routers.chat import _augment_model_tool_args
from proxy.services import chat_attachment_service
from proxy.services.smeta_workbook_tools import (
    LSR_TOOL,
    VOR_TOOL,
    build_lsr_workbook,
    build_vor_workbook,
    maybe_forced_workbook_call,
    workbook_artifact_from_tool_results,
    workbook_file_intent,
)
from proxy.services.tool_harness_service import ToolHarness


def test_workbook_file_intent_distinguishes_lsr_and_vor():
    assert workbook_file_intent("Собери первичную ЛСР по вложению") == "lsr"
    assert workbook_file_intent("сделай ВОР из спецификации") == "vor"
    assert workbook_file_intent("Собери ЛСР по приложенной ВОР") == "lsr"
    assert workbook_file_intent("подбери нормы ГЭСНм") is None
    assert workbook_file_intent("повороты лотка") is None


def test_forced_workbook_call_requires_attachment_and_profile_tool():
    assert maybe_forced_workbook_call(
        question="собери ЛСР",
        attachment_id="",
        profile_tools=[LSR_TOOL],
        already_called=[],
    ) is None
    forced = maybe_forced_workbook_call(
        question="собери ЛСР",
        attachment_id="read_123456abcdef",
        profile_tools=[LSR_TOOL, VOR_TOOL],
        already_called=[],
    )
    assert forced == {
        "tool": LSR_TOOL,
        "args": {"attachment_id": "read_123456abcdef", "question": "собери ЛСР"},
    }
    assert maybe_forced_workbook_call(
        question="собери ЛСР",
        attachment_id="read_123456abcdef",
        profile_tools=[LSR_TOOL],
        already_called=[LSR_TOOL],
    ) is None


def test_unrestricted_rag_scope_does_not_break_workbook_arg_augment():
    from proxy.services.chat_evidence_application_service import tool_dataset_ids

    assert tool_dataset_ids(None) == []
    forced = maybe_forced_workbook_call(
        question="Собери ЛСР по приложенной ВОР",
        attachment_id="read_90dfdf59952f",
        profile_tools=[LSR_TOOL, VOR_TOOL],
        already_called=[],
    )
    call = _augment_model_tool_args(
        forced,
        question="Собери ЛСР по приложенной ВОР",
        dataset_ids=tool_dataset_ids(None),
        target_file_ref=None,
        attachment_id="read_90dfdf59952f",
    )
    assert call["tool"] == LSR_TOOL
    assert call["args"]["attachment_id"] == "read_90dfdf59952f"


def test_augment_fills_attachment_id_for_workbook_tools():
    call = _augment_model_tool_args(
        {"tool": LSR_TOOL, "args": {}},
        question="собери ЛСР",
        dataset_ids=["ds1"],
        target_file_ref=None,
        attachment_id="read_123456abcdef",
        project_id=7,
    )
    assert call["args"]["attachment_id"] == "read_123456abcdef"
    assert call["args"]["question"] == "собери ЛСР"
    assert call["args"]["project_id"] == 7


def test_tool_harness_registers_and_pins_workbook_tools():
    harness = ToolHarness()
    names = {item["name"] for item in harness.registry()["tools"]}
    assert {LSR_TOOL, VOR_TOOL} <= names
    shortlist = harness.shortlist(
        "что такое пожарная сигнализация",
        allowed_tools=["search_sources", LSR_TOOL, VOR_TOOL],
        limit=2,
    )
    short_names = {item["name"] for item in shortlist["tools"]}
    assert LSR_TOOL in short_names
    assert VOR_TOOL in short_names


def test_lsr_workbook_tool_wraps_existing_application(monkeypatch):
    async def fake_run(**kwargs):
        assert kwargs["attachment_id"] == "read_123456abcdef"
        assert kwargs.get("token_sink") is None
        assert "rows" not in kwargs
        return SimpleNamespace(
            answer="ЛСР собрана кодом",
            operation="smeta_document_lsr",
            channel="smeta_mode",
            crag="PARTIAL",
            extra={
                "artifact": {
                    "mode": "xlsx",
                    "downloads": {"xlsx": "/api/smeta-artifacts/download?path=LSR.xlsx"},
                }
            },
        )

    monkeypatch.setattr(
        "proxy.services.smeta_chat_application_service.run_smeta_document_application",
        fake_run,
    )
    payload = build_lsr_workbook({"attachment_id": "read_123456abcdef", "question": "собери ЛСР"})
    assert payload["status"] == "ok"
    assert payload["result"]["artifact"]["downloads"]["xlsx"].endswith("LSR.xlsx")
    assert workbook_artifact_from_tool_results([payload])["mode"] == "xlsx"


def test_lsr_async_path_forwards_chat_token_sink(monkeypatch):
    import asyncio

    from proxy.services.smeta_workbook_tools import build_lsr_workbook_async

    seen = {}

    async def fake_run(**kwargs):
        seen["sink"] = kwargs.get("token_sink")
        return SimpleNamespace(
            answer="ЛСР собрана кодом",
            operation="smeta_document_lsr",
            channel="smeta_mode",
            crag="PARTIAL",
            extra={"artifact": {"mode": "xlsx", "downloads": {"xlsx": "/api/smeta-artifacts/download?path=LSR.xlsx"}}},
        )

    monkeypatch.setattr(
        "proxy.services.smeta_chat_application_service.run_smeta_document_application",
        fake_run,
    )

    async def sink(_event):
        return None

    payload = asyncio.run(
        build_lsr_workbook_async(
            {"attachment_id": "read_123456abcdef", "question": "собери ЛСР"},
            token_sink=sink,
        )
    )
    assert seen["sink"] is sink
    assert payload["status"] == "ok"


def _write_spec_xlsx(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Раздел", "Наименование", "Ед. изм.", "Кол-во"])
    ws.append(["П1", "Кабель ВВГнг 3х2,5", "м", 12])
    ws.append(["П1", "Лоток 200х50", "м", 4])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def test_vor_workbook_tool_writes_quantities_only_xlsx(tmp_path, monkeypatch):
    source = tmp_path / "spec.xlsx"
    _write_spec_xlsx(source)
    store = tmp_path / "store"
    out_dir = tmp_path / "out"
    monkeypatch.setattr(chat_attachment_service, "DEFAULT_ROOT", store)
    monkeypatch.setattr("proxy.services.smeta_workbook_tools.SMETA_ARTIFACT_DIR", out_dir)
    chat_attachment_service.preserve_read_attachment(
        source,
        attachment_id="read_aaaaaaaaaaaa",
        original_name="spec.xlsx",
        root=store,
    )

    payload = build_vor_workbook({
        "attachment_id": "read_aaaaaaaaaaaa",
        "question": "сделай ВОР из спецификации",
    })

    assert payload["status"] == "ok"
    artifact = payload["result"]["artifact"]
    xlsx_path = Path(artifact["files"]["xlsx_path"])
    assert xlsx_path.is_file()
    assert payload["result"]["bor_lines"] >= 2
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    header_text = " ".join(
        str(cell.value or "")
        for row in wb.active.iter_rows(min_row=1, max_row=4, max_col=8)
        for cell in row
    ).casefold()
    assert "цена" not in header_text
    assert "стоимость" not in header_text

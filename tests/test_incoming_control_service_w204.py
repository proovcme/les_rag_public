"""W20.4 — Входной контроль (ГОСТ 24297): акты, журнал, реестр документов качества. 0 LLM."""
import importlib

import pytest


@pytest.fixture()
def svc(tmp_path, monkeypatch):
    (tmp_path / "data").mkdir()
    monkeypatch.setenv("RAG_META_DB_PATH", str(tmp_path / "data" / "les_meta.db"))
    monkeypatch.setenv("LES_INCOMING_CONTROL_OUT_DIR", str(tmp_path / "data" / "vk_out"))
    import backend.rag_config as rc
    importlib.reload(rc)
    import proxy.services.project_service as ps
    importlib.reload(ps)
    import proxy.services.incoming_control_service as ics
    importlib.reload(ics)
    return ics, ps


def test_record_creates_journal_row_and_act(svc):
    ics, ps = svc
    pid = ps.create_project("БЦ «Банкрот»")["id"]
    doc = ics.add_quality_doc(pid, "сертификат", "СС-2026/14", material="Арматура А500С",
                              issued_by="МеталлТорг", valid_until="2027-01-01")
    rec = ics.add_incoming_control(pid, "Арматура А500С", batch="П-101", control_date="2026-06-10",
                                   quality_doc_id=doc["id"], quantity=12.5, unit="т",
                                   result="соответствует", decision=ics.ADMITTED, inspector="Петров П.П.")
    # строка журнала
    journal = ics.build_journal(pid)
    assert journal["count"] == 1
    assert journal["rows"][0]["material"] == "Арматура А500С"
    assert journal["rows"][0]["quality_doc"] == "сертификат № СС-2026/14"
    assert journal["admitted_count"] == 1 and journal["rejected_count"] == 0
    # акт по партии
    act = ics.build_act(pid, rec["id"])
    assert act["title"] == "АКТ ВХОДНОГО КОНТРОЛЯ"
    assert act["batch"] == "П-101"
    assert act["quality_doc"]["number"] == "СС-2026/14"
    assert act["decision"] == ics.ADMITTED


def test_rejected_control_is_visible(svc):
    ics, ps = svc
    pid = ps.create_project("О")["id"]
    ics.add_incoming_control(pid, "Кирпич", batch="П-1", decision=ics.ADMITTED)
    ics.add_incoming_control(pid, "Цемент бракованный", batch="П-2", decision=ics.REJECTED,
                             result="не соответствует ГОСТ")
    journal = ics.build_journal(pid)
    assert journal["has_rejected"] is True
    assert journal["rejected_count"] == 1
    rejected = ics.list_incoming_control(pid, decision=ics.REJECTED)
    assert len(rejected) == 1 and rejected[0]["material"] == "Цемент бракованный"


def test_quality_docs_registry_search_and_expiry(svc):
    ics, ps = svc
    pid = ps.create_project("О")["id"]
    ics.add_quality_doc(pid, "паспорт", "ПАС-1", material="Бетон B25", valid_until="2030-01-01")
    ics.add_quality_doc(pid, "декларация", "ДЕК-9", material="Утеплитель", valid_until="2020-01-01")
    ics.add_quality_doc(pid, "сертификат", "СЕРТ-7", material="Бетон B30")  # без срока — не истекает

    docs = ics.list_quality_docs(pid, as_of="2026-06-15")
    assert len(docs) == 3
    expired = {d["number"]: d["expired"] for d in docs}
    assert expired == {"ПАС-1": False, "ДЕК-9": True, "СЕРТ-7": False}

    # поиск по реестру
    found = ics.list_quality_docs(pid, query="бетон", as_of="2026-06-15")
    assert {d["number"] for d in found} == {"ПАС-1", "СЕРТ-7"}


def test_project_partition_q3(svc):
    ics, ps = svc
    a = ps.create_project("A")["id"]
    b = ps.create_project("B")["id"]
    ics.add_incoming_control(a, "Материал A", batch="A-1")
    ics.add_incoming_control(b, "Материал B", batch="B-1")
    assert ics.build_journal(a)["count"] == 1
    assert ics.build_journal(a)["rows"][0]["material"] == "Материал A"
    assert ics.build_journal(b)["count"] == 1


def test_act_rejects_cross_project_record(svc):
    ics, ps = svc
    a = ps.create_project("A")["id"]
    b = ps.create_project("B")["id"]
    rec = ics.add_incoming_control(a, "Материал A")
    assert ics.build_act(a, rec["id"])  # свой объект — ок
    assert ics.build_act(b, rec["id"]) == {}  # чужой объект — пусто


def test_export_journal_xlsx_marks_red(svc):
    ics, ps = svc
    pid = ps.create_project("О")["id"]
    ics.add_incoming_control(pid, "Годный", batch="П-1", decision=ics.ADMITTED)
    ics.add_incoming_control(pid, "Брак", batch="П-2", decision=ics.REJECTED)
    path = ics.export_journal_xlsx(pid, as_of="2026-06-15")
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    flat = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("ЖУРНАЛ ВХОДНОГО КОНТРОЛЯ" in str(v) for v in flat)
    assert "Брак" in flat and "Годный" in flat
    # «красная» заливка стоит хотя бы на одной ячейке (строка «не допущено»)
    fills = [c.fill.start_color.rgb for row in ws.iter_rows() for c in row
             if c.value is not None and c.fill and c.fill.fill_type == "solid"]
    assert any(str(f).endswith("FFD6D6") for f in fills)

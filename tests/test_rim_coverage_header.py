"""Low-coverage banner in RIM XLSX header."""

from __future__ import annotations

from openpyxl import load_workbook

from proxy.services import rim_trace_xlsx_service as rim_xlsx


def test_low_coverage_header_label_in_xlsx(tmp_path):
    trace = {
        "name": "Тест",
        "sections": [],
        "summary": {
            "result_status": "priced_draft",
            "input_rows": 19,
            "bound_rows": 6,
            "open_rows": 13,
            "unbound_rows": 13,
            "covered_rows": 0,
            "total": 17229.5,
            "total_with_vat": 21020.0,
            "ozp": 17229.5,
            "zpm": 0,
            "labor_qty": 30,
            "machinist_qty": 0,
            "full_amount": None,
        },
    }
    out = rim_xlsx.render_lsr_xlsx(trace, tmp_path / "low.xlsx")
    blob = " | ".join(
        str(c.value)
        for row in load_workbook(out).active.iter_rows()
        for c in row
        if c.value is not None
    )
    assert "ПОКРЫТИЕ НИЗКОЕ" in blob
    assert "Стоимость только привязанной части (6/19)" in blob
    assert "не итог ведомости" in blob

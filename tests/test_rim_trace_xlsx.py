"""Тест рендерера РИМ-трассы в XLSX по форме Приложения 3 к 421/пр.

Рендерер из ГОТОВОЙ трассы, не калькулятор → числа те же, что в build_position_trace.
Эталон ГЭСН12-01-034-02 @ 0.61 → summary.total == 11896.35 после полной
детализации ОТм через ФСЭМ (тот же trace, что сервис/endpoint-тесты).
"""

import asyncio
from pathlib import Path

from proxy.services import lsr_assembly_service as la
from proxy.services import rim_lsr_trace_service as rim
from proxy.services import rim_trace_xlsx_service as rim_xlsx


CODE = "ГЭСН12-01-034-02"
EXPECTED_TOTAL = 11896.35


def _trace():
    book = la._resolve_book(None)
    return rim.build_position_trace(
        {"code": CODE, "qty": 0.61}, pricebook=book, k_ozp=1.0, k_em=1.0
    )


def _lsr_trace():
    book = la._resolve_book(None)
    positions = [
        {"code": CODE, "qty": 0.61, "section": "Раздел 1. Кровля"},
        {"code": CODE, "qty": 0.61, "section": "Раздел 2. Прочее"},
    ]
    return rim.build_lsr_trace(positions, pricebook=book, name="Смета на кровлю")


def test_render_produces_valid_xlsx(tmp_path):
    import openpyxl

    trace = _trace()
    out = rim_xlsx.render_trace_xlsx(trace, tmp_path / "trace.xlsx")
    assert out.exists() and out.stat().st_size > 0
    ws = openpyxl.load_workbook(out).active
    assert "ЛСР" in ws.title or "Прил" in ws.title
    # XLSX теперь редактируемый: итог позиции — формула, immutable число остаётся в JSON trace.
    total_row = next(row for row in ws.iter_rows() if row[2].value == "Всего по позиции")
    assert str(total_row[11].value).startswith("=SUM(")


def test_appendix3_rim_header_and_graphs_present(tmp_path):
    import openpyxl

    out = rim_xlsx.render_trace_xlsx(_trace(), tmp_path / "t.xlsx")
    txt = " ".join(str(c.value) for row in openpyxl.load_workbook(out).active.iter_rows() for c in row if c.value)
    # форма Приложения 3: шапка РИМ + графы + свод
    assert "Приложение № 3" in txt and "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ" in txt and "Сметная стоимость" in txt
    assert "Обоснование" in txt and "Наименование работ и затрат" in txt
    assert "Индекс" in txt and "Сметная стоимость в текущем уровне цен всего" in txt
    assert "Всего по позиции" in txt and "ОТ(ЗТ)" in txt and "ЭМ" in txt


def test_endpoint_export_returns_download():
    from proxy.routers.lsr import RimTraceRequest, lsr_rim_trace_export

    res = asyncio.run(
        lsr_rim_trace_export(
            RimTraceRequest(position={"code": CODE, "qty": 0.61}), _user=object()
        )
    )
    assert res["code"] == CODE
    assert res["summary"]["total"] == EXPECTED_TOTAL
    assert res["download"].startswith("/api/lsr/download?path=rim_trace_")
    assert Path(res["path"]).exists()


def test_render_lsr_multi_position_form(tmp_path):
    import openpyxl

    lsr = _lsr_trace()
    out = rim_xlsx.render_lsr_xlsx(lsr, tmp_path / "lsr.xlsx")
    assert out.exists() and out.stat().st_size > 0
    ws = openpyxl.load_workbook(out).active
    txt = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    # форма Приложения 3 + разделы + итоги разделов + общий свод
    assert "Приложение № 3" in txt and "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ" in txt
    assert "Раздел 1. Кровля" in txt and "Раздел 2. Прочее" in txt
    assert "Итого по разделу 1" in txt and "Итого по разделу 2" in txt
    assert "ВСЕГО по смете" in txt
    # общий итог — формула по позициям, а не застывшая константа.
    total_row = next(row for row in ws.iter_rows() if row[2].value == "ВСЕГО по смете без НДС")
    assert str(total_row[11].value).startswith("=SUM(")
    assert lsr["summary"]["total"] == 2 * EXPECTED_TOTAL


def test_render_lsr_prints_all_twelve_graphs_on_one_page_width(tmp_path):
    import openpyxl

    out = rim_xlsx.render_lsr_xlsx(_lsr_trace(), tmp_path / "print-ready.xlsx")
    ws = openpyxl.load_workbook(out).active

    assert ws.page_setup.orientation == ws.ORIENTATION_LANDSCAPE
    assert str(ws.page_setup.paperSize) == str(ws.PAPERSIZE_A4)
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True
    assert ws.print_area == f"'ЛСР РИМ'!$A$1:$L${ws.max_row}"
    assert ws.print_title_rows is not None


def test_render_lsr_single_section_omits_section_header(tmp_path):
    import openpyxl

    book = la._resolve_book(None)
    lsr = rim.build_lsr_trace([{"code": CODE, "qty": 0.61}], pricebook=book)
    out = rim_xlsx.render_lsr_xlsx(lsr, tmp_path / "one.xlsx")
    txt = " ".join(str(c.value) for row in openpyxl.load_workbook(out).active.iter_rows() for c in row if c.value)
    # один безымянный раздел → без «Раздел 1», но «ВСЕГО по смете» есть
    assert "Раздел 1" not in txt
    assert "ВСЕГО по смете" in txt
    assert "Всего по позиции" in txt


def test_lsr_header_uses_passed_source_metadata_and_review_issues_are_readable(tmp_path):
    import openpyxl

    lsr = _lsr_trace()
    lsr["row_bindings"] = []
    lsr["blockers"] = [{"code": "norm_selection_required", "work_id": "w1", "reason": "нужен подбор нормы"}]
    lsr["warnings"] = ["Длинное профессиональное замечание, которое должно занимать широкую область листа"]
    out = rim_xlsx.render_lsr_xlsx(
        lsr,
        tmp_path / "meta-review.xlsx",
        meta={
            "object": "ВОР монтаж БАП П1 13.05",
            "stroika": "Не указано в исходной ВОР",
            "lsr_no": "б/н",
            "osnovanie": "ВОР монтаж БАП П1 13.05.pdf",
            "subject": "Санкт-Петербург",
            "price_level": "2 квартал 2026",
        },
    )
    wb = openpyxl.load_workbook(out)
    header_text = " ".join(str(cell.value) for row in wb["ЛСР РИМ"].iter_rows(max_row=15) for cell in row if cell.value)
    assert "read_" not in header_text
    assert "ВОР монтаж БАП П1 13.05.pdf" in header_text
    assert "Санкт-Петербург" in header_text
    review = wb["Проверка"]
    issue_row = next(row[0].row for row in review.iter_rows() if row[0].value == "warning")
    assert review.cell(issue_row, 4).value.startswith("Длинное профессиональное замечание")
    assert f"D{issue_row}:M{issue_row}" in {str(value) for value in review.merged_cells.ranges}
    assert review.column_dimensions["A"].width >= 20


def test_endpoint_lsr_trace_export_returns_download():
    from proxy.routers.lsr import LsrTraceRequest, lsr_multi_trace_export

    res = asyncio.run(
        lsr_multi_trace_export(
            LsrTraceRequest(
                positions=[
                    {"code": CODE, "qty": 0.61, "section": "Раздел 1"},
                    {"code": CODE, "qty": 0.61, "section": "Раздел 2"},
                ],
                name="Тест",
            ),
            _user=object(),
        )
    )
    assert res["name"] == "Тест"
    assert res["summary"]["total"] == 2 * EXPECTED_TOTAL
    assert len(res["sections"]) == 2
    assert res["download"].startswith("/api/lsr/download?path=lsr_trace_")
    assert Path(res["path"]).exists()

from pathlib import Path

import pytest

import backend.converter as converter


def test_indexing_uses_isolated_converter_for_pdf_and_excel(monkeypatch):
    calls = []

    monkeypatch.setattr(
        converter,
        "convert_to_markdown_isolated",
        lambda path, route=None: calls.append(("isolated", path.suffix)) or "isolated",
    )
    monkeypatch.setattr(
        converter,
        "convert_to_markdown",
        lambda path, route=None: calls.append(("direct", path.suffix)) or "direct",
    )

    assert converter.convert_to_markdown_for_indexing(Path("doc.pdf")) == "isolated"
    assert converter.convert_to_markdown_for_indexing(Path("table.xlsx")) == "isolated"
    assert converter.convert_to_markdown_for_indexing(Path("note.md")) == "direct"
    assert calls == [("isolated", ".pdf"), ("isolated", ".xlsx"), ("direct", ".md")]


def test_isolated_converter_can_be_disabled(monkeypatch):
    monkeypatch.setenv("RAG_CONVERT_SUBPROCESS_ENABLED", "0")
    monkeypatch.setattr(converter, "convert_to_markdown_isolated", lambda *args, **kwargs: "isolated")
    monkeypatch.setattr(converter, "convert_to_markdown", lambda *args, **kwargs: "direct")

    assert converter.convert_to_markdown_for_indexing(Path("doc.pdf")) == "direct"


def test_large_pdf_indexing_uses_fast_text_first(tmp_path, monkeypatch):
    path = tmp_path / "project.pdf"
    path.write_bytes(b"%PDF placeholder")
    calls = []

    monkeypatch.setattr(converter, "_pdf_page_count", lambda _path: 83)
    monkeypatch.setattr(
        converter,
        "_parse_pdf_fast_text_layer",
        lambda p, reason="": calls.append(("fast", reason)) or "page text",
    )
    monkeypatch.setattr(
        converter,
        "convert_to_markdown_isolated",
        lambda *args, **kwargs: calls.append(("isolated", "")) or "isolated",
    )

    assert converter.convert_to_markdown_for_indexing(path) == "page text"
    assert calls == [("fast", "pdf_index_text_first")]


def test_pdf_timeout_falls_back_to_fast_text(tmp_path, monkeypatch):
    path = tmp_path / "project.pdf"
    path.write_bytes(b"%PDF placeholder")
    calls = []

    monkeypatch.setenv("RAG_PDF_INDEX_FAST_TEXT_FIRST", "0")
    monkeypatch.setattr(
        converter,
        "convert_to_markdown_isolated",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("convert subprocess timeout: >162s")),
    )
    monkeypatch.setattr(
        converter,
        "_parse_pdf_fast_text_layer",
        lambda p, reason="": calls.append(reason) or "fallback page text",
    )

    assert converter.convert_to_markdown_for_indexing(path) == "fallback page text"
    assert calls == ["isolated_convert_failed: convert subprocess timeout: >162s"]


def test_fast_pdf_text_does_not_run_global_boilerplate_regex(tmp_path, monkeypatch):
    path = tmp_path / "project.pdf"
    path.write_bytes(b"%PDF placeholder")

    class FakePage:
        def get_text(self, kind, sort=False):
            assert kind == "text"
            assert sort is False
            return "page text"

    class FakeDoc:
        page_count = 1

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return None

        def __iter__(self):
            return iter([FakePage()])

    monkeypatch.setattr(converter, "strip_legal_boilerplate", lambda _text: (_ for _ in ()).throw(AssertionError()))
    monkeypatch.setitem(__import__("sys").modules, "fitz", type("FakeFitz", (), {"open": lambda _path: FakeDoc()}))

    out = converter._parse_pdf_fast_text_layer(path)

    assert "## Page 1" in out
    assert "page text" in out


def test_spreadsheet_parser_runs_before_markitdown(tmp_path, monkeypatch):
    path = tmp_path / "table.xlsx"
    path.write_bytes(b"placeholder")
    calls = []

    monkeypatch.setattr(
        converter,
        "_parse_spreadsheet",
        lambda p: calls.append("spreadsheet") or "spreadsheet markdown",
    )
    monkeypatch.setattr(
        converter,
        "_parse_with_markitdown",
        lambda p: calls.append("markitdown") or "markitdown markdown",
    )

    assert converter.convert_to_markdown(path) == "spreadsheet markdown"
    assert calls == ["spreadsheet"]


def test_isolated_converter_returns_excel_markdown(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Qty"])
    sheet.append(["Cable", 12])
    path = tmp_path / "sample.xlsx"
    workbook.save(path)

    markdown = converter.convert_to_markdown_isolated(path, timeout_sec=30)

    assert markdown
    assert "Cable" in markdown
    assert "12" in markdown


def test_small_spreadsheet_keeps_full_table(tmp_path, monkeypatch):
    from openpyxl import Workbook

    monkeypatch.setenv("RAG_CONVERT_SUBPROCESS_ENABLED", "0")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "VOR"
    sheet.append(["Name", "Qty"])
    sheet.append(["Cable", 12])
    path = tmp_path / "small.xlsx"
    workbook.save(path)

    markdown = converter.convert_to_markdown(path)

    assert "Тип: spreadsheet_navigation_projection" not in markdown
    assert "| Name" in markdown
    assert "Cable" in markdown


def test_large_spreadsheet_uses_navigation_projection(tmp_path, monkeypatch):
    from openpyxl import Workbook

    monkeypatch.setenv("RAG_CONVERT_SUBPROCESS_ENABLED", "0")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Big"
    sheet.append(["Code", "Name", "Qty"])
    for idx in range(1, 2500):
        sheet.append([f"C-{idx:04d}", f"Position {idx}", idx])
    path = tmp_path / "large.xlsx"
    workbook.save(path)

    markdown = converter.convert_to_markdown(path)

    assert "Тип: spreadsheet_navigation_projection" in markdown
    assert "Профили колонок" in markdown
    assert "Code, Name, Qty" in markdown
    assert "C-0001" in markdown
    assert "Position 2499" not in markdown
    assert len(markdown) < 20_000


def test_isolated_converter_timeout_terminates_child(monkeypatch):
    created = {}

    class EmptyQueue:
        def get(self, timeout=None):
            raise converter.queue.Empty

        def get_nowait(self):
            raise converter.queue.Empty

    class HangingProcess:
        exitcode = None

        def __init__(self, **kwargs):
            self.alive = True
            self.terminated = False
            created["process"] = self

        def start(self):
            pass

        def join(self, timeout=None):
            pass

        def is_alive(self):
            return self.alive

        def terminate(self):
            self.terminated = True
            self.alive = False

        def kill(self):
            self.alive = False

    class FakeContext:
        def Queue(self, maxsize=1):
            return EmptyQueue()

        def Process(self, target, args, name):
            return HangingProcess(target=target, args=args, name=name)

    monkeypatch.setattr(converter.multiprocessing, "get_context", lambda method: FakeContext())

    with pytest.raises(RuntimeError, match="convert subprocess timeout"):
        converter.convert_to_markdown_isolated(Path("doc.pdf"), timeout_sec=1)

    assert created["process"].terminated is True


def test_isolated_converter_reads_queue_before_join(monkeypatch):
    class ResultQueue:
        def get(self, timeout=None):
            return ("ok", "x" * 300_000)

        def get_nowait(self):
            raise AssertionError("result should be read before process join")

    class AliveProcess:
        exitcode = None

        def __init__(self, **kwargs):
            self.alive = True
            self.joined = False

        def start(self):
            pass

        def join(self, timeout=None):
            self.joined = True
            self.alive = False

        def is_alive(self):
            return self.alive

        def terminate(self):
            raise AssertionError("process should not be terminated after result")

        def kill(self):
            raise AssertionError("process should not be killed after result")

    class FakeContext:
        def Queue(self, maxsize=1):
            return ResultQueue()

        def Process(self, target, args, name):
            return AliveProcess(target=target, args=args, name=name)

    monkeypatch.setattr(converter.multiprocessing, "get_context", lambda method: FakeContext())

    assert converter.convert_to_markdown_isolated(Path("doc.xlsx"), timeout_sec=1) == "x" * 300_000

"""Ц14, Ярус 3 — action-сервис: смета→документ (save) и запись→журнал (append).

Проверяем композицию (assemble→save), безопасность (create/append, не overwrite),
идемпотентность журнала и валидацию входа. 0 LLM, всё детерминированно.
"""

from __future__ import annotations

from pathlib import Path

import pytest

import proxy.services.field_intake_service as F
import proxy.services.project_service as P
from proxy.services import les_action_service as A


@pytest.fixture(autouse=True)
def _isolated(tmp_path, monkeypatch):
    # журнал/проекты — во временную метабазу; документы — во временный storage.
    db = str(tmp_path / "meta.db")
    monkeypatch.setattr(F, "rag_meta_db_path", lambda: db)
    monkeypatch.setattr(P, "rag_meta_db_path", lambda: db)
    monkeypatch.setenv("LES_PROJECTS_DIR", str(tmp_path / "projects"))


def _assembled() -> dict:
    """Минимальный выход les_lsr_assemble (positions + summary)."""
    return {
        "positions": [
            {"code": "08-01-001", "name": "Кладка стен", "unit": "м3", "qty": 12.5, "total": 84210.0},
            {"code": "", "name": "Монолитная плита", "unit": "м3", "qty": 30, "total": 150000.0},
        ],
        "summary": {"positions": 2, "total": 234210.0},
    }


# ── les_smeta_save: композиция assemble → save ──

def test_save_smeta_creates_vor_document():
    res = A.save_smeta(_assembled(), project_id=2, form_id="vor", fmt="xlsx",
                       doc_code="4-02-АС19", link=False)
    assert res["ok"] is True
    assert res["positions"] == 2
    assert res["total"] == 234210.0
    path = Path(res["path"])
    assert path.is_file()
    # документ лежит в storage проекта, расширение верное
    assert path.suffix == ".xlsx"
    assert "/2/smeta/" in str(path).replace("\\", "/")

    from openpyxl import load_workbook
    blob = " | ".join(
        str(c.value) for row in load_workbook(path).active.iter_rows()
        for c in row if c.value is not None
    )
    assert "Ведомость объёмов работ" in blob   # бланк ВОР
    assert "Кладка стен" in blob                # позиция сметы попала строкой
    assert "ИТОГО по смете" in blob             # строка итога
    assert "{{" not in blob                     # плейсхолдеры не утекли


def test_save_smeta_does_not_overwrite():
    a = A.save_smeta(_assembled(), project_id=5, link=False)
    b = A.save_smeta(_assembled(), project_id=5, link=False)
    # два сохранения → два разных файла (явность/без слепого overwrite)
    assert a["path"] != b["path"]
    assert Path(a["path"]).is_file() and Path(b["path"]).is_file()


def test_save_smeta_links_to_existing_project():
    proj = P.create_project("БЦ Банкрот", code="BANKROT")
    res = A.save_smeta(_assembled(), project_id=proj["id"], link=True)
    assert res["linked"] is True
    folders = [l["ref"] for l in P.list_links(proj["id"], kind="folder")]
    assert res["path"] in folders


def test_save_smeta_rejects_empty():
    with pytest.raises(ValueError):
        A.save_smeta({"positions": [], "summary": {}}, project_id=1, link=False)


def test_save_smeta_validates_form_and_project():
    with pytest.raises(ValueError):
        A.save_smeta(_assembled(), project_id=1, form_id="aosr", link=False)
    with pytest.raises(ValueError):
        A.save_smeta(_assembled(), project_id=0, link=False)


# ── les_journal_append: append + pending + идемпотентность ──

def test_journal_append_is_pending():
    res = A.journal_append("монолитная плита", 50, "м3", zahvatka="3", project_id=2)
    assert res["ok"] is True
    assert res["status"] == "pending"
    assert res["idempotent"] is False
    entry = F.get_entry(res["entry_id"])
    assert entry["status"] == "pending"           # ждёт подтверждения (как приёмка ИД)
    assert entry["volume"] == 50.0
    assert entry["project_id"] == 2


def test_journal_append_does_not_drop_existing():
    A.journal_append("кладка", 10, "м2", project_id=1)
    A.journal_append("стяжка", 20, "м2", project_id=1)
    # append, не overwrite — обе записи существуют
    assert len(F.list_entries(project_id=1)) == 2


def test_journal_append_idempotent():
    a = A.journal_append("плита", 100, "м3", project_id=1, idem_key="act-77")
    b = A.journal_append("плита", 100, "м3", project_id=1, idem_key="act-77")
    assert a["entry_id"] == b["entry_id"]          # дубль не создан
    assert b["idempotent"] is True
    assert len(F.list_entries(project_id=1)) == 1


def test_journal_append_validates_volume():
    with pytest.raises(ValueError):
        A.journal_append("плита", 0, "м3")
    with pytest.raises(ValueError):
        A.journal_append("", 5, "м3")

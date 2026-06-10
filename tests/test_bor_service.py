"""W11.1: генератор ВОР из спецификаций — офлайн-тесты, без LLM и живых сервисов."""

from pathlib import Path

import pytest

from backend.parquet_writer import save_parquet
from proxy.services.bor_service import (
    build_bor,
    collect_spec_rows,
    generate_bor,
    normalize_unit,
)


def _spec_row(**overrides) -> dict:
    row = {
        "doc_type": "SPEC",
        "doc_title": "Спецификация ОВ",
        "source_file": "spec_ov.xlsx",
        "pos": "1",
        "section": "Вентиляция",
        "name": "Воздуховод оцинкованный 200x200",
        "code": "В-200",
        "mark": "",
        "unit": "м",
        "qty": 10.0,
    }
    row.update(overrides)
    return row


# ── normalize_unit ──

def test_normalize_unit_aliases():
    assert normalize_unit("шт.") == "шт"
    assert normalize_unit("Штук") == "шт"
    assert normalize_unit("кв.м") == "м²"
    assert normalize_unit("м2") == "м²"
    assert normalize_unit("куб.м") == "м³"
    assert normalize_unit("п.м") == "пог.м"
    assert normalize_unit(None) == ""


def test_normalize_unit_unknown_passthrough():
    assert normalize_unit("рулон") == "рулон"


# ── build_bor ──

def test_build_bor_sums_identical_positions():
    rows = [
        _spec_row(qty=10.0, pos="1"),
        _spec_row(qty=5.5, pos="7", source_file="spec_ov2.xlsx"),
    ]
    lines = build_bor(rows)
    assert len(lines) == 1
    assert lines[0].qty == pytest.approx(15.5)
    assert lines[0].source_rows == 2
    assert "spec_ov.xlsx#1" in lines[0].sources
    assert "spec_ov2.xlsx#7" in lines[0].sources


def test_build_bor_groups_by_unit_after_normalization():
    rows = [
        _spec_row(unit="шт", name="Клапан", qty=2.0),
        _spec_row(unit="шт.", name="Клапан", qty=3.0),
    ]
    lines = build_bor(rows)
    assert len(lines) == 1
    assert lines[0].qty == pytest.approx(5.0)
    assert lines[0].unit == "шт"


def test_build_bor_separates_different_marks_and_sections():
    rows = [
        _spec_row(mark="М1", qty=1.0),
        _spec_row(mark="М2", qty=2.0),
        _spec_row(section="Отопление", qty=3.0),
    ]
    assert len(build_bor(rows)) == 3


def test_build_bor_name_whitespace_insensitive():
    rows = [
        _spec_row(name="Воздуховод  оцинкованный   200x200", qty=1.0),
        _spec_row(name="воздуховод оцинкованный 200x200", qty=2.0),
    ]
    lines = build_bor(rows)
    assert len(lines) == 1
    assert lines[0].qty == pytest.approx(3.0)


def test_build_bor_qty_missing_tracked():
    rows = [_spec_row(qty=None), _spec_row(qty=4.0)]
    lines = build_bor(rows)
    assert len(lines) == 1
    assert lines[0].qty == pytest.approx(4.0)
    assert lines[0].qty_missing_rows == 1


def test_build_bor_all_qty_missing_gives_none():
    lines = build_bor([_spec_row(qty=None)])
    assert lines[0].qty is None
    assert lines[0].qty_missing_rows == 1


def test_build_bor_sorted_by_section_then_name():
    rows = [
        _spec_row(section="Отопление", name="Радиатор", qty=1.0),
        _spec_row(section="Вентиляция", name="Клапан", qty=1.0),
        _spec_row(section="Вентиляция", name="Воздуховод", qty=1.0),
    ]
    lines = build_bor(rows)
    assert [(l.section, l.name) for l in lines] == [
        ("Вентиляция", "Воздуховод"),
        ("Вентиляция", "Клапан"),
        ("Отопление", "Радиатор"),
    ]


# ── parquet → ВОР → xlsx (полный цикл на синтетике) ──

def _make_dataset(tmp_path: Path, dataset_id: str, rows: list[dict]) -> Path:
    parquet_dir = tmp_path / dataset_id / "_parquet"
    parquet_dir.mkdir(parents=True)
    save_parquet(rows, str(parquet_dir / "spec.parquet"))
    return tmp_path


def test_collect_spec_rows_filters_doc_type_and_empty_names(tmp_path):
    rows = [
        _spec_row(),
        _spec_row(doc_type="SMETA", name="Не должна попасть"),
        _spec_row(name=""),
        _spec_row(doc_type="VEDOMOST", name="Ведомостная позиция"),
    ]
    storage = _make_dataset(tmp_path, "ds1", rows)
    collected = collect_spec_rows("ds1", storage_root=storage)
    names = {r["name"] for r in collected}
    assert names == {"Воздуховод оцинкованный 200x200", "Ведомостная позиция"}


def test_generate_bor_end_to_end_xlsx(tmp_path):
    rows = [_spec_row(qty=10.0), _spec_row(qty=2.5, pos="3")]
    storage = _make_dataset(tmp_path, "ds2", rows)
    out_dir = tmp_path / "out"
    result = generate_bor("ds2", storage_root=storage, output_dir=out_dir)
    assert result["bor_lines"] == 1
    assert result["lines"][0]["qty"] == pytest.approx(12.5)
    xlsx = Path(result["xlsx_path"])
    assert xlsx.exists()

    from openpyxl import load_workbook

    ws = load_workbook(xlsx).active
    data_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    assert data_row[6] == pytest.approx(12.5)  # Кол-во сходится бит-в-бит


def test_generate_bor_empty_dataset(tmp_path):
    result = generate_bor("missing", storage_root=tmp_path, output_dir=tmp_path / "out")
    assert result["bor_lines"] == 0
    assert result["xlsx_path"] is None


def test_bor_service_uses_no_llm():
    """ADR-11: модуль ВОР не импортирует HTTP/LLM-клиентов и не зовёт чат-эндпоинты."""
    import inspect

    import proxy.services.bor_service as bor

    source = inspect.getsource(bor)
    for marker in ("import httpx", "import openai", "import requests", "/api/chat", "completions"):
        assert marker not in source, f"LLM/HTTP-маркер '{marker}' в bor_service"

"""Filled KS-2 / KS-3 / KS-6а mappers and chat intent (0 LLM)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from proxy.services import field_intake_service as F
from proxy.services import ks_forms_chat_service as chat
from proxy.services import ks_forms_service as K
from proxy.services import les_action_service as A


@pytest.fixture(autouse=True)
def _isolated(tmp_path, monkeypatch):
    db = str(tmp_path / "meta.db")
    monkeypatch.setattr(F, "rag_meta_db_path", lambda: db)
    monkeypatch.setenv("LES_PROJECTS_DIR", str(tmp_path / "projects"))
    monkeypatch.setenv("LES_FORMS_OUT_DIR", str(tmp_path / "forms_out"))
    monkeypatch.setenv("LES_FORMS_DIR", "config/forms")
    # Default disk fallback would otherwise read real storage/smeta_artifacts.
    empty_artifacts = tmp_path / "smeta_artifacts"
    empty_artifacts.mkdir()
    _orig_disk = K.load_latest_assembled_from_disk

    def _disk(artifact_dir: str | Path = "storage/smeta_artifacts"):
        root = Path(artifact_dir)
        if root == Path("storage/smeta_artifacts"):
            return _orig_disk(empty_artifacts)
        return _orig_disk(root)

    monkeypatch.setattr(K, "load_latest_assembled_from_disk", _disk)


def _assembled():
    return {
        "positions": [
            {"code": "ГЭСН08-01-001-01", "name": "Кладка стен", "unit": "м3", "qty": 12.5, "total": 84210.0},
            {"code": "ГЭСН06-01-001-01", "name": "Монолит", "unit": "м3", "qty": 30, "total": 150000.0},
        ],
        "summary": {"positions": 2, "total": 234210.0},
    }


def _rim():
    return {
        "rows": [
            {"basis": "ГЭСН08-01-001-01", "title": "Кладка стен", "unit": "м3", "quantity": 12.5, "amount": 84210.0},
            {"basis": "ГЭСН06-01-001-01", "title": "Монолит", "unit": "м3", "quantity": 30, "amount": 150000.0},
        ],
        "amount_total": 234210.0,
    }


def test_assembled_from_rim_form():
    assembled = K.assembled_from_rim_form(_rim())
    assert len(assembled["positions"]) == 2
    assert assembled["summary"]["total"] == pytest.approx(234210.0)
    assert assembled["positions"][0]["code"].startswith("ГЭСН")


def test_assembled_from_document_rim_trace():
    """Document-workflow LSR stores sections under rim_trace — KS must read it."""
    artifact = {
        "mode": "xlsx",
        "title": "ЛСР — vor.pdf",
        "rim_trace": {
            "sections": [
                {
                    "section": "Раздел 1",
                    "positions": [
                        {
                            "code": "ГЭСН08-01-001-01",
                            "name": "Кладка стен",
                            "unit": "м3",
                            "qty": 12.5,
                            "summary": {"total": 84210.0},
                        },
                        {
                            "code": "ГЭСН06-01-001-01",
                            "name": "Монолит",
                            "unit": "м3",
                            "qty": 30,
                            "summary": {"total": 150000.0},
                        },
                    ],
                }
            ],
            "summary": {"total": 234210.0, "bound_rows": 2, "input_rows": 2},
        },
    }
    assembled = K.assembled_from_artifact(artifact)
    assert assembled is not None
    assert len(assembled["positions"]) == 2
    assert assembled["positions"][0]["name"] == "Кладка стен"
    assert assembled["summary"]["total"] == pytest.approx(234210.0)
    rows = K.ks2_rows(assembled)
    assert rows[0][2] == "Кладка стен"
    assert rows[-1][-1] == "234210.00"


def test_assembled_from_workflow_report_lsr_key(tmp_path):
    """Disk fallback reads workflow report JSON (lsr key), not only rim_lsr_form."""
    report = {
        "schema": "smeta_document_workflow_v2",
        "lsr": {
            "sections": [
                {
                    "section": "A",
                    "positions": [
                        {
                            "code": "ГЭСН01-01-001-01",
                            "name": "Земляные",
                            "unit": "м3",
                            "qty": 10,
                            "summary": {"total": 5000.0},
                        }
                    ],
                }
            ],
            "summary": {"total": 5000.0},
        },
    }
    path = tmp_path / "LSR_test.json"
    path.write_text(__import__("json").dumps(report, ensure_ascii=False), encoding="utf-8")
    loaded = K.load_latest_assembled_from_disk(tmp_path)
    assert loaded is not None
    assert loaded["positions"][0]["name"] == "Земляные"
    assert loaded["summary"]["total"] == pytest.approx(5000.0)


def test_ks2_rows_columns_and_total():
    rows = K.ks2_rows(_assembled())
    # Official 8 cols: № | № сметы | name | code | unit | qty | price | total
    assert rows[0][2] == "Кладка стен"
    assert rows[0][3] == "ГЭСН08-01-001-01"
    assert rows[0][5] == "12.5"
    assert rows[0][6] == "6736.80"  # 84210/12.5
    assert rows[-1][2] == "Итого"
    assert rows[-1][-1] == "234210.00"


def test_ks2_xlsx_official_layout(tmp_path):
    from openpyxl import load_workbook

    out = tmp_path / "ks2_official.xlsx"
    res = K.build_ks_document(
        "ks2", project_id=1, assembled=_assembled(), out_path=out, source="last_lsr",
    )
    assert Path(res["path"]).suffix == ".xlsx"
    ws = load_workbook(res["path"]).active
    blob = " | ".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "Унифицированная форма № КС-2" in blob
    assert "0322005" in blob
    assert "ЧЕРНОВИК ИЗ ЛСР" in blob
    assert "НЕ ПОДТВЕРЖДАЕТ ФАКТИЧЕСКОЕ ВЫПОЛНЕНИЕ" in blob
    assert "Номер позиции по смете" in blob
    assert "Кладка стен" in blob
    assert "Выполнено работ" in blob


def test_chat_ks2_filename_has_xlsx_extension():
    res = chat.answer_ks_forms_query("собери КС-2", project_id=1, assembled=_assembled())
    assert res["ok"] is True
    assert str(res["filename"]).endswith(".xlsx")
    assert str(res["command"]["filename"]).endswith(".xlsx")
    assert not str(res["command"].get("title", "")).endswith(".xlsx")  # title stays human


def test_ks3_rows_period_totals():
    rows = K.ks3_rows(_assembled())
    assert len(rows) == 1
    assert rows[0][2] == ""
    assert rows[0][3] == ""
    assert rows[0][4] == "234210.00"


def test_build_ks2_xlsx_non_blank(tmp_path):
    out = tmp_path / "ks2.xlsx"
    res = K.build_ks_document(
        "ks2", project_id=1, assembled=_assembled(), out_path=out, source="last_lsr",
    )
    assert res["filled"] is True
    assert Path(res["path"]).is_file()
    blob = " | ".join(
        str(c.value) for row in load_workbook(res["path"]).active.iter_rows()
        for c in row if c.value is not None
    )
    assert "Кладка стен" in blob
    assert "234210.00" in blob or "234210" in blob


def test_ks6a_requires_confirmed_journal():
    with pytest.raises(ValueError, match="Журнал"):
        K.build_ks_document("ks6a", source="field_journal", project_id=1)


def test_filled_forms_never_fall_back_to_another_project(monkeypatch):
    def _must_not_read_shared_artifacts(*args, **kwargs):
        raise AssertionError("shared artifact fallback must not run")

    monkeypatch.setattr(K, "load_latest_assembled_from_disk", _must_not_read_shared_artifacts)
    with pytest.raises(ValueError, match="project_id"):
        K.resolve_assembled()
    assert K.resolve_assembled(project_id=7) is None


def test_ks6a_requires_project_scope_even_with_existing_journal():
    F.create_entry("Кабель", 10, "м", entry_date="2026-07-01", status="confirmed", project_id=1)
    with pytest.raises(ValueError, match="project_id"):
        K.load_confirmed_journal()


def test_ks6a_from_confirmed_entries(tmp_path):
    F.create_entry("Кабель", 10, "м", entry_date="2026-07-01", status="confirmed", project_id=1)
    F.create_entry("Кабель", 5, "м", entry_date="2026-07-02", status="confirmed", project_id=1)
    F.create_entry("Лоток", 2, "м", entry_date="2026-07-01", status="pending", project_id=1)
    out = tmp_path / "ks6a.xlsx"
    res = K.build_ks_document("ks6a", project_id=1, out_path=out, source="field_journal")
    assert res["entries"] == 2
    blob = " | ".join(
        str(c.value) for row in load_workbook(res["path"]).active.iter_rows()
        for c in row if c.value is not None
    )
    assert "Кабель" in blob
    assert "15" in blob  # cumulative 10+5
    assert "Лоток" not in blob


def test_save_smeta_ks2(tmp_path, monkeypatch):
    monkeypatch.setenv("LES_PROJECTS_DIR", str(tmp_path / "projects"))
    res = A.save_smeta(_assembled(), project_id=3, form_id="ks2", fmt="xlsx", link=False)
    assert res["ok"] is True
    blob = " | ".join(
        str(c.value) for row in load_workbook(res["path"]).active.iter_rows()
        for c in row if c.value is not None
    )
    assert "Кладка стен" in blob


def test_chat_detect_and_clarify():
    assert chat.detect_ks_form("собери КС-2 из сметы") == "ks2"
    assert chat.detect_ks_form("сформируй кс-3") == "ks3"
    assert chat.detect_ks_form("выгрузи КС-6а") == "ks6a"
    assert chat.detect_ks_form("сверь вор и кс-2") is None
    clarified = chat.answer_ks_forms_query("собери кс-2", session_id="")
    assert clarified["clarify"] is True
    assert "project_id" in clarified["answer"]


def test_chat_builds_ks2_from_assembled():
    res = chat.answer_ks_forms_query("собери КС-2", project_id=1, assembled=_assembled())
    assert res["ok"] is True
    assert res["command"]["action"] == "generate_filled_form"
    assert Path(res["path"]).is_file()


def test_ks2_xlsx_escapes_formula_like_project_text(tmp_path):
    assembled = _assembled()
    assembled["positions"][0]["name"] = '=HYPERLINK("https://example.invalid","x")'
    out = tmp_path / "ks2_safe.xlsx"
    result = K.build_ks_document(
        "ks2", project_id=1, assembled=assembled, out_path=out, source="last_lsr",
    )
    ws = load_workbook(result["path"], data_only=False).active
    cells = [cell for row in ws.iter_rows() for cell in row]
    injected = next(cell for cell in cells if "HYPERLINK" in str(cell.value or ""))
    assert str(injected.value).startswith("'=")
    assert injected.data_type != "f"


def test_filled_command_does_not_claim_to_be_blank():
    from proxy.services.command_service import handle_command

    response = handle_command("/кс-2", project_id=1)
    assert response is not None
    assert response["command"]["action"] == "generate_filled_form"
    assert "ПУСТОЙ" not in response["answer"]


def test_mcp_filled_ks_preserves_draft_status():
    from tools.les_mcp_server import les_form_generate

    result = les_form_generate(
        "ks2",
        project_id=1,
        source="last_lsr",
        assembled=_assembled(),
    )
    assert result["ok"] is True
    assert result["filled"] is True
    assert result["draft"] is True
    assert result["document_status"] == "draft_from_lsr_not_execution_fact"


def test_natural_language_ks_request_has_no_pre_model_hook():
    """Only an explicit slash command may build a form before the model."""
    import inspect

    from proxy.routers import chat as chat_router

    source = inspect.getsource(chat_router._run_chat)
    assert "is_ks_forms_query(req.question)" not in source
    assert "run_chat_evidence_application(" in source

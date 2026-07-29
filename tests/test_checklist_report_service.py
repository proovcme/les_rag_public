"""Тесты checklist_report_service (T4.2): XLSX/HTML отчёт по checklist_review_v1, TDD red->green.

По образцу doc_review_service.review_to_xlsx (:710) / review_to_html (:696) — openpyxl-паттерн
(лист + шапка + жирный заголовок + normalized_remarks лист), но контракт входа — dict
``checklist_review_v1`` (как из ``run_checklist_review``/``run.json`` персиста), не
``list[ReviewItem]``. 0 LLM — модуль только рендерит уже посчитанный результат.

Формат отчёта — implementation_plan.md §5 + docs/CHECKLIST_REVIEW_PD_TASK.md §8:
  XLSX лист «Чек-лист» = форма исходника (Раздел/№/Критерий) + системные колонки
    (Предложенный ответ рус.ярлык / Уверенность / Evidence / Примечание системы /
    Ответ инженера / Примечание инженера), сгруппировано по discipline с строкой-заголовком
    раздела; лист «Сводка»; лист «normalized_remarks».
  HTML — валидный html с summary-таблицей и всеми items, computed_issue строки визуально
  помечены css-классом.

Никакого общего вердикта «соответствует/не соответствует» ни в XLSX, ни в HTML — правило 10
canonical doc (§7): полного автоматического вердикта нет by design.
"""

from __future__ import annotations

import re

import pytest
from openpyxl import load_workbook

from proxy.services.checklist_report_service import report_to_html, report_to_xlsx

# ── фикстура: run_payload (форма checklist_review_v1, как из run_checklist_review) ─────────


def _run_payload(items=None, normalized_remarks=None, summary=None) -> dict:
    default_items = [
        {
            "item_id": "PD-OB-001",
            "sheet_name": "Общее",
            "row": 1,
            "item_no": "1",
            "section_path": [],
            "criterion": "Приложен отчет об инженерно-геологических изысканиях",
            "kind": "presence",
            "discipline": "Общее",
            "allowed_answers": ["Да", "Нет"],
            "status": "supported_by_evidence",
            "suggested_answer": "yes",
            "confidence": 0.75,
            "requirement_refs": [],
            "document_evidence": [
                {
                    "kind": "document",
                    "source_ref": "ds-pd:otchet_igi.pdf#page=1",
                    "snippet": "Отчет об инженерно-геологических изысканиях выполнен в 2026 году.",
                    "value": "", "unit": "", "bbox": None,
                    "reason": "контентный хит подтверждает наличие документа/раздела",
                },
            ],
            "computed_check": {"name": "presence", "status": "ok"},
            "model_note": "Найден контентный хит, подтверждающий наличие документа.",
            "human_decision": "unset",
            "human_answer": "unset",
            "human_comment": "",
            "human_note": "",
            "doc_review_item_ids": [],
            "formal_check_ids": [],
            "normalized_remark_ids": ["REM-CHK-PD-OB-001"],
        },
        {
            "item_id": "PD-KR-020",
            "sheet_name": "КР",
            "row": 20,
            "item_no": "6",
            "section_path": ["Конструктивные решения"],
            "criterion": "Марка бетона соответствует требованиям по водонепроницаемости для ростверка",
            "kind": "parametric",
            "discipline": "КР",
            "allowed_answers": ["Да", "Нет"],
            "status": "computed_issue",
            "suggested_answer": "no",
            "confidence": 0.7,
            "requirement_refs": [],
            "document_evidence": [
                {
                    "kind": "document",
                    "source_ref": "ds-pd:kr_zapiska.pdf#page=4",
                    "snippet": "Марка бетона ростверка W6.",
                    "value": "W6", "unit": "", "bbox": None,
                    "reason": "извлечено значение параметра «марка водонепроницаемости» из содержимого документа",
                },
            ],
            "computed_check": {"name": "concrete_w_rostverk", "status": "issue"},
            "model_note": "Марка W6 не удовлетворяет требованию >= W8.",
            "human_decision": "unset",
            "human_answer": "unset",
            "human_comment": "",
            "human_note": "",
            "doc_review_item_ids": [],
            "formal_check_ids": [],
            "normalized_remark_ids": ["REM-CHK-PD-KR-020"],
        },
        {
            "item_id": "PD-AR-049",
            "sheet_name": "АР",
            "row": 49,
            "item_no": "8",
            "section_path": ["Архитектурные решения"],
            "criterion": "Кухни расположены не над жилыми комнатами",
            "kind": "spatial_visual",
            "discipline": "АР",
            "allowed_answers": ["Да", "Нет"],
            "status": "manual_required",
            "suggested_answer": "manual_required",
            "confidence": 0.0,
            "requirement_refs": [],
            "document_evidence": [],
            "computed_check": {"name": "spatial_visual", "status": "not_run"},
            "model_note": "Пространственная/визуальная проверка — честная граница v1, требуется инженер.",
            "human_decision": "human_decided",
            "human_answer": "yes",
            "human_comment": "",
            "human_note": "Проверено инженером по плану этажа.",
            "doc_review_item_ids": [],
            "formal_check_ids": [],
            "normalized_remark_ids": ["REM-CHK-PD-AR-049"],
        },
        {
            "item_id": "PD-SPOZU-010",
            "sheet_name": "СПОЗУ",
            "row": 10,
            "item_no": "7",
            "section_path": [],
            "criterion": "Соответствие пирогов благоустройства ТЗ и техническому стандарту Glorax",
            "kind": "cross_section",
            "discipline": "СПОЗУ",
            "allowed_answers": ["Да", "Нет"],
            "status": "review_needed",
            "suggested_answer": "unknown",
            "confidence": 0.1,
            "requirement_refs": [],
            "document_evidence": [],
            "computed_check": {"name": "cross_section", "status": "not_run"},
            "model_note": "Исходники (ТЗ/ОПР/АГО/СТУ) не выбраны для проверки.",
            "human_decision": "unset",
            "human_answer": "unset",
            "human_comment": "",
            "human_note": "",
            "doc_review_item_ids": [],
            "formal_check_ids": [],
            "normalized_remark_ids": [],
        },
    ]
    default_remarks = [
        {
            "schema": "normalized_remark_v1",
            "id": "REM-CHK-PD-OB-001",
            "source": "checklist",
            "category": "checklist",
            "severity": "info",
            "status": "supported_by_evidence",
            "target": "Приложен отчет об инженерно-геологических изысканиях",
            "clause": "1",
            "requirement_ref": "",
            "document_refs": [],
            "source_refs": ["ds-pd:otchet_igi.pdf#page=1"],
            "checklist_ref": {"template": "mini_report_template", "item_id": "PD-OB-001"},
            "computed_check": {"name": "presence", "status": "ok"},
            "message": "Найден контентный хит, подтверждающий наличие документа.",
            "human_decision": "unset",
            "human_comment": "",
            "human_decided_at": "",
            "finality": "proposed",
            "requires_human": True,
            "confidence": 0.75,
        },
        {
            "schema": "normalized_remark_v1",
            "id": "REM-CHK-PD-KR-020",
            "source": "checklist",
            "category": "checklist",
            "severity": "info",
            "status": "computed_issue",
            "target": "Марка бетона соответствует требованиям по водонепроницаемости для ростверка",
            "clause": "6",
            "requirement_ref": "concrete_w_rostverk",
            "document_refs": [],
            "source_refs": ["ds-pd:kr_zapiska.pdf#page=4"],
            "checklist_ref": {"template": "mini_report_template", "item_id": "PD-KR-020"},
            "computed_check": {"name": "concrete_w_rostverk", "status": "issue"},
            "message": "Марка W6 не удовлетворяет требованию >= W8.",
            "human_decision": "unset",
            "human_comment": "",
            "human_decided_at": "",
            "finality": "proposed",
            "requires_human": True,
            "confidence": 0.7,
        },
        {
            "schema": "normalized_remark_v1",
            "id": "REM-CHK-PD-AR-049",
            "source": "checklist",
            "category": "checklist",
            "severity": "info",
            "status": "manual_required",
            "target": "Кухни расположены не над жилыми комнатами",
            "clause": "8",
            "requirement_ref": "",
            "document_refs": [],
            "source_refs": [],
            "checklist_ref": {"template": "mini_report_template", "item_id": "PD-AR-049"},
            "computed_check": {"name": "spatial_visual", "status": "not_run"},
            "message": "Пространственная/визуальная проверка — честная граница v1, требуется инженер.",
            "human_decision": "human_decided",
            "human_comment": "Проверено инженером по плану этажа.",
            "human_decided_at": "2026-07-04T12:00:00+00:00",
            "finality": "human_decided",
            "requires_human": True,
            "confidence": 0.0,
        },
    ]
    default_summary = {
        "total": 4,
        "by_status": {
            "supported_by_evidence": 1,
            "computed_issue": 1,
            "manual_required": 1,
            "review_needed": 1,
        },
        "by_kind": {"presence": 1, "parametric": 1, "spatial_visual": 1, "cross_section": 1},
        "by_discipline": {"Общее": 1, "КР": 1, "АР": 1, "СПОЗУ": 1},
        "suggested": {"yes": 1, "no": 1, "not_required": 0, "manual_required": 1, "unknown": 1},
        "source_backed": 2,
        "without_evidence": 2,
        "human_final_required": True,
        "human": {"yes": 1, "no": 0, "not_required": 0, "unset": 3},
    }
    return {
        "schema": "checklist_review_v1",
        "run_id": "run-test-001",
        "dataset_id": "ds-pd",
        "source_dataset_ids": [],
        "template": "mini_report_template",
        "stage": "PD",
        "discipline": "all",
        "status": "done",
        "summary": summary if summary is not None else default_summary,
        "items": items if items is not None else default_items,
        "normalized_remarks": normalized_remarks if normalized_remarks is not None else default_remarks,
        "defense": {"schema": "defense_contract_v1", "claims": [], "human_final_required": True},
        "workflow_plan": {
            "schema": "workflow_plan_v1", "workflow_id": "checklist_review",
            "finality": "human_required", "blockers": [], "missing_inputs": [], "next_actions": [],
        },
    }


_ANSWER_LABELS = {
    "yes": "Да",
    "no": "Нет",
    "not_required": "Не требуется",
    "unknown": "—",
    "manual_required": "Требует инженера",
}


# ── report_to_xlsx: базовая структура книги ─────────────────────────────────────────────


def test_report_to_xlsx_creates_valid_workbook(tmp_path):
    payload = _run_payload()
    out_path = tmp_path / "report.xlsx"

    report_to_xlsx(payload, out_path)

    assert out_path.exists()
    wb = load_workbook(out_path)
    assert "Чек-лист" in wb.sheetnames
    assert "Сводка" in wb.sheetnames
    assert "normalized_remarks" in wb.sheetnames


def test_report_to_xlsx_checklist_sheet_has_system_columns_header():
    payload = _run_payload()
    import io

    from openpyxl import Workbook  # noqa: F401 — sanity, ensures dependency importable

    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory() as d:
        out_path = Path(d) / "report.xlsx"
        report_to_xlsx(payload, out_path)
        wb = load_workbook(out_path)
        ws = wb["Чек-лист"]

        header_row = None
        header_values = None
        for row in ws.iter_rows(min_row=1, max_row=5):
            values = [c.value for c in row]
            if values and values[0] == "ID пункта":
                header_row = row[0].row
                header_values = values
                break
        assert header_row is not None, "не найдена строка шапки с 'ID пункта'"
        expected = [
            "ID пункта", "№", "Критерий", "Предложенный ответ", "Уверенность",
            "Evidence", "Примечание системы", "Ответ инженера", "Примечание инженера",
        ]
        assert header_values[: len(expected)] == expected


def test_report_to_xlsx_checklist_sheet_groups_by_discipline_with_header_rows(tmp_path):
    payload = _run_payload()
    out_path = tmp_path / "report.xlsx"
    report_to_xlsx(payload, out_path)

    wb = load_workbook(out_path)
    ws = wb["Чек-лист"]
    all_values = [
        [c.value for c in row]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
    ]
    # раздел-заголовки для каждой дисциплины, представленной в items
    disciplines_in_items = {"Общее", "КР", "АР", "СПОЗУ"}
    section_header_texts = {
        row[0] for row in all_values
        if row and isinstance(row[0], str) and row[0] in disciplines_in_items
    }
    assert disciplines_in_items <= section_header_texts


def test_report_to_xlsx_russian_answer_labels_mapping(tmp_path):
    payload = _run_payload()
    out_path = tmp_path / "report.xlsx"
    report_to_xlsx(payload, out_path)

    wb = load_workbook(out_path)
    ws = wb["Чек-лист"]
    rows_by_item = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = [c.value for c in row]
        if values and values[0] in {"PD-OB-001", "PD-KR-020", "PD-AR-049", "PD-SPOZU-010"}:
            rows_by_item[values[0]] = values

    # ищем по критерию, если item_id не в первой колонке — пробуем найти строку по критерию
    if not rows_by_item:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            values = [c.value for c in row]
            for it in payload["items"]:
                if values and it["criterion"] in values:
                    rows_by_item[it["item_id"]] = values

    assert rows_by_item, "не удалось найти строки items в листе Чек-лист"
    yes_row = rows_by_item["PD-OB-001"]
    no_row = rows_by_item["PD-KR-020"]
    unknown_row = rows_by_item["PD-SPOZU-010"]
    assert "Да" in yes_row
    assert "Нет" in no_row
    assert "—" in unknown_row


def test_report_to_xlsx_item_with_human_answer_shown_in_engineer_column(tmp_path):
    payload = _run_payload()
    out_path = tmp_path / "report.xlsx"
    report_to_xlsx(payload, out_path)

    wb = load_workbook(out_path)
    ws = wb["Чек-лист"]
    found_yes_in_engineer_col = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = [c.value for c in row]
        if values and "Кухни расположены не над жилыми комнатами" in values:
            # human_answer=yes -> где-то в строке должна быть русская метка "Да" помимо
            # предложенного system-ответа (который тут manual_required -> "Требует инженера")
            assert "Требует инженера" in values
            assert values.count("Да") >= 1
            found_yes_in_engineer_col = True
    assert found_yes_in_engineer_col


def test_report_to_xlsx_evidence_joins_source_refs_with_newline_max_5():
    payload = _run_payload()
    many_refs_item = {
        "item_id": "PD-OB-999",
        "sheet_name": "Общее",
        "row": 99,
        "item_no": "99",
        "section_path": [],
        "criterion": "Критерий с многими evidence",
        "kind": "presence",
        "discipline": "Общее",
        "allowed_answers": ["Да", "Нет"],
        "status": "supported_by_evidence",
        "suggested_answer": "yes",
        "confidence": 0.75,
        "requirement_refs": [],
        "document_evidence": [
            {"kind": "document", "source_ref": f"ds-pd:doc{i}.pdf#page={i}",
             "snippet": f"snippet {i}", "value": "", "unit": "", "bbox": None, "reason": ""}
            for i in range(1, 8)  # 7 refs -> max 5 in report
        ],
        "computed_check": {"name": "presence", "status": "ok"},
        "model_note": "",
        "human_decision": "unset",
        "human_answer": "unset",
        "human_comment": "",
        "human_note": "",
        "doc_review_item_ids": [],
        "formal_check_ids": [],
        "normalized_remark_ids": [],
    }
    payload["items"] = [many_refs_item]

    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory() as d:
        out_path = Path(d) / "report.xlsx"
        report_to_xlsx(payload, out_path)
        wb = load_workbook(out_path)
        ws = wb["Чек-лист"]
        evidence_cell_value = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            values = [c.value for c in row]
            if values and "Критерий с многими evidence" in values:
                idx = values.index("Критерий с многими evidence")
                evidence_cell_value = values[idx + 1] if False else None
                # находим evidence-колонку по заголовку
                header = [c.value for c in ws[1]]
                if "Evidence" not in header:
                    # заголовок может быть на другой строке (после раздела) — ищем по всем строкам
                    for hrow in ws.iter_rows(min_row=1, max_row=ws.max_row):
                        hvals = [c.value for c in hrow]
                        if "Evidence" in hvals:
                            header = hvals
                            break
                ev_idx = header.index("Evidence")
                evidence_cell_value = values[ev_idx]
                break
        assert evidence_cell_value is not None
        refs_in_cell = [line for line in evidence_cell_value.split("\n") if line.strip()]
        assert len(refs_in_cell) <= 5
        assert "\n" in evidence_cell_value


# ── report_to_xlsx: лист «Сводка» ────────────────────────────────────────────────────────


def test_report_to_xlsx_summary_sheet_contains_key_counts(tmp_path):
    payload = _run_payload()
    out_path = tmp_path / "report.xlsx"
    report_to_xlsx(payload, out_path)

    wb = load_workbook(out_path)
    ws = wb["Сводка"]
    flat_text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "4" in flat_text  # total
    assert "computed_issue" in flat_text or "Замечание" in flat_text
    assert "without_evidence" in flat_text or "без evidence" in flat_text.lower() or "Evidence" in flat_text


def test_report_to_xlsx_summary_lists_open_manual_required_item_ids(tmp_path):
    # PD-AR-049 manual_required, но human_decision=human_decided -> закрыт, не должен попасть
    # в список "незакрытые manual_required". Добавим второй manual_required без решения.
    payload = _run_payload()
    open_manual_item = {
        "item_id": "PD-EN-777",
        "sheet_name": "ЭН",
        "row": 77,
        "item_no": "77",
        "section_path": [],
        "criterion": "Экспертный критерий без решения инженера",
        "kind": "manual_required",
        "discipline": "ЭН",
        "allowed_answers": ["Да", "Нет"],
        "status": "manual_required",
        "suggested_answer": "manual_required",
        "confidence": 0.0,
        "requirement_refs": [],
        "document_evidence": [],
        "computed_check": {"name": "manual_required", "status": "not_run"},
        "model_note": "Пункт требует инженерного решения, не формализуется.",
        "human_decision": "unset",
        "human_answer": "unset",
        "human_comment": "",
        "human_note": "",
        "doc_review_item_ids": [],
        "formal_check_ids": [],
        "normalized_remark_ids": [],
    }
    payload["items"] = payload["items"] + [open_manual_item]

    out_path = tmp_path / "report.xlsx"
    report_to_xlsx(payload, out_path)

    wb = load_workbook(out_path)
    ws = wb["Сводка"]
    flat_text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "PD-EN-777" in flat_text
    assert "PD-AR-049" not in flat_text.split("PD-EN-777")[0][-200:] or True  # см. проверку ниже

    # Более строгая проверка: собираем именно ячейки колонки со списком item_id.
    unresolved_cells = [
        str(c.value) for row in ws.iter_rows() for c in row
        if c.value and "PD-EN-777" in str(c.value)
    ]
    assert unresolved_cells
    for cell_text in unresolved_cells:
        assert "PD-AR-049" not in cell_text


# ── report_to_xlsx: лист «normalized_remarks» ───────────────────────────────────────────


def test_report_to_xlsx_normalized_remarks_sheet_has_expected_columns_and_rows(tmp_path):
    payload = _run_payload()
    out_path = tmp_path / "report.xlsx"
    report_to_xlsx(payload, out_path)

    wb = load_workbook(out_path)
    ws = wb["normalized_remarks"]
    header = [c.value for c in ws[1]]
    for col in ("id", "status", "severity", "source_refs", "message", "human_decision"):
        assert col in header

    ids_col = header.index("id")
    ids_in_sheet = {row[ids_col].value for row in ws.iter_rows(min_row=2) if row[ids_col].value}
    expected_ids = {r["id"] for r in payload["normalized_remarks"]}
    assert expected_ids <= ids_in_sheet


# ── report_to_xlsx: никакого общего вердикта ────────────────────────────────────────────

_VERDICT_PATTERNS = [
    # Запрещённые формулировки ОБЩЕГО вердикта по всему комплекту/прогону (implementation_plan.md
    # §7 правило 10) — не путать с легитимным текстом отдельных критериев чек-листа, где обороты
    # вида «марка бетона соответствует требованиям...» (конкретному пункту) — нормальная формулировка
    # исходного Glorax-чек-листа, не вердикт движка. Паттерны нарочно завязаны на слова уровня
    # документа/комплекта/проекта в целом, а не на само слово «соответствует».
    re.compile(r"(документация|комплект|проект)\w*\s+(в\s+целом\s+)?(не\s+)?соответствует", re.IGNORECASE),
    re.compile(r"итоговый\s+вердикт", re.IGNORECASE),
    re.compile(r"общий\s+вывод", re.IGNORECASE),
    re.compile(r"общее\s+заключение", re.IGNORECASE),
]


def test_report_to_xlsx_has_no_overall_verdict_string(tmp_path):
    payload = _run_payload()
    out_path = tmp_path / "report.xlsx"
    report_to_xlsx(payload, out_path)

    wb = load_workbook(out_path)
    all_text = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    all_text.append(c.value)
    joined = "\n".join(all_text)
    for pattern in _VERDICT_PATTERNS:
        assert not pattern.search(joined), f"найден запрещённый вердикт-паттерн: {pattern.pattern}"


# ── report_to_html ───────────────────────────────────────────────────────────────────────


def test_report_to_html_is_valid_and_contains_summary_table():
    payload = _run_payload()
    html = report_to_html(payload)

    assert isinstance(html, str)
    assert "<html" in html.lower() or "<!doctype" in html.lower()
    assert "<table" in html.lower()
    assert "4" in html  # total из summary


def test_report_to_html_contains_all_items():
    payload = _run_payload()
    html = report_to_html(payload)

    for item in payload["items"]:
        assert item["item_id"] in html
        assert item["criterion"] in html


def test_report_to_html_marks_computed_issue_rows_with_css_class():
    payload = _run_payload()
    html = report_to_html(payload)

    # находим строку с PD-KR-020 (computed_issue) — она должна иметь css-класс, отличающийся
    # от обычной строки (маркировка "визуально помечены")
    rows = re.findall(r"<tr[^>]*>.*?</tr>", html, flags=re.DOTALL)
    issue_rows = [r for r in rows if "PD-KR-020" in r]
    assert issue_rows, "строка PD-KR-020 не найдена в html"
    issue_row = issue_rows[0]
    assert 'class=' in issue_row
    # у обычной supported_by_evidence строки не должно быть того же класса, что у issue
    supported_rows = [r for r in rows if "PD-OB-001" in r]
    assert supported_rows
    supported_row = supported_rows[0]

    issue_class_match = re.search(r'class="([^"]*)"', issue_row)
    supported_class_match = re.search(r'class="([^"]*)"', supported_row)
    issue_classes = set((issue_class_match.group(1) if issue_class_match else "").split())
    supported_classes = set((supported_class_match.group(1) if supported_class_match else "").split())
    assert issue_classes - supported_classes, "computed_issue строка не отличается css-классом от обычной"


def test_report_to_html_no_binary_garbage():
    payload = _run_payload()
    html = report_to_html(payload)

    # html должен быть чисто текстовым/печатным — никаких суррогатов/непечатных управляющих
    # символов (кроме обычных \n, \r, \t)
    for ch in html:
        codepoint = ord(ch)
        if codepoint < 0x20 and ch not in "\n\r\t":
            pytest.fail(f"обнаружен управляющий символ 0x{codepoint:02x} в html-отчёте")
    # не должно быть суррогатных пар/replacement character — признак битой кодировки
    assert "�" not in html


def test_report_to_html_has_no_overall_verdict_string():
    payload = _run_payload()
    html = report_to_html(payload)
    for pattern in _VERDICT_PATTERNS:
        assert not pattern.search(html), f"найден запрещённый вердикт-паттерн: {pattern.pattern}"


def test_report_to_html_reflects_applied_human_decision():
    payload = _run_payload()
    html = report_to_html(payload)

    # PD-AR-049 имеет human_answer=yes + human_note — должно быть видно в html
    assert "Проверено инженером по плану этажа." in html


# ── decisions применяются в обоих форматах (сквозной сценарий) ─────────────────────────────


def test_decisions_reflected_in_both_xlsx_and_html(tmp_path):
    payload = _run_payload()
    # применим decision «руками» к одному из review_needed items — как это делает роутер
    # (_apply_decisions) перед вызовом отчёта
    for item in payload["items"]:
        if item["item_id"] == "PD-SPOZU-010":
            item["human_answer"] = "not_required"
            item["human_note"] = "Не требуется — уточнено с ГИПом"
            item["human_decision"] = "human_decided"
    payload["summary"]["human"] = {"yes": 1, "no": 0, "not_required": 1, "unset": 2}

    out_path = tmp_path / "report.xlsx"
    report_to_xlsx(payload, out_path)
    wb = load_workbook(out_path)
    ws = wb["Чек-лист"]
    xlsx_text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "Не уточнено с ГИПом" not in xlsx_text  # sanity: не мусорим случайными строками
    assert "Не требуется — уточнено с ГИПом" in xlsx_text

    html = report_to_html(payload)
    assert "Не требуется — уточнено с ГИПом" in html

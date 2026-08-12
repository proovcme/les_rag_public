"""W11.10 — спецификация (Ф9) → ВОР работ. Офлайн, без LLM."""

from __future__ import annotations

from pathlib import Path

import pytest

from backend.parquet_writer import save_parquet
from proxy.services.spec_to_bor_service import (
    generate_spec_bor,
    generate_spec_bor_from_rows,
    is_spec_to_bor_query,
    rows_from_spec_document,
    rows_from_spec_pdf,
    rows_from_spec_xlsx,
    spec_rows_to_work_lines,
    work_verb,
)


def test_is_spec_to_bor_query_word_boundary():
    # «повороты» содержит подстроку «вор» — НЕ должно триггерить канал ВОР
    assert is_spec_to_bor_query("Собери спецификацию лотков 200х50, повороты, крышки") is False
    assert is_spec_to_bor_query("спецификация светильников, забор, творог") is False
    # легитимные «ВОР из спецификации» — должны
    assert is_spec_to_bor_query("сделай ВОР из спецификации формы 9") is True
    assert is_spec_to_bor_query("ведомость объёмов работ из спецификации") is True
    # с tabular-вложением достаточно «ВОР», без слова «спецификация»
    assert is_spec_to_bor_query("Собери ВОР в разрезе ЛВЖ, Компрессор и зарядка", has_attachment=True) is True
    assert is_spec_to_bor_query("Собери ВОР в разрезе ЛВЖ, Компрессор и зарядка", has_attachment=False) is False
    # ЛСР по готовой ВОР — document LSR, не spec→ВОР (в т.ч. PDF)
    assert is_spec_to_bor_query(
        "Собери первую ЛСР по приложенной ВОР", has_attachment=True,
    ) is False
    assert is_spec_to_bor_query(
        "сделай смету по ВОР", has_attachment=True,
    ) is False


def _spec(name, unit="шт", qty=1.0, **kw):
    row = {"doc_type": "SPEC", "name": name, "unit": unit, "qty": qty, "source_file": "spec.xlsx"}
    row.update(kw)
    return row


# ── словарь глаголов ──

def test_verb_cable_is_prokladka():
    assert work_verb("Кабель медный ВВГнг 3х1,5") == "Прокладка"
    assert work_verb("Провод ПВ-1") == "Прокладка"


def test_verb_equipment_is_montazh():
    assert work_verb("Светильник NOTOR78 LED") == "Монтаж"
    assert work_verb("Щит распределительный ЩР") == "Монтаж"
    assert work_verb("Лоток кабельный 200х50") == "Монтаж"


def test_verb_fasteners_is_ustanovka():
    assert work_verb("Коробка установочная") == "Установка"
    assert work_verb("Наконечник кабельный") == "Установка"


def test_verb_default_montazh():
    assert work_verb("Нечто непонятное X") == "Монтаж"


# ── свод работ ──

def test_work_name_format_and_qty_carry():
    lines = spec_rows_to_work_lines([_spec("Светильник LED", "шт", 280.0)])
    assert len(lines) == 1
    assert lines[0].name == "Монтаж: Светильник LED"
    assert lines[0].qty == pytest.approx(280.0)
    assert lines[0].unit == "шт"


def test_cable_in_meters_prokladka():
    lines = spec_rows_to_work_lines([_spec("Кабель ВВГнг 3х1,5", "м", 744.93)])
    assert lines[0].name == "Прокладка: Кабель ВВГнг 3х1,5"
    assert lines[0].unit == "м"
    assert lines[0].qty == pytest.approx(744.93)


def test_identical_works_summed():
    lines = spec_rows_to_work_lines([
        _spec("Светильник LED", "шт", 280.0),
        _spec("Светильник LED", "шт", 109.0),
    ])
    assert len(lines) == 1
    assert lines[0].qty == pytest.approx(389.0)


def test_noise_rows_skipped():
    lines = spec_rows_to_work_lines([
        _spec("1. Раздел освещение", "", None),
        _spec("2", "шт", 5.0),
        _spec("Розетка 220В", "шт", 7.0),
    ])
    assert [l.name for l in lines] == ["Установка: Розетка 220В"]


def test_qty_missing_tracked():
    lines = spec_rows_to_work_lines([_spec("Прибор учёта", "шт", None)])
    assert lines[0].qty is None
    assert lines[0].qty_missing_rows == 1


# ── полный цикл Parquet → xlsx ──

def test_generate_spec_bor_end_to_end(tmp_path):
    parquet_dir = tmp_path / "ds" / "_parquet"
    parquet_dir.mkdir(parents=True)
    rows = [
        _spec("Кабель ВВГнг 3х1,5", "м", 744.93),
        _spec("Светильник LED", "шт", 280.0),
        _spec("Коробка установочная", "шт", 60.0),
    ]
    save_parquet(rows, str(parquet_dir / "spec.parquet"))
    out = tmp_path / "out"
    res = generate_spec_bor("ds", storage_root=tmp_path, output_dir=out, decompose=False)  # v1
    assert res["bor_lines"] == 3 and res["mode"] == "simple"
    names = {l["name"] for l in res["lines"]}
    assert "Прокладка: Кабель ВВГнг 3х1,5" in names
    assert "Монтаж: Светильник LED" in names
    assert "Установка: Коробка установочная" in names
    assert Path(res["xlsx_path"]).exists()


def test_generate_spec_bor_v2_decompose(tmp_path):
    parquet_dir = tmp_path / "ds" / "_parquet"
    parquet_dir.mkdir(parents=True)
    rows = [
        _spec("Кабель ВВГнг 3х1,5", "м", 744.93, section="ЭОМ"),
        _spec("Светильник LED", "шт", 280.0, section="ЭОМ"),
    ]
    save_parquet(rows, str(parquet_dir / "spec.parquet"))
    res = generate_spec_bor("ds", storage_root=tmp_path, output_dir=tmp_path / "out")  # decompose=default
    assert res["mode"] == "decompose"
    assert res["bor_lines"] > 2                # позиции декомпозированы в набор работ
    assert Path(res["xlsx_path"]).exists()     # xlsx с графами ГОСТ 21.111


def test_uses_no_llm():
    import inspect

    import proxy.services.spec_to_bor_service as svc

    src = inspect.getsource(svc)
    for marker in ("import httpx", "import openai", "/api/chat", "completions"):
        assert marker not in src


# ── v2: декомпозиция (методика ГОСТ 21.111) ──

from proxy.services.spec_to_bor_service import (  # noqa: E402
    _decompose,
    _has_cable_noun,
    spec_rows_to_work_lines_v2,
    work_lines_to_xlsx,
)


def test_decompose_cable_into_works():
    works, note = _decompose("кабель ВВГнг 3х2,5")
    assert "Разметка трассы" in works and any("Прокладка" in w for w in works)
    assert "конц" in note  # доп. работы по числу концов — в примечании


def test_decompose_device_install_connect():
    works, _ = _decompose("щит распределительный ЩР-1")
    assert any("Установка" in w for w in works) and "Подключение" in works


def test_decompose_rejects_cable_adjective_and_compound_wire():
    """«кабельный» / «водогазопроводная» must not become Прокладка кабеля."""
    box = "Коробка ответвительная 100x100x50 с 6 кабельными вводами, IP55"
    works_box, note_box = _decompose(box)
    assert works_box == ("Установка {предмет}",)
    assert "Прокладка" not in " ".join(works_box)
    assert note_box == ""
    assert work_verb(box) == "Установка"

    pipe = "Труба стальная водогазопроводная - Ø 15 мм"
    works_pipe, note_pipe = _decompose(pipe)
    assert "Прокладка кабеля" not in works_pipe
    assert any("Монтаж" in w for w in works_pipe)
    assert "Разметка трассы" in works_pipe
    assert note_pipe == ""
    assert work_verb(pipe) == "Монтаж"

    assert _has_cable_noun("Кабель силовой ВВГнг 3х1,5") is True
    assert _has_cable_noun("Провод ПВ-1") is True
    assert _has_cable_noun(box) is False
    assert _has_cable_noun(pipe) is False
    assert _has_cable_noun("Лоток кабельный 200х50") is False


def test_v2_nd_like_rows_no_false_cable_prokladka():
    rows = [
        _spec(
            "Коробка ответвительная 100x100x50 с 6 кабельными вводами, IP55",
            unit="шт",
            qty=70.0,
            pos="3.9",
        ),
        _spec("Кабель силовой ВВГнг 3х1,5", unit="м", qty=800.0, pos="4.1", mark="ППГ"),
        _spec("Труба стальная водогазопроводная Ø15", unit="м", qty=10.0, pos="5.2"),
        _spec("Труба стальная водогазопроводная Ø50", unit="м", qty=30.0, pos="5.2"),
    ]
    lines = spec_rows_to_work_lines_v2(rows)
    prokladka = [l for l in lines if l.work == "Прокладка кабеля"]
    assert len(prokladka) == 1
    assert prokladka[0].unit == "м"
    assert prokladka[0].qty == pytest.approx(800.0)
    assert not any(l.work == "Прокладка кабеля" and l.unit == "шт" for l in lines)
    pipe_lines = [l for l in lines if "Труба стальная" in l.work]
    assert pipe_lines and all(l.qty in (10.0, 30.0) for l in pipe_lines)
    assert any("Установка" in l.work and "кабельными вводами" in l.work for l in lines)


def test_v2_one_position_many_works_qty_inherited():
    rows = [_spec("кабель ВВГнг 3х2,5", unit="м", qty=1003.0, section="ЭОМ", mark="Э-1")]
    lines = spec_rows_to_work_lines_v2(rows)
    # одна позиция → несколько работ, у каждой объём = кол-ву позиции
    assert len(lines) >= 2
    for l in lines:
        assert l.unit == "м" and l.qty == 1003.0
        assert l.chertezh == "Э-1" and l.section == "ЭОМ"
        assert "поз" in l.note or "спецификац" in l.note


def test_v2_groups_and_sums_same_work():
    rows = [
        _spec("кабель А", unit="м", qty=100.0, section="ЭОМ"),
        _spec("кабель Б", unit="м", qty=50.0, section="ЭОМ"),
    ]
    lines = spec_rows_to_work_lines_v2(rows)
    razm = [l for l in lines if l.work.startswith("Разметка трассы")]
    assert len(razm) == 1 and razm[0].qty == 150.0  # свод одинаковой работы


def _write_kamenka_like_xlsx(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист2"
    ws.append(["Раздел", "Объем", "Наименование", "Артикул", "Производитель", "Ед. изм.", "Кол-во ФАКТ"])
    ws.append(["ЛВЖ", "осн.", "Кабель КПСнг(А)-FRLS 1*2*1,0", "1*2*1,0", "Россия", "м.", 2000])
    ws.append(["ЛВЖ", "доп.", "Шайба кузовная оцинкованная 12x37 мм DIN 9021 (30 шт.)", "", "", "шт", 1])
    ws.append(["Компрессор", "осн.", "Кабель ППГ нг(А) HF 1*150", "", "Кабэкс", "м.", 200])
    ws.append(["Компрессор", "осн.", "Лоток глухой 200*50*3000", "35024", "ДКС", "м.", 33])
    ws.append(["Зарядка", "осн.", "Кабель силовой ВВГнг(А)-LS 4х16 ок", "", "", "м.", 80])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def test_rows_from_spec_xlsx_kamenka_like(tmp_path):
    src = tmp_path / "kamenka.xlsx"
    _write_kamenka_like_xlsx(src)
    rows = rows_from_spec_xlsx(src, source_label="Общая по Каменке.xlsx")
    assert len(rows) == 5
    sections = {r["section"] for r in rows}
    assert sections == {"ЛВЖ", "Компрессор", "Зарядка"}
    by_name = {r["name"]: r for r in rows}
    assert by_name["Кабель КПСнг(А)-FRLS 1*2*1,0"]["qty"] == 2000.0
    assert by_name["Кабель силовой ВВГнг(А)-LS 4х16 ок"]["section"] == "Зарядка"


def test_rows_from_spec_document_dispatches_xlsx(tmp_path):
    src = tmp_path / "kamenka.xlsx"
    _write_kamenka_like_xlsx(src)
    rows = rows_from_spec_document(src, source_label="Общая по Каменке.xlsx")
    assert len(rows) == 5


def test_rows_from_spec_pdf_reads_tables(monkeypatch, tmp_path):
    class _Page:
        def extract_tables(self):
            return [[
                ["Раздел", "Наименование", "Ед. изм.", "Кол-во"],
                ["ЛВЖ", "Кабель КПСнг(А)-FRLS", "м.", "100"],
                ["Компрессор", "Лоток глухой 200*50", "м.", "33"],
            ]]

    class _Pdf:
        pages = [_Page()]

        def __enter__(self):
            return self

        def __exit__(self, *_a):
            return False

    import sys
    from types import ModuleType

    fake = ModuleType("pdfplumber")
    fake.open = lambda *_a, **_k: _Pdf()
    monkeypatch.setitem(sys.modules, "pdfplumber", fake)

    src = tmp_path / "spec.pdf"
    src.write_bytes(b"%PDF-1.4 placeholder")
    rows = rows_from_spec_pdf(src, source_label="spec.pdf")
    assert len(rows) == 2
    assert rows[0]["name"] == "Кабель КПСнг(А)-FRLS"
    assert rows[0]["qty"] == 100.0
    assert rows[1]["section"] == "Компрессор"
    via_dispatch = rows_from_spec_document(src, source_label="spec.pdf")
    assert len(via_dispatch) == 2


def test_generate_spec_bor_from_xlsx_has_works_no_prices(tmp_path):
    src = tmp_path / "kamenka.xlsx"
    _write_kamenka_like_xlsx(src)
    rows = rows_from_spec_xlsx(src)
    res = generate_spec_bor_from_rows(
        rows, output_dir=tmp_path / "out", title="ВОР Каменка", source_id="read_test"
    )
    assert res["source_rows"] == 5
    assert res["bor_lines"] >= 5
    assert Path(res["xlsx_path"]).exists()
    joined = "\n".join(l["name"] for l in res["lines"])
    assert "Прокладка" in joined or "Разметка" in joined
    assert "Монтаж" in joined or "Установка" in joined
    # no price columns in ГОСТ VOR xlsx
    import openpyxl

    ws = openpyxl.load_workbook(res["xlsx_path"]).active
    header = [ws.cell(2, c).value for c in range(1, 8)]
    header_text = " ".join(str(h or "") for h in header).lower()
    assert "цена" not in header_text and "стоимость" not in header_text
    assert any(l.get("section") == "ЛВЖ" for l in res["lines"])
    assert any(l.get("section") == "Компрессор" for l in res["lines"])


def test_v2_xlsx_has_gost_columns(tmp_path):
    rows = [_spec("извещатель ИП-212", unit="шт", qty=85.0, section="АУПС")]
    lines = spec_rows_to_work_lines_v2(rows)
    out = tmp_path / "vor.xlsx"
    n = work_lines_to_xlsx(lines, out, title="ВОР тест")
    assert out.exists() and n == len(lines)
    import openpyxl
    ws = openpyxl.load_workbook(out).active
    hdr = [ws.cell(row=2, column=c).value for c in range(1, 7)]
    assert hdr == ["№", "Наименование работ", "Ед. изм.", "Кол-во", "Ссылка на чертёж", "Примечание"]


# ── Form-9 PDF wrap coalesce / RU qty / км→м (ADR-11) ──

from proxy.services.spec_to_bor_service import (  # noqa: E402
    _coalesce_form9_rows,
    _spec_rows_from_matrix,
)


def test_coalesce_form9_cable_split_matrix():
    """pdfplumber-style wrap: prefix + soft lines + size branch → one cable row."""
    matrix = [
        ["Поз.", "Наименование и техническая характеристика", "Тип", "Код", "Поставщик", "Ед. изм.", "Кол.", "Масса"],
        ["1", "Кабель силовой с изоляцией из", "", "", "", "", "", ""],
        ["", "полимерных материалов, не распространяющих горение,", "", "", "", "", "", ""],
        ["", "сечением:", "", "", "", "", "", ""],
        ["", "- 3х1,5 мм2", "ВВГнг(А)-LS", "", "", "км", "1,5", ""],
        ["2", "Светильник LED", "N78", "", "", "шт", "10", "0,5"],
    ]
    rows, _ = _spec_rows_from_matrix(matrix, source="nd.pdf")
    assert len(rows) == 2
    cable = rows[0]
    assert "Кабель силовой" in cable["name"]
    assert "3х1,5" in cable["name"]
    assert "полимерных" in cable["name"]
    assert cable["unit"] == "м"
    assert cable["qty"] == pytest.approx(1500.0)  # 1,5 км → м
    assert "км→м" in (cable.get("note") or "")
    assert cable["pos"] == "1"
    assert rows[1]["name"] == "Светильник LED"
    assert rows[1]["qty"] == pytest.approx(10.0)  # mass column ignored


def test_coalesce_form9_km_without_qty_no_invent():
    rows = _coalesce_form9_rows([
        _spec("Кабель ВВГнг 3х1,5", unit="км", qty=None),
    ])
    assert len(rows) == 1
    assert rows[0]["unit"] == "м"
    assert rows[0]["qty"] is None
    assert "без выдуманного" in (rows[0].get("note") or "")


def test_coalesce_form9_size_branches_separate_positions():
    rows = _coalesce_form9_rows([
        {"doc_type": "SPEC", "name": "Кабель силовой сечением:", "unit": "", "qty": None, "pos": "5"},
        {"doc_type": "SPEC", "name": "- 3х1,5 мм2", "unit": "км", "qty": "0,2", "pos": "5.1"},
        {"doc_type": "SPEC", "name": "- 3х2,5 мм2", "unit": "км", "qty": "0,1", "pos": "5.2"},
    ])
    assert len(rows) == 2
    assert rows[0]["qty"] == pytest.approx(200.0) and rows[0]["unit"] == "м"
    assert rows[1]["qty"] == pytest.approx(100.0) and rows[1]["unit"] == "м"
    assert "3х1,5" in rows[0]["name"] and "Кабель силовой" in rows[0]["name"]
    assert "3х2,5" in rows[1]["name"]


def test_coalesce_form9_parent_qty_not_double_counted():
    rows = _coalesce_form9_rows([
        {"doc_type": "SPEC", "name": "Кабель силовой сечением:", "unit": "км", "qty": 1.0, "pos": "1"},
        {"doc_type": "SPEC", "name": "- 3х1,5 мм2", "unit": "", "qty": None},
        {"doc_type": "SPEC", "name": "- 3х2,5 мм2", "unit": "", "qty": None},
    ])
    assert len(rows) == 2
    assert rows[0]["qty"] == pytest.approx(1000.0)
    assert rows[1]["qty"] is None  # parent qty already consumed


def test_parse_ru_qty_in_matrix():
    matrix = [
        ["Наименование", "Ед. изм.", "Кол-во"],
        ["Лоток кабельный", "м", "12,5"],
    ]
    rows, _ = _spec_rows_from_matrix(matrix, source="t.xlsx")
    assert rows[0]["qty"] == pytest.approx(12.5)


def test_v2_propagates_km_note():
    rows = _coalesce_form9_rows([_spec("Кабель ВВГнг", unit="км", qty=0.5)])
    lines = spec_rows_to_work_lines_v2(rows)
    assert any("км→м" in (l.note or "") for l in lines)
    assert all(l.qty == pytest.approx(500.0) for l in lines)


def test_coalesce_form9_cross_table_size_branch():
    """Size branch in the next pdfplumber table joins prior cable prefix."""
    batch = [
        {
            "doc_type": "SPEC",
            "name": "Кабель силовой сечением:",
            "unit": "",
            "qty": None,
            "pos": "4.2",
            "mark": "ППГ-нг(А)-FRHF",
        },
        {
            "doc_type": "SPEC",
            "name": "- 3х1,5 мм2",
            "unit": "км",
            "qty": "0,24",
            "pos": "1",
        },
    ]
    rows = _coalesce_form9_rows(batch)
    assert len(rows) == 1
    assert rows[0]["qty"] == pytest.approx(240.0)
    assert rows[0]["unit"] == "м"
    assert rows[0]["pos"] == "4.2"
    assert "3х1,5" in rows[0]["name"]
    assert "Кабель силовой" in rows[0]["name"]


def test_rating_only_fragment_is_noise():
    from proxy.services.spec_to_bor_service import _is_noise_name

    assert _is_noise_name("16А, ~230В, IP20") is True
    assert _is_noise_name("Розетка 16А, ~230В, IP20") is False

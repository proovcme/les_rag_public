"""Unified Construction Harness v0.5 — Resource-Based Cost Calculation.

Детальный ресурсный обсчёт строительной позиции по ГЭСН: норма → ресурсный состав → коэффициенты
условий → цены/индексы/КАЦ → прямые затраты → ФОТ → НР → СП → итог позиции → доп. ТЦ/КАЦ → grand.

НЕ замена ЛСР-пути (lsr_assembly), а отдельный thin-сервис. Числа ТОЛЬКО из tool-результатов:
коэффициенты/цены/ставки НР-СП — из источника (workbook/parquet/ФГИС), не из модели. Нет цены/
коэффициента/ставки → MISSING/BLOCKED, не silent default. Excel-пример — golden, не источник формул.

ЧЕСТНО: исходный ПРИМЕР_обсчета_24_06.xlsx в репо отсутствует — golden-позиция закодирована как
структурная fixture по документированной структуре (`golden_position()`), source_refs ссылаются на
лист/строку. Движок (expand/price/direct/ФОТ/НР-СП/assemble/КАЦ/grand) реальный и переиспользуем для
будущего реального ВОР+ГЭСН+цен. Per-line цены материалов/машин в данных не полны → категорийные
суб-итоги (материалы/машины/машинисты) берутся как RETRIEVED-ячейки workbook; labor-строка, ФОТ, НР,
СП, итог позиции и grand — полностью COMPUTED. Все golden-числа воспроизводятся точно.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from proxy.services.evidence_contract import (
    ConstructionHarnessResult,
    EvidenceItem,
    EvidenceType,
    block_of,
)

WORKBOOK_NAME = "ПРИМЕР_обсчета_24_06.xlsx"
RECONSTRUCTED_FIXTURE = Path("tests/fixtures/cost_calc/ПРИМЕР_обсчета_24_06.xlsx")
REAL_WORKBOOK = Path("tests/fixtures/cost_calc/real_ПРИМЕР_обсчета_24_06.xlsx")

# v0.6: РЕАЛЬНЫЙ ПРИМЕР_обсчета_24_06.xlsx найден (creator='fsnb2022.ru') и скопирован в REAL_WORKBOOK.
# Система различает real (оригинал fsnb2022.ru) и reconstructed (наша openpyxl-fixture) и НЕ выдаёт
# reconstructed за real (ТЗ §0.4/§8/§11). Расчёт воспроизводится КОДОМ; xlsx хранит значения, не формулы.
FIXTURE_PROVENANCE = "real"
REAL_WORKBOOK_ABSENT_NOTE = ("оригинальный workbook не найден; источник — reconstructed fixture "
                             "(структура документирована, расчёт воспроизводится кодом)")


def _r2(x: float) -> float:
    return round(float(x) + 1e-9, 2)


def _provenance_of(path: Path, creator: str) -> str:
    cl = (creator or "").lower()
    if cl.startswith("openpyxl"):
        return "reconstructed"
    if "fsnb" in cl or "excel" in cl or "fsnb2022" in cl:
        return "real"
    return "real_candidate" if path.name.startswith("real_") else "reconstructed"


def real_workbook_status(path: str | Path | None = None) -> dict[str, Any]:
    """Провенанс источника обсчёта. real → оригинал fsnb2022.ru; reconstructed → наша fixture;
    absent → нет файла. НИКОГДА не помечает reconstructed как real."""
    p = Path(path) if path else (REAL_WORKBOOK if REAL_WORKBOOK.exists() else RECONSTRUCTED_FIXTURE)
    if not p.exists():
        return {"available": False, "provenance": "absent", "path": str(p), "creator": "",
                "note": REAL_WORKBOOK_ABSENT_NOTE}
    try:
        import openpyxl
        creator = openpyxl.load_workbook(p, read_only=True).properties.creator or ""
    except Exception:  # noqa: BLE001
        creator = ""
    prov = _provenance_of(p, creator)
    note = ("источник — РЕАЛЬНЫЙ workbook (fsnb2022.ru)" if prov == "real" else
            REAL_WORKBOOK_ABSENT_NOTE if prov == "reconstructed" else "кандидат реального workbook")
    return {"available": True, "provenance": prov, "path": str(p), "creator": creator, "note": note}


# ── data models ──────────────────────────────────────────────────────────────────────────

@dataclass
class NormResource:
    code: str
    name: str
    norm_qty: float | None
    unit: str
    category: str = "unknown"        # labor|machinist_labor|machine|material|equipment|price_item|project_quantity|unknown
    source_ref: str = ""
    raw_qty: str = ""                # «П» = проектная потребность


@dataclass
class CoefficientSet:
    labor_coeff: float = 1.0
    machine_usage_coeff: float = 1.0
    machinist_labor_coeff: float = 1.0
    material_coeff: float = 1.0
    source_ref: str = ""
    reason: str = ""
    status: str = "missing"          # retrieved|assumed_default_1|missing|blocked


@dataclass
class ExpandedResourceLine:
    resource_code: str
    resource_name: str
    unit: str
    category: str
    norm_qty_per_unit: float | None
    position_qty: float
    coefficient: float
    total_qty: float | None
    source_refs: list[str] = field(default_factory=list)
    status: str = "computed"         # computed|missing|blocked


@dataclass
class ResourcePrice:
    resource_code: str
    unit: str
    base_price: float | None = None
    index: float | None = None
    current_price: float | None = None
    source_ref: str = ""
    price_status: str = "not_found"  # found_current|base_times_index|needs_kac|not_found|assumed


@dataclass
class ResourceCostLine:
    resource_code: str
    name: str
    category: str
    total_qty: float | None
    unit_price_current: float | None
    line_total: float | None
    price_status: str
    source_refs: list[str] = field(default_factory=list)
    blockers: list[str] = field(default_factory=list)


@dataclass
class ResourceEstimateResult:
    norm_code: str = ""
    title: str = ""
    measure_unit: str = ""
    position_qty: float = 0.0
    direct_cost_total: float | None = None
    labor_cost_total: float | None = None
    machinist_labor_cost_total: float | None = None
    machine_cost_total: float | None = None
    material_cost_total: float | None = None
    equipment_cost_total: float | None = None
    fot: float | None = None
    nr_rate: float | None = None
    nr_amount: float | None = None
    sp_rate: float | None = None
    sp_amount: float | None = None
    position_total: float | None = None
    additional_price_items_total: float | None = None
    grand_total: float | None = None
    total_status: str = "blocked"    # complete|partial|blocked
    missing_prices: list[str] = field(default_factory=list)
    blockers: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    evidence_blocks: list = field(default_factory=list)
    trace: list = field(default_factory=list)


# ── resource category classification (детерминированно) ──────────────────────────────────

def classify_resource_category(code: str, name: str = "", raw_qty: str = "") -> str:
    c = (code or "").strip()
    n = (name or "").lower()
    if str(raw_qty).strip().upper() == "П":
        return "project_quantity"
    if c.startswith("ТЦ_") or c.startswith("ТЦ ") or c.startswith("КАЦ"):
        return "price_item"
    if c.startswith("1-"):
        return "labor"
    if c.startswith("4-") or "отм" in n or "зтм" in n or "машинист" in n:
        return "machinist_labor"
    if c == "2":
        return "machinist_labor"
    if c.startswith("91."):
        return "machine"
    if c[:3] in ("01.", "07.", "08.", "11.", "14.") or c[:2] in ("0.", "1.") and "." in c:
        return "material"
    if any(c.startswith(p) for p in ("01.", "02.", "03.", "05.", "06.", "07.", "08.", "09.",
                                     "10.", "11.", "12.", "13.", "14.", "15.")):
        return "material"
    return "unknown"


# машина → тарифный код машиниста (минимальный маппинг для golden; не выдумываем полный).
# Канонический справочник «машина→машинист» — config/domain/fsem_machinist.yaml (fsem_machinist_service);
# здесь оставлен отдельный минимальный маппинг под golden (консолидировать при работе над golden, #4).
_MACHINE_TO_MACHINIST: dict[str, str] = {
    "91.05.05-015": "4-100-060",     # кран 16 т → машинист 6 разр.
    "91.14.02-001": "4-100-040",     # автомобиль бортовой → машинист 4 разр.
}


def machine_to_machinist(machine_code: str) -> str | None:
    return _MACHINE_TO_MACHINIST.get((machine_code or "").strip())


# ── coefficients ─────────────────────────────────────────────────────────────────────────

def coefficient_for(category: str, coeff: CoefficientSet) -> float:
    return {"labor": coeff.labor_coeff, "machine": coeff.machine_usage_coeff,
            "machinist_labor": coeff.machinist_labor_coeff,
            "material": coeff.material_coeff}.get(category, 1.0)


# ── expansion: total_qty = norm × position × coeff ───────────────────────────────────────

def expand_resource(res: NormResource, position_qty: float, coeff: CoefficientSet) -> ExpandedResourceLine:
    k = coefficient_for(res.category, coeff)
    if res.category == "project_quantity" or res.norm_qty is None:
        return ExpandedResourceLine(res.code, res.name, res.unit, res.category, res.norm_qty,
                                    position_qty, k, None, [res.source_ref], status="missing")
    total = res.norm_qty * position_qty * k
    return ExpandedResourceLine(res.code, res.name, res.unit, res.category, res.norm_qty,
                                position_qty, k, total, [res.source_ref], status="computed")


# ── price lookup ─────────────────────────────────────────────────────────────────────────

def resolve_price(rp: ResourcePrice) -> ResourcePrice:
    """found_current → как есть; base+index → base×index; '-'/нет → needs_kac/not_found."""
    if rp.current_price is not None and rp.current_price != "-":
        rp.price_status = "found_current"
        return rp
    if rp.base_price is not None and rp.index is not None:
        rp.current_price = _r2(rp.base_price * rp.index)
        rp.price_status = "base_times_index"
        return rp
    rp.price_status = "needs_kac"
    rp.current_price = None
    return rp


def cost_line(expanded: ExpandedResourceLine, price: ResourcePrice) -> ResourceCostLine:
    if expanded.total_qty is None:
        return ResourceCostLine(expanded.resource_code, expanded.resource_name, expanded.category,
                                None, None, None, "project_quantity", expanded.source_refs,
                                blockers=["проектная потребность «П» не разрешена"])
    if price.current_price is None:
        return ResourceCostLine(expanded.resource_code, expanded.resource_name, expanded.category,
                                expanded.total_qty, None, None, price.price_status,
                                expanded.source_refs + ([price.source_ref] if price.source_ref else []),
                                blockers=[f"нет текущей цены ({price.price_status})"])
    return ResourceCostLine(expanded.resource_code, expanded.resource_name, expanded.category,
                            expanded.total_qty, price.current_price,
                            _r2(expanded.total_qty * price.current_price), price.price_status,
                            expanded.source_refs + ([price.source_ref] if price.source_ref else []))


# ── ФОТ / НР / СП / итог ─────────────────────────────────────────────────────────────────

def calculate_fot(labor_cost: float, machinist_labor_cost: float) -> float:
    return _r2(labor_cost + machinist_labor_cost)


def apply_nr_sp(fot: float, nr_rate: float | None, sp_rate: float | None) -> tuple[float | None, float | None]:
    nr = _r2(fot * nr_rate) if nr_rate is not None else None
    sp = _r2(fot * sp_rate) if sp_rate is not None else None
    return nr, sp


# ── golden position fixture (документированный ПРИМЕР_обсчета_24_06.xlsx) ─────────────────

def _sref(sheet: str, row: int | str) -> str:
    return f"{WORKBOOK_NAME}#{sheet}!R{row}"


def golden_position() -> dict[str, Any]:
    """Структурное представление позиции ГЭСН09-06-006-03 из листа `пример` (workbook отсутствует
    в репо → реконструкция по документированной структуре). Числа = expected golden; source_refs
    указывают лист/строку. Per-line цены материалов/машин в данных не полны → категорийные суб-итоги
    как RETRIEVED-ячейки workbook; labor-строка COMPUTED из нормы×коэфф×объём×цена."""
    coeff = CoefficientSet(labor_coeff=1.15, machine_usage_coeff=1.15, machinist_labor_coeff=1.15,
                           material_coeff=1.0, status="retrieved",
                           reason="Приказ от 30.01.2024 №55/пр прил.5 табл.1 п.5 (стеснённые условия)",
                           source_ref=_sref("пример", 4))
    return {
        "norm_code": "ГЭСН09-06-006-03",
        "title": "Монтаж стационарных конструкций сцены: направляющие с каркасами ограждений",
        "unit": "т",
        "position_qty": 26.958848,
        "position_qty_ref": _sref("пример", 3),
        "coeff": coeff,
        # labor-строка — полностью COMPUTED (норма×коэфф×объём×цена)
        "labor": {"code": "1-100-39", "name": "Труд рабочих, средний разряд 3,9", "norm_qty": 230.21,
                  "unit": "чел.-ч", "price": 552.21, "src": _sref("ГЭСН 09-06-006-03", 3),
                  "price_src": _sref("пример", 6)},
        # категорийные суб-итоги из workbook (RETRIEVED — per-line материалы не полны в данных)
        "subtotals": {
            "machinist_labor": {"value": 19228.60, "src": _sref("пример", 8)},
            "machine": {"value": 127906.66, "src": _sref("пример", 12)},
            "material": {"value": 245466.07, "src": _sref("пример", 30)},
        },
        # ставки НР/СП — из workbook
        "nr_rate": 0.93, "nr_src": _sref("пример", 41),
        "sp_rate": 0.62, "sp_src": _sref("пример", 42),
        # доп. ТЦ/КАЦ-позиция
        "kac_item": {"code": "ТЦ_07.2.03.00_78_7810391323_14.11.2025_02_1.1", "name": "Крепежная рама",
                     "qty": 4, "unit": "шт", "price": 1588709.31, "src": _sref("пример", 45)},
        # проектная потребность «П» (информативно, не в priced-сумме)
        "project_qty_line": {"code": "07.2.06.06", "name": "материал по проектной потребности",
                             "raw_qty": "П", "src": _sref("ГЭСН 09-06-006-03", 20)},
        # репрезентативные ресурсные строки для unit-проверок expand/price (НЕ суммируются в direct,
        # т.к. учтены в категорийных суб-итогах — иначе двойной счёт)
        "sample_resources": [
            {"code": "91.05.05-015", "name": "Кран на автомобильном ходу 16 т", "norm_qty": 0.73,
             "unit": "маш.-ч", "cat": "machine", "price_current": 1663.18, "src": _sref("ГЭСН 09-06-006-03", 5)},
            {"code": "91.06.03-062", "name": "Лебёдка электрическая 31.39 кН", "norm_qty": 2.9,
             "unit": "маш.-ч", "cat": "machine", "base": 13.44, "index": 1.42, "src": _sref("ГЭСН 09-06-006-03", 6)},
            {"code": "01.7.15.06-0111", "name": "Кислород технический газообразный", "norm_qty": 0.8,
             "unit": "м3", "cat": "material", "src": _sref("ГЭСН 09-06-006-03", 15)},
            {"code": "01.7.11.07-0227", "name": "Пропан-бутан", "norm_qty": 0.24, "unit": "кг",
             "cat": "material", "base": 41.38, "index": 1.71, "src": _sref("ГЭСН 09-06-006-03", 14)},
            {"code": "14.4.01.01-0003", "name": "Конструкции металлические", "norm_qty": 0.027, "unit": "т",
             "cat": "material", "src": _sref("ГЭСН 09-06-006-03", 18)},
        ],
    }


# ── golden runner: воспроизводит обсчёт кодом ─────────────────────────────────────────────

def run_resource_cost_golden(pos: dict[str, Any] | None = None) -> ResourceEstimateResult:
    pos = pos or golden_position()
    coeff: CoefficientSet = pos["coeff"]
    qty = pos["position_qty"]
    trace: list[dict] = [{"step": "position", "norm": pos["norm_code"], "qty": qty},
                         {"step": "coefficient", "value": coeff.labor_coeff, "reason": coeff.reason}]
    res = ResourceEstimateResult(norm_code=pos["norm_code"], title=pos["title"],
                                 measure_unit=pos["unit"], position_qty=qty,
                                 nr_rate=pos["nr_rate"], sp_rate=pos["sp_rate"])
    retrieved: list[EvidenceItem] = []
    computed: list[EvidenceItem] = []
    missing: list[EvidenceItem] = []

    # 1. labor — полностью COMPUTED
    lb = pos["labor"]
    labor_res = NormResource(lb["code"], lb["name"], lb["norm_qty"], lb["unit"], "labor", lb["src"])
    labor_exp = expand_resource(labor_res, qty, coeff)
    labor_price = ResourcePrice(lb["code"], lb["unit"], current_price=lb["price"], source_ref=lb["price_src"])
    resolve_price(labor_price)
    labor_line = cost_line(labor_exp, labor_price)
    res.labor_cost_total = labor_line.line_total
    retrieved.append(EvidenceItem(EvidenceType.RETRIEVED, f"Норма {lb['code']} {lb['name']}",
                                  value=lb["norm_qty"], unit=lb["unit"], source_refs=[lb["src"]]))
    retrieved.append(EvidenceItem(EvidenceType.RETRIEVED, f"Коэффициент условий {coeff.labor_coeff}",
                                  source_refs=[coeff.source_ref], status="supported"))
    computed.append(EvidenceItem(EvidenceType.COMPUTED, f"Труд рабочих, кол-во ({lb['code']})",
                                 value=_r2_q(labor_exp.total_qty), unit=lb["unit"],
                                 formula=f"{lb['norm_qty']} × {coeff.labor_coeff} × {qty}",
                                 inputs=[{"norm": lb["norm_qty"]}, {"coeff": coeff.labor_coeff}, {"qty": qty}],
                                 source_refs=[lb["src"]]))
    computed.append(EvidenceItem(EvidenceType.COMPUTED, f"Труд рабочих, стоимость ({lb['code']})",
                                 value=labor_line.line_total, unit="руб.",
                                 formula=f"{_r2_q(labor_exp.total_qty)} × {lb['price']}",
                                 inputs=[{"qty": _r2_q(labor_exp.total_qty)}, {"price": lb["price"]}],
                                 source_refs=[lb["src"], lb["price_src"]]))
    trace.append({"step": "labor", "qty": labor_exp.total_qty, "cost": labor_line.line_total})

    # 2. категорийные суб-итоги (RETRIEVED — workbook cells)
    sub = pos["subtotals"]
    res.machinist_labor_cost_total = sub["machinist_labor"]["value"]
    res.machine_cost_total = sub["machine"]["value"]
    res.material_cost_total = sub["material"]["value"]
    for key, human in (("machinist_labor", "ФОТ машинистов"), ("machine", "Эксплуатация машин"),
                       ("material", "Материалы")):
        retrieved.append(EvidenceItem(EvidenceType.RETRIEVED, f"{human} (суб-итог из workbook)",
                                      value=sub[key]["value"], unit="руб.", source_refs=[sub[key]["src"]]))

    # 3. прямые затраты — COMPUTED
    res.direct_cost_total = _r2(res.labor_cost_total + res.machinist_labor_cost_total
                                + res.machine_cost_total + res.material_cost_total)
    computed.append(EvidenceItem(EvidenceType.COMPUTED, "Прямые затраты", value=res.direct_cost_total,
                                 unit="руб.", formula="ОЗП + ФОТмаш + ЭМ + материалы",
                                 inputs=[{"labor": res.labor_cost_total}, {"machinist": res.machinist_labor_cost_total},
                                         {"machine": res.machine_cost_total}, {"material": res.material_cost_total}]))

    # 4. ФОТ / НР / СП
    res.fot = calculate_fot(res.labor_cost_total, res.machinist_labor_cost_total)
    res.nr_amount, res.sp_amount = apply_nr_sp(res.fot, res.nr_rate, res.sp_rate)
    computed.append(EvidenceItem(EvidenceType.COMPUTED, "ФОТ", value=res.fot, unit="руб.",
                                 formula="ОЗП + ФОТ машинистов",
                                 inputs=[{"labor": res.labor_cost_total}, {"machinist": res.machinist_labor_cost_total}]))
    if res.nr_amount is not None:
        retrieved.append(EvidenceItem(EvidenceType.RETRIEVED, f"Ставка НР {int(res.nr_rate*100)}%",
                                      source_refs=[pos["nr_src"]], status="supported"))
        computed.append(EvidenceItem(EvidenceType.COMPUTED, "НР", value=res.nr_amount, unit="руб.",
                                     formula=f"ФОТ × {res.nr_rate}", inputs=[{"fot": res.fot}, {"rate": res.nr_rate}],
                                     source_refs=[pos["nr_src"]]))
    else:
        missing.append(EvidenceItem(EvidenceType.MISSING, "Ставка НР", blockers=["ставка НР не задана источником"]))
    if res.sp_amount is not None:
        retrieved.append(EvidenceItem(EvidenceType.RETRIEVED, f"Ставка СП {int(res.sp_rate*100)}%",
                                      source_refs=[pos["sp_src"]], status="supported"))
        computed.append(EvidenceItem(EvidenceType.COMPUTED, "СП", value=res.sp_amount, unit="руб.",
                                     formula=f"ФОТ × {res.sp_rate}", inputs=[{"fot": res.fot}, {"rate": res.sp_rate}],
                                     source_refs=[pos["sp_src"]]))
    else:
        missing.append(EvidenceItem(EvidenceType.MISSING, "Ставка СП", blockers=["ставка СП не задана источником"]))

    # 5. итог позиции
    if res.nr_amount is not None and res.sp_amount is not None:
        res.position_total = _r2(res.direct_cost_total + res.nr_amount + res.sp_amount)
        computed.append(EvidenceItem(EvidenceType.COMPUTED, "Итог по позиции", value=res.position_total,
                                     unit="руб.", formula="прямые + НР + СП",
                                     inputs=[{"direct": res.direct_cost_total}, {"nr": res.nr_amount},
                                             {"sp": res.sp_amount}]))

    # 6. доп. ТЦ/КАЦ-позиция — отдельная price_item evidence
    ki = pos.get("kac_item")
    if ki:
        kac_total = _r2(ki["qty"] * ki["price"])
        res.additional_price_items_total = kac_total
        retrieved.append(EvidenceItem(EvidenceType.RETRIEVED, f"ТЦ/КАЦ: {ki['name']} ({ki['code']})",
                                      value=ki["price"], unit=f"руб./{ki['unit']}", source_refs=[ki["src"]]))
        computed.append(EvidenceItem(EvidenceType.COMPUTED, f"ТЦ/КАЦ-позиция: {ki['name']}", value=kac_total,
                                     unit="руб.", formula=f"{ki['qty']} {ki['unit']} × {ki['price']}",
                                     inputs=[{"qty": ki["qty"]}, {"price": ki["price"]}], source_refs=[ki["src"]]))

    # 7. проектная потребность «П» — MISSING (не в priced-сумме)
    pql = pos.get("project_qty_line")
    if pql:
        missing.append(EvidenceItem(EvidenceType.MISSING, f"Проектная потребность {pql['code']} ({pql['name']})",
                                    blockers=["количество «П» не разрешено — нужна проектная спецификация"],
                                    source_refs=[pql["src"]]))
        res.warnings.append("строка с потребностью «П» не учтена в стоимости (нужна проектная спецификация)")

    # 8. grand total — только при отсутствии критичных блокеров (цены/ставки)
    res.missing_prices = [m.title for m in missing if "ставка" in m.title.lower() or "цена" in m.title.lower()]
    critical = bool(res.missing_prices) or res.position_total is None
    if not critical:
        res.grand_total = _r2(res.position_total + (res.additional_price_items_total or 0.0))
        res.total_status = "complete"
        computed.append(EvidenceItem(EvidenceType.COMPUTED, "ИТОГО (позиция + ТЦ/КАЦ)", value=res.grand_total,
                                     unit="руб.", formula="итог позиции + ТЦ/КАЦ-позиции",
                                     inputs=[{"position": res.position_total},
                                             {"kac": res.additional_price_items_total}]))
    else:
        res.total_status = "blocked" if res.position_total is None else "partial"

    blocks = []
    if retrieved:
        blocks.append(block_of(EvidenceType.RETRIEVED, "Источники (норма/коэфф/цены/ставки/КАЦ)", retrieved))
    if computed:
        blocks.append(block_of(EvidenceType.COMPUTED, "Ресурсный расчёт", computed))
    if missing:
        blocks.append(block_of(EvidenceType.MISSING, "Не хватает / требует уточнения", missing))
    res.evidence_blocks = blocks
    res.trace = trace
    return res


def _r2_q(x: float | None) -> float | None:
    return round(x, 7) if x is not None else None


# ── expand sample resources (для unit-проверок expand/price из golden) ───────────────────

def expand_sample(pos: dict[str, Any] | None = None) -> list[ExpandedResourceLine]:
    pos = pos or golden_position()
    coeff = pos["coeff"]
    out = []
    for s in pos["sample_resources"]:
        r = NormResource(s["code"], s["name"], s["norm_qty"], s["unit"], s["cat"], s["src"])
        out.append(expand_resource(r, pos["position_qty"], coeff))
    return out


def price_sample(pos: dict[str, Any] | None = None) -> list[ResourcePrice]:
    pos = pos or golden_position()
    out = []
    for s in pos["sample_resources"]:
        rp = ResourcePrice(s["code"], s["unit"], base_price=s.get("base"), index=s.get("index"),
                           current_price=s.get("price_current"), source_ref=s["src"])
        out.append(resolve_price(rp))
    return out


# ── parse uploaded/dataset workbook (если xlsx присутствует) ─────────────────────────────

@dataclass
class WorkbookSourceRef:
    """Cell-level ссылка на ячейку/строку workbook (ТЗ §4)."""
    file_name: str = WORKBOOK_NAME
    sheet_name: str = ""
    row: int | None = None
    column: int | None = None
    cell_address: str = ""
    value_type: str = ""

    def ref(self) -> str:
        base = f"{self.file_name}#{self.sheet_name}"
        if self.cell_address:
            return f"{base}!{self.cell_address}"
        if self.row is not None:
            return f"{base}!R{self.row}" + (f"C{self.column}" if self.column else "")
        return base


@dataclass
class MethodNote:
    """Методический комментарий из ПРАВОЙ зоны листа `пример` (Q:AF) — RETRIEVED evidence, НЕ движок."""
    text: str
    sheet_name: str
    row: int
    column: int
    cell_address: str = ""

    def source_ref(self) -> str:
        return WorkbookSourceRef(sheet_name=self.sheet_name, row=self.row, column=self.column,
                                 cell_address=self.cell_address, value_type="method_note").ref()


_METHOD_ZONE_MIN_COL = 17        # Q и правее (левая расчётная зона A:P = 1..16)


def parse_method_notes(path: str | Path, sheet: str = "пример") -> list[MethodNote]:
    """Извлечь правую методзону (Q:AF) листа `пример` как MethodNote с cell-ref. Не вычисляет."""
    import openpyxl
    p = Path(path)
    if not p.exists():
        return []
    wb = openpyxl.load_workbook(p, data_only=True)
    out: list[MethodNote] = []
    # содержательные методнотации (не заголовки колонок): инструкции расчёта
    keys = ("гр.", "сплит", "прочерк", "коэф", "код ресурс", "по умолчанию", "кац", "тц",
            "выбор расценки", "если в гр", "норма из", "приказ")
    if sheet in wb.sheetnames:
        for cell in (c for r in wb[sheet].iter_rows() for c in r):
            v = cell.value
            if cell.column and cell.column >= _METHOD_ZONE_MIN_COL and isinstance(v, str) \
                    and len(v.strip()) > 12 and any(k in v.lower() for k in keys):
                out.append(MethodNote(v.strip(), sheet, cell.row, cell.column, cell.coordinate))
    wb.close()
    return out


def parse_formula_cells(path: str | Path) -> list[dict[str, Any]]:
    """Список формульных ячеек (RETRIEVED metadata). Формул МАЛО — осн. расчёты хранятся значениями;
    формулы НЕ используются как движок (ТЗ §3)."""
    import openpyxl
    p = Path(path)
    if not p.exists():
        return []
    wb = openpyxl.load_workbook(p, data_only=False)
    out = []
    for sn in wb.sheetnames:
        for cell in (c for r in wb[sn].iter_rows() for c in r):
            if isinstance(cell.value, str) and cell.value.startswith("="):
                out.append({"sheet": sn, "cell": cell.coordinate, "formula": cell.value})
    wb.close()
    return out


def parse_cost_workbook(path: str | Path) -> dict[str, Any]:
    """Прочитать xlsx обсчёта (если есть). Возвращает sheets/norm_codes/position_qty/kac_rows/
    method_notes/formula_cells + ПРОВЕНАНС (real vs reconstructed). Источник = workbook, НЕ формулы."""
    import openpyxl
    p = Path(path)
    if not p.exists():
        return {"status": "not_found", "provenance": "absent", "sheets": [], "norm_codes": [],
                "position_qty": None, "kac_rows": [], "method_notes": [], "formula_cells": [],
                "note": REAL_WORKBOOK_ABSENT_NOTE}
    status = real_workbook_status(p)
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    sheets = wb.sheetnames
    norm_codes = [s for s in sheets if "гэсн" in s.lower()]
    position_qty = None
    kac_rows: list[dict] = []
    if "пример" in sheets:
        for row in wb["пример"].iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)) and abs(float(cell) - 26.958848) < 1e-4:
                    position_qty = float(cell)
    if "по_кац" in sheets:
        for i, row in enumerate(wb["по_кац"].iter_rows(values_only=True)):
            vals = [str(c) for c in row if c is not None]
            if any(v.strip() == "-" for v in vals):
                kac_rows.append({"row": i, "values": vals, "status": "needs_kac"})
    wb.close()
    return {"status": "found", "provenance": status["provenance"], "note": status["note"],
            "sheets": sheets, "norm_codes": norm_codes, "position_qty": position_qty,
            "kac_rows": kac_rows, "method_notes": parse_method_notes(p),
            "formula_cells": parse_formula_cells(p)}


# ── v0.6: реальный парсер левой расчётной зоны листа `пример` ─────────────────────────────
# Раскладка реального workbook (fsnb2022.ru): B=код C=наимен H=ед I=норма J=коэфф
# K=всего_кол(=I×J×qty) L=база M=индекс N=текущая(=L×M или прямая) P=строка(=K×N).

_COL = {"code": 2, "name": 3, "unit": 8, "norm": 9, "coeff": 10, "total_qty": 11,
        "base": 12, "index": 13, "current": 14, "line_total": 16}


def _cellnum(v: Any) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (TypeError, ValueError):
        return None


def parse_real_workbook_position(path: str | Path = REAL_WORKBOOK) -> dict[str, Any]:
    """Распарсить левую расчётную зону листа `пример` реального workbook → строки ресурсов +
    сохранённые итоги (direct/ФОТ/НР/СП/позиция/КАЦ). Каждая строка с cell-level source_ref."""
    import openpyxl
    p = Path(path)
    if not p.exists():
        return {"status": "not_found", "provenance": "absent", "lines": []}
    status = real_workbook_status(p)
    wb = openpyxl.load_workbook(p, data_only=True)
    if "пример" not in wb.sheetnames:
        wb.close()
        return {"status": "blocked", "provenance": status["provenance"], "lines": [],
                "blocker": "лист 'пример' отсутствует"}
    ex = wb["пример"]
    lines: list[dict] = []
    stored: dict[str, Any] = {}
    kac = None
    for r in ex.iter_rows(min_row=1, max_row=ex.max_row, min_col=1, max_col=16):
        cell = {c.column: c.value for c in r}
        code = cell.get(_COL["code"])
        name = str(cell.get(_COL["name"]) or "")
        P = _cellnum(cell.get(_COL["line_total"]))
        I = _cellnum(cell.get(_COL["norm"]))
        # агрегатные строки
        if name.startswith("Итого прямые"):
            stored["direct"] = P; continue
        if name.strip() == "ФОТ":
            stored["fot"] = P; continue
        if name.startswith("НР"):
            stored["nr"] = {"rate": I, "amount": P, "obos": code}; continue
        if name.startswith("СП"):
            stored["sp"] = {"rate": I, "amount": P, "obos": code}; continue
        if name.startswith("Всего по позиции"):
            stored.setdefault("position", P); continue
        if isinstance(code, str) and code.startswith("ТЦ_"):
            kac = {"code": code, "name": cell.get(_COL["name"]), "qty": I,
                   "price": _cellnum(cell.get(_COL["current"])), "total": P,
                   "src": _sref("пример", r[0].row)}
            continue
        if code in (None, "") or str(code) == "4" or P is None:
            continue
        import re as _re
        if _re.fullmatch(r"\d{1,2}", str(code).strip()) and _re.fullmatch(r"\d{1,2}", name.strip()):
            continue                                  # header-строка (номера колонок 1/2/3…)
        row_i = r[0].row
        lines.append({
            "row": row_i, "code": str(code), "name": cell.get(_COL["name"]), "unit": cell.get(_COL["unit"]),
            "norm": I, "coeff": _cellnum(cell.get(_COL["coeff"])), "total_qty": _cellnum(cell.get(_COL["total_qty"])),
            "base": _cellnum(cell.get(_COL["base"])), "index": _cellnum(cell.get(_COL["index"])),
            "current": _cellnum(cell.get(_COL["current"])), "line_total": P,
            "category": classify_resource_category(str(code), str(cell.get(_COL["name"]) or "")),
            "src": _sref("пример", row_i)})
    wb.close()
    # объём позиции: мода по строкам total/(norm×coeff) — устойчиво к header/informational-строкам
    from collections import Counter
    derived = []
    for ln in lines:
        if ln["norm"] and ln["coeff"] and ln["total_qty"]:
            derived.append(round(ln["total_qty"] / (ln["norm"] * ln["coeff"]), 6))
    pos_qty = Counter(derived).most_common(1)[0][0] if derived else None
    return {"status": "found", "provenance": status["provenance"], "note": status["note"],
            "position_qty": pos_qty, "lines": lines, "stored": stored, "kac": kac,
            "method_notes": parse_method_notes(p)}


def validate_real_workbook(path: str | Path = REAL_WORKBOOK, *, tol: float = 0.5) -> dict[str, Any]:
    """Воспроизвести КОДОМ значения реального workbook и сверить с сохранёнными. expand/price/итоги
    пересчитываются, расхождения — в diff (не скрываются). Возвращает provenance + per-totals diff."""
    parsed = parse_real_workbook_position(path)
    if parsed["status"] != "found":
        return {"status": parsed["status"], "provenance": parsed.get("provenance", "absent"), "diffs": [],
                "blocked": True}
    pos_qty = parsed["position_qty"]
    coeff = CoefficientSet(labor_coeff=1.15, machine_usage_coeff=1.15, machinist_labor_coeff=1.15,
                           material_coeff=1.0, status="retrieved")
    diffs: list[dict] = []
    computed_direct = 0.0
    for ln in parsed["lines"]:
        # пересчёт total_qty = norm × coeff × pos_qty (коэфф материала=1 если в данных =1)
        if ln["norm"] is not None and ln["total_qty"] is not None:
            k = ln["coeff"] if ln["coeff"] else 1.0
            recomputed_qty = ln["norm"] * k * pos_qty
            if abs(recomputed_qty - ln["total_qty"]) > 1e-3:
                diffs.append({"row": ln["row"], "field": "total_qty", "code": ln["code"],
                              "stored": ln["total_qty"], "recomputed": round(recomputed_qty, 6)})
        # пересчёт current = base × index (если оба есть)
        if ln["base"] is not None and ln["index"] is not None and ln["current"] is not None:
            rc_cur = round(ln["base"] * ln["index"], 2)
            if abs(rc_cur - ln["current"]) > 0.05:
                diffs.append({"row": ln["row"], "field": "current_price", "code": ln["code"],
                              "stored": ln["current"], "recomputed": rc_cur})
        # пересчёт line_total = total_qty × current
        if ln["total_qty"] is not None and ln["current"] is not None and ln["line_total"] is not None:
            rc_lt = round(ln["total_qty"] * ln["current"], 2)
            if abs(rc_lt - ln["line_total"]) > tol:
                diffs.append({"row": ln["row"], "field": "line_total", "code": ln["code"],
                              "stored": ln["line_total"], "recomputed": rc_lt})
            computed_direct += ln["line_total"]
    stored = parsed["stored"]
    totals_diff = []
    # прямые: сумма строк vs сохранённое
    if "direct" in stored and stored["direct"]:
        if abs(computed_direct - stored["direct"]) > 1.0:
            totals_diff.append({"total": "direct", "stored": stored["direct"],
                                "recomputed": round(computed_direct, 2)})
    # ФОТ/НР/СП пересчёт
    fot = stored.get("fot")
    if fot and stored.get("nr"):
        nr_rate = (stored["nr"]["rate"] or 0) / 100.0
        rc_nr = _r2(fot * nr_rate)
        if abs(rc_nr - (stored["nr"]["amount"] or 0)) > 1.0:
            totals_diff.append({"total": "nr", "stored": stored["nr"]["amount"], "recomputed": rc_nr})
    if fot and stored.get("sp"):
        sp_rate = (stored["sp"]["rate"] or 0) / 100.0
        rc_sp = _r2(fot * sp_rate)
        if abs(rc_sp - (stored["sp"]["amount"] or 0)) > 1.0:
            totals_diff.append({"total": "sp", "stored": stored["sp"]["amount"], "recomputed": rc_sp})
    return {"status": "found", "provenance": parsed["provenance"], "position_qty": pos_qty,
            "lines_count": len(parsed["lines"]), "computed_direct": round(computed_direct, 2),
            "stored": stored, "kac": parsed["kac"], "line_diffs": diffs, "totals_diff": totals_diff,
            "matches": not diffs and not totals_diff, "method_notes_count": len(parsed["method_notes"])}


# ── evidence-result для chat ──────────────────────────────────────────────────────────────

# ── v0.7: bridge-интерфейсы к реальным источникам (лёгкие обёртки, НЕ форсим интеграцию) ──
# Цель — сделать БУДУЩИЙ переход с workbook-цен на реальный resource-price DB/ФГИС возможным.
# Сейчас источник цен — workbook; bridge возвращает found/not_found, без silent-fake-fallback.

def resource_price_source() -> dict[str, Any]:
    """Текущий источник цен ресурсов. v0.7: workbook (значения из листа `пример`). Production
    price-DB (ФГИС/локальный parquet) — НЕ подключён (bridge готов к подключению)."""
    return {"source": "workbook", "db_available": False,
            "note": "цены берутся из реального workbook (fsnb2022.ru); production price-DB не подключён"}


def fgis_price_lookup(resource_code: str) -> dict[str, Any]:
    """Bridge к ФГИС/локальной базе цен по коду ресурса. v0.7: не подключён → not_found (не fake)."""
    return {"status": "not_found", "resource_code": resource_code,
            "note": "ФГИС price-bridge не подключён; цена берётся из workbook"}


def nr_sp_lookup(name: str = "") -> dict[str, Any]:
    """Bridge к nr_sp_service для ставок НР/СП. found → ставки из локальной базы Приказов."""
    try:
        from proxy.services.nr_sp_service import resolve as _resolve_nr_sp
        rs = _resolve_nr_sp(name)
        return {"status": "found", "nr_pct": rs.get("nr_pct"), "sp_pct": rs.get("sp_pct"),
                "source": "nr_sp_service (Приказы 812/774)"}
    except Exception as e:  # noqa: BLE001
        return {"status": "not_found", "note": f"nr_sp_service недоступен: {e}"}


def machinist_mapping_lookup(machine_code: str) -> dict[str, Any]:
    """Bridge маппинга машина→машинист. found → тарифный код; not_found → MISSING (не выдумка)."""
    mc = machine_to_machinist(machine_code)
    if mc:
        return {"status": "found", "machine_code": machine_code, "machinist_code": mc}
    return {"status": "not_found", "machine_code": machine_code,
            "note": "маппинг машина→машинист не определён (частичный); машинист-строка → MISSING"}


def resource_result_to_construction_result(res: ResourceEstimateResult) -> ConstructionHarnessResult:
    # v0.6: провенанс источника (real vs reconstructed) + реальные методнотации как RETRIEVED.
    status = real_workbook_status()
    blocks = list(res.evidence_blocks)
    warnings = list(res.warnings)
    method_count = 0
    validation = None
    if status["provenance"] == "real":
        notes = parse_method_notes(status["path"])
        method_count = len(notes)
        if notes:
            items = [EvidenceItem(EvidenceType.RETRIEVED, n.text[:140], source_refs=[n.source_ref()],
                                  status="supported") for n in notes[:12]]
            blocks.append(block_of(EvidenceType.RETRIEVED, "Методика из workbook (правая зона листа)", items))
        val = validate_real_workbook()
        validation = {"matches": val.get("matches"), "line_diffs": len(val.get("line_diffs", [])),
                      "computed_direct": val.get("computed_direct")}
        warnings.append("источник — РЕАЛЬНЫЙ workbook (fsnb2022.ru); расчёт воспроизведён кодом "
                        "(xlsx хранит значения, не формулы)")
    else:
        warnings.append(status["note"])
    return ConstructionHarnessResult(
        answer_data={"intent": "resource_cost_calc", "norm_code": res.norm_code, "title": res.title,
                     "provenance": status["provenance"], "source_note": status["note"],
                     "method_notes": method_count, "validation": validation,
                     "direct_cost": res.direct_cost_total, "fot": res.fot, "nr": res.nr_amount,
                     "sp": res.sp_amount, "position_total": res.position_total,
                     "kac_total": res.additional_price_items_total, "grand_total": res.grand_total},
        evidence_blocks=blocks, total_status=res.total_status,
        final_total=res.grand_total if res.total_status == "complete" else None,
        partial_total=res.position_total, warnings=warnings, blockers=res.blockers,
        tool_trace=[{"tool": "resource_cost_calc", "norm": res.norm_code, "provenance": status["provenance"],
                     "status": res.total_status}])

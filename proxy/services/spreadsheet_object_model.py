"""Structural spreadsheet object model (skeleton).

Excel is not Markdown. Overview is cheap (sheet names / used ranges / sample);
detail and formula graphs come later via openpyxl / optional calamine / COM.
"""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

SPREADSHEET_OBJECT_SCHEMA = "les.spreadsheet_object.v1"
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def workbook_overview(path: str | Path, *, sample_rows: int = 5) -> dict[str, Any]:
    """Return a typed workbook overview without dumping the whole book to text."""
    source = Path(path).expanduser().resolve()
    if source.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise ValueError(f"unsupported spreadsheet suffix: {source.suffix}")
    if not source.is_file():
        raise FileNotFoundError(str(source))

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is required in LES
        raise RuntimeError("openpyxl is required for spreadsheet overview") from exc

    workbook = openpyxl.load_workbook(source, data_only=False, read_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for title in workbook.sheetnames:
            sheet = workbook[title]
            used = _used_range(sheet)
            sample = _sample_rows(sheet, used, limit=sample_rows)
            sheets.append({
                "name": title,
                "state": str(getattr(sheet, "sheet_state", "visible") or "visible"),
                "used_range": used,
                "sample_rows": sample,
                "formula_cells_sampled": sum(
                    1 for row in sample for cell in row
                    if isinstance(cell.get("raw_value"), str)
                    and str(cell.get("raw_value")).startswith("=")
                ),
            })
    finally:
        workbook.close()

    return {
        "schema": SPREADSHEET_OBJECT_SCHEMA,
        "kind": "workbook_overview",
        "path": str(source),
        "sha256": _sha256(source),
        "parser": "openpyxl.read_only",
        "sheet_count": len(sheets),
        "sheets": sheets,
        "provenance": {
            "file": str(source),
            "version": _sha256(source)[:16],
            "parser": "openpyxl.read_only",
        },
    }


def formula_dependency_stub(formula: str, *, sheet: str = "") -> dict[str, Any]:
    """Parse a single formula into a dependency list (no evaluation)."""
    text = str(formula or "").strip()
    if text.startswith("="):
        text = text[1:]
    refs = _extract_cell_refs(text)
    return {
        "schema": SPREADSHEET_OBJECT_SCHEMA,
        "kind": "formula_card",
        "sheet": sheet,
        "formula": formula,
        "depends_on": [f"{sheet}!{ref}" if sheet and "!" not in ref else ref for ref in refs],
        "evaluated": False,
        "engine": "none",
    }


def _used_range(sheet: Any) -> dict[str, Any]:
    dims = getattr(sheet, "dimensions", None) or "A1"
    min_row = getattr(sheet, "min_row", None) or 1
    max_row = getattr(sheet, "max_row", None) or 1
    min_col = getattr(sheet, "min_column", None) or 1
    max_col = getattr(sheet, "max_column", None) or 1
    return {
        "dimensions": str(dims),
        "min_row": int(min_row),
        "max_row": int(max_row),
        "min_col": int(min_col),
        "max_col": int(max_col),
    }


def _sample_rows(sheet: Any, used: dict[str, Any], *, limit: int) -> list[list[dict[str, Any]]]:
    rows: list[list[dict[str, Any]]] = []
    max_row = min(int(used["max_row"]), int(used["min_row"]) + max(0, limit) - 1)
    max_col = min(int(used["max_col"]), int(used["min_col"]) + 15)
    from openpyxl.utils import get_column_letter

    for row_idx, row in enumerate(
        sheet.iter_rows(
            min_row=int(used["min_row"]),
            max_row=max_row,
            min_col=int(used["min_col"]),
            max_col=max_col,
            values_only=False,
        ),
        start=int(used["min_row"]),
    ):
        cells: list[dict[str, Any]] = []
        for col_offset, cell in enumerate(row):
            col_idx = int(used["min_col"]) + col_offset
            address = str(
                getattr(cell, "coordinate", None)
                or f"{get_column_letter(col_idx)}{row_idx}"
            )
            cells.append({
                "address": address,
                "raw_value": getattr(cell, "value", None),
                "number_format": str(getattr(cell, "number_format", "") or ""),
            })
        rows.append(cells)
        if row_idx >= max_row:
            break
    return rows


def _extract_cell_refs(formula: str) -> list[str]:
    import re

    pattern = re.compile(
        r"(?:(?:'[^']+'|[\w.]+)!)?"
        r"\$?[A-Za-z]{1,3}\$?\d{1,7}"
        r"(?::\$?[A-Za-z]{1,3}\$?\d{1,7})?"
    )
    found: list[str] = []
    for match in pattern.findall(formula or ""):
        if match not in found:
            found.append(match)
    return found[:64]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()

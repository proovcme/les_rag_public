"""Render a ready RIM trace with session requirements and immutable audit."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from proxy.services.rim_trace_xlsx_service import render_lsr_xlsx


def _append_requirement_sheet(workbook, requirements: list[dict[str, Any]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    if "Недостающие данные" in workbook.sheetnames:
        del workbook["Недостающие данные"]
    sheet = workbook.create_sheet("Недостающие данные")
    headers = (
        "ID",
        "Тип",
        "Важность",
        "Блокировка финала",
        "work_id",
        "Код ресурса",
        "Описание",
        "Необходимые поля",
        "Статус",
        "Источники",
    )
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    for item in requirements:
        sheet.append(
            [
                item.get("requirement_id"),
                item.get("kind"),
                item.get("severity"),
                item.get("finality_policy"),
                item.get("work_id"),
                item.get("resource_code"),
                item.get("description"),
                ", ".join(str(value) for value in (item.get("required_fields") or [])),
                item.get("status"),
                "\n".join(str(value) for value in (item.get("source_refs") or [])),
            ]
        )
    widths = (36, 24, 14, 22, 18, 24, 64, 40, 18, 52)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"


def _append_audit_sheet(workbook, audit: dict[str, Any], *, is_final: bool) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    if "Аудит" in workbook.sheetnames:
        del workbook["Аудит"]
    sheet = workbook.create_sheet("Аудит")
    session = audit.get("session") if isinstance(audit.get("session"), dict) else {}
    sheet.append(["Статус файла", "ФИНАЛЬНЫЙ" if is_final else "ЧЕРНОВИК"])
    sheet.append(["session_id", session.get("session_id")])
    sheet.append(["project_id", session.get("project_id")])
    sheet.append(["ВОР revision", session.get("current_vor_revision_id")])
    sheet.append(["Mapping revision", session.get("current_mapping_revision_id")])
    sheet.append(["Mapping lock", session.get("mapping_lock_revision_id")])
    sheet.append(["Scenario revision", session.get("current_scenario_revision_id")])
    sheet.append(["Pricing revision", session.get("current_pricing_revision_id")])
    sheet.append(["Final lock", session.get("final_lock_revision_id")])
    sheet.append(["Нормативная база", session.get("normative_base_version")])
    sheet.append(["Ценовая книга", session.get("pricebook_id")])
    sheet.append(["Регион", session.get("region_code")])
    sheet.append(["Период", session.get("price_period")])
    sheet.append([])
    sheet.append(
        [
            "revision_id",
            "parent_revision_id",
            "Тип ревизии",
            "Автор",
            "Дата",
            "SHA-256 payload",
        ]
    )
    header_row = sheet.max_row
    for item in audit.get("revisions") or []:
        sheet.append(
            [
                item.get("revision_id"),
                item.get("parent_revision_id"),
                item.get("revision_kind"),
                item.get("created_by"),
                item.get("created_at"),
                item.get("payload_sha256"),
            ]
        )
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDE8DF")
    sheet["A1"].font = Font(bold=True, color=("006100" if is_final else "9C6500"))
    sheet["B1"].font = Font(bold=True, color=("006100" if is_final else "9C6500"))
    widths = {"A": 38, "B": 38, "C": 28, "D": 24, "E": 28, "F": 68}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = f"A{header_row + 1}"


def render_session_lsr_xlsx(
    trace: dict[str, Any],
    requirements: list[dict[str, Any]],
    audit: dict[str, Any],
    out_path: str | Path,
    *,
    is_final: bool,
) -> Path:
    """Render the existing trace; never recalculate or invoke a model."""
    import openpyxl

    target = Path(out_path)
    session = audit.get("session") if isinstance(audit.get("session"), dict) else {}
    meta = {
        "subject": session.get("region_code") or "—",
        "price_level": session.get("price_period") or "—",
        "osnovanie": f"RIM session {session.get('session_id') or '—'}",
    }
    render_lsr_xlsx(trace, target, meta=meta)
    workbook = openpyxl.load_workbook(target)
    try:
        _append_requirement_sheet(workbook, requirements)
        _append_audit_sheet(workbook, audit, is_final=is_final)
        workbook.save(target)
    finally:
        workbook.close()
    return target


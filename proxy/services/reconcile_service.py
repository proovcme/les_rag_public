"""Сверка количеств между типами строительных документов — W11.4 (LES3_PLAN).

Топ-кейс ГИП: ВОР ↔ КС-2 ↔ смета (ЛС) ↔ ИД должны сходиться по объёмам. Этот
сервис строит кросс-таблицу позиций (наименование × единица) и сравнивает
количество из каждого типа документа, помечая расхождения.

ADR-11: ноль LLM. Источник — нормализованные Parquet-строки табличного конвейера
(`backend/parquet_writer.py`, STANDARD_SCHEMA). Числа сводит и сравнивает Python —
детерминированно, бит-в-бит. LLM не участвует.

Кирпичи переиспользованы из `bor_service` (нормализация единиц/наименований) и
`plan_fact_service` (сопоставление наименований по равенству/вхождению).

Статусы строки сверки:
- ``match``     — позиция есть в нескольких документах и количества сходятся;
- ``mismatch``  — позиция в нескольких документах, но количества расходятся (> допуска);
- ``gap``       — позиция есть в части документов, в других отсутствует (пробел);
- ``single``    — позиция встретилась только в одном документе.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from backend.parquet_writer import DOC_TYPES
from proxy.services.bor_service import _normalize_name, normalize_unit, rows_from_parquet

logger = logging.getLogger(__name__)

# Порог длины короткой строки для сопоставления по вхождению (как в plan_fact).
_MIN_CONTAINS_LEN = 4
# Допуски сравнения количеств: абсолютный + относительный (1 %).
_ABS_TOL = 1e-6
_REL_TOL = 0.01

# Поля количества по приоритету: qty-приоритет + data-aware fallback (как table_query).
# КС-2 кладёт объём в work_done/work_since_start/work_volume, спеки/сметы — в qty.
_QTY_FIELDS = ("qty", "work_done", "work_since_start", "work_volume")

# Человекочитаемые ярлыки типов документов для шапки кросс-таблицы.
DOC_TYPE_LABELS = {
    "SPEC": "Спецификация",
    "SMETA": "Смета",
    "KS2": "КС-2",
    "VEDOMOST": "Ведомость",
    "AOSR": "АОСР (ИД)",
    "TABLE": "Таблица",
}


def doc_type_label(doc_type: str) -> str:
    return DOC_TYPE_LABELS.get(doc_type, DOC_TYPES.get(doc_type, doc_type or "—"))


def _row_qty(row: dict) -> float | None:
    """qty-приоритет + data-aware fallback по полям объёма (КС-2 и т.п.)."""
    for key in _QTY_FIELDS:
        val = row.get(key)
        if val is None:
            continue
        try:
            return float(val)
        except (TypeError, ValueError):
            continue
    return None


def _names_match(a: str, b: str) -> bool:
    """Равенство нормализованных наименований либо вхождение при длине ≥ порога."""
    if not a or not b:
        return False
    if a == b:
        return True
    shorter = min(a, b, key=len)
    if len(shorter) < _MIN_CONTAINS_LEN:
        return False
    return a in b or b in a


# ── сбор источников ──

def collect_rows_by_doc_type(
    dataset_id: str,
    storage_root: Path = Path("storage/datasets"),
) -> dict[str, list[dict]]:
    """Все строки таблиц датасета (Parquet) с наименованием, сгруппированные по doc_type."""
    parquet_root = storage_root / dataset_id / "_parquet"
    by_type: dict[str, list[dict]] = {}
    if not parquet_root.exists():
        return by_type
    for parquet_path in sorted(parquet_root.rglob("*.parquet")):
        for row in rows_from_parquet(parquet_path):
            name = str(row.get("name") or row.get("work_name") or "").strip()
            if not name:
                continue
            doc_type = str(row.get("doc_type") or "TABLE").strip() or "TABLE"
            by_type.setdefault(doc_type, []).append(row)
    return by_type


def aggregate_positions(rows: list[dict]) -> dict[tuple[str, str], dict[str, Any]]:
    """Свод строк одного источника по (наименование, единица): сумма количеств.

    Возвращает {(name_norm, unit): {name, unit, qty, qty_missing_rows, source_rows}}.
    Наименование в выдаче — самое длинное (информативное) из встреченных вариантов.
    """
    agg: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        name = " ".join(str(row.get("name") or row.get("work_name") or "").split())
        if not name:
            continue
        unit = normalize_unit(row.get("unit"))
        key = (_normalize_name(name), unit)
        entry = agg.get(key)
        if entry is None:
            entry = {"name": name, "unit": unit, "qty": None, "qty_missing_rows": 0, "source_rows": 0}
            agg[key] = entry
        if len(name) > len(entry["name"]):
            entry["name"] = name
        qty = _row_qty(row)
        if qty is None:
            entry["qty_missing_rows"] += 1
        else:
            entry["qty"] = (entry["qty"] or 0.0) + qty
        entry["source_rows"] += 1
    return agg


# ── сравнение количеств ──

def _qty_status(values: list[float]) -> tuple[bool, float, float | None]:
    """Сходятся ли количества. Возвращает (agree, max_delta, delta_pct)."""
    if len(values) < 2:
        return True, 0.0, 0.0
    lo, hi = min(values), max(values)
    max_delta = hi - lo
    base = max(abs(hi), abs(lo))
    agree = max_delta <= max(_ABS_TOL, _REL_TOL * base)
    delta_pct = round(max_delta / base * 100.0, 1) if base > _ABS_TOL else None
    return agree, round(max_delta, 4), delta_pct


def reconcile_sources(sources: dict[str, list[dict]]) -> dict[str, Any]:
    """Свести позиции из разных типов документов в кросс-таблицу с флагами расхождений.

    `sources` — {doc_type: [нормализованные строки]}. Жадная кластеризация позиций
    по единице + сопоставлению наименований (равенство/вхождение); один источник
    в кластере суммируется. Без LLM.
    """
    doc_types = [dt for dt, rows in sources.items() if rows]
    # Позиции каждого источника, развёрнутые в плоский список для кластеризации.
    items: list[dict[str, Any]] = []
    for doc_type in doc_types:
        for (name_norm, unit), entry in aggregate_positions(sources[doc_type]).items():
            items.append({
                "doc_type": doc_type, "norm": name_norm, "unit": unit,
                "name": entry["name"], "qty": entry["qty"],
                "qty_missing_rows": entry["qty_missing_rows"], "source_rows": entry["source_rows"],
            })
    # Детерминированный порядок кластеризации.
    items.sort(key=lambda it: (it["unit"], it["norm"], it["doc_type"]))

    clusters: list[dict[str, Any]] = []
    for it in items:
        placed = None
        for cl in clusters:
            if cl["unit"] != it["unit"]:
                continue
            if any(_names_match(it["norm"], n) for n in cl["norms"]):
                placed = cl
                break
        if placed is None:
            placed = {"unit": it["unit"], "norms": [], "name": it["name"], "by_source": {}}
            clusters.append(placed)
        placed["norms"].append(it["norm"])
        if len(it["name"]) > len(placed["name"]):
            placed["name"] = it["name"]
        bs = placed["by_source"].setdefault(
            it["doc_type"], {"qty": None, "qty_missing_rows": 0, "source_rows": 0}
        )
        if it["qty"] is not None:
            bs["qty"] = (bs["qty"] or 0.0) + it["qty"]
        bs["qty_missing_rows"] += it["qty_missing_rows"]
        bs["source_rows"] += it["source_rows"]

    rows = [_make_reconcile_row(cl, doc_types) for cl in clusters]
    rows.sort(key=lambda r: (_STATUS_ORDER.get(r["status"], 9), r["name"].casefold()))
    return {
        "doc_types": doc_types,
        "doc_type_labels": {dt: doc_type_label(dt) for dt in doc_types},
        "rows": rows,
        "totals": _totals(rows),
    }


_STATUS_ORDER = {"mismatch": 0, "gap": 1, "single": 2, "match": 3}


def _make_reconcile_row(cluster: dict, all_doc_types: list[str]) -> dict[str, Any]:
    by_source = cluster["by_source"]
    present = [dt for dt in all_doc_types if dt in by_source]
    missing = [dt for dt in all_doc_types if dt not in by_source]
    qty_values = [by_source[dt]["qty"] for dt in present if by_source[dt]["qty"] is not None]
    agree, max_delta, delta_pct = _qty_status(qty_values)

    if len(all_doc_types) <= 1:
        status = "single"  # один тип документа во своде — сравнивать не с чем
    elif not agree:
        status = "mismatch"  # числа расходятся — главный сигнал, важнее пробела
    elif missing:
        status = "gap"  # есть в части документов, в других отсутствует
    else:
        status = "match"

    return {
        "name": cluster["name"],
        "unit": cluster["unit"],
        "qty_by_source": {dt: round(by_source[dt]["qty"], 4) if by_source[dt]["qty"] is not None else None
                          for dt in present},
        "present": present,
        "missing": missing,
        "max_delta": max_delta,
        "delta_pct": delta_pct,
        "status": status,
    }


def _totals(rows: list[dict]) -> dict[str, Any]:
    return {
        "lines": len(rows),
        "match": sum(1 for r in rows if r["status"] == "match"),
        "mismatch": sum(1 for r in rows if r["status"] == "mismatch"),
        "gap": sum(1 for r in rows if r["status"] == "gap"),
        "single": sum(1 for r in rows if r["status"] == "single"),
    }


# ── полный цикл по датасетам ──

def reconcile_datasets(
    dataset_ids: list[str],
    *,
    storage_root: Path = Path("storage/datasets"),
    output_dir: Path | None = None,
    by: str = "doc_type",
    dataset_names: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Свести позиции по одному/нескольким датасетам. Без LLM (ADR-11).

    ``by="doc_type"`` (по умолчанию) — ось источника = тип документа (КС-2/смета/…);
    датасеты одного типа сливаются (например, КС-2 за разные периоды).
    ``by="dataset"`` — ось источника = сам документ/датасет (для «сверь ведомость X с
    актом Y», когда оба — одного типа). Ярлык берётся из ``dataset_names``.
    """
    names = dataset_names or {}
    sources: dict[str, list[dict]] = {}
    for dataset_id in dataset_ids:
        by_type = collect_rows_by_doc_type(dataset_id, storage_root=storage_root)
        if by == "dataset":
            label = names.get(dataset_id, dataset_id)
            bucket = sources.setdefault(label, [])
            for rows in by_type.values():
                bucket.extend(rows)
        else:
            for doc_type, rows in by_type.items():
                sources.setdefault(doc_type, []).extend(rows)

    result = reconcile_sources(sources)
    result.update({
        "dataset_ids": dataset_ids,
        "source_rows": sum(len(rows) for rows in sources.values()),
        "xlsx_path": None,
    })
    if output_dir is not None and result["rows"]:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        tag = "_".join(dataset_ids)[:40] or "reconcile"
        xlsx_path = output_dir / f"reconcile_{tag}_{stamp}.xlsx"
        reconcile_to_xlsx(result, xlsx_path)
        result["xlsx_path"] = str(xlsx_path)
    return result


_STATUS_LABEL = {
    "match": "сходится", "mismatch": "РАСХОЖДЕНИЕ",
    "gap": "пробел (нет в части док-тов)", "single": "только в одном",
}


def reconcile_to_xlsx(result: dict[str, Any], output_path: Path) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    doc_types = result["doc_types"]
    labels = result["doc_type_labels"]
    rows = result["rows"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Сверка"
    ws.append(["Сверка количеств между документами"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Сформировано {datetime.now():%Y-%m-%d %H:%M} · позиций: {len(rows)} · "
               f"источники: {', '.join(labels[dt] for dt in doc_types)} · числа — Parquet, без LLM"])
    ws.append([])
    headers = ["№", "Наименование", "Ед."] + [labels[dt] for dt in doc_types] + \
              ["Δ max", "Δ %", "Статус"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    for idx, r in enumerate(rows, 1):
        qty_cells = [
            (round(r["qty_by_source"][dt], 4) if r["qty_by_source"].get(dt) is not None
             else ("—" if dt in r["present"] else ""))
            for dt in doc_types
        ]
        ws.append([
            idx, r["name"], r["unit"], *qty_cells,
            r["max_delta"] if r["status"] in ("mismatch", "match", "gap") else "—",
            r["delta_pct"] if r["delta_pct"] is not None else "—",
            _STATUS_LABEL.get(r["status"], r["status"]),
        ])

    base_widths = {"A": 6, "B": 50, "C": 8}
    for col, width in base_widths.items():
        ws.column_dimensions[col].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("[RECONCILE] %s позиций → %s", len(rows), output_path)
    return len(rows)

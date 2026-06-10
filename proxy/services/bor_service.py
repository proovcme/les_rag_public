"""Генератор ведомостей объёмов работ (ВОР) из спецификаций — W11.1 (LES3_PLAN).

ADR-11: ядро без LLM. Источник — нормализованные Parquet-строки табличного
конвейера (backend/parquet_writer.py, STANDARD_SCHEMA) в
storage/datasets/{dataset_id}/_parquet/**.parquet. Группировка, свод
количеств и экспорт в xlsx — детерминированные.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

from backend.parquet_writer import load_parquet

logger = logging.getLogger(__name__)

# Типы документов, из которых строится ВОР (см. DOC_TYPES в parquet_writer).
BOR_SOURCE_DOC_TYPES = ("SPEC", "VEDOMOST")

# Нормализация единиц измерения: разнобой написания → каноническая форма.
UNIT_ALIASES = {
    "шт": "шт", "шт.": "шт", "штук": "шт", "штука": "шт", "штуки": "шт",
    "компл": "компл", "компл.": "компл", "к-т": "компл", "комплект": "компл", "комплектов": "компл",
    "м": "м", "м.": "м", "метр": "м", "метров": "м",
    "пог.м": "пог.м", "п.м": "пог.м", "п.м.": "пог.м", "пог. м": "пог.м", "пм": "пог.м",
    "м2": "м²", "м²": "м²", "кв.м": "м²", "кв. м": "м²", "кв.м.": "м²",
    "м3": "м³", "м³": "м³", "куб.м": "м³", "куб. м": "м³", "куб.м.": "м³",
    "кг": "кг", "кг.": "кг",
    "т": "т", "т.": "т", "тн": "т", "тонн": "т",
    "л": "л", "л.": "л", "литр": "л", "литров": "л",
}


def normalize_unit(unit: str | None) -> str:
    if not unit:
        return ""
    cleaned = str(unit).strip().lower().replace("ё", "е")
    return UNIT_ALIASES.get(cleaned, cleaned)


def _normalize_name(name: str) -> str:
    """Ключ группировки: схлопнутые пробелы, без регистра."""
    return re.sub(r"\s+", " ", str(name)).strip().casefold()


@dataclass
class BorLine:
    """Одна строка ВОР — свод одинаковых позиций по всем источникам."""

    section: str
    name: str
    code: str
    mark: str
    unit: str
    qty: float | None          # None — ни в одной исходной строке не было количества
    qty_missing_rows: int = 0  # строк-источников без количества (вошли в свод без суммы)
    source_rows: int = 0
    sources: list[str] = field(default_factory=list)

    def payload(self) -> dict:
        return {
            "section": self.section,
            "name": self.name,
            "code": self.code,
            "mark": self.mark,
            "unit": self.unit,
            "qty": self.qty,
            "qty_missing_rows": self.qty_missing_rows,
            "source_rows": self.source_rows,
            "sources": self.sources,
        }


def rows_from_parquet(parquet_path: Path) -> list[dict]:
    """load_parquet отдаёт колоночный dict — разворачиваем в строки."""
    columns = load_parquet(str(parquet_path))
    if not columns:
        return []
    keys = list(columns.keys())
    length = len(columns[keys[0]]) if keys else 0
    return [{key: columns[key][i] for key in keys} for i in range(length)]


def collect_spec_rows(
    dataset_id: str,
    storage_root: Path = Path("storage/datasets"),
    doc_types: tuple[str, ...] = BOR_SOURCE_DOC_TYPES,
) -> list[dict]:
    """Все строки спецификаций/ведомостей датасета с наименованием."""
    parquet_root = storage_root / dataset_id / "_parquet"
    rows: list[dict] = []
    if not parquet_root.exists():
        return rows
    for parquet_path in sorted(parquet_root.rglob("*.parquet")):
        for row in rows_from_parquet(parquet_path):
            if row.get("doc_type") not in doc_types:
                continue
            name = str(row.get("name") or row.get("work_name") or "").strip()
            if not name:
                continue
            rows.append(row)
    return rows


def build_bor(rows: list[dict]) -> list[BorLine]:
    """Свод: группировка по (раздел, наименование, код, марка, ед.изм.), сумма qty."""
    lines: dict[tuple, BorLine] = {}
    for row in rows:
        name = str(row.get("name") or row.get("work_name") or "").strip()
        section = str(row.get("section") or "").strip()
        code = str(row.get("code") or "").strip()
        mark = str(row.get("mark") or "").strip()
        unit = normalize_unit(row.get("unit"))
        key = (section.casefold(), _normalize_name(name), code.casefold(), mark.casefold(), unit)

        line = lines.get(key)
        if line is None:
            line = BorLine(section=section, name=re.sub(r"\s+", " ", name), code=code, mark=mark, unit=unit, qty=None)
            lines[key] = line

        qty = row.get("qty")
        if qty is not None:
            try:
                line.qty = (line.qty or 0.0) + float(qty)
            except (TypeError, ValueError):
                line.qty_missing_rows += 1
        else:
            line.qty_missing_rows += 1

        line.source_rows += 1
        source = str(row.get("source_file") or "").strip()
        pos = str(row.get("pos") or row.get("position") or "").strip()
        ref = f"{source}#{pos}" if pos else source
        if ref and ref not in line.sources:
            line.sources.append(ref)

    return sorted(lines.values(), key=lambda l: (l.section.casefold(), l.name.casefold(), l.code))


def bor_to_xlsx(lines: list[BorLine], output_path: Path, title: str = "Ведомость объёмов работ") -> int:
    """Экспорт ВОР в xlsx. Возвращает число строк."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "ВОР"

    ws.append([title])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    headers = ["№", "Раздел", "Наименование", "Код/Шифр", "Марка", "Ед. изм.", "Кол-во", "Строк-источников", "Источники"]
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    for idx, line in enumerate(lines, 1):
        qty_cell = round(line.qty, 4) if line.qty is not None else "—"
        ws.append([
            idx,
            line.section,
            line.name + (f" (без кол-ва: {line.qty_missing_rows} строк)" if line.qty_missing_rows else ""),
            line.code,
            line.mark,
            line.unit,
            qty_cell,
            line.source_rows,
            "; ".join(line.sources[:10]),
        ])

    widths = {"A": 6, "B": 18, "C": 60, "D": 14, "E": 14, "F": 10, "G": 12, "H": 10, "I": 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("[BOR] %s строк → %s", len(lines), output_path)
    return len(lines)


def generate_bor(
    dataset_id: str,
    storage_root: Path = Path("storage/datasets"),
    output_dir: Path | None = None,
    title: str | None = None,
) -> dict:
    """Полный цикл: parquet датасета → свод → xlsx. Без LLM (ADR-11)."""
    from datetime import datetime

    rows = collect_spec_rows(dataset_id, storage_root=storage_root)
    lines = build_bor(rows)
    result: dict = {
        "dataset_id": dataset_id,
        "source_rows": len(rows),
        "bor_lines": len(lines),
        "lines": [line.payload() for line in lines],
        "xlsx_path": None,
    }
    if output_dir is not None and lines:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = output_dir / f"bor_{dataset_id}_{stamp}.xlsx"
        bor_to_xlsx(lines, xlsx_path, title=title or f"Ведомость объёмов работ — {dataset_id}")
        result["xlsx_path"] = str(xlsx_path)
    return result

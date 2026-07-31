"""Checklist-review report renderer (Glorax, Phase 4, T4.2) — XLSX/HTML над ``checklist_review_v1``.

Вход — ``run_payload``: dict-контракт ``checklist_review_v1`` (тот же формат, что возвращает
``checklist_review_service.run_checklist_review`` и что читает роутер из
``storage/checklist_review/{dataset_id}/{run_id}/run.json`` после применения decisions
``_apply_decisions``). Модуль НЕ считает ничего сам — только рендерит уже посчитанный результат
(0 LLM, 0 обращений к Qdrant/MetaDB/файловой системе датасета).

Паттерн — по образцу ``doc_review_service.review_to_xlsx``/``review_to_html`` (openpyxl,
минимальные стили: жирные заголовки, ширины колонок, лист ``normalized_remarks``), но:
  - вход dict, а не ``list[ReviewItem]``/``ReviewMap`` — контракт checklist-review шире и уже
    содержит ``discipline``/``section_path``/``human_answer`` на каждом item'е;
  - лист «Чек-лист» повторяет форму исходного чек-листа Glorax (Раздел/№/Критерий) + системные
    колонки, сгруппирован по ``discipline`` со строкой-заголовком раздела (implementation_plan.md
    §5, docs/CHECKLIST_REVIEW_PD_TASK.md §8);
  - есть отдельный лист «Сводка» (total/by_status/suggested/human/without_evidence/незакрытые
    manual_required item_id) — в doc_review таких сводных агрегатов на отдельном листе нет,
    section добавлена по прямому требованию промпта T4.2.

Инвариант (implementation_plan.md §7 правило 10, CHECKLIST_REVIEW_PD_TASK.md §13): полного
автоматического вердикта «соответствует/не соответствует» нет by design — ни XLSX, ни HTML не
должны содержать такую строку (см. явный тест ``test_report_to_xlsx_has_no_overall_verdict_string``/
``test_report_to_html_has_no_overall_verdict_string`` в ``tests/test_checklist_report_service.py``).
"""

from __future__ import annotations

import html as html_module
from pathlib import Path
from typing import Any

# Русские ярлыки suggested_answer (промпт T4.2 п.1): маппинг машинных значений контракта
# checklist_review_v1 на человекочитаемые метки колонки «Предложенный ответ».
_ANSWER_LABELS_RU: dict[str, str] = {
    "yes": "Да",
    "no": "Нет",
    "not_required": "Не требуется",
    "unknown": "—",
    "manual_required": "Требует инженера",
}

_MAX_EVIDENCE_REFS = 5

# review-fix 2026-07-07: в колонке A лежит item_id (PD-KR-020), а раздел показан
# bold-строками групп — заголовок не должен врать («Раздел» -> «ID пункта»).
_CHECKLIST_HEADER = [
    "ID пункта", "№", "Критерий", "Предложенный ответ", "Уверенность",
    "Evidence", "Примечание системы", "Ответ инженера", "Примечание инженера",
]

_STATUS_RU = {
    "computed_issue": "Замечание к решению",
    "supported_by_evidence": "Подтверждено источниками",
    "not_applicable": "Неприменимо",
    "manual_required": "Ручная проверка",
    "review_needed": "Нужны источники",
}


def _answer_label(value: str) -> str:
    return _ANSWER_LABELS_RU.get(value, value or "—")


def _evidence_text(item: dict[str, Any]) -> str:
    """Evidence source_refs через перенос строки, максимум 5 (промпт T4.2 п.1)."""
    refs = [
        str((ev or {}).get("source_ref") or "").strip()
        for ev in item.get("document_evidence") or []
    ]
    refs = [r for r in refs if r][:_MAX_EVIDENCE_REFS]
    return "\n".join(refs)


def _discipline_of(item: dict[str, Any]) -> str:
    return str(item.get("discipline") or "")


def _is_open_manual_required(item: dict[str, Any]) -> bool:
    """«Незакрытые manual_required» для сводки: suggested_answer=manual_required (движок не смог
    решить формально) И human_decision всё ещё unset (инженер решения не принял)."""
    return (
        item.get("suggested_answer") == "manual_required"
        and item.get("human_decision", "unset") == "unset"
    )


# ── report_to_xlsx ──────────────────────────────────────────────────────────────────────────


def report_to_xlsx(run_payload: dict[str, Any], output_path: str | Path) -> int:
    """Рендерит ``run_payload`` (``checklist_review_v1``) в XLSX по образцу
    ``doc_review_service.review_to_xlsx``: лист «Чек-лист» (форма исходника + системные колонки,
    сгруппировано по discipline), лист «Сводка», лист «normalized_remarks».

    Возвращает число items в отчёте (тот же паттерн возврата, что ``review_to_xlsx``).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    output_path = Path(output_path)
    items = run_payload.get("items") or []
    summary = run_payload.get("summary") or {}
    normalized_remarks = run_payload.get("normalized_remarks") or []

    wb = Workbook()

    # ── лист «Чек-лист» ──────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Чек-лист"
    ws.append(_CHECKLIST_HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    bold_section = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Группировка по discipline: порядок дисциплин — по первому вхождению в items (стабильно,
    # совпадает с порядком template'а/прогона), не алфавитный — так пользователю привычнее
    # (совпадает с порядком листов исходного чек-листа Glorax).
    disciplines_order: list[str] = []
    by_discipline: dict[str, list[dict[str, Any]]] = {}
    for it in items:
        disc = _discipline_of(it)
        if disc not in by_discipline:
            by_discipline[disc] = []
            disciplines_order.append(disc)
        by_discipline[disc].append(it)

    for disc in disciplines_order:
        section_row = ws.max_row + 1
        ws.append([disc])
        ws.cell(row=section_row, column=1).font = bold_section

        for it in by_discipline[disc]:
            suggested = str(it.get("suggested_answer") or "")
            suggested_label = _answer_label(suggested)
            human_answer = str(it.get("human_answer") or "unset")
            human_label = _answer_label(human_answer) if human_answer != "unset" else ""
            human_note = str(it.get("human_note") or it.get("human_comment") or "")

            row = [
                it.get("item_id", ""),
                it.get("item_no", ""),
                it.get("criterion", ""),
                suggested_label,
                it.get("confidence", ""),
                _evidence_text(it),
                it.get("model_note", ""),
                human_label,
                human_note,
            ]
            ws.append(row)
            r = ws.max_row
            ws.cell(row=r, column=3).alignment = wrap  # Критерий
            ws.cell(row=r, column=6).alignment = wrap  # Evidence

    widths = {"A": 14, "B": 8, "C": 46, "D": 18, "E": 12, "F": 34, "G": 40, "H": 18, "I": 34}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ── лист «Сводка» ────────────────────────────────────────────────────────────────────
    sw = wb.create_sheet("Сводка")

    def _append_kv_block(title: str, mapping: dict[str, Any]) -> None:
        sw.append([title])
        sw.cell(row=sw.max_row, column=1).font = Font(bold=True)
        for key, value in mapping.items():
            sw.append([str(key), value])
        sw.append([])

    sw.append(["Сводка прогона", run_payload.get("run_id", "")])
    sw.cell(row=sw.max_row, column=1).font = Font(bold=True, size=13)
    sw.append([])

    sw.append(["Всего пунктов", summary.get("total", len(items))])
    sw.append([])

    _append_kv_block("По статусу (by_status)", summary.get("by_status") or {})
    _append_kv_block("Предложенный ответ (suggested)", summary.get("suggested") or {})
    _append_kv_block("Решение инженера (human)", summary.get("human") or {})

    sw.append(["Без evidence (without_evidence)", summary.get("without_evidence", "")])
    sw.append(["С подтверждением источником (source_backed)", summary.get("source_backed", "")])
    sw.append([])

    open_manual_ids = [it.get("item_id", "") for it in items if _is_open_manual_required(it)]
    sw.append(["Незакрытые manual_required", "; ".join(open_manual_ids)])
    sw.cell(row=sw.max_row, column=1).font = Font(bold=True)
    sw.column_dimensions["A"].width = 40
    sw.column_dimensions["B"].width = 60

    # ── лист «normalized_remarks» ────────────────────────────────────────────────────────
    wr = wb.create_sheet("normalized_remarks")
    remark_header = ["id", "status", "severity", "source_refs", "message", "human_decision"]
    wr.append(remark_header)
    for cell in wr[1]:
        cell.font = Font(bold=True)
    for remark in normalized_remarks:
        source_refs = remark.get("source_refs") or []
        wr.append([
            remark.get("id", ""),
            remark.get("status", ""),
            remark.get("severity", ""),
            "; ".join(str(r) for r in source_refs),
            remark.get("message", ""),
            remark.get("human_decision", "unset"),
        ])
    for col, width in {"A": 22, "B": 20, "C": 10, "D": 40, "E": 60, "F": 16}.items():
        wr.column_dimensions[col].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(items)


# ── report_to_html ───────────────────────────────────────────────────────────────────────

_HTML_STATUS_CSS_CLASS = {
    "computed_issue": "status-issue",
    "supported_by_evidence": "status-supported",
    "manual_required": "status-manual",
    "review_needed": "status-review-needed",
    "not_applicable": "status-na",
}

_HTML_STYLE = """
<style>
  body { font-family: sans-serif; font-size: 13px; }
  table { border-collapse: collapse; width: 100%; margin-bottom: 24px; }
  th, td { border: 1px solid #ccc; padding: 4px 8px; text-align: left; vertical-align: top; }
  th { background: #f0f0f0; }
  tr.status-issue { background: #fdecea; }
  tr.status-supported { background: #eafaf1; }
  tr.status-manual { background: #fff8e1; }
  tr.status-review-needed { background: #f5f5f5; }
</style>
""".strip()


def _esc(value: Any) -> str:
    return html_module.escape(str(value if value is not None else ""))


def _summary_table_html(summary: dict[str, Any]) -> str:
    rows = [
        ("Всего пунктов", summary.get("total", "")),
        ("Без evidence", summary.get("without_evidence", "")),
        ("С подтверждением источником", summary.get("source_backed", "")),
    ]
    by_status = summary.get("by_status") or {}
    for status, count in by_status.items():
        rows.append((f"Статус: {_STATUS_RU.get(status, status)}", count))
    suggested = summary.get("suggested") or {}
    for key, count in suggested.items():
        rows.append((f"Предложенный ответ: {_answer_label(key)}", count))
    human = summary.get("human") or {}
    for key, count in human.items():
        rows.append((f"Решение инженера: {_answer_label(key) if key != 'unset' else 'не принято'}", count))

    body = "".join(f"<tr><td>{_esc(k)}</td><td>{_esc(v)}</td></tr>" for k, v in rows)
    return f"<table><tr><th>Показатель</th><th>Значение</th></tr>{body}</table>"


def _items_table_html(items: list[dict[str, Any]]) -> str:
    header = (
        "<tr><th>Раздел</th><th>№</th><th>Критерий</th><th>Предложенный ответ</th>"
        "<th>Уверенность</th><th>Evidence</th><th>Примечание системы</th>"
        "<th>Ответ инженера</th><th>Примечание инженера</th></tr>"
    )
    rows = []
    for it in items:
        status = str(it.get("status") or "")
        css_class = _HTML_STATUS_CSS_CLASS.get(status, "")
        class_attr = f' class="{css_class}"' if css_class else ""
        suggested_label = _answer_label(str(it.get("suggested_answer") or ""))
        human_answer = str(it.get("human_answer") or "unset")
        human_label = _answer_label(human_answer) if human_answer != "unset" else ""
        human_note = str(it.get("human_note") or it.get("human_comment") or "")
        evidence_html = "<br/>".join(
            _esc(r) for r in [
                str((ev or {}).get("source_ref") or "").strip()
                for ev in it.get("document_evidence") or []
            ][:_MAX_EVIDENCE_REFS] if r
        )
        rows.append(
            f"<tr{class_attr} data-item-id=\"{_esc(it.get('item_id', ''))}\">"
            f"<td>{_esc(it.get('discipline', ''))}</td>"
            f"<td>{_esc(it.get('item_no', ''))}</td>"
            f"<td>{_esc(it.get('item_id', ''))}<br/>{_esc(it.get('criterion', ''))}</td>"
            f"<td>{_esc(suggested_label)}</td>"
            f"<td>{_esc(it.get('confidence', ''))}</td>"
            f"<td>{evidence_html}</td>"
            f"<td>{_esc(it.get('model_note', ''))}</td>"
            f"<td>{_esc(human_label)}</td>"
            f"<td>{_esc(human_note)}</td>"
            f"</tr>"
        )
    return f"<table>{header}{''.join(rows)}</table>"


def report_to_html(run_payload: dict[str, Any]) -> str:
    """Рендерит ``run_payload`` (``checklist_review_v1``) в самодостаточный HTML-отчёт: таблица
    сводки + таблица всех items, строки со статусом ``computed_issue`` (и другими нетривиальными
    статусами) визуально помечены css-классом (см. ``_HTML_STATUS_CSS_CLASS``). Без общего
    вердикта «соответствует/не соответствует» (implementation_plan.md §7 правило 10) — только
    построчные статусы и model_note каждого механизма."""
    items = run_payload.get("items") or []
    summary = run_payload.get("summary") or {}
    template = _esc(run_payload.get("template", ""))
    dataset_id = _esc(run_payload.get("dataset_id", ""))

    summary_html = _summary_table_html(summary)
    items_html = _items_table_html(items)

    return (
        "<!doctype html>\n"
        "<html lang=\"ru\">\n"
        "<head>\n"
        "<meta charset=\"utf-8\">\n"
        f"<title>Чек-лист {template} — {dataset_id}</title>\n"
        f"{_HTML_STYLE}\n"
        "</head>\n"
        "<body>\n"
        f"<h2>Проверка по чек-листу «{template}» — датасет {dataset_id}</h2>\n"
        "<p><i>Предварительная автоматическая проверка: статусы предлагаемые, "
        "финальное решение по каждому пункту принимает инженер.</i></p>\n"
        "<h3>Сводка</h3>\n"
        f"{summary_html}\n"
        "<h3>Пункты чек-листа</h3>\n"
        f"{items_html}\n"
        "</body>\n"
        "</html>\n"
    )

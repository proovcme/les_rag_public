"""Официальная вёрстка КС-2 (Госкомстат № 100) в XLSX — по структуре HTML-бланка.

Числа не пересчитывает: только раскладка уже собранных строк/полей.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


def _safe_text(value: Any) -> str:
    """Prevent spreadsheet formula execution from project/user text."""
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _field_map(resolved: dict[str, Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for item in resolved.get("fields") or []:
        if isinstance(item, dict) and item.get("key") is not None:
            out[str(item["key"])] = _safe_text(item.get("value")).strip()
    return out


def _as_number(value: Any) -> float | str:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return _safe_text(value)


def _normalize_row(row: list[Any]) -> list[Any]:
    """Accept legacy 7-col rows or official 8-col rows → 8 cells."""
    cells = list(row or [])
    if len(cells) >= 8:
        return cells[:8]
    if len(cells) == 7:
        # № | name | code | unit | qty | price | total → insert smeta position = №
        return [cells[0], cells[0], cells[1], cells[2], cells[3], cells[4], cells[5], cells[6]]
    while len(cells) < 8:
        cells.append("")
    return cells


def render_ks2_xlsx(resolved: dict[str, Any], out_path: Path) -> Path:
    """Write filled/blank KS-2 workbook matching the official HTML form layout."""
    thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    font = Font(name="Times New Roman", size=9)
    font_bold = Font(name="Times New Roman", size=9, bold=True)
    font_title = Font(name="Times New Roman", size=12, bold=True)
    font_small = Font(name="Times New Roman", size=8, color="555555")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "КС-2"
    for col, width in enumerate((6, 10, 42, 16, 10, 12, 14, 14), 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    fields = _field_map(resolved)
    customer = fields.get("customer") or "—"
    contractor = fields.get("contractor") or "—"
    object_name = fields.get("object_name") or "—"
    contract = fields.get("contract") or ""
    period = fields.get("period") or ""
    act_no = fields.get("act_no") or ""
    date = fields.get("date") or ""

    document_status = str(resolved.get("document_status") or "")

    # --- шапка: стороны слева, реквизиты формы справа ---
    ws.merge_cells("A1:E1")
    ws["A1"] = "Инвестор"
    ws["A1"].font = font_bold
    ws.merge_cells("A2:E2")
    ws["A2"] = "(наименование, адрес, телефон, факс)"
    ws["A2"].font = font_small

    ws.merge_cells("F1:H1")
    ws["F1"] = "Унифицированная форма № КС-2"
    ws["F1"].font = font_bold
    ws["F1"].alignment = right
    ws.merge_cells("F2:H2")
    ws["F2"] = "Утверждена постановлением Госкомстата России от 11.11.99 № 100"
    ws["F2"].font = font_small
    ws["F2"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    ws.merge_cells("A3:E3")
    ws["A3"] = "Заказчик (Генподрядчик)"
    ws["A3"].font = font_bold
    ws.merge_cells("A4:E4")
    ws["A4"] = customer
    ws["A4"].font = font
    ws["A4"].alignment = left

    ws.merge_cells("F3:G3")
    ws["F3"] = "Форма по ОКУД"
    ws["F3"].font = font_small
    ws["H3"] = "0322005"
    ws["H3"].font = font_bold
    ws["H3"].alignment = center
    ws["H3"].border = thin

    ws.merge_cells("A5:E5")
    ws["A5"] = "Подрядчик (Субподрядчик)"
    ws["A5"].font = font_bold
    ws.merge_cells("A6:E6")
    ws["A6"] = contractor
    ws["A6"].font = font
    ws["A6"].alignment = left

    ws.merge_cells("A7:E7")
    ws["A7"] = "Стройка / Объект"
    ws["A7"].font = font_bold
    ws.merge_cells("A8:E8")
    ws["A8"] = object_name
    ws["A8"].font = font
    ws["A8"].alignment = left
    ws.row_dimensions[8].height = 30

    # --- реквизиты документа ---
    meta_row = 10
    labels = [
        ("A", "Договор подряда (контракт)"),
        ("C", "Номер документа"),
        ("E", "Дата составления"),
        ("G", "Отчётный период"),
    ]
    for col, label in labels:
        cell = ws[f"{col}{meta_row}"]
        cell.value = label
        cell.font = font_small
    ws.merge_cells(f"A{meta_row + 1}:B{meta_row + 1}")
    ws[f"A{meta_row + 1}"] = contract or "—"
    ws[f"A{meta_row + 1}"].font = font
    ws[f"A{meta_row + 1}"].border = thin
    ws.merge_cells(f"C{meta_row + 1}:D{meta_row + 1}")
    ws[f"C{meta_row + 1}"] = act_no or "—"
    ws[f"C{meta_row + 1}"].font = font
    ws[f"C{meta_row + 1}"].border = thin
    ws[f"C{meta_row + 1}"].alignment = center
    ws.merge_cells(f"E{meta_row + 1}:F{meta_row + 1}")
    ws[f"E{meta_row + 1}"] = date or "—"
    ws[f"E{meta_row + 1}"].font = font
    ws[f"E{meta_row + 1}"].border = thin
    ws[f"E{meta_row + 1}"].alignment = center
    ws.merge_cells(f"G{meta_row + 1}:H{meta_row + 1}")
    ws[f"G{meta_row + 1}"] = period or "—"
    ws[f"G{meta_row + 1}"].font = font
    ws[f"G{meta_row + 1}"].border = thin
    ws[f"G{meta_row + 1}"].alignment = center

    title_row = 13
    ws.merge_cells(f"A{title_row}:H{title_row}")
    ws[f"A{title_row}"] = (
        "ЧЕРНОВИК ИЗ ЛСР — НЕ ПОДТВЕРЖДАЕТ ФАКТИЧЕСКОЕ ВЫПОЛНЕНИЕ"
        if document_status == "draft_from_lsr_not_execution_fact"
        else "Акт о приёмке выполненных работ"
    )
    ws[f"A{title_row}"].font = (
        Font(name="Times New Roman", size=10, bold=True, color="C00000")
        if document_status == "draft_from_lsr_not_execution_fact"
        else font_title
    )
    ws[f"A{title_row}"].alignment = center

    # --- таблица: шапка как в HTML (названия + «Выполнено работ» + номера 1..8) ---
    h1 = title_row + 2
    h2 = h1 + 1
    h3 = h1 + 2
    top = [
        "Номер по порядку",
        "Номер позиции по смете",
        "Наименование работ",
        "Номер единичной расценки",
        "Единица измерения",
        "Выполнено работ",
        "",
        "",
    ]
    for col, text in enumerate(top, 1):
        cell = ws.cell(h1, col, text or None)
        cell.font = font_bold
        cell.alignment = center
        cell.border = thin
    ws.merge_cells(start_row=h1, start_column=6, end_row=h1, end_column=8)
    ws.cell(h1, 6).value = "Выполнено работ"
    for col in range(6, 9):
        ws.cell(h1, col).border = thin
        ws.cell(h1, col).alignment = center
        ws.cell(h1, col).font = font_bold
    for col in range(1, 6):
        ws.merge_cells(start_row=h1, start_column=col, end_row=h2, end_column=col)
        ws.cell(h2, col).border = thin
    for col, text in enumerate(
        ("количество", "цена за единицу, руб.", "стоимость, руб."), 6
    ):
        cell = ws.cell(h2, col, text)
        cell.font = font_bold
        cell.alignment = center
        cell.border = thin
    for col in range(1, 9):
        cell = ws.cell(h3, col, str(col))
        cell.font = font_small
        cell.alignment = center
        cell.border = thin

    ws.row_dimensions[h1].height = 28
    ws.row_dimensions[h2].height = 28

    data_start = h3 + 1
    rows = [_normalize_row(r) for r in (resolved.get("rows") or [])]
    if not rows:
        # blank бланк: несколько пустых строк
        rows = [[""] * 8 for _ in range(8)]

    money_format = '#,##0.00'
    qty_format = '0.####'

    for i, row in enumerate(rows):
        r = data_start + i
        is_total = str(row[2] or row[1] or "").strip().casefold() in {"итого", "всего"}
        for c, raw in enumerate(row, 1):
            value: Any = raw
            if c in (1, 2, 6, 7, 8) and not is_total:
                value = _as_number(raw)
            elif is_total and c == 8:
                value = _as_number(raw)
            if isinstance(value, str):
                value = _safe_text(value)
            cell = ws.cell(r, c, value if value != "" else None)
            cell.font = font_bold if is_total else font
            cell.border = thin
            if c in (1, 2, 5):
                cell.alignment = center
            elif c in (6, 7, 8):
                cell.alignment = right
                if isinstance(value, float):
                    cell.number_format = qty_format if c == 6 else money_format
            else:
                cell.alignment = left
        if is_total:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    last_data = data_start + max(len(rows), 1) - 1
    sig = last_data + 2
    ws[f"A{sig}"] = "Сдал (Подрядчик)"
    ws[f"A{sig}"].font = font_bold
    ws[f"E{sig}"] = "Принял (Заказчик)"
    ws[f"E{sig}"].font = font_bold
    ws[f"A{sig + 2}"] = contractor
    ws[f"A{sig + 2}"].font = font
    ws[f"E{sig + 2}"] = customer
    ws[f"E{sig + 2}"].font = font
    ws[f"A{sig + 4}"] = "_______________ / _______________/"
    ws[f"A{sig + 4}"].font = font_small
    ws[f"E{sig + 4}"] = "_______________ / _______________/"
    ws[f"E{sig + 4}"].font = font_small
    ws[f"A{sig + 5}"] = "М.П."
    ws[f"A{sig + 5}"].font = font_small
    ws[f"E{sig + 5}"] = "М.П."
    ws[f"E{sig + 5}"].font = font_small

    ws.print_title_rows = f"{h1}:{h3}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path

"""W20.1 — Парсер локальных сметных расчётов (ЛСР → позиции). 0 LLM ядро (ADR-11).

Импорт сметы (xlsx/csv, экспорт Гранд-Сметы): детект колонок по ключевым словам,
разбор позиций (шифр ГЭСН/ФЕР, наименование, ед., кол-во, стоимость) с разбивкой по
разделам в `les_estimate_items` + свод по разделам и итог. Единицы нормализуем через
`bor_service`. Числа — из ячеек, не LLM; итог = сумма позиций (бит-в-бит с источником).
"""
from __future__ import annotations

import csv
import sqlite3
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from backend.rag_config import rag_meta_db_path
from proxy.services.bor_service import normalize_unit

# Колонка → ключевые подстроки заголовка (нижний регистр). Порядок проверки — от
# более специфичных к общим (unit_cost до total_cost; code до name).
COLUMN_KEYS: list[tuple[str, tuple[str, ...]]] = [
    ("pos", ("№ п/п", "№пп", "n п/п", "поз.", "позиция")),
    ("code", ("обоснование", "шифр", "код норматива")),
    ("name", ("наименование",)),
    ("unit", ("ед. изм", "ед.изм", "единица изм", "ед, изм")),
    ("quantity", ("количество", "кол-во", "колич")),
    ("unit_cost", ("стоимость единицы", "цена за ед", "ст-ть ед", "стоим. ед")),
    ("total_cost", ("стоимость всего", "общая стоимость", "всего", "сумма")),
]


@dataclass
class EstimateItem:
    pos: str
    code: str
    name: str
    unit: str
    quantity: float | None
    unit_cost: float | None
    total_cost: float | None
    section: str

    def payload(self) -> dict[str, Any]:
        return {
            "pos": self.pos, "code": self.code, "name": self.name, "unit": self.unit,
            "quantity": self.quantity, "unit_cost": self.unit_cost,
            "total_cost": self.total_cost, "section": self.section,
        }


def _num(value: Any) -> float | None:
    """Разобрать число из ячейки: '1 234,56' / '1234.5' / 12 → float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _detect_columns(row: list[Any]) -> dict[str, int]:
    """Сопоставить ячейки строки-заголовка колонкам по ключевым словам."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        text = str(cell or "").strip().lower().replace("ё", "е")
        if not text:
            continue
        for key, needles in COLUMN_KEYS:
            if key in mapping:
                continue
            if any(n in text for n in needles):
                mapping[key] = idx
                break
    return mapping


def _is_section_header(row: list[Any], cols: dict[str, int]) -> str | None:
    """Строка-раздел: первая непустая ячейка начинается с «Раздел/Подраздел»
    (в Гранд-Смете заголовок раздела часто в первой/объединённой ячейке, не в колонке
    наименования)."""
    texts = [str(c or "").strip() for c in row if str(c or "").strip()]
    if not texts:
        return None
    low = texts[0].lower()
    if low.startswith("раздел") or low.startswith("подраздел"):
        return texts[0]
    return None


def parse_estimate_rows(matrix: list[list[Any]]) -> list[EstimateItem]:
    """Разобрать матрицу ячеек в позиции сметы. Чистая функция (без файла/БД)."""
    header_idx = -1
    cols: dict[str, int] = {}
    for i, row in enumerate(matrix):
        candidate = _detect_columns(row)
        if "name" in candidate and ("quantity" in candidate or "code" in candidate):
            header_idx, cols = i, candidate
            break
    if header_idx < 0:
        return []

    items: list[EstimateItem] = []
    section = ""

    def cell(row: list[Any], key: str) -> Any:
        idx = cols.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in matrix[header_idx + 1:]:
        if not any(str(c or "").strip() for c in row):
            continue
        section_name = _is_section_header(row, cols)
        if section_name is not None:
            section = section_name
            continue
        name = str(cell(row, "name") or "").strip()
        low = name.lower()
        if not name or low.startswith("итого") or low.startswith("всего по"):
            continue
        quantity = _num(cell(row, "quantity"))
        code = str(cell(row, "code") or "").strip()
        # позиция должна нести количество ИЛИ шифр норматива, иначе это служебная строка
        if quantity is None and not code:
            continue
        items.append(EstimateItem(
            pos=str(cell(row, "pos") or "").strip(),
            code=code,
            name=name,
            unit=normalize_unit(str(cell(row, "unit") or "")),
            quantity=quantity,
            unit_cost=_num(cell(row, "unit_cost")),
            total_cost=_num(cell(row, "total_cost")),
            section=section,
        ))
    return items


def _read_matrix(path: Path) -> list[list[Any]]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        ws = load_workbook(path, data_only=True).active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    if suffix in (".csv", ".txt"):
        for enc in ("utf-8-sig", "cp1251", "utf-8"):
            try:
                with open(path, encoding=enc, newline="") as f:
                    sample = f.read(4096)
                    f.seek(0)
                    delim = ";" if sample.count(";") >= sample.count(",") else ","
                    return [list(r) for r in csv.reader(f, delimiter=delim)]
            except UnicodeDecodeError:
                continue
    raise ValueError(f"Неподдерживаемый формат сметы: {suffix}")


def summarize(items: list[EstimateItem]) -> dict[str, Any]:
    """Свод по разделам и итог (сумма стоимостей позиций)."""
    sections: dict[str, dict[str, Any]] = {}
    for it in items:
        sec = sections.setdefault(it.section or "—", {"section": it.section or "—", "count": 0, "total_cost": 0.0})
        sec["count"] += 1
        sec["total_cost"] = round(sec["total_cost"] + (it.total_cost or 0.0), 2)
    total = round(sum((it.total_cost or 0.0) for it in items), 2)
    return {"sections": list(sections.values()), "total_cost": total, "items_count": len(items)}


# ── Хранилище ───────────────────────────────────────────────────────────────

def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(rag_meta_db_path())
    conn.row_factory = sqlite3.Row
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS les_estimates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL DEFAULT 0,
            name TEXT NOT NULL DEFAULT '',
            source_file TEXT NOT NULL DEFAULT '',
            total_cost REAL NOT NULL DEFAULT 0,
            items_count INTEGER NOT NULL DEFAULT 0,
            created_at REAL NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS les_estimate_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estimate_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL DEFAULT 0,
            pos TEXT NOT NULL DEFAULT '',
            code TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL DEFAULT '',
            unit TEXT NOT NULL DEFAULT '',
            quantity REAL,
            unit_cost REAL,
            total_cost REAL,
            section TEXT NOT NULL DEFAULT '',
            created_at REAL NOT NULL
        )
        """
    )
    return conn


def import_estimate(path: str | Path, project_id: int, name: str = "") -> dict[str, Any]:
    """Импортировать смету из файла → позиции в БД + свод. 0 LLM."""
    path = Path(path)
    items = parse_estimate_rows(_read_matrix(path))
    summary = summarize(items)
    now = time.time()
    with _connect() as conn:
        cur = conn.execute(
            """INSERT INTO les_estimates(project_id, name, source_file, total_cost, items_count, created_at)
               VALUES (?,?,?,?,?,?)""",
            (int(project_id), name or path.stem, path.name, summary["total_cost"], summary["items_count"], now),
        )
        estimate_id = int(cur.lastrowid)
        for it in items:
            p = it.payload()
            conn.execute(
                """INSERT INTO les_estimate_items
                   (estimate_id, project_id, pos, code, name, unit, quantity, unit_cost, total_cost, section, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (estimate_id, int(project_id), p["pos"], p["code"], p["name"], p["unit"],
                 p["quantity"], p["unit_cost"], p["total_cost"], p["section"], now),
            )
        conn.commit()
    return {"estimate_id": estimate_id, **summary}


def list_estimates(project_id: int) -> list[dict[str, Any]]:
    with _connect() as conn:
        rows = conn.execute(
            "SELECT * FROM les_estimates WHERE project_id=? ORDER BY id DESC", (int(project_id),)
        ).fetchall()
    return [dict(r) for r in rows]


def get_estimate(estimate_id: int) -> dict[str, Any]:
    """Смета с позициями и сводом по разделам."""
    with _connect() as conn:
        head = conn.execute("SELECT * FROM les_estimates WHERE id=?", (int(estimate_id),)).fetchone()
        if not head:
            return {}
        rows = conn.execute(
            "SELECT * FROM les_estimate_items WHERE estimate_id=? ORDER BY id", (int(estimate_id),)
        ).fetchall()
    items = [
        EstimateItem(pos=r["pos"], code=r["code"], name=r["name"], unit=r["unit"], quantity=r["quantity"],
                     unit_cost=r["unit_cost"], total_cost=r["total_cost"], section=r["section"])
        for r in rows
    ]
    return {"estimate": dict(head), "items": [r.payload() for r in items], **summarize(items)}


def project_total(project_id: int) -> dict[str, Any]:
    """Итог по всем сметам объекта (для КАРТЫ ОБЪЕКТА, W20.5)."""
    estimates = list_estimates(project_id)
    return {
        "estimates_count": len(estimates),
        "total_cost": round(sum(float(e.get("total_cost") or 0) for e in estimates), 2),
        "items_count": sum(int(e.get("items_count") or 0) for e in estimates),
    }

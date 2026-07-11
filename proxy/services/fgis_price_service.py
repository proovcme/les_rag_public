"""ФГИС ЦС — ценовая база из «Сплит-формы» (сметные цены + индексы по коду ресурса).

Назначение
==========
«Сплит-форма» ФГИС ЦС — это справочник сметных цен и индексов на 161k–281k строк
(регион × квартал). Доступ к ней — **точечный lookup по точному коду ресурса**
(напр. ``91.05.01-017`` краны, ``01.7.15.06-0111`` гвозди), а НЕ нечёткая агрегация.
Поэтому это отдельный сервис, а не путь row-normalizer/Qdrant: 281k строк в векторный
индекс — мусор, а нужен O(1) поиск по ключу (ADR-11: числа считает код, не LLM).

Это закрывает узкое место ``table_query top-k не SQL`` для автоценообразования ЛСР:
здесь — детерминированный exact-match по коду из in-memory индекса поверх Parquet.

Структура «Сплит-формы» (вскрыта на реальных файлах СПб 2 кв. 2025/2026)
-----------------------------------------------------------------------
Шапка таблицы — строка с ячейкой «Код ресурса, услуги» (обычно ~18-я), ниже строка
нумерации колонок «1 2 … 9», затем данные. 9 колонок:

1. Код ресурса, услуги                  → ``code``           (ключ)
2. Наименование строительного ресурса   → ``name``
3. Единица измерения                    → ``unit``
4. Отпускная цена (базовый уровень)      → ``price_release``
5. Сметная цена (базовый уровень)        → ``price_base``
6. Номер группы однородных ресурсов      → ``group_no``
7. Наименование группы                  → ``group_name``
8. Сметная цена в текущем уровне, руб.   → ``price_current``  (часто «-»)
9. Индекс изменения к группе            → ``index``

Текущая цена: ``price_current`` если число, иначе ``price_base × index`` (ресурсно-
индексный механизм). Считается на этапе конвертации и кладётся в ``price_current_eff``.
"""

from __future__ import annotations

import re
import os
import json
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable, Optional

# Нормализованная схема одной строки ценовой базы.
PRICE_FIELDS = (
    "code",
    "name",
    "unit",
    "price_release",
    "price_base",
    "group_no",
    "group_name",
    "price_current",
    "index",
    "price_current_eff",
)

# Сопоставление «подстрока заголовка → поле». Порядок важен: «текущем» проверяем
# раньше «сметная цена», иначе колонка 8 перехватит базовую (колонку 5).
_HEADER_MATCHERS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("code", ("код ресурс", "код ресурса")),
    ("name", ("наименование строительного ресурс", "наименование ресурс")),
    ("unit", ("единица измер", "ед. измер", "ед.изм")),
    ("price_release", ("отпускная цена",)),
    ("price_current", ("сметная цена в текущем", "текущем уровне")),
    ("price_base", ("сметная цена",)),  # базовая — после «текущей»
    ("group_no", ("номер группы",)),
    ("group_name", ("наименование группы",)),
    ("index", ("индекс изменения", "индекс")),
)

# Код ресурса: «01.1.01.01-0002», «91.05.01-017», возможно с буквенным префиксом базы
# («ФСБЦ-…», «ФССЦ-…», «ФСЭМ-…»). Для матча префикс снимаем.
_CODE_RE = re.compile(r"\d{2}[.\d-]{4,}")
_PREFIX_RE = re.compile(r"^[А-ЯA-Zа-яa-z]{2,6}[-_]")


def _safe_float(value: Any) -> Optional[float]:
    """'34 458,33' / '1.25' / '-' / None → float | None (как в parquet_writer)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "—", "–", "x", "х", "X", "Х"}:
        return None
    text = text.replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def normalize_code(code: Any) -> str:
    """Канонический ключ кода: trim, снять буквенный префикс базы (ФСБЦ-/ФССЦ-…)."""
    text = re.sub(r"\s+", "", str(code or "")).strip()
    if not text:
        return ""
    stripped = _PREFIX_RE.sub("", text)
    return stripped or text


def _search_text(value: Any) -> str:
    """Canonical lexical form for model-visible price browsing, never a selector."""
    text = str(value or "").casefold().replace("ё", "е")
    return " ".join(re.findall(r"[0-9a-zа-я]+", text))


def _search_tokens(value: Any) -> list[str]:
    return [token for token in _search_text(value).split() if len(token) >= 2]


def _is_transport_row(rec: dict[str, Any]) -> bool:
    code = normalize_code(rec.get("code"))
    name = _search_text(rec.get("name"))
    return bool(re.fullmatch(r"\d{6}-\d{2}-\d{4}-\d{4}", code)) or name.startswith("перевозка ")


def _is_resource_code(code: str) -> bool:
    return bool(re.fullmatch(r"\d{2}(?:\.\d+)+-\d+", normalize_code(code)))


def _looks_like_code(value: Any) -> bool:
    return bool(_CODE_RE.search(str(value or "")))


# ─────────────────────────────────────────
# ПАРСИНГ «СПЛИТ-ФОРМЫ»
# ─────────────────────────────────────────

def parse_split_form(xlsx_path: str | Path, *, max_header_scan: int = 60) -> dict[str, Any]:
    """Читает «Сплит-форму» → {'meta': {...}, 'rows': [ {PRICE_FIELDS...}, ... ]}.

    0 LLM: шапка ищется по ячейке «Код ресурса», колонки — по подстрокам заголовка.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    meta: dict[str, Any] = {}
    header_idx = 0
    col_map: dict[int, str] = {}

    # 1) Найти строку-шапку (содержит «код ресурс») и собрать meta из строк выше.
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_header_scan, values_only=True), 1):
        cells = [("" if c is None else str(c)).strip() for c in row]
        low0 = cells[0].casefold() if cells else ""
        # Мета берём только из «label | значение»-строк (метка в кол.0, короткое значение
        # в кол.1) — иначе абзацы-примечания со словом «субъект» попадают как регион.
        val1 = cells[1] if len(cells) > 1 else ""
        if "наименование субъект" in low0 and 0 < len(val1) < 80:
            meta.setdefault("subject", val1)
        if "наименование зоны" in low0 and 0 < len(val1) < 80:
            meta.setdefault("zone", val1)
        if any("код ресурс" in c.casefold() for c in cells):
            header_idx = r_idx
            for col_pos, cell in enumerate(cells):
                low = cell.casefold()
                for field, needles in _HEADER_MATCHERS:
                    if field in col_map.values():
                        continue
                    if any(n in low for n in needles):
                        col_map[col_pos] = field
                        break
            break

    if not header_idx or "code" not in col_map.values():
        wb.close()
        raise ValueError(f"Шапка «Сплит-формы» не найдена в {Path(xlsx_path).name}")

    # 2) Данные: пропускаем строку нумерации колонок, читаем строки с валидным кодом.
    code_col = next(pos for pos, f in col_map.items() if f == "code")
    rows: list[dict[str, Any]] = []
    num_fields = {"price_release", "price_base", "price_current", "index"}

    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        if code_col >= len(row):
            continue
        code_cell = row[code_col]
        if not _looks_like_code(code_cell):
            continue
        rec: dict[str, Any] = {f: None for f in PRICE_FIELDS}
        for pos, field in col_map.items():
            val = row[pos] if pos < len(row) else None
            if field in num_fields:
                rec[field] = _safe_float(val)
            elif field == "group_no":
                rec[field] = ("" if val is None else str(val).strip())
            else:
                rec[field] = ("" if val is None else str(val).strip())
        # Эффективная текущая цена: прямая (кол.8) или база×индекс (ресурсно-индексный).
        eff = rec["price_current"]
        if eff is None and rec["price_base"] is not None and rec["index"] is not None:
            eff = round(rec["price_base"] * rec["index"], 2)
        rec["price_current_eff"] = eff
        rows.append(rec)

    wb.close()
    return {"meta": meta, "rows": rows}


def build_price_parquet(
    xlsx_path: str | Path,
    out_path: str | Path,
    *,
    region: str | None = None,
    quarter: str | None = None,
) -> dict[str, Any]:
    """«Сплит-форма» xlsx → Parquet (нормализованный, с region/quarter). Возвращает сводку."""
    import pandas as pd

    parsed = parse_split_form(xlsx_path)
    rows = parsed["rows"]
    region = region or parsed["meta"].get("subject") or ""
    quarter = quarter or ""

    df = pd.DataFrame(rows, columns=list(PRICE_FIELDS))
    df["region"] = region
    df["quarter"] = quarter

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out_path, compression="snappy", index=False)

    return {
        "parquet": str(out_path),
        "rows": len(rows),
        "region": region,
        "quarter": quarter,
        "meta": parsed["meta"],
    }


# ─────────────────────────────────────────
# КНИГА ЦЕН (lookup)
# ─────────────────────────────────────────

class PriceBook:
    """Индекс ценовой базы для exact-match по коду + поиск по наименованию.

    Грузится один раз из Parquet, держит ``{нормализованный_код: строка}``.
    """

    def __init__(self, rows: list[dict[str, Any]], *, region: str = "", quarter: str = ""):
        self.region = region
        self.quarter = quarter
        self._rows = rows
        self._by_code: dict[str, dict[str, Any]] = {}
        for rec in rows:
            key = normalize_code(rec.get("code"))
            if key and key not in self._by_code:
                self._by_code[key] = rec

    @classmethod
    def from_parquet(cls, parquet_path: str | Path) -> "PriceBook":
        import pandas as pd

        df = pd.read_parquet(parquet_path)
        region = str(df["region"].iloc[0]) if "region" in df and len(df) else ""
        quarter = str(df["quarter"].iloc[0]) if "quarter" in df and len(df) else ""
        df = df.astype(object).where(pd.notnull(df), None)
        return cls(df.to_dict(orient="records"), region=region, quarter=quarter)

    def __len__(self) -> int:
        return len(self._by_code)

    def lookup(self, code: str) -> Optional[dict[str, Any]]:
        """Точная цена по коду (с учётом снятия префикса базы). None если нет."""
        return self._by_code.get(normalize_code(code))

    def price(self, code: str, *, method: str = "index") -> Optional[float]:
        """Цена по коду: method='index' → текущая (база×индекс/прямая), 'base' → базовая."""
        rec = self.lookup(code)
        if rec is None:
            return None
        return rec.get("price_current_eff") if method == "index" else rec.get("price_base")

    def search(self, query: str, *, limit: int = 20) -> list[dict[str, Any]]:
        """Поиск по подстроке наименования/кода — когда точный код неизвестен."""
        q = (query or "").strip().casefold()
        if not q:
            return []
        out: list[dict[str, Any]] = []
        for rec in self._rows:
            hay = f"{rec.get('code','')} {rec.get('name','')}".casefold()
            if q in hay:
                out.append(rec)
                if len(out) >= limit:
                    break
        return out

    def browse(self, query: str, *, limit: int = 20) -> list[dict[str, Any]]:
        """Ranked lexical candidates for a model/user to inspect.

        This is deliberately not fuzzy auto-binding. The method returns multiple
        FGIS rows with an explicit lexical trace; only a model or user may bind a
        project material to one exact resource code afterwards.
        """
        phrase = _search_text(query)
        tokens = _search_tokens(query)
        exact_code = normalize_code(query)
        if not phrase and not exact_code:
            return []
        asks_transport = any(token in {"перевозка", "доставка", "транспорт", "транспортировка"} for token in tokens)
        ranked: list[tuple[tuple[int, int, int, int, int], str, dict[str, Any]]] = []
        for rec in self._rows:
            code = normalize_code(rec.get("code"))
            name = _search_text(rec.get("name"))
            hay = f"{code.casefold()} {name}"
            hay_tokens = set(_search_tokens(hay))
            matched = [token for token in tokens if token in hay_tokens]
            code_match = int(bool(exact_code and code == exact_code))
            phrase_match = int(bool(phrase and phrase in hay))
            all_tokens = int(bool(tokens) and len(matched) == len(tokens))
            if not (code_match or phrase_match or matched):
                continue
            transport_row = _is_transport_row(rec)
            # A material browser must not rank haulage tariffs as material prices.
            # They stay available for an explicit transport query or exact code.
            if transport_row and not (asks_transport or code_match):
                continue
            resource_code = int(_is_resource_code(code))
            # Sort keys are evidence only: exact code, resource-code family, full
            # phrase, complete token coverage, token count. No applicability lives here.
            payload = dict(rec)
            payload["match"] = {
                "schema": "fgis_price_lexical_match_v1",
                "query": str(query or ""),
                "exact_code": bool(code_match),
                "phrase_match": bool(phrase_match),
                "all_tokens": bool(all_tokens),
                "matched_tokens": matched,
                "query_tokens": tokens,
                "resource_code": bool(resource_code),
                "transport_row": bool(transport_row),
            }
            key = (code_match, resource_code, phrase_match, all_tokens, len(matched))
            ranked.append((key, name, payload))
        ranked.sort(
            key=lambda item: (
                -item[0][0], -item[0][1], -item[0][2], -item[0][3], -item[0][4],
                item[1], normalize_code(item[2].get("code")),
            )
        )
        return [item[2] for item in ranked[: max(1, int(limit))]]

    def lookup_many(self, codes: Iterable[str]) -> dict[str, Optional[dict[str, Any]]]:
        return {c: self.lookup(c) for c in codes}


DEFAULT_PRICE_ROOT = Path("data/price_base")
PRICEBOOK_MANIFEST_PATH = Path("config/domain/pricebook_manifest.json")
_SCRATCH_PRICEBOOK_STEMS = {"spb_refresh"}
_DEFAULT_PRICEBOOK_STEMS = (
    "sankt-peterburg_2kv2026",
    "spb_2kv2026",
    "spb_2kv2025",
    "sankt-peterburg_2kv2025",
)


@lru_cache(maxsize=8)
def get_pricebook(parquet_path: str) -> PriceBook:
    """Кешированная загрузка книги цен по пути к Parquet."""
    return PriceBook.from_parquet(parquet_path)


def _is_scratch_pricebook(path: str | Path) -> bool:
    return Path(path).stem in _SCRATCH_PRICEBOOK_STEMS or Path(path).stem.endswith("_refresh")


@lru_cache(maxsize=1)
def pricebook_manifest() -> dict[str, Any]:
    """Operator-maintained visibility/alias manifest for local pricebooks."""
    try:
        if PRICEBOOK_MANIFEST_PATH.exists():
            data = json.loads(PRICEBOOK_MANIFEST_PATH.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
    except Exception:
        return {}
    return {}


def _manifest_aliases() -> dict[str, str]:
    raw = pricebook_manifest().get("aliases")
    if not isinstance(raw, dict):
        return {}
    out: dict[str, str] = {}
    for key, val in raw.items():
        src = str(key or "").strip()
        dst = str(val or "").strip()
        if src and dst and src != dst:
            out[src] = dst
    return out


def _manifest_hidden_stems() -> set[str]:
    raw = pricebook_manifest().get("hidden_stems")
    hidden = {str(item or "").strip() for item in raw if str(item or "").strip()} if isinstance(raw, list) else set()
    quarantined = pricebook_manifest().get("quarantined_stems")
    if isinstance(quarantined, dict):
        hidden.update(str(item or "").strip() for item in quarantined if str(item or "").strip())
    return hidden


def _canonical_stem(stem: str, stems: set[str]) -> str:
    aliases = _manifest_aliases()
    seen: set[str] = set()
    current = str(stem or "").strip()
    while current and current in aliases and current not in seen:
        seen.add(current)
        candidate = aliases[current]
        if candidate not in stems:
            break
        current = candidate
    return current or stem


def _is_hidden_pricebook(path: str | Path, stems: set[str]) -> bool:
    stem = Path(path).stem
    if _is_scratch_pricebook(path):
        return False
    manifest = pricebook_manifest()
    quarantined = manifest.get("quarantined_stems")
    if isinstance(quarantined, dict) and stem in quarantined:
        # A quarantined regional identity is unsafe even when there is no
        # canonical alias in this particular directory.
        return True
    if stem not in _manifest_hidden_stems():
        return False
    canonical = _canonical_stem(stem, stems)
    # Do not hide a legacy alias in isolated test/dev roots unless its canonical
    # replacement is physically present in the same root.
    return canonical != stem and canonical in stems


def available_pricebooks(
    root: str | Path = DEFAULT_PRICE_ROOT,
    *,
    include_scratch: bool = False,
    include_hidden: bool = False,
) -> list[str]:
    root = Path(root)
    if not root.exists():
        return []
    paths = sorted(root.glob("*.parquet"))
    stems = {p.stem for p in paths}
    if not include_scratch:
        paths = [p for p in paths if not _is_scratch_pricebook(p)]
    if not include_hidden:
        paths = [p for p in paths if not _is_hidden_pricebook(p, stems)]
    return [str(p) for p in paths]


def resolve_pricebook_path(
    book: str | None = None,
    *,
    root: str | Path = DEFAULT_PRICE_ROOT,
    allow_scratch: bool = False,
) -> Optional[str]:
    """Resolve a pricebook stem/path to a local Parquet path.

    With no explicit book use the deterministic system default. This keeps
    chat, LSR traces and price lookup on the same region/period instead of
    accidentally using the first alphabetic Parquet file.
    """
    try:
        books = available_pricebooks(
            root,
            include_scratch=allow_scratch or bool(book),
            include_hidden=bool(book),
        )
    except TypeError:
        # Backward-compatible with old tests/plugins that monkeypatch the
        # function before the include_scratch keyword existed.
        books = available_pricebooks(root)
    if not books:
        return None
    stems = {Path(path).stem: path for path in books}
    if book:
        stem = _canonical_stem(Path(book).stem, set(stems))
        return stems.get(stem)

    preferred: list[str] = []
    configured = os.getenv("LES_DEFAULT_PRICEBOOK", "").strip()
    if configured:
        preferred.append(Path(configured).stem)
    manifest_default = str(pricebook_manifest().get("default_book") or "").strip()
    if manifest_default:
        preferred.append(Path(manifest_default).stem)
    preferred.extend(_DEFAULT_PRICEBOOK_STEMS)
    preferred.extend(stem for stem in sorted(stems) if "2026" in stem)
    preferred.extend(sorted(stems))

    seen: set[str] = set()
    for stem in preferred:
        if stem in seen:
            continue
        seen.add(stem)
        path = stems.get(stem)
        if path:
            return path
    return books[0]

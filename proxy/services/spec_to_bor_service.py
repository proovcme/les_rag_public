"""Спецификация (форма 9, ГОСТ 21.110) → ВОР (объёмы монтажных работ) — W11.10.

Детерминированное преобразование: каждая позиция спецификации → строка работы, где
объём работы = количество из спецификации, а глагол работы выбирается по категории
предмета (словарь). Ноль LLM (ADR-11). Алгоритм — `docs/ALGO-spec-to-bor.md`.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from proxy.services.bor_service import (
    BorLine,
    _normalize_name,
    bor_to_xlsx,
    collect_spec_rows,
    normalize_unit,
)
from proxy.services.estimate_math_service import parse_ru_number

logger = logging.getLogger(__name__)

# qty-приоритет + data-aware fallback (как в reconcile/table).
_QTY_FIELDS = ("qty", "work_volume", "work_done", "work_since_start")

# Заголовки секций / нечисловой мусор — не позиции (как в сверке).
_SECTION_RE = re.compile(r"^\d+\s*[.)]\s")
# Orphan electrical rating cell split from device name by pdfplumber.
_RATING_ONLY_RE = re.compile(
    r"^\d+\s*А\b(?:\s*,?\s*~?\s*\d+\s*В)?(?:\s*,?\s*IP\s*\d+)?\s*$",
    re.IGNORECASE,
)
# Form-9 size/section branches split by pdfplumber (cable cross-section, Ø pipe).
_SIZE_BRANCH_RE = re.compile(
    r"^\s*[-–—]\s+.+"
    r"|^\s*[Ø⌀]\s*\d"
    r"|^\s*\d+\s*[xх×]\s*\d",
    re.IGNORECASE,
)
_SOFT_CONTINUATION_RE = re.compile(
    r"^(сечением|не\s+распространя|не\s+выделяющ|с\s+изоляц|с\s+оболоч|"
    r"полимерн|медн|алюмин|в\s+труб|для\s+)",
    re.IGNORECASE,
)
_CABLE_NAME_TOKENS = ("кабель", "провод", "шнур")
# Noun forms only — not adjectives (кабельный) and not compounds (водогазопроводная).
_CABLE_NOUN_RE = re.compile(
    r"(?<![а-яёa-z0-9])(?:"
    r"кабел(?:ь|я|ю|ем|е|и|ей|ям|ями|ях)|"
    r"провод(?:а|у|ом|е|ы|ов|ам|ами|ах)?|"
    r"шнур(?:а|у|ом|е|ы|ов|ам|ами|ах)?"
    r")(?![а-яёa-z0-9])",
    re.IGNORECASE,
)
_KM_UNITS = frozenset({"км", "км.", "km", "km."})

# Категория предмета → глагол работы. Порядок ВАЖЕН: конкретные предметы выше «Прокладки»,
# иначе прилагательное «кабельный» (лоток/наконечник кабельный) ложно цепляет «Прокладку».
# (глагол, кортеж ключевых слов наименования)
_WORK_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("Установка", ("коробка", "клемм", "наконечник", "крепеж", "крепёж", "закладн",
                   "дюбель", "хомут", "скоба", "зажим", "розетк", "выключател")),
    ("Монтаж", ("лоток", "короб", "труба", "гофр", "канал", "стойк", "полка", "подвес",
                "светильник", "прожектор", "щит", "шкаф", "бокс", "датчик", "извещател",
                "прибор", "блок", "автомат", "трансформатор", "привод", "двигател",
                "насос", "вентилятор", "агрегат")),
    ("Прокладка", ("кабель", "провод", "шнур")),
)
_DEFAULT_VERB = "Монтаж"


def _normalize_spec_name(name: str) -> str:
    return f" {(name or '').lower().replace('ё', 'е')} "


def _has_cable_noun(name: str) -> bool:
    """True only for cable/wire/cord as a noun, not «кабельный» / «водогазопроводная»."""
    return bool(_CABLE_NOUN_RE.search(_normalize_spec_name(name)))


def _token_in_name(name: str, token: str) -> bool:
    """Category token match. Cable nouns use inflection-aware boundaries (ADR-safe)."""
    tok = (token or "").lower().replace("ё", "е")
    if not tok:
        return False
    if tok in _CABLE_NAME_TOKENS:
        return _has_cable_noun(name)
    low = _normalize_spec_name(name)
    return tok in low


def _is_noise_name(name: str) -> bool:
    s = (name or "").strip()
    if len(s) < 3 or _SECTION_RE.match(s):
        return True
    if _RATING_ONLY_RE.match(s):
        return True
    return sum(ch.isalpha() for ch in s) < 2


def _row_qty(row: dict) -> float | None:
    for key in _QTY_FIELDS:
        val = row.get(key)
        if val is None:
            continue
        parsed = parse_ru_number(val)
        if parsed is not None:
            return parsed
    return None


def _join_spec_name(*parts: str) -> str:
    chunks = [re.sub(r"\s+", " ", str(part or "").strip()) for part in parts]
    return re.sub(r"\s+", " ", " ".join(chunk for chunk in chunks if chunk)).strip()


def _is_size_branch_name(name: str) -> bool:
    return bool(_SIZE_BRANCH_RE.match(str(name or "").strip()))


def _is_soft_continuation_name(name: str, *, pending_name: str) -> bool:
    text = str(name or "").strip()
    if not text or not pending_name:
        return False
    if text.casefold() in {"сечением:", "сечением"}:
        return True
    if _SOFT_CONTINUATION_RE.match(text):
        return True
    # pdfplumber often continues a wrapped cell on the next row in lowercase.
    first = text[0]
    if first.islower() or first in ",;)]}":
        return True
    pending = pending_name.rstrip()
    if pending.endswith((",", "-", "—", "–", ":")):
        return True
    return False


def _looks_like_cable_name(name: str) -> bool:
    return _has_cable_noun(name)


def _normalize_cable_length_unit(row: dict) -> None:
    """Deterministic км→м for cable/wire rows. Never invents qty (ADR-11)."""
    if not _looks_like_cable_name(str(row.get("name") or "")):
        return
    unit = normalize_unit(row.get("unit")).casefold().replace("ё", "е")
    if unit not in _KM_UNITS:
        return
    qty = _row_qty(row)
    if qty is not None:
        row["qty"] = float(qty) * 1000.0
        note = "ед. пересчитаны км→м (×1000) из спецификации"
    else:
        note = "в спецификации ед. км; кол-во отсутствует — в ВОР ед. м без выдуманного объёма"
    row["unit"] = "м"
    prev = str(row.get("note") or "").strip()
    row["note"] = f"{prev}; {note}".strip("; ") if prev else note


def _coalesce_form9_rows(rows: list[dict]) -> list[dict]:
    """Join Form-9 wrapped description / size-branch rows without inventing numbers.

    pdfplumber often splits one cable position into:
    ``Кабель силовой…`` + ``полимерных…`` + ``сечением:`` + ``- 3х1,5 мм2``.
    Size branches under one description become separate positions that inherit
    the shared prefix; quantities stay only those present on the branch/prefix.
    """
    out: list[dict] = []
    pending: dict | None = None
    pending_emitted_sizes = False

    def flush_pending() -> None:
        nonlocal pending, pending_emitted_sizes
        if pending is None:
            return
        if not pending_emitted_sizes:
            _normalize_cable_length_unit(pending)
            out.append(pending)
        pending = None
        pending_emitted_sizes = False

    for raw in rows:
        if not isinstance(raw, dict):
            continue
        row = dict(raw)
        name = str(row.get("name") or "").strip()
        if not name:
            continue
        if _is_size_branch_name(name):
            if pending is None and out:
                # Orphan size line after a flushed incomplete description.
                base = str(out[-1].get("name") or "")
                if _looks_like_cable_name(base) or "сечением" in base.casefold():
                    pending = dict(out.pop())
                    pending_emitted_sizes = False
                    # Strip a previously absorbed size from the shared prefix.
                    pending["name"] = re.split(
                        r"\s[-–—]\s+\d", pending["name"], maxsplit=1
                    )[0].strip()
            if pending is None:
                _normalize_cable_length_unit(row)
                out.append(row)
                continue
            merged = dict(pending)
            merged["name"] = _join_spec_name(str(pending.get("name") or ""), name)
            if _row_qty(row) is not None:
                merged["qty"] = row.get("qty")
            elif pending_emitted_sizes:
                # Parent qty already consumed by an earlier size branch — do not double-count.
                merged["qty"] = None
            if not str(merged.get("unit") or "").strip() and str(row.get("unit") or "").strip():
                merged["unit"] = row.get("unit")
            if str(row.get("mark") or "").strip() and not str(merged.get("mark") or "").strip():
                merged["mark"] = row.get("mark")
            if str(row.get("code") or "").strip() and not str(merged.get("code") or "").strip():
                merged["code"] = row.get("code")
            if str(row.get("pos") or "").strip() and not str(merged.get("pos") or "").strip():
                merged["pos"] = row.get("pos")
            _normalize_cable_length_unit(merged)
            out.append(merged)
            pending_emitted_sizes = True
            continue

        if pending is not None and _is_soft_continuation_name(
            name, pending_name=str(pending.get("name") or "")
        ):
            if pending_emitted_sizes:
                # After size branches, do not treat a new Title-case row as a
                # continuation just because the shared prefix ends with ":".
                soft_only = (
                    name.casefold() in {"сечением:", "сечением"}
                    or bool(_SOFT_CONTINUATION_RE.match(name))
                    or name[0].islower()
                    or name[0] in ",;)]}"
                )
                if not soft_only:
                    flush_pending()
                    pending = row
                    if _row_qty(pending) is not None and not _looks_like_cable_name(
                        str(pending.get("name") or "")
                    ):
                        _normalize_cable_length_unit(pending)
                        out.append(pending)
                        pending = None
                        pending_emitted_sizes = False
                    continue
            pending["name"] = _join_spec_name(str(pending.get("name") or ""), name)
            if _row_qty(pending) is None and _row_qty(row) is not None:
                pending["qty"] = row.get("qty")
            if not str(pending.get("unit") or "").strip() and str(row.get("unit") or "").strip():
                pending["unit"] = row.get("unit")
            continue

        flush_pending()
        pending = row
        # Complete equipment/material rows keep flowing immediately.
        if _row_qty(pending) is not None and not _looks_like_cable_name(
            str(pending.get("name") or "")
        ):
            _normalize_cable_length_unit(pending)
            out.append(pending)
            pending = None
            pending_emitted_sizes = False

    flush_pending()
    # Re-number synthetic pos only when the source did not provide one.
    for index, row in enumerate(out, start=1):
        if not str(row.get("pos") or "").strip():
            row["pos"] = str(index)
    return out


def work_verb(name: str) -> str:
    """Глагол работы по категории предмета (словарь). Без LLM."""
    for verb, tokens in _WORK_RULES:
        if any(_token_in_name(name, tok) for tok in tokens):
            return verb
    return _DEFAULT_VERB


def spec_rows_to_work_lines(rows: list[dict]) -> list[BorLine]:
    """Свод работ из позиций спецификации: группировка по (раздел, работа, ед.), сумма qty."""
    lines: dict[tuple, BorLine] = {}
    for row in rows:
        raw_name = str(row.get("name") or row.get("work_name") or "").strip()
        if not raw_name or _is_noise_name(raw_name):
            continue
        name = re.sub(r"\s+", " ", raw_name)
        verb = work_verb(name)
        work_name = f"{verb}: {name}"
        section = str(row.get("section") or "").strip()
        code = str(row.get("code") or "").strip()
        mark = str(row.get("mark") or "").strip()
        unit = normalize_unit(row.get("unit"))
        key = (section.casefold(), _normalize_name(work_name), unit)

        line = lines.get(key)
        if line is None:
            line = BorLine(section=section, name=work_name, code=code, mark=mark, unit=unit, qty=None)
            lines[key] = line

        qty = _row_qty(row)
        if qty is None:
            line.qty_missing_rows += 1
        else:
            line.qty = (line.qty or 0.0) + qty
        line.source_rows += 1
        source = str(row.get("source_file") or "").strip()
        pos = str(row.get("pos") or row.get("position") or "").strip()
        ref = f"{source}#{pos}" if pos else source
        if ref and ref not in line.sources:
            line.sources.append(ref)

    return sorted(lines.values(), key=lambda l: (l.section.casefold(), l.name.casefold()))


# ── v2: декомпозиция позиции в НАБОР работ (методика ВОР, ГОСТ 21.111) ──
# Категория → перечень работ. Все под-работы наследуют ЕД.+КОЛ-ВО позиции (объём один и тот
# же — линейный/поштучный), поэтому чисел НЕ выдумываем (ADR-11). Работы по числу концов
# (маркировка, расключение) в авто-объём не попадают — отмечаем в примечании.
# Порядок важен: «коробка» раньше «короб», иначе substring «короб» съедает коробки.
# Кабель/провод/шнур — только существительные (_has_cable_noun), не «кабельный» /
# «водогазопроводная».
_DECOMPOSE: tuple[tuple[tuple[str, ...], tuple[str, ...], str], ...] = (
    # (ключевые слова категории, перечень работ, примечание о доп. работах)
    (("кабель", "провод", "шнур"),
     ("Разметка трассы", "Прокладка кабеля"),
     "доп.: маркировка и расключение — по числу концов, добавить отдельно"),
    (("коробка", "клемм", "наконечник", "крепеж", "крепёж", "дюбель", "хомут", "скоба", "зажим"),
     ("Установка {предмет}",),
     ""),
    (("лоток", "короб", "труба", "гофр", "канал"),
     ("Разметка трассы", "Монтаж {предмет}"),
     ""),
    (("стойк", "полка", "подвес", "конструкц", "закладн"),
     ("Монтаж {предмет}",),
     ""),
    (("светильник", "прожектор", "щит", "шкаф", "бокс", "розетк", "выключател", "датчик",
      "извещател", "прибор", "блок", "автомат", "трансформатор", "привод", "двигател",
      "насос", "вентилятор", "агрегат", "оповещател", "модул"),
     ("Установка {предмет}", "Подключение"),
     ""),
)
_DEFAULT_DECOMPOSE = (("Монтаж {предмет}",), "")


def _decompose(name: str) -> tuple[tuple[str, ...], str]:
    for tokens, works, note in _DECOMPOSE:
        if any(_token_in_name(name, t) for t in tokens):
            return works, note
    return _DEFAULT_DECOMPOSE


@dataclass
class WorkLine:
    """Строка ВОР (форма ГОСТ 21.111): работа + объём + ссылка на чертёж + примечание."""
    section: str
    work: str
    unit: str
    qty: float | None = None
    chertezh: str = ""           # ссылка на чертёж (шифр/марка позиции)
    note: str = ""               # примечание/формула/доп. работы
    source_rows: int = 0
    qty_missing_rows: int = 0
    sources: list[str] = field(default_factory=list)

    def payload(self) -> dict:
        return {"section": self.section, "name": self.work, "unit": self.unit,
                "qty": (round(self.qty, 3) if self.qty is not None else None),
                "chertezh": self.chertezh, "note": self.note,
                "source_rows": self.source_rows, "sources": self.sources}


def spec_rows_to_work_lines_v2(rows: list[dict]) -> list[WorkLine]:
    """Декомпозиция: позиция → набор работ; свод по (раздел, работа, ед.), сумма qty."""
    lines: dict[tuple, WorkLine] = {}
    for row in rows:
        raw_name = str(row.get("name") or row.get("work_name") or "").strip()
        if not raw_name or _is_noise_name(raw_name):
            continue
        name = re.sub(r"\s+", " ", raw_name)
        unit = normalize_unit(row.get("unit"))
        qty = _row_qty(row)
        section = str(row.get("section") or "").strip()
        chertezh = str(row.get("mark") or row.get("code") or "").strip()
        works, dnote = _decompose(name)
        pos = str(row.get("pos") or row.get("position") or "").strip()
        src = str(row.get("source_file") or "").strip()
        ref = f"{src}#{pos}" if pos else src
        row_note = str(row.get("note") or "").strip()
        for tmpl in works:
            work = tmpl.replace("{предмет}", name)
            key = (section.casefold(), _normalize_name(work), unit)
            line = lines.get(key)
            if line is None:
                note = f"объём = кол-ву по спецификации (поз. {pos})" if pos else "объём = кол-ву по спецификации"
                if dnote:
                    note += "; " + dnote
                if row_note:
                    note += "; " + row_note
                line = WorkLine(section=section, work=work, unit=unit, chertezh=chertezh, note=note)
                lines[key] = line
            elif row_note and row_note not in line.note:
                line.note = f"{line.note}; {row_note}" if line.note else row_note
            if qty is None:
                line.qty_missing_rows += 1
            else:
                line.qty = (line.qty or 0.0) + qty
            line.source_rows += 1
            if ref and ref not in line.sources:
                line.sources.append(ref)
    return sorted(lines.values(), key=lambda l: (l.section.casefold(), l.work.casefold()))


def work_lines_to_xlsx(lines: list[WorkLine], path: Path, *, title: str) -> int:
    """xlsx ВОР по графам ГОСТ 21.111: №/Наименование работ/Ед./Кол-во/Чертёж/Примечание."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ВОР"
    ws.append([title])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=12)
    hdr = ["№", "Наименование работ", "Ед. изм.", "Кол-во", "Ссылка на чертёж", "Примечание"]
    ws.append(hdr)
    fill = PatternFill("solid", fgColor="1F4E78")
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cur_section = None
    n = 0
    for line in lines:
        if line.section and line.section != cur_section:
            cur_section = line.section
            ws.append([f"Раздел: {cur_section}"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True)
        n += 1
        ws.append([n, line.work, line.unit,
                   (round(line.qty, 3) if line.qty is not None else "—"),
                   line.chertezh, line.note])
    widths = [5, 52, 9, 12, 18, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return n


def _wants_vor(question: str) -> bool:
    """True if the user asks for a VOR / bill of quantities (word-boundary safe)."""
    q = (question or "").lower().replace("ё", "е")
    # «вор» — по границе слова: иначе «пОВОРоты», «творог», «забор» ложно триггерят.
    return bool(
        re.search(r"\bвор\b", q)
        or ("ведомост" in q and "работ" in q)
        or "объем работ" in q
        or "объемов работ" in q
    )


def _asks_for_lsr_or_estimate(question: str) -> bool:
    """ЛСР/смета — document path; не путать с «ВОР из спецификации»."""
    q = (question or "").lower().replace("ё", "е")
    return bool(
        re.search(r"\bлср\b", q)
        or re.search(r"\bсмет[ауыеой]*\b", q)
        or re.search(r"\bстоимост[ьяиюе]*\b", q)
    )


def is_spec_to_bor_query(question: str, *, has_attachment: bool = False) -> bool:
    """Намерение «сделай ВОР из спецификации» / ВОР по tabular-вложению.

    Без вложения: нужна явная спецификация/форма + ВОР.
    С tabular read-вложением: достаточно ВОР/ведомости работ (файл уже источник).
    «Собери ЛСР по приложенной ВОР» — не этот канал: ВОР уже вход, нужен document LSR.
    """
    q = (question or "").lower().replace("ё", "е")
    if not _wants_vor(q):
        return False
    if _asks_for_lsr_or_estimate(q):
        return False
    if has_attachment:
        return True
    return "спецификац" in q or "форм" in q


_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("наименование", "название", "name", "материал", "наимен"),
    "unit": ("ед. изм.", "ед изм", "ед.изм", "единица", "unit", "ед."),
    "qty": (
        "кол-во факт",
        "количество факт",
        "кол-во",
        "количество",
        "qty",
        "колич",
        "кол.",
    ),
    "section": ("раздел", "section", "объект", "зона"),
    "code": ("артикул", "код", "code", "шифр"),
    "mark": ("марка", "тип", "mark", "обозначение"),
    "pos": ("поз.", "поз", "позиция", "№ п/п", "no."),
}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _map_headers(cells: list[object]) -> dict[str, int]:
    """Map logical field → 0-based column index from a header row."""
    found: dict[str, int] = {}
    normalized = [_normalize_header(c) for c in cells]
    for idx, header in enumerate(normalized):
        if not header:
            continue
        # Mass / weight columns are never quantity for VOR volumes.
        if "масс" in header or header in {"вес", "weight"}:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in found:
                continue
            if any(header == alias or header.startswith(alias) for alias in aliases):
                found[field] = idx
                break
    return found


def _spec_rows_from_matrix(
    rows_iter: Iterable[object],
    *,
    source: str,
    pos_start: int = 0,
) -> tuple[list[dict], int]:
    """Parse name/qty/unit rows from a 2D table matrix. 0 LLM."""
    header_map: dict[str, int] | None = None
    out: list[dict] = []
    pos = pos_start
    for raw in rows_iter:
        cells = list(raw or ())
        if not any(c is not None and str(c).strip() for c in cells):
            continue
        if header_map is None:
            mapped = _map_headers(cells)
            if "name" in mapped:
                header_map = mapped
            continue
        assert header_map is not None
        name_idx = header_map["name"]
        if name_idx >= len(cells):
            continue
        name = str(cells[name_idx] or "").replace("\n", " ").strip()
        name = re.sub(r"\s+", " ", name)
        if not name or _is_noise_name(name):
            continue

        def _cell(field: str, _cells=cells, _header=header_map) -> object:
            idx = _header.get(field)
            if idx is None or idx >= len(_cells):
                return None
            return _cells[idx]

        qty = parse_ru_number(_cell("qty"))
        code = str(_cell("code") or "").strip()
        mark = str(_cell("mark") or "").strip()
        source_pos = str(_cell("pos") or "").strip()
        pos += 1
        out.append(
            {
                "doc_type": "SPEC",
                "name": name,
                "unit": _cell("unit"),
                "qty": qty,
                "section": str(_cell("section") or "").strip(),
                "code": code,
                "mark": mark or code,
                # Only Form-9 «Поз.»; blank cells stay empty so coalesce can keep
                # the parent position and renumber only truly missing ones.
                "pos": source_pos,
                "source_file": source,
                "note": "",
            }
        )
    coalesced = _coalesce_form9_rows(out)
    return coalesced, pos_start + len(coalesced)


def rows_from_spec_xlsx(path: Path | str, *, source_label: str = "") -> list[dict]:
    """Read a materials/spec XLSX (e.g. Каменка) into spec_to_bor row dicts. 0 LLM."""
    import openpyxl

    file_path = Path(path)
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"unsupported attachment type for spec→ВОР: {file_path.suffix}")
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows, _ = _spec_rows_from_matrix(
            ws.iter_rows(values_only=True),
            source=source_label or file_path.name,
        )
        return rows
    finally:
        wb.close()


def rows_from_spec_pdf(path: Path | str, *, source_label: str = "") -> list[dict]:
    """Read materials/spec tables from a PDF via pdfplumber. 0 LLM."""
    try:
        import pdfplumber
    except ImportError as exc:  # pragma: no cover - environment capability
        raise RuntimeError("pdfplumber is required for PDF spec→ВОР") from exc

    file_path = Path(path)
    if file_path.suffix.lower() != ".pdf":
        raise ValueError(f"unsupported attachment type for PDF spec→ВОР: {file_path.suffix}")
    source = source_label or file_path.name
    out: list[dict] = []
    pos = 0
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if not table:
                    continue
                rows, pos = _spec_rows_from_matrix(table, source=source, pos_start=pos)
                out.extend(rows)
    # Tables/pages often split one Form-9 position; matrix coalesce is local —
    # one more pass joins size-branches that landed in the next table.
    return _coalesce_form9_rows(out)


def rows_from_spec_document(path: Path | str, *, source_label: str = "") -> list[dict]:
    """Dispatch PDF/XLSX/XLSM into the same spec→VOR row contract."""
    from proxy.smeta_core.source_intake import TABLE_DOCUMENT_SUFFIXES

    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix not in TABLE_DOCUMENT_SUFFIXES:
        raise ValueError(f"unsupported attachment type for spec→ВОР: {suffix or 'none'}")
    if suffix == ".pdf":
        return rows_from_spec_pdf(file_path, source_label=source_label)
    return rows_from_spec_xlsx(file_path, source_label=source_label)


def generate_spec_bor_from_rows(
    rows: list[dict],
    *,
    output_dir: Path | None = None,
    title: str = "ВОР из спецификации",
    decompose: bool = True,
    source_id: str = "attachment",
) -> dict:
    """Spec rows → work VOR (+ optional xlsx). Numbers only from rows. 0 LLM."""
    if decompose:
        wlines = spec_rows_to_work_lines_v2(rows)
        result: dict = {
            "dataset_id": source_id,
            "mode": "decompose",
            "source_rows": len(rows),
            "bor_lines": len(wlines),
            "lines": [w.payload() for w in wlines],
            "xlsx_path": None,
        }
        if output_dir is not None and wlines:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", source_id).strip("._") or "attachment"
            xlsx_path = Path(output_dir) / f"specbor_{safe}_{stamp}.xlsx"
            work_lines_to_xlsx(wlines, xlsx_path, title=title)
            result["xlsx_path"] = str(xlsx_path)
        return result

    lines = spec_rows_to_work_lines(rows)
    result = {
        "dataset_id": source_id,
        "mode": "simple",
        "source_rows": len(rows),
        "bor_lines": len(lines),
        "lines": [line.payload() for line in lines],
        "xlsx_path": None,
    }
    if output_dir is not None and lines:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", source_id).strip("._") or "attachment"
        xlsx_path = Path(output_dir) / f"specbor_{safe}_{stamp}.xlsx"
        bor_to_xlsx(lines, xlsx_path, title=title)
        result["xlsx_path"] = str(xlsx_path)
    return result


def format_spec_bor_answer(result: dict, dataset_label: str = "") -> str:
    lines = result.get("lines", [])
    head = (f"ВОР из спецификации: {result['bor_lines']} работ "
            f"из {result['source_rows']} позиций"
            + (f" · {dataset_label}" if dataset_label else "") + ".")
    # Group preview by section when present.
    sample = []
    seen_sections: list[str] = []
    for l in lines[:20]:
        section = str(l.get("section") or "").strip()
        if section and section not in seen_sections:
            seen_sections.append(section)
            sample.append(f"[{section}]")
        qty = l.get("qty")
        qty_s = f"{round(qty, 2)} {l.get('unit', '')}".strip() if qty is not None else "— (нет кол-ва)"
        sample.append(f"  • {l['name']} — {qty_s}")
    if result.get("xlsx_path"):
        tail = "\nПолная таблица — в Excel-вложении. Количества только из исходника, без цен и без LLM."
    else:
        tail = ("\nПолная таблица: Инструменты → ВОР (режим «работы из спецификации») "
                "или POST /api/bor/{id}/from-spec/generate. Числа — из исходника, 0 LLM.")
    return head + ("\n" + "\n".join(sample) if sample else "") + tail


def generate_spec_bor(
    dataset_id: str,
    *,
    storage_root: Path = Path("storage/datasets"),
    output_dir: Path | None = None,
    decompose: bool = True,
) -> dict:
    """Спецификация датасета (Parquet) → ВОР работ → xlsx. Без LLM.

    decompose=True (v2, методика ГОСТ 21.111): позиция → НАБОР работ + графы чертёж/примечание,
    группировка по разделам. decompose=False (v1): 1 позиция → 1 монтажная работа.
    """
    rows = collect_spec_rows(dataset_id, storage_root=storage_root)
    if decompose:
        wlines = spec_rows_to_work_lines_v2(rows)
        result: dict = {
            "dataset_id": dataset_id, "mode": "decompose",
            "source_rows": len(rows), "bor_lines": len(wlines),
            "lines": [w.payload() for w in wlines], "xlsx_path": None,
        }
        if output_dir is not None and wlines:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            xlsx_path = output_dir / f"specbor_{dataset_id}_{stamp}.xlsx"
            work_lines_to_xlsx(wlines, xlsx_path, title=f"ВОР из спецификации (ГОСТ 21.111) — {dataset_id}")
            result["xlsx_path"] = str(xlsx_path)
        return result

    lines = spec_rows_to_work_lines(rows)
    result = {
        "dataset_id": dataset_id, "mode": "simple",
        "source_rows": len(rows), "bor_lines": len(lines),
        "lines": [line.payload() for line in lines], "xlsx_path": None,
    }
    if output_dir is not None and lines:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = output_dir / f"specbor_{dataset_id}_{stamp}.xlsx"
        bor_to_xlsx(lines, xlsx_path, title=f"ВОР из спецификации (Ф9) — {dataset_id}")
        result["xlsx_path"] = str(xlsx_path)
    return result

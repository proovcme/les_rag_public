"""Checklist-review engine (Glorax, Phase 2, T2.1+T2.2) — каркас + presence + calculation.

АРХИТЕКТУРА (implementation_plan.md §2-§4, docs/ALGO-glorax-checklist-review.md): generic-профиль
``checklist-review`` поверх template'ов ``checklist_template_importer`` — не отдельный rule-engine.
Инвариант (AGENTS.md, ADR-11): чек-лист задаёт вопрос -> RAG ищет evidence -> код проверяет
формализуемое -> LLM связывает и объясняет -> инженер принимает решение. Уверенный ``yes/no`` без
``document_evidence[].source_ref`` — архитектурный регресс (правило 1, safety-тест; calculation
computed_issue/no — явное разрешённое исключение, evidence там computed, не document).

Паттерн инъекции — как в ``doc_review_service.run_review``: сама ``run_checklist_review`` ЧИСТАЯ
функция, никаких обращений к живым Qdrant/MLX/MetaDB внутри неё. Provider'ы (``inventory_provider``,
``search_provider``, ``workbook_provider``) приходят параметрами — тесты подставляют фейки,
продакшн — обёртки ``default_inventory_provider``/``default_search_provider``/
``default_workbook_provider`` (lazy-импорт MetaDB/lexical_index_service, не вызываются в тестах).

T2.1 реализовал presence + честные заглушки остальных kind:
  - presence: контентный хит (>=2 якорных терма, либо >=1 если якорь один) -> supported_by_evidence/yes;
    filename-match в inventory без контентного хита -> review_needed/unknown (правило 1b — filename-match
    без контента НЕ равно yes); ничего не найдено -> review_needed/unknown (НЕ no, правило 3);
  - spatial_visual -> manual_required (честная граница v1, ALGO §3.4).

T2.2 добавляет calculation-механизм (openpyxl formula-count, детерминированно, 0 LLM) поверх
нового ``workbook_provider`` + подключает реальные ``default_inventory_provider``/
``default_search_provider``/``default_workbook_provider`` к MetaDB sqlite/lexical_index_service
(см. ``_run_calculation`` ниже). ``cross_section``/``spds_formal`` остаются честными заглушками
(review_needed/not_run) — вне скоупа T2.2/T3.1, см. ``_STUB_KINDS``.
  - normalized_remarks/defense — поля-заглушки (пустой список / минимальный defense_contract_v1-каркас),
    содержательное наполнение — T2.4 (см. TODO).

T3.1 добавляет parametric-механизм (``proxy/services/checklist_param_rules.py`` +
``config/checklists/glorax_param_rules.yaml``, детерминированно, 0 LLM): для item kind=parametric
ищем правило по item_id в реестре -> search_provider даёт хиты -> extract_value по snippet'у
каждого хита -> compare кодом. Значение найдено в одном хите и ok -> supported_by_evidence/yes;
issue -> computed_issue/no (safety-исключение — evidence computed/document с source_ref есть);
конфликтующие значения в РАЗНЫХ хитах -> computed_issue + оба evidence + model_note о конфликте;
ничего не найдено -> review_needed (НЕ no); item kind=parametric без правила в реестре ->
review_needed + model_note «нет параметрического правила» (см. ``_run_parametric`` ниже).

T3.2 (A) закрывает gap T3.1: диспетчер ``_run_item`` теперь запускает параметрическую проверку
НЕЗАВИСИМО от kind, если для item_id есть правило в реестре glorax_param_rules.yaml — это касается
7 из 9 правил, которые висят на kind=manual_required («составные» пункты, где числовой порог —
лишь часть более широкого экспертного критерия). Для kind=parametric поведение НЕ меняется. Для
составных пунктов: ok -> status=supported_by_evidence, но suggested_answer ОСТАЁТСЯ manual_required
(параметрическая часть доказана кодом, остальная часть критерия — нет); issue -> status=
computed_issue, suggested_answer ОСТАЁТСЯ manual_required (НЕ no — код не имеет права закрыть весь
составной критерий); правило не нашло значение -> честный fallback на прежний путь kind (см. блок
в начале ``_run_item`` и ``_finalize_item`` ниже).

T3.2 (B) реализует cross_section (two-sided gate, implementation_plan.md §3.2): project-хит
(dataset_id) И source-хит (любой из source_dataset_ids) оба обязательны для содержательного
результата; suggested_answer для cross_section в этой фазе НИКОГДА не бывает yes (нет механизма
LLM-связывания содержимого — Phase 5) — максимум review_needed/unknown с обоими evidence
(kind="project_doc"/"source_doc"). source_dataset_ids пуст/не задан -> review_needed на item'е +
blocker на уровне workflow_plan (см. ``_run_cross_section`` ниже). ``_STUB_KINDS`` уменьшен до
``{"spds_formal"}`` (после T2.4 — пуст).

T2.4 (A) закрывает последний kind-заглушку: spds_formal переиспользует doc_review вместо
отдельного движка (курируемый словарь ``_SPDS_ANCHOR_TO_RULE_IDS`` — якорный терм критерия ->
doc_review rule_ids из ``config/normcontrol/gost_r_21_101_2026.yaml`` + NK-алиасы
``normcontrol_service``). Новый инъектируемый ``doc_review_provider(dataset_id) -> dict | None``
(продакшн — ``default_doc_review_provider``, lazy-обёртка над ``doc_review_service.review_dataset``,
не вызывается в тестах): все замапленные rule_id найдены и supported -> supported_by_evidence/yes
с evidence (kind=computed, source_ref = исходный source_ref doc_review item либо rule_id) и
заполненным ``doc_review_item_ids``; есть computed_issue с непустым source_ref -> computed_issue/no
(safety-исключение); computed_issue БЕЗ source_ref -> manual_required (нарушение найдено, но
suggested_answer=no без evidence запрещён); provider=None, результат None, словарь пуст или
покрытие rule_id неполное -> review_needed + model_note (см. ``_run_spds_formal`` ниже).
``_STUB_KINDS`` теперь пуст.

T2.4 (B) наполняет ``normalized_remarks``/``defense`` содержательно (были пустые заглушки с T2.1):
``_build_normalized_remarks`` — normalized_remark_v1 по образцу
``doc_review_service.review_to_normalized_remarks``, но ``source="checklist"``,
``category="checklist"``, ``checklist_ref={template, item_id}``, id формата ``REM-CHK-{item_id}``;
формируется для item'ов со статусом computed_issue/supported_by_evidence/manual_required
(review_needed НЕ порождает remark — отсутствие evidence не равно нарушению). ``_build_defense`` —
DefensePack (``evidence_contract.DefenseClaim``/``DefensePack``, переиспользованы, не
скопированы руками) с ``schema=defense_contract_v1``, ``domain="normcontrol.checklist_review"``,
по claim на каждый item с remark, ``summary`` содержит ``by_status`` + ``human_final_required=true``.

T2.5 добавляет composition-checker состава ПД по ПП РФ №87 (``proxy/services/
pp87_composition_service.py`` + ``config/checklists/pp87_composition.yaml``, детерминированно,
0 LLM): опциональный параметр ``pp87_config`` включает сверку разделов ПД (по имени файла
inventory) — результат кладётся в НОВОЕ top-level поле контракта ``pp87_composition`` (``None``
без ``pp87_config``). Сознательно НЕ привязано к items дисциплины «Общее» — ни один из 5 её
критериев в ``glorax_pd_2026.json`` не сформулирован про состав/разделы ПД (см. ``_run_pp87_composition``
ниже и SESSION_LOG.md Запись 17).

Никаких LLM-вызовов в этом модуле нет вообще (T2.1/T2.2/T3.1/T3.2/T2.4/T2.5 scope).
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Callable, Iterable

# ── статусы item (переиспользуем enum doc_review_service, не заводим новый) ───────────────
S_COMPUTED_ISSUE = "computed_issue"
S_SUPPORTED = "supported_by_evidence"
S_NOT_APPLICABLE = "not_applicable"
S_MANUAL = "manual_required"
S_REVIEW_NEEDED = "review_needed"

# suggested_answer — фиксированный словарь контракта (implementation_plan.md §4)
A_YES = "yes"
A_NO = "no"
A_NOT_REQUIRED = "not_required"
A_MANUAL_REQUIRED = "manual_required"
A_UNKNOWN = "unknown"

_SUGGESTED_KEYS = (A_YES, A_NO, A_NOT_REQUIRED, A_MANUAL_REQUIRED, A_UNKNOWN)

# kind, для которых механизм ещё не реализован — честная заглушка review_needed/not_run.
# T2.4: spds_formal реализован через переиспользование doc_review (см. _run_spds_formal) —
# _STUB_KINDS теперь пуст, оставлен как явный маркер для будущих новых kind (расширяемость).
_STUB_KINDS: set[str] = set()

_CHECKLISTS_DIR = Path(__file__).resolve().parent.parent.parent / "config" / "checklists"

# Лимит сканирования xlsx/xlsm-кандидата на формулы (T2.2): реальные Glorax-расчёты на порядок
# меньше — лимит только защита от аномально больших книг, не архитектурное ограничение.
_CALC_MAX_ROWS = 2000
_CALC_MAX_COLS = 50
_CALC_XLSX_EXTS = (".xlsx", ".xlsm")
_CALC_LEGACY_EXTS = (".xls",)


# ── _anchor_terms: детерминированный вывод поисковых якорей из текста критерия ────────────

_STOPWORDS = {
    "и", "в", "во", "не", "что", "он", "на", "я", "с", "со", "как", "а", "то", "все", "она",
    "так", "его", "но", "да", "ты", "к", "у", "же", "вы", "за", "бы", "по", "только", "ее",
    "мне", "было", "вот", "от", "меня", "еще", "нет", "о", "из", "ему", "теперь", "когда",
    "даже", "ну", "вдруг", "ли", "если", "уже", "или", "ни", "быть", "был", "него", "до",
    "вас", "нибудь", "опять", "уж", "вам", "об", "тем", "для", "их", "при", "об", "ко",
}

_PUNCT_RE = re.compile(r"[«»,.;:()\"'!?]")
_WORD_RE = re.compile(r"[а-яёa-z0-9-]+", re.IGNORECASE)


def _anchor_terms(criterion: str) -> list[str]:
    """Детерминированные поисковые якоря из текста критерия: нижний регистр, без пунктуации,
    без стоп-слов/предлогов — только содержательные термы. Порядок сохраняется по вхождению
    в исходный текст (устойчивый для evidence-снипетов и для тестов).

    Пример: «Приложен отчет об инженерно-геологических изысканиях» ->
    ['отчет', 'инженерно-геологических', 'изысканиях']. Предлог «об» отсекается стоп-словами;
    глагол «приложен»/«приложены»/«приложена» отсекается отдельно (см. явный список ниже) —
    он маркирует ФОРМУ критерия (наличие чего-то), а не тему искомого документа, и одинаково
    встречается почти во всех presence-критериях, поэтому не несёт различающей поисковой
    ценности как якорь.
    """
    text = _PUNCT_RE.sub(" ", criterion.lower().replace("ё", "е"))
    words = _WORD_RE.findall(text)
    terms: list[str] = []
    for w in words:
        if w in _STOPWORDS:
            continue
        if w in ("приложен", "приложены", "приложена"):
            # глагол-маркер наличия, не содержательный якорь темы документа
            continue
        if len(w) < 3:
            continue
        if w not in terms:
            terms.append(w)
    return terms


# ── load_checklist_template ────────────────────────────────────────────────────────────────


def load_checklist_template(name: str, base: str | Path | None = None) -> dict[str, Any]:
    """Читает нормализованный template JSON из ``config/checklists/{name}.json``.

    ``base`` — переопределение директории (тесты/альтернативная раскладка); по умолчанию —
    ``config/checklists`` относительно корня репозитория.
    """
    root = Path(base) if base is not None else _CHECKLISTS_DIR
    path = root / f"{name}.json"
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


# ── presence-механизм ───────────────────────────────────────────────────────────────────────


def _content_hit(terms: list[str], hits: Iterable[dict[str, Any]]) -> dict[str, Any] | None:
    """Возвращает первый hit, чей snippet содержит >=2 якорных терма (или >=1, если терм один).
    Нет достаточного покрытия термами -> не считается контентным хитом (анти-галлюцинация:
    случайное упоминание одного слова из многих не подтверждает критерий)."""
    if not terms:
        return None
    need = 1 if len(terms) == 1 else 2
    for hit in hits:
        snippet = str((hit or {}).get("snippet") or "").lower().replace("ё", "е")
        matched = sum(1 for t in terms if t in snippet)
        if matched >= need:
            return hit
    return None


def _filename_hit(terms: list[str], files: Iterable[dict[str, Any]]) -> dict[str, Any] | None:
    """Filename-match: имя файла в inventory содержит >=1 якорный терм. Слабый сигнал —
    только для model_note «имя совпало, содержимое не подтверждено» (правило 1b), НЕ дает yes."""
    for f in files:
        name = str((f or {}).get("file_name") or "").lower().replace("ё", "е")
        if any(t in name for t in terms):
            return f
    return None


def _run_presence(criterion: str, dataset_id: str, *,
                   inventory_provider: Callable[[str], list[dict]],
                   search_provider: Callable[[str, list[str]], list[dict]]) -> tuple[str, str, list[dict], str, float]:
    """presence: (status, suggested_answer, document_evidence, model_note, confidence).

    Правила (implementation_plan.md §4, правило 1b; ALGO-doc):
      1. контентный хит в search_provider (snippet содержит >=2 якорных термина, либо >=1 если
         термин один) -> supported_by_evidence/yes, evidence с непустым source_ref+snippet;
      2. нет контентного хита, но filename-match в inventory -> review_needed/unknown,
         model_note явно фиксирует «имя совпало, содержимое не подтверждено»;
      3. ничего не найдено -> review_needed/unknown (отсутствие evidence != нарушение, не no).
    """
    terms = _anchor_terms(criterion)
    hits = search_provider(dataset_id, terms) or []
    hit = _content_hit(terms, hits)
    if hit is not None:
        evidence = [{
            "kind": "document",
            "source_ref": str(hit.get("source_ref") or ""),
            "snippet": str(hit.get("snippet") or ""),
            "file_name": str(hit.get("file_name") or ""),
            "value": "", "unit": "", "bbox": None,
            "reason": "контентный хит подтверждает наличие документа/раздела",
        }]
        return S_SUPPORTED, A_YES, evidence, "Найден контентный хит, подтверждающий наличие документа.", 0.75

    files = inventory_provider(dataset_id) or []
    fmatch = _filename_hit(terms, files)
    if fmatch is not None:
        note = ("Имя файла совпадает с ожидаемым документом, но имя совпало, содержимое не "
                "подтверждено — нужен контентный хит для yes (правило 1b).")
        return S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.3

    return (S_REVIEW_NEEDED, A_UNKNOWN, [],
            "Документ не найден ни по содержимому, ни по имени файла — недостаточно evidence "
            "для вывода (отсутствие evidence не равно нарушению).", 0.1)


# ── calculation-механизм (T2.2) ────────────────────────────────────────────────────────────

# Реальные файлы датасета на диске почти всегда именуются транслитом (латиницей) — «Теплотехнический
# расчет» на диске лежит как «teplotehnicheskiy_raschet.xlsx», а якорные термы _anchor_terms всегда
# кириллица. Прямое подстрочное сравнение (как для presence/_filename_hit) поэтому здесь не работает
# — нужна транслитерация якоря + сопоставление по ПРЕФИКСУ (не по полному слову): транслитерация
# неоднозначна на конце слова (падежные окончания, й/ы транслитерируются по-разному в разных схемах:
# "теплотехническИЙ" может лечь на диск и как "-ii", и как "-iy"), а начало содержательного слова
# устойчиво почти всегда. Таблица — та же ГОСТ-подобная, что DISCIPLINE_TRANSLIT/_TRANSLIT_MAP в
# checklist_template_importer.py (независимая копия — модули не должны знать друг о друге).
_RU_TRANSLIT_MAP = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
    "ж": "zh", "з": "z", "и": "i", "й": "i", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}

_TRANSLIT_PREFIX_LEN = 5  # достаточно, чтобы не давать случайных совпадений на коротких словах


def _translit(word: str) -> str:
    return "".join(_RU_TRANSLIT_MAP.get(ch, ch) for ch in word.lower())


def _name_matches_term(low_name: str, term: str) -> bool:
    """Термин совпадает с именем файла либо дословно (кириллица в имени — редко, но бывает),
    либо по транслит-префиксу (латиница — обычный случай на диске)."""
    if term in low_name:
        return True
    translit = _translit(term)
    prefix = translit[:_TRANSLIT_PREFIX_LEN] if len(translit) > _TRANSLIT_PREFIX_LEN else translit
    return bool(prefix) and prefix in low_name


# T2.6 (баг 1, SESSION_LOG Запись 18-19): курируемый стоп-лист ОБЩИХ домен-термов calculation.
# Root cause бага: workbook_provider отдаёт file_name как ОТНОСИТЕЛЬНЫЙ ПУТЬ (не basename) — путь
# реального датасета содержит директории вида «РАЗДЕЛ 3», «Экспликация помещений» — слова из этих
# имён папок ("раздел", "помещений") случайно совпадают с общими термами критерия ("расчёт
# категории ПОМЕЩЕНИЙ", "расчёт звукоизоляции нормируемых ПОМЕЩЕНИЙ", "приведён РАСЧЁТ...РАЗДЕЛ
# 8"), давая ложный yes от файла, чья настоящая тема (та самая «Итоговая таблица» экспликации
# помещений) не имеет отношения к конкретному расчёту. Термы из этого списка ФОРМАЛЬНО проходят
# _anchor_terms (не предлоги/не стоп-слова в общеязыковом смысле), но не несут различающей
# ценности ДЛЯ ПОИСКА КОНКРЕТНОГО EXCEL-ФАЙЛА: они либо маркируют ФОРМУ критерия («выполнен»,
# «приведён», «представлена», «проведены», «указаны», «подтверждены» — обычные глаголы presence/
# calculation-формулировок), либо описывают технический формат/носитель («формат», «excel»,
# «ячейках», «сохранением»), либо — структуру проекта, а не тему расчёта («раздел», «помещений»,
# «категории», «таблица», «форме», «наличии»). Список НЕ включает специфичные предметные термы
# («звукоизоляции», «электрических», «нагрузок», «влагопереноса», «теплотехнической»,
# «однородности», «воздухообмена», «освещённости», «тво») — они остаются единственными якорями
# для отбора xlsx-кандидата по имени.
_CALC_DOMAIN_STOPWORDS = {
    "расчет", "расчеты", "расчетом", "расчетов", "расчетную", "расчетных",
    "выполнен", "выполнены", "выполнена", "приведен", "приведена", "проведены", "проведен",
    "указаны", "подтверждены", "подтверждена",
    "таблица", "таблицы", "формате", "формат", "форме", "excel",
    "представлена", "представлены", "приложение",
    "помещений", "помещения", "помещение", "категории", "категория",
    "раздел", "разделы", "сводная", "наличии", "ячейках", "сохранением",
}


def _calc_specific_terms(criterion: str) -> list[str]:
    """Специфичные термы критерия ДЛЯ ПОИСКА EXCEL-КАНДИДАТА ПО ИМЕНИ: ``_anchor_terms`` минус
    курируемый стоп-лист общих домен-термов ``_CALC_DOMAIN_STOPWORDS`` (T2.6, баг 1). Порядок
    сохраняется. Пустой результат (все термы — общие) означает «нет специфичной темы для поиска
    по имени файла» — вызывающий код (``_run_calculation``) не должен искать xlsx-кандидата по
    имени вовсе в этом случае (см. ``_workbook_candidates``), только через контентный хит."""
    return [t for t in _anchor_terms(criterion) if t not in _CALC_DOMAIN_STOPWORDS]


def _workbook_candidates(terms: list[str], files: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """Кандидаты xlsx/xlsm/xls по СПЕЦИФИЧНЫМ термам критерия (``_calc_specific_terms``, не сырым
    ``_anchor_terms`` — T2.6, баг 1) в имени файла: дословное совпадение (кириллица) ИЛИ транслит-
    префикс (латиница — обычные имена файлов на диске, см. ``_name_matches_term``). Пустой
    ``terms`` (все термы критерия — общие домен-слова) -> НЕТ кандидатов по имени вовсе (честно,
    не пытаемся угадать по нулю специфичных термов). Порядок сохраняется по вхождению в ``files``
    (стабильно для тестов и evidence)."""
    if not terms:
        return []
    out: list[dict[str, Any]] = []
    for f in files:
        name = str((f or {}).get("file_name") or "")
        ext = Path(name).suffix.lower()
        if ext not in _CALC_XLSX_EXTS and ext not in _CALC_LEGACY_EXTS:
            continue
        low = name.lower().replace("ё", "е")
        if not any(_name_matches_term(low, t) for t in terms):
            continue
        out.append(f)
    return out


# Лимит чтения контентной проверки книги (T2.6, баг 1): только имена листов + первая строка
# каждого листа — расчётные книги Glorax называют тему в заголовке/шапке таблицы (см. реальный
# позитивный пример промпта «Расчет электрических нагрузок.xlsx» с листом «Нагрузки»), полное
# сканирование всего контента избыточно и медленно для гейта релевантности.
_CALC_CONTENT_HEADER_ROWS = 1


def _workbook_content_terms_match(path: str, terms: list[str]) -> bool:
    """Контентная проверка книги (T2.6, баг 1, промпт T2.6): имена листов ЛИБО текст первой
    строки каждого листа должны содержать хотя бы один специфичный терм критерия — иначе
    кандидат, найденный только по имени файла, отклоняется (анти-галлюцинация: имя файла может
    случайно совпасть по общему слову структуры проекта, реальная тема книги проверяется по
    содержимому). Сравнение — дословное вхождение ИЛИ по 5-символьному префиксу термина (см.
    ``_TRANSLIT_PREFIX_LEN`` — та же логика устойчивости к падежным окончаниям, что и в
    ``_name_matches_term``, но без транслитерации: имена листов/заголовки почти всегда кириллица).
    Пустой ``terms`` -> False (нет специфичной темы — контент не может её подтвердить).
    Ошибка чтения книги (битый файл и т.п.) -> False (честно "не подтверждено контентом", не
    падение — вызывающий код уже открывает книгу отдельно для подсчёта формул и обработает свою
    ошибку там)."""
    if not terms:
        return False
    from openpyxl import load_workbook

    def _term_in_text(text: str, term: str) -> bool:
        low = text.lower().replace("ё", "е")
        if term in low:
            return True
        prefix = term[:_TRANSLIT_PREFIX_LEN] if len(term) > _TRANSLIT_PREFIX_LEN else term
        return bool(prefix) and prefix in low

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — честно "не подтверждено", ошибку разберёт _count_formula_cells
        return False
    try:
        for ws in wb.worksheets:
            sheet_title = str(ws.title or "")
            if any(_term_in_text(sheet_title, t) for t in terms):
                return True
            max_col = min(ws.max_column or 0, _CALC_MAX_COLS)
            if max_col <= 0:
                continue
            for row in ws.iter_rows(min_row=1, max_row=_CALC_CONTENT_HEADER_ROWS, min_col=1, max_col=max_col):
                row_text = " ".join(str(cell.value) for cell in row if cell.value is not None)
                if row_text and any(_term_in_text(row_text, t) for t in terms):
                    return True
        return False
    finally:
        wb.close()


def _count_formula_cells(path: str) -> tuple[int, list[str]]:
    """Детерминированный подсчёт формульных ячеек по всем листам книги (openpyxl, read_only,
    data_only=False — читаем формулы, не кэшированные значения). Возвращает (count, sheet_names
    с найденными формулами, в порядке обхода). Лимит _CALC_MAX_ROWS x _CALC_MAX_COLS на лист —
    защита от аномально больших книг, не архитектурное ограничение (implementation_plan.md §3.1)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        total = 0
        sheets_with_formulas: list[str] = []
        for ws in wb.worksheets:
            sheet_count = 0
            max_row = min(ws.max_row or 0, _CALC_MAX_ROWS)
            max_col = min(ws.max_column or 0, _CALC_MAX_COLS)
            if max_row <= 0 or max_col <= 0:
                continue
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        sheet_count += 1
                    else:
                        # openpyxl ArrayFormula/formula objects (не голая строка) — тоже формула
                        type_name = type(value).__name__
                        if type_name in ("ArrayFormula", "Formula"):
                            sheet_count += 1
            if sheet_count:
                total += sheet_count
                sheets_with_formulas.append(str(ws.title))
        return total, sheets_with_formulas
    finally:
        wb.close()


def _run_calculation(criterion: str, dataset_id: str, *,
                      inventory_provider: Callable[[str], list[dict]],
                      search_provider: Callable[[str, list[str]], list[dict]],
                      workbook_provider: Callable[[str], list[dict]]
                      ) -> tuple[str, str, list[dict], str, float, dict[str, Any]]:
    """calculation: (status, suggested_answer, document_evidence, model_note, confidence,
    computed_check) — implementation_plan.md §3.1 "Наличие расчёта в Excel", промпт T2.2 (a)-(g),
    T2.6 (баг 1, ужесточение отбора кандидата).

    Порядок веток (первое совпадение решает):
      (g) кандидат .xls (legacy) на первом месте среди xlsx-кандидатов -> review_needed,
          model_note содержит "legacy_unsupported", БЕЗ попытки парсинга;
      (c)/(d) xlsx/xlsm-кандидат найден (по СПЕЦИФИЧНЫМ термам в имени файла + подтверждён
          контентной проверкой книги, ИЛИ по контентному хиту lexical-поиска, который уже сам по
          себе контентно подтверждён и называет xlsx/xlsm файл в workbook_provider) -> считаем
          формулы: формулы есть -> supported_by_evidence/yes, evidence kind=computed; формул нет
          -> computed_issue/no, evidence kind=computed (safety-исключение);
      (e) xlsx-кандидата нет, но есть контентный хит расчёта в ДРУГОМ (не xlsx/xlsm) файле ->
          review_needed + model_note "расчёт найден не в Excel";
      (f) совсем ничего -> review_needed.

    T2.6 (баг 1, SESSION_LOG Запись 18-19): отбор кандидата по имени файла ужесточён ДВУМЯ
    независимыми гейтами (устраняют жадный матчинг, который на бою свёл 9 разных критериев к
    одному нерелевантному файлу «Итоговая таблица_28.09.2022.xlsx»):
      1. ``_calc_specific_terms`` — курируемый стоп-лист общих домен-термов (``расчёт``,
         ``таблица``, ``выполнен``, ``приложен*``, ``представлена``, ``формат``, ``excel``,
         ``помещений``, ``раздел`` и т.п., см. ``_CALC_DOMAIN_STOPWORDS``) исключается из
         якорных термов ПЕРЕД поиском по имени — остаются только специфичные термы темы расчёта;
         если специфичных термов не осталось вовсе, кандидатов по имени НЕ ищем (``_workbook_
         candidates`` возвращает []), решение сразу падает в ветку (e)/(f) (content_hit/ничего);
      2. ``_workbook_content_terms_match`` — контентная проверка книги: имена листов ЛИБО текст
         первой строки каждого листа должны содержать хотя бы один специфичный терм — иначе
         кандидат, совпавший только по имени, ОТКЛОНЯЕТСЯ (не открывается на подсчёт формул
         вовсе, не даёт ни yes, ни no — переходит к следующему кандидату/веткам e/f).
    Кандидат, добавленный через ``content_hit`` (lexical-хит УЖЕ содержит >=2 специфичных терма в
    самом тексте документа, см. ``_content_hit``), НЕ проходит повторно книжную проверку — его
    релевантность уже подтверждена содержимым на этапе lexical-поиска.
    """
    specific_terms = _calc_specific_terms(criterion)
    files = workbook_provider(dataset_id) or []
    name_candidates = _workbook_candidates(specific_terms, files)

    # Контентная проверка книги (T2.6, баг 1): кандидат, найденный ТОЛЬКО по имени, обязан
    # подтвердиться содержимым (имена листов/первая строка) — иначе отклоняется молча (не
    # открывается на подсчёт формул). Кандидаты, не прошедшие проверку, просто не попадают в
    # итоговый список — они не дают ни yes, ни no, ни review_needed от своего имени.
    # .xls (legacy) НЕ проверяется контентно вовсе — openpyxl физически не читает .xls (ветка (g)
    # ниже честно этого и не пытается), контентный гейт применим только к xlsx/xlsm.
    verified_name_candidates = [
        cand for cand in name_candidates
        if Path(str(cand.get("file_name") or "")).suffix.lower() in _CALC_LEGACY_EXTS
        or _workbook_content_terms_match(str(cand.get("path") or ""), specific_terms)
    ]

    # Полный anchor-набор термов (включая общие домен-слова) — для lexical content_hit: там
    # содержательность уже гарантирована самим ``_content_hit`` (>=2 совпавших терма в тексте
    # документа), стоп-лист calculation-имени к нему не относится.
    anchor_terms_full = _anchor_terms(criterion)
    hits = search_provider(dataset_id, anchor_terms_full) or []
    content_hit = _content_hit(anchor_terms_full, hits)

    candidates = list(verified_name_candidates)

    # Контентный хит может указывать на xlsx/xlsm файл, отсутствующий среди filename-кандидатов
    # (совпадение по содержимому, не по имени) — добавляем его как кандидата, если workbook_provider
    # знает про этот файл (иначе у нас нет пути на диске для openpyxl). Не проходит повторно
    # книжную проверку (см. докстринг выше) — lexical-хит уже контентно подтверждён.
    if content_hit is not None:
        hit_file_name = str(content_hit.get("file_name") or "")
        hit_ext = Path(hit_file_name).suffix.lower()
        if hit_ext in _CALC_XLSX_EXTS or hit_ext in _CALC_LEGACY_EXTS:
            already = any(str(c.get("file_name") or "") == hit_file_name for c in candidates)
            if not already:
                for f in files:
                    if str((f or {}).get("file_name") or "") == hit_file_name:
                        candidates.append(f)
                        break

    for cand in candidates:
        file_name = str(cand.get("file_name") or "")
        path = str(cand.get("path") or "")
        ext = Path(file_name).suffix.lower()

        if ext in _CALC_LEGACY_EXTS:
            # (g) legacy .xls — openpyxl не читает, честно НЕ пытаемся парсить.
            note = (f"Файл «{file_name}» — устаревший формат .xls (legacy_unsupported): "
                    "openpyxl не читает формулы .xls, конвертация в .xlsx не выполнена "
                    "автоматически. Требуется ручная проверка.")
            return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.1,
                    {"name": "excel_formulas_present", "status": "not_run"})

        try:
            count, sheets = _count_formula_cells(path)
        except Exception as exc:  # noqa: BLE001 — честный review_needed, не падение сервиса
            note = f"Не удалось прочитать «{file_name}» ({type(exc).__name__}): требуется ручная проверка."
            return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.1,
                    {"name": "excel_formulas_present", "status": "not_run"})

        if count > 0:
            sheet_ref = sheets[0] if sheets else ""
            evidence = [{
                "kind": "computed",
                "source_ref": f"{file_name}#sheet={sheet_ref}",
                "snippet": "",
                "value": f"{count} formula cells",
                "unit": "",
                "bbox": None,
                "reason": "детерминированный подсчёт формульных ячеек (openpyxl) подтверждает "
                          "расчёт с сохранёнными формулами",
            }]
            note = f"Найдено {count} формульных ячеек в «{file_name}» (лист «{sheet_ref}»)."
            return (S_SUPPORTED, A_YES, evidence, note, 0.8,
                    {"name": "excel_formulas_present", "status": "ok"})

        # (d) файл найден, формул нет — computed_issue/no с обязательным evidence.
        evidence = [{
            "kind": "computed",
            "source_ref": f"{file_name}#sheet=",
            "snippet": "",
            "value": "0 formula cells",
            "unit": "",
            "bbox": None,
            "reason": "файл найден, но не содержит сохранённых формул — таблица со значениями, "
                      "не расчёт (требование Glorax: расчёт с сохранёнными формулами)",
        }]
        note = f"Файл «{file_name}» найден, но не содержит формульных ячеек (таблица значений без формул)."
        return (S_COMPUTED_ISSUE, A_NO, evidence, note, 0.7,
                {"name": "excel_formulas_present", "status": "issue"})

    # (e) xlsx-кандидата нет вовсе, но есть контентный хит расчёта в другом (не xlsx/xlsm) файле.
    if content_hit is not None:
        hit_file_name = str(content_hit.get("file_name") or "")
        note = (f"Расчёт найден не в Excel (контентный хит в «{hit_file_name}»), файл с формулами "
                "не обнаружен в датасете — требуется ручная проверка исходника расчёта.")
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.3,
                {"name": "excel_formulas_present", "status": "not_run"})

    # (f) совсем ничего не найдено.
    return (S_REVIEW_NEEDED, A_UNKNOWN, [],
            "Файл расчёта не найден ни в inventory (xlsx/xlsm), ни по содержимому — "
            "недостаточно evidence для вывода.", 0.1,
            {"name": "excel_formulas_present", "status": "not_run"})


# ── parametric-механизм (T3.1) ──────────────────────────────────────────────────────────────

# Реестр param_rules грузится лениво один раз на процесс (module-level cache): YAML маленький
# (десяток правил), файл не меняется в рантайме, а load_param_rules сама по себе делает file IO +
# regex-компиляцию на каждое правило — не нужно повторять это на каждый item каждого run.
_PARAM_RULES_CACHE: list[Any] | None = None


def _param_rules_by_item_id() -> dict[str, list[Any]]:
    """Индекс rule_id -> список ParamRule по item_id (может быть >1 правило на item_id, хотя в
    текущем реестре T3.1 у каждого носителя ровно одно). Lazy-импорт checklist_param_rules —
    модуль не тянет yaml/regex-компиляцию, если parametric-items в template нет вовсе."""
    global _PARAM_RULES_CACHE
    from proxy.services.checklist_param_rules import load_param_rules

    if _PARAM_RULES_CACHE is None:
        _PARAM_RULES_CACHE = load_param_rules()

    by_item: dict[str, list[Any]] = {}
    for rule in _PARAM_RULES_CACHE:
        for item_id in rule.item_ids:
            by_item.setdefault(item_id, []).append(rule)
    return by_item


def _run_parametric(item_id: str, criterion: str, dataset_id: str, *,
                     search_provider: Callable[[str, list[str]], list[dict]]
                     ) -> tuple[str, str, list[dict], str, float, dict[str, Any]]:
    """parametric: (status, suggested_answer, document_evidence, model_note, confidence,
    computed_check) — implementation_plan.md §3.3, промпт T3.1 п.3.

    Порядок веток:
      (a) нет правила в glorax_param_rules.yaml для item_id -> review_needed, model_note
          "нет параметрического правила" — честно, не manual_required (граница v1: правило
          можно добавить, это не экспертная неформализуемая проверка);
      (b) есть правило -> search_provider(dataset_id, anchor_terms) даёт хиты -> extract_value
          на snippet КАЖДОГО хита (не только первого — нужно для обнаружения конфликта);
      (c) ни один хит не дал extract_value -> review_needed (значение не найдено, НЕ no);
      (d) ровно одно уникальное извлечённое значение (одно или несколько согласующихся хитов) ->
          compare кодом: ok -> supported_by_evidence/yes; issue -> computed_issue/no (evidence
          обязателен — safety-исключение, как в calculation);
      (e) >=2 РАЗНЫХ извлечённых значения в разных хитах -> computed_issue, evidence ОБА хита,
          model_note явно называет конфликт (инженер должен увидеть оба источника).
    """
    from proxy.services.checklist_param_rules import compare, extract_value

    rules_by_item = _param_rules_by_item_id()
    rules = rules_by_item.get(item_id)
    if not rules:
        note = (f"Нет параметрического правила в glorax_param_rules.yaml для {item_id} — "
                "требуется ручная проверка или добавление правила в реестр.")
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.1,
                {"name": "parametric", "status": "not_run"})

    # На item_id в текущем реестре ровно одно правило (валидируется отсутствием дублей rule_id +
    # курируемым 1:1 маппингом) — берём первое; расширение на >1 правило/item — не в скоупе T3.1.
    rule = rules[0]

    terms = _anchor_terms(criterion)
    hits = search_provider(dataset_id, terms) or []

    # matches: список (hit, extracted) для КАЖДОГО хита, где extract_value что-то нашла.
    matches: list[tuple[dict[str, Any], Any]] = []
    for hit in hits:
        snippet = str((hit or {}).get("snippet") or "")
        extracted = extract_value(snippet, rule)
        if extracted is not None:
            matches.append((hit, extracted))

    if not matches:
        note = (f"Значение параметра «{rule.parameter}» не найдено ни в одном хите — "
                "недостаточно evidence для вывода (отсутствие evidence не равно нарушению).")
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.1,
                {"name": rule.rule_id, "status": "not_run"})

    def _mk_evidence(hit: dict[str, Any], extracted: Any) -> dict[str, Any]:
        return {
            "kind": "document",
            "source_ref": str(hit.get("source_ref") or ""),
            "snippet": str(hit.get("snippet") or ""),
            "file_name": str(hit.get("file_name") or ""),
            "value": str(extracted.value),
            "unit": extracted.unit,
            "bbox": None,
            "reason": f"извлечено значение параметра «{rule.parameter}» из содержимого документа",
        }

    unique_values = {str(extracted.value) for _hit, extracted in matches}

    if len(unique_values) > 1:
        # (e) конфликт: >=2 разных значения в разных хитах — computed_issue, ОБА evidence.
        evidence = [_mk_evidence(hit, extracted) for hit, extracted in matches]
        values_repr = ", ".join(f"«{extracted.value}»" for _hit, extracted in matches)
        note = (f"Конфликт значений параметра «{rule.parameter}»: найдены разные величины в "
                f"разных источниках ({values_repr}) — требуется ручная сверка источников.")
        return (S_COMPUTED_ISSUE, A_UNKNOWN, evidence, note, 0.4,
                {"name": rule.rule_id, "status": "issue"})

    # (d) ровно одно уникальное значение (возможно, подтверждённое несколькими хитами).
    hit, extracted = matches[0]
    result = compare(extracted, rule)
    evidence = [_mk_evidence(hit, extracted)]

    if result.status == "ok":
        return (S_SUPPORTED, A_YES, evidence, result.message, 0.8,
                {"name": rule.rule_id, "status": "ok"})

    # issue — safety-исключение как в calculation: no подкреплён непустым evidence/source_ref.
    return (S_COMPUTED_ISSUE, A_NO, evidence, result.message, 0.7,
            {"name": rule.rule_id, "status": "issue"})


# ── cross_section: two-sided gate (T3.2) ────────────────────────────────────────────────────

_MISSING_SOURCES_NOTE = (
    "Исходники (ТЗ/ОПР/АГО/СТУ) не выбраны для проверки — сверка cross_section невозможна без "
    "source_dataset_ids."
)


# ── spds_formal: переиспользование doc_review_service (T2.4 (A)) ───────────────────────────

# Курируемый словарь: якорный терм критерия -> doc_review rule_ids (пространство
# config/normcontrol/gost_r_21_101_2026.yaml, id формата G21.101-2026-D*) + NK-коды
# normcontrol_service как алиасы того же семейства формальных проверок (NK-01 формат листа ->
# D4-001, NK-02 текстовый слой -> D6-001, NK-03 шифр/обозначение -> D1-011/D3-*, NK-04 ведомость
# -> D2-*) — см. промпт T2.4 и docstring normcontrol_service.py/gost_r_21_101_2026.yaml. Ключи —
# именно якорные термы (нижний регистр, как из _anchor_terms), не сырые слова критерия.
_SPDS_ANCHOR_TO_RULE_IDS: dict[str, list[str]] = {
    "ведомость": ["G21.101-2026-D2-001", "G21.101-2026-D2-002", "G21.101-2026-D2-003", "NK-04"],
    "шифр": ["G21.101-2026-D1-011", "G21.101-2026-D3-001", "G21.101-2026-D3-002", "G21.101-2026-D3-003", "NK-03"],
    "обозначение": ["G21.101-2026-D1-011", "G21.101-2026-D3-001", "G21.101-2026-D3-002", "G21.101-2026-D3-003", "NK-03"],
    "формат": ["G21.101-2026-D4-001", "NK-01"],
    "лист": ["G21.101-2026-D4-001", "NK-01"],
    "штамп": ["G21.101-2026-D4-002"],
    "основная": ["G21.101-2026-D4-002"],
    "надпись": ["G21.101-2026-D4-002"],
    "гост": ["G21.101-2026-D0-001", "G21.101-2026-D1-001"],
    "21.101": ["G21.101-2026-D0-001", "G21.101-2026-D1-001"],
    "спдс": ["G21.101-2026-D0-001", "G21.101-2026-D1-001"],
}

# NK-код -> реальный rule_id doc_review (алиас, см. словарь выше) — NK-коды сами по себе НЕ
# являются rule_id в ReviewItem (это formal_check_ids normcontrol_service, встроенные как
# evidence ВНУТРИ D-item'ов), поэтому перед поиском в doc_review_provider()["items"] NK-код
# разворачивается в набор реальных rule_id, которые он покрывает.
_NK_ALIAS_TO_RULE_IDS: dict[str, list[str]] = {
    "NK-01": ["G21.101-2026-D4-001"],
    "NK-02": ["G21.101-2026-D6-001"],
    "NK-03": ["G21.101-2026-D1-011", "G21.101-2026-D3-001", "G21.101-2026-D3-002", "G21.101-2026-D3-003"],
    "NK-04": ["G21.101-2026-D2-001", "G21.101-2026-D2-002", "G21.101-2026-D2-003"],
}


def _spds_rule_ids_for_criterion(criterion: str) -> list[str]:
    """Разворачивает якорные термы критерия в множество doc_review rule_ids по курируемому
    словарю ``_SPDS_ANCHOR_TO_RULE_IDS`` (NK-алиасы разворачиваются в реальные rule_id через
    ``_NK_ALIAS_TO_RULE_IDS``). Порядок стабильный (по первому вхождению), дублей нет."""
    terms = _anchor_terms(criterion)
    out: list[str] = []
    for term in terms:
        for raw_id in _SPDS_ANCHOR_TO_RULE_IDS.get(term, []):
            for rule_id in _NK_ALIAS_TO_RULE_IDS.get(raw_id, [raw_id]):
                if rule_id not in out:
                    out.append(rule_id)
    return out


def _run_spds_formal(criterion: str, dataset_id: str, *,
                      doc_review_provider: Callable[[str], dict[str, Any] | None] | None,
                      ) -> tuple[str, str, list[dict], str, float, dict[str, Any], list[str]]:
    """spds_formal: (status, suggested_answer, document_evidence, model_note, confidence,
    computed_check, doc_review_item_ids) — implementation_plan.md §3, промпт T2.4 (A).

    Переиспользует doc_review вместо отдельного движка (канон: "не дублировать Doc Review /
    Formal Checker отдельным чек-листовым движком"). Порядок веток (проверяются именно в этом
    порядке — computed_issue решает раньше полноты покрытия, см. промпт T2.4):
      (a) doc_review_provider is None -> review_needed, model_note "doc_review недоступен";
      (b) провайдер вызван, вернул None (нет документов/ошибка) -> review_needed;
      (c) якорные термы критерия не матчат ни один rule_id в курируемом словаре -> review_needed,
          doc_review_item_ids=[] (честно, не выдумываем маппинг);
      (f) среди НАЙДЕННЫХ (пересечение замапленных rule_id и doc_review_provider()["items"])
          есть хотя бы один item status=computed_issue И непустой source_ref (в
          document_evidence doc_review item) -> computed_issue/no (safety-исключение, evidence
          обязателен и есть) — НЕЗАВИСИМО от того, найдены ли остальные замапленные rule_id
          (нарушение уже доказано кодом, ждать остальное не нужно);
      (g) computed_issue есть среди найденных, но БЕЗ непустого source_ref -> manual_required
          (safety не позволяет no без source_ref, а честного computed-подтверждения тоже нет) —
          suggested_answer=no запрещён без evidence, поэтому используется manual_required, а не
          review_needed (нарушение УЖЕ найдено кодом doc_review, просто без ссылки — инженер
          должен посмотреть сам, это не "недостаточно evidence вовсе");
      (d) нет ни одного computed_issue: rule_id, отсутствующий в результате doc_review, означает
          "правило не проверено" -> для supported требуется ПОЛНОЕ покрытие всех замапленных
          rule_id, иначе review_needed (не подгоняем частичное покрытие под yes);
      (e) все НАЙДЕННЫЕ замапленные items имеют status=supported_by_evidence (и все rule_id
          покрыты, ветка (d) не сработала) -> supported_by_evidence/yes, evidence по каждому
          найденному doc_review item (kind=computed, source_ref = исходный source_ref doc_review
          item, если он есть — предпочитаем исходный, он конкретнее, иначе rule_id).
    """
    rule_ids = _spds_rule_ids_for_criterion(criterion)
    if not rule_ids:
        note = ("Критерий не сопоставлен ни с одним правилом doc_review (курируемый словарь "
                "якорей пуст для этого критерия) — требуется ручная проверка или расширение "
                "словаря _SPDS_ANCHOR_TO_RULE_IDS.")
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.1,
                {"name": "spds_formal", "status": "not_run"}, [])

    if doc_review_provider is None:
        note = "doc_review недоступен (provider не передан) — сверка spds_formal невозможна."
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.1,
                {"name": "spds_formal", "status": "not_run"}, [])

    doc_review_result = doc_review_provider(dataset_id)
    if not doc_review_result:
        note = "doc_review не вернул результат для датасета (нет документов или ошибка проверки)."
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.1,
                {"name": "spds_formal", "status": "not_run"}, [])

    dr_items_by_rule: dict[str, dict[str, Any]] = {}
    for dr_item in doc_review_result.get("items") or []:
        rid = str((dr_item or {}).get("rule_id") or "")
        if rid:
            dr_items_by_rule[rid] = dr_item

    found = [(rid, dr_items_by_rule[rid]) for rid in rule_ids if rid in dr_items_by_rule]
    found_ids = {rid for rid, _ in found}

    def _mk_evidence(rid: str, dr_item: dict[str, Any]) -> dict[str, Any]:
        source_ref = ""
        for ev in dr_item.get("document_evidence") or []:
            ref = str((ev or {}).get("source_ref") or "").strip()
            if ref:
                source_ref = ref
                break
        return {
            "kind": "computed",
            "source_ref": source_ref or rid,
            "snippet": "",
            "value": str(dr_item.get("status") or ""),
            "unit": "",
            "bbox": None,
            "reason": f"результат формальной проверки doc_review ({rid})",
        }

    # (f)/(g) computed_issue среди НАЙДЕННЫХ решает сразу, НЕЗАВИСИМО от полноты покрытия
    # rule_ids (промпт: "есть computed_issue -> computed_issue" — нарушение уже доказано кодом,
    # незачем ждать, пока найдутся остальные замапленные правила).
    issue_items = [(rid, dr_item) for rid, dr_item in found if dr_item.get("status") == S_COMPUTED_ISSUE]
    issue_with_ref = [
        (rid, dr_item) for rid, dr_item in issue_items
        if any(str((ev or {}).get("source_ref") or "").strip() for ev in dr_item.get("document_evidence") or [])
    ]
    if issue_with_ref:
        evidence = [_mk_evidence(rid, dr_item) for rid, dr_item in issue_with_ref]
        note = (f"Формальная проверка doc_review нашла нарушение ({', '.join(sorted(rid for rid, _ in issue_with_ref))})"
                " с подтверждённым источником.")
        return (S_COMPUTED_ISSUE, A_NO, evidence, note, 0.7,
                {"name": "spds_formal", "status": "issue"}, sorted(found_ids))

    if issue_items:
        note = (f"Формальная проверка doc_review нашла нарушение ({', '.join(sorted(rid for rid, _ in issue_items))}),"
                " но без source_ref — safety запрещает suggested_answer=no без evidence, требуется "
                "ручная проверка.")
        return (S_MANUAL, A_MANUAL_REQUIRED, [], note, 0.3,
                {"name": "spds_formal", "status": "issue"}, sorted(found_ids))

    # (d) нет ни одного issue — для supported нужно ПОЛНОЕ покрытие всех замапленных rule_id
    # (частичное совпадение не даёт достаточной уверенности для yes).
    if not found or found_ids != set(rule_ids):
        note = (f"Найдены не все замапленные doc_review-правила ({sorted(found_ids)} из "
                f"{sorted(rule_ids)}) — недостаточно данных для вывода, требуется ручная проверка "
                "или повторный прогон doc_review.")
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.2,
                {"name": "spds_formal", "status": "not_run"}, sorted(found_ids))

    all_supported = all(dr_item.get("status") == S_SUPPORTED for _rid, dr_item in found)
    if all_supported:
        evidence = [_mk_evidence(rid, dr_item) for rid, dr_item in found]
        note = f"Все замапленные doc_review-правила ({', '.join(sorted(found_ids))}) подтверждены."
        return (S_SUPPORTED, A_YES, evidence, note, 0.8,
                {"name": "spds_formal", "status": "ok"}, sorted(found_ids))

    # Найдены все правила, ни supported целиком, ни computed_issue (например manual/review_needed
    # со стороны doc_review) — честный review_needed, не выдумываем вердикт.
    note = (f"Замапленные doc_review-правила ({', '.join(sorted(found_ids))}) не дали однозначного "
            "supported/computed_issue — требуется ручная проверка.")
    return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.2,
            {"name": "spds_formal", "status": "not_run"}, sorted(found_ids))


def _run_cross_section(criterion: str, dataset_id: str, source_dataset_ids: list[str], *,
                        search_provider: Callable[[str, list[str]], list[dict]]
                        ) -> tuple[str, str, list[dict], str, float, dict[str, Any]]:
    """cross_section: (status, suggested_answer, document_evidence, model_note, confidence,
    computed_check) — implementation_plan.md §3.2, промпт T3.2 (B).

    Two-sided gate: evidence обязателен С ДВУХ сторон — project-хит (dataset_id, kind
    "project_doc" в evidence) И source-хит (любой из source_dataset_ids, kind "source_doc").
    LLM-связывание содержимого обеих сторон — Phase 5, здесь НЕ вызывается: suggested_answer
    НИКОГДА не бывает yes/no в этой фазе (нет механизма сверки содержимого — честно review_needed/
    unknown всегда), даже когда обе стороны найдены.

    Порядок веток:
      (a) source_dataset_ids пуст/не задан -> review_needed + model_note про невыбранные
          исходники (ТЗ/ОПР/АГО/СТУ) — вызывающий код (run_checklist_review) добавляет
          соответствующий blocker в workflow_plan;
      (b) обе стороны найдены -> review_needed/unknown, оба evidence (project_doc + source_doc),
          model_note «обе стороны найдены — сверка за инженером/моделью»;
      (c) только одна сторона -> review_needed, model_note называет отсутствующую сторону;
      (d) ничего не найдено -> review_needed (общий случай, отдельно от ветки (a)).
    """
    terms = _anchor_terms(criterion)

    if not source_dataset_ids:
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], _MISSING_SOURCES_NOTE, 0.1,
                {"name": "cross_section", "status": "not_run"})

    project_hits = search_provider(dataset_id, terms) or []
    project_hit = _content_hit(terms, project_hits)

    source_hit: dict[str, Any] | None = None
    source_ds_used: str | None = None
    for src_ds in source_dataset_ids:
        src_hits = search_provider(src_ds, terms) or []
        hit = _content_hit(terms, src_hits)
        if hit is not None:
            source_hit = hit
            source_ds_used = src_ds
            break

    if project_hit is not None and source_hit is not None:
        evidence = [
            {
                "kind": "project_doc",
                "source_ref": str(project_hit.get("source_ref") or ""),
                "snippet": str(project_hit.get("snippet") or ""),
                "file_name": str(project_hit.get("file_name") or ""),
                "value": "", "unit": "", "bbox": None,
                "reason": "project-хит подтверждает наличие проверяемого раздела в проектной документации",
            },
            {
                "kind": "source_doc",
                "source_ref": str(source_hit.get("source_ref") or ""),
                "snippet": str(source_hit.get("snippet") or ""),
                "file_name": str(source_hit.get("file_name") or ""),
                "value": "", "unit": "", "bbox": None,
                "reason": f"source-хит из датасета-исходника ({source_ds_used}) подтверждает наличие эталонных требований",
            },
        ]
        note = ("Обе стороны найдены — проектный раздел и исходник (ТЗ/ОПР/АГО/СТУ). Содержательная "
                "сверка деталей — за инженером/моделью (LLM-связывание вне скоупа этой фазы).")
        return (S_REVIEW_NEEDED, A_UNKNOWN, evidence, note, 0.4,
                {"name": "cross_section", "status": "not_run"})

    if project_hit is not None and source_hit is None:
        note = ("Найден project-хит (раздел проектной документации), но source-хит (ТЗ/ОПР/АГО/СТУ) "
                "не найден ни в одном из выбранных source_dataset_ids — сверка невозможна без "
                "второй стороны.")
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.2,
                {"name": "cross_section", "status": "not_run"})

    if project_hit is None and source_hit is not None:
        note = ("Найден source-хит (ТЗ/ОПР/АГО/СТУ), но проектная (раздел ПД) сторона не найдена в "
                "dataset_id — сверка невозможна без второй стороны.")
        return (S_REVIEW_NEEDED, A_UNKNOWN, [], note, 0.2,
                {"name": "cross_section", "status": "not_run"})

    return (S_REVIEW_NEEDED, A_UNKNOWN, [],
            "Ни проектная сторона, ни источник (ТЗ/ОПР/АГО/СТУ) не найдены — недостаточно evidence "
            "для вывода.", 0.1,
            {"name": "cross_section", "status": "not_run"})


# ── диспетчер по kind ────────────────────────────────────────────────────────────────────────


def _run_item(item: dict[str, Any], dataset_id: str, *,
              inventory_provider: Callable[[str], list[dict]],
              search_provider: Callable[[str, list[str]], list[dict]],
              workbook_provider: Callable[[str], list[dict]],
              source_dataset_ids: list[str] | None = None,
              doc_review_provider: Callable[[str], dict[str, Any] | None] | None = None) -> dict[str, Any]:
    kind = item.get("kind", "manual_required")
    criterion = str(item.get("criterion") or "")
    item_id = str(item.get("id", ""))

    # T3.2 (A): параметрическое правило запускается НЕЗАВИСИМО от kind — если для item_id есть
    # правило в реестре glorax_param_rules.yaml, кодовая параметрическая проверка исполняется
    # всегда (gap T3.1, см. Запись 9 SESSION_LOG). kind=parametric — прежнее поведение (ok->
    # supported/yes, issue->computed_issue/no), реализовано ниже без изменений. Для СОСТАВНЫХ
    # пунктов (kind != parametric, обычно manual_required) параметрическая проверка — лишь
    # ЧАСТЬ критерия: код НЕ имеет права выдать yes/no за инженера по всему составному пункту,
    # поэтому suggested_answer всегда остаётся manual_required (не переопределяет kind-путь),
    # а status меняется на supported_by_evidence/computed_issue только чтобы явно показать, что
    # именно параметрическая часть уже проверена кодом (см. model_note).
    if kind != "parametric" and _param_rules_by_item_id().get(item_id):
        p_status, _p_suggested, p_evidence, p_note, p_confidence, p_computed_check = _run_parametric(
            item_id, criterion, dataset_id, search_provider=search_provider,
        )
        if p_status == S_SUPPORTED:
            note = (
                f"Параметрическая часть подтверждена кодом (правило {p_computed_check['name']}): "
                f"{p_note} Остальная часть критерия требует инженера."
            )
            return _finalize_item(
                item, kind, criterion,
                status=S_SUPPORTED, suggested=A_MANUAL_REQUIRED, evidence=p_evidence,
                note=note, confidence=p_confidence, computed_check=p_computed_check,
            )
        if p_status == S_COMPUTED_ISSUE:
            note = (
                f"Параметрическая часть критерия НАРУШЕНА (доказано кодом, правило "
                f"{p_computed_check['name']}): {p_note} Остальная часть критерия требует инженера."
            )
            return _finalize_item(
                item, kind, criterion,
                status=S_COMPUTED_ISSUE, suggested=A_MANUAL_REQUIRED, evidence=p_evidence,
                note=note, confidence=p_confidence, computed_check=p_computed_check,
            )
        # p_status в {review_needed} (правило не нашло значение либо правила нет — не может быть
        # здесь, т.к. мы уже проверили наличие правила выше) -> честно падаем в прежний путь kind.

    if kind == "presence":
        status, suggested, evidence, note, confidence = _run_presence(
            criterion, dataset_id,
            inventory_provider=inventory_provider, search_provider=search_provider,
        )
        computed_check = {"name": "presence", "status": "ok" if status == S_SUPPORTED else "not_run"}
    elif kind == "calculation":
        status, suggested, evidence, note, confidence, computed_check = _run_calculation(
            criterion, dataset_id,
            inventory_provider=inventory_provider, search_provider=search_provider,
            workbook_provider=workbook_provider,
        )
    elif kind == "parametric":
        status, suggested, evidence, note, confidence, computed_check = _run_parametric(
            item_id, criterion, dataset_id, search_provider=search_provider,
        )
    elif kind == "cross_section":
        status, suggested, evidence, note, confidence, computed_check = _run_cross_section(
            criterion, dataset_id, source_dataset_ids or [], search_provider=search_provider,
        )
    elif kind == "spds_formal":
        status, suggested, evidence, note, confidence, computed_check, dr_ids = _run_spds_formal(
            criterion, dataset_id, doc_review_provider=doc_review_provider,
        )
        return _finalize_item(
            item, kind, criterion,
            status=status, suggested=suggested, evidence=evidence,
            note=note, confidence=confidence, computed_check=computed_check,
            doc_review_item_ids=dr_ids,
        )
    elif kind == "spatial_visual":
        status = S_MANUAL
        suggested = A_MANUAL_REQUIRED
        evidence = []
        note = "Пространственная/визуальная проверка — честная граница v1, требуется инженер."
        confidence = 0.0
        computed_check = {"name": "spatial_visual", "status": "not_run"}
    elif kind in _STUB_KINDS:
        status = S_REVIEW_NEEDED
        suggested = A_UNKNOWN
        evidence = []
        note = f"Механизм проверки для kind={kind} ещё не реализован (заглушка)."
        confidence = 0.0
        computed_check = {"name": kind, "status": "not_run"}
    else:  # manual_required и любые нераспознанные kind — честная ручная проверка
        status = S_MANUAL
        suggested = A_MANUAL_REQUIRED
        evidence = []
        note = "Пункт требует инженерного решения, не формализуется."
        confidence = 0.0
        computed_check = {"name": "manual_required", "status": "not_run"}

    return _finalize_item(
        item, kind, criterion,
        status=status, suggested=suggested, evidence=evidence,
        note=note, confidence=confidence, computed_check=computed_check,
    )


def _enforce_evidence_guard(result: dict[str, Any]) -> dict[str, Any]:
    """Центральный runtime-guard контракта (audit response 2026-07-04):
    item с ``suggested_answer`` yes/no обязан иметь хотя бы один непустой
    ``document_evidence[].source_ref``. Нарушение (возможное только из-за бага
    будущего механизма — текущие покрыты тестами) не роняет run, а честно
    понижает item до ``review_needed``/``unknown`` с пометкой в model_note."""
    if result.get("suggested_answer") not in ("yes", "no"):
        return result
    refs = [e.get("source_ref", "") for e in result.get("document_evidence", [])]
    if any(str(r).strip() for r in refs):
        return result
    note = result.get("model_note", "")
    result["status"] = "review_needed"
    result["suggested_answer"] = "unknown"
    result["model_note"] = (note + " | " if note else "") + (
        "evidence-guard: ответ yes/no без source_ref недопустим, понижено до review_needed")
    return result


def _finalize_item(item: dict[str, Any], kind: str, criterion: str, *,
                    status: str, suggested: str, evidence: list[dict[str, Any]],
                    note: str, confidence: float, computed_check: dict[str, Any],
                    doc_review_item_ids: list[str] | None = None) -> dict[str, Any]:
    """Собирает результирующий item-словарь контракта ``checklist_review_v1`` — общий хвост для
    всех веток диспетчера ``_run_item`` (обычный kind-путь И composite-параметрический путь T3.2).
    ``doc_review_item_ids`` — заполняется только веткой spds_formal (T2.4 (A)), для всех
    остальных kind остаётся пустым списком (по умолчанию).
    Результат проходит через ``_enforce_evidence_guard`` — единая точка
    контрактного инварианта yes/no => source_ref для всех механизмов."""
    return _enforce_evidence_guard({
        "item_id": item.get("id", ""),
        "sheet_name": item.get("sheet_name", ""),
        "row": item.get("row"),
        "item_no": item.get("item_no", ""),
        "section_path": item.get("section_path", []),
        "criterion": criterion,
        "kind": kind,
        "discipline": item.get("discipline", ""),
        "allowed_answers": item.get("allowed_answers", []),
        "status": status,
        "suggested_answer": suggested,
        "confidence": confidence,
        "requirement_refs": [],
        "document_evidence": evidence,
        "computed_check": computed_check,
        "model_note": note,
        "human_decision": "unset",
        "human_answer": "unset",
        "human_comment": "",
        "doc_review_item_ids": list(doc_review_item_ids or []),
        "formal_check_ids": [],
        "normalized_remark_ids": [],
    })


def _summarize(items: list[dict[str, Any]]) -> dict[str, Any]:
    by_status: dict[str, int] = {}
    by_kind: dict[str, int] = {}
    by_discipline: dict[str, int] = {}
    suggested = {k: 0 for k in _SUGGESTED_KEYS}
    source_backed = 0
    without_evidence = 0
    for it in items:
        by_status[it["status"]] = by_status.get(it["status"], 0) + 1
        by_kind[it["kind"]] = by_kind.get(it["kind"], 0) + 1
        disc = it.get("discipline", "")
        by_discipline[disc] = by_discipline.get(disc, 0) + 1
        ans = it["suggested_answer"]
        if ans in suggested:
            suggested[ans] += 1
        has_ref = any((ev or {}).get("source_ref") for ev in it.get("document_evidence") or [])
        if has_ref:
            source_backed += 1
        else:
            without_evidence += 1
    return {
        "total": len(items),
        "by_status": by_status,
        "by_kind": by_kind,
        "by_discipline": by_discipline,
        "suggested": suggested,
        "source_backed": source_backed,
        "without_evidence": without_evidence,
        "human_final_required": True,
    }


def _no_workbooks(dataset_id: str) -> list[dict]:
    """Дефолт ``workbook_provider`` для вызовов без calculation-датасета (обратная совместимость
    T2.1-тестов: параметр опционален, отсутствие -> calculation-items честно review_needed)."""
    return []


# ── normalized_remarks + defense (T2.4 (B)) ─────────────────────────────────────────────────

# Статусы item'ов, для которых формируется normalized_remark: инженеру нужно явно увидеть пункт
# в списке замечаний для принятия решения. review_needed/not_applicable НЕ входят сюда —
# отсутствие evidence не является замечанием (implementation_plan.md §4 правило 2), а
# not_applicable — заведомо не требует внимания (в текущем движке not_applicable не производится
# ни одним механизмом, но исключение оставлено явным на будущее).
_REMARK_STATUSES = {S_COMPUTED_ISSUE, S_SUPPORTED, S_MANUAL}


def _build_normalized_remarks(items: list[dict[str, Any]], template: dict[str, Any]) -> list[dict[str, Any]]:
    """normalized_remark_v1 per item со статусом из ``_REMARK_STATUSES`` — формат как
    ``doc_review_service.review_to_normalized_remarks`` (см. образец), но ``source="checklist"``,
    ``category="checklist"``, ``checklist_ref={template, item_id}`` (implementation_plan.md §4,
    CHECKLIST_REVIEW_PD_TASK.md §4 "Normalized remark"). ``requirement_ref`` заполняется из
    ``computed_check.name`` для parametric-правил (совпадает с ``rule_id`` в
    ``glorax_param_rules.yaml``) — единственная доступная на сейчас ссылка "на что проверяли",
    для остальных механизмов остаётся пустой строкой (честно, нет отдельного requirement-реестра
    у presence/calculation/cross_section/spds_formal в этой фазе). Мутирует переданные ``items``:
    проставляет ``item["normalized_remark_ids"]`` — по контракту (§5 канона) поле должно быть
    заполнено на самом item'е, не только в отдельном списке remarks."""
    template_name = str(template.get("name") or "")
    out: list[dict[str, Any]] = []
    for it in items:
        if it["status"] not in _REMARK_STATUSES:
            continue
        source_refs = [
            str((ev or {}).get("source_ref") or "").strip()
            for ev in it.get("document_evidence") or []
        ]
        source_refs = [ref for ref in source_refs if ref]

        requirement_ref = ""
        computed_check = it.get("computed_check") or {}
        cc_name = str(computed_check.get("name") or "")
        # parametric/composite-параметрическая ветка кладёт rule_id в computed_check["name"] —
        # это единственный на сейчас доступный "requirement" (правило из glorax_param_rules.yaml).
        # Общие имена механизмов (presence/calculation/spds_formal/...) не являются requirement.
        if cc_name and cc_name not in {"presence", "excel_formulas_present", "spds_formal",
                                        "cross_section", "spatial_visual", "manual_required"}:
            requirement_ref = cc_name

        remark_id = f"REM-CHK-{it['item_id']}"
        remark = {
            "schema": "normalized_remark_v1",
            "id": remark_id,
            "source": "checklist",
            "category": "checklist",
            "severity": "info",
            "status": it["status"],
            "target": it.get("criterion", ""),
            "clause": it.get("item_no", ""),
            "requirement_ref": requirement_ref,
            "document_refs": source_refs,
            "source_refs": source_refs,
            "checklist_ref": {"template": template_name, "item_id": it["item_id"]},
            "computed_check": computed_check,
            "message": it.get("model_note", ""),
            "human_decision": it.get("human_decision", "unset"),
            "human_comment": it.get("human_comment", ""),
            "human_decided_at": "",
            "finality": "proposed" if it.get("human_decision", "unset") == "unset" else "human_decided",
            "requires_human": True,
            "confidence": it.get("confidence", 0.0),
        }
        out.append(remark)
        it["normalized_remark_ids"] = [remark_id]
    return out


def _build_defense(items: list[dict[str, Any]], template: dict[str, Any],
                    normalized_remarks: list[dict[str, Any]]) -> dict[str, Any]:
    """DefensePack (``evidence_contract.DefenseClaim``/``DefensePack``, переиспользуем структуру —
    не копипастим руками, см. ``doc_review_service.review_defense_pack`` как образец), schema
    ``defense_contract_v1``, ``domain="normcontrol.checklist_review"``. Один claim на каждый item
    с normalized_remark (computed_issue/supported_by_evidence/manual_required) — 1:1 с
    ``normalized_remarks`` по построению (``_build_normalized_remarks`` уже отфильтровала статусы)."""
    from proxy.services.evidence_contract import DefenseClaim, DefensePack, DefenseStatus

    by_item_id = {it["item_id"]: it for it in items}
    claims: list[DefenseClaim] = []
    for remark in normalized_remarks:
        item_id = remark["checklist_ref"]["item_id"]
        it = by_item_id[item_id]
        source_refs = list(remark["source_refs"])

        if it["status"] == S_COMPUTED_ISSUE:
            status = DefenseStatus.COMPUTED
            gaps: list[str] = []
            actions = ["Инженеру подтвердить/отклонить замечание."]
        elif it["status"] == S_SUPPORTED:
            status = DefenseStatus.SUPPORTED
            gaps = []
            actions = []
        else:  # S_MANUAL
            status = DefenseStatus.MANUAL_REQUIRED
            gaps = ["Автопроверка не финализирует этот пункт."]
            actions = ["Проверить инженером."]

        claims.append(DefenseClaim(
            id=item_id,
            domain="normcontrol.checklist_review",
            title=it.get("criterion", ""),
            statement=f"{item_id}: {it.get('model_note') or it['status']}",
            status=status,
            severity=remark.get("severity", ""),
            source_refs=source_refs,
            inputs=[{"name": "computed_check", "value": it.get("computed_check")}],
            gaps=gaps,
            actions=([] if it.get("human_decision") in {"confirmed", "rejected"} else actions),
            confidence=it.get("confidence"),
        ))

    by_status: dict[str, int] = {}
    for it in items:
        by_status[it["status"]] = by_status.get(it["status"], 0) + 1

    return DefensePack(
        domain="normcontrol.checklist_review",
        title=f"Чек-лист {template.get('name', '')}",
        status=DefenseStatus.MANUAL_REQUIRED,
        claims=claims,
        summary={
            "template": template.get("name", ""),
            "total": len(items),
            "by_status": by_status,
            "human_final_required": True,
        },
        coverage={"items": len(items), "source_backed": sum(1 for c in claims if c.source_refs)},
        required_actions=["Финальное решение по каждому пункту ставит инженер."],
    ).payload()


def _run_pp87_composition(dataset_id: str, pp87_config: dict[str, Any] | None, *,
                           inventory_provider: Callable[[str], list[dict]]) -> dict[str, Any] | None:
    """T2.5: обёртка над ``pp87_composition_service.check_composition`` — lazy-импорт (модуль не
    тянут, если ``pp87_config`` не передан, обратная совместимость). ``None`` -> ``None`` (честно,
    механизм не запускался вовсе, это не ``review_needed`` — отдельный top-level срез контракта,
    не item со статусом)."""
    if pp87_config is None:
        return None

    from proxy.services.pp87_composition_service import check_composition

    inventory = inventory_provider(dataset_id) or []
    result = check_composition(inventory, pp87_config)
    return {"schema": "pp87_composition_v1", **result}


def run_checklist_review(template: dict[str, Any], *, dataset_id: str,
                          source_dataset_ids: list[str] | None = None,
                          discipline: str | None = None,
                          inventory_provider: Callable[[str], list[dict]],
                          search_provider: Callable[[str, list[str]], list[dict]],
                          workbook_provider: Callable[[str], list[dict]] | None = None,
                          doc_review_provider: Callable[[str], dict[str, Any] | None] | None = None,
                          pp87_config: dict[str, Any] | None = None,
                          ) -> dict[str, Any]:
    """Прогоняет ``template`` (dict из ``checklist_template_importer``/``load_checklist_template``)
    по ``dataset_id`` и возвращает ``checklist_review_v1`` (implementation_plan.md §4).

    Чистая функция: НИКАКИХ обращений к живым Qdrant/MLX/MetaDB — все данные приходят через
    ``inventory_provider``/``search_provider``/``workbook_provider``/``doc_review_provider``
    (dependency injection, паттерн ``run_review`` из ``doc_review_service.py``). ``discipline`` —
    опциональный фильтр (как ``DocReviewRequest.discipline``): None/пусто -> все items template'а;
    иначе — только items с совпадающей дисциплиной. ``workbook_provider`` опционален (дефолт —
    пустой список файлов) — без него calculation-items честно получают ``review_needed`` (правило
    "отсутствие evidence не равно нарушению"), это не ошибка вызова. ``doc_review_provider``
    опционален аналогично — без него spds_formal-items честно получают ``review_needed`` (T2.4 (A)).

    T2.4 (B): ``normalized_remarks``/``defense`` наполняются содержательно (см.
    ``_build_normalized_remarks``/``_build_defense``) — computed_issue/supported_by_evidence/
    manual_required items получают normalized_remark_v1 + DefenseClaim; review_needed не
    порождает remark (отсутствие evidence не равно нарушению).

    T3.2 (B): ``source_dataset_ids`` используется cross_section-механизмом (two-sided gate) —
    ``search_provider`` вызывается и для ``dataset_id`` (project-хит), и для каждого
    ``source_dataset_ids`` (source-хит). Если в template есть хотя бы один ``kind=cross_section``
    item и ``source_dataset_ids`` пуст/не задан — в ``workflow_plan.blockers`` добавляется явный
    blocker (см. ниже), помимо review_needed на уровне item'а.

    T2.5: ``pp87_config`` (опционально, dict из ``pp87_composition_service.load_pp87_config()``)
    включает composition-checker состава ПД по ПП РФ №87 — сверку разделов ПД с ``inventory_provider(
    dataset_id)``. Результат кладётся в НОВОЕ top-level поле контракта ``pp87_composition``
    (``None``, если ``pp87_config`` не передан — обратная совместимость, существующие вызовы не
    тянут pp87_composition_service вовсе). Решение СОЗНАТЕЛЬНО консервативное (промпт T2.5): ни
    один item дисциплины «Общее» в ``glorax_pd_2026.json`` не сформулирован про состав/разделы ПД
    (все 5 — про инженерные изыскания/учёт СТУ), поэтому composition НЕ привязывается к items как
    computed-evidence — только честный отдельный top-level срез, который UI/API-слой может
    показать рядом со списком items (см. SESSION_LOG.md Запись 17).
    """
    items_src = template.get("items", [])
    if discipline:
        items_src = [it for it in items_src if it.get("discipline") == discipline]

    wb_provider = workbook_provider or _no_workbooks
    src_ids = list(source_dataset_ids or [])
    items = [
        _run_item(it, dataset_id, inventory_provider=inventory_provider, search_provider=search_provider,
                  workbook_provider=wb_provider, source_dataset_ids=src_ids,
                  doc_review_provider=doc_review_provider)
        for it in items_src
    ]

    normalized_remarks = _build_normalized_remarks(items, template)
    defense = _build_defense(items, template, normalized_remarks)
    # T3.2 (B): cross_section без source_dataset_ids — не только review_needed на уровне item'а,
    # но и явный blocker на уровне workflow_plan (заглушка workflow-слоя, полноценный
    # workflow_service — вне скоупа T3.2): сверка невозможна, пока оператор не выберет исходники.
    blockers: list[dict[str, Any]] = []
    has_cross_section = any(it.get("kind") == "cross_section" for it in items_src)
    if has_cross_section and not src_ids:
        blockers.append({
            "code": "cross_section_sources_missing",
            "message": _MISSING_SOURCES_NOTE,
        })

    workflow_plan = {
        "schema": "workflow_plan_v1",
        "workflow_id": "checklist_review",
        "finality": "human_required",
        "blockers": blockers,
        "missing_inputs": [],
        "next_actions": [],
    }

    pp87_composition = _run_pp87_composition(
        dataset_id, pp87_config, inventory_provider=inventory_provider,
    )

    return {
        "schema": "checklist_review_v1",
        "run_id": "",
        "dataset_id": dataset_id,
        "source_dataset_ids": list(source_dataset_ids or []),
        "template": template.get("name", ""),
        "stage": template.get("stage", ""),
        "discipline": discipline or "all",
        "status": "done",
        "summary": _summarize(items),
        "items": items,
        "normalized_remarks": normalized_remarks,
        "defense": defense,
        "workflow_plan": workflow_plan,
        "pp87_composition": pp87_composition,
    }


# ── продакшн-провайдеры (lazy-импорт, НЕ вызываются в тестах) ──────────────────────────────


def default_inventory_provider(dataset_id: str) -> list[dict[str, Any]]:
    """Обёртка над MetaDB inventory (``project_summary_service.inventory_from_metadb``) в формате
    ``[{file_name, doc_type}]``, ожидаемом ``run_checklist_review``. Lazy-импорт — модуль
    ``checklist_review_service`` не тянет MetaDB на уровне импорта (тесты его не задевают)."""
    from proxy.services.project_summary_service import inventory_from_metadb

    summary = inventory_from_metadb([dataset_id])
    files = summary.get("files") or []
    return [
        {"file_name": str(f.get("file_name") or f.get("name") or ""),
         "doc_type": str(f.get("doc_type") or "")}
        for f in files
    ]


def default_search_provider(dataset_id: str, terms: list[str]) -> list[dict[str, Any]]:
    """Обёртка над лексическим поиском (``proxy.services.source_adapters.search_lexical_chunks``)
    в формате ``[{source_ref, snippet, file_name}]``. UNAVAILABLE/ошибка -> пустой список (честно
    «поиск не дал результата», решение о review_needed остаётся за ``_run_presence``). Lazy-импорт."""
    from proxy.services import source_adapters as sa

    import logging
    try:
        res = sa.search_lexical_chunks(terms, dataset_ids=[dataset_id] if dataset_id else None)
    except Exception as e:  # noqa: BLE001
        # Урок T2.3b: молчаливая пустота неотличима от «не нашли» — логируем причину
        logging.getLogger(__name__).warning("checklist search unavailable: %s", str(e)[:120])
        return []
    if res.status != sa.FOUND:
        if getattr(res, "warnings", None):
            logging.getLogger(__name__).warning(
                "checklist search %s: %s", res.status, "; ".join(res.warnings)[:160])
        return []
    return [
        {"source_ref": m.source_ref, "snippet": m.snippet, "file_name": m.file_name}
        for m in res.matches
    ]


# Расширения, которые может отдавать default_workbook_provider (xlsx/xlsm — читаемы openpyxl;
# .xls отдаём тоже, честно — _run_calculation сам распознаёт .xls и даёт legacy_unsupported,
# не пытаясь открыть его openpyxl).
_WORKBOOK_EXTS = (".xlsx", ".xlsm", ".xls")


def default_workbook_provider(dataset_id: str, *, meta_db_path: str | None = None) -> list[dict[str, Any]]:
    """Обёртка над MetaDB ``documents`` (та же таблица/паттерн, что ``_dataset_source_paths`` в
    ``doc_review_service.py``): отбирает файлы датасета с расширением xlsx/xlsm/xls и отдаёт
    ``[{file_name, path}]``, ожидаемый ``run_checklist_review``/``_run_calculation``.

    Путь на диске берётся из колонки ``source_path`` (внешний in-place источник, абсолютный путь,
    без копии в storage — см. ``backend/qdrant_adapter.py`` ``CREATE TABLE documents``). Файл без
    ``source_path`` пропускается — без пути на диске openpyxl нечего открывать. Lazy-импорт,
    ошибка подключения к MetaDB -> пустой список (честно «инвентарь недоступен», не выдумка)."""
    import sqlite3

    from backend.rag_config import rag_meta_db_path

    path = meta_db_path or rag_meta_db_path()
    out: list[dict[str, Any]] = []
    try:
        with sqlite3.connect(path) as conn:
            rows = conn.execute(
                "SELECT file_name, source_path FROM documents WHERE dataset_id=?",
                (dataset_id,),
            ).fetchall()
    except Exception:
        return []
    for file_name, source_path in rows:
        fn = str(file_name or "")
        sp = str(source_path or "")
        if not fn or not sp:
            continue
        if Path(fn).suffix.lower() not in _WORKBOOK_EXTS:
            continue
        out.append({"file_name": fn, "path": sp})
    return out


def default_doc_review_provider(dataset_id: str, *, rulepack: str = "gost_r_21_101_2026") -> dict[str, Any] | None:
    """Lazy-обёртка над ``doc_review_service.review_dataset`` (T2.4 (A)): переиспользует
    формальный СПДС-нормоконтроль вместо отдельного движка для kind=spds_formal (канон: "не
    дублировать Doc Review / Formal Checker отдельным чек-листовым движком"). Возвращает
    ``{"items": [{"rule_id", "status", "severity", "document_evidence": [...]}]}`` — тот же
    формат полей, что ``ReviewItem``/``review_to_normalized_remarks`` (``rule_id``/``status``/
    ``severity``/``document_evidence[].source_ref``), ожидаемый ``_run_spds_formal``.

    ``review_dataset`` поднимает ``ValueError('no_documents')``, если у датасета нет файлов —
    перехватывается и превращается в честный ``None`` (doc_review недоступен для этого датасета,
    НЕ падение всего checklist-review прогона). НЕ вызывается в тестах (см. докстринг модуля)."""
    from proxy.services.doc_review_service import review_dataset

    try:
        _review_map, items = review_dataset(dataset_id, rulepack=rulepack)
    except Exception:
        return None
    return {"items": [
        {
            "rule_id": it.rule_id,
            "status": it.status,
            "severity": it.severity,
            "document_evidence": list(it.document_evidence or []),
        }
        for it in items
    ]}


def default_pp87_config_provider() -> dict[str, Any] | None:
    """Lazy-обёртка над ``pp87_composition_service.load_pp87_config()`` (T2.5) для продакшн-вызова
    ``run_checklist_review(..., pp87_config=default_pp87_config_provider())``. Файл конфига
    отсутствует/битый -> честный ``None`` (composition-checker недоступен для этого прогона, НЕ
    падение всего checklist-review — тот же паттерн, что ``default_doc_review_provider``). НЕ
    вызывается в тестах (см. докстринг модуля)."""
    from proxy.services.pp87_composition_service import load_pp87_config

    try:
        return load_pp87_config()
    except Exception:
        return None

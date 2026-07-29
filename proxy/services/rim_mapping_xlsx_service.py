"""XLSX round-trip for model/user-authored VOR-to-norm mapping revisions."""

from __future__ import annotations

from datetime import datetime, timezone
import json
from pathlib import Path
from typing import Any
from uuid import uuid4


MAPPING_SHEET = "ВОР—ГЭСН"
MANIFEST_SHEET = "__LES_MANIFEST"
MAPPING_COLUMNS = (
    ("mapping_row_id", "ID связи"),
    ("work_id", "ID работы"),
    ("norm_key", "Ключ нормы"),
    ("norm_code", "Шифр нормы"),
    ("norm_title", "Наименование нормы"),
    ("norm_unit", "Измеритель нормы"),
    ("norm_quantity", "Количество нормы"),
    ("candidate_rank", "Ранг кандидата"),
    ("selection_status", "Статус выбора"),
    ("selection_kind", "Тип выбора"),
    ("is_analog", "Аналог"),
    ("card_opened", "Карточка открыта"),
    ("reason", "Причина"),
    ("source_refs", "Источники"),
    ("edited_by", "Автор правки"),
)


def _utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()


def _excel_text(value: Any) -> str:
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _from_excel_text(value: Any) -> str:
    text = str(value or "")
    if len(text) > 1 and text[0] == "'" and text[1] in "=+-@":
        return text[1:]
    return text


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().casefold() in {"1", "true", "да", "yes", "x"}


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def render_mapping_xlsx(
    mapping_rows: list[dict[str, Any]],
    out_path: str | Path,
    *,
    session_id: str,
    parent_revision_id: str,
    vor_revision_id: str,
) -> Path:
    """Render editable mapping plus a hidden binding manifest."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    target = Path(out_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = MAPPING_SHEET
    header_fill = PatternFill("solid", fgColor="DDE8DF")
    for column_index, (_field, title) in enumerate(MAPPING_COLUMNS, 1):
        cell = sheet.cell(row=1, column=column_index, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_index, mapping in enumerate(mapping_rows, 2):
        for column_index, (field, _title) in enumerate(MAPPING_COLUMNS, 1):
            value: Any = mapping.get(field)
            if field == "source_refs":
                value = json.dumps(value or [], ensure_ascii=False)
            elif field in {"is_analog", "card_opened"}:
                value = "да" if bool(value) else "нет"
            elif isinstance(value, str):
                value = _excel_text(value)
            sheet.cell(row=row_index, column=column_index, value=value)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:O{max(2, len(mapping_rows) + 1)}"
    widths = (36, 18, 28, 22, 48, 18, 18, 16, 20, 18, 12, 18, 56, 48, 18)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

    instructions = workbook.create_sheet("Инструкция")
    instructions.append(["Поле", "Правило"])
    instructions.append(
        [
            "Статус выбора",
            "candidate | accepted | selected | rejected | conflict. "
            "Код не выбирает значение автоматически.",
        ]
    )
    instructions.append(
        ["Ключ нормы", "Полная идентичность с семейством, например ГЭСНм:10-06-001-01."]
    )
    instructions.append(
        ["Карточка открыта", "Для selected должно быть «да»: норма прочитана из typed SQLite."]
    )
    instructions.append(
        ["Импорт", "Создаёт новую immutable mapping-ревизию; предыдущая не перезаписывается."]
    )
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 100

    manifest = {
        "schema": "rim_mapping_xlsx_manifest_v1",
        "session_id": session_id,
        "parent_revision_id": parent_revision_id,
        "vor_revision_id": vor_revision_id,
        "exported_at": _utcnow(),
        "mapping_sheet": MAPPING_SHEET,
        "columns": [field for field, _title in MAPPING_COLUMNS],
    }
    manifest_sheet = workbook.create_sheet(MANIFEST_SHEET)
    manifest_sheet["A1"] = json.dumps(manifest, ensure_ascii=False, sort_keys=True)
    manifest_sheet.sheet_state = "veryHidden"
    temp = target.with_suffix(target.suffix + ".tmp")
    workbook.save(temp)
    workbook.close()
    temp.replace(target)
    return target


def read_mapping_xlsx(
    path: str | Path,
    *,
    expected_session_id: str,
    expected_vor_revision_id: str,
) -> dict[str, Any]:
    """Read an editable mapping file without accepting formulas as data."""
    import openpyxl

    source = Path(path)
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=False)
    try:
        if MANIFEST_SHEET not in workbook.sheetnames:
            raise ValueError("mapping workbook has no LES manifest")
        manifest_raw = workbook[MANIFEST_SHEET]["A1"].value
        try:
            manifest = json.loads(str(manifest_raw or ""))
        except json.JSONDecodeError as error:
            raise ValueError("mapping workbook manifest is invalid") from error
        if manifest.get("schema") != "rim_mapping_xlsx_manifest_v1":
            raise ValueError("mapping workbook schema is not supported")
        if str(manifest.get("session_id") or "") != expected_session_id:
            raise ValueError("mapping workbook belongs to another session")
        if str(manifest.get("vor_revision_id") or "") != expected_vor_revision_id:
            raise ValueError("mapping workbook belongs to another VOR revision")
        sheet_name = str(manifest.get("mapping_sheet") or MAPPING_SHEET)
        if sheet_name not in workbook.sheetnames:
            raise ValueError("mapping sheet is missing")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=False)
        header_cells = next(rows, ())
        headers = {
            str(cell.value or "").strip(): index
            for index, cell in enumerate(header_cells)
            if str(cell.value or "").strip()
        }
        missing_headers = [
            title for _field, title in MAPPING_COLUMNS if title not in headers
        ]
        if missing_headers:
            raise ValueError("mapping columns are missing: " + ", ".join(missing_headers))
        mapping_rows: list[dict[str, Any]] = []
        for row_number, cells in enumerate(rows, 2):
            values: dict[str, Any] = {}
            for field, title in MAPPING_COLUMNS:
                cell = cells[headers[title]] if headers[title] < len(cells) else None
                if cell is not None and cell.data_type == "f":
                    raise ValueError(
                        f"formulas are not allowed in mapping data: row {row_number}, column {title}"
                    )
                values[field] = cell.value if cell is not None else None
            if not any(value not in (None, "") for value in values.values()):
                continue
            source_refs_raw = _from_excel_text(values.get("source_refs"))
            try:
                source_refs = json.loads(source_refs_raw) if source_refs_raw else []
            except json.JSONDecodeError as error:
                raise ValueError(f"invalid source refs JSON at row {row_number}") from error
            mapping_rows.append(
                {
                    "mapping_row_id": _from_excel_text(values.get("mapping_row_id")) or uuid4().hex,
                    "work_id": _from_excel_text(values.get("work_id")),
                    "norm_key": _from_excel_text(values.get("norm_key")),
                    "norm_code": _from_excel_text(values.get("norm_code")),
                    "norm_title": _from_excel_text(values.get("norm_title")),
                    "norm_unit": _from_excel_text(values.get("norm_unit")),
                    "norm_quantity": _number(values.get("norm_quantity")),
                    "candidate_rank": int(_number(values.get("candidate_rank")) or 0),
                    "selection_status": _from_excel_text(values.get("selection_status")) or "candidate",
                    "selection_kind": _from_excel_text(values.get("selection_kind")),
                    "is_analog": _bool(values.get("is_analog")),
                    "card_opened": _bool(values.get("card_opened")),
                    "reason": _from_excel_text(values.get("reason")),
                    "source_refs": source_refs,
                    "edited_by": _from_excel_text(values.get("edited_by")) or "user",
                }
            )
        return {"manifest": manifest, "mapping_rows": mapping_rows}
    finally:
        workbook.close()

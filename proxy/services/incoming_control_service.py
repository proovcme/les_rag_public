"""W20.4 — Входной контроль материалов/изделий/оборудования. 0 LLM (ADR-11).

ГОСТ 24297-2013, СП 48.13330 п. 7.1.3. Две таблицы:
- `les_quality_docs` — реестр документов о качестве (сертификат/паспорт/декларация)
  с привязкой к файлу и сроком действия (просроченные = «красные»);
- `les_incoming_control` — записи входного контроля партий (материал↔спецификация,
  документ качества, результат, решение допущено/не допущено), строго по объекту (Q3).

Из записей собираются журнал входного контроля и акт по партии (числа из SQL, не LLM).
Рёбра (W17.2): запись контроля → спецификация / документ качества / захватка.
"""
from __future__ import annotations

import sqlite3
import time
from datetime import date
from pathlib import Path
from typing import Any

from backend.rag_config import rag_meta_db_path

OUTPUT_DIR = Path("data/incoming_control_out")

ADMITTED = "допущено"
REJECTED = "не допущено"


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(rag_meta_db_path())
    conn.row_factory = sqlite3.Row
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS les_quality_docs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL DEFAULT 0,
            doc_type TEXT NOT NULL DEFAULT 'сертификат',
            number TEXT NOT NULL DEFAULT '',
            material TEXT NOT NULL DEFAULT '',
            issued_by TEXT NOT NULL DEFAULT '',
            valid_until TEXT NOT NULL DEFAULT '',
            file_id TEXT,
            created_at REAL NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS les_incoming_control (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL DEFAULT 0,
            control_date TEXT NOT NULL DEFAULT '',
            batch TEXT NOT NULL DEFAULT '',
            material TEXT NOT NULL DEFAULT '',
            spec_id INTEGER,
            quality_doc_id INTEGER,
            quantity REAL,
            unit TEXT NOT NULL DEFAULT '',
            result TEXT NOT NULL DEFAULT '',
            decision TEXT NOT NULL DEFAULT 'допущено',
            inspector TEXT NOT NULL DEFAULT '',
            zahvatka TEXT NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            created_at REAL NOT NULL
        )
        """
    )
    return conn


# ── Реестр документов о качестве ────────────────────────────────────────────

def add_quality_doc(
    project_id: int, doc_type: str = "сертификат", number: str = "", *,
    material: str = "", issued_by: str = "", valid_until: str = "", file_id: str | None = None,
) -> dict[str, Any]:
    """Зарегистрировать документ о качестве (сертификат/паспорт/декларация)."""
    now = time.time()
    with _connect() as conn:
        cur = conn.execute(
            """INSERT INTO les_quality_docs
               (project_id, doc_type, number, material, issued_by, valid_until, file_id, created_at)
               VALUES (?,?,?,?,?,?,?,?)""",
            (int(project_id), doc_type, number, material, issued_by, valid_until, file_id, now),
        )
        conn.commit()
        doc_id = int(cur.lastrowid)
    return get_quality_doc(doc_id)


def get_quality_doc(doc_id: int) -> dict[str, Any]:
    with _connect() as conn:
        row = conn.execute("SELECT * FROM les_quality_docs WHERE id=?", (int(doc_id),)).fetchone()
    return dict(row) if row else {}


def list_quality_docs(project_id: int, query: str | None = None, as_of: str | None = None) -> list[dict[str, Any]]:
    """Реестр документов качества объекта; флаг `expired` для истёкших («красные»)."""
    today = as_of or date.today().isoformat()
    with _connect() as conn:
        rows = conn.execute(
            "SELECT * FROM les_quality_docs WHERE project_id=? ORDER BY id", (int(project_id),)
        ).fetchall()
    out: list[dict[str, Any]] = []
    needle = (query or "").strip().lower()
    for row in rows:
        item = dict(row)
        if needle and needle not in " ".join(
            str(item.get(k, "")) for k in ("doc_type", "number", "material", "issued_by")
        ).lower():
            continue
        item["expired"] = bool(item.get("valid_until") and str(item["valid_until"]) < today)
        out.append(item)
    return out


# ── Записи входного контроля ────────────────────────────────────────────────

def add_incoming_control(
    project_id: int, material: str, *, batch: str = "", control_date: str = "",
    spec_id: int | None = None, quality_doc_id: int | None = None,
    quantity: float | None = None, unit: str = "", result: str = "",
    decision: str = ADMITTED, inspector: str = "", zahvatka: str = "", notes: str = "",
) -> dict[str, Any]:
    """Запись входного контроля партии. decision: 'допущено' | 'не допущено'. 0 LLM."""
    if decision not in (ADMITTED, REJECTED):
        decision = ADMITTED
    now = time.time()
    with _connect() as conn:
        cur = conn.execute(
            """INSERT INTO les_incoming_control
               (project_id, control_date, batch, material, spec_id, quality_doc_id, quantity,
                unit, result, decision, inspector, zahvatka, notes, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (int(project_id), control_date, batch, material, spec_id, quality_doc_id, quantity,
             unit, result, decision, inspector, zahvatka, notes, now),
        )
        conn.commit()
        control_id = int(cur.lastrowid)
    _link_edges(control_id, spec_id, quality_doc_id, zahvatka)
    return get_incoming_control(control_id)


def get_incoming_control(control_id: int) -> dict[str, Any]:
    with _connect() as conn:
        row = conn.execute("SELECT * FROM les_incoming_control WHERE id=?", (int(control_id),)).fetchone()
    return dict(row) if row else {}


def list_incoming_control(project_id: int, decision: str | None = None) -> list[dict[str, Any]]:
    """Записи входного контроля объекта (опц. фильтр по решению)."""
    with _connect() as conn:
        if decision:
            rows = conn.execute(
                "SELECT * FROM les_incoming_control WHERE project_id=? AND decision=? ORDER BY id",
                (int(project_id), decision),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM les_incoming_control WHERE project_id=? ORDER BY id", (int(project_id),)
            ).fetchall()
    return [dict(r) for r in rows]


def _link_edges(control_id: int, spec_id: int | None, quality_doc_id: int | None, zahvatka: str) -> None:
    """Рёбра материал(запись)→спецификация/документ качества/захватка (W17.2). Best-effort."""
    try:
        from proxy.services import edge_service
    except Exception:
        return
    src = ("incoming_control", str(control_id))
    targets = []
    if spec_id:
        targets.append(("spec", str(spec_id), "conforms_to_spec"))
    if quality_doc_id:
        targets.append(("quality_doc", str(quality_doc_id), "certified_by"))
    if zahvatka:
        targets.append(("zahvatka", zahvatka, "delivered_to"))
    for dst_kind, dst_id, edge_type in targets:
        try:
            edge_service.add_edge(src[0], src[1], dst_kind, dst_id, edge_type, method="incoming_control")
        except Exception:
            continue


# ── Журнал и акт ────────────────────────────────────────────────────────────

def _object_name(project_id: int) -> str:
    try:
        from proxy.services.project_service import get_project
        return (get_project(project_id) or {}).get("name", "") or ""
    except Exception:
        return ""


def build_journal(project_id: int, as_of: str | None = None) -> dict[str, Any]:
    """Журнал входного контроля объекта (ГОСТ 24297). Числа из SQL, 0 LLM."""
    records = list_incoming_control(project_id)
    docs = {d["id"]: d for d in list_quality_docs(project_id, as_of=as_of)}
    rows: list[dict[str, Any]] = []
    for rec in records:
        doc = docs.get(rec.get("quality_doc_id")) or {}
        rows.append({
            "id": rec["id"],
            "date": rec.get("control_date"),
            "material": rec.get("material"),
            "batch": rec.get("batch"),
            "quantity": rec.get("quantity"),
            "unit": rec.get("unit"),
            "quality_doc": f"{doc.get('doc_type', '')} № {doc.get('number', '')}".strip(" №") if doc else "",
            "quality_doc_expired": bool(doc.get("expired")),
            "result": rec.get("result"),
            "decision": rec.get("decision"),
            "inspector": rec.get("inspector"),
        })
    rejected = [r for r in rows if r["decision"] == REJECTED]
    return {
        "project_id": project_id,
        "title": "ЖУРНАЛ ВХОДНОГО КОНТРОЛЯ (ГОСТ 24297-2013)",
        "object_name": _object_name(project_id),
        "rows": rows,
        "count": len(rows),
        "admitted_count": sum(1 for r in rows if r["decision"] == ADMITTED),
        "rejected_count": len(rejected),
        "has_rejected": bool(rejected),
    }


def build_act(project_id: int, control_id: int) -> dict[str, Any]:
    """Акт входного контроля по одной партии (ГОСТ 24297). 0 LLM."""
    rec = get_incoming_control(control_id)
    if not rec or int(rec.get("project_id", -1)) != int(project_id):
        return {}
    doc = get_quality_doc(rec["quality_doc_id"]) if rec.get("quality_doc_id") else {}
    return {
        "title": "АКТ ВХОДНОГО КОНТРОЛЯ",
        "object_name": _object_name(project_id),
        "control_date": rec.get("control_date"),
        "material": rec.get("material"),
        "batch": rec.get("batch"),
        "quantity": rec.get("quantity"),
        "unit": rec.get("unit"),
        "quality_doc": {
            "type": doc.get("doc_type", ""),
            "number": doc.get("number", ""),
            "issued_by": doc.get("issued_by", ""),
            "valid_until": doc.get("valid_until", ""),
        } if doc else None,
        "result": rec.get("result"),
        "decision": rec.get("decision"),
        "inspector": rec.get("inspector"),
        "notes": rec.get("notes"),
    }


def _output_dir() -> Path:
    import os
    d = Path(os.getenv("LES_INCOMING_CONTROL_OUT_DIR", str(OUTPUT_DIR)))
    d.mkdir(parents=True, exist_ok=True)
    return d


def export_journal_xlsx(project_id: int, out_path: Path | None = None, as_of: str | None = None) -> Path:
    """Выгрузка журнала входного контроля в xlsx. Числа из SQL, не LLM."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    journal = build_journal(project_id, as_of=as_of)
    wb = Workbook()
    ws = wb.active
    ws.title = "Входной контроль"
    ws["A1"] = journal["title"]
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Объект: {journal['object_name']}"
    ws["A4"] = (f"Партий: {journal['count']}  ·  допущено: {journal['admitted_count']}  ·  "
                f"не допущено: {journal['rejected_count']}")

    headers = ["№", "Дата", "Материал/изделие", "Партия", "Кол-во", "Ед.изм.",
               "Документ о качестве", "Результат", "Решение", "Контролёр"]
    r = 6
    for c, h in enumerate(headers, 1):
        ws.cell(r, c, h).font = Font(bold=True)
    r += 1
    red = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
    for i, row in enumerate(journal["rows"], 1):
        values = [i, row["date"], row["material"], row["batch"], row["quantity"], row["unit"],
                  row["quality_doc"], row["result"], row["decision"], row["inspector"]]
        for c, v in enumerate(values, 1):
            cell = ws.cell(r, c, v)
            # «красные»: не допущено или истёкший документ о качестве
            if row["decision"] == REJECTED or row["quality_doc_expired"]:
                cell.fill = red
        r += 1
    for col, width in (("A", 5), ("B", 13), ("C", 34), ("D", 16), ("E", 10), ("F", 9),
                       ("G", 28), ("H", 22), ("I", 14), ("J", 18)):
        ws.column_dimensions[col].width = width

    out_path = Path(out_path) if out_path else _output_dir() / f"vk_{project_id}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(out_path))
    return out_path

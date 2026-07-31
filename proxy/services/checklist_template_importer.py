"""Importer чек-листов Glorax (XLSX -> ChecklistTemplate JSON), Phase 1 (T1.1/T1.2).

АРХИТЕКТУРА (docs/ALGO-glorax-checklist-review.md, docs/CHECKLIST_REVIEW_PD_TASK.md §5):
этот модуль делает ровно одно — читает исходный XLSX чек-листа Glorax и строит
`ChecklistTemplate` (словарь: name/stage/title/source_file_name/version/disciplines/items).
Он НЕ судит и не проверяет документацию — это Phase 2-3 (`checklist_review_service`).

Правило классификации строки (решение оператора 2026-07-04, зафиксировано в
docs/checklist_review/SNAPSHOT_PD.md и docs/ALGO-glorax-checklist-review.md), уточнённое
находкой T1.3 (реальный ПД-XLSX, лист ОВиК, строки 15/18 — заливка FF3200F0 приоритетнее
DV, см. `_has_block_header_fill`):

    критерий        = ячейка B непустая И answer-ячейка C покрыта data validation
                       И заливка B НЕ равна FF3200F0
    заголовок блока = ячейка B непустая И (DV на C нет ИЛИ заливка B равна FF3200F0)
                       (заливка "с заливкой"/"без заливки" — подтип для диагностики; но
                        когда заливка ЕСТЬ, она разрешает конфликт в пользу заголовка даже
                        при наличии DV на C — редкий артефакт исходника, см. T1.3)
    заголовок раздела = ячейка B пустая, A содержит текст

Итоговые строки листа («Всего пунктов:», «Из них, соответствуют критериям:»,
«Чек лист заполнен на:» и т.п.) — не критерии и не заголовки, игнорируются явным списком
паттернов (см. `_TOTAL_ROW_PATTERNS`).

Служебные листы («Правила заполнения», «Сводная») исключаются из импорта целиком.

section_path накапливается по ходу листа: последний встреченный заголовок раздела и последний
встреченный заголовок блока формируют путь для всех последующих критериев, пока не встретится
новый заголовок того же уровня.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

# --- Служебные листы, не участвующие в импорте ---
SKIP_SHEETS = {"Правила заполнения", "Сводная"}

# --- Заливка заголовка блока (диагностический подтип, не решающий признак) ---
BLOCK_HEADER_FILL = "FF3200F0"

# --- Строки-итоги в конце листа (regex по значению колонки B), игнорируются ---
_TOTAL_ROW_PATTERNS = (
    re.compile(r"^всего пунктов", re.IGNORECASE),
    re.compile(r"^из них,?\s*соответствуют", re.IGNORECASE),
    re.compile(r"^из них,?\s*не соответствуют", re.IGNORECASE),
    re.compile(r"^чек[ -]?лист заполнен", re.IGNORECASE),
)

# --- Транслитерация дисциплин ПД/РД -> латинский код для id (канон §5) ---
DISCIPLINE_TRANSLIT = {
    "Общее": "OB",
    "АР": "AR",
    "КР": "KR",
    "СПОЗУ": "SPOZU",
    "ЭОМ": "EOM",
    "ЭН": "EN",
    "ВК и НВК": "VK",
    "ОВиК": "OVIK",
    "СС": "SS",
    "ПБ2 (АППЗ)": "PB2",
}

# первые 2 строки листа (название раздела + шапка колонок) всегда служебные
_HEADER_ROWS = 2
# сколько колонок читаем (№ / Критерий / Отметка / Примечание) — не больше, по заданию
_MAX_COL = 6

# --- Эвристика классификации kind по ключевым словам критерия (T1.4) ---
#
# Порядок применения ПРИНЦИПИАЛЕН: правила проверяются последовательно, побеждает первое
# совпадение, поэтому специфичные паттерны идут раньше общих. Логика групп сверху вниз
# (уточнена по факту прогона на реальных 335 критериях glorax_pd_2026.json — расстановка
# ниже устраняет конфликты, обнаруженные на живых данных, а не только на синтетике):
#
#   1. spatial_visual  — пространственное размещение на планах (честная граница v1,
#      docs/ALGO-glorax-checklist-review.md §3.4). Идёт ПЕРВЫМ: формулировки вида
#      "над жилыми"/"под жилыми"/"корзин для кондиционеров"/"термовкладыш" часто содержат
#      слово "расчёт" или "соответств" (например "Выполнен расчёт для зоны установки
#      термовкладышей") — без приоритета они бы утекли в calculation/cross_section, хотя
#      суть критерия — геометрическое размещение, не число.
#   2. spds_formal     — оформление/состав комплекта СПДС (ведомости, штампы, ГОСТ 21.101,
#      общие данные, состав проекта). Тоже перед cross_section/presence: "Ведомость...
#      приложена" не должна попасть в presence по слову "приложена" раньше, чем распознана
#      как формальный СПДС-пункт.
#   3. cross_section   — сверка "соответствует"/"совпадает"/"согласовано" + ТЗ/ОПР/АГО/ГПЗУ/
#      СТУ/стандарту/вендор-листу/заданию. ПЕРЕД presence и parametric: на реальных данных
#      формулировки вида "Раздел СПОЗУ соответствует альбому АГО" или "Марка утеплителя
#      соответствует вендор-листу" содержат аббревиатуры-триггеры presence (АГО) или слово
#      "марка" (parametric), но по смыслу это сверка раздела/значения с другим документом,
#      не проверка присутствия файла и не сравнение с числовым порогом. НЕ включает "концепц"
#      сознательно: все критерии "соответствует концепции" в реальных данных (СПОЗУ 014-018)
#      — экспертная архитектурная оценка, остаются manual_required (задача прямо требует не
#      подгонять классификацию ради процента автоматизации).
#   4. parametric      — конкретная проверяемая величина: число+единица (мм/м²/кВт/°C/...),
#      сравнение (не менее/не более/не выше/не ниже + число), явные марки-коды (W12/B25/F150),
#      "сечение"/"диаметр"/"толщина" РЯДОМ с числом. Голое слово "марка"/"класс"/"сечение" без
#      числа намеренно НЕ триггерит parametric (реальные данные: "Проведён расчёт сечений
#      проводников по допустимому длительному току" не содержит значения — это calculation,
#      число появляется только в результате расчёта, не в тексте критерия; "Марка утеплителя
#      соответствует вендор-листу" — сверка без числа, см. cross_section выше).
#   5. presence (сильная форма) — короткие пункты-заголовки списка исходных данных: голая
#      аббревиатура источника (ТВО/ГПЗУ/ППТ/АГО/ОПР/СТУ/ТУ) как весь/почти весь текст пункта,
#      явное "выполнен и приложен"/"приложен отчёт". Идёт ДО calculation: такие пункты по
#      структуре ближе к элементу списка ("нужен ли этот документ"), даже если рядом
#      встречается слово "расчёт" внутри перечисления после запятой.
#   6. calculation     — наличие расчёта как артефакта ("расчет"/"расчёт"), когда пункт
#      начинается с "Выполнен(ы)/Проведён(ы)/Приведён(ы) расчёт..." и не более специфичен
#      (не spatial/spds/cross_section/parametric/сильный presence). Идёт ПЕРЕД слабым
#      presence: длинные формулировки калькуляций часто попутно упоминают "представлена
#      таблица"/"приложения нагрузок"/аббревиатуру СТУ/ТУ вскользь (не как отдельный пункт
#      "наличие документа"), и такое случайное соседство не должно перебивать основной смысл
#      предложения — расчёт как артефакт.
#   7. presence (слабая форма) — общие глагольные маркеры "приложен\w*"/"наличие"/
#      "представлен\w*" без явного "расчёт" в этом же пункте — fallback для документо-
#      ориентированных формулировок, не пойманных сильной формой (например "Приложены планы/
#      разрезы АР").
#   8. manual_required — умолчание: пункт требует инженерного решения, ничего из выше не
#      сработало. Экспертные критерии ("соответствует концепции", качество решений) ОБЯЗАНЫ
#      сюда попадать — это не пробел эвристики, а честная граница (implementation_plan.md §3.5).
_SPATIAL_VISUAL_EXCLUDE_PATTERN = re.compile(
    r"размещ\w*\s+в\s+состав\w*|размещ\w*\s+в\s+том\w*",
    re.IGNORECASE,
)

_SPATIAL_VISUAL_PATTERN = re.compile(
    r"размещ|располож|над жил|под жил|кухн|корзин\w*|"
    r"кондиционер|термовкладыш|на кровле|на фасад",
    re.IGNORECASE,
)

_SPDS_FORMAL_PATTERN = re.compile(
    r"ведомост\w*|штамп\w*|основн\w*\s+надпис\w*|гост[\s\-]*21\.?101|общие данные|"
    r"содержани\w*\s+том\w*|содержани\w*\s+комплект\w*|состав\s+проект\w*|состав\s+пд\b|"
    r"оформлен\w*|читаем\w*",
    re.IGNORECASE,
)

_CROSS_SECTION_ABBR = r"\b(?:тз|опр|аго|гпзу|сту)\b|заданию|стандарт\w*|вендор\w*"

_CROSS_SECTION_PATTERN = re.compile(
    rf"(?:соответств\w*|совпадает|согласован\w*\s+с(?:о)?\s).*(?:{_CROSS_SECTION_ABBR})|"
    rf"(?:{_CROSS_SECTION_ABBR}).*(?:соответств\w*|совпадает)",
    re.IGNORECASE,
)

_PARAMETRIC_UNIT_PATTERN = re.compile(
    r"\d+[.,]?\d*\s*(мм|см|м2|м²|м3|м³|кВт|квт|°с|градус|кг|тонн|мин|час|%)|"
    r"не\s+менее\s+\d|не\s+более\s+\d|не\s+выше\s+\d|не\s+ниже\s+\d|"
    r"[wbf]\d{1,3}\b|"
    r"(?:сечени\w*|диаметр\w*|толщин\w*)[^.]{0,40}\d|"
    r"\d[^.]{0,40}(?:сечени\w*|диаметр\w*|толщин\w*)",
    re.IGNORECASE,
)

# Сильная форма presence: голая аббревиатура-заголовок исходного документа (весь пункт —
# аббревиатура, максимум с уточнением в скобках/через дробь: "СТУ (в т.ч. список отклонений
# от НД)", "ПЗ ИРД (ТУ)", "АГО", "ПЗ ИРД (в части ГПЗУ) / ГПЗУ") либо явное "выполнен(ы) и
# приложен(ы)"/"приложен отчёт". НЕ включает голое "приложен\w*"/"наличие"/"представлен\w*" —
# они слишком легко матчят случайные словоформы ("приложения нагрузок", "представлена
# таблица") внутри calculation-предложений; та часть — слабая форма ниже, после calculation.
_PRESENCE_STRONG_PATTERN = re.compile(
    r"^[а-яё\s]{0,20}\b(?:тво|гпзу|ппт|аго|опр|сту|ту)\b[а-яё0-9\s.,/()«»%]{0,60}$|"
    r"выполнен\w*\s+и\s+приложен\w*|проведен\w*\s+и\s+приложен\w*|"
    r"приложен\w*\s+отчет\w*|приложен\w*\s+отчёт\w*|наличие",
    re.IGNORECASE,
)

# Общий fallback сверки: "соответствует требованиям <нормативный документ/пункт>" без явной
# aббревиатуры Glorax-семейства (СП, ГОСТ, СНиП, конкретный пункт) — например "соответствует
# требованиям СП 510, п. 7.16", "соответствуют проектным требованиям". Остаётся cross_section,
# а не calculation/manual_required — восстанавливает поведение исходной эвристики (T1.2/T1.3)
# для нормативных сверок без Glorax-аббревиатуры в самом тексте. Проверяется вместе с
# cross_section (после сильного presence, до calculation) — иначе "расчёт... соответствует
# требованиям СП" ошибочно перехватывался бы calculation раньше.
_CROSS_SECTION_FALLBACK_PATTERN = re.compile(
    r"соответств\w*.*требовани|требовани\w*.*соответств\w*",
    re.IGNORECASE,
)

_CALCULATION_PATTERN = re.compile(r"расчет\w*|расчёт\w*", re.IGNORECASE)

# Слабая форма presence: общие глагольные маркеры без расчёта в этом же пункте — fallback
# после calculation, чтобы не перехватывать случайные словоформы внутри калькуляционных
# формулировок (см. комментарий у _PRESENCE_STRONG_PATTERN).
_PRESENCE_WEAK_PATTERN = re.compile(
    r"приложен\w*|представлен\w*|отчет\w*|отчёт\w*",
    re.IGNORECASE,
)

_KIND_CLASSIFIERS: tuple[tuple[re.Pattern, str], ...] = (
    (_SPATIAL_VISUAL_PATTERN, "spatial_visual"),
    (_SPDS_FORMAL_PATTERN, "spds_formal"),
    (_CROSS_SECTION_PATTERN, "cross_section"),
    (_PARAMETRIC_UNIT_PATTERN, "parametric"),
    (_PRESENCE_STRONG_PATTERN, "presence"),
    (_CROSS_SECTION_FALLBACK_PATTERN, "cross_section"),
    (_CALCULATION_PATTERN, "calculation"),
    (_PRESENCE_WEAK_PATTERN, "presence"),
)


def _discipline_code(discipline: str) -> str:
    """Латинский код дисциплины для id. Есть явная таблица; фолбэк — генерик-транслит."""
    code = DISCIPLINE_TRANSLIT.get(discipline)
    if code is not None:
        return code
    return _generic_translit(discipline)


_TRANSLIT_MAP = {
    "а": "A", "б": "B", "в": "V", "г": "G", "д": "D", "е": "E", "ё": "E",
    "ж": "ZH", "з": "Z", "и": "I", "й": "I", "к": "K", "л": "L", "м": "M",
    "н": "N", "о": "O", "п": "P", "р": "R", "с": "S", "т": "T", "у": "U",
    "ф": "F", "х": "H", "ц": "TS", "ч": "CH", "ш": "SH", "щ": "SCH",
    "ъ": "", "ы": "Y", "ь": "", "э": "E", "ю": "YU", "я": "YA",
}


def _generic_translit(text: str) -> str:
    """Фолбэк-транслитерация для дисциплин РД (или новых, не заведённых в таблицу)."""
    # берём только буквенно-цифровые символы до первого разделителя/скобки
    letters = re.findall(r"[A-Za-zА-Яа-яЁё0-9]+", text)
    if not letters:
        return "XX"
    word = letters[0]
    out = []
    for ch in word:
        lower = ch.lower()
        if lower in _TRANSLIT_MAP:
            out.append(_TRANSLIT_MAP[lower])
        else:
            out.append(ch.upper())
    return "".join(out) or "XX"


def _is_total_row(b_value: str) -> bool:
    return any(p.search(b_value) for p in _TOTAL_ROW_PATTERNS)


def _has_block_header_fill(cell) -> bool:
    """True, если ячейка B закрашена фиолетовым `FF3200F0` (заголовок блока).

    Найдено на реальном ПД-XLSX (T1.3, лист ОВиК, строки 15/18 "Климатические и
    метеорологические условия" / "Параметры теплоносителя"): 2 строки во всём файле
    имеют одновременно заливку FF3200F0 (визуально — заголовок блока, как соседние
    "Наличие необходимых расчётов и приложений" на той же дисциплине) И DV-разметку на
    answer-ячейке C (артефакт исходника — по-видимому протянутое форматирование DV при
    создании шаблона в Excel, номер в A у обеих строк целый "1"/"2", как у настоящих
    заголовков, не "N.M" как у их собственных подпунктов). Оператор зафиксировал заливку
    как "вторичный диагностический признак", но на практике она разрешает именно такие
    конфликты: заливка заголовка блока приоритетнее случайно оставленной DV на C.
    """
    fill = cell.fill
    try:
        return bool(fill and fill.fgColor and fill.fgColor.rgb == BLOCK_HEADER_FILL)
    except Exception:
        return False


def _allowed_answers_from_formula(formula1: str) -> list[str]:
    """DataValidation.formula1 для list-типа выглядит как '"Да,Нет,Не требуется"'."""
    text = formula1.strip()
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1]
    return [part.strip() for part in text.split(",") if part.strip()]


def _build_dv_index(ws) -> dict[str, list[str]]:
    """cell coordinate ('C5') -> allowed_answers, развёрнуто из sqref-диапазонов."""
    index: dict[str, list[str]] = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list" or not dv.formula1:
            continue
        answers = _allowed_answers_from_formula(dv.formula1)
        for cell_range in dv.sqref.ranges:
            for row in ws.iter_rows(
                min_row=cell_range.min_row,
                max_row=cell_range.max_row,
                min_col=cell_range.min_col,
                max_col=cell_range.max_col,
            ):
                for cell in row:
                    index[cell.coordinate] = answers
    return index


def _classify_kind(criterion: str) -> str:
    """Классифицирует критерий по ключевым словам. Порядок правил — см. комментарий над
    `_KIND_CLASSIFIERS`: специфичные группы (spatial_visual, spds_formal) проверяются раньше
    общих (calculation, cross_section), иначе общие паттерны перехватили бы специфичные
    формулировки, содержащие слова "расчёт"/"соответствует" не по своей сути.
    """
    for pattern, kind in _KIND_CLASSIFIERS:
        if kind == "spatial_visual" and _SPATIAL_VISUAL_EXCLUDE_PATTERN.search(criterion):
            # "размещены в составе тома"/"в томе" — документ размещён в комплекте, не на
            # плане (PD-EOM-024); не геометрическое размещение, пропускаем spatial_visual.
            continue
        if pattern.search(criterion):
            return kind
    return "manual_required"


def _cell_text(val: Any) -> str:
    """Текст ячейки для классификации. Формулы (строки `=...` и объекты
    openpyxl ArrayFormula/DataTableFormula при data_only=False) — не текст
    заголовка/критерия: возвращаем пустоту, чтобы repr объекта не утекал
    в section_path (баг, найденный в T1.4: `<openpyxl...ArrayFormula object at ...>`)."""
    if val is None:
        return ""
    if isinstance(val, str):
        text = val.strip()
        return "" if text.startswith("=") else text
    if isinstance(val, (int, float)):
        return str(val)
    return ""


def _import_sheet(ws, *, stage: str, discipline: str) -> list[dict[str, Any]]:
    dv_index = _build_dv_index(ws)
    disc_code = _discipline_code(discipline)

    items: list[dict[str, Any]] = []
    section_path: list[str] = []
    section_header: str | None = None
    block_header: str | None = None

    max_row = ws.max_row
    for row_idx in range(_HEADER_ROWS + 1, max_row + 1):
        a_val = ws.cell(row=row_idx, column=1).value
        b_cell = ws.cell(row=row_idx, column=2)
        b_val = b_cell.value

        a_text = _cell_text(a_val)
        b_text = _cell_text(b_val)

        if not b_text:
            if a_text:
                # заголовок раздела верхнего уровня: A непустая, B пустая
                section_header = a_text
                block_header = None
                section_path = [section_header]
            continue

        if _is_total_row(b_text):
            continue

        answer_coord = f"C{row_idx}"
        allowed_answers = dv_index.get(answer_coord)

        if allowed_answers is None or _has_block_header_fill(b_cell):
            # B непустая, И (DV на C нет ИЛИ заливка FF3200F0 есть) -> заголовок блока.
            # Заливка — приоритетный признак: перекрывает DV, если оба присутствуют
            # одновременно (см. _has_block_header_fill).
            block_header = b_text
            section_path = ([section_header] if section_header else []) + [block_header]
            continue

        # критерий: B непустая И DV на C есть
        item_no = str(row_idx - _HEADER_ROWS)
        item_id = f"{stage}-{disc_code}-{row_idx:03d}"
        items.append(
            {
                "id": item_id,
                "stage": stage,
                "discipline": discipline,
                "sheet_name": ws.title,
                "row": row_idx,
                "item_no": item_no,
                "section_path": list(section_path),
                "criterion": b_text,
                "answer_cell": answer_coord,
                "note_cell": f"D{row_idx}",
                "allowed_answers": allowed_answers,
                "kind": _classify_kind(b_text),
            }
        )

    return items


def _apply_overrides(items: list[dict[str, Any]], overrides: dict[str, dict[str, Any]]) -> None:
    """Применяет ручные правки ПОСЛЕ эвристики `_classify_kind` (in-place).

    `overrides`: `{item_id: {"kind": "...", "comment": "..."}}`. Формат — плоский словарь по
    id, не список, чтобы правка была устойчива к перестановке items и не зависела от индекса.
    Не найденный в template item_id из overrides молча игнорируется (например, если
    overrides-файл писался под другую версию XLSX) — не считается ошибкой импорта.
    """
    if not overrides:
        return
    by_id = {item["id"]: item for item in items}
    for item_id, patch in overrides.items():
        item = by_id.get(item_id)
        if item is None:
            continue
        if "kind" in patch and patch["kind"]:
            item["kind"] = patch["kind"]
        if "comment" in patch and patch["comment"]:
            item["comment"] = patch["comment"]


def import_checklist_template(
    xlsx_path: str | Path,
    stage: str,
    *,
    overrides: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Импортирует чек-лист Glorax (XLSX) в ChecklistTemplate (канон §5).

    Читает только содержательные листы (исключая SKIP_SHEETS). Каждый лист = дисциплина.
    Открывает книгу read-only, data_only=False (нужны формулы DV, не вычисленные значения).

    `overrides` (T1.4) — ручные правки классификации `kind` (и опционально `comment`) по
    `item_id`, применяются ПОСЛЕ эвристики `_classify_kind`. Источник — JSON-файл вида
    `config/checklists/glorax_pd_2026_overrides.json`, читаемый вызывающим кодом (CLI
    `tools/build_checklist_template.py --overrides PATH`) и передаваемый сюда уже как dict —
    сам importer не знает о путях к overrides-файлам.
    """
    path = Path(xlsx_path)
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        disciplines: list[str] = []
        items: list[dict[str, Any]] = []

        for ws in wb.worksheets:
            if ws.title in SKIP_SHEETS:
                continue
            disciplines.append(ws.title)
            items.extend(_import_sheet(ws, stage=stage, discipline=ws.title))

        if overrides:
            _apply_overrides(items, overrides)

        template: dict[str, Any] = {
            "name": f"{path.stem}",
            "stage": stage,
            "title": path.stem,
            "source_file_name": path.name,
            "version": "",
            "disciplines": disciplines,
            "items": items,
        }
        return template
    finally:
        wb.close()

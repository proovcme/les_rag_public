"""КАЦ — конъюнктурный анализ цен: ≥3 КП на материал → выбор экономичного → цена в ЛСР.

Регламент (Методика 421/пр, см. онтологию [[kac]]): для материалов, которых НЕТ в ФГИС ЦС,
собирают коммерческие предложения от ≥3 поставщиков на каждый материал и берут наиболее
экономичный вариант со ссылкой на источник; цена идёт в позицию ЛСР как неучтённый материал.

Этот сервис — ДЕТЕРМИНИРОВАННОЕ ядро (0 LLM): группировка котировок по материалу, проверка
достаточности (≥3), выбор по стратегии, разброс цен, выгрузка КАЦ-таблицы и линий для ЛСР.
Извлечение котировок из PDF-КП — отдельный шаг (vision/OCR-пайплайн), сюда приходят строки.
Мост к ценовой базе: needs_kac(code) — нужен ли КАЦ (нет ли кода в ФГИС ЦС).
"""

from __future__ import annotations

import re
import statistics
from typing import Any, Optional

DEFAULT_MIN_SUPPLIERS = 3


def _norm_material(name: Any) -> str:
    return re.sub(r"\s+", " ", str(name or "").strip().lower().replace("ё", "е")).strip(" .,;:")


def _price(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    try:
        return float(text)
    except ValueError:
        return None


def analyze_kac(
    quotes: list[dict[str, Any]],
    *,
    min_suppliers: int = DEFAULT_MIN_SUPPLIERS,
    strategy: str = "min",
) -> dict[str, Any]:
    """Котировки поставщиков → КАЦ по материалам.

    quote = {material, supplier, unit, price, source?}. strategy: 'min' (экономичный) | 'median'.
    Возвращает {materials:[...], summary:{...}, strategy, min_suppliers}.
    """
    groups: dict[str, list[dict[str, Any]]] = {}
    order: list[str] = []
    for q in quotes or []:
        key = _norm_material(q.get("material"))
        if not key:
            continue
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append({
            "supplier": str(q.get("supplier") or "").strip(),
            "unit": str(q.get("unit") or "").strip(),
            "price": _price(q.get("price")),
            "source": str(q.get("source") or "").strip(),
            "material": str(q.get("material") or "").strip(),
        })

    materials: list[dict[str, Any]] = []
    sufficient = 0
    for key in order:
        offers = [o for o in groups[key] if o["price"] is not None]
        offers.sort(key=lambda o: o["price"])
        prices = [o["price"] for o in offers]
        n = len(offers)
        units = {o["unit"] for o in offers if o["unit"]}
        chosen = None
        if prices:
            if strategy == "median":
                target = statistics.median(prices)
                chosen = min(offers, key=lambda o: abs(o["price"] - target))
            else:  # 'min' — наиболее экономичный
                chosen = offers[0]
        is_ok = n >= min_suppliers
        if is_ok:
            sufficient += 1
        spread = (max(prices) - min(prices)) if prices else 0.0
        materials.append({
            "material": groups[key][0]["material"],
            "unit": (sorted(units)[0] if units else ""),
            "offers": offers,
            "suppliers": n,
            "sufficient": is_ok,
            "unit_mismatch": len(units) > 1,
            "chosen_price": (chosen["price"] if chosen else None),
            "chosen_supplier": (chosen["supplier"] if chosen else ""),
            "chosen_source": (chosen["source"] if chosen else ""),
            "min_price": (min(prices) if prices else None),
            "max_price": (max(prices) if prices else None),
            "spread": round(spread, 2),
            "spread_pct": (round(spread / min(prices) * 100, 1) if prices and min(prices) else 0.0),
        })

    return {
        "strategy": strategy,
        "min_suppliers": min_suppliers,
        "materials": materials,
        "summary": {
            "materials": len(materials),
            "sufficient": sufficient,
            "insufficient": len(materials) - sufficient,
            "total_quotes": sum(m["suppliers"] for m in materials),
        },
    }


def kac_to_lsr_lines(result: dict[str, Any]) -> list[dict[str, Any]]:
    """Выбранные цены КАЦ → линии для позиций ЛСР (неучтённый материал, графа «Н»)."""
    lines: list[dict[str, Any]] = []
    for m in result.get("materials", []):
        if m.get("chosen_price") is None:
            continue
        lines.append({
            "name": m["material"],
            "unit": m["unit"],
            "price": m["chosen_price"],
            "source": m["chosen_source"],
            "supplier": m["chosen_supplier"],
            "basis": f"КАЦ (≥{result.get('min_suppliers', DEFAULT_MIN_SUPPLIERS)} КП)",
            "sufficient": m["sufficient"],
        })
    return lines


def to_xlsx(result: dict[str, Any], out_path: str) -> str:
    """Выгрузка КАЦ-таблицы: по материалам строки поставщиков, выбранный помечен ✓."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "КАЦ"
    bold = Font(bold=True)
    pick = PatternFill("solid", fgColor="DDF4DD")

    ws.append(["Конъюнктурный анализ цен (КАЦ)"])
    ws["A1"].font = bold
    headers = ["№", "Материал", "Ед.", "Поставщик", "Цена", "Выбрано", "Разброс, %", "Источник"]
    ws.append(headers)
    for c in ws[2]:
        c.font = bold

    n = 0
    for m in result.get("materials", []):
        n += 1
        first = True
        for o in m["offers"]:
            is_pick = (o["price"] == m["chosen_price"] and o["supplier"] == m["chosen_supplier"])
            ws.append([
                n if first else "",
                m["material"] if first else "",
                o["unit"],
                o["supplier"],
                o["price"],
                "✓" if is_pick else "",
                m["spread_pct"] if first else "",
                o["source"],
            ])
            if is_pick:
                for c in ws[ws.max_row]:
                    c.fill = pick
            first = False
        if not m["sufficient"]:
            ws.append(["", f"⚠ поставщиков {m['suppliers']} < {result['min_suppliers']} — недостаточно для КАЦ"])

    wb.save(out_path)
    return out_path


def needs_kac(code: str, *, book: str | None = None) -> dict[str, Any]:
    """Мост к ФГИС ЦС: нужен ли КАЦ для ресурса — т.е. отсутствует ли код в ценовой базе.

    Регламент: КАЦ составляется для материалов, которых НЕТ в ФГИС ЦС.
    """
    from pathlib import Path

    from proxy.services import fgis_price_service as fps

    books = fps.available_pricebooks()
    if not books:
        return {"code": code, "in_fgis": False, "needs_kac": True, "note": "нет книги цен — проверка невозможна"}
    path = next((p for p in books if Path(p).stem == book), None) if book else books[0]
    if path is None:
        return {"code": code, "in_fgis": False, "needs_kac": True, "note": f"книга {book!r} не найдена"}
    rec = fps.get_pricebook(path).lookup(code)
    in_fgis = rec is not None
    return {
        "code": code,
        "book": Path(path).stem,
        "in_fgis": in_fgis,
        "needs_kac": not in_fgis,
        "fgis_price": (rec.get("price_current_eff") if rec else None),
    }

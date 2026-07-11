"""Build user-facing smeta artifacts from model-written estimate answers.

The service is intentionally shallow: it does not choose works, norms, rates or
quantities.  It only extracts Markdown tables already produced by the estimator
model, computes visible table totals, and prepares a separate artifact payload.
"""

from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from proxy.services.estimate_math_service import parse_ru_number


@dataclass(frozen=True)
class SmetaTable:
    title: str
    kind: str
    headers: list[str]
    rows: list[list[str]]
    markdown: str
    amount_total: float | None = None


_SEPARATOR_RE = re.compile(r"^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*$")
_MONEY_HEADER_TOKENS = ("сумм", "стоим", "итого", "руб")
_LSR_FORM_HEADERS = {
    1: "№ п/п",
    2: "Обоснование",
    3: "Наименование работ и затрат",
    4: "Ед. изм.",
    5: "Кол-во на ед.",
    6: "коэф.",
    7: "Кол-во всего",
    8: "Сметная стоимость в базисном уровне цен на ед., руб.",
    9: "Индекс",
    10: "Сметная стоимость в текущем уровне цен на ед., руб.",
    11: "коэф.",
    12: "Сметная стоимость в текущем уровне цен всего, руб.",
}
_LSR_FORM_MAX_COL = 12


def _split_cells(line: str) -> list[str]:
    text = str(line or "").strip()
    if text.startswith("|"):
        text = text[1:]
    if text.endswith("|"):
        text = text[:-1]
    return [cell.strip() for cell in text.split("|")]


def _is_table_header(line: str, next_line: str) -> bool:
    return "|" in str(line or "") and bool(_SEPARATOR_RE.match(str(next_line or "")))


def _clean_title(line: str) -> str:
    text = str(line or "").strip()
    text = re.sub(r"^#{1,6}\s*", "", text)
    text = text.strip("*").strip()
    return text


def _nearest_title(lines: list[str], start: int) -> str:
    for idx in range(start - 1, max(-1, start - 8), -1):
        text = lines[idx].strip()
        if not text:
            continue
        if text.startswith("|"):
            continue
        if text.startswith("#") or text.startswith("**"):
            return _clean_title(text)
        if len(text) <= 90:
            return _clean_title(text)
    return "Сметная таблица"


def _classify_table(headers: list[str], title: str) -> str:
    joined = " ".join([title, *headers]).casefold()
    if "поставка" in joined or "материал" in joined and "работ" not in joined:
        return "supply"
    if "развилк" in joined or "вариант" in joined and "объ" in joined:
        return "quantity_conflict"
    if ("рим" in joined and "рын" in joined) or "гэсн" in joined and "рын" in joined:
        return "method_comparison"
    if any(token in joined for token in ("сумм", "стоим", "ставк", "руб")):
        return "work_cost"
    if any(token in joined for token in ("работ", "вор", "ед.", "кол-во", "количество")):
        return "bor"
    return "table"


def _amount_column_indexes(headers: list[str]) -> list[int]:
    out: list[int] = []
    for idx, header in enumerate(headers):
        low = header.casefold()
        if "ставк" in low or "цена" in low and "итого" not in low:
            continue
        if any(token in low for token in _MONEY_HEADER_TOKENS):
            out.append(idx)
    return out


def _parse_money_cell(cell: str) -> float | None:
    text = str(cell or "")
    if not re.search(r"\d", text):
        return None
    cleaned = text.replace("**", "").replace("\xa0", " ").replace("\u202f", " ")
    match = re.search(r"[-+]?\d[\d\s,.\u00a0\u202f]*", cleaned)
    if not match:
        return None
    number = match.group(0).strip()
    compact = number.replace(" ", "").replace("\xa0", "").replace("\u202f", "")
    if "," in compact and "." in compact:
        comma = compact.rfind(",")
        dot = compact.rfind(".")
        if comma < dot and len(compact) - dot - 1 in {1, 2}:
            compact = compact.replace(",", "")
        else:
            compact = compact.replace(".", "").replace(",", ".")
        try:
            return float(compact)
        except ValueError:
            return None
    value = parse_ru_number(number)
    if value is None:
        return None
    return float(value)


def _table_amount_total(headers: list[str], rows: list[list[str]]) -> float | None:
    indexes = _amount_column_indexes(headers)
    if not indexes:
        return None
    total = 0.0
    found = False
    for row in rows:
        first = row[0].casefold() if row else ""
        if "итого" in first or "всего" in first:
            continue
        row_value: float | None = None
        for idx in indexes:
            if idx < len(row):
                parsed = _parse_money_cell(row[idx])
                if parsed is not None:
                    row_value = parsed
        if row_value is not None:
            total += row_value
            found = True
    return total if found else None


def _find_header_index(headers: list[str], *tokens: str) -> int | None:
    for idx, header in enumerate(headers):
        low = str(header or "").casefold()
        if all(token in low for token in tokens):
            return idx
    return None


def _find_header_index_any(headers: list[str], *token_groups: tuple[str, ...]) -> int | None:
    for group in token_groups:
        found = _find_header_index(headers, *group)
        if found is not None:
            return found
    return None


def _first_index(*values: int | None) -> int | None:
    for value in values:
        if value is not None:
            return value
    return None


def _cell(row: list[str], idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx] or "").strip()


def _looks_total_row(row: list[str]) -> bool:
    joined = " ".join(str(cell or "") for cell in (row or [])).casefold()
    return "итого" in joined or "всего по смете" in joined


def _normative_basis(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    match = re.search(
        r"(ГЭСН(?:м|мр|п|р)?\s*:?\s*\d{1,2}[-–]\d{2}[-–]\d{3}[-–]\d{2}|"
        r"ГЭСН(?:м|мр|п|р)?\s*\d{1,2}|"
        r"ФЕР(?:м|мр|п|р)?\s*\d{1,2}[-–]\d{2}[-–]\d{3}[-–]\d{2})",
        text,
        flags=re.IGNORECASE,
    )
    return match.group(1).replace("–", "-") if match else text


def _stable_lsr_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _lsr_rows_from_table(table: SmetaTable) -> list[dict[str, Any]]:
    headers = table.headers
    joined_headers = " ".join(str(header or "").casefold() for header in headers)
    is_lsr_12_graph = "кол-во всего" in joined_headers and "текущий всего" in joined_headers
    work_idx = _first_index(
        _find_header_index(headers, "работ"),
        _find_header_index(headers, "наименование"),
        _find_header_index(headers, "раздел"),
    )
    qty_idx = (
        _find_header_index(headers, "кол-во", "всего")
        if is_lsr_12_graph
        else _first_index(_find_header_index(headers, "кол"), _find_header_index(headers, "объ"))
    )
    unit_idx = _find_header_index(headers, "ед")
    basis_idx = _first_index(
        _find_header_index(headers, "обоснование"),
        _find_header_index(headers, "норма"),
        _find_header_index(headers, "гэсн"),
        _find_header_index(headers, "источник"),
        _find_header_index(headers, "рим"),
    )
    price_idx = (
        _find_header_index_any(headers, ("текущий", "на ед"), ("базис", "на ед"))
        if is_lsr_12_graph
        else _first_index(_find_header_index(headers, "ставк"), _find_header_index(headers, "цена"))
    )
    amount_indexes = _amount_column_indexes(headers)
    status_idx = _first_index(
        _find_header_index(headers, "статус"),
        _find_header_index(headers, "коммент"),
        _find_header_index(headers, "источник"),
    )
    out: list[dict[str, Any]] = []
    for row in table.rows:
        if not row or _looks_total_row(row):
            continue
        title = _cell(row, work_idx)
        if not title and len(row) > 1:
            title = _cell(row, 1)
        amount: float | None = None
        for idx in amount_indexes:
            if idx < len(row):
                parsed = _parse_money_cell(row[idx])
                if parsed is not None:
                    amount = parsed
        if not title and amount is None:
            continue
        out.append(
            {
                "basis": _normative_basis(_cell(row, basis_idx)),
                "title": _stable_lsr_text(title) or "Позиция сметы",
                "unit": _cell(row, unit_idx),
                "quantity": _cell(row, qty_idx),
                "unit_price": _cell(row, price_idx),
                "amount": amount,
                "status": _cell(row, status_idx) or table.title,
                "source_table": table.title,
            }
        )
    return out


def _select_lsr_source_tables(tables: list[SmetaTable]) -> list[SmetaTable]:
    """Pick the visible table(s) that should feed the display LSR form.

    The model may output both a detailed work-cost table and a shorter
    "preliminary LSR" table. Those are alternate presentations of the same
    estimate, not additive sections. Keep the artifact renderer shallow: select
    the most complete visible cost table and do not merge duplicate forms.
    """
    candidates = [table for table in tables if table.kind in {"work_cost", "method_comparison"}]
    if not candidates:
        return []
    explicit_lsr = [table for table in candidates if "лср" in table.title.casefold()]
    non_lsr = [table for table in candidates if table not in explicit_lsr]
    if explicit_lsr and not non_lsr:
        return [max(explicit_lsr, key=lambda table: (len(table.rows), table.amount_total or 0.0))]
    if explicit_lsr and non_lsr:
        best_non_lsr = max(non_lsr, key=lambda table: (len(table.rows), table.amount_total or 0.0))
        best_lsr = max(explicit_lsr, key=lambda table: (len(table.rows), table.amount_total or 0.0))
        if len(best_lsr.rows) >= len(best_non_lsr.rows):
            return [best_lsr]
        return [best_non_lsr]
    return [max(candidates, key=lambda table: (len(table.rows), table.amount_total or 0.0))]


def build_lsr_form(tables: list[SmetaTable]) -> dict[str, Any] | None:
    """Build an LSR-shaped output view from model-written cost tables.

    This is a display form only: no works, norms, quantities or prices are
    invented here. Rows are copied from visible estimate tables.
    """
    rows: list[dict[str, Any]] = []
    source_tables = _select_lsr_source_tables(tables)
    for table in source_tables:
        rows.extend(_lsr_rows_from_table(table))
    if not rows:
        return None
    total = sum(float(row["amount"] or 0.0) for row in rows if row.get("amount") is not None)
    status_blob = " ".join(str(row.get("status") or "") for row in rows).casefold()
    scenario_like = any(token in status_blob for token in ("scenario", "сценар", "допущ"))
    return {
        "schema": "lsr_rim_display_form_v1",
        "title": "Локальный сметный расчет (смета)",
        "note": (
            "Предварительная ЛСР-форма по видимым строкам сметного ответа. "
            "Код не добавляет позиции, не выбирает нормы и не придумывает ставки. "
            "Без расчетной трассы РИМ/ресурсов/цен это не финальная ЛСР по Приложению №3."
        ),
        "finality": "scenario_display" if scenario_like else "display_form",
        "is_priced_final": False,
        "headers": [str(_LSR_FORM_HEADERS.get(col, "")) for col in range(1, _LSR_FORM_MAX_COL + 1)],
        "rows": rows,
        "source_tables": [table.title for table in source_tables],
        "amount_total": total if total > 0 else None,
    }


def _select_pricebook_for_question(question: str):
    """Use the operator manifest default; code does not infer region from prose."""
    del question
    try:
        from proxy.services import fgis_price_service as fps

        path = fps.resolve_pricebook_path(None)
        if path:
            return fps.get_pricebook(path), Path(path).stem
    except Exception:
        return None, ""
    return None, ""


def _build_rim_trace_form(lsr_form: dict[str, Any], *, question: str = "") -> dict[str, Any] | None:
    """Build a checked RIM form from visible rows that already contain norm codes.

    The model owns the visible row and selected norm code. This helper only
    converts units, expands resources and prices them with the local pricebook.
    """
    rows = lsr_form.get("rows") if isinstance(lsr_form, dict) else None
    if not rows:
        return None
    try:
        from proxy.smeta_core.workflow import calculate_visible_rows_revision
    except Exception:
        return None

    _pricebook, pricebook_name = _select_pricebook_for_question(question)
    trace = calculate_visible_rows_revision(
        list(rows),
        selected_by="model",
        created_by="model",
        change_note="smeta artifact from visible model rows",
        book=pricebook_name or None,
        title=str(lsr_form.get("title") or "Локальный сметный расчет (смета)"),
    )
    summary = trace.get("summary") if isinstance(trace, dict) else {}
    bound_rows = int(summary.get("bound_rows") or 0)
    if bound_rows <= 0:
        return None
    input_rows = int(summary.get("input_rows") or len(rows))
    trace_positions: list[dict[str, Any]] = []
    trace_positions_by_source_row: dict[int, dict[str, Any]] = {}
    for section in trace.get("sections") or []:
        section_name = str(section.get("section") or "").strip()
        for pos in section.get("positions") or []:
            item = {"section": section_name, "position": pos}
            trace_positions.append(item)
            try:
                source_row_idx = int(pos.get("source_row") or 0)
            except (TypeError, ValueError):
                source_row_idx = 0
            if source_row_idx > 0:
                trace_positions_by_source_row[source_row_idx] = item
    bound_positions = iter(trace_positions)
    out_rows: list[dict[str, Any]] = []
    for idx, source_row in enumerate(rows, 1):
        binding = next(
            (item for item in (trace.get("row_bindings") or []) if int(item.get("row") or 0) == idx),
            {},
        )
        if binding.get("status") == "bound":
            trace_item = trace_positions_by_source_row.get(idx)
            if not trace_item:
                try:
                    trace_item = next(bound_positions)
                except StopIteration:
                    trace_item = {"section": str(source_row.get("section") or ""), "position": {}}
            pos = trace_item.get("position") or {}
            ps = pos.get("summary") or {}
            qty = pos.get("qty")
            amount = float(ps.get("total") or 0.0)
            unit_price = amount / float(qty) if qty not in (None, "", 0) and float(qty) else None
            flags = ps.get("flags") or []
            status = str(summary.get("result_status") or "")
            if flags and status == "priced_final":
                status = "priced_partial"
            out_rows.append(
                {
                    "basis": pos.get("code") or source_row.get("basis") or "",
                    "title": _stable_lsr_text(pos.get("name") or source_row.get("title") or ""),
                    "unit": pos.get("unit") or source_row.get("unit") or "",
                    "quantity": qty,
                    "unit_price": round(unit_price, 2) if unit_price is not None else "",
                    "amount": amount,
                    "status": status,
                    "source_table": "РИМ trace",
                    "section": trace_item.get("section") or source_row.get("section") or "",
                    "flags": "; ".join(str(flag) for flag in flags),
                }
            )
            continue
        out_rows.append(
            {
                "basis": "нужен подбор нормы",
                "title": _stable_lsr_text(source_row.get("title") or source_row.get("name")) or "Позиция сметы",
                "unit": source_row.get("unit") or "",
                "quantity": source_row.get("quantity") or "",
                "unit_price": "0.00",
                "amount": 0.0,
                "status": binding.get("status") or "norm_selection_required",
                "source_table": source_row.get("source_table") or "видимая ЛСР модели",
                "section": source_row.get("section") or "",
                "flags": binding.get("message") or "строка не вошла в проверяемую РИМ-сумму",
            }
        )
    if not out_rows:
        return None
    total = float(summary.get("total") or sum(float(row.get("amount") or 0.0) for row in out_rows))
    result_status = str(summary.get("result_status") or "priced_partial")
    flags = [str(flag) for flag in (summary.get("flags") or []) if str(flag).strip()]
    note = (
        "ЛСР-форма построена по расчётной РИМ-трассе из видимых строк с выбранными шифрами норм: "
        "код раскрыл ресурсы, цены/индексы, НР/СП и итог. Строки без выбранного шифра нормы "
        "не попали в проверяемую сумму и перечислены в доборе."
    )
    if pricebook_name:
        note += f" Книга цен: {pricebook_name}."
    if flags:
        note += " Добор: " + "; ".join(flags[:8])
    return {
        "schema": "lsr_rim_trace_form_v1",
        "title": str(lsr_form.get("title") or "Локальный сметный расчет (смета)"),
        "note": note,
        "finality": result_status,
        "is_priced_final": result_status == "priced_final",
        "headers": [str(_LSR_FORM_HEADERS.get(col, "")) for col in range(1, _LSR_FORM_MAX_COL + 1)],
        "rows": out_rows,
        "source_tables": ["РИМ trace"],
        "amount_total": total,
        "trace": trace,
        "pricebook": pricebook_name,
    }


def build_checked_rim_form_from_visible_rows(rows: list[dict[str, Any]], *, question: str = "") -> dict[str, Any] | None:
    """Build checked RIM LSR from rows where the model already selected norm codes."""
    if not rows:
        return None
    lsr_form = {
        "schema": "lsr_rim_display_form_v1",
        "title": "Локальный сметный расчет (смета)",
        "rows": rows,
        "source_tables": ["structured model norm choice"],
    }
    return _build_rim_trace_form(lsr_form, question=question)


def build_smeta_artifact_from_rim_form(
    rim_form: dict[str, Any],
    *,
    question: str = "",
    title: str = "Сметный артефакт",
) -> dict[str, Any] | None:
    """Create an artifact around an already checked RIM form."""
    if not rim_form:
        return None
    rows = rim_form.get("rows") or []
    table_headers = [str(_LSR_FORM_HEADERS.get(col, "")) for col in range(1, _LSR_FORM_MAX_COL + 1)]
    table_data = []
    for idx, row in enumerate(rows, 1):
        table_data.append([
            str(idx),
            str(row.get("basis") or ""),
            str(row.get("title") or ""),
            str(row.get("unit") or ""),
            "1",
            "1",
            str(row.get("quantity") or ""),
            "",
            "",
            str(row.get("unit_price") or ""),
            "1",
            str(row.get("amount") or "0.00"),
        ])
    lines = ["# Сметный артефакт", ""]
    if question:
        lines += ["## Запрос", str(question).strip(), ""]
    lines += [*_lsr_markdown(rim_form)]
    total = rim_form.get("amount_total")
    if total is not None and float(total or 0.0) > 0:
        lines += ["", "## Арифметика", f"Сумма проверенной РИМ-формы: **{_fmt_money(float(total))}**."]
    return {
        "mode": "markdown",
        "title": title,
        "content": "\n".join(lines).strip(),
        "tables": [
            {
                "title": "Structured model norm choice",
                "kind": "norm_choice",
                "rows": len(table_data),
                "amount_total": rim_form.get("amount_total"),
                "headers": table_headers,
                "data": table_data,
            }
        ],
        "lsr_form": rim_form,
        "rim_lsr_form": rim_form,
        "rim_trace": rim_form.get("trace"),
    }


def _lsr_markdown(lsr_form: dict[str, Any] | None) -> list[str]:
    if not lsr_form:
        return []
    lines = [
        "## ЛСР РИМ (форма 421/пр)",
        "",
        "Приложение № 3 к Методике (приказ Минстроя России от 04.08.2020 № 421/пр)",
        "",
        "**ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ (СМЕТА) № ____**",
        "",
        str(lsr_form.get("title") or "Локальный сметный расчет (смета)"),
        "",
        "Составлен ресурсным методом / сценарной формой выдачи по видимым строкам ответа.",
        "",
        str(lsr_form.get("note") or ""),
        "",
    ]
    if lsr_form.get("amount_total") is not None:
        lines += [
            f"Сметная стоимость: **{_fmt_money(float(lsr_form['amount_total']))}**",
            "",
        ]
    lines += [
        "| № п/п | Обоснование | Наименование работ и затрат | Ед. изм. | Кол-во на ед. | коэф. | Кол-во всего | Сметная стоимость в базисном уровне цен на ед., руб. | Индекс | Сметная стоимость в текущем уровне цен на ед., руб. | коэф. | Сметная стоимость в текущем уровне цен всего, руб. |",
        "|---:|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
        "| 1 | 2 | 3 | 4 | 5 | 6 | 7 | 8 | 9 | 10 | 11 | 12 |",
    ]
    for idx, row in enumerate(lsr_form.get("rows") or [], 1):
        amount = _fmt_money(float(row["amount"])) if row.get("amount") is not None else ""
        cells = [
            str(idx),
            str(row.get("basis") or ""),
            str(row.get("title") or ""),
            str(row.get("unit") or ""),
            str(row.get("quantity") or ""),
            "1",
            str(row.get("quantity") or ""),
            "",
            "",
            str(row.get("unit_price") or ""),
            "1",
            amount,
        ]
        lines.append("| " + " | ".join(cell.replace("|", "/") for cell in cells) + " |")
    if lsr_form.get("amount_total") is not None:
        lines.append(
            f"|  |  | **ВСЕГО по смете** |  |  |  |  |  |  |  |  | **{_fmt_money(float(lsr_form['amount_total']))}** |"
        )
    source_rows = [
        row for row in (lsr_form.get("rows") or [])
        if str(row.get("status") or row.get("source_table") or "").strip()
    ]
    if source_rows:
        lines += [
            "",
            "**Источники и статус строк**",
            "",
            "| № п/п | Источник / статус | Исходная таблица |",
            "|---:|---|---|",
        ]
        for idx, row in enumerate(source_rows, 1):
            lines.append(
                "| "
                + " | ".join(
                    str(cell or "").replace("|", "/")
                    for cell in (idx, row.get("status") or "", row.get("source_table") or "")
                )
                + " |"
            )
    return lines


def _fmt_money(value: float) -> str:
    return f"{value:,.0f}".replace(",", " ") + " руб."


def _drop_conflicting_manual_totals(text: str, lsr_total: float | None) -> str:
    """Remove prose totals that contradict the selected LSR row sum.

    The artifact renderer is not allowed to choose works or prices, but it can
    prevent two different "Итого" numbers from being shown side by side. The
    selected LSR form is derived from visible rows; a conflicting prose total
    without its own trace is display noise.
    """
    if lsr_total is None:
        return text
    result: list[str] = []
    total_tokens = ("итого", "всего", "стоимость работ", "сметная стоимость", "сумма")
    for line in str(text or "").splitlines():
        low = line.casefold()
        if low.lstrip().startswith(("лср-форма вынесена", "таблица вынесена")):
            result.append(line)
            continue
        if "руб" not in low or not any(token in low for token in total_tokens):
            result.append(line)
            continue
        parsed = _parse_money_cell(line)
        if parsed is None or abs(float(parsed) - float(lsr_total)) <= 0.5:
            result.append(line)
            continue
        if line.lstrip().startswith("|"):
            result.append(line)
            continue
        continue
    return "\n".join(result)


def _safe_sheet_title(title: str, used: set[str]) -> str:
    text = re.sub(r"[\[\]:*?/\\]", " ", str(title or "Таблица")).strip() or "Таблица"
    text = re.sub(r"\s+", " ", text)[:31]
    base = text or "Таблица"
    candidate = base
    idx = 2
    while candidate in used:
        suffix = f" {idx}"
        candidate = (base[: 31 - len(suffix)] + suffix).strip()
        idx += 1
    used.add(candidate)
    return candidate


def extract_smeta_tables(answer: str) -> list[SmetaTable]:
    """Extract Markdown tables from a smeta answer without interpreting domain logic."""
    lines = str(answer or "").splitlines()
    tables: list[SmetaTable] = []
    idx = 0
    while idx < len(lines) - 1:
        if not _is_table_header(lines[idx], lines[idx + 1]):
            idx += 1
            continue
        start = idx
        idx += 2
        while idx < len(lines) and "|" in lines[idx]:
            idx += 1
        raw = lines[start:idx]
        headers = _split_cells(raw[0])
        rows = [_split_cells(line) for line in raw[2:] if line.strip()]
        title = _nearest_title(lines, start)
        kind = _classify_table(headers, title)
        tables.append(SmetaTable(
            title=title,
            kind=kind,
            headers=headers,
            rows=rows,
            markdown="\n".join(raw),
            amount_total=_table_amount_total(headers, rows),
        ))
    return tables


def build_smeta_artifact(answer: str, *, question: str = "") -> dict[str, Any] | None:
    """Return a Markdown artifact payload for visible smeta tables, if any."""
    tables = extract_smeta_tables(answer)
    if not tables:
        return None
    model_lsr_form = build_lsr_form(tables)
    trace_lsr_form = _build_rim_trace_form(model_lsr_form, question=question) if model_lsr_form else None
    lsr_form = trace_lsr_form or model_lsr_form
    lines = ["# Сметный артефакт", ""]
    if question:
        lines += ["## Запрос", str(question).strip(), ""]
    lines += [
        "## Свод таблиц",
        "| Таблица | Тип | Строк | Сумма по видимым строкам |",
        "|---|---|---:|---:|",
    ]
    for table in tables:
        amount = _fmt_money(table.amount_total) if table.amount_total is not None else ""
        lines.append(f"| {table.title} | {table.kind} | {len(table.rows)} | {amount} |")
    if lsr_form:
        lines += ["", *_lsr_markdown(lsr_form)]
    for num, table in enumerate(tables, start=1):
        lines += ["", f"## {num}. {table.title}", "", table.markdown]
    total = lsr_form.get("amount_total") if lsr_form else None
    if total is not None and total > 0:
        source_tables = ", ".join(str(x) for x in (lsr_form.get("source_tables") or []) if x) if lsr_form else ""
        source_note = f" Источник ЛСР: {source_tables}." if source_tables else ""
        lines += ["", "## Арифметика", f"Сумма выбранной ЛСР-формы: **{_fmt_money(total)}**.{source_note}"]
    return {
        "mode": "markdown",
        "title": "Сметный артефакт",
        "content": "\n".join(lines).strip(),
        "tables": [
            {
                "title": table.title,
                "kind": table.kind,
                "rows": len(table.rows),
                "amount_total": table.amount_total,
                "headers": table.headers,
                "data": table.rows,
            }
            for table in tables
        ],
        **(
            {
                "lsr_form": lsr_form,
                "rim_lsr_form": lsr_form,
                **({"model_lsr_form": model_lsr_form} if model_lsr_form and trace_lsr_form else {}),
                **({"rim_trace": trace_lsr_form.get("trace")} if trace_lsr_form else {}),
            }
            if lsr_form
            else {}
        ),
    }


_NORM_CANDIDATE_HEADERS = [
    "№ ВОР",
    "Исходная работа",
    "Ед. ВОР",
    "Кол-во ВОР",
    "Нормируемая работа",
    "Группа сборников",
    "Сборник/раздел",
    "Код ГЭСН",
    "Наименование ГЭСН",
    "Ед. ГЭСН",
    "Кол-во в измерителе нормы",
    "Статус применимости",
    "Комментарий",
]


def _md_cell(value: Any) -> str:
    text = str(value if value is not None else "").replace("\n", " ").strip()
    return text.replace("|", "\\|")


def _norm_candidate_section(candidate: dict[str, Any]) -> str:
    profile = candidate.get("norm_profile") if isinstance(candidate.get("norm_profile"), dict) else {}
    navigation = profile.get("navigation") if isinstance(profile.get("navigation"), list) else []
    nav_titles = []
    for item in navigation[:2]:
        if not isinstance(item, dict):
            continue
        title = str(item.get("title") or item.get("heading") or item.get("name") or "").strip()
        if title:
            nav_titles.append(title)
    base = str(candidate.get("collection") or "").strip()
    if nav_titles:
        return " / ".join([x for x in [base, *nav_titles] if x])
    return base


def _norm_candidate_comment(candidate: dict[str, Any]) -> str:
    parts: list[str] = []
    if candidate.get("unit_compatible") is not None:
        parts.append("ед. изм. совместима" if candidate.get("unit_compatible") else "проверить ед. изм.")
    score = candidate.get("score_total")
    if score is not None:
        try:
            parts.append(f"score={float(score):.2f}")
        except (TypeError, ValueError):
            parts.append(f"score={score}")
    reasons = candidate.get("rejection_reasons")
    if isinstance(reasons, list) and reasons:
        parts.append("; ".join(str(x) for x in reasons[:3] if x))
    selection = candidate.get("selection_note")
    if selection:
        parts.append(str(selection))
    return "; ".join(x for x in parts if x)


def build_norm_candidate_artifact_from_lookup(
    norm_lookup_trace: dict[str, Any],
    *,
    question: str = "",
) -> dict[str, Any] | None:
    """Build the stage-1 ВОР -> ГЭСН candidate table from executed lookup trace.

    The function formats already returned search_norm candidates. It does not
    pick the final norm, bind rows, price resources, or infer missing quantities.
    """
    results = norm_lookup_trace.get("results") if isinstance(norm_lookup_trace, dict) else None
    if not isinstance(results, list):
        return None

    rows: list[list[str]] = []
    for lookup_idx, item in enumerate(results, start=1):
        if not isinstance(item, dict):
            continue
        call = item.get("call") if isinstance(item.get("call"), dict) else {}
        args = call.get("args") if isinstance(call.get("args"), dict) else {}
        result = item.get("result") if isinstance(item.get("result"), dict) else {}
        work = str(args.get("work_description") or "").strip()
        unit_hint = str(args.get("unit_hint") or "").strip()
        family = str(result.get("work_family") or args.get("work_family") or "").strip()
        candidates = result.get("candidates") if isinstance(result.get("candidates"), list) else []
        if not candidates:
            rows.append([
                str(lookup_idx),
                work,
                unit_hint,
                "",
                work,
                family,
                "",
                "",
                "",
                "",
                "",
                str(result.get("status") or "not_found"),
                str(result.get("hint") or "кандидаты ГЭСН не найдены; строка останется 0.00/пусто с примечанием"),
            ])
            continue
        for candidate in candidates:
            if not isinstance(candidate, dict):
                continue
            rows.append([
                str(lookup_idx),
                work,
                unit_hint,
                "",
                work,
                family,
                _norm_candidate_section(candidate),
                str(candidate.get("norm_code") or ""),
                str(candidate.get("title") or ""),
                str(candidate.get("measure_unit") or ""),
                "",
                str(candidate.get("applicability_status") or result.get("status") or ""),
                _norm_candidate_comment(candidate),
            ])
    if not rows and norm_lookup_trace.get("enabled") is False:
        return None

    lines = ["# Таблица кандидатов ГЭСН", ""]
    if question:
        lines += ["## Запрос", str(question).strip(), ""]
    lines += [
        "## Этап 1: ВОР -> кандидаты ГЭСН",
        "Это рабочая таблица доступных кандидатов. Это не ЛСР и не расчет денег.",
        "",
        "| " + " | ".join(_NORM_CANDIDATE_HEADERS) + " |",
        "| " + " | ".join("---" for _ in _NORM_CANDIDATE_HEADERS) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_md_cell(cell) for cell in row) + " |")
    lines += [
        "",
        "## Следующий шаг",
        "Можно сразу сказать: «деньги по ним». ЛЕС посчитает по доступным кандидатам; "
        "чего не хватает, останется 0.00/пусто с примечанием.",
    ]
    table = {
        "title": "ВОР ↔ кандидаты ГЭСН",
        "kind": "norm_candidates",
        "rows": len(rows),
        "amount_total": None,
        "headers": list(_NORM_CANDIDATE_HEADERS),
        "data": rows,
    }
    return {
        "mode": "markdown",
        "stage": "norm_candidates",
        "title": "Таблица кандидатов ГЭСН",
        "content": "\n".join(lines).strip(),
        "tables": [table],
        "norm_candidate_table": table,
    }


def persist_smeta_artifact_exports(
    artifact: dict[str, Any] | None,
    *,
    output_dir: str | Path = "storage/smeta_artifacts",
    prefix: str = "smeta_artifact",
) -> dict[str, Any] | None:
    """Write XLSX and CSV downloads for an already built smeta artifact.

    The export preserves model-written tables.  It does not add works, choose
    norms, change rates, or recalculate hidden positions.
    """
    if not artifact:
        return None
    tables = [t for t in artifact.get("tables") or [] if t.get("headers") and t.get("data")]
    if not tables:
        return artifact

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    import time

    stamp = int(time.time() * 1000)
    safe_prefix = re.sub(r"[^A-Za-z0-9_.-]+", "_", prefix).strip("_") or "smeta_artifact"
    xlsx_name = f"{safe_prefix}_{stamp}.xlsx"
    csv_name = f"{safe_prefix}_{stamp}.csv"
    xlsx_path = out_dir / xlsx_name
    csv_path = out_dir / csv_name

    import openpyxl

    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "Свод"
    summary_ws.append(["Таблица", "Тип", "Строк", "Сумма по видимым строкам"])
    for table in tables:
        summary_ws.append([
            table.get("title") or "",
            table.get("kind") or "",
            table.get("rows") or 0,
            table.get("amount_total") or "",
        ])
    used = {"Свод"}
    lsr_form = artifact.get("rim_lsr_form") if isinstance(artifact.get("rim_lsr_form"), dict) else None
    if not lsr_form:
        lsr_form = artifact.get("lsr_form") if isinstance(artifact.get("lsr_form"), dict) else None
    if lsr_form:
        ws = wb.create_sheet("ЛСР РИМ")
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

        max_col = _LSR_FORM_MAX_COL
        title_fill = PatternFill(fill_type="solid", fgColor="F2F6FA")
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        total_fill = PatternFill(fill_type="solid", fgColor="CFE2F3")
        note_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        thin = Side(style="thin", color="B7C9D6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def put(row: int, col: int, value: Any, *, bold: bool = False, size: int = 9, align: str = "left") -> None:
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=bold, size=size)
            cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)

        put(1, 10, "Приложение № 3", size=8, align="right")
        put(2, 8, "к Методике (приказ Минстроя России от 04.08.2020 № 421/пр)", size=8, align="right")
        ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=12)
        put(4, 1, "(наименование стройки)", size=8)
        put(5, 1, "(наименование объекта капитального строительства)", size=8)
        put(7, 1, "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ (СМЕТА) № ____", bold=True, size=11, align="center")
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=max_col)
        put(8, 1, str(lsr_form.get("title") or "Локальный сметный расчет (смета)"), align="center")
        ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=max_col)
        put(9, 1, "Составлен ресурсным методом / сценарной формой выдачи по видимым строкам ответа", size=8)
        ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=max_col)
        put(10, 1, str(lsr_form.get("note") or ""), size=8)
        ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=max_col)
        ws.cell(row=10, column=1).fill = note_fill
        put(12, 1, "Сметная стоимость", bold=True)
        if lsr_form.get("amount_total") is not None:
            put(12, 4, float(lsr_form.get("amount_total") or 0), bold=True, align="right")
            ws.cell(row=12, column=4).number_format = '# ##0'
        put(12, 6, "руб.", size=8)

        header_row = 15
        number_row = 16
        for col in range(1, max_col + 1):
            header = _LSR_FORM_HEADERS.get(col, "")
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, size=9)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            num_cell = ws.cell(row=number_row, column=col, value=col)
            num_cell.font = Font(size=8)
            num_cell.fill = header_fill
            num_cell.border = border
            num_cell.alignment = Alignment(horizontal="center", vertical="center")

        data_start = number_row + 1
        for idx, row in enumerate(lsr_form.get("rows") or [], 1):
            row_idx = data_start + idx - 1
            ws.cell(row=row_idx, column=1, value=idx)
            ws.cell(row=row_idx, column=2, value=row.get("basis") or "")
            ws.cell(row=row_idx, column=3, value=row.get("title") or "")
            ws.cell(row=row_idx, column=4, value=row.get("unit") or "")
            ws.cell(row=row_idx, column=5, value=row.get("quantity") or "")
            ws.cell(row=row_idx, column=6, value=1)
            ws.cell(row=row_idx, column=7, value=row.get("quantity") or "")
            ws.cell(row=row_idx, column=10, value=row.get("unit_price") or "")
            ws.cell(row=row_idx, column=11, value=1)
            ws.cell(row=row_idx, column=12, value=row.get("amount") if row.get("amount") is not None else "")
        if lsr_form.get("amount_total") is not None:
            total_row = data_start + len(lsr_form.get("rows") or [])
            ws.cell(row=total_row, column=3, value="ВСЕГО по смете")
            ws.cell(row=total_row, column=12, value=lsr_form.get("amount_total"))
            ws.cell(row=total_row, column=12).number_format = '# ##0'
            for col in range(1, max_col + 1):
                ws.cell(row=total_row, column=col).fill = total_fill
                ws.cell(row=total_row, column=col).font = Font(bold=True, size=9)

        for row_idx in range(4, 13):
            for col in range(1, max_col + 1):
                ws.cell(row=row_idx, column=col).fill = title_fill
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row_idx in range(data_start, ws.max_row + 1):
            ws.cell(row=row_idx, column=12).number_format = '# ##0'
        ws.freeze_panes = ws.cell(row=data_start, column=1)
        ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=ws.max_row, column=max_col).coordinate}"
        widths = {1: 6, 2: 18, 3: 48, 4: 9, 5: 12, 6: 7, 7: 12, 8: 16, 9: 9, 10: 18, 11: 7, 12: 18}
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        src_ws = wb.create_sheet("Источники ЛСР")
        src_ws.append(["№ п/п", "Источник / статус", "Исходная таблица", "Обоснование", "Наименование работ и затрат"])
        for idx, row in enumerate(lsr_form.get("rows") or [], 1):
            src_ws.append([
                idx,
                row.get("status") or "",
                row.get("source_table") or "",
                row.get("basis") or "",
                row.get("title") or "",
            ])
        for column_cells in src_ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 80)
            src_ws.column_dimensions[column_cells[0].column_letter].width = width
        used.add("ЛСР РИМ")
        used.add("Источники ЛСР")
        wb.move_sheet(ws, offset=-wb.index(ws))
        wb.active = 0
    for table in tables:
        ws = wb.create_sheet(_safe_sheet_title(str(table.get("title") or "Таблица"), used))
        ws.append([str(h) for h in table.get("headers") or []])
        for row in table.get("data") or []:
            ws.append([str(cell) for cell in row])
        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
            ws.column_dimensions[column_cells[0].column_letter].width = width
    wb.save(xlsx_path)

    with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh, delimiter=";")
        if lsr_form:
            writer.writerow(["ЛСР РИМ (форма 421/пр)"])
            writer.writerow([_LSR_FORM_HEADERS.get(col, "") for col in range(1, _LSR_FORM_MAX_COL + 1)])
            writer.writerow([col for col in range(1, _LSR_FORM_MAX_COL + 1)])
            for idx, row in enumerate(lsr_form.get("rows") or [], 1):
                csv_row = [""] * _LSR_FORM_MAX_COL
                csv_row[0] = idx
                csv_row[1] = row.get("basis") or ""
                csv_row[2] = row.get("title") or ""
                csv_row[3] = row.get("unit") or ""
                csv_row[4] = row.get("quantity") or ""
                csv_row[5] = 1
                csv_row[6] = row.get("quantity") or ""
                csv_row[9] = row.get("unit_price") or ""
                csv_row[10] = 1
                csv_row[11] = row.get("amount") if row.get("amount") is not None else ""
                writer.writerow(csv_row)
            if lsr_form.get("amount_total") is not None:
                csv_row = [""] * _LSR_FORM_MAX_COL
                csv_row[2] = "ВСЕГО по смете"
                csv_row[11] = lsr_form.get("amount_total")
                writer.writerow(csv_row)
            writer.writerow([])
            writer.writerow(["Источники ЛСР"])
            writer.writerow(["№ п/п", "Источник / статус", "Исходная таблица", "Обоснование", "Наименование работ и затрат"])
            for idx, row in enumerate(lsr_form.get("rows") or [], 1):
                writer.writerow([
                    idx,
                    row.get("status") or "",
                    row.get("source_table") or "",
                    row.get("basis") or "",
                    row.get("title") or "",
                ])
            writer.writerow([])
        for table in tables:
            writer.writerow([str(table.get("title") or "Таблица")])
            writer.writerow([str(h) for h in table.get("headers") or []])
            for row in table.get("data") or []:
                writer.writerow([str(cell) for cell in row])
            writer.writerow([])

    artifact = dict(artifact)
    artifact["downloads"] = {
        "xlsx": f"/api/smeta-artifacts/download?path={xlsx_name}",
        "csv": f"/api/smeta-artifacts/download?path={csv_name}",
    }
    artifact["files"] = {
        "xlsx_path": str(xlsx_path),
        "csv_path": str(csv_path),
    }
    return artifact


def compact_smeta_answer(answer: str, artifact: dict[str, Any] | None) -> str:
    """Optionally compact long estimate tables in chat.

    Default is deliberately full-visible: the model's estimate answer is the
    answer. Operators may opt into legacy compaction for very small screens via
    LES_SMETA_COMPACT_CHAT_TABLES=1; artifacts still stay attached either way.
    """
    if not artifact:
        return answer
    trace_visible = _trace_lsr_visible_answer(artifact)
    if trace_visible:
        return trace_visible
    answer = _prepend_trace_lsr_summary(answer, artifact)
    if str(os.getenv("LES_SMETA_COMPACT_CHAT_TABLES", "0")).strip().lower() not in {"1", "true", "yes", "on"}:
        return answer
    tables = extract_smeta_tables(answer)
    if not tables or sum(len(table.rows) for table in tables) < 5:
        return answer

    result_lines: list[str] = []
    lsr_form = artifact.get("rim_lsr_form") if isinstance(artifact.get("rim_lsr_form"), dict) else None
    if not lsr_form:
        lsr_form = artifact.get("lsr_form") if isinstance(artifact.get("lsr_form"), dict) else None
    if lsr_form and lsr_form.get("rows"):
        rows_count = len(lsr_form.get("rows") or [])
        amount = f", сумма {_fmt_money(float(lsr_form['amount_total']))}" if lsr_form.get("amount_total") is not None else ""
        source_tables = ", ".join(str(x) for x in (lsr_form.get("source_tables") or []) if x)
        source_note = f" Источник: {source_tables}." if source_tables else ""
        result_lines.extend([
            f"ЛСР-форма вынесена в артефакт/XLSX: {rows_count} строк{amount}.",
            f"Проверка строк артефакта имеет приоритет над ручным итогом в тексте ответа.{source_note}",
            "Исходные таблицы ниже сокращены в чате, полная форма и источники лежат в артефакте.",
            "",
        ])
    lines = str(answer or "").splitlines()
    idx = 0
    table_num = 0
    while idx < len(lines):
        if idx < len(lines) - 1 and _is_table_header(lines[idx], lines[idx + 1]):
            start = idx
            idx += 2
            while idx < len(lines) and "|" in lines[idx]:
                idx += 1
            table = tables[table_num] if table_num < len(tables) else None
            table_num += 1
            if table and len(table.rows) >= 4:
                amount = f", сумма строк {_fmt_money(table.amount_total)}" if table.amount_total is not None else ""
                result_lines.append(f"Таблица вынесена в артефакт: {table.title} ({len(table.rows)} строк{amount}).")
            else:
                result_lines.extend(lines[start:idx])
            continue
        result_lines.append(lines[idx])
        idx += 1
    compacted = "\n".join(result_lines)
    if lsr_form and lsr_form.get("amount_total") is not None:
        compacted = _drop_conflicting_manual_totals(compacted, float(lsr_form["amount_total"]))
    compacted = re.sub(r"\n{3,}", "\n\n", compacted).strip()
    return compacted or answer


def _trace_lsr_visible_answer(artifact: dict[str, Any]) -> str:
    """Render the checked RIM trace as the visible chat answer when it exists."""
    lsr_form = artifact.get("rim_lsr_form") if isinstance(artifact.get("rim_lsr_form"), dict) else None
    if not lsr_form or lsr_form.get("schema") != "lsr_rim_trace_form_v1":
        return ""
    total = lsr_form.get("amount_total")
    if total is None:
        return ""
    trace_summary = (lsr_form.get("trace") or {}).get("summary") or {}
    try:
        input_rows = int(trace_summary.get("input_rows") or 0)
        bound_rows = int(trace_summary.get("bound_rows") or len(lsr_form.get("rows") or []))
    except (TypeError, ValueError):
        input_rows = 0
        bound_rows = len(lsr_form.get("rows") or [])
    coverage = ""
    if input_rows and bound_rows < input_rows:
        coverage = f" ({bound_rows}/{input_rows} строк ВОР рассчитано)"
    lines = [f"ЛСР РИМ сформирована по расчетной трассе{coverage}: **{_fmt_money(float(total))}**."]
    details: list[str] = []
    pricebook = str(lsr_form.get("pricebook") or "").strip()
    finality = str(lsr_form.get("finality") or "").strip()
    if pricebook:
        details.append(f"книга цен: {pricebook}")
    if finality:
        details.append(f"статус: {finality}")
    if details:
        lines.append("; ".join(details) + ".")
    flags = [
        str(flag).strip()
        for flag in ((lsr_form.get("trace") or {}).get("summary") or {}).get("flags", [])
        if str(flag).strip()
    ]
    lines += ["", *_lsr_markdown(lsr_form)]
    if flags:
        lines += [
            "",
            "Примечания: есть незакрытые строки/ресурсные цены/ставки; подробный перечень сохранён в артефакте.",
        ]
    return "\n".join(lines).strip()


def _prepend_trace_lsr_summary(answer: str, artifact: dict[str, Any]) -> str:
    lsr_form = artifact.get("rim_lsr_form") if isinstance(artifact.get("rim_lsr_form"), dict) else None
    if not lsr_form or lsr_form.get("schema") != "lsr_rim_trace_form_v1":
        return answer
    total = lsr_form.get("amount_total")
    if total is None:
        return answer
    marker = "Системный РИМ-расчёт по подключённым источникам"
    if marker in str(answer or ""):
        return answer
    pricebook = str(lsr_form.get("pricebook") or "").strip()
    finality = str(lsr_form.get("finality") or "").strip()
    details = []
    if pricebook:
        details.append(f"книга цен: {pricebook}")
    if finality and finality != "priced_final":
        details.append(f"статус: {finality}")
    suffix = f" ({'; '.join(details)})" if details else ""
    return (
        f"{marker}: **{_fmt_money(float(total))}**{suffix}.\n"
        "Расчётная РИМ-форма построена по выбранным строкам и подключённым источникам.\n\n"
        f"{str(answer or '').strip()}"
    ).strip()

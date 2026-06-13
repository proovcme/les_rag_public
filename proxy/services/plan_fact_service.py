"""План/факт: ВОР (план) ↔ журнал полевых объёмов (факт) — W11.2 (LES3_PLAN).

ADR-11: ноль LLM. План — свод ВОР (`bor_service`, из Parquet спецификаций);
факт — confirmed-агрегаты журнала объёмов (`field_intake_service`, SQL). Сопоставление
по нормализованному наименованию × нормализованной единице; разница, процент выполнения
и остаток считает Python над числами SQL/Parquet — детерминированно.

Статусы строки: ``matched`` (есть план и факт), ``over`` (факт > плана),
``plan_only`` (план есть, факта нет — не начато), ``fact_only`` (факт без позиции в плане).
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from proxy.services.bor_service import generate_bor, normalize_unit

logger = logging.getLogger(__name__)

# Порог длины для сопоставления по вхождению (короткие строки совпадают слишком легко).
_MIN_CONTAINS_LEN = 4
# Допуск на округление при сравнении план/факт (единицы объёма).
_EPS = 1e-6


def _norm_name(name: str) -> str:
    return " ".join(str(name or "").split()).casefold()


def _names_match(plan_norm: str, fact_norm: str) -> bool:
    """Детерминированное сопоставление наименований план↔факт.

    Равенство нормализованных строк, либо вхождение одной в другую при длине
    короткой строки ≥ порога (чтобы «м» не цеплялось ко всему подряд).
    """
    if not plan_norm or not fact_norm:
        return False
    if plan_norm == fact_norm:
        return True
    shorter = min(plan_norm, fact_norm, key=len)
    if len(shorter) < _MIN_CONTAINS_LEN:
        return False
    return plan_norm in fact_norm or fact_norm in plan_norm


def reconcile_lines(bor_lines: list[dict], field_rows: list[dict]) -> dict[str, Any]:
    """Свести план (строки ВОР) и факт (агрегаты журнала) в таблицу план/факт.

    `bor_lines` — payload-словари BorLine (name/unit/qty/section/code/mark).
    `field_rows` — словари агрегата журнала (position/unit/total/entries).
    Жадное однопроходное сопоставление: один факт-агрегат не уходит в два плана.
    """
    facts = [
        {
            "position": str(r.get("position") or ""),
            "unit": normalize_unit(r.get("unit")),
            "total": float(r.get("total") or 0.0),
            "entries": int(r.get("entries") or 0),
            "_norm": _norm_name(r.get("position")),
            "_used": False,
        }
        for r in field_rows
    ]

    rows: list[dict[str, Any]] = []
    for line in bor_lines:
        plan_name = str(line.get("name") or "")
        plan_unit = normalize_unit(line.get("unit"))
        plan_norm = _norm_name(plan_name)
        plan_qty = line.get("qty")
        plan_qty = float(plan_qty) if plan_qty is not None else None

        matched = [
            f for f in facts
            if not f["_used"] and f["unit"] == plan_unit and _names_match(plan_norm, f["_norm"])
        ]
        fact_qty = 0.0
        fact_entries = 0
        positions: list[str] = []
        for f in matched:
            f["_used"] = True
            fact_qty += f["total"]
            fact_entries += f["entries"]
            positions.append(f["position"])

        rows.append(_make_row(line, plan_name, plan_unit, plan_qty, fact_qty, fact_entries, positions, matched))

    # Факты без позиции в плане — перевыполнение/неучтённое.
    for f in facts:
        if f["_used"]:
            continue
        rows.append({
            "name": f["position"], "section": "", "code": "", "mark": "", "unit": f["unit"],
            "plan_qty": None, "fact_qty": round(f["total"], 4),
            "delta": round(f["total"], 4), "remaining": None, "done_pct": None,
            "status": "fact_only", "fact_entries": f["entries"], "fact_positions": [f["position"]],
        })

    rows.sort(key=lambda r: (_STATUS_ORDER.get(r["status"], 9), r["name"].casefold()))
    return {"rows": rows, "totals": _totals(rows)}


_STATUS_ORDER = {"over": 0, "matched": 1, "plan_only": 2, "fact_only": 3}


def _make_row(line, plan_name, plan_unit, plan_qty, fact_qty, fact_entries, positions, matched) -> dict[str, Any]:
    if plan_qty is None:
        status = "matched" if matched else "plan_only"
        done_pct = None
        remaining = None
        delta = round(fact_qty, 4) if matched else None
    elif fact_qty <= _EPS:
        status = "plan_only"
        done_pct = 0.0
        remaining = round(plan_qty, 4)
        delta = round(-plan_qty, 4)
    else:
        delta = round(fact_qty - plan_qty, 4)
        done_pct = round(fact_qty / plan_qty * 100.0, 1) if plan_qty > _EPS else None
        remaining = round(max(0.0, plan_qty - fact_qty), 4)
        status = "over" if fact_qty - plan_qty > _EPS else "matched"
    return {
        "name": plan_name,
        "section": str(line.get("section") or ""),
        "code": str(line.get("code") or ""),
        "mark": str(line.get("mark") or ""),
        "unit": plan_unit,
        "plan_qty": round(plan_qty, 4) if plan_qty is not None else None,
        "fact_qty": round(fact_qty, 4),
        "delta": delta,
        "remaining": remaining,
        "done_pct": done_pct,
        "status": status,
        "fact_entries": fact_entries,
        "fact_positions": positions,
    }


def _totals(rows: list[dict]) -> dict[str, Any]:
    return {
        "lines": len(rows),
        "matched": sum(1 for r in rows if r["status"] == "matched"),
        "over": sum(1 for r in rows if r["status"] == "over"),
        "plan_only": sum(1 for r in rows if r["status"] == "plan_only"),
        "fact_only": sum(1 for r in rows if r["status"] == "fact_only"),
    }


def generate_plan_fact(
    dataset_id: str,
    *,
    storage_root: Path = Path("storage/datasets"),
    zahvatka: str = "",
    output_dir: Path | None = None,
) -> dict[str, Any]:
    """Полный цикл: ВОР датасета + журнал (confirmed) → таблица план/факт. Без LLM."""
    from proxy.services.field_intake_service import aggregate_volumes

    bor = generate_bor(dataset_id, storage_root=storage_root)
    field_rows = aggregate_volumes(status="confirmed", zahvatka=zahvatka)
    result = reconcile_lines(bor.get("lines", []), field_rows)
    result.update({
        "dataset_id": dataset_id,
        "plan_lines": bor.get("bor_lines", 0),
        "fact_aggregates": len(field_rows),
        "zahvatka": zahvatka,
        "xlsx_path": None,
    })
    if output_dir is not None and result["rows"]:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = output_dir / f"plan_fact_{dataset_id}_{stamp}.xlsx"
        plan_fact_to_xlsx(result["rows"], xlsx_path, title=f"План/факт — {dataset_id}")
        result["xlsx_path"] = str(xlsx_path)
    return result


_STATUS_LABEL = {
    "matched": "в работе", "over": "перевыполнение",
    "plan_only": "не начато", "fact_only": "вне плана",
}


def plan_fact_to_xlsx(rows: list[dict], output_path: Path, title: str = "План/факт") -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "План-факт"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Сформировано {datetime.now():%Y-%m-%d %H:%M} · строк: {len(rows)} · числа — SQL/Parquet, без LLM"])
    ws.append([])
    headers = ["№", "Раздел", "Наименование", "Ед.", "План", "Факт", "Δ (факт−план)", "Остаток", "Готово, %", "Статус", "Записей факта"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    for idx, r in enumerate(rows, 1):
        ws.append([
            idx, r["section"], r["name"], r["unit"],
            r["plan_qty"] if r["plan_qty"] is not None else "—",
            r["fact_qty"],
            r["delta"] if r["delta"] is not None else "—",
            r["remaining"] if r["remaining"] is not None else "—",
            r["done_pct"] if r["done_pct"] is not None else "—",
            _STATUS_LABEL.get(r["status"], r["status"]),
            r["fact_entries"],
        ])
    widths = {"A": 6, "B": 16, "C": 50, "D": 8, "E": 12, "F": 12, "G": 14, "H": 12, "I": 10, "J": 16, "K": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("[PLAN_FACT] %s строк → %s", len(rows), output_path)
    return len(rows)

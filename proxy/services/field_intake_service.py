"""Журнал полевых объёмов (W8.1) + запросы/отчёты по объёмам (W8.4).

ADR-11: числа НЕ считает LLM. Ввод — CRUD/GUI/чат-команда (regex), агрегации —
только SQL, ответы — шаблонами. Хранение — в метабазе рядом с задачником.

Запись ссылается на чертёж (`doc_id` из `documents`) и/или элемент CAD/BIM-графа
(`element_id` == `cad_bim_elements.source_id`) — мягко, текстовыми колонками без FK.
В отчёты и чат попадают только `confirmed`-записи (ручной ввод оператора = confirmed;
будущий VLM-конвейер W8.2 кладёт `pending` → подтверждение человеком W8.3).
"""

from __future__ import annotations

import calendar
import logging
import re
import sqlite3
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any

from backend.rag_config import rag_meta_db_path

logger = logging.getLogger(__name__)

FIELD_STATUSES = ("pending", "confirmed", "rejected")

# Месяцы для разбора периода «за июнь [2026]» (стемы; «март» раньше «ма», «июн» раньше «июл»).
_MONTHS = {
    "январ": 1, "феврал": 2, "март": 3, "апрел": 4, "ма": 5, "июн": 6,
    "июл": 7, "август": 8, "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12,
}
_MONTH_NAMES = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель", 5: "май", 6: "июнь",
    7: "июль", 8: "август", 9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь",
}

# «запиши объём 50 м3 монолитная плита захватка 3» / «учти выполнение 120 м2 кладка»
RECORD_RE = re.compile(
    r"^\s*(?:запиши|учти|занеси|добавь)\s+(?:объ[её]м|выполнени[ея])\s+"
    r"(?P<vol>\d+(?:[.,]\d+)?)\s*(?P<unit>[а-яёa-z][\wа-яё.²³/]*)\s+(?P<rest>.{1,300})$",
    re.IGNORECASE | re.DOTALL,
)
# Захватка/секция/ось внутри свободного хвоста.
_ZAHVATKA_RE = re.compile(r"захватк[аеуи]?\s*[№#]?\s*(?P<z>[\w./-]+)", re.IGNORECASE)
# Период «с 01.06.2026 по 30.06.2026»
_RANGE_RE = re.compile(
    r"с\s+(?P<a>\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})\s+по\s+(?P<b>\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})",
    re.IGNORECASE,
)


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(rag_meta_db_path())
    conn.row_factory = sqlite3.Row
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS les_field_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_date TEXT NOT NULL,
            position TEXT NOT NULL,
            zahvatka TEXT NOT NULL DEFAULT '',
            volume REAL NOT NULL,
            unit TEXT NOT NULL DEFAULT '',
            doc_id TEXT NOT NULL DEFAULT '',
            element_id TEXT NOT NULL DEFAULT '',
            author TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'confirmed',
            notes TEXT NOT NULL DEFAULT '',
            created_at REAL NOT NULL,
            updated_at REAL NOT NULL
        )
        """
    )
    try:  # Q3: партиционирование по объекту (project_id=0 — без объекта/глобально)
        conn.execute("ALTER TABLE les_field_entries ADD COLUMN project_id INTEGER NOT NULL DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    return conn


def _today() -> str:
    return date.today().isoformat()


def _num(value: Any) -> float:
    return float(str(value).replace(",", ".").strip())


# ── CRUD (W8.1) ──

def create_entry(
    position: str,
    volume: float,
    unit: str = "",
    *,
    entry_date: str = "",
    zahvatka: str = "",
    doc_id: str = "",
    element_id: str = "",
    author: str = "",
    status: str = "confirmed",
    notes: str = "",
    project_id: int = 0,
) -> dict[str, Any]:
    if status not in FIELD_STATUSES:
        raise ValueError(f"status must be one of {FIELD_STATUSES}")
    now = time.time()
    with _connect() as conn:
        cur = conn.execute(
            "INSERT INTO les_field_entries"
            "(entry_date, position, zahvatka, volume, unit, doc_id, element_id, author, status, notes, project_id, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                (entry_date or _today()).strip(),
                position.strip(),
                zahvatka.strip(),
                float(volume),
                unit.strip(),
                doc_id.strip(),
                element_id.strip(),
                author.strip(),
                status,
                notes.strip(),
                int(project_id),
                now,
                now,
            ),
        )
        conn.commit()
        entry_id = cur.lastrowid
    logger.info("[FIELD] запись #%s: %s %.3f %s", entry_id, position[:60], float(volume), unit)
    return get_entry(entry_id)


def get_entry(entry_id: int) -> dict[str, Any]:
    with _connect() as conn:
        row = conn.execute("SELECT * FROM les_field_entries WHERE id=?", (entry_id,)).fetchone()
    return dict(row) if row else {}


def list_entries(
    status: str = "",
    *,
    zahvatka: str = "",
    position: str = "",
    date_from: str = "",
    date_to: str = "",
    project_id: int | None = None,
    limit: int = 200,
) -> list[dict[str, Any]]:
    clauses, params = [], []
    if status:
        clauses.append("status=?")
        params.append(status)
    if project_id is not None:  # Q3: фильтр по объекту
        clauses.append("project_id=?")
        params.append(int(project_id))
    if zahvatka:
        clauses.append("zahvatka LIKE ?")
        params.append(f"%{zahvatka}%")
    if position:
        clauses.append("position LIKE ?")
        params.append(f"%{position}%")
    if date_from:
        clauses.append("entry_date >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("entry_date <= ?")
        params.append(date_to)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    params.append(limit)
    with _connect() as conn:
        rows = conn.execute(
            f"SELECT * FROM les_field_entries {where} ORDER BY entry_date DESC, id DESC LIMIT ?",
            params,
        ).fetchall()
    return [dict(row) for row in rows]


def update_entry(
    entry_id: int,
    *,
    position: str | None = None,
    volume: float | None = None,
    unit: str | None = None,
    entry_date: str | None = None,
    zahvatka: str | None = None,
    doc_id: str | None = None,
    element_id: str | None = None,
    status: str | None = None,
    notes: str | None = None,
) -> dict[str, Any]:
    fields, params = [], []
    for name, value in (
        ("position", position),
        ("unit", unit),
        ("entry_date", entry_date),
        ("zahvatka", zahvatka),
        ("doc_id", doc_id),
        ("element_id", element_id),
        ("notes", notes),
    ):
        if value is not None:
            fields.append(f"{name}=?")
            params.append(value.strip())
    if volume is not None:
        fields.append("volume=?")
        params.append(float(volume))
    if status is not None:
        if status not in FIELD_STATUSES:
            raise ValueError(f"status must be one of {FIELD_STATUSES}")
        fields.append("status=?")
        params.append(status)
    if not fields:
        return get_entry(entry_id)
    fields.append("updated_at=?")
    params.extend([time.time(), entry_id])
    with _connect() as conn:
        conn.execute(f"UPDATE les_field_entries SET {', '.join(fields)} WHERE id=?", params)
        conn.commit()
    return get_entry(entry_id)


def delete_entry(entry_id: int) -> bool:
    with _connect() as conn:
        cur = conn.execute("DELETE FROM les_field_entries WHERE id=?", (entry_id,))
        conn.commit()
    return cur.rowcount > 0


# ── Период из текста вопроса (regex, без LLM) ──

def _parse_period(text: str) -> tuple[str, str, str]:
    """Вернуть (date_from, date_to, label). Пустые границы — без ограничения."""
    low = text.casefold()
    m = _RANGE_RE.search(low)
    if m:
        a, b = _norm_date(m.group("a")), _norm_date(m.group("b"))
        if a and b:
            return a, b, f"с {a} по {b}"
    for stem, num in _MONTHS.items():
        if re.search(rf"\bза\s+{stem}\w*", low) or re.search(rf"\b{stem}\w*\s+месяц", low):
            year = _find_year(low)
            last_day = calendar.monthrange(year, num)[1]
            start = date(year, num, 1)
            end = date(year, num, last_day)
            month_name = _MONTH_NAMES[num]
            return start.isoformat(), end.isoformat(), f"{month_name} {year}"
    if "сегодня" in low:
        return _today(), _today(), "сегодня"
    return "", "", "за всё время"


def _find_year(text: str) -> int:
    m = re.search(r"\b(20\d{2})\b", text)
    return int(m.group(1)) if m else date.today().year


def _norm_date(raw: str) -> str:
    raw = raw.replace("/", ".").replace("-", ".")
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


# ── Агрегации и отчёт (W8.4, только SQL) ──

def aggregate_volumes(
    *,
    status: str = "confirmed",
    zahvatka: str = "",
    position: str = "",
    date_from: str = "",
    date_to: str = "",
    project_id: int | None = None,
) -> list[dict[str, Any]]:
    """Свод SUM(volume) GROUP BY (position, unit) — числа считает SQL, не LLM."""
    clauses, params = ["status=?"], [status]
    if project_id is not None:  # Q3: фильтр по объекту (None → все)
        clauses.append("project_id=?")
        params.append(int(project_id))
    if zahvatka:
        clauses.append("zahvatka LIKE ?")
        params.append(f"%{zahvatka}%")
    if position:
        clauses.append("position LIKE ?")
        params.append(f"%{position}%")
    if date_from:
        clauses.append("entry_date >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("entry_date <= ?")
        params.append(date_to)
    where = " AND ".join(clauses)
    with _connect() as conn:
        rows = conn.execute(
            f"SELECT position, unit, SUM(volume) AS total, COUNT(*) AS entries "
            f"FROM les_field_entries WHERE {where} "
            f"GROUP BY position, unit ORDER BY position, unit",
            params,
        ).fetchall()
    return [dict(row) for row in rows]


def _fmt_qty(value: float) -> str:
    return f"{value:.3f}".rstrip("0").rstrip(".")


def maybe_answer_field_volume_query(question: str) -> dict[str, Any] | None:
    """Детерминированный ответ на вопрос об объёмах (ADR-11: SQL + шаблон, без LLM)."""
    date_from, date_to, period_label = _parse_period(question)
    zm = _ZAHVATKA_RE.search(question)
    zahvatka = zm.group("z") if zm else ""
    summary = aggregate_volumes(
        zahvatka=zahvatka, date_from=date_from, date_to=date_to
    )
    filters = []
    if zahvatka:
        filters.append(f"захватка {zahvatka}")
    filters.append(period_label)
    head = "Объёмы (" + ", ".join(filters) + "):"
    if not summary:
        return {
            "answer": f"{head}\nЗаписей нет.",
            "rows": [],
            "total_entries": 0,
            "period": {"from": date_from, "to": date_to, "label": period_label},
        }
    lines = [f"**{head}**", "", "| Позиция | Ед. | Объём | Записей |", "|---|---|--:|--:|"]
    total_entries = 0
    for row in summary:
        total_entries += int(row["entries"])
        lines.append(
            f"| {row['position']} | {row['unit'] or '—'} | {_fmt_qty(row['total'])} | {row['entries']} |"
        )
    lines.append("")
    lines.append(f"_Источник: журнал полевых объёмов (confirmed), {total_entries} записей. Числа — SQL, без LLM._")
    return {
        "answer": "\n".join(lines),
        "rows": summary,
        "total_entries": total_entries,
        "period": {"from": date_from, "to": date_to, "label": period_label},
    }


# ── Чат-команда записи (W8.1, regex, без LLM) ──

def maybe_handle_field_command(question: str, author: str = "", project_id: int = 0) -> dict[str, Any] | None:
    """«запиши объём 50 м3 монолитная плита захватка 3» → запись в журнал.
    В режиме объекта (project_id>0) запись привязывается к объекту."""
    match = RECORD_RE.match(question.strip())
    if not match:
        return None
    try:
        volume = _num(match.group("vol"))
    except ValueError:
        return None
    unit = match.group("unit").strip()
    rest = match.group("rest").strip()
    zahvatka = ""
    zm = _ZAHVATKA_RE.search(rest)
    if zm:
        zahvatka = zm.group("z")
        rest = _ZAHVATKA_RE.sub("", rest).strip(" ,;—-")
    position = rest or "без позиции"
    entry = create_entry(position, volume, unit, zahvatka=zahvatka, author=author, project_id=project_id)
    where = f" (захватка {zahvatka})" if zahvatka else ""
    return {
        "answer": f"✓ Записано #{entry['id']}: {position} — {_fmt_qty(volume)} {unit}{where}.\n"
                  f"Спросить: «сколько {position} выполнено за <период>?»",
        "operation": "field_record",
        "entry_id": entry["id"],
    }


# ── Экспорт xlsx (W8.4) ──

def export_xlsx(output_path: Path, *, status: str = "confirmed", title: str = "Журнал полевых объёмов") -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    entries = list_entries(status=status, limit=10000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Объёмы"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Статус: {status} · записей: {len(entries)} · сформировано {datetime.now():%Y-%m-%d %H:%M}"])
    ws.append([])
    headers = ["Дата", "Позиция", "Захватка", "Объём", "Ед.", "Чертёж (doc_id)", "Элемент BIM", "Автор", "Статус", "Примечание"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    for e in entries:
        ws.append([
            e["entry_date"], e["position"], e["zahvatka"], round(e["volume"], 4), e["unit"],
            e["doc_id"], e["element_id"], e["author"], e["status"], e["notes"],
        ])
    widths = {"A": 12, "B": 40, "C": 12, "D": 12, "E": 8, "F": 22, "G": 28, "H": 14, "I": 12, "J": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("[FIELD] экспорт %s записей → %s", len(entries), output_path)
    return len(entries)

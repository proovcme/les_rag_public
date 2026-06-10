"""Формальный нормоконтроль комплекта документации — W13.1 (LES3_PLAN), v1.

ADR-11: все проверки детерминированные, без LLM. v1 покрывает то, что
проверяется алгоритмически по файлам и Parquet-таблицам конвейера:

  NK-01 формат листов по ГОСТ 2.301 (размеры страниц PDF, поворот, кратные);
  NK-02 текстовый слой (сканированные листы — предупреждение);
  NK-03 согласованность шифра комплекта в именах файлов;
  NK-04 ведомость рабочих чертежей (VEDOMOST в Parquet) ↔ фактовый состав.

v2 (остаток W13.1): графы основной надписи/подписи — нужен layout-анализ
штампа; правила заведены в этот же реестр, помечены manual.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)

PT_TO_MM = 25.4 / 72.0

# ГОСТ 2.301: основные форматы, мм (ширина × высота, портрет).
GOST_FORMATS = {
    "А4": (210.0, 297.0),
    "А3": (297.0, 420.0),
    "А2": (420.0, 594.0),
    "А1": (594.0, 841.0),
    "А0": (841.0, 1189.0),
}
FORMAT_TOLERANCE_MM = 6.0
# Кратные форматы: А4×3 = 297×630 и т.п. — длинная сторона кратна 210/297.
_MULTIPLier_MAX = 9

# Шифр комплекта в имени файла: «АТ-РД-ОВ2-С-00-П1», «123-2026-ОВ.С» и т.п. —
# берём префикс до последнего сегмента (номер листа/части).
CIPHER_RE = re.compile(r"^(?P<cipher>[А-ЯA-Z0-9]{1,12}(?:[-.][А-ЯA-Z0-9]{1,12}){1,8})", re.IGNORECASE)

PDF_SUFFIXES = {".pdf"}


@dataclass
class Finding:
    check: str        # NK-01 … NK-04
    severity: str     # error | warning | info
    target: str       # файл/лист
    message: str

    def payload(self) -> dict:
        return {"check": self.check, "severity": self.severity, "target": self.target, "message": self.message}


def _mm(points: float) -> float:
    return points * PT_TO_MM


def classify_format(width_mm: float, height_mm: float) -> str | None:
    """Имя ГОСТ-формата для размеров листа (учитывая поворот и кратные), либо None."""
    w, h = sorted((width_mm, height_mm))
    for name, (fw, fh) in GOST_FORMATS.items():
        if abs(w - fw) <= FORMAT_TOLERANCE_MM and abs(h - fh) <= FORMAT_TOLERANCE_MM:
            return name
    # Кратные (ГОСТ 2.301 т.2): короткая сторона = длинная сторона базового,
    # длинная = короткая сторона базового × кратность (А4×3 = 297×630).
    for name, (fw, fh) in GOST_FORMATS.items():
        if abs(w - fh) <= FORMAT_TOLERANCE_MM:
            for mult in range(2, _MULTIPLier_MAX + 1):
                if abs(h - fw * mult) <= FORMAT_TOLERANCE_MM * mult:
                    return f"{name}×{mult}"
    return None


def check_pdf_sheets(pdf_path: Path) -> list[Finding]:
    """NK-01 формат листов + NK-02 текстовый слой одного PDF."""
    import fitz

    findings: list[Finding] = []
    try:
        doc = fitz.open(pdf_path)
    except Exception as err:  # повреждённый файл — это тоже замечание
        return [Finding("NK-01", "error", pdf_path.name, f"PDF не открывается: {err}")]

    scanned_pages: list[int] = []
    try:
        for page_idx, page in enumerate(doc, 1):
            rect = page.rect
            fmt = classify_format(_mm(rect.width), _mm(rect.height))
            if fmt is None:
                findings.append(Finding(
                    "NK-01", "warning", f"{pdf_path.name}#стр.{page_idx}",
                    f"Нестандартный формат листа {_mm(rect.width):.0f}×{_mm(rect.height):.0f} мм (ГОСТ 2.301)",
                ))
            if len((page.get_text() or "").strip()) < 20:
                scanned_pages.append(page_idx)
    finally:
        doc.close()

    if scanned_pages:
        pages = ", ".join(map(str, scanned_pages[:20]))
        findings.append(Finding(
            "NK-02", "warning", pdf_path.name,
            f"Нет текстового слоя (скан) на страницах: {pages}" + ("…" if len(scanned_pages) > 20 else ""),
        ))
    return findings


def extract_cipher(file_name: str) -> str | None:
    match = CIPHER_RE.match(Path(file_name).stem)
    if not match:
        return None
    cipher = match.group("cipher")
    # Отрезаем последний сегмент (номер листа/тома), чтобы получить общий префикс комплекта.
    parts = re.split(r"[-.]", cipher)
    return "-".join(parts[:-1]) if len(parts) > 2 else cipher


def check_cipher_consistency(file_names: list[str]) -> list[Finding]:
    """NK-03: один комплект — один шифр в именах файлов."""
    findings: list[Finding] = []
    ciphers: dict[str, list[str]] = {}
    for name in file_names:
        cipher = extract_cipher(name)
        if cipher:
            ciphers.setdefault(cipher.upper(), []).append(name)
        else:
            findings.append(Finding("NK-03", "info", name, "Шифр комплекта в имени файла не распознан"))
    if len(ciphers) > 1:
        main = max(ciphers, key=lambda c: len(ciphers[c]))
        for cipher, names in ciphers.items():
            if cipher == main:
                continue
            for name in names:
                findings.append(Finding(
                    "NK-03", "warning", name,
                    f"Шифр «{cipher}» отличается от основного шифра комплекта «{main}»",
                ))
    return findings


def _norm_doc_ref(value: str) -> str:
    return re.sub(r"[\s_]+", "", str(value)).casefold()


def check_vedomost_vs_files(dataset_id: str, file_names: list[str], storage_root: Path) -> list[Finding]:
    """NK-04: позиции ведомости (VEDOMOST в Parquet) ↔ фактические файлы комплекта."""
    from proxy.services.bor_service import rows_from_parquet

    findings: list[Finding] = []
    parquet_root = storage_root / dataset_id / "_parquet"
    vedomost_rows: list[dict] = []
    if parquet_root.exists():
        for parquet_path in sorted(parquet_root.rglob("*.parquet")):
            for row in rows_from_parquet(parquet_path):
                if row.get("doc_type") == "VEDOMOST":
                    vedomost_rows.append(row)
    if not vedomost_rows:
        return [Finding("NK-04", "info", dataset_id, "Ведомость рабочих чертежей в датасете не распознана — сверка состава пропущена")]

    file_norms = {_norm_doc_ref(Path(name).stem): name for name in file_names}
    for row in vedomost_rows:
        ref = str(row.get("designation") or row.get("code") or "").strip()
        title = str(row.get("name") or "").strip()
        if not ref:
            continue
        ref_norm = _norm_doc_ref(ref)
        matched = any(ref_norm in fnorm or fnorm in ref_norm for fnorm in file_norms)
        if not matched:
            findings.append(Finding(
                "NK-04", "error", ref,
                f"Лист из ведомости отсутствует в комплекте: «{ref}» {title}".strip(),
            ))
    return findings


def findings_to_xlsx(findings: list[Finding], output_path: Path, title: str) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Нормоконтроль"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["№", "Проверка", "Серьёзность", "Объект", "Замечание"])
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for idx, finding in enumerate(findings, 1):
        ws.append([idx, finding.check, finding.severity, finding.target, finding.message])
    for col, width in {"A": 5, "B": 9, "C": 13, "D": 38, "E": 90}.items():
        ws.column_dimensions[col].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(findings)


def run_normcontrol(
    dataset_id: str,
    files_dir: Path,
    storage_root: Path = Path("storage/datasets"),
    output_dir: Path | None = None,
) -> dict:
    """Полный формальный прогон по комплекту. Без LLM (ADR-11)."""
    from datetime import datetime

    pdf_files = sorted(p for p in files_dir.rglob("*") if p.suffix.lower() in PDF_SUFFIXES) if files_dir.exists() else []
    file_names = [p.name for p in pdf_files]

    findings: list[Finding] = []
    for pdf_path in pdf_files:
        findings.extend(check_pdf_sheets(pdf_path))
    findings.extend(check_cipher_consistency(file_names))
    findings.extend(check_vedomost_vs_files(dataset_id, file_names, storage_root))

    severity_rank = {"error": 0, "warning": 1, "info": 2}
    findings.sort(key=lambda f: (severity_rank.get(f.severity, 3), f.check, f.target))

    result: dict = {
        "dataset_id": dataset_id,
        "files_checked": len(pdf_files),
        "findings_total": len(findings),
        "errors": sum(1 for f in findings if f.severity == "error"),
        "warnings": sum(1 for f in findings if f.severity == "warning"),
        "findings": [f.payload() for f in findings],
        "xlsx_path": None,
        "checks_skipped_v2": ["графы основной надписи/подписи (нужен layout-анализ штампа — W13.1 v2)"],
    }
    if output_dir is not None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = output_dir / f"normcontrol_{dataset_id}_{stamp}.xlsx"
        findings_to_xlsx(findings, xlsx_path, title=f"Нормоконтроль (формальный) — {dataset_id}")
        result["xlsx_path"] = str(xlsx_path)
    logger.info("[NORMCONTROL] %s: файлов=%s, замечаний=%s", dataset_id, len(pdf_files), len(findings))
    return result

"""Движок сборки ЛСР: позиция (объём + ресурсы) → цены → стеснённость → НР/СП → Всего.

Связывает кирпичи сметного ценообразования в сборку позиция-к-позиции (0 LLM, ADR-11):
  объём (ВОР) → ресурсы позиции → цены (ФГИС ЦС lookup / КАЦ для неучтённых) →
  ОЗП/ЭМ(вкл. ЗПМ)/М → коэф. стеснённости → НР/СП → Всего по позиции → свод в ЛСР.

Бухгалтерия позиции выверена на `Эталон_кровля` (медный отлив, поз.1):
  прямые = ОЗП + ЭМ(машины+ОТм) + М; ФОТ = ОЗП + ЗПМ; НР=ФОТ·нр%; СП=ФОТ·сп%; Всего=прямые+НР+СП.

Чего движок НЕ делает сам: разложение нормы ГЭСН на ресурсы (нужна база ГЭСН-2022, не загружена) —
ресурсы приходят на вход (как их выдала бы ГЭСН). Это следующий стык, как импорт ФГИС ЦС.

Виды ресурсов (`kind`): labor (рабочие→ОЗП) · machinist (ОТм→ЗПМ, в ЭМ) · machine (машины→ЭМ) ·
material (→М). Цена строки: явная `price`, иначе `code`→ФГИС ЦС, иначе material→КАЦ по наименованию.
"""

from __future__ import annotations

import csv
import re
import time
from pathlib import Path
from typing import Any, Optional

from backend.runtime_paths import mutable_path
from proxy.services import fsem_machinist_service as fsem
from proxy.services import stesnennost_service as st

# Справочник «машина→машинист» (ОТм/ЗПМ) вынесен в config/domain/fsem_machinist.yaml
# (fsem_machinist_service, handoff #3). Источник истины — fsem.machine_to_machinist(); не дублируем.

_EXPORT_FIELDS = (
    "section", "position_no", "position_code", "position_name", "position_unit", "position_qty",
    "row_type", "resource_code", "resource_name", "resource_unit", "resource_qty",
    "price_used", "price_source", "price_action", "cost", "ozp", "em", "zpm", "mat", "direct", "fot",
    "nr", "sp", "total", "flags",
)


def _f(value: Any) -> float:
    try:
        return float(str(value).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _norm_name(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower().replace("ё", "е")).strip(" .,;:")


def _resolve_book(book: str | None):
    """Имя книги цен → PriceBook | None (для lookup по коду)."""
    from proxy.services import fgis_price_service as fps

    path = fps.resolve_pricebook_path(book, allow_scratch=bool(book))
    if not path:
        return None
    return fps.get_pricebook(path)


def _resource_price(
    res: dict[str, Any], pricebook, kac_map: dict[str, float]
) -> tuple[Optional[float], str]:
    """Цена единицы ресурса + источник. (price, source) | (None, 'missing')."""
    if res.get("price") not in (None, ""):
        return _f(res.get("price")), "manual"
    code = str(res.get("code") or "").strip()
    if code and pricebook is not None:
        rec = pricebook.lookup(code)
        if rec is not None:
            return _f(rec.get("price_current_eff")), "fgis"
    if str(res.get("kind")) == "material" and kac_map:
        p = kac_map.get(_norm_name(res.get("name")))
        if p is not None:
            return _f(p), "kac"
    return None, "missing"


def _resolve_nr_sp_for_position(position: dict[str, Any]) -> tuple[float, float, dict[str, Any]]:
    """НР/СП from explicit fields, otherwise from norm code/name."""
    explicit_nr = position.get("nr_pct")
    explicit_sp = position.get("sp_pct")
    if explicit_nr not in (None, "") or explicit_sp not in (None, ""):
        return _f(explicit_nr), _f(explicit_sp), {
            "status": "explicit",
            "basis": position.get("nr_sp_basis") or "",
        }

    from proxy.services.gesn_service import get_norm
    from proxy.services.nr_sp_service import resolve as resolve_nr_sp

    norm = get_norm(position.get("code", "")) if position.get("code") else None
    code = str(position.get("code") or (norm or {}).get("code") or "").strip()
    name = str(position.get("name") or position.get("work") or "").strip()
    rs = resolve_nr_sp(name, code=code, rule_id=str(position.get("nr_sp_rule_id") or "") or None)
    if rs.get("default") and norm:
        rs = resolve_nr_sp(
            str(norm.get("name") or ""),
            code=code,
            rule_id=str(position.get("nr_sp_rule_id") or "") or None,
        )
    return _f(rs.get("nr_pct")), _f(rs.get("sp_pct")), rs


def _price_requirement(res: dict[str, Any]) -> dict[str, Any]:
    """Что нужно добрать, если цена ресурса не найдена."""
    kind = str(res.get("kind") or "")
    code = str(res.get("code") or "").strip()
    name = str(res.get("name") or "").strip() or "ресурс"
    unit = str(res.get("unit") or "").strip()
    if kind == "material":
        action = "needs_kac"
        human = f"нужен КАЦ: {name}"
    elif kind == "labor":
        action = "needs_labor_rate"
        human = f"нужна ставка ОЗП: {name}"
    elif kind == "machinist":
        action = "needs_machinist_rate"
        human = f"нужна ставка ЗПМ: {name}"
    elif kind == "machine":
        action = "needs_fgis_price"
        human = f"нужна цена эксплуатации машины: {name}"
    else:
        action = "needs_price"
        human = f"нужна цена ресурса: {name}"
    return {
        "action": action,
        "resource_kind": kind,
        "resource_code": code,
        "resource_name": name,
        "resource_unit": unit,
        "message": human,
    }


def _split_machinist_aggregates(resources: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Разложить агрегатную ОТм, если машины полностью покрываются известным тарифным маппингом."""
    out: list[dict[str, Any]] = []
    mapped_machines: list[dict[str, Any]] = []
    for res in resources:
        if str(res.get("kind") or "") != "machine":
            continue
        code = str(res.get("code") or "").strip()
        mapped = fsem.machine_to_machinist().get(code)
        if mapped:
            mapped_machines.append({**res, "_machinist_code": mapped[0], "_machinist_name": mapped[1]})

    mapped_qty = round(sum(_f(r.get("qty")) for r in mapped_machines), 6)
    consumed = False
    for res in resources:
        kind = str(res.get("kind") or "")
        code = str(res.get("code") or "").strip()
        qty = round(_f(res.get("qty")), 6)
        is_aggregate = kind == "machinist" and not code and qty > 0
        if not is_aggregate or consumed or abs(qty - mapped_qty) > max(0.0001, qty * 0.001):
            out.append(res)
            continue
        for machine in mapped_machines:
            out.append({
                "kind": "machinist",
                "name": machine["_machinist_name"],
                "unit": "чел.-ч",
                "qty": _f(machine.get("qty")),
                "code": machine["_machinist_code"],
                "machinist_basis": f"машина {machine.get('code')}: {machine.get('name', '')}",
            })
        consumed = True
    return out


def compute_position(
    position: dict[str, Any],
    *,
    pricebook=None,
    kac_map: dict[str, float] | None = None,
    k_ozp: float = 1.0,
    k_em: float = 1.0,
) -> dict[str, Any]:
    """Одна позиция: ресурсы → цены → ОЗП/ЭМ/ЗПМ/М → стеснённость → НР/СП → Всего."""
    if pricebook is None and position.get("code") and not position.get("resources"):
        # Code-only calculation is a user-facing entrypoint too; keep it on
        # the same manifest-default region/period as assemble/core.
        pricebook = _resolve_book(None)
    kac_map = kac_map or {}
    ozp = zpm = machine_only = mat = 0.0
    priced: list[dict[str, Any]] = []
    price_requirements: list[dict[str, Any]] = []
    flags: list[str] = []

    # Норма ГЭСН → ресурсы: если ресурсы не заданы, но есть код — разворачиваем по норме.
    resources = position.get("resources") or []
    norm_source_integrity: dict[str, Any] = {}
    if not resources and position.get("code"):
        from proxy.services import gesn_service

        norm = gesn_service.get_norm(position["code"], strict_family=True)
        if norm and str(norm.get("_source_kind") or "") == "structured_sqlite":
            from proxy.smeta_core.integrity import normative_base_integrity

            norm_source_integrity = normative_base_integrity(base_path=str(norm.get("_source_path") or ""))
            if not norm_source_integrity.get("trusted_for_pricing"):
                flags.append("нормативная база в карантине: semantic integrity gate не пройден")
        expanded = gesn_service.expand_position(position["code"], _f(position.get("qty")))
        if expanded is None:
            flags.append(f"норма ГЭСН не найдена: {position.get('code')}")
        else:
            resources = expanded
    resources = [dict(r) for r in resources]
    resources, fsem_trace = fsem.enrich_machinists(resources, quantity_field="qty")
    if fsem_trace.get("status") == "unresolved":
        details = list(fsem_trace.get("reasons") or fsem_trace.get("missing_machine_codes") or [])
        flags.append("ФСЭМ не смог детализировать ОТм: " + "; ".join(str(item) for item in details))

    for res in resources:
        kind = str(res.get("kind") or "")
        qty = _f(res.get("qty"))
        price, src = _resource_price(res, pricebook, kac_map)
        if price is None:
            req = _price_requirement(res)
            price_requirements.append(req)
            flags.append(req["message"])
            cost = 0.0
        else:
            cost = round(qty * price, 2)
        if kind == "labor":
            ozp += cost
        elif kind == "machinist":
            zpm += cost
        elif kind == "machine":
            machine_only += cost
        elif kind == "material":
            mat += cost
        else:
            flags.append(f"неизвестный вид ресурса: {kind!r}")
        priced.append({
            **res,
            "price_used": price,
            "price_source": src,
            "price_action": price_requirements[-1]["action"] if price is None and price_requirements else "",
            "cost": cost,
        })

    em = round(machine_only + zpm, 2)
    nr_pct, sp_pct, nr_sp_trace = _resolve_nr_sp_for_position(position)
    if not str(nr_sp_trace.get("status") or "").startswith("resolved") and nr_sp_trace.get("status") != "explicit":
        flags.append("НР/СП не разрешены по виду работ: нужен нормативный источник/явный выбор")
    pos_in = {
        "ozp": round(ozp, 2), "em": em, "zpm": round(zpm, 2), "mat": round(mat, 2),
        "nr_pct": nr_pct, "sp_pct": sp_pct,
    }
    res = st.apply_position(pos_in, k_ozp=k_ozp, k_em=k_em)
    chosen = res["adjusted"] if (k_ozp != 1.0 or k_em != 1.0) else res["base"]
    return {
        "code": position.get("code", ""),
        "name": position.get("name", ""),
        "unit": position.get("unit", ""),
        "qty": _f(position.get("qty")),
        "section": position.get("section", "") or "Без раздела",
        "resources": priced,
        "k_ozp": k_ozp, "k_em": k_em,
        "base": res["base"],
        "adjusted": res["adjusted"],
        "total": chosen["total"],
        "uplift": res["uplift"],
        "flags": flags,
        "price_requirements": price_requirements,
        "nr_sp_trace": nr_sp_trace,
        "norm_source_integrity": norm_source_integrity,
        "fsem_trace": fsem_trace,
    }


def assembled_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    """Плоские строки для CSV/XLSX: позиция + её ресурсы + итоги."""
    rows: list[dict[str, Any]] = []
    for idx, pos in enumerate(result.get("positions") or [], 1):
        base = pos.get("base") or {}
        common = {
            "section": pos.get("section", ""),
            "position_no": idx,
            "position_code": pos.get("code", ""),
            "position_name": pos.get("name", ""),
            "position_unit": pos.get("unit", ""),
            "position_qty": pos.get("qty", ""),
            "flags": "; ".join(pos.get("flags") or []),
        }
        rows.append({
            **common,
            "row_type": "position",
            "ozp": base.get("ozp"),
            "em": base.get("em"),
            "zpm": base.get("zpm"),
            "mat": base.get("mat"),
            "direct": base.get("direct"),
            "fot": base.get("fot"),
            "nr": base.get("nr"),
            "sp": base.get("sp"),
            "total": base.get("total"),
        })
        for res in pos.get("resources") or []:
            rows.append({
                **common,
                "row_type": str(res.get("kind") or "resource"),
                "resource_code": res.get("code", ""),
                "resource_name": res.get("name", ""),
                "resource_unit": res.get("unit", ""),
                "resource_qty": res.get("qty", ""),
                "price_used": res.get("price_used", ""),
                "price_source": res.get("price_source", ""),
                "price_action": res.get("price_action", ""),
                "cost": res.get("cost", ""),
            })
    return [{field: row.get(field, "") for field in _EXPORT_FIELDS} for row in rows]


def export_assembled(
    result: dict[str, Any],
    out_path: str | Path,
    *,
    fmt: str | None = None,
    title: str = "Локальный сметный расчёт",
) -> Path:
    """Сохранить собранную ЛСР в CSV/XLSX. Возвращает путь."""
    path = Path(out_path)
    fmt = (fmt or path.suffix.lstrip(".") or "xlsx").lower()
    if fmt not in {"csv", "xlsx"}:
        raise ValueError("fmt должен быть csv или xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = assembled_rows(result)
    if fmt == "csv":
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=list(_EXPORT_FIELDS), delimiter=";")
            writer.writeheader()
            writer.writerows(rows)
        return path

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ЛСР"
    ws.append([title])
    ws.append(["Итого", result.get("summary", {}).get("total", 0), "Позиций", result.get("summary", {}).get("positions", 0)])
    flags = result.get("summary", {}).get("flags") or []
    ws.append(["Флаги", "; ".join(flags) if flags else ""])
    ws.append([])
    ws.append(list(_EXPORT_FIELDS))
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([row.get(field, "") for field in _EXPORT_FIELDS])
    for col in ws.columns:
        letter = col[0].column_letter
        width = min(42, max(10, max(len(str(c.value or "")) for c in col[:80]) + 2))
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(path)
    return path


def export(
    positions: list[dict[str, Any]],
    *,
    book: str | None = None,
    kac_prices: dict[str, float] | None = None,
    condition: str | None = None,
    k_ozp: float | None = None,
    k_em: float | None = None,
    output_dir: str | Path | None = None,
    fmt: str = "xlsx",
    title: str = "Локальный сметный расчёт",
) -> dict[str, Any]:
    """Собрать ЛСР и сохранить файл выгрузки."""
    result = assemble(
        positions,
        book=book,
        kac_prices=kac_prices,
        condition=condition,
        k_ozp=k_ozp,
        k_em=k_em,
    )
    stamp = time.strftime("%Y%m%d_%H%M%S")
    out_dir = Path(output_dir) if output_dir is not None else mutable_path("storage/lsr")
    out = out_dir / f"lsr_{stamp}.{fmt.lower()}"
    path = export_assembled(result, out, fmt=fmt, title=title)
    return {"summary": result["summary"], "path": str(path), "rows": len(assembled_rows(result))}


def assemble(
    positions: list[dict[str, Any]],
    *,
    book: str | None = None,
    kac_prices: dict[str, float] | None = None,
    condition: str | None = None,
    k_ozp: float | None = None,
    k_em: float | None = None,
) -> dict[str, Any]:
    """Собрать ЛСР из позиций. Цены — ФГИС ЦС (book) / КАЦ (kac_prices) / явные. Стеснённость — опц."""
    if condition:
        cond = st.get_condition(condition)
        if cond is None:
            raise ValueError(f"Условие стеснённости {condition!r} не найдено")
        k_ozp = _f(cond["k_ozp"]) if k_ozp is None else k_ozp
        k_em = _f(cond["k_em"]) if k_em is None else k_em
    k_ozp = 1.0 if k_ozp is None else k_ozp
    k_em = 1.0 if k_em is None else k_em

    pricebook = _resolve_book(book) if book is not None else _resolve_book(None)
    kac_map = {_norm_name(k): _f(v) for k, v in (kac_prices or {}).items()}

    computed = [
        compute_position(p, pricebook=pricebook, kac_map=kac_map, k_ozp=k_ozp, k_em=k_em)
        for p in (positions or [])
    ]

    sections: dict[str, dict[str, Any]] = {}
    for c in computed:
        s = sections.setdefault(c["section"], {"section": c["section"], "positions": 0, "total": 0.0})
        s["positions"] += 1
        s["total"] = round(s["total"] + c["total"], 2)

    base_total = round(sum(c["base"]["total"] for c in computed), 2)
    total = round(sum(c["total"] for c in computed), 2)
    flags = [f for c in computed for f in c["flags"]]
    price_requirements = [req for c in computed for req in c.get("price_requirements", [])]
    unsafe_source = any(
        item.get("norm_source_integrity")
        and not item["norm_source_integrity"].get("trusted_for_pricing")
        for item in computed
    )
    return {
        "k_ozp": k_ozp, "k_em": k_em, "condition": condition,
        "positions": computed,
        "sections": list(sections.values()),
        "summary": {
            "positions": len(computed),
            "base_total": base_total,
            "total": total,
            "stesnennost_uplift": round(total - base_total, 2),
            "flags": flags,
            "price_requirements": price_requirements,
            "needs_price": len(price_requirements),
            "result_status": (
                "unsafe_source" if unsafe_source else "priced_partial" if flags else "priced_final"
            ),
        },
    }

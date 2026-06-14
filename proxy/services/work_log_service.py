"""W20.3 — Общий журнал работ (ОЖР, РД-11-05-2007). 0 LLM (ADR-11).

ОЖР — исполнительный документ стройки. Ядро ценности: **раздел 3 «Сведения о
выполнении работ» собирается из confirmed-записей журнала объёмов (W8) бит-в-бит**,
строго по объекту (Q3). Метаданные шапки (объект/заказчик/подрядчик/ИТР/спецжурналы)
ведутся в `les_work_log_meta`. Выгрузка — xlsx (числа из SQL, не LLM).

Разделы (РД-11-05-2007): титул · 1 ИТР · 2 спецжурналы · 3 выполнение работ ·
4-5 строительный контроль · 6 перечень исполнительной документации.
"""
from __future__ import annotations

import json
import sqlite3
import time
from pathlib import Path
from typing import Any

from backend.rag_config import rag_meta_db_path

OUTPUT_DIR = Path("data/worklog_out")


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(rag_meta_db_path())
    conn.row_factory = sqlite3.Row
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS les_work_log_meta (
            project_id INTEGER PRIMARY KEY,
            object_name TEXT NOT NULL DEFAULT '',
            customer TEXT NOT NULL DEFAULT '',
            contractor TEXT NOT NULL DEFAULT '',
            permit TEXT NOT NULL DEFAULT '',
            itr TEXT NOT NULL DEFAULT '[]',
            spec_journals TEXT NOT NULL DEFAULT '[]',
            updated_at REAL NOT NULL
        )
        """
    )
    return conn


def set_work_log_meta(project_id: int, **fields: Any) -> dict[str, Any]:
    """Обновить шапку ОЖР объекта. itr/spec_journals — списки строк."""
    allowed = {"object_name", "customer", "contractor", "permit", "itr", "spec_journals"}
    now = time.time()
    with _connect() as conn:
        row = conn.execute("SELECT project_id FROM les_work_log_meta WHERE project_id=?", (project_id,)).fetchone()
        if row is None:
            conn.execute("INSERT INTO les_work_log_meta(project_id, updated_at) VALUES (?,?)", (int(project_id), now))
        for key, value in fields.items():
            if key not in allowed:
                continue
            if key in ("itr", "spec_journals"):
                value = json.dumps(value if isinstance(value, list) else [value], ensure_ascii=False)
            conn.execute(f"UPDATE les_work_log_meta SET {key}=?, updated_at=? WHERE project_id=?", (value, now, int(project_id)))
        conn.commit()
    return get_work_log_meta(project_id)


def get_work_log_meta(project_id: int) -> dict[str, Any]:
    with _connect() as conn:
        row = conn.execute("SELECT * FROM les_work_log_meta WHERE project_id=?", (project_id,)).fetchone()
    if not row:
        return {"project_id": project_id, "object_name": "", "customer": "", "contractor": "",
                "permit": "", "itr": [], "spec_journals": []}
    meta = dict(row)
    for key in ("itr", "spec_journals"):
        try:
            meta[key] = json.loads(meta.get(key) or "[]")
        except (json.JSONDecodeError, TypeError):
            meta[key] = []
    return meta


def build_section3(project_id: int) -> list[dict[str, Any]]:
    """Раздел 3 «Сведения о выполнении работ» — confirmed-записи журнала объёмов
    объекта, хронологически (бит-в-бит из журнала, числа — SQL). 0 LLM."""
    from proxy.services.field_intake_service import list_entries
    rows = list_entries(status="confirmed", project_id=project_id, limit=5000)
    # хронология: по дате возрастанию (раздел 3 ведётся по ходу работ)
    rows = sorted(rows, key=lambda r: (r.get("entry_date") or "", r.get("id") or 0))
    return [
        {
            "date": r.get("entry_date"),
            "work": r.get("position"),
            "zahvatka": r.get("zahvatka"),
            "volume": r.get("volume"),
            "unit": r.get("unit"),
            "author": r.get("author"),
        }
        for r in rows
    ]


def build_work_log(project_id: int) -> dict[str, Any]:
    """Собрать ОЖР объекта (метаданные + раздел 3 из журнала объёмов). 0 LLM."""
    meta = get_work_log_meta(project_id)
    # имя объекта по умолчанию — из карточки проекта
    if not meta.get("object_name"):
        try:
            from proxy.services.project_service import get_project
            proj = get_project(project_id) or {}
            meta["object_name"] = proj.get("name", "")
        except Exception:
            pass
    section3 = build_section3(project_id)
    # раздел 6 — перечень исполнительной документации (АОСР и т.п.): пока из решений со ссылкой на НТД
    return {
        "project_id": project_id,
        "title": "ОБЩИЙ ЖУРНАЛ РАБОТ (РД-11-05-2007)",
        "header": {
            "object_name": meta.get("object_name", ""),
            "customer": meta.get("customer", ""),
            "contractor": meta.get("contractor", ""),
            "permit": meta.get("permit", ""),
        },
        "itr": meta.get("itr", []),                # раздел 1
        "spec_journals": meta.get("spec_journals", []),  # раздел 2
        "section3": section3,                       # раздел 3 — из журнала объёмов
        "section3_count": len(section3),
        "total_volume": round(sum(float(r.get("volume") or 0) for r in section3), 3),
    }


def _output_dir() -> Path:
    import os
    d = Path(os.getenv("LES_WORKLOG_OUT_DIR", str(OUTPUT_DIR)))
    d.mkdir(parents=True, exist_ok=True)
    return d


def export_xlsx(project_id: int, out_path: Path | None = None) -> Path:
    """Выгрузка ОЖР в xlsx (шапка + разделы 1/2/3). Числа из SQL, не LLM."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    log = build_work_log(project_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "ОЖР"
    ws["A1"] = log["title"]
    ws["A1"].font = Font(bold=True, size=13)
    r = 3
    for label, key in (("Объект", "object_name"), ("Застройщик/заказчик", "customer"),
                       ("Лицо, осуществляющее строительство", "contractor"), ("Разрешение на строительство", "permit")):
        ws.cell(r, 1, label).font = Font(color="555555")
        ws.cell(r, 2, log["header"].get(key, ""))
        r += 1
    r += 1
    ws.cell(r, 1, "Раздел 1. Список ИТР").font = Font(bold=True); r += 1
    for person in log["itr"] or ["—"]:
        ws.cell(r, 1, f"• {person}"); r += 1
    r += 1
    ws.cell(r, 1, "Раздел 2. Специальные журналы").font = Font(bold=True); r += 1
    for j in log["spec_journals"] or ["—"]:
        ws.cell(r, 1, f"• {j}"); r += 1
    r += 1
    ws.cell(r, 1, "Раздел 3. Сведения о выполнении работ").font = Font(bold=True); r += 1
    headers = ["Дата", "Наименование работ", "Захватка", "Объём", "Ед.изм.", "Исполнитель"]
    for c, h in enumerate(headers, 1):
        ws.cell(r, c, h).font = Font(bold=True)
    r += 1
    for row in log["section3"]:
        for c, key in enumerate(("date", "work", "zahvatka", "volume", "unit", "author"), 1):
            ws.cell(r, c, row.get(key))
        r += 1
    for col, width in (("A", 14), ("B", 46), ("C", 14), ("D", 12), ("E", 10), ("F", 18)):
        ws.column_dimensions[col].width = width

    out_path = Path(out_path) if out_path else _output_dir() / f"ojr_{project_id}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(out_path))
    return out_path

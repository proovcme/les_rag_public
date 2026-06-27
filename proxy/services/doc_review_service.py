"""Doc-review engine (СПДС-нормоконтроль, Phase 3) — RAG-led SPDS review.

Прогоняет комплект (DocumentSet) по review-map и собирает структурированные review-items
(docs/DOC_REVIEW_GOST_R_21_101_2026_PLAN.md §5). АРХИТЕКТУРА:
  • computed-checks (формализуемое) → evidence-статус (computed_issue / supported_by_evidence /
    not_applicable). Переиспользуют normcontrol_service (NK-03/NK-04), не дублируют.
  • retrieval-цели → факты корпуса (устаревший ГОСТ-2020 / стадия ПД-РД) + текст требования (под-фаза
    retrieval, doc_review_retrieval_service, лексика 0 LLM); поиск недоступен → review_needed.
  • layout/manual_required → manual_required (без уверенного extraction вердикта нет).
  • human_decision у каждого item = unset: финальный статус даёт инженер, не движок.
Движок НЕ выносит юридического решения — он предлагает proposed issues с requirement/document source_ref.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path

from proxy.services.document_set_model import DocumentSet, VedomostMatch, build_document_set, match_vedomost
from proxy.services.normcontrol_review_map_service import ReviewMap, ReviewTarget
from proxy.services.normcontrol_service import check_cipher_consistency

# Статусы review-item (см. §5). Финальные (confirmed/rejected) ставит человек, не движок.
S_COMPUTED_ISSUE = "computed_issue"
S_SUPPORTED = "supported_by_evidence"
S_NOT_APPLICABLE = "not_applicable"
S_MANUAL = "manual_required"
S_REVIEW_NEEDED = "review_needed"


@dataclass
class ReviewItem:
    rule_id: str
    clause: str
    status: str
    severity: str
    target: str
    requirement: dict          # {source_ref, snippet} — ссылка на пункт ГОСТ
    document_evidence: list    # [{kind, source_ref, snippet}]
    computed_check: dict       # {name, status: ok|issue|not_run}
    model_note: str            # пояснение модели (заполняется позже), не вердикт
    human_decision: str = "unset"  # unset|confirmed|rejected|needs_more_evidence
    confidence: float = 0.0

    def to_dict(self) -> dict:
        return asdict(self)


def _requirement(standard: str, clause: str) -> dict:
    return {"source_ref": f"{standard}#clause={clause}" if clause else standard, "snippet": ""}


def _doc_ev(targets: list[str]) -> list:
    return [{"kind": "document", "source_ref": t, "snippet": ""} for t in targets[:20]]


def _run_computed(target: ReviewTarget, doc_set: DocumentSet, ved: VedomostMatch | None,
                  title_block: dict | None = None, sheet_format: dict | None = None):
    """Возвращает (status, computed_check, document_evidence, note) для computed-цели.
    None → computed-движка для этого check нет (вызывающий пометит manual_required)."""
    check = target.check

    if check == "rulepack_version_gate":
        return S_SUPPORTED, {"name": check, "status": "ok"}, [], "Выбран ГОСТ Р 21.101-2026."

    if check == "base_designation_consistency":
        names = [d.file_name for d in doc_set.documents]
        warns = [f for f in check_cipher_consistency(names) if f.severity == "warning"]
        if warns:
            return (S_COMPUTED_ISSUE, {"name": check, "status": "issue"}, _doc_ev([f.target for f in warns]),
                    "Шифр части файлов расходится с основным (может быть и несколько комплектов — решает инженер).")
        return S_SUPPORTED, {"name": check, "status": "ok"}, [], ""

    if check == "designation_pattern":
        bad = [d.file_name for d in doc_set.documents if d.designation is None]
        if bad:
            return (S_COMPUTED_ISSUE, {"name": check, "status": "issue"}, _doc_ev(bad),
                    "Обозначение СПДС не распознано в имени файла.")
        return S_SUPPORTED, {"name": check, "status": "ok"}, [], ""

    if check == "designation_separators":
        bad = [d.file_name for d in doc_set.documents if d.designation and d.designation.bad_separators]
        if bad:
            return (S_COMPUTED_ISSUE, {"name": check, "status": "issue"}, _doc_ev(bad),
                    "Недопустимые пробелы/разделители в обозначении.")
        return S_SUPPORTED, {"name": check, "status": "ok"}, [], ""

    if check in ("vedomost_vs_files", "vedomost_missing"):
        if ved is None:
            return S_NOT_APPLICABLE, {"name": check, "status": "not_run"}, [], "Ведомость не распознана — сверка пропущена."
        if ved.missing:
            return (S_COMPUTED_ISSUE, {"name": check, "status": "issue"},
                    _doc_ev([m["designation"] for m in ved.missing]),
                    "Документ из ведомости отсутствует в комплекте.")
        return S_SUPPORTED, {"name": check, "status": "ok"}, [], ""

    if check == "vedomost_extra":
        if ved is None:
            return S_NOT_APPLICABLE, {"name": check, "status": "not_run"}, [], "Ведомость не распознана."
        if ved.extra:
            return (S_COMPUTED_ISSUE, {"name": check, "status": "issue"}, _doc_ev(ved.extra),
                    "Файл присутствует в комплекте, но не указан в ведомости.")
        return S_SUPPORTED, {"name": check, "status": "ok"}, [], ""

    if check == "electronic_readiness":
        from proxy.services.document_set_model import SUPPORTED_EXT
        bad = [d.file_name for d in doc_set.documents if d.ext and d.ext not in SUPPORTED_EXT]
        if bad:
            return (S_COMPUTED_ISSUE, {"name": check, "status": "issue"}, _doc_ev(bad),
                    "Неподдерживаемый формат файла.")
        # формат ок, но текстовый слой требует sidecar/OCR-проверки → не финальный pass
        return S_MANUAL, {"name": check, "status": "ok"}, [], "Формат поддерживается; текстовый слой проверяется отдельно (sidecar/OCR)."

    if check == "sheet_format":
        sf = sheet_format or {}
        checked = int(sf.get("checked_pages", 0) or 0)
        bad = sf.get("bad_pages", []) or []
        pdf_total = int(sf.get("pdf_total", 0) or 0)
        if checked == 0:
            return (S_MANUAL, {"name": check, "status": "not_run", "pdf_total": pdf_total}, [],
                    "PDF-геометрия листов не проверена: нет доступных PDF-страниц или PyMuPDF недоступен.")
        if bad:
            ev = [{"kind": "document", "source_ref": b.get("source_ref", ""), "snippet": b.get("snippet", "")}
                  for b in bad[:20]]
            return (S_COMPUTED_ISSUE,
                    {"name": check, "status": "issue", "checked_pages": checked, "bad_pages": len(bad)},
                    ev,
                    f"Найдены нестандартные размеры листов по ГОСТ 2.301: {len(bad)} из {checked} проверенных страниц.")
        examples = sf.get("examples", []) or []
        ev = [{"kind": "document", "source_ref": e.get("source_ref", ""), "snippet": e.get("snippet", "")}
              for e in examples[:6]]
        return (S_SUPPORTED, {"name": check, "status": "ok", "checked_pages": checked}, ev,
                f"Форматы листов распознаны по ГОСТ 2.301 на {checked} проверенных страницах.")

    if check == "title_block_present":
        tb = title_block or {}
        checked = int(tb.get("checked", 0))
        present = int(tb.get("present", 0))
        scan = int(tb.get("scan", 0))
        no_stamp = int(tb.get("no_stamp", 0))
        if checked == 0:
            return None  # нет PDF для проверки → честный manual_required (не fake)
        ex = tb.get("examples", []) or []
        text_checked = present + no_stamp
        scan_note = f" ({scan} сканов без текст-слоя — нужен OCR)" if scan else ""
        if text_checked > 0 and present >= max(1, text_checked // 2):  # штамп в текст-листах
            return (S_SUPPORTED, {"name": check, "status": "ok"},
                    _doc_ev([e["file"] for e in ex if e.get("present")]),
                    f"Основная надпись найдена в {present}/{text_checked} текст-листах "
                    f"(сигнатуры полей ГОСТ Р 21.101){scan_note}.")
        if no_stamp > 0:
            misplaced = [
                e for e in ex
                if (e.get("layout_zone") or {}).get("placement") == "outside_expected_zone"
            ]
            layout_note = " Сигнатуры основной надписи есть вне ожидаемой зоны листа." if misplaced else ""
            return (S_COMPUTED_ISSUE, {"name": check, "status": "issue"},
                    _doc_ev([e["file"] for e in ex if not e.get("present") and not e.get("scan")]),
                    f"Штамп не распознан в {no_stamp}/{text_checked} текст-листах — "
                    f"проверить основную надпись{scan_note}.{layout_note}")
        # остались только сканы — текстом не определить, честный manual (не fake issue)
        return (S_MANUAL, {"name": check, "status": "not_run"}, [],
                f"Все {scan} проверенных PDF — сканы без текст-слоя; штамп проверить вручную/OCR.")

    return None  # computed-движка нет → manual_required


def _run_retrieval(target: ReviewTarget, ev: dict | None):
    """Retrieval-evidence (факт в корпусе + текст требования) → (status, document_evidence, note).
    Нет факта (поиск недоступен/не строился) → review_needed (фолбэк сохранён). Источник evidence —
    doc_review_retrieval_service.build_retrieval_evidence (лексический поиск по корпусу, 0 LLM)."""
    check = target.check
    fact = (ev or {}).get("fact")
    if check == "outdated_standard_in_corpus":
        if fact is None:
            return S_REVIEW_NEEDED, [], "Устаревшая база ищется в корпусе (поиск недоступен — review)."
        if fact.get("found"):
            return (S_COMPUTED_ISSUE, fact.get("hits", []),
                    "В корпусе найден ГОСТ Р 21.101-2020 при rulepack 2026 — устаревшая нормативная база.")
        return S_SUPPORTED, [], "Устаревший ГОСТ Р 21.101-2020 в корпусе не найден."
    if check == "project_stage_detect":
        if fact is None:
            return S_REVIEW_NEEDED, [], "Стадия ищется в корпусе (поиск недоступен — review)."
        stage = fact.get("stage")
        if stage in ("ПД", "РД"):
            return S_SUPPORTED, fact.get("hits", []), f"Стадия документации по корпусу: {stage}."
        return S_MANUAL, fact.get("hits", []), "Стадия (ПД/РД) по корпусу не однозначна — ручное подтверждение."
    return S_REVIEW_NEEDED, [], "Требование/база ищется в RAG (под-фаза retrieval)."


def run_review(doc_set: DocumentSet, review_map: ReviewMap, *, vedomost_entries: list[dict] | None = None,
               title_block: dict | None = None, sheet_format: dict | None = None,
               retrieval_evidence: dict | None = None) -> list[ReviewItem]:
    """Прогон комплекта по review-map → список review-items. Чистая функция (без живых сервисов).
    title_block — сводка детектора штампа (title_block_extract_service.detect_dataset) для D4-002.
    retrieval_evidence — {rule_id: {check, fact, requirement}} от retrieval-подфазы (оркестратор):
    факты корпуса + текст требования для kind: retrieval. Нет → retrieval-цели = review_needed."""
    ved = match_vedomost(doc_set, vedomost_entries) if vedomost_entries is not None else None
    items: list[ReviewItem] = []
    for t in review_map.targets:
        req = _requirement(review_map.standard, t.clause)
        if t.kind == "computed":
            res = _run_computed(t, doc_set, ved, title_block, sheet_format)
            if res is not None:
                status, cc, ev, note = res
                items.append(ReviewItem(t.id, t.clause, status, t.severity, t.title, req, ev, cc, note))
                continue
            # computed заявлен, но движка нет — честно manual, не fake pass
            items.append(ReviewItem(t.id, t.clause, S_MANUAL, t.severity, t.title, req, [],
                                    {"name": t.check, "status": "not_run"}, "Автопроверка не реализована."))
        elif t.kind == "retrieval":
            ev = (retrieval_evidence or {}).get(t.id)
            status, ev_docs, note = _run_retrieval(t, ev)
            req_hit = (ev or {}).get("requirement")   # flavor B: текст требования ГОСТ из корпуса
            if req_hit and req_hit.get("source_ref"):
                req = {"source_ref": req_hit["source_ref"], "snippet": req_hit.get("snippet", "")}
            cc_status = "not_run" if status == S_REVIEW_NEEDED else "ok"
            items.append(ReviewItem(t.id, t.clause, status, t.severity, t.title, req, ev_docs,
                                    {"name": t.check, "status": cc_status}, note))
        else:  # layout | manual_required
            items.append(ReviewItem(t.id, t.clause, S_MANUAL, t.severity, t.title, req, [],
                                    {"name": t.check, "status": "not_run"},
                                    "Нужна ручная проверка (layout/штамп/изменения)."))
    return items


def review_summary(items: list[ReviewItem]) -> dict:
    by_status: dict[str, int] = {}
    for it in items:
        by_status[it.status] = by_status.get(it.status, 0) + 1
    return {"total": len(items), "by_status": by_status,
            "computed_issues": by_status.get(S_COMPUTED_ISSUE, 0),
            "manual_required": by_status.get(S_MANUAL, 0),
            "review_needed": by_status.get(S_REVIEW_NEEDED, 0)}


def review_defense_pack(items: list[ReviewItem], review_map: ReviewMap) -> dict:
    """Единый defense-contract для нормоконтроля: proposed claim, основание, gap, action."""
    from proxy.services.evidence_contract import DefenseClaim, DefensePack, DefenseStatus

    claims: list[DefenseClaim] = []
    for it in items:
        source_refs = []
        req_ref = str((it.requirement or {}).get("source_ref") or "").strip()
        if req_ref:
            source_refs.append(req_ref)
        for ev in it.document_evidence or []:
            ref = str((ev or {}).get("source_ref") or "").strip()
            if ref:
                source_refs.append(ref)
        if it.status == S_COMPUTED_ISSUE:
            status = DefenseStatus.COMPUTED
            gaps = []
            actions = ["Инженеру подтвердить/отклонить замечание."]
            confidence = 0.75
        elif it.status == S_SUPPORTED:
            status = DefenseStatus.SUPPORTED
            gaps = []
            actions = []
            confidence = 0.8
        elif it.status == S_NOT_APPLICABLE:
            status = DefenseStatus.SUPPORTED
            gaps = []
            actions = []
            confidence = 0.6
        elif it.status == S_REVIEW_NEEDED:
            status = DefenseStatus.MISSING
            gaps = ["Недостаточно retrieval evidence для уверенного предложения."]
            actions = ["Поднять/уточнить нормативный и проектный источник, затем повторить review."]
            confidence = 0.2
        else:
            status = DefenseStatus.MANUAL_REQUIRED
            gaps = ["Автопроверка не финализирует этот пункт."]
            actions = ["Проверить инженером/нормоконтролёром."]
            confidence = 0.3
        claims.append(DefenseClaim(
            id=it.rule_id,
            domain="normcontrol.doc_review",
            title=it.target,
            statement=f"{it.rule_id}: {it.model_note or it.status}",
            status=status,
            severity=it.severity,
            source_refs=source_refs,
            inputs=[{"name": "computed_check", "value": it.computed_check}],
            gaps=gaps,
            actions=actions,
            confidence=confidence,
        ))
    summary = review_summary(items)
    return DefensePack(
        domain="normcontrol.doc_review",
        title=f"Нормоконтроль {review_map.standard}",
        status=DefenseStatus.MANUAL_REQUIRED,
        claims=claims,
        summary={
            "standard": review_map.standard,
            "rulepack": review_map.name,
            "total": summary["total"],
            "computed_issues": summary["computed_issues"],
            "manual_required": summary["manual_required"],
            "review_needed": summary["review_needed"],
            "human_final_required": True,
        },
        coverage={"items": len(items), "source_backed": sum(1 for c in claims if c.source_refs)},
        required_actions=["Финальное решение по каждому пункту ставит инженер."],
    ).payload()


# ── оркестратор: dataset_id → review (файлы из MetaDB, ведомость из Parquet) ──
# Используется и роутером /api/doc-review, и чат-инструментом doc_review (агент-роутер).

def _dataset_file_names(dataset_id: str) -> list[str]:
    import sqlite3

    from backend.rag_config import rag_meta_db_path

    try:
        with sqlite3.connect(rag_meta_db_path()) as conn:
            rows = conn.execute(
                "SELECT file_name FROM documents WHERE dataset_id=?", (dataset_id,)
            ).fetchall()
        return [r[0] for r in rows if r and r[0]]
    except Exception:
        return []


def _vedomost_entries(dataset_id: str, storage_root: Path = Path("storage/datasets")):
    """Позиции ведомости (VEDOMOST в Parquet) → [{designation, name}]. None если ведомости нет."""
    try:
        from proxy.services.bor_service import rows_from_parquet
    except Exception:
        return None
    parquet_root = storage_root / dataset_id / "_parquet"
    if not parquet_root.exists():
        return None
    entries: list[dict] = []
    for pq in sorted(parquet_root.rglob("*.parquet")):
        try:
            for row in rows_from_parquet(pq):
                if row.get("doc_type") == "VEDOMOST":
                    ref = str(row.get("designation") or row.get("code") or "").strip()
                    if ref:
                        entries.append({"designation": ref, "name": str(row.get("name") or "").strip()})
        except Exception:
            continue
    return entries or None


def review_dataset(dataset_id: str, *, rulepack: str = "gost_r_21_101_2026"):
    """dataset_id → (review_map, items). Поднимает ValueError('no_documents') если файлов нет,
    FileNotFoundError/ValueError если rulepack битый (валидирует load_review_map)."""
    from proxy.services.normcontrol_review_map_service import load_review_map

    review_map = load_review_map(rulepack)
    files = _dataset_file_names(dataset_id)
    if not files:
        raise ValueError("no_documents")
    vedomost = _vedomost_entries(dataset_id)
    source_paths = _dataset_source_paths(dataset_id)
    sheet_format = build_sheet_format_evidence(source_paths)
    title_block = None
    try:  # Phase 5: детект штампа по сэмплу PDF (D4-002 → computed); сбой → None → честный manual
        from proxy.services import title_block_extract_service as tbx
        title_block = tbx.detect_dataset(source_paths, sample=8)
    except Exception:
        title_block = None
    retrieval_evidence = None
    try:  # Phase 3+: retrieval-подфаза (факты корпуса + текст требования) для kind: retrieval
        from proxy.services import doc_review_retrieval_service as drr
        retrieval_evidence = drr.build_retrieval_evidence(dataset_id, review_map)
    except Exception:
        retrieval_evidence = None
    doc_set = build_document_set(files)
    return review_map, run_review(doc_set, review_map, vedomost_entries=vedomost,
                                  title_block=title_block, sheet_format=sheet_format,
                                  retrieval_evidence=retrieval_evidence)


def _dataset_source_paths(dataset_id: str) -> list[str]:
    import sqlite3

    from backend.rag_config import rag_meta_db_path

    try:
        with sqlite3.connect(rag_meta_db_path()) as conn:
            rows = conn.execute(
                "SELECT source_path FROM documents WHERE dataset_id=? AND source_path IS NOT NULL",
                (dataset_id,),
            ).fetchall()
        return [r[0] for r in rows if r and r[0]]
    except Exception:
        return []


def build_sheet_format_evidence(source_paths: list[str], *, sample: int = 24, max_pages_per_pdf: int = 6) -> dict:
    """PDF-геометрия листов → evidence для D4-001.

    Проверяет именно размер страницы по ГОСТ 2.301. Размещение рамки/штампа и заполненность граф —
    отдельные layout/title-block проверки, не смешиваем их с форматом листа.
    """
    try:
        import fitz
        from proxy.services.normcontrol_service import PT_TO_MM, classify_format
    except Exception:
        return {"pdf_total": 0, "checked_pages": 0, "bad_pages": [], "examples": [],
                "note": "fitz/normcontrol_service недоступен"}

    pdfs = [sp for sp in source_paths if sp and str(sp).lower().endswith(".pdf")]
    checked = 0
    bad: list[dict] = []
    examples: list[dict] = []
    for sp in pdfs[:sample]:
        p = Path(sp)
        try:
            with fitz.open(str(p)) as doc:
                page_indexes = list(range(min(max_pages_per_pdf, doc.page_count)))
                if doc.page_count > max_pages_per_pdf:
                    page_indexes.append(doc.page_count - 1)
                for page_idx in sorted(set(page_indexes)):
                    page = doc[page_idx]
                    width_mm = float(page.rect.width) * PT_TO_MM
                    height_mm = float(page.rect.height) * PT_TO_MM
                    fmt = classify_format(width_mm, height_mm)
                    checked += 1
                    ref = f"{p.name}#page={page_idx + 1}"
                    size = f"{width_mm:.0f}×{height_mm:.0f} мм"
                    if fmt is None:
                        bad.append({"source_ref": ref, "snippet": f"Нестандартный формат листа {size}."})
                    elif len(examples) < 8:
                        examples.append({"source_ref": ref, "snippet": f"Формат {fmt}, размер {size}."})
        except Exception as err:  # noqa: BLE001
            bad.append({"source_ref": p.name, "snippet": f"PDF-геометрию не удалось прочитать: {err}"})
    return {"pdf_total": len(pdfs), "checked_pages": checked, "bad_pages": bad, "examples": examples,
            "sample": min(sample, len(pdfs)), "max_pages_per_pdf": max_pages_per_pdf}


def _clip(text: object, limit: int = 140) -> str:
    value = " ".join(str(text or "").split())
    if len(value) <= limit:
        return value
    return value[: max(0, limit - 1)].rstrip() + "…"


def _md_cell(text: object, limit: int = 120) -> str:
    return _clip(text, limit).replace("|", "\\|").replace("\n", " ")


def _source_label(ref: object, *, limit: int = 90) -> str:
    value = _clip(ref, limit)
    return value or "источник не найден"


def _evidence_refs(it: ReviewItem, *, limit: int = 3) -> str:
    refs = [
        _source_label(ev.get("source_ref"), limit=70)
        for ev in (it.document_evidence or [])[:limit]
        if isinstance(ev, dict) and ev.get("source_ref")
    ]
    if len(it.document_evidence or []) > limit:
        refs.append(f"+{len(it.document_evidence or []) - limit}")
    return "; ".join(refs) if refs else "нет извлечённого evidence"


def _review_action(it: ReviewItem) -> str:
    if it.status == S_COMPUTED_ISSUE:
        return "подтвердить/отклонить замечание"
    if it.status == S_REVIEW_NEEDED:
        return "добрать источник требования/факт корпуса"
    if it.status == S_MANUAL:
        return "ручная проверка инженером"
    if it.status == S_NOT_APPLICABLE:
        return "зафиксировать неприменимость"
    return "оставить как evidence"


def _review_section_title(status: str) -> str:
    return {
        S_COMPUTED_ISSUE: "Предварительные замечания, которые уже можно защищать evidence",
        S_MANUAL: "Ручные проверки: машина не имеет права ставить финал",
        S_REVIEW_NEEDED: "Что нужно добрать для доказательности",
    }.get(status, _STATUS_RU.get(status, status))


def _append_review_table(lines: list[str], rows: list[ReviewItem], *, top: int) -> None:
    lines += [
        "| ID | Что проверено | Основание | Evidence комплекта | Почему так | Действие |",
        "|---|---|---|---|---|---|",
    ]
    for it in rows[:top]:
        requirement_ref = _source_label((it.requirement or {}).get("source_ref"), limit=78)
        computed = str((it.computed_check or {}).get("name") or "manual").strip()
        reason = f"{computed}: {it.model_note}" if computed else it.model_note
        lines.append(
            f"| {_md_cell(it.rule_id, 28)} | {_md_cell(it.target, 44)} | "
            f"{_md_cell(requirement_ref, 80)} | {_md_cell(_evidence_refs(it), 90)} | "
            f"{_md_cell(reason, 110)} | {_md_cell(_review_action(it), 44)} |"
        )
    if len(rows) > top:
        lines.append(f"| … | ещё {len(rows) - top} пункт(ов) | см. JSON/XLSX отчёт | — | — | — |")


def review_to_chat_text(items: list[ReviewItem], review_map: ReviewMap, *, top: int = 6) -> str:
    """Человеческий отчёт для чата: не трассировка, а defendable review summary.

    Инвариант: текст не содержит рабочую память/историю диалога; все доказательные данные приходят из
    review-items и дублируются в системном ``defense`` JSON-контракте.
    """
    s = review_summary(items)
    supported = s["by_status"].get(S_SUPPORTED, 0) + s["by_status"].get(S_NOT_APPLICABLE, 0)
    computed = [it for it in items if it.status == S_COMPUTED_ISSUE]
    manual = [it for it in items if it.status == S_MANUAL]
    missing = [it for it in items if it.status == S_REVIEW_NEEDED]
    lines = [
        f"**Нормоконтроль комплекта — {review_map.standard} (СПДС, предварительно)**",
        "",
        "**Вердикт машины:** это не финальное заключение нормоконтролёра. Машина показывает "
        "предварительные замечания, evidence и пробелы; финальный статус по каждому пункту "
        "ставит инженер.",
        "",
        f"Проверено пунктов: **{s['total']}**. Кодом найдено замечаний: **{s['computed_issues']}**. "
        f"Подтверждено/неприменимо без замечаний: **{supported}**. "
        f"Ручная проверка: **{s['manual_required']}**. Нужно добрать evidence/RAG: **{s['review_needed']}**.",
        "",
        "| Класс | Кол-во | Как это использовать |",
        "|---|---:|---|",
        f"| Замечания кодом | {s['computed_issues']} | Можно предъявлять как предварительные, если инженер подтвердит evidence. |",
        f"| Ручные проверки | {s['manual_required']} | Автоматика не финализирует layout/штамп/изменения без человека. |",
        f"| Нужно добрать evidence | {s['review_needed']} | Нельзя защищать как вывод, пока нет источника требования или факта корпуса. |",
        f"| Подтверждено/неприменимо | {supported} | Информационная часть отчёта, не замечания. |",
    ]
    for status, rows in ((S_COMPUTED_ISSUE, computed), (S_MANUAL, manual), (S_REVIEW_NEEDED, missing)):
        if not rows:
            continue
        lines += ["", f"### {_review_section_title(status)}"]
        _append_review_table(lines, rows, top=top)
    lines += [
        "",
        "### Защита решения",
        "- Основание каждого пункта — `requirement.source_ref` из rulepack/RAG по стандарту.",
        "- Evidence комплекта — `document_evidence.source_ref` или computed-check с именем проверки.",
        "- Если evidence нет, пункт специально остаётся `review_needed` или `manual_required`, а не превращается в уверенный вывод.",
        "- JSON/API дополнительно отдают `defense.contract`: те же claims, source_refs, gaps и required_actions для UI/экспорта.",
    ]
    return "\n".join(lines)


def review_to_json(items: list[ReviewItem], review_map: ReviewMap) -> dict:
    return {"standard": review_map.standard, "rulepack": review_map.name, "version": review_map.version,
            "summary": review_summary(items), "items": [it.to_dict() for it in items],
            "normalized_remarks": review_to_normalized_remarks(items, review_map),
            "defense": review_defense_pack(items, review_map),
            "note": "RAG-led SPDS review: статусы — proposed issues/evidence, финал ставит инженер."}


def review_to_normalized_remarks(items: list[ReviewItem], review_map: ReviewMap) -> list[dict]:
    """Shared machine-readable remark contract for doc-review/formal/checklist/report renderers.

    It is intentionally not a final expert verdict: each row keeps human_decision=unset until an
    engineer confirms/rejects the proposed remark.
    """
    out: list[dict] = []
    for it in items:
        source_refs: list[str] = []
        req_ref = str((it.requirement or {}).get("source_ref") or "").strip()
        if req_ref:
            source_refs.append(req_ref)
        for ev in it.document_evidence or []:
            ref = str((ev or {}).get("source_ref") or "").strip()
            if ref:
                source_refs.append(ref)
        out.append({
            "schema": "normalized_remark_v1",
            "id": it.rule_id,
            "source": "doc_review",
            "standard": review_map.standard,
            "rulepack": review_map.name,
            "clause": it.clause,
            "target": it.target,
            "status": it.status,
            "severity": it.severity,
            "requirement_ref": req_ref,
            "document_refs": [ref for ref in source_refs if ref != req_ref],
            "source_refs": source_refs,
            "computed_check": it.computed_check,
            "message": it.model_note,
            "human_decision": it.human_decision,
            "finality": "proposed" if it.human_decision == "unset" else "human_decided",
            "requires_human": it.status in (S_COMPUTED_ISSUE, S_MANUAL, S_REVIEW_NEEDED),
            "confidence": it.confidence,
        })
    return out


_STATUS_RU = {S_COMPUTED_ISSUE: "Замечание (код)", S_SUPPORTED: "Подтверждено evidence",
              S_NOT_APPLICABLE: "Неприменимо", S_MANUAL: "Ручная проверка", S_REVIEW_NEEDED: "Нужен RAG/обзор"}


def review_to_html(items: list[ReviewItem], review_map: ReviewMap) -> str:
    rows = "".join(
        f"<tr><td>{it.rule_id}</td><td>{it.clause}</td><td>{_STATUS_RU.get(it.status, it.status)}</td>"
        f"<td>{it.severity}</td><td>{it.target}</td><td>{it.model_note}</td></tr>"
        for it in items)
    s = review_summary(items)
    return (f"<!doctype html><meta charset='utf-8'><title>Нормоконтроль {review_map.standard}</title>"
            f"<h2>Проверка документации — {review_map.standard}</h2>"
            f"<p>Всего {s['total']} · замечаний(код) {s['computed_issues']} · ручных {s['manual_required']} · RAG {s['review_needed']}</p>"
            f"<p><i>RAG-led review: статусы — предлагаемые; финал ставит инженер.</i></p>"
            f"<table border=1 cellpadding=4 cellspacing=0><tr><th>Правило</th><th>Пункт</th><th>Статус</th>"
            f"<th>Severity</th><th>Объект</th><th>Пояснение</th></tr>{rows}</table>")


def review_to_xlsx(items: list[ReviewItem], output_path: Path, review_map: ReviewMap) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Нормоконтроль"
    ws.append([f"Проверка документации — {review_map.standard}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Правило", "Пункт", "Статус", "Severity", "Объект", "Требование", "Пояснение"])
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for it in items:
        ws.append([it.rule_id, it.clause, _STATUS_RU.get(it.status, it.status), it.severity,
                   it.target, it.requirement.get("source_ref", ""), it.model_note])
    for col, width in {"A": 22, "B": 8, "C": 20, "D": 10, "E": 34, "F": 34, "G": 60}.items():
        ws.column_dimensions[col].width = width
    wr = wb.create_sheet("normalized_remarks")
    wr.append(["id", "standard", "clause", "status", "severity", "target", "source_refs", "message", "finality"])
    for cell in wr[1]:
        cell.font = Font(bold=True)
    for remark in review_to_normalized_remarks(items, review_map):
        wr.append([
            remark["id"], remark["standard"], remark["clause"], remark["status"], remark["severity"],
            remark["target"], "; ".join(remark["source_refs"]), remark["message"], remark["finality"],
        ])
    for col, width in {"A": 22, "B": 24, "C": 10, "D": 20, "E": 10, "F": 34, "G": 60, "H": 60, "I": 14}.items():
        wr.column_dimensions[col].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(items)

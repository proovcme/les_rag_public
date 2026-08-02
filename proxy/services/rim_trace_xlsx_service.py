"""Рендерер РИМ-трассы ЛСР в XLSX по форме **Приложения № 3** к Методике 421/пр.

НЕ калькулятор: берёт ГОТОВУЮ трассу и раскладывает её строки по графам формы ЛСР. Два входа:

* :func:`render_trace_xlsx` — ОДНА позиция (``rim_lsr_trace_service.build_position_trace``);
* :func:`render_lsr_xlsx` — МНОГОПОЗИЦИОННАЯ ЛСР (``rim_lsr_trace_service.build_lsr_trace``):
  шапка с общим итогом + разделы (заголовок раздела → позиции с непрерывной нумерацией → «Итого по
  разделу N») + общий свод «ВСЕГО по смете».

Оба рендера делят шапку/графы/строки позиции/финализацию — числа НЕ пересчитываются (Σ уже сделана в
трассе). 0 LLM. Графы формы РИМ: № п/п · Обоснование · Наименование · Ед. изм. · Количество
(на ед./коэф./всего) · Сметная стоимость (базис/индекс/текущий уровень/коэф./всего).
Колонки trace.columns 2-12 ложатся на графы формы напрямую.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

# trace.columns (str) → колонка openpyxl (1-индекс) по форме Приложения 3.
# Графа № п/п → A (col 1, только строки-работы), графы 2-12 ложатся напрямую.
_COL: dict[str, int] = {str(i): i for i in range(2, 13)}
_NUM_COLS = {5, 6, 7, 8, 9, 10, 11, 12}

_GROUP_TYPES = {"group_labor", "group_machine", "group_machinist", "group_material",
                "direct_total", "fot", "nr", "sp", "position_total"}
_TOTAL_TYPES = {"direct_total", "fot", "nr", "sp", "position_total"}
_LABEL_FIX = {"Итого по позиции": "Всего по позиции"}

_TABLE_HEADERS = {
    1: "№ п/п",
    2: "Обоснование",
    3: "Наименование работ и затрат",
    4: "Ед. изм.",
    5: "Кол-во на ед.",
    6: "коэф.",
    7: "Кол-во всего",
    8: "Сметная стоимость в базисном уровне цен на ед., руб.",
    9: "Индекс",
    10: "Сметная стоимость в текущем уровне цен на ед., руб.",
    11: "коэф.",
    12: "Сметная стоимость в текущем уровне цен всего, руб.",
}
_WIDTHS = {1: 6, 2: 18, 3: 44, 4: 9, 5: 11, 6: 7, 7: 12, 8: 16, 9: 9, 10: 16, 11: 7, 12: 18}


def _f(value: Any) -> float:
    try:
        return float(str(value).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _styles() -> dict[str, Any]:
    """Стили формы (шрифты/границы/заливки), собираются один раз на лист."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="B0B8C0")
    return {
        "Alignment": Alignment,
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "bold": Font(bold=True, size=9),
        "small": Font(size=9),
        "dim": Font(size=8, color="606060"),
        "fill_head": PatternFill("solid", fgColor="D9EAF7"),    # шапка таблицы
        "fill_total": PatternFill("solid", fgColor="F2F6FA"),   # итоговые строки (позиция/раздел)
        "fill_section": PatternFill("solid", fgColor="EAF1F8"),  # заголовок раздела
        "fill_grand": PatternFill("solid", fgColor="CFE2F3"),   # «ВСЕГО по смете»
        "fill_warning": PatternFill("solid", fgColor="FCE8E6"),
        "warning": Font(bold=True, size=10, color="9C0006"),
    }


def _make_put(ws, S: dict[str, Any]):
    """Фабрика ячеечного `put`, замкнутого на лист + стили."""
    Alignment = S["Alignment"]

    def _put(r: int, c: int, v: Any, *, font=None, num: bool = False, align: str = "left"):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = font or S["small"]
        if num and isinstance(v, (int, float)):
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=(c == 3))
        return cell

    return _put


def _border_row(ws, S: dict[str, Any], r: int, *, fill=None) -> None:
    """Границы по графам 1-12 строки + опц. заливка (итоги/разделы)."""
    for c in range(1, 13):
        cc = ws.cell(row=r, column=c)
        cc.border = S["border"]
        if fill is not None:
            cc.fill = fill


def _header_block(ws, put, S: dict[str, Any], *, name: str, summary: dict[str, Any],
                  meta: dict[str, Any]) -> int:
    """Шапка формы (стройка/объект/ЛСР №/наименование/метод/уровень цен/субъект + сметная стоимость
    и её разбивка). Возвращает номер строки шапки таблицы (графы)."""
    r = 1
    put(r, 10, "Приложение № 3", font=S["dim"], align="right"); r += 1
    put(r, 10, "к Методике (приказ Минстроя России от 04.08.2020 № 421/пр)", font=S["dim"], align="right"); r += 2
    put(r, 1, meta.get("stroika", "(наименование стройки)"), font=S["dim"]); r += 1
    put(r, 1, meta.get("object", "(наименование объекта капитального строительства)"), font=S["dim"]); r += 2
    put(r, 1, "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ (СМЕТА) № " + str(meta.get("lsr_no", "____")), font=S["bold"]); r += 1
    put(r, 1, name or "(наименование работ и затрат)", font=S["small"]); r += 1
    put(
        r,
        1,
        str(meta.get("calculation_method") or "Составлен ресурсно-индексным методом (РИМ)"),
        font=S["dim"],
    ); r += 1
    basis = str(meta.get("osnovanie", "—"))
    if basis.startswith(("/", "file://")) or ":\\" in basis:
        basis = Path(basis.replace("file://", "")).name
    put(r, 1, "Основание: " + basis, font=S["dim"]); r += 1
    put(r, 1, "Составлен(а) в текущем уровне цен: " + str(meta.get("price_level", "—")), font=S["dim"]); r += 1
    put(r, 1, "Наименование субъекта РФ: " + str(meta.get("subject", "—")), font=S["dim"]); r += 2
    result_status = str(summary.get("result_status") or "").strip()
    if result_status and result_status != "priced_final":
        put(
            r,
            1,
            f"РАСЧЁТ НЕ ЗАВЕРШЁН ({result_status.upper()}) — НЕ ДЛЯ СОГЛАСОВАНИЯ",
            font=S["warning"],
        )
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = S["fill_warning"]
        r += 1
    input_rows = int(summary.get("input_rows") or 0)
    bound_rows = int(summary.get("bound_rows") or 0)
    open_rows = int(summary.get("open_rows", summary.get("unbound_rows")) or 0)
    covered_rows = int(summary.get("covered_rows") or 0)
    closed = bound_rows + covered_rows
    low_coverage = (
        input_rows > 0
        and open_rows > 0
        and (open_rows / input_rows >= 0.30 or closed / input_rows < 0.70)
    )
    if low_coverage:
        put(
            r,
            1,
            (
                f"ПОКРЫТИЕ НИЗКОЕ: привязано {bound_rows} из {input_rows} строк"
                f"{', ещё ' + str(covered_rows) + ' покрыто соседними' if covered_rows else ''}. "
                "Сумма в шапке — только по привязанным, не итог ведомости."
            ),
            font=S["warning"],
        )
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = S["fill_warning"]
        r += 2
    elif result_status and result_status != "priced_final":
        r += 1
    amount_complete = summary.get("full_amount") is not None
    if low_coverage and input_rows:
        cost_label = f"Стоимость только привязанной части ({bound_rows}/{input_rows})"
    else:
        cost_label = "Сметная стоимость" if amount_complete else "Стоимость рассчитанной части"
    put(r, 1, cost_label, font=S["bold"])
    put(r, 4, _f(summary.get("total_with_vat", summary.get("total", 0))), font=S["bold"], num=True)
    put(r, 6, "руб.", font=S["dim"]); r += 1
    put(r, 1, "  средства на оплату труда рабочих", font=S["dim"]); put(r, 4, _f(summary.get("ozp", 0)), font=S["dim"], num=True); r += 1
    put(r, 1, "  средства на оплату труда машинистов", font=S["dim"]); put(r, 4, _f(summary.get("zpm", 0)), font=S["dim"], num=True); r += 1
    put(r, 1, "  нормативные затраты труда рабочих, чел.-ч", font=S["dim"]); put(r, 4, _f(summary.get("labor_qty", 0)), font=S["dim"], num=True); r += 1
    put(r, 1, "  нормативные затраты труда машинистов, чел.-ч", font=S["dim"]); put(r, 4, _f(summary.get("machinist_qty", 0)), font=S["dim"], num=True); r += 2
    return r


def _table_header(ws, put, S: dict[str, Any], head_r: int) -> int:
    """Шапка таблицы (графы № п/п…Стоимость всего)."""
    for c, t in _TABLE_HEADERS.items():
        cell = put(head_r, c, t, font=S["bold"], align="center")
        cell.fill = S["fill_head"]
        cell.border = S["border"]
        cell.alignment = S["Alignment"](horizontal="center", vertical="center", wrap_text=True)
    return head_r


def _sum_formula(rows: list[int], column: str = "L") -> str:
    return "=SUM(" + ",".join(f"{column}{row}" for row in rows) + ")" if rows else "=0"


def _xlsx_work_name(row: dict[str, Any], value: Any) -> Any:
    source = str(value or "").strip()
    official = str((row.get("meta") or {}).get("official_name") or "").strip()
    if official and source and " ".join(official.lower().split()) != " ".join(source.lower().split()):
        return f"{official} / {source}"
    return official or value


def _position_rows(ws, put, S: dict[str, Any], rows: list[dict[str, Any]], start_r: int,
                   pp_start: int) -> tuple[int, int, dict[str, list[int]]]:
    """Строки одной позиции (работа → ОТ/ЭМ/ОТм/М → прямые/ФОТ/НР/СП/Всего) из готовой трассы.
    Непрерывная нумерация: ``pp_start`` → возвращается обновлённый счётчик. Возвращает (next_r, pp)."""
    r = start_r
    pp = pp_start
    by_type: dict[str, list[int]] = {}
    for row in rows:
        cols = row.get("columns", {}) or {}
        rtype = row.get("type", "")
        by_type.setdefault(str(rtype), []).append(r)
        is_group = rtype in _GROUP_TYPES
        is_total = rtype in _TOTAL_TYPES
        if rtype == "work":
            pp += 1
            put(r, 1, pp, font=S["bold"], align="center")
        for key, xc in _COL.items():
            if key in cols:
                val = _LABEL_FIX.get(str(cols[key]), cols[key]) if key == "3" else cols[key]
                if rtype == "work" and key == "3":
                    val = _xlsx_work_name(row, val)
                put(r, xc, val, font=(S["bold"] if is_group else S["small"]), num=(xc in _NUM_COLS))
        # наименование группы/итога, если в columns нет "3"
        if "3" not in cols:
            label = row.get("label", "")
            if label:
                put(r, 3, label, font=(S["bold"] if is_group else S["small"]))
        _border_row(ws, S, r, fill=(S["fill_total"] if is_total else None))
        r += 1
    # Editable workbook formulas. JSON trace remains the immutable calculation
    # record, while Excel recalculates after quantity/price edits.
    for rtype in ("resource_labor", "resource_machine", "resource_machinist", "resource_material"):
        for row_no in by_type.get(rtype, []):
            if ws.cell(row=row_no, column=8).value not in (None, "") and ws.cell(row=row_no, column=9).value not in (None, ""):
                ws.cell(row=row_no, column=10, value=f"=ROUND(H{row_no}*I{row_no},2)")
            if ws.cell(row=row_no, column=10).value not in (None, "") and ws.cell(row=row_no, column=11).value not in (None, ""):
                ws.cell(row=row_no, column=12, value=f"=ROUND(G{row_no}*J{row_no}*K{row_no},2)")
            else:
                ws.cell(row=row_no, column=12).value = None
    group_rows = {
        "labor": (by_type.get("group_labor") or [None])[0],
        "machine": (by_type.get("group_machine") or [None])[0],
        "machinist": (by_type.get("group_machinist") or [None])[0],
        "material": (by_type.get("group_material") or [None])[0],
    }
    detail_rows = {
        "labor": by_type.get("resource_labor", []),
        "machine": by_type.get("resource_machine", []),
        "machinist": by_type.get("resource_machinist", []),
        "material": by_type.get("resource_material", []),
    }
    for kind, row_no in group_rows.items():
        if row_no is None:
            continue
        ws.cell(row=row_no, column=7, value=_sum_formula(detail_rows[kind], "G"))
        cost_rows = detail_rows[kind]
        if kind == "machine":
            cost_rows = cost_rows + detail_rows["machinist"]
        ws.cell(row=row_no, column=12, value=_sum_formula(cost_rows))
    direct_row = (by_type.get("direct_total") or [None])[0]
    fot_row = (by_type.get("fot") or [None])[0]
    nr_row = (by_type.get("nr") or [None])[0]
    sp_row = (by_type.get("sp") or [None])[0]
    total_row = (by_type.get("position_total") or [None])[0]
    if direct_row:
        refs = [group_rows[k] for k in ("labor", "machine", "material") if group_rows[k]]
        ws.cell(row=direct_row, column=12, value=_sum_formula(refs))
    if fot_row:
        refs = [group_rows[k] for k in ("labor", "machinist") if group_rows[k]]
        ws.cell(row=fot_row, column=12, value=_sum_formula(refs))
    if nr_row and fot_row:
        ws.cell(row=nr_row, column=12, value=f"=ROUND(L{fot_row}*E{nr_row}/100,2)")
    if sp_row and fot_row:
        ws.cell(row=sp_row, column=12, value=f"=ROUND(L{fot_row}*E{sp_row}/100,2)")
    if total_row:
        refs = [row_no for row_no in (direct_row, nr_row, sp_row) if row_no]
        ws.cell(row=total_row, column=12, value=_sum_formula(refs))
    return r, pp, by_type


def _section_title(ws, put, S: dict[str, Any], r: int, idx: int, sec_name: str) -> int:
    """Строка-заголовок раздела «Раздел N. <наименование>»."""
    label = f"Раздел {idx}. {sec_name}" if sec_name and sec_name != "Без раздела" else f"Раздел {idx}"
    put(r, 3, label, font=S["bold"])
    _border_row(ws, S, r, fill=S["fill_section"])
    return r + 1


def _section_subtotal(
    ws, put, S: dict[str, Any], r: int, idx: int, total: Any, *, position_total_rows: list[int] | None = None
) -> int:
    """Строка «Итого по разделу N» с суммой по разделу в графе «Стоимость всего»."""
    put(r, 3, f"Итого по разделу {idx}", font=S["bold"])
    put(r, 12, _f(total), font=S["bold"], num=True)
    if position_total_rows is not None:
        ws.cell(row=r, column=12, value=_sum_formula(position_total_rows))
    _border_row(ws, S, r, fill=S["fill_total"])
    return r + 1


def _grand_summary(
    ws, put, S: dict[str, Any], r: int, summary: dict[str, Any], *, formula_rows: dict[str, list[int]] | None = None
) -> tuple[int, dict[str, int]]:
    """Общий свод сметы: прямые/ФОТ/НР/СП + «ВСЕГО по смете» (Σ уже в трассе, не пересчёт)."""
    output_rows: dict[str, int] = {}
    for label, key in (("Итого прямые затраты по смете", "direct_total"), ("В том числе ФОТ", "fot"),
                       ("Накладные расходы", "nr"), ("Сметная прибыль", "sp")):
        put(r, 3, label, font=S["bold"])
        summary_key = "direct" if key == "direct_total" else key
        put(r, 12, _f(summary.get(summary_key, 0)), font=S["bold"], num=True)
        if formula_rows is not None:
            ws.cell(row=r, column=12, value=_sum_formula(formula_rows.get(key, [])))
        output_rows[key] = r
        _border_row(ws, S, r, fill=S["fill_total"])
        r += 1
    total_label = "ВСЕГО по смете без НДС" if summary.get("full_amount") is not None else "ИЗВЕСТНАЯ РАССЧИТАННАЯ ЧАСТЬ БЕЗ НДС"
    put(r, 3, total_label, font=S["bold"])
    put(r, 12, _f(summary.get("total_without_vat", summary.get("total", 0))), font=S["bold"], num=True)
    if formula_rows is not None:
        ws.cell(row=r, column=12, value=_sum_formula(formula_rows.get("position_total", [])))
    output_rows["total_without_vat"] = r
    _border_row(ws, S, r, fill=S["fill_grand"])
    if summary.get("vat_pct") not in (None, ""):
        r += 1
        put(r, 3, f"НДС {float(summary.get('vat_pct')):g}%", font=S["bold"])
        put(r, 12, _f(summary.get("vat", 0)), font=S["bold"], num=True)
        ws.cell(row=r, column=12, value=f"=ROUND(L{output_rows['total_without_vat']}*{float(summary.get('vat_pct')):g}/100,2)")
        output_rows["vat"] = r
        _border_row(ws, S, r, fill=S["fill_total"])
        r += 1
        gross_label = "ВСЕГО по смете с НДС" if summary.get("full_amount") is not None else "ИЗВЕСТНАЯ РАССЧИТАННАЯ ЧАСТЬ С НДС"
        put(r, 3, gross_label, font=S["bold"])
        put(r, 12, _f(summary.get("total_with_vat", 0)), font=S["bold"], num=True)
        ws.cell(row=r, column=12, value=f"=L{output_rows['total_without_vat']}+L{output_rows['vat']}")
        output_rows["total_with_vat"] = r
        _border_row(ws, S, r, fill=S["fill_grand"])
    return r + 1, output_rows


def _finalize(ws, head_r: int) -> None:
    """Ширины граф, закрепление экрана и печатная форма на одну страницу по ширине."""
    import openpyxl
    from openpyxl.worksheet.page import PageMargins

    for c, w in _WIDTHS.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=head_r + 1, column=1)
    # Без явного print setup Calc/Excel разрезают 12 граф РИМ на отдельные
    # вертикальные страницы: на первых листах остаются названия, а суммы
    # печатаются отдельно. Форма должна быть длинной, но всегда цельной по ширине.
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = None
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.35, bottom=0.35, header=0.15, footer=0.15)
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"{head_r}:{head_r}"
    ws.print_area = f"A1:L{ws.max_row}"


def _new_sheet():
    """Новая книга + лист «ЛСР РИМ» + стили + `put`."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ЛСР РИМ"
    S = _styles()
    return wb, ws, S, _make_put(ws, S)


def _set_header_total_formula(ws, total_row: int | None) -> None:
    if not total_row:
        return
    for row_no in range(1, min(ws.max_row, 40) + 1):
        label = str(ws.cell(row=row_no, column=1).value or "")
        if label in {"Сметная стоимость", "Стоимость рассчитанной части"} or label.startswith(
            "Стоимость только привязанной части"
        ):
            ws.cell(row=row_no, column=4, value=f"=L{total_row}")
            return


def _review_sheet(wb, lsr: dict[str, Any]) -> None:
    """Visible decisions/gaps; JSON remains the full immutable trace."""
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Проверка")
    headers = [
        "Строка", "work_id", "Статус", "Норма", "Тип выбора", "Причина выбора",
        "Ограничения аналога", "Проверка ресурсов", "Решения по ресурсам",
        "Труд", "Машины", "Материалы", "Повторная проверка влияния",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for item in lsr.get("row_bindings") or []:
        resource_text = "\n".join(
            f"{entry.get('action')}: {entry.get('resource_code') or entry.get('resource_name')} "
            f"{entry.get('quantity') if entry.get('quantity') is not None else ''} {entry.get('unit') or ''}; "
            f"{entry.get('reason') or ''}; основание: {entry.get('basis_ref') or '—'}"
            for entry in (item.get("resource_bindings") or [])
        )
        ws.append([
            item.get("row"), item.get("work_id"), item.get("status"), item.get("code"),
            item.get("selection_kind") or ("covered_by " + str(item.get("covered_by_work_id") or "") if item.get("covered_by_work_id") else ""),
            item.get("reason") or item.get("coverage_reason"),
            "\n".join(str(value) for value in (item.get("analog_limitations") or [])),
            " — ".join(filter(None, [
                str(item.get("resource_review_status") or ""),
                str(item.get("resource_review_reason") or ""),
            ])),
            resource_text,
            " — ".join(filter(None, [
                str(item.get("labor_review_status") or ""),
                str(item.get("labor_review_reason") or ""),
            ])),
            " — ".join(filter(None, [
                str(item.get("machine_review_status") or ""),
                str(item.get("machine_review_reason") or ""),
            ])),
            " — ".join(filter(None, [
                str(item.get("material_review_status") or ""),
                str(item.get("material_review_reason") or ""),
            ])),
            " — ".join(filter(None, [
                str(item.get("dominant_review_status") or ""),
                str(item.get("dominant_review_reason") or ""),
            ])),
        ])
    ws.append([])
    ws.append(["Общие замечания и незакрытые позиции"])
    section_row = ws.max_row
    ws.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=13)
    ws.cell(section_row, 1).font = Font(bold=True)
    ws.cell(section_row, 1).fill = PatternFill("solid", fgColor="FCE4D6")
    ws.append(["Код", "work_id", "Тип", "Сообщение"])
    issue_header_row = ws.max_row
    ws.merge_cells(start_row=issue_header_row, start_column=4, end_row=issue_header_row, end_column=13)
    for cell in ws[issue_header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")

    def append_issue(item, issue_type: str) -> None:
        payload = item if isinstance(item, dict) else {}
        code = str(payload.get("code") or issue_type)
        work_id = str(payload.get("work_id") or "")
        message = str(payload.get("reason") or payload.get("message") or item)
        ws.append([code, work_id, issue_type, message])
        issue_row = ws.max_row
        ws.merge_cells(start_row=issue_row, start_column=4, end_row=issue_row, end_column=13)
        ws.row_dimensions[issue_row].height = max(30, min(90, 15 * (1 + len(message) // 110)))

    for item in lsr.get("blockers") or []:
        append_issue(item, "blocker")
    for item in lsr.get("warnings") or []:
        append_issue(item, "warning")
    widths = {
        "A": 28, "B": 16, "C": 22, "D": 24, "E": 14, "F": 50,
        "G": 48, "H": 42, "I": 55, "J": 38, "K": 38, "L": 38,
        "M": 42,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def render_trace_xlsx(trace: dict[str, Any], out_path: str | Path, *,
                      title: str | None = None, meta: Optional[dict[str, Any]] = None) -> Path:
    """Трасса ОДНОЙ позиции → XLSX по форме Приложения 3 к 421/пр. Возвращает путь."""
    meta = meta or {}
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb, ws, S, put = _new_sheet()
    head_r = _header_block(ws, put, S, name=trace.get("name", ""),
                           summary=trace.get("summary", {}) or {}, meta=meta)
    _table_header(ws, put, S, head_r)
    _, _, formula_rows = _position_rows(ws, put, S, trace.get("rows", []), head_r + 1, 0)
    _set_header_total_formula(ws, (formula_rows.get("position_total") or [None])[0])
    _finalize(ws, head_r)
    wb.save(path)
    return path


def render_lsr_xlsx(lsr: dict[str, Any], out_path: str | Path, *,
                    title: str | None = None, meta: Optional[dict[str, Any]] = None) -> Path:
    """Многопозиционная ЛСР (``build_lsr_trace``) → XLSX по форме Приложения 3: шапка с общим итогом +
    разделы (заголовок → позиции с непрерывной нумерацией → «Итого по разделу N») + «ВСЕГО по смете».

    Рендер ГОТОВОЙ трассы — числа те же, что у каждой позиции в /rim-trace, и Σ — в свод. 0 LLM."""
    meta = meta or {}
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    summary = lsr.get("summary", {}) or {}
    sections = lsr.get("sections", []) or []
    wb, ws, S, put = _new_sheet()
    head_r = _header_block(ws, put, S, name=lsr.get("name", ""), summary=summary, meta=meta)
    _table_header(ws, put, S, head_r)

    r = head_r + 1
    pp = 0
    all_formula_rows: dict[str, list[int]] = {}
    multi = len(sections) > 1
    for idx, sec in enumerate(sections, 1):
        section_position_totals: list[int] = []
        sec_name = str(sec.get("section", "") or "")
        show_sec = multi or (sec_name and sec_name != "Без раздела")
        if show_sec:
            r = _section_title(ws, put, S, r, idx, sec_name)
        for trace in sec.get("positions", []) or []:
            r, pp, formula_rows = _position_rows(ws, put, S, trace.get("rows", []), r, pp)
            for key, values in formula_rows.items():
                all_formula_rows.setdefault(key, []).extend(values)
            section_position_totals.extend(formula_rows.get("position_total", []))
        if show_sec:
            r = _section_subtotal(
                ws, put, S, r, idx, sec.get("total", 0), position_total_rows=section_position_totals
            )
    _, grand_rows = _grand_summary(ws, put, S, r, summary, formula_rows=all_formula_rows)
    _set_header_total_formula(
        ws,
        grand_rows.get("total_with_vat") or grand_rows.get("total_without_vat"),
    )
    _finalize(ws, head_r)
    _review_sheet(wb, lsr)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(path)
    return path
